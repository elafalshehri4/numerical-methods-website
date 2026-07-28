from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Sequence
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css

# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
        value: float | int | None,
        decimals: int = 3,
        unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
        value: float | int | None,
        decimals: int = 3,
        unavailable: str = "—",
) -> str:
    """Show three decimals and use × 10ⁿ when fixed notation loses meaning."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Newton Divided Differences Interpolation"
SUPPORTED_POINT_COUNTS = tuple(range(2, 13))
DEFAULT_POINT_COUNT = 4
DISPLAY_DECIMALS = 3
DUPLICATE_X_TOLERANCE = 100.0 * np.finfo(float).eps
NODE_RESIDUAL_TOLERANCE = 1.0e-10
CLOSE_NODE_WARNING_RATIO = math.sqrt(np.finfo(float).eps)
EXACT_NODE_MISMATCH_TOLERANCE = 1.0e-9
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
NEWTON_BASIS_GROWTH_WARNING = 1.0e8
CONDITION_NUMBER_WARNING = 1.0e12
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DEFAULT_X_VALUES = np.array(
    [-1.0, 0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0],
    dtype=float,
)
DEFAULT_Y_VALUES = np.array(
    [2.0, 1.0, 2.0, 5.0, 10.0, 17.0, 26.0, 37.0, 50.0, 65.0, 82.0, 101.0],
    dtype=float,
)


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class DividedDifferenceStep:
    """One manually calculated entry in the divided-difference table."""

    step_number: int
    order: int
    start_index: int
    end_index: int
    left_value: float
    right_value: float
    numerator: float
    denominator: float
    result: float
    notation: str
    operation: str


@dataclass(frozen=True)
class NewtonTermStep:
    """One Newton-form term evaluated at the requested target coordinate."""

    term_number: int
    order: int
    coefficient: float
    factor_text: str
    product_value: float
    contribution: float
    cumulative_value: float


@dataclass(frozen=True)
class InterpolationResult:
    """Complete result shared by Streamlit renderers and Excel export."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str
    point_count: int
    polynomial_degree: int
    original_x: np.ndarray
    original_y: np.ndarray
    target_x: float
    target_matches_node: bool
    minimum_node_separation: float | None
    scaled_minimum_node_separation: float | None
    divided_difference_table: np.ndarray | None
    divided_difference_history: tuple[DividedDifferenceStep, ...]
    newton_coefficients: np.ndarray | None
    evaluation_history: tuple[NewtonTermStep, ...]
    expanded_coefficients: np.ndarray | None
    newton_equation_text: str
    expanded_equation_text: str
    interpolated_value: float | None
    explicit_term_value: float | None
    nested_evaluation_difference: float | None
    newton_basis_growth_at_target: float | None
    node_predictions: np.ndarray | None
    node_residuals: np.ndarray | None
    node_absolute_residuals: np.ndarray | None
    node_residual_norm: float | None
    maximum_node_residual: float | None
    vandermonde_condition_number: float | None
    exact_function_text: str
    exact_expression_text: str
    exact_value_at_target: float | None
    exact_node_values: np.ndarray | None
    exact_node_errors: np.ndarray | None
    maximum_exact_node_mismatch: float | None
    absolute_error_at_target: float | None
    relative_error_at_target: float | None
    warnings: tuple[str, ...]
    duplicate_x_tolerance: float
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def format_number(
        value: float | int | None,
        decimals: int = 3,
) -> str:
    """Format displayed values with three decimals and × 10ⁿ notation."""

    return format_display_number(value, decimals)


def round_numeric_dataframe(
        dataframe: pd.DataFrame,
        decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns for display without changing stored precision."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def create_input_signature(
        data: pd.DataFrame,
        target_x: Any,
        exact_function_text: str,
) -> str:
    """Create a stable signature used to prevent stale Streamlit results."""

    serialized_rows = []
    for _, row in data.iterrows():
        serialized_rows.append((str(row.get("x", "")), str(row.get("y", ""))))
    payload = repr(
        (
            serialized_rows,
            str(target_x),
            str(exact_function_text).strip(),
            DUPLICATE_X_TOLERANCE,
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def safe_numeric_preview(values: Sequence[Any]) -> np.ndarray:
    """Return finite values that can safely be shown in an error result."""

    preview: list[float] = []
    for raw_value in values:
        if raw_value is None:
            continue
        if isinstance(raw_value, str) and not raw_value.strip():
            continue
        try:
            numeric_value = float(raw_value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(numeric_value):
            preview.append(numeric_value)
    return np.asarray(preview, dtype=float)


def default_points_dataframe(point_count: int) -> pd.DataFrame:
    """Return default editable interpolation points."""

    return pd.DataFrame(
        {
            "x": DEFAULT_X_VALUES[:point_count],
            "y": DEFAULT_Y_VALUES[:point_count],
        }
    )


# =============================================================================
# Input validation
# =============================================================================
def coerce_real_finite_vector(
        values: Sequence[Any],
        vector_name: str,
) -> np.ndarray:
    """Convert user values to a finite real one-dimensional NumPy vector."""

    converted: list[float] = []
    for index, raw_value in enumerate(values, start=1):
        if raw_value is None:
            raise ValueError(f"{vector_name} contains an empty value at row {index}.")
        if isinstance(raw_value, str) and not raw_value.strip():
            raise ValueError(f"{vector_name} contains an empty value at row {index}.")
        try:
            numeric_value = float(raw_value)
        except (TypeError, ValueError) as error:
            raise ValueError(
                f"{vector_name} contains a non-numeric value at row {index}."
            ) from error
        if not math.isfinite(numeric_value):
            raise ValueError(
                f"{vector_name} contains NaN or infinity at row {index}."
            )
        converted.append(numeric_value)

    vector = np.asarray(converted, dtype=float).reshape(-1)
    if vector.size == 0:
        raise ValueError(f"{vector_name} must contain at least one value.")
    return vector


def validate_distinct_x_values(x_values: np.ndarray) -> None:
    """Reject only true floating-point duplicates.

    Closely spaced but distinct nodes are permitted and diagnosed separately,
    because divided differences are mathematically defined for every distinct
    set of abscissas even when the problem is ill-conditioned.
    """

    nodes = np.asarray(x_values, dtype=float).reshape(-1)

    for first_index in range(len(nodes)):
        for second_index in range(first_index + 1, len(nodes)):
            first_x = float(nodes[first_index])
            second_x = float(nodes[second_index])
            scale = max(1.0, abs(first_x), abs(second_x))
            threshold = DUPLICATE_X_TOLERANCE * scale

            if abs(first_x - second_x) <= threshold:
                raise ValueError(
                    "Newton divided differences require distinct x-values. "
                    f"Rows {first_index + 1} and {second_index + 1} are equal "
                    "or indistinguishable at double precision."
                )


def calculate_node_spacing_diagnostics(
        x_values: np.ndarray,
) -> tuple[float, float]:
    """Return minimum absolute and scale-normalized pairwise node spacing."""

    nodes = np.asarray(x_values, dtype=float).reshape(-1)
    minimum_absolute = math.inf
    minimum_scaled = math.inf

    for first_index in range(len(nodes)):
        for second_index in range(first_index + 1, len(nodes)):
            difference = abs(float(nodes[first_index] - nodes[second_index]))
            scale = max(
                1.0,
                abs(float(nodes[first_index])),
                abs(float(nodes[second_index])),
            )
            minimum_absolute = min(minimum_absolute, difference)
            minimum_scaled = min(minimum_scaled, difference / scale)

    if not math.isfinite(minimum_absolute) or not math.isfinite(minimum_scaled):
        raise ValueError("At least two distinct interpolation nodes are required.")

    return float(minimum_absolute), float(minimum_scaled)


def validate_and_prepare_points(
        raw_data: pd.DataFrame,
) -> tuple[np.ndarray, np.ndarray]:
    """Validate the editable data table and return x and y vectors."""

    if not isinstance(raw_data, pd.DataFrame):
        raise ValueError("Data points must be supplied in a table.")
    if "x" not in raw_data.columns or "y" not in raw_data.columns:
        raise ValueError("The data table must contain x and y columns.")
    if len(raw_data) not in SUPPORTED_POINT_COUNTS:
        raise ValueError(
            "The number of interpolation points must be between "
            f"{min(SUPPORTED_POINT_COUNTS)} and {max(SUPPORTED_POINT_COUNTS)}."
        )

    x_values = coerce_real_finite_vector(raw_data["x"].tolist(), "x-values")
    y_values = coerce_real_finite_vector(raw_data["y"].tolist(), "y-values")

    if x_values.size != y_values.size:
        raise ValueError("The x and y vectors must contain the same number of values.")

    validate_distinct_x_values(x_values)
    return x_values, y_values


def validate_target_x(raw_target_x: Any) -> float:
    """Convert and validate the requested interpolation coordinate."""

    if raw_target_x is None:
        raise ValueError("The target x-value is required.")
    if isinstance(raw_target_x, str) and not raw_target_x.strip():
        raise ValueError("The target x-value is required.")
    try:
        target_x = float(raw_target_x)
    except (TypeError, ValueError) as error:
        raise ValueError("The target x-value must be numeric.") from error
    if not math.isfinite(target_x):
        raise ValueError("The target x-value must be finite.")
    return target_x


# =============================================================================
# Optional exact-function parsing
# =============================================================================
def parse_exact_function(
        function_text: str,
) -> tuple[sp.Expr | None, Callable[..., Any] | None]:
    """Parse a safe optional exact function of x for comparison only."""

    cleaned_text = str(function_text).strip()
    if not cleaned_text:
        return None, None
    if "=" in cleaned_text:
        raise ValueError(
            "Enter the exact function as an expression only, without f(x)= or an equals sign."
        )

    x_symbol = sp.Symbol("x", real=True)
    allowed_locals = {
        "x": x_symbol,
        "sin": sp.sin,
        "cos": sp.cos,
        "tan": sp.tan,
        "asin": sp.asin,
        "acos": sp.acos,
        "atan": sp.atan,
        "sinh": sp.sinh,
        "cosh": sp.cosh,
        "tanh": sp.tanh,
        "exp": sp.exp,
        "log": sp.log,
        "ln": sp.log,
        "sqrt": sp.sqrt,
        "Abs": sp.Abs,
        "abs": sp.Abs,
        "pi": sp.pi,
        "E": sp.E,
    }

    try:
        expression = sp.sympify(
            cleaned_text.replace("^", "**"),
            locals=allowed_locals,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The exact function has an invalid format. Use syntax such as "
            "sin(x), exp(x), or x**3 - 2*x + 1."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError("The exact function could not be interpreted as an expression.")

    unsupported_atoms = (
        AppliedUndef,
        sp.Derivative,
        sp.Integral,
        sp.Sum,
        sp.Product,
        sp.Limit,
    )
    if expression.has(*unsupported_atoms):
        raise ValueError("The exact function contains unsupported symbolic operations.")
    if isinstance(expression, Relational) or expression.has(Relational):
        raise ValueError("The exact function must be an expression, not a relation.")

    unexpected_symbols = expression.free_symbols.difference({x_symbol})
    if unexpected_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unexpected_symbols))
        raise ValueError(
            "The exact function may contain only x. "
            f"Unexpected symbol(s): {names}."
        )
    if expression.has(sp.I, sp.zoo, sp.nan, sp.oo, -sp.oo):
        raise ValueError("The exact function contains a complex or undefined value.")

    try:
        numeric_function = sp.lambdify(x_symbol, expression, modules=["numpy"])
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The exact function could not be converted to numerical form."
        ) from error

    return sp.simplify(expression), numeric_function


def evaluate_exact_function_scalar(
        numeric_function: Callable[..., Any],
        x_value: float,
) -> float:
    """Evaluate the optional exact function safely at one real point."""

    try:
        with np.errstate(all="raise"):
            raw_value = numeric_function(float(x_value))
        array = np.asarray(raw_value)
    except (
            ArithmeticError,
            TypeError,
            ValueError,
            OverflowError,
            ZeroDivisionError,
            FloatingPointError,
    ) as error:
        raise ValueError(
            f"The exact function is undefined at x = {x_value:.12g}. Reason: {error}"
        ) from error

    if array.size != 1:
        raise ValueError(
            f"The exact function did not return a scalar at x = {x_value:.12g}."
        )

    scalar = array.reshape(-1)[0]
    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)
        if abs(complex_value.imag) > 1.0e-12:
            raise ValueError(
                f"The exact function is complex at x = {x_value:.12g}."
            )
        scalar = complex_value.real

    try:
        value = float(scalar)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The exact function value at x = {x_value:.12g} is not real."
        ) from error

    if not math.isfinite(value):
        raise ValueError(
            f"The exact function is undefined or non-finite at x = {x_value:.12g}."
        )
    return value


def evaluate_exact_function_array(
        numeric_function: Callable[..., Any],
        x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate an exact function on a grid, preserving invalid values as NaN."""

    x_array = np.asarray(x_values, dtype=float)
    try:
        with np.errstate(all="ignore"):
            raw_values = numeric_function(x_array)
    except Exception:
        return np.full_like(x_array, np.nan, dtype=float)

    array = np.asarray(raw_values)
    if array.ndim == 0:
        array = np.full_like(
            x_array,
            array,
            dtype=complex if np.iscomplexobj(array) else float,
        )
    else:
        try:
            array = np.broadcast_to(array, x_array.shape)
        except ValueError:
            return np.full_like(x_array, np.nan, dtype=float)

    if np.iscomplexobj(array):
        complex_values = np.asarray(array, dtype=complex)
        real_mask = np.abs(np.imag(complex_values)) <= 1.0e-12
        result = np.real(complex_values).astype(float)
        result[~real_mask] = np.nan
    else:
        try:
            result = np.asarray(array, dtype=float).copy()
        except (TypeError, ValueError):
            return np.full_like(x_array, np.nan, dtype=float)

    result[~np.isfinite(result)] = np.nan
    return result


# =============================================================================
# Polynomial arithmetic and Newton divided differences
# =============================================================================
def multiply_polynomials_ascending(
        first: np.ndarray,
        second: np.ndarray,
) -> np.ndarray:
    """Multiply two polynomials stored in ascending-power coefficient order."""

    first_coefficients = np.asarray(first, dtype=float).reshape(-1)
    second_coefficients = np.asarray(second, dtype=float).reshape(-1)
    result = np.zeros(
        first_coefficients.size + second_coefficients.size - 1,
        dtype=float,
    )
    for first_power, first_value in enumerate(first_coefficients):
        for second_power, second_value in enumerate(second_coefficients):
            result[first_power + second_power] += first_value * second_value
    return result


def add_polynomials_ascending(
        first: np.ndarray,
        second: np.ndarray,
) -> np.ndarray:
    """Add ascending-power polynomial coefficient vectors safely."""

    result_size = max(len(first), len(second))
    result = np.zeros(result_size, dtype=float)
    result[: len(first)] += np.asarray(first, dtype=float)
    result[: len(second)] += np.asarray(second, dtype=float)
    return result


def evaluate_polynomial(
        coefficients: np.ndarray,
        x_values: float | np.ndarray,
) -> float | np.ndarray:
    """Evaluate ascending-power coefficients with Horner's method."""

    coefficient_vector = np.asarray(coefficients, dtype=float).reshape(-1)
    x_array = np.asarray(x_values, dtype=float)
    result = np.zeros_like(x_array, dtype=float)
    for coefficient in coefficient_vector[::-1]:
        result = result * x_array + coefficient
    if np.isscalar(x_values):
        return float(np.asarray(result))
    return result


def evaluate_newton_nested(
        x_values: np.ndarray,
        newton_coefficients: np.ndarray,
        evaluation_points: float | np.ndarray,
) -> float | np.ndarray:
    """Evaluate Newton form with stable nested multiplication."""

    nodes = np.asarray(x_values, dtype=float).reshape(-1)
    coefficients = np.asarray(newton_coefficients, dtype=float).reshape(-1)
    if len(nodes) != len(coefficients):
        raise ValueError("The Newton nodes and coefficient vectors must have equal length.")

    points = np.asarray(evaluation_points, dtype=float)
    result = np.full_like(points, coefficients[-1], dtype=float)

    for index in range(len(coefficients) - 2, -1, -1):
        result = coefficients[index] + (points - nodes[index]) * result

    if not np.all(np.isfinite(result)):
        raise ValueError("Nested Newton evaluation produced NaN or infinity.")

    if np.ndim(evaluation_points) == 0:
        return float(np.asarray(result).reshape(-1)[0])
    return result


def calculate_newton_basis_growth(
        evaluation_history: tuple[NewtonTermStep, ...],
        interpolated_value: float,
) -> float:
    """Measure cancellation by comparing absolute term sum with final value."""

    absolute_sum = math.fsum(abs(item.contribution) for item in evaluation_history)
    denominator = max(abs(float(interpolated_value)), np.finfo(float).tiny)
    growth = absolute_sum / denominator
    return float(growth) if math.isfinite(growth) else math.inf


def format_equation_number(value: float) -> str:
    """Format polynomial coefficients without truncating them to display precision."""

    number = float(value)
    if number == 0.0:
        return "0"
    return f"{number:.12g}"


def build_divided_difference_table(
        x_values: np.ndarray,
        y_values: np.ndarray,
) -> tuple[np.ndarray, tuple[DividedDifferenceStep, ...]]:
    """Build the recursive Newton divided-difference table manually."""

    nodes = np.asarray(x_values, dtype=float).reshape(-1)
    values = np.asarray(y_values, dtype=float).reshape(-1)
    point_count = len(nodes)

    table_work = np.full(
        (point_count, point_count),
        np.longdouble(np.nan),
        dtype=np.longdouble,
    )
    table_work[:, 0] = values.astype(np.longdouble)
    history: list[DividedDifferenceStep] = []
    step_number = 0

    for order in range(1, point_count):
        for start_index in range(point_count - order):
            end_index = start_index + order
            left_value = table_work[start_index, order - 1]
            right_value = table_work[start_index + 1, order - 1]
            numerator = right_value - left_value
            denominator = (
                    np.longdouble(nodes[end_index])
                    - np.longdouble(nodes[start_index])
            )

            scale = max(
                1.0,
                abs(float(nodes[end_index])),
                abs(float(nodes[start_index])),
            )
            if abs(float(denominator)) <= DUPLICATE_X_TOLERANCE * scale:
                raise ValueError(
                    "A divided difference denominator is zero at double precision. "
                    "Check for repeated x-values."
                )

            result = numerator / denominator
            if not np.isfinite(result):
                raise ValueError(
                    "A non-finite divided difference was produced. Rescale the "
                    "x-values or reduce their magnitude."
                )

            table_work[start_index, order] = result
            step_number += 1
            notation = (
                    "f["
                    + ", ".join(
                f"x{node_index}"
                for node_index in range(start_index, end_index + 1)
            )
                    + "]"
            )
            operation = (
                f"({float(right_value):.12g} - {float(left_value):.12g}) / "
                f"({nodes[end_index]:.12g} - {nodes[start_index]:.12g})"
            )
            history.append(
                DividedDifferenceStep(
                    step_number=step_number,
                    order=order,
                    start_index=start_index,
                    end_index=end_index,
                    left_value=float(left_value),
                    right_value=float(right_value),
                    numerator=float(numerator),
                    denominator=float(denominator),
                    result=float(result),
                    notation=notation,
                    operation=operation,
                )
            )

    table = np.asarray(table_work, dtype=float)
    if not np.all(np.isfinite(table[~np.isnan(table)])):
        raise ValueError("The divided-difference table contains non-finite values.")
    return table, tuple(history)


def evaluate_newton_form(
        x_values: np.ndarray,
        newton_coefficients: np.ndarray,
        target_x: float,
) -> tuple[float, tuple[NewtonTermStep, ...]]:
    """Accumulate textbook Newton terms at the target for explanation."""

    nodes = np.asarray(x_values, dtype=float).reshape(-1)
    coefficients = np.asarray(newton_coefficients, dtype=float).reshape(-1)

    cumulative_value = float(coefficients[0])
    product_value = 1.0
    history: list[NewtonTermStep] = [
        NewtonTermStep(
            term_number=1,
            order=0,
            coefficient=float(coefficients[0]),
            factor_text="1",
            product_value=1.0,
            contribution=float(coefficients[0]),
            cumulative_value=cumulative_value,
        )
    ]

    factor_parts: list[str] = []
    for order in range(1, len(coefficients)):
        product_value *= float(target_x - nodes[order - 1])
        if not math.isfinite(product_value):
            raise ValueError("Newton term evaluation produced a non-finite product.")

        factor_parts.append(f"(x - {format_equation_number(nodes[order - 1])})")
        contribution = float(coefficients[order]) * product_value
        cumulative_value = math.fsum([cumulative_value, contribution])
        if not math.isfinite(cumulative_value):
            raise ValueError("Newton term accumulation produced a non-finite result.")

        history.append(
            NewtonTermStep(
                term_number=order + 1,
                order=order,
                coefficient=float(coefficients[order]),
                factor_text=" × ".join(factor_parts),
                product_value=float(product_value),
                contribution=float(contribution),
                cumulative_value=float(cumulative_value),
            )
        )

    return float(cumulative_value), tuple(history)


def expand_newton_polynomial(
        x_values: np.ndarray,
        newton_coefficients: np.ndarray,
) -> np.ndarray:
    """Expand Newton form into ascending-power coefficients for display/export."""

    nodes = np.asarray(x_values, dtype=float).reshape(-1)
    coefficients = np.asarray(newton_coefficients, dtype=float).reshape(-1)
    expanded = np.array([np.longdouble(coefficients[0])], dtype=np.longdouble)
    product_polynomial = np.array([np.longdouble(1.0)], dtype=np.longdouble)

    for order in range(1, len(coefficients)):
        next_factor = np.array(
            [-np.longdouble(nodes[order - 1]), np.longdouble(1.0)],
            dtype=np.longdouble,
        )
        product_polynomial = np.convolve(product_polynomial, next_factor)
        contribution = np.longdouble(coefficients[order]) * product_polynomial

        result_size = max(len(expanded), len(contribution))
        combined = np.zeros(result_size, dtype=np.longdouble)
        combined[: len(expanded)] += expanded
        combined[: len(contribution)] += contribution
        expanded = combined

    expanded_float = np.asarray(expanded, dtype=float)
    cleanup_scale = max(1.0, float(np.max(np.abs(expanded_float))))
    expanded_float[
        np.abs(expanded_float)
        <= 100.0 * np.finfo(float).eps * cleanup_scale
        ] = 0.0

    if not np.all(np.isfinite(expanded_float)):
        raise ValueError(
            "Expanded coefficients overflowed. The Newton form remains the preferred representation."
        )
    return expanded_float


def newton_equation_text(
        x_values: np.ndarray,
        coefficients: np.ndarray | None,
) -> str:
    """Return a readable Newton-form polynomial without display rounding loss."""

    if coefficients is None or len(coefficients) == 0:
        return "P(x) is not available."

    nodes = np.asarray(x_values, dtype=float)
    coefficient_array = np.asarray(coefficients, dtype=float)
    terms: list[str] = []
    factors: list[str] = []

    for order, coefficient in enumerate(coefficient_array):
        coefficient_value = float(coefficient)
        if order > 0:
            factors.append(f"(x - {format_equation_number(nodes[order - 1])})")

        coefficient_scale = max(1.0, float(np.max(np.abs(coefficient_array))))
        if abs(coefficient_value) <= 100.0 * np.finfo(float).eps * coefficient_scale:
            continue

        magnitude = format_equation_number(abs(coefficient_value))
        body = magnitude if order == 0 else magnitude + "".join(factors)

        if not terms:
            terms.append(("-" if coefficient_value < 0 else "") + body)
        else:
            terms.append(("- " if coefficient_value < 0 else "+ ") + body)

    return "P(x) = " + (" ".join(terms) if terms else "0")


def expanded_equation_text(coefficients: np.ndarray | None) -> str:
    """Return the expanded power form without truncating coefficients."""

    if coefficients is None or len(coefficients) == 0:
        return "P(x) is not available."

    coefficient_array = np.asarray(coefficients, dtype=float)
    scale = max(1.0, float(np.max(np.abs(coefficient_array))))
    terms: list[str] = []

    for power in range(len(coefficient_array) - 1, -1, -1):
        coefficient = float(coefficient_array[power])
        if abs(coefficient) <= 100.0 * np.finfo(float).eps * scale:
            continue

        magnitude = abs(coefficient)
        if power == 0:
            body = format_equation_number(magnitude)
        elif power == 1:
            body = "x" if math.isclose(magnitude, 1.0) else f"{format_equation_number(magnitude)}x"
        else:
            body = f"x^{power}" if math.isclose(magnitude, 1.0) else f"{format_equation_number(magnitude)}x^{power}"

        if not terms:
            terms.append(("-" if coefficient < 0 else "") + body)
        else:
            terms.append(("- " if coefficient < 0 else "+ ") + body)

    return "P(x) = " + (" ".join(terms) if terms else "0")


def calculate_relative_error(absolute_error: float, exact_value: float) -> float:
    """Return absolute relative true error as a percentage."""

    if abs(exact_value) <= RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
        return float("nan")
    return float(absolute_error / abs(exact_value) * 100.0)


# =============================================================================
# Solver controller
# =============================================================================
def empty_error_result(
        message: str,
        raw_data: pd.DataFrame,
        raw_target_x: Any,
        exact_function_text: str,
        input_signature: str,
) -> InterpolationResult:
    """Create a structured error result without crashing the Streamlit page."""

    x_preview = (
        safe_numeric_preview(raw_data["x"].tolist())
        if isinstance(raw_data, pd.DataFrame) and "x" in raw_data.columns
        else np.asarray([], dtype=float)
    )
    y_preview = (
        safe_numeric_preview(raw_data["y"].tolist())
        if isinstance(raw_data, pd.DataFrame) and "y" in raw_data.columns
        else np.asarray([], dtype=float)
    )
    try:
        target_preview = float(raw_target_x)
        if not math.isfinite(target_preview):
            target_preview = 0.0
    except (TypeError, ValueError):
        target_preview = 0.0

    return InterpolationResult(
        status="Error",
        success=False,
        method=METHOD_NAME,
        message=message,
        stopping_reason="Input validation or numerical computation failed.",
        point_count=min(len(x_preview), len(y_preview)),
        polynomial_degree=max(min(len(x_preview), len(y_preview)) - 1, 0),
        original_x=x_preview,
        original_y=y_preview,
        target_x=target_preview,
        target_matches_node=False,
        minimum_node_separation=None,
        scaled_minimum_node_separation=None,
        divided_difference_table=None,
        divided_difference_history=tuple(),
        newton_coefficients=None,
        evaluation_history=tuple(),
        expanded_coefficients=None,
        newton_equation_text="P(x) is not available.",
        expanded_equation_text="P(x) is not available.",
        interpolated_value=None,
        explicit_term_value=None,
        nested_evaluation_difference=None,
        newton_basis_growth_at_target=None,
        node_predictions=None,
        node_residuals=None,
        node_absolute_residuals=None,
        node_residual_norm=None,
        maximum_node_residual=None,
        vandermonde_condition_number=None,
        exact_function_text=str(exact_function_text).strip(),
        exact_expression_text="",
        exact_value_at_target=None,
        exact_node_values=None,
        exact_node_errors=None,
        maximum_exact_node_mismatch=None,
        absolute_error_at_target=None,
        relative_error_at_target=None,
        warnings=tuple(),
        duplicate_x_tolerance=DUPLICATE_X_TOLERANCE,
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


def solve_newton_divided_differences(
        raw_data: pd.DataFrame,
        raw_target_x: Any,
        exact_function_text: str,
        input_signature: str,
) -> InterpolationResult:
    """Run the complete Newton divided-difference interpolation workflow."""

    try:
        x_values, y_values = validate_and_prepare_points(raw_data)
        target_x = validate_target_x(raw_target_x)
        minimum_separation, scaled_minimum_separation = (
            calculate_node_spacing_diagnostics(x_values)
        )
        exact_expression, exact_function = parse_exact_function(exact_function_text)

        difference_table, difference_history = build_divided_difference_table(
            x_values,
            y_values,
        )
        newton_coefficients = difference_table[0, :].copy()
        if not np.all(np.isfinite(newton_coefficients)):
            raise ValueError("The Newton coefficients contain NaN or infinity.")

        explicit_term_value, evaluation_history = evaluate_newton_form(
            x_values,
            newton_coefficients,
            target_x,
        )
        interpolated_value = float(
            evaluate_newton_nested(x_values, newton_coefficients, target_x)
        )
        nested_evaluation_difference = abs(interpolated_value - explicit_term_value)
        basis_growth = calculate_newton_basis_growth(
            evaluation_history,
            interpolated_value,
        )
        expanded_coefficients = expand_newton_polynomial(
            x_values,
            newton_coefficients,
        )

        node_predictions = np.asarray(
            evaluate_newton_nested(x_values, newton_coefficients, x_values),
            dtype=float,
        )
        node_residuals = node_predictions - y_values
        node_absolute_residuals = np.abs(node_residuals)
        node_residual_norm = float(np.linalg.norm(node_residuals, ord=2))
        maximum_node_residual = float(np.max(node_absolute_residuals))

        vandermonde_matrix = np.vander(
            x_values,
            N=len(x_values),
            increasing=True,
        )
        with np.errstate(all="ignore"):
            vandermonde_condition_number = float(np.linalg.cond(vandermonde_matrix))

        target_scale = np.maximum(
            1.0,
            np.maximum(abs(target_x), np.abs(x_values)),
        )
        target_matches_node = bool(
            np.any(
                np.abs(target_x - x_values)
                <= DUPLICATE_X_TOLERANCE * target_scale
            )
        )

        exact_value_at_target: float | None = None
        exact_node_values: np.ndarray | None = None
        exact_node_errors: np.ndarray | None = None
        maximum_exact_node_mismatch: float | None = None
        absolute_error_at_target: float | None = None
        relative_error_at_target: float | None = None

        warnings: list[str] = []

        if exact_function is not None:
            exact_node_values = np.asarray(
                [
                    evaluate_exact_function_scalar(exact_function, node)
                    for node in x_values
                ],
                dtype=float,
            )
            exact_node_errors = np.abs(exact_node_values - y_values)
            maximum_exact_node_mismatch = float(np.max(exact_node_errors))
            mismatch_scale = max(
                1.0,
                float(np.max(np.abs(y_values))),
                float(np.max(np.abs(exact_node_values))),
            )
            if maximum_exact_node_mismatch > EXACT_NODE_MISMATCH_TOLERANCE * mismatch_scale:
                warnings.append(
                    "The optional exact function does not reproduce all supplied y-values. "
                    "The interpolation polynomial is based on the table data, while the "
                    "exact function is used only as an external comparison."
                )

            exact_value_at_target = evaluate_exact_function_scalar(
                exact_function,
                target_x,
            )
            absolute_error_at_target = abs(
                interpolated_value - exact_value_at_target
            )
            relative_error_at_target = calculate_relative_error(
                absolute_error_at_target,
                exact_value_at_target,
            )

        minimum_x = float(np.min(x_values))
        maximum_x = float(np.max(x_values))
        if target_x < minimum_x or target_x > maximum_x:
            warnings.append(
                "The target lies outside the node interval. This is polynomial "
                "extrapolation and may be substantially less reliable."
            )
        if scaled_minimum_separation <= CLOSE_NODE_WARNING_RATIO:
            warnings.append(
                "Some x-values are extremely close relative to their scale. "
                "Divided differences and high-order coefficients may be sensitive "
                "to rounding even though the nodes are distinct."
            )
        if not math.isfinite(vandermonde_condition_number):
            warnings.append(
                "The equivalent power-basis interpolation problem is numerically singular."
            )
        elif vandermonde_condition_number >= CONDITION_NUMBER_WARNING:
            warnings.append(
                "The Vandermonde condition number is very large. The Newton form is "
                "preferred; expanded power coefficients may be highly sensitive."
            )
        if maximum_node_residual > NODE_RESIDUAL_TOLERANCE * max(1.0, float(np.max(np.abs(y_values)))):
            warnings.append(
                "The nested Newton polynomial does not reproduce the supplied nodes "
                "within the expected floating-point tolerance."
            )
        if basis_growth >= NEWTON_BASIS_GROWTH_WARNING:
            warnings.append(
                "Large cancellation occurred among Newton terms at the target. "
                "The final value may be sensitive to coefficient rounding and node order."
            )
        if nested_evaluation_difference > 1.0e-10 * max(1.0, abs(interpolated_value)):
            warnings.append(
                "Direct term accumulation and nested Newton evaluation differ noticeably. "
                "The nested value is reported because it is numerically more stable."
            )
        if result_degree := len(x_values) - 1:
            if result_degree >= 9:
                warnings.append(
                    "A high-degree global interpolation polynomial can oscillate between nodes. "
                    "Inspect the graph and consider local interpolation when appropriate."
                )

        return InterpolationResult(
            status="Success",
            success=True,
            method=METHOD_NAME,
            message="Execution completed successfully.",
            stopping_reason=(
                "The recursive divided-difference table was completed and the "
                "Newton polynomial was evaluated with nested multiplication."
            ),
            point_count=len(x_values),
            polynomial_degree=len(x_values) - 1,
            original_x=x_values.copy(),
            original_y=y_values.copy(),
            target_x=target_x,
            target_matches_node=target_matches_node,
            minimum_node_separation=minimum_separation,
            scaled_minimum_node_separation=scaled_minimum_separation,
            divided_difference_table=difference_table.copy(),
            divided_difference_history=difference_history,
            newton_coefficients=newton_coefficients.copy(),
            evaluation_history=evaluation_history,
            expanded_coefficients=expanded_coefficients.copy(),
            newton_equation_text=newton_equation_text(x_values, newton_coefficients),
            expanded_equation_text=expanded_equation_text(expanded_coefficients),
            interpolated_value=interpolated_value,
            explicit_term_value=explicit_term_value,
            nested_evaluation_difference=nested_evaluation_difference,
            newton_basis_growth_at_target=basis_growth,
            node_predictions=node_predictions.copy(),
            node_residuals=node_residuals.copy(),
            node_absolute_residuals=node_absolute_residuals.copy(),
            node_residual_norm=node_residual_norm,
            maximum_node_residual=maximum_node_residual,
            vandermonde_condition_number=vandermonde_condition_number,
            exact_function_text=str(exact_function_text).strip(),
            exact_expression_text=str(exact_expression) if exact_expression is not None else "",
            exact_value_at_target=exact_value_at_target,
            exact_node_values=(None if exact_node_values is None else exact_node_values.copy()),
            exact_node_errors=(None if exact_node_errors is None else exact_node_errors.copy()),
            maximum_exact_node_mismatch=maximum_exact_node_mismatch,
            absolute_error_at_target=absolute_error_at_target,
            relative_error_at_target=relative_error_at_target,
            warnings=tuple(dict.fromkeys(warnings)),
            duplicate_x_tolerance=DUPLICATE_X_TOLERANCE,
            input_signature=input_signature,
            execution_datetime=current_report_datetime(),
        )

    except (ValueError, TypeError, OverflowError, FloatingPointError) as error:
        return empty_error_result(
            message=str(error),
            raw_data=raw_data,
            raw_target_x=raw_target_x,
            exact_function_text=exact_function_text,
            input_signature=input_signature,
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def input_points_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return the original interpolation nodes as a labeled DataFrame."""

    return pd.DataFrame(
        {
            "Node": [f"x{i}" for i in range(result.point_count)],
            "x": result.original_x,
            "y = f(x)": result.original_y,
        }
    )


def divided_difference_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return the complete triangular divided-difference table."""

    if result.divided_difference_table is None:
        return pd.DataFrame()

    data: dict[str, Any] = {
        "Node": [f"x{i}" for i in range(result.point_count)],
        "x": result.original_x,
        "f[x]": result.divided_difference_table[:, 0],
    }
    for order in range(1, result.point_count):
        data[f"Order {order} Difference"] = result.divided_difference_table[:, order]
    return pd.DataFrame(data)


def divided_difference_steps_dataframe(
        result: InterpolationResult,
) -> pd.DataFrame:
    """Return every divided-difference calculation as one table row."""

    rows = []
    for step in result.divided_difference_history:
        rows.append(
            {
                "Step": step.step_number,
                "Order": step.order,
                "Start Node": f"x{step.start_index}",
                "End Node": f"x{step.end_index}",
                "Notation": step.notation,
                "Left Previous Difference": step.left_value,
                "Right Previous Difference": step.right_value,
                "Numerator": step.numerator,
                "Denominator": step.denominator,
                "Result": step.result,
                "Operation": step.operation,
            }
        )
    return pd.DataFrame(rows)


def newton_coefficients_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return the Newton coefficients from the first table row."""

    if result.newton_coefficients is None:
        return pd.DataFrame()

    factor_texts = ["1"]
    factors: list[str] = []
    for order in range(1, result.point_count):
        factors.append(f"(x - {format_number(result.original_x[order - 1])})")
        factor_texts.append(" × ".join(factors))

    return pd.DataFrame(
        {
            "Order": np.arange(result.point_count),
            "Newton Coefficient": result.newton_coefficients,
            "Coefficient Notation": [
                "f[x0]"
                if order == 0
                else "f["
                     + ", ".join(f"x{i}" for i in range(order + 1))
                     + "]"
                for order in range(result.point_count)
            ],
            "Newton Factor": factor_texts,
        }
    )


def evaluation_steps_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return Newton-term accumulation at the target coordinate."""

    rows = []
    for step in result.evaluation_history:
        rows.append(
            {
                "Term": step.term_number,
                "Order": step.order,
                "Coefficient": step.coefficient,
                "Factor Product": step.factor_text,
                "Product at Target": step.product_value,
                "Term Contribution": step.contribution,
                "Cumulative P(x)": step.cumulative_value,
            }
        )
    return pd.DataFrame(rows)


def expanded_coefficients_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return expanded polynomial coefficients in descending display order."""

    if result.expanded_coefficients is None:
        return pd.DataFrame()

    powers = np.arange(len(result.expanded_coefficients) - 1, -1, -1)
    return pd.DataFrame(
        {
            "Power": powers,
            "Term": ["1" if power == 0 else ("x" if power == 1 else f"x^{power}") for power in powers],
            "Coefficient": result.expanded_coefficients[powers],
        }
    )


def node_residuals_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return interpolation residuals at the supplied nodes."""

    if result.node_predictions is None or result.node_residuals is None:
        return pd.DataFrame()
    return pd.DataFrame(
        {
            "Node": [f"x{i}" for i in range(result.point_count)],
            "x": result.original_x,
            "Observed y": result.original_y,
            "Predicted P(x)": result.node_predictions,
            "Residual P(x) - y": result.node_residuals,
            "Absolute Residual": result.node_absolute_residuals,
        }
    )


def evaluation_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return target evaluation, stability checks, and optional exact error."""

    rows: list[dict[str, Any]] = [
        {"Quantity": "Target x", "Value": result.target_x},
        {"Quantity": "Target matches an input node", "Value": result.target_matches_node},
        {"Quantity": "Nested Newton P(x)", "Value": result.interpolated_value},
        {"Quantity": "Explicit term accumulation", "Value": result.explicit_term_value},
        {"Quantity": "Nested/direct difference", "Value": result.nested_evaluation_difference},
        {"Quantity": "Newton basis growth indicator", "Value": result.newton_basis_growth_at_target},
    ]
    if result.exact_value_at_target is not None:
        rows.extend(
            [
                {"Quantity": "Exact f(x)", "Value": result.exact_value_at_target},
                {"Quantity": "Absolute Error", "Value": result.absolute_error_at_target},
                {"Quantity": "Relative Error (%)", "Value": result.relative_error_at_target},
            ]
        )
    return pd.DataFrame(rows)


def exact_function_check_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return optional exact-function consistency at all supplied nodes."""

    if result.exact_node_values is None or result.exact_node_errors is None:
        return pd.DataFrame(
            {"Message": ["No exact function was supplied."]}
        )

    return pd.DataFrame(
        {
            "Node": [f"x{i}" for i in range(result.point_count)],
            "x": result.original_x,
            "Supplied y": result.original_y,
            "Exact f(x)": result.exact_node_values,
            "Absolute Data Mismatch": result.exact_node_errors,
        }
    )


def method_formula_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return the textbook formulas and numerical implementation notes."""

    return pd.DataFrame(
        {
            "Item": [
                "Newton polynomial",
                "First divided difference",
                "Recursive higher-order difference",
                "Newton coefficients",
                "Nested evaluation",
                "Polynomial degree",
                "Node requirement",
                "True-error availability",
            ],
            "Formula / Meaning": [
                "P_n(x)=b_0+b_1(x-x_0)+...+b_n∏_{j=0}^{n-1}(x-x_j)",
                "f[x_i,x_j]=(f(x_i)-f(x_j))/(x_i-x_j)",
                "f[x_i,...,x_{i+k}]=(right previous-left previous)/(x_{i+k}-x_i)",
                "b_k=f[x_0,...,x_k] from the first row of the table",
                "P=b_n; P=b_k+(x-x_k)P for k=n-1,...,0",
                f"At most {result.polynomial_degree}",
                "All x-values must be distinct; equal spacing is not required",
                "Requires an independent exact function; node residuals alone are not true error",
            ],
        }
    )


# =============================================================================
# Plotting
# =============================================================================
def create_plot_interval(result: InterpolationResult) -> tuple[float, float]:
    """Create a readable graph interval containing nodes and target x."""

    all_x = np.concatenate([result.original_x, np.array([result.target_x])])
    minimum_x = float(np.min(all_x))
    maximum_x = float(np.max(all_x))
    span = maximum_x - minimum_x
    padding = 1.0 if span == 0.0 else 0.15 * span
    return minimum_x - padding, maximum_x + padding


def build_plot_dataframe(
        result: InterpolationResult,
        point_count: int = 500,
) -> pd.DataFrame:
    """Create dense Newton-form and optional exact-function plot data."""

    if result.newton_coefficients is None:
        return pd.DataFrame()

    x_minimum, x_maximum = create_plot_interval(result)
    graph_x = np.linspace(x_minimum, x_maximum, point_count)
    graph_y = np.asarray(
        evaluate_newton_nested(
            result.original_x,
            result.newton_coefficients,
            graph_x,
        ),
        dtype=float,
    )
    dataframe = pd.DataFrame({"x": graph_x, "Newton P(x)": graph_y})

    if result.exact_expression_text:
        _, exact_function = parse_exact_function(result.exact_function_text)
        if exact_function is not None:
            exact_values = evaluate_exact_function_array(exact_function, graph_x)
            dataframe["Exact f(x)"] = exact_values
            dataframe["Absolute Error"] = np.abs(graph_y - exact_values)

    return dataframe


def create_interpolation_figure(result: InterpolationResult) -> plt.Figure:
    """Create the interpolation curve, nodes, target, and optional exact curve."""

    plot_data = build_plot_dataframe(result)
    figure, axis = plt.subplots(figsize=(11, 6.5))
    axis.plot(
        plot_data["x"],
        plot_data["Newton P(x)"],
        linewidth=2.2,
        label="Newton interpolation polynomial",
    )

    if "Exact f(x)" in plot_data.columns:
        exact_values = plot_data["Exact f(x)"].to_numpy(dtype=float)
        valid = np.isfinite(exact_values)
        if np.any(valid):
            axis.plot(
                plot_data.loc[valid, "x"],
                exact_values[valid],
                linestyle="--",
                linewidth=1.8,
                label="Exact function",
            )

    axis.scatter(
        result.original_x,
        result.original_y,
        s=70,
        zorder=5,
        label="Interpolation nodes",
    )
    axis.scatter(
        [result.target_x],
        [result.interpolated_value],
        s=120,
        marker="*",
        zorder=6,
        label=(
            f"P({result.target_x:.6g}) = {result.interpolated_value:.6g}"
        ),
    )
    axis.axvline(result.target_x, linestyle=":", linewidth=1.2)
    axis.axhline(0.0, linewidth=1.0)
    axis.set_title("Newton Divided-Difference Interpolation")
    axis.set_xlabel("x")
    axis.set_ylabel("y")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_error_figure(result: InterpolationResult) -> plt.Figure | None:
    """Create a semilog exact-function error graph when available."""

    if not result.exact_expression_text:
        return None
    plot_data = build_plot_dataframe(result)
    if "Absolute Error" not in plot_data.columns:
        return None

    finite_mask = np.isfinite(plot_data["Absolute Error"].to_numpy(dtype=float))
    if not np.any(finite_mask):
        return None

    x_values = plot_data.loc[finite_mask, "x"].to_numpy(dtype=float)
    error_values = plot_data.loc[finite_mask, "Absolute Error"].to_numpy(dtype=float)
    safe_errors = np.maximum(error_values, np.finfo(float).tiny)

    figure, axis = plt.subplots(figsize=(11, 6))
    axis.semilogy(x_values, safe_errors, linewidth=2.0)
    axis.axvline(result.target_x, linestyle=":", linewidth=1.2)
    axis.set_title("Absolute Interpolation Error")
    axis.set_xlabel("x")
    axis.set_ylabel("|P(x) - f(x)| (log scale)")
    axis.grid(True, which="both", alpha=0.3)
    figure.tight_layout()
    return figure


# =============================================================================
# Excel export
# =============================================================================
def style_excel_workbook(workbook: Any) -> None:
    """Apply professional formatting to every generated worksheet."""

    header_fill = PatternFill("solid", fgColor="0D3151")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions

        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            maximum_length = 0
            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                value_length = len(str(cell.value)) if cell.value is not None else 0
                maximum_length = max(maximum_length, value_length)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"
            worksheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 12),
                55,
            )


def figure_to_png_bytes(figure: plt.Figure) -> bytes:
    """Serialize a matplotlib figure as PNG bytes."""

    buffer = BytesIO()
    figure.savefig(buffer, format="png", dpi=180, bbox_inches="tight")
    buffer.seek(0)
    return buffer.getvalue()


def add_excel_image(
        worksheet: Any,
        image_bytes: bytes,
        anchor: str,
        width: int = 780,
        height: int = 500,
) -> None:
    """Insert a PNG image into an Excel worksheet."""

    image_stream = BytesIO(image_bytes)
    excel_image = ExcelImage(image_stream)
    excel_image.width = width
    excel_image.height = height
    worksheet.add_image(excel_image, anchor)


def create_excel_report(result: InterpolationResult) -> bytes:
    """Generate a complete formatted XLSX Newton interpolation report."""

    if not result.success:
        raise ValueError("Only successful interpolation results can be exported.")

    summary_df = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "Message",
                "Number of Points",
                "Polynomial Degree",
                "Target x",
                "Target Matches Node",
                "Interpolated Value (Nested Newton)",
                "Explicit Term Value",
                "Nested/Direct Difference",
                "Newton Basis Growth Indicator",
                "Minimum Node Separation",
                "Scaled Minimum Node Separation",
                "Newton Polynomial",
                "Expanded Polynomial",
                "Node Residual 2-Norm",
                "Maximum Absolute Node Residual",
                "Vandermonde Condition Number",
                "Exact Function",
                "Exact Value at Target",
                "Maximum Exact-Node Mismatch",
                "Absolute Error at Target",
                "Relative Error at Target (%)",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                result.message,
                result.point_count,
                result.polynomial_degree,
                result.target_x,
                result.target_matches_node,
                result.interpolated_value,
                result.explicit_term_value,
                result.nested_evaluation_difference,
                result.newton_basis_growth_at_target,
                result.minimum_node_separation,
                result.scaled_minimum_node_separation,
                result.newton_equation_text,
                result.expanded_equation_text,
                result.node_residual_norm,
                result.maximum_node_residual,
                result.vandermonde_condition_number,
                result.exact_expression_text or "Not provided",
                result.exact_value_at_target,
                result.maximum_exact_node_mismatch,
                result.absolute_error_at_target,
                result.relative_error_at_target,
                " | ".join(result.warnings) if result.warnings else "None",
                result.stopping_reason,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )

    plot_df = build_plot_dataframe(result)
    error_metrics_df = pd.DataFrame(
        {
            "Metric": [
                "Node Residual 2-Norm",
                "Maximum Absolute Node Residual",
                "Nested/Direct Difference",
                "Newton Basis Growth Indicator",
                "Maximum Exact-Node Mismatch",
                "Absolute Error at Target",
                "Relative Error at Target (%)",
            ],
            "Value": [
                result.node_residual_norm,
                result.maximum_node_residual,
                result.nested_evaluation_difference,
                result.newton_basis_growth_at_target,
                result.maximum_exact_node_mismatch,
                result.absolute_error_at_target,
                result.relative_error_at_target,
            ],
        }
    )

    interpolation_figure = create_interpolation_figure(result)
    interpolation_png = figure_to_png_bytes(interpolation_figure)
    plt.close(interpolation_figure)

    error_figure = create_error_figure(result)
    error_png: bytes | None = None
    if error_figure is not None:
        error_png = figure_to_png_bytes(error_figure)
        plt.close(error_figure)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        method_formula_dataframe(result).to_excel(
            writer, sheet_name="Method Formulas", index=False
        )
        input_points_dataframe(result).to_excel(
            writer, sheet_name="Input Points", index=False
        )
        divided_difference_dataframe(result).to_excel(
            writer, sheet_name="Divided Difference Table", index=False
        )
        divided_difference_steps_dataframe(result).to_excel(
            writer, sheet_name="Difference Steps", index=False
        )
        newton_coefficients_dataframe(result).to_excel(
            writer, sheet_name="Newton Coefficients", index=False
        )
        evaluation_steps_dataframe(result).to_excel(
            writer, sheet_name="Evaluation Steps", index=False
        )
        expanded_coefficients_dataframe(result).to_excel(
            writer, sheet_name="Expanded Polynomial", index=False
        )
        evaluation_dataframe(result).to_excel(
            writer, sheet_name="Target Evaluation", index=False
        )
        node_residuals_dataframe(result).to_excel(
            writer, sheet_name="Node Residuals", index=False
        )
        exact_function_check_dataframe(result).to_excel(
            writer, sheet_name="Exact Function Check", index=False
        )
        plot_df.to_excel(writer, sheet_name="Plot Data", index=False)
        error_metrics_df.to_excel(writer, sheet_name="Error Analysis", index=False)

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Newton Divided Differences Report Plots"
        plots_sheet["A1"].font = Font(bold=True, size=14)
        add_excel_image(plots_sheet, interpolation_png, "A3")
        if error_png is not None:
            add_excel_image(plots_sheet, error_png, "A31")

        summary_sheet = workbook["Summary"]
        plot_sheet = workbook["Plot Data"]
        if len(plot_df) > 1:
            chart = ScatterChart()
            chart.title = "Newton Interpolation Polynomial"
            chart.x_axis.title = "x"
            chart.y_axis.title = "y"
            chart.height = 8
            chart.width = 15

            x_reference = Reference(
                plot_sheet,
                min_col=1,
                min_row=2,
                max_row=len(plot_df) + 1,
            )
            polynomial_reference = Reference(
                plot_sheet,
                min_col=2,
                min_row=2,
                max_row=len(plot_df) + 1,
            )
            chart.series.append(
                Series(polynomial_reference, x_reference, title="Newton P(x)")
            )

            if "Exact f(x)" in plot_df.columns:
                exact_reference = Reference(
                    plot_sheet,
                    min_col=3,
                    min_row=2,
                    max_row=len(plot_df) + 1,
                )
                chart.series.append(
                    Series(exact_reference, x_reference, title="Exact f(x)")
                )
            summary_sheet.add_chart(chart, "D2")

            if "Absolute Error" in plot_df.columns:
                error_chart = ScatterChart()
                error_chart.title = "Absolute Interpolation Error"
                error_chart.x_axis.title = "x"
                error_chart.y_axis.title = "Absolute Error"
                error_chart.height = 8
                error_chart.width = 15
                error_reference = Reference(
                    plot_sheet,
                    min_col=4,
                    min_row=2,
                    max_row=len(plot_df) + 1,
                )
                error_chart.series.append(
                    Series(error_reference, x_reference, title="|P(x)-f(x)|")
                )
                summary_sheet.add_chart(error_chart, "D20")

        style_excel_workbook(workbook)
        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit result renderers
# =============================================================================

def finalize_excel_report_with_visible_charts(report_bytes: bytes) -> bytes:
    """Compatibility helper; charts are already placed on Summary."""

    return report_bytes


def render_final_result(result: InterpolationResult) -> None:
    """Render the compact final-result card."""

    if not result.success:
        st.error(result.message)
        return

    st.success(result.message)
    st.markdown(f"**{result.newton_equation_text}**")

    metric_columns = st.columns(2)
    metric_columns[0].metric("Target x", format_number(result.target_x))
    metric_columns[1].metric(
        "Interpolated Value",
        format_number(result.interpolated_value),
    )

    detail_columns = st.columns(2)
    detail_columns[0].metric("Points", result.point_count)
    detail_columns[1].metric("Polynomial Degree", result.polynomial_degree)

    stability_columns = st.columns(2)
    stability_columns[0].metric(
        "Nested/Direct Difference",
        format_scientific_power(result.nested_evaluation_difference),
    )
    stability_columns[1].metric(
        "Basis Growth Indicator",
        format_number(result.newton_basis_growth_at_target),
    )

    if result.target_matches_node:
        st.info("The target coincides with an interpolation node; the supplied y-value is reproduced.")

    if result.exact_value_at_target is not None:
        exact_columns = st.columns(3)
        exact_columns[0].metric(
            "Exact Value",
            format_number(result.exact_value_at_target),
        )
        exact_columns[1].metric(
            "Absolute Error",
            format_number(result.absolute_error_at_target),
        )
        relative_value = result.relative_error_at_target
        exact_columns[2].metric(
            "Relative Error (%)",
            "Undefined" if relative_value is not None and math.isnan(relative_value)
            else format_number(relative_value),
        )

    for warning in result.warnings:
        st.warning(warning)


def render_input_points(result: InterpolationResult) -> None:
    """Render original nodes and interpolation diagnostics."""

    st.subheader("Original Interpolation Points")
    st.dataframe(
        round_numeric_dataframe(input_points_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    diagnostic_columns = st.columns(4)
    diagnostic_columns[0].metric("Number of Points", result.point_count)
    diagnostic_columns[1].metric("Polynomial Degree", result.polynomial_degree)
    diagnostic_columns[2].metric(
        "Minimum Node Separation",
        format_scientific_power(result.minimum_node_separation),
    )
    diagnostic_columns[3].metric(
        "Vandermonde Condition Number",
        format_scientific_power(result.vandermonde_condition_number),
    )


def render_divided_difference_table(result: InterpolationResult) -> None:
    """Render the complete triangular divided-difference calculation."""

    st.subheader("Divided Difference Table")
    st.latex(
        r"f[x_i,\ldots,x_{i+k}]="
        r"\frac{f[x_{i+1},\ldots,x_{i+k}]-"
        r"f[x_i,\ldots,x_{i+k-1}]}{x_{i+k}-x_i}"
    )
    st.dataframe(
        round_numeric_dataframe(divided_difference_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Divided Difference Calculations")
    st.dataframe(
        round_numeric_dataframe(divided_difference_steps_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    for step in result.divided_difference_history:
        with st.expander(
                f"Step {step.step_number}: Order {step.order} — {step.notation}"
        ):
            st.markdown(f"**Formula substitution:** `{step.operation}`")
            calculation_df = pd.DataFrame(
                {
                    "Quantity": [
                        "Right previous difference",
                        "Left previous difference",
                        "Numerator",
                        "Denominator",
                        "Result",
                    ],
                    "Value": [
                        step.right_value,
                        step.left_value,
                        step.numerator,
                        step.denominator,
                        step.result,
                    ],
                }
            )
            st.dataframe(
                round_numeric_dataframe(calculation_df),
                use_container_width=True,
                hide_index=True,
            )


def render_polynomial_evaluation(result: InterpolationResult) -> None:
    """Render Newton coefficients, both polynomial forms, and evaluation steps."""

    st.subheader("Newton Polynomial Coefficients")
    st.dataframe(
        round_numeric_dataframe(newton_coefficients_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Newton Form")
    st.markdown(f"### {result.newton_equation_text}")

    st.subheader("Expanded Power Form")
    st.markdown(f"### {result.expanded_equation_text}")
    st.dataframe(
        round_numeric_dataframe(expanded_coefficients_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Evaluation at the Target x-value")
    st.caption(
        "Each Newton term is evaluated using the ordered node products and "
        "added to the cumulative polynomial value."
    )
    st.dataframe(
        round_numeric_dataframe(evaluation_steps_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )
    st.dataframe(
        round_numeric_dataframe(evaluation_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Function Graph")
    interpolation_figure = create_interpolation_figure(result)
    st.pyplot(interpolation_figure, use_container_width=True)
    plt.close(interpolation_figure)


def render_error_analysis(result: InterpolationResult) -> None:
    """Render node residuals and optional exact-function comparison."""

    st.subheader("Interpolation Residual Analysis")
    st.caption(
        "Node residuals verify that the polynomial reproduces the supplied data. "
        "They are not the unknown interpolation error between nodes."
    )
    st.dataframe(
        round_numeric_dataframe(node_residuals_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    metric_columns = st.columns(3)
    metric_columns[0].metric(
        "Node Residual 2-Norm",
        format_scientific_power(result.node_residual_norm),
    )
    metric_columns[1].metric(
        "Maximum Node Residual",
        format_scientific_power(result.maximum_node_residual),
    )
    metric_columns[2].metric(
        "Nested/Direct Difference",
        format_scientific_power(result.nested_evaluation_difference),
    )

    st.subheader("Exact-Function Consistency")
    st.dataframe(
        round_numeric_dataframe(exact_function_check_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    if result.exact_value_at_target is None:
        st.info(
            "No exact function was provided. True absolute and relative errors "
            "cannot be calculated from interpolation data alone."
        )
        return

    comparison_columns = st.columns(3)
    comparison_columns[0].metric(
        "Exact f(x)",
        format_number(result.exact_value_at_target),
    )
    comparison_columns[1].metric(
        "Absolute Error",
        format_number(result.absolute_error_at_target),
    )
    relative_error = result.relative_error_at_target
    comparison_columns[2].metric(
        "Relative Error (%)",
        "Undefined (exact value is zero)"
        if relative_error is not None and math.isnan(relative_error)
        else format_number(relative_error),
    )

    error_figure = create_error_figure(result)
    if error_figure is not None:
        st.pyplot(error_figure, use_container_width=True)
        plt.close(error_figure)
    else:
        st.warning(
            "The exact-function error graph could not be created because the "
            "comparison function is not finite over the plotted interval."
        )


def render_excel_download(result: InterpolationResult) -> None:
    """Generate and render the Excel report download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature
    signature_key = "newton_divided_differences_excel_signature"
    report_key = "newton_divided_differences_excel_report"

    if st.session_state.get(signature_key) != report_signature:
        try:
            st.session_state[report_key] = create_excel_report(result)
            st.session_state[signature_key] = report_signature
        except (ValueError, OSError, PermissionError) as error:
            st.error(f"Excel report generation failed: {error}")
            return

    report_bytes = st.session_state.get(report_key)
    if report_bytes is None:
        st.error("The Excel report is not available.")
        return

    date_text = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=f"newton_divided_differences_report_{date_text}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="newton_divided_differences_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Streamlit Newton interpolation solver page."""

    st.set_page_config(
        page_title="Newton Divided Differences Solver | Numerical Methods",
        page_icon="📈",
        layout="wide",
    )
    load_css()

    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">INTERPOLATION TOOL</div>
                <h1>Newton Divided Differences Solver</h1>
                <p>
                    Enter distinct data points, choose a target x-value, and
                    follow the complete divided-difference table, Newton
                    coefficients, polynomial evaluation, graph, and error analysis.
                </p>

                <div class="method-actions">
                    <a href="/Newton_Divided_Differences" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Polynomial_Interpolation_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Match the centered Bisection solver content width.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader('How to Enter the Data')
                st.markdown(
                    """
                Enter the interpolation nodes as paired numerical values **(xᵢ, yᵢ)**.

                - Each row contains one finite data point.
                - All **x** values must be distinct.
                - The entered node order is preserved; changing the order changes the Newton coefficients but not the final polynomial in exact arithmetic.
                - Enter the target **x** value for evaluation.
                - The optional exact function is used only for reference error analysis.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader('Before Solving')
                st.markdown(
                    """
                - Repeated **x** values cause division by zero in the divided differences.
                - With **n** points, the Newton polynomial has degree at most **n − 1**.
                - The solver evaluates the Newton form with nested multiplication; the expanded coefficients are provided mainly for interpretation.
                - Extrapolation beyond the data range may be inaccurate even when all node residuals are zero.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">Number of data points</div>',
                    unsafe_allow_html=True,
                )
                point_count = st.selectbox(
                    "Number of data points",
                    options=SUPPORTED_POINT_COUNTS,
                    index=SUPPORTED_POINT_COUNTS.index(DEFAULT_POINT_COUNT),
                    label_visibility="collapsed",
                    key="newton_divided_differences_point_count",
                )

                st.markdown(
                    '<div class="input-label-ui">Interpolation points</div>',
                    unsafe_allow_html=True,
                )
                data_editor_key = f"newton_divided_differences_data_editor_{point_count}"
                edited_data = st.data_editor(
                    default_points_dataframe(point_count),
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    column_config={
                        "x": st.column_config.NumberColumn(
                            "x",
                            help="Distinct interpolation-node coordinate",
                            format="%.10g",
                        ),
                        "y": st.column_config.NumberColumn(
                            "y",
                            help="Function or observed value at the node",
                            format="%.10g",
                        ),
                    },
                    key=data_editor_key,
                )

                target_columns = st.columns(2)
                with target_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Target x-value</div>',
                        unsafe_allow_html=True,
                    )
                    target_x = st.number_input(
                        "Target x-value",
                        value=1.5,
                        format="%.10g",
                        label_visibility="collapsed",
                        key="newton_divided_differences_target_x",
                    )
                with target_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Optional exact function</div>',
                        unsafe_allow_html=True,
                    )
                    exact_function_text = st.text_input(
                        "Optional exact function",
                        value="",
                        placeholder="Example: x**2 + 1",
                        label_visibility="collapsed",
                        key="newton_divided_differences_exact_function",
                    )

                st.caption(
                    "All x-values must be distinct. Their entered order is preserved "
                    "in the divided-difference table. The exact function is optional."
                )

                solve_button_clicked = st.button(
                    "Solve",
                    use_container_width=True,
                    key="newton_divided_differences_solve_button",
                )

        current_input_signature = create_input_signature(
            edited_data,
            target_x,
            exact_function_text,
        )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                stored_result = st.session_state.get("newton_divided_differences_result")
                if stored_result is None:
                    st.info("Enter the data points and click Solve to see the result.")
                elif stored_result.input_signature != current_input_signature:
                    st.info(
                        "The points, target value, or exact function has changed. "
                        "Click Solve to calculate a new result."
                    )
                else:
                    render_final_result(stored_result)

        if solve_button_clicked:
            st.session_state.newton_divided_differences_result = (
                solve_newton_divided_differences(
                    raw_data=edited_data.copy(),
                    raw_target_x=target_x,
                    exact_function_text=exact_function_text,
                    input_signature=current_input_signature,
                )
            )
            st.session_state.pop("newton_divided_differences_excel_report", None)
            st.session_state.pop("newton_divided_differences_excel_signature", None)
            st.rerun()

        active_result = st.session_state.get("newton_divided_differences_result")
        if (
                active_result is not None
                and active_result.input_signature == current_input_signature
        ):
            if active_result.success:
                st.divider()
                render_input_points(active_result)

                st.divider()
                render_divided_difference_table(active_result)

                st.divider()
                render_polynomial_evaluation(active_result)

                st.divider()
                render_error_analysis(active_result)

                st.divider()
                render_excel_download(active_result)

                st.divider()
                navigation_left_column, navigation_right_column = st.columns(2)

                with navigation_left_column:
                    if st.button(
                            "Review Newton Divided Differences Lesson",
                            use_container_width=True,
                            key="review_newton_divided_differences_lesson",
                    ):
                        st.switch_page("pages/Newton_Divided_Differences.py")

                with navigation_right_column:
                    if st.button(
                            "Back to Solver Menu",
                            use_container_width=True,
                            key="back_to_solver_menu_newton_divided_differences",
                    ):
                        st.switch_page("pages/Numerical_Solver.py")

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Interpolation • Newton Divided Differences</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()