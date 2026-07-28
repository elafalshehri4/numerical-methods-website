from __future__ import annotations

import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any, Callable
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css


# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0**exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Show fixed notation, using scientific notation when needed."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6
    ):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


def format_number(
    value: float | int | None,
    decimals: int = 3,
) -> str:
    """Format a displayed numerical value."""

    return format_display_number(value, decimals)


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Fourth-Order Runge–Kutta Method (Classical RK4)"
DISPLAY_DECIMALS = 3
DEFAULT_ORDER = 1
DEFAULT_ODE = "x + y"
DEFAULT_X0 = 0.0
DEFAULT_Y0 = 1.0
DEFAULT_YP0 = 0.0
DEFAULT_YPP0 = 0.0
DEFAULT_X_END = 1.0
DEFAULT_STEPS = 10
MIN_STEPS = 1
MAX_STEPS = 10000
DEFAULT_CONVERGENCE_LEVELS = 5
MIN_CONVERGENCE_LEVELS = 3
MAX_CONVERGENCE_LEVELS = 7
MAX_REFINED_STEPS = 200000
ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
VALUE_MAGNITUDE_WARNING = 1.0e12
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

X_SYMBOL = sp.Symbol("x", real=True)
Y_SYMBOL = sp.Symbol("y", real=True)
YP_SYMBOL = sp.Symbol("yp", real=True)
YPP_SYMBOL = sp.Symbol("ypp", real=True)

ALLOWED_FUNCTION_NAMES = {
    "x": X_SYMBOL,
    "y": Y_SYMBOL,
    "yp": YP_SYMBOL,
    "v": YP_SYMBOL,
    "ypp": YPP_SYMBOL,
    "a": YPP_SYMBOL,
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}

EXACT_SOLUTION_NAMES = {
    key: value
    for key, value in ALLOWED_FUNCTION_NAMES.items()
    if key not in {"y", "yp", "v", "ypp", "a"}
}



# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class RK4Iteration:
    """One classical RK4 step for an ODE of order 1, 2, or 3."""

    iteration: int
    x_n: float
    y_n: float
    yp_n: float | None
    ypp_n: float | None
    step_size: float

    k1_y: float
    k1_yp: float | None
    k1_ypp: float | None

    stage2_x: float
    stage2_y: float
    stage2_yp: float | None
    stage2_ypp: float | None
    k2_y: float
    k2_yp: float | None
    k2_ypp: float | None

    stage3_x: float
    stage3_y: float
    stage3_yp: float | None
    stage3_ypp: float | None
    k3_y: float
    k3_yp: float | None
    k3_ypp: float | None

    stage4_x: float
    stage4_y: float
    stage4_yp: float | None
    stage4_ypp: float | None
    k4_y: float
    k4_yp: float | None
    k4_ypp: float | None

    weighted_slope_y: float
    weighted_slope_yp: float | None
    weighted_slope_ypp: float | None

    x_next: float
    y_next: float
    yp_next: float | None
    ypp_next: float | None

    exact_y_next: float | None
    exact_yp_next: float | None
    exact_ypp_next: float | None

    signed_error_y: float | None
    signed_error_yp: float | None
    signed_error_ypp: float | None

    absolute_error_y: float | None
    absolute_error_yp: float | None
    absolute_error_ypp: float | None

    relative_error_y_percent: float | None
    relative_error_yp_percent: float | None
    relative_error_ypp_percent: float | None

    local_truncation_error_y: float | None
    local_truncation_error_yp: float | None
    local_truncation_error_ypp: float | None

    stage_operation: str
    update_operation: str
    status: str


@dataclass(frozen=True)
class ConvergenceRecord:
    """One grid-refinement result for the classical RK4 method."""

    level: int
    steps: int
    step_size: float

    final_y: float
    final_yp: float | None
    final_ypp: float | None

    exact_final_y: float | None
    exact_final_yp: float | None
    exact_final_ypp: float | None

    absolute_error_y: float | None
    absolute_error_yp: float | None
    absolute_error_ypp: float | None

    successive_difference_y: float | None
    observed_order: float | None


@dataclass(frozen=True)
class RK4Result:
    """Complete RK4 result shared by Streamlit and Excel renderers."""

    status: str
    success: bool
    method: str
    ode_order: int
    message: str
    stopping_reason: str

    ode_text: str
    ode_expression: sp.Expr | None
    exact_solution_text: str
    exact_solution_expression: sp.Expr | None

    x0: float | None
    y0: float | None
    yp0: float | None
    ypp0: float | None
    x_end: float | None
    steps: int
    step_size: float | None
    direction: str

    iterations: tuple[RK4Iteration, ...]

    final_x: float | None
    final_y: float | None
    final_yp: float | None
    final_ypp: float | None

    exact_final_y: float | None
    exact_final_yp: float | None
    exact_final_ypp: float | None

    signed_final_error_y: float | None
    signed_final_error_yp: float | None
    signed_final_error_ypp: float | None

    absolute_final_error_y: float | None
    absolute_final_error_yp: float | None
    absolute_final_error_ypp: float | None

    relative_final_error_y_percent: float | None
    relative_final_error_yp_percent: float | None
    relative_final_error_ypp_percent: float | None

    maximum_absolute_error_y: float | None
    rmse_y: float | None

    convergence_records: tuple[ConvergenceRecord, ...]
    latest_observed_order: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def current_report_datetime() -> datetime:
    """Return the current date and time in the report time zone."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def create_input_signature(
    ode_order: int,
    ode_text: str,
    x0: float,
    y0: float,
    yp0: float | None,
    ypp0: float | None,
    x_end: float,
    steps: int,
    exact_solution_text: str,
    convergence_levels: int,
) -> str:
    """Create a stable signature used to detect stale Streamlit results."""

    payload = "|".join(
        [
            str(int(ode_order)),
            str(ode_text).strip(),
            f"{float(x0):.17g}",
            f"{float(y0):.17g}",
            "" if yp0 is None else f"{float(yp0):.17g}",
            "" if ypp0 is None else f"{float(ypp0):.17g}",
            f"{float(x_end):.17g}",
            str(int(steps)),
            str(exact_solution_text).strip(),
            str(int(convergence_levels)),
        ]
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def validate_finite_real(value: Any, value_name: str) -> float:
    """Convert a scalar to a finite real float or raise ValueError."""

    try:
        complex_value = complex(value)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(f"{value_name} is not a valid numerical value.") from error

    if abs(complex_value.imag) > ZERO_TOLERANCE:
        raise ValueError(f"{value_name} is complex; a real value is required.")

    real_value = float(complex_value.real)
    if not math.isfinite(real_value):
        raise ValueError(f"{value_name} is NaN or infinity.")
    return real_value


def round_numeric_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Return a display copy with numerical columns rounded."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(DISPLAY_DECIMALS)
    return rounded


def order_label(order: int) -> str:
    """Return a readable differential-equation order label."""

    return {
        1: "First-order ODE",
        2: "Second-order ODE",
        3: "Third-order ODE",
    }[order]


def highest_derivative_label(order: int) -> str:
    """Return the highest derivative label."""

    return {1: "y′", 2: "y″", 3: "y‴"}[order]


def rhs_variables_text(order: int) -> str:
    """Return the variables accepted in the right-hand side."""

    return {
        1: "x, y",
        2: "x, y, yp",
        3: "x, y, yp, ypp",
    }[order]


# =============================================================================
# Safe symbolic parsing and evaluation
# =============================================================================
def _reject_unsupported_symbolic_constructs(
    expression: sp.Expr,
    allowed_symbols: set[sp.Symbol],
    expression_name: str,
) -> None:
    """Reject symbolic constructs inappropriate for numerical evaluation."""

    unsupported_atoms = (
        AppliedUndef,
        sp.Derivative,
        sp.Integral,
        sp.Sum,
        sp.Product,
        sp.Limit,
    )

    if expression.has(*unsupported_atoms):
        raise ValueError(
            f"{expression_name} contains unsupported symbolic operations."
        )

    if isinstance(expression, Relational) or expression.has(Relational):
        raise ValueError(
            f"{expression_name} must be an expression, not an equation or inequality."
        )

    unexpected_symbols = expression.free_symbols.difference(allowed_symbols)
    if unexpected_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unexpected_symbols))
        raise ValueError(
            f"{expression_name} contains unsupported variable(s): {names}."
        )

    if expression.has(sp.zoo, sp.oo, -sp.oo, sp.nan):
        raise ValueError(f"{expression_name} contains a non-finite symbolic value.")


def parse_ode_expression(
    ode_text: str,
    ode_order: int,
) -> tuple[sp.Expr, Callable[..., Any]]:
    """Parse the highest derivative equation for order one, two, or three."""

    if ode_order not in {1, 2, 3}:
        raise ValueError("The ODE order must be 1, 2, or 3.")

    if not isinstance(ode_text, str) or not ode_text.strip():
        raise ValueError("The differential-equation expression cannot be empty.")

    text = ode_text.strip().replace("^", "**")
    if "=" in text:
        raise ValueError(
            "Enter only the right-hand side F(...), without an equals sign."
        )

    try:
        expression = sp.sympify(
            text,
            locals=ALLOWED_FUNCTION_NAMES,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The differential equation has an invalid format. "
            "Examples: x + y, -y, -yp, or x + y + yp + ypp."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError("The differential equation could not be interpreted.")

    allowed_symbols = {
        1: {X_SYMBOL, Y_SYMBOL},
        2: {X_SYMBOL, Y_SYMBOL, YP_SYMBOL},
        3: {X_SYMBOL, Y_SYMBOL, YP_SYMBOL, YPP_SYMBOL},
    }[ode_order]

    _reject_unsupported_symbolic_constructs(
        expression,
        allowed_symbols=allowed_symbols,
        expression_name="The differential equation",
    )

    arguments = {
        1: (X_SYMBOL, Y_SYMBOL),
        2: (X_SYMBOL, Y_SYMBOL, YP_SYMBOL),
        3: (X_SYMBOL, Y_SYMBOL, YP_SYMBOL, YPP_SYMBOL),
    }[ode_order]

    try:
        numerical_function = sp.lambdify(
            arguments,
            expression,
            modules=["numpy"],
        )
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The differential equation could not be converted to a numerical function."
        ) from error

    return expression, numerical_function


def parse_exact_solution(
    exact_solution_text: str,
) -> tuple[sp.Expr | None, Callable[[float], Any] | None]:
    """Parse an optional exact solution y(x)."""

    if not str(exact_solution_text).strip():
        return None, None

    text = str(exact_solution_text).strip().replace("^", "**")
    if "=" in text:
        left, right = text.split("=", maxsplit=1)
        if left.strip().lower() not in {"y", "y(x)"}:
            raise ValueError(
                "Enter the exact solution as an expression or as y = expression."
            )
        text = right.strip()

    try:
        expression = sp.sympify(
            text,
            locals=EXACT_SOLUTION_NAMES,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The exact solution has an invalid format. "
            "Example: 2*exp(x) - x - 1 or cos(x)."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError("The exact solution could not be interpreted.")

    _reject_unsupported_symbolic_constructs(
        expression,
        allowed_symbols={X_SYMBOL},
        expression_name="The exact solution",
    )

    try:
        numerical_function = sp.lambdify(
            X_SYMBOL,
            expression,
            modules=["numpy"],
        )
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The exact solution could not be converted to a numerical function."
        ) from error

    return expression, numerical_function


def evaluate_exact_function(
    numerical_function: Callable[[float], Any],
    x_value: float,
    description: str,
) -> float:
    """Evaluate an exact solution or derivative safely."""

    try:
        with np.errstate(all="raise"):
            raw_value = numerical_function(x_value)
    except (
        ValueError,
        TypeError,
        OverflowError,
        ZeroDivisionError,
        FloatingPointError,
    ) as error:
        raise ValueError(
            f"{description} is undefined at x = {x_value:.12g}. Reason: {error}"
        ) from error

    return validate_finite_real(raw_value, description)


def evaluate_highest_derivative(
    numerical_function: Callable[..., Any],
    ode_order: int,
    x_value: float,
    y_value: float,
    yp_value: float | None,
    ypp_value: float | None,
) -> float:
    """Evaluate F at one state of the converted first-order system."""

    try:
        with np.errstate(all="raise"):
            if ode_order == 1:
                raw_value = numerical_function(x_value, y_value)
            elif ode_order == 2:
                if yp_value is None:
                    raise ValueError("The state y′ is missing.")
                raw_value = numerical_function(x_value, y_value, yp_value)
            else:
                if yp_value is None or ypp_value is None:
                    raise ValueError("The states y′ and y″ are required.")
                raw_value = numerical_function(
                    x_value,
                    y_value,
                    yp_value,
                    ypp_value,
                )
    except (
        ValueError,
        TypeError,
        OverflowError,
        ZeroDivisionError,
        FloatingPointError,
    ) as error:
        raise ValueError(
            f"The differential equation is undefined at x = {x_value:.12g}. "
            f"Reason: {error}"
        ) from error

    return validate_finite_real(raw_value, f"F at x = {x_value:.12g}")


def build_exact_functions(
    exact_expression: sp.Expr | None,
) -> tuple[
    Callable[[float], Any] | None,
    Callable[[float], Any] | None,
    Callable[[float], Any] | None,
    Callable[[float], Any] | None,
]:
    """Create exact y, y′, y″, and y‴ functions."""

    if exact_expression is None:
        return None, None, None, None

    expressions = [
        exact_expression,
        sp.diff(exact_expression, X_SYMBOL, 1),
        sp.diff(exact_expression, X_SYMBOL, 2),
        sp.diff(exact_expression, X_SYMBOL, 3),
    ]
    functions = tuple(
        sp.lambdify(X_SYMBOL, expression, modules=["numpy"])
        for expression in expressions
    )
    return functions  # type: ignore[return-value]


def verify_exact_solution(
    ode_order: int,
    ode_expression: sp.Expr,
    exact_expression: sp.Expr,
    x0: float,
    y0: float,
    yp0: float | None,
    ypp0: float | None,
    x_end: float,
) -> None:
    """Verify all required initial conditions and the ODE residual."""

    exact_functions = build_exact_functions(exact_expression)
    exact_y_function = exact_functions[0]
    exact_yp_function = exact_functions[1]
    exact_ypp_function = exact_functions[2]
    assert exact_y_function is not None

    exact_y0 = evaluate_exact_function(exact_y_function, x0, "Exact y(x0)")
    if abs(exact_y0 - y0) > 1.0e-8 * max(1.0, abs(exact_y0), abs(y0)):
        raise ValueError(
            "The exact solution does not satisfy the supplied y(x0) value."
        )

    exact_yp_expression = sp.diff(exact_expression, X_SYMBOL, 1)
    exact_ypp_expression = sp.diff(exact_expression, X_SYMBOL, 2)

    if ode_order >= 2:
        if yp0 is None or exact_yp_function is None:
            raise ValueError("The initial value y′(x0) is required.")
        exact_yp0 = evaluate_exact_function(
            exact_yp_function,
            x0,
            "Exact y′(x0)",
        )
        if abs(exact_yp0 - yp0) > 1.0e-8 * max(
            1.0,
            abs(exact_yp0),
            abs(yp0),
        ):
            raise ValueError(
                "The exact solution does not satisfy the supplied y′(x0) value."
            )

    if ode_order >= 3:
        if ypp0 is None or exact_ypp_function is None:
            raise ValueError("The initial value y″(x0) is required.")
        exact_ypp0 = evaluate_exact_function(
            exact_ypp_function,
            x0,
            "Exact y″(x0)",
        )
        if abs(exact_ypp0 - ypp0) > 1.0e-8 * max(
            1.0,
            abs(exact_ypp0),
            abs(ypp0),
        ):
            raise ValueError(
                "The exact solution does not satisfy the supplied y″(x0) value."
            )

    exact_highest_expression = sp.diff(
        exact_expression,
        X_SYMBOL,
        ode_order,
    )
    substituted_rhs = ode_expression.subs(
        {
            Y_SYMBOL: exact_expression,
            YP_SYMBOL: exact_yp_expression,
            YPP_SYMBOL: exact_ypp_expression,
        }
    )
    residual_expression = sp.simplify(
        exact_highest_expression - substituted_rhs
    )

    if residual_expression == 0:
        return

    residual_function = sp.lambdify(
        X_SYMBOL,
        residual_expression,
        modules=["numpy"],
    )
    residual_values = [
        abs(
            evaluate_exact_function(
                residual_function,
                float(point),
                "Exact-solution residual",
            )
        )
        for point in np.linspace(x0, x_end, 7)
    ]

    if max(residual_values) > 1.0e-7:
        raise ValueError(
            "The supplied exact solution does not satisfy the differential equation."
        )





# =============================================================================
# RK4 algorithm
# =============================================================================
def evaluate_system_rhs(
    ode_function: Callable[..., Any],
    ode_order: int,
    x_value: float,
    y_value: float,
    yp_value: float | None,
    ypp_value: float | None,
) -> tuple[float, float | None, float | None]:
    """Evaluate the equivalent first-order system U′ = G(x, U)."""

    highest_derivative = evaluate_highest_derivative(
        numerical_function=ode_function,
        ode_order=ode_order,
        x_value=x_value,
        y_value=y_value,
        yp_value=yp_value,
        ypp_value=ypp_value,
    )

    if ode_order == 1:
        return highest_derivative, None, None

    if yp_value is None:
        raise ValueError("The state y′ is required for this ODE order.")

    if ode_order == 2:
        return float(yp_value), highest_derivative, None

    if ypp_value is None:
        raise ValueError("The state y″ is required for a third-order ODE.")

    return float(yp_value), float(ypp_value), highest_derivative


def state_to_array(
    ode_order: int,
    y_value: float,
    yp_value: float | None,
    ypp_value: float | None,
) -> np.ndarray:
    """Create a state vector of the required length."""

    if ode_order == 1:
        values = [y_value]
    elif ode_order == 2:
        if yp_value is None:
            raise ValueError("The state y′ is missing.")
        values = [y_value, yp_value]
    else:
        if yp_value is None or ypp_value is None:
            raise ValueError("The states y′ and y″ are missing.")
        values = [y_value, yp_value, ypp_value]

    state = np.asarray(values, dtype=float)

    if not np.all(np.isfinite(state)):
        raise ValueError("The ODE state contains NaN or infinity.")

    return state


def unpack_state(
    ode_order: int,
    state: np.ndarray,
) -> tuple[float, float | None, float | None]:
    """Return y, y′, and y″ from a state vector."""

    y_value = float(state[0])
    yp_value = float(state[1]) if ode_order >= 2 else None
    ypp_value = float(state[2]) if ode_order >= 3 else None
    return y_value, yp_value, ypp_value


def rhs_to_array(
    ode_order: int,
    rhs: tuple[float, float | None, float | None],
) -> np.ndarray:
    """Convert system derivatives to a dense numerical vector."""

    values: list[float] = [float(rhs[0])]

    if ode_order >= 2:
        if rhs[1] is None:
            raise ValueError("The derivative of y′ is missing.")
        values.append(float(rhs[1]))

    if ode_order >= 3:
        if rhs[2] is None:
            raise ValueError("The derivative of y″ is missing.")
        values.append(float(rhs[2]))

    array = np.asarray(values, dtype=float)

    if not np.all(np.isfinite(array)):
        raise ValueError("The system derivative contains NaN or infinity.")

    return array


def exact_state_at(
    exact_functions: tuple[
        Callable[[float], Any] | None,
        Callable[[float], Any] | None,
        Callable[[float], Any] | None,
        Callable[[float], Any] | None,
    ],
    ode_order: int,
    x_value: float,
) -> np.ndarray:
    """Evaluate the exact state [y, y′, y″] at one point."""

    values: list[float] = []

    for component_index in range(ode_order):
        function = exact_functions[component_index]

        if function is None:
            raise ValueError("An exact-state derivative is unavailable.")

        labels = ["Exact y", "Exact y′", "Exact y″"]
        values.append(
            evaluate_exact_function(
                function,
                x_value,
                labels[component_index],
            )
        )

    return np.asarray(values, dtype=float)


def safe_relative_percent(
    absolute_error: float | None,
    exact_value: float | None,
) -> float | None:
    """Return absolute relative error as a percentage when defined."""

    if absolute_error is None or exact_value is None:
        return None

    if abs(exact_value) <= RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
        return None

    return (absolute_error / abs(exact_value)) * 100.0


def component_stage_text(
    label: str,
    current_value: float,
    step_size: float,
    k1_value: float,
    stage2_value: float,
    k2_value: float,
    stage3_value: float,
    k3_value: float,
    stage4_value: float,
    k4_value: float,
    weighted_slope_value: float,
    next_value: float,
) -> tuple[str, str]:
    """Build detailed stage and accepted-update descriptions for one component."""

    stages = (
        f"K1_{label} = {k1_value:.12g}; "
        f"{label}^(2) = {current_value:.12g} + "
        f"({step_size:.12g}/2)({k1_value:.12g}) = {stage2_value:.12g}; "
        f"K2_{label} = {k2_value:.12g}; "
        f"{label}^(3) = {current_value:.12g} + "
        f"({step_size:.12g}/2)({k2_value:.12g}) = {stage3_value:.12g}; "
        f"K3_{label} = {k3_value:.12g}; "
        f"{label}^(4) = {current_value:.12g} + "
        f"({step_size:.12g})({k3_value:.12g}) = {stage4_value:.12g}; "
        f"K4_{label} = {k4_value:.12g}"
    )
    update = (
        f"weighted_{label} = ({k1_value:.12g} + 2({k2_value:.12g}) + "
        f"2({k3_value:.12g}) + {k4_value:.12g})/6 = "
        f"{weighted_slope_value:.12g}; "
        f"{label}_next = {current_value:.12g} + "
        f"({step_size:.12g})({weighted_slope_value:.12g}) = {next_value:.12g}"
    )
    return stages, update


def run_rk4_grid(
    ode_function: Callable[..., Any],
    ode_order: int,
    x0: float,
    y0: float,
    yp0: float | None,
    ypp0: float | None,
    x_end: float,
    steps: int,
    exact_expression: sp.Expr | None = None,
    store_history: bool = True,
) -> tuple[
    tuple[float, float | None, float | None],
    tuple[RK4Iteration, ...],
]:
    """Run classical fourth-order Runge–Kutta on the equivalent first-order system.

    For U′ = G(x,U):

        K1 = G(x_n, U_n)
        K2 = G(x_n+h/2, U_n+h K1/2)
        K3 = G(x_n+h/2, U_n+h K2/2)
        K4 = G(x_n+h, U_n+h K3)
        U_(n+1) = U_n + h(K1+2K2+2K3+K4)/6
    """

    if steps < 1:
        raise ValueError("The number of steps must be at least 1.")

    step_size = (x_end - x0) / steps
    if not math.isfinite(step_size) or abs(step_size) <= ZERO_TOLERANCE:
        raise ValueError("The computed step size is zero or numerically unusable.")

    current_x = float(x0)
    current_state = state_to_array(
        ode_order,
        float(y0),
        None if ode_order == 1 else float(yp0),
        None if ode_order < 3 else float(ypp0),
    )

    exact_functions = build_exact_functions(exact_expression)
    history: list[RK4Iteration] = []

    for iteration in range(1, steps + 1):
        current_y, current_yp, current_ypp = unpack_state(
            ode_order,
            current_state,
        )

        k1 = rhs_to_array(
            ode_order,
            evaluate_system_rhs(
                ode_function,
                ode_order,
                current_x,
                current_y,
                current_yp,
                current_ypp,
            ),
        )

        stage2_x = current_x + 0.5 * step_size
        stage2_state = current_state + 0.5 * step_size * k1
        if not np.all(np.isfinite(stage2_state)) or not math.isfinite(stage2_x):
            raise ValueError(
                f"RK4 produced a non-finite second-stage state at iteration {iteration}."
            )
        stage2_y, stage2_yp, stage2_ypp = unpack_state(ode_order, stage2_state)
        k2 = rhs_to_array(
            ode_order,
            evaluate_system_rhs(
                ode_function,
                ode_order,
                stage2_x,
                stage2_y,
                stage2_yp,
                stage2_ypp,
            ),
        )

        stage3_x = current_x + 0.5 * step_size
        stage3_state = current_state + 0.5 * step_size * k2
        if not np.all(np.isfinite(stage3_state)) or not math.isfinite(stage3_x):
            raise ValueError(
                f"RK4 produced a non-finite third-stage state at iteration {iteration}."
            )
        stage3_y, stage3_yp, stage3_ypp = unpack_state(ode_order, stage3_state)
        k3 = rhs_to_array(
            ode_order,
            evaluate_system_rhs(
                ode_function,
                ode_order,
                stage3_x,
                stage3_y,
                stage3_yp,
                stage3_ypp,
            ),
        )

        stage4_x = current_x + step_size
        stage4_state = current_state + step_size * k3
        if not np.all(np.isfinite(stage4_state)) or not math.isfinite(stage4_x):
            raise ValueError(
                f"RK4 produced a non-finite fourth-stage state at iteration {iteration}."
            )
        stage4_y, stage4_yp, stage4_ypp = unpack_state(ode_order, stage4_state)
        k4 = rhs_to_array(
            ode_order,
            evaluate_system_rhs(
                ode_function,
                ode_order,
                stage4_x,
                stage4_y,
                stage4_yp,
                stage4_ypp,
            ),
        )

        weighted_slope = (k1 + 2.0 * k2 + 2.0 * k3 + k4) / 6.0
        next_state = current_state + step_size * weighted_slope
        next_x = x0 + iteration * step_size

        if not np.all(np.isfinite(weighted_slope)) or not np.all(np.isfinite(next_state)):
            raise ValueError(
                f"RK4 produced a non-finite accepted state at iteration {iteration}."
            )

        next_y, next_yp, next_ypp = unpack_state(ode_order, next_state)

        exact_next_state: np.ndarray | None = None
        signed_errors: np.ndarray | None = None
        absolute_errors: np.ndarray | None = None
        relative_errors_percent: list[float | None] = [None] * ode_order
        local_truncation_errors: np.ndarray | None = None

        if exact_expression is not None:
            exact_current_state = exact_state_at(
                exact_functions,
                ode_order,
                current_x,
            )
            exact_next_state = exact_state_at(
                exact_functions,
                ode_order,
                next_x,
            )

            signed_errors = next_state - exact_next_state
            absolute_errors = np.abs(signed_errors)
            relative_errors_percent = [
                safe_relative_percent(
                    float(absolute_errors[index]),
                    float(exact_next_state[index]),
                )
                for index in range(ode_order)
            ]

            exact_y, exact_yp, exact_ypp = unpack_state(
                ode_order,
                exact_current_state,
            )
            exact_k1 = rhs_to_array(
                ode_order,
                evaluate_system_rhs(
                    ode_function,
                    ode_order,
                    current_x,
                    exact_y,
                    exact_yp,
                    exact_ypp,
                ),
            )

            exact_stage2_state = exact_current_state + 0.5 * step_size * exact_k1
            exact_stage2_y, exact_stage2_yp, exact_stage2_ypp = unpack_state(
                ode_order,
                exact_stage2_state,
            )
            exact_k2 = rhs_to_array(
                ode_order,
                evaluate_system_rhs(
                    ode_function,
                    ode_order,
                    stage2_x,
                    exact_stage2_y,
                    exact_stage2_yp,
                    exact_stage2_ypp,
                ),
            )

            exact_stage3_state = exact_current_state + 0.5 * step_size * exact_k2
            exact_stage3_y, exact_stage3_yp, exact_stage3_ypp = unpack_state(
                ode_order,
                exact_stage3_state,
            )
            exact_k3 = rhs_to_array(
                ode_order,
                evaluate_system_rhs(
                    ode_function,
                    ode_order,
                    stage3_x,
                    exact_stage3_y,
                    exact_stage3_yp,
                    exact_stage3_ypp,
                ),
            )

            exact_stage4_state = exact_current_state + step_size * exact_k3
            exact_stage4_y, exact_stage4_yp, exact_stage4_ypp = unpack_state(
                ode_order,
                exact_stage4_state,
            )
            exact_k4 = rhs_to_array(
                ode_order,
                evaluate_system_rhs(
                    ode_function,
                    ode_order,
                    stage4_x,
                    exact_stage4_y,
                    exact_stage4_yp,
                    exact_stage4_ypp,
                ),
            )

            exact_one_step_state = exact_current_state + step_size * (
                exact_k1 + 2.0 * exact_k2 + 2.0 * exact_k3 + exact_k4
            ) / 6.0
            local_truncation_errors = exact_next_state - exact_one_step_state

        def optional_component(
            values: np.ndarray | None,
            component_index: int,
        ) -> float | None:
            if values is None or component_index >= len(values):
                return None
            return float(values[component_index])

        component_names = ["y", "yp", "ypp"]
        stage_parts: list[str] = []
        update_parts: list[str] = []
        for component_index in range(ode_order):
            stage_text, update_text = component_stage_text(
                label=component_names[component_index],
                current_value=float(current_state[component_index]),
                step_size=float(step_size),
                k1_value=float(k1[component_index]),
                stage2_value=float(stage2_state[component_index]),
                k2_value=float(k2[component_index]),
                stage3_value=float(stage3_state[component_index]),
                k3_value=float(k3[component_index]),
                stage4_value=float(stage4_state[component_index]),
                k4_value=float(k4[component_index]),
                weighted_slope_value=float(weighted_slope[component_index]),
                next_value=float(next_state[component_index]),
            )
            stage_parts.append(stage_text)
            update_parts.append(update_text)

        if store_history:
            history.append(
                RK4Iteration(
                    iteration=iteration,
                    x_n=float(current_x),
                    y_n=current_y,
                    yp_n=current_yp,
                    ypp_n=current_ypp,
                    step_size=float(step_size),

                    k1_y=float(k1[0]),
                    k1_yp=float(k1[1]) if ode_order >= 2 else None,
                    k1_ypp=float(k1[2]) if ode_order >= 3 else None,

                    stage2_x=float(stage2_x),
                    stage2_y=stage2_y,
                    stage2_yp=stage2_yp,
                    stage2_ypp=stage2_ypp,
                    k2_y=float(k2[0]),
                    k2_yp=float(k2[1]) if ode_order >= 2 else None,
                    k2_ypp=float(k2[2]) if ode_order >= 3 else None,

                    stage3_x=float(stage3_x),
                    stage3_y=stage3_y,
                    stage3_yp=stage3_yp,
                    stage3_ypp=stage3_ypp,
                    k3_y=float(k3[0]),
                    k3_yp=float(k3[1]) if ode_order >= 2 else None,
                    k3_ypp=float(k3[2]) if ode_order >= 3 else None,

                    stage4_x=float(stage4_x),
                    stage4_y=stage4_y,
                    stage4_yp=stage4_yp,
                    stage4_ypp=stage4_ypp,
                    k4_y=float(k4[0]),
                    k4_yp=float(k4[1]) if ode_order >= 2 else None,
                    k4_ypp=float(k4[2]) if ode_order >= 3 else None,

                    weighted_slope_y=float(weighted_slope[0]),
                    weighted_slope_yp=(
                        float(weighted_slope[1]) if ode_order >= 2 else None
                    ),
                    weighted_slope_ypp=(
                        float(weighted_slope[2]) if ode_order >= 3 else None
                    ),

                    x_next=float(next_x),
                    y_next=next_y,
                    yp_next=next_yp,
                    ypp_next=next_ypp,

                    exact_y_next=optional_component(exact_next_state, 0),
                    exact_yp_next=optional_component(exact_next_state, 1),
                    exact_ypp_next=optional_component(exact_next_state, 2),

                    signed_error_y=optional_component(signed_errors, 0),
                    signed_error_yp=optional_component(signed_errors, 1),
                    signed_error_ypp=optional_component(signed_errors, 2),

                    absolute_error_y=optional_component(absolute_errors, 0),
                    absolute_error_yp=optional_component(absolute_errors, 1),
                    absolute_error_ypp=optional_component(absolute_errors, 2),

                    relative_error_y_percent=relative_errors_percent[0],
                    relative_error_yp_percent=(
                        relative_errors_percent[1] if ode_order >= 2 else None
                    ),
                    relative_error_ypp_percent=(
                        relative_errors_percent[2] if ode_order >= 3 else None
                    ),

                    local_truncation_error_y=optional_component(
                        local_truncation_errors, 0
                    ),
                    local_truncation_error_yp=optional_component(
                        local_truncation_errors, 1
                    ),
                    local_truncation_error_ypp=optional_component(
                        local_truncation_errors, 2
                    ),

                    stage_operation="; ".join(stage_parts),
                    update_operation="; ".join(update_parts),
                    status="Completed",
                )
            )

        current_x = float(next_x)
        current_state = next_state

    final_y, final_yp, final_ypp = unpack_state(ode_order, current_state)
    return (final_y, final_yp, final_ypp), tuple(history)


def build_convergence_analysis(
    ode_function: Callable[..., Any],
    ode_order: int,
    exact_expression: sp.Expr | None,
    x0: float,
    y0: float,
    yp0: float | None,
    ypp0: float | None,
    x_end: float,
    base_steps: int,
    levels: int,
) -> tuple[ConvergenceRecord, ...]:
    """Refine the grid and estimate classical RK4 fourth-order convergence."""

    raw_records: list[dict[str, Any]] = []
    exact_functions = build_exact_functions(exact_expression)

    exact_final_state = (
        exact_state_at(
            exact_functions,
            ode_order,
            x_end,
        )
        if exact_expression is not None
        else None
    )

    for level in range(1, levels + 1):
        refined_steps = base_steps * (2 ** (level - 1))

        final_state_tuple, _ = run_rk4_grid(
            ode_function=ode_function,
            ode_order=ode_order,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            steps=refined_steps,
            exact_expression=None,
            store_history=False,
        )
        final_state = state_to_array(
            ode_order,
            final_state_tuple[0],
            final_state_tuple[1],
            final_state_tuple[2],
        )

        absolute_errors = (
            np.abs(final_state - exact_final_state)
            if exact_final_state is not None
            else None
        )

        raw_records.append(
            {
                "level": level,
                "steps": refined_steps,
                "step_size": (x_end - x0) / refined_steps,

                "final_y": float(final_state[0]),
                "final_yp": (
                    float(final_state[1])
                    if ode_order >= 2
                    else None
                ),
                "final_ypp": (
                    float(final_state[2])
                    if ode_order >= 3
                    else None
                ),

                "exact_final_y": (
                    float(exact_final_state[0])
                    if exact_final_state is not None
                    else None
                ),
                "exact_final_yp": (
                    float(exact_final_state[1])
                    if exact_final_state is not None
                    and ode_order >= 2
                    else None
                ),
                "exact_final_ypp": (
                    float(exact_final_state[2])
                    if exact_final_state is not None
                    and ode_order >= 3
                    else None
                ),

                "absolute_error_y": (
                    float(absolute_errors[0])
                    if absolute_errors is not None
                    else None
                ),
                "absolute_error_yp": (
                    float(absolute_errors[1])
                    if absolute_errors is not None
                    and ode_order >= 2
                    else None
                ),
                "absolute_error_ypp": (
                    float(absolute_errors[2])
                    if absolute_errors is not None
                    and ode_order >= 3
                    else None
                ),

                "successive_difference_y": None,
                "observed_order": None,
            }
        )

    for index in range(1, len(raw_records)):
        raw_records[index]["successive_difference_y"] = abs(
            float(raw_records[index]["final_y"])
            - float(raw_records[index - 1]["final_y"])
        )

    if exact_final_state is not None:
        for index in range(1, len(raw_records)):
            previous_error = raw_records[index - 1]["absolute_error_y"]
            current_error = raw_records[index]["absolute_error_y"]

            if (
                previous_error is not None
                and current_error is not None
                and float(previous_error) > ZERO_TOLERANCE
                and float(current_error) > ZERO_TOLERANCE
            ):
                raw_records[index]["observed_order"] = math.log(
                    float(previous_error) / float(current_error),
                    2.0,
                )
    else:
        for index in range(2, len(raw_records)):
            previous_difference = raw_records[index - 1][
                "successive_difference_y"
            ]
            current_difference = raw_records[index][
                "successive_difference_y"
            ]

            if (
                previous_difference is not None
                and current_difference is not None
                and float(previous_difference) > ZERO_TOLERANCE
                and float(current_difference) > ZERO_TOLERANCE
            ):
                raw_records[index]["observed_order"] = math.log(
                    float(previous_difference) / float(current_difference),
                    2.0,
                )

    return tuple(
        ConvergenceRecord(**record)
        for record in raw_records
    )


def error_result(
    message: str,
    ode_order: int = 1,
    input_signature: str = "",
) -> RK4Result:
    """Create a consistent failed result."""

    return RK4Result(
        status="error",
        success=False,
        method=METHOD_NAME,
        ode_order=ode_order,
        message=message,
        stopping_reason=(
            "Execution stopped because input validation or numerical "
            "evaluation failed."
        ),
        ode_text="",
        ode_expression=None,
        exact_solution_text="",
        exact_solution_expression=None,
        x0=None,
        y0=None,
        yp0=None,
        ypp0=None,
        x_end=None,
        steps=0,
        step_size=None,
        direction="Not available",
        iterations=(),
        final_x=None,
        final_y=None,
        final_yp=None,
        final_ypp=None,
        exact_final_y=None,
        exact_final_yp=None,
        exact_final_ypp=None,
        signed_final_error_y=None,
        signed_final_error_yp=None,
        signed_final_error_ypp=None,
        absolute_final_error_y=None,
        absolute_final_error_yp=None,
        absolute_final_error_ypp=None,
        relative_final_error_y_percent=None,
        relative_final_error_yp_percent=None,
        relative_final_error_ypp_percent=None,
        maximum_absolute_error_y=None,
        rmse_y=None,
        convergence_records=(),
        latest_observed_order=None,
        warnings=(),
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


def solve_rk4_method(
    ode_order_input: Any,
    ode_text: str,
    x0_input: Any,
    y0_input: Any,
    yp0_input: Any = None,
    ypp0_input: Any = None,
    x_end_input: Any = DEFAULT_X_END,
    steps_input: Any = DEFAULT_STEPS,
    exact_solution_text: str = "",
    convergence_levels_input: Any = DEFAULT_CONVERGENCE_LEVELS,
) -> RK4Result:
    """Validate inputs, solve the IVP, and analyze RK4 errors and convergence."""

    ode_order = 1
    input_signature = ""

    try:
        try:
            ode_order_value = float(ode_order_input)
        except (TypeError, ValueError, OverflowError) as error:
            raise ValueError("The ODE order must be an integer 1, 2, or 3.") from error
        if not math.isfinite(ode_order_value) or not ode_order_value.is_integer():
            raise ValueError("The ODE order must be an integer 1, 2, or 3.")
        ode_order = int(ode_order_value)
        if ode_order not in {1, 2, 3}:
            raise ValueError("The ODE order must be 1, 2, or 3.")

        x0 = validate_finite_real(x0_input, "Initial x value x0")
        y0 = validate_finite_real(y0_input, "Initial value y(x0)")
        x_end = validate_finite_real(x_end_input, "Final x value")

        yp0 = (
            validate_finite_real(yp0_input, "Initial value y′(x0)")
            if ode_order >= 2
            else None
        )
        ypp0 = (
            validate_finite_real(ypp0_input, "Initial value y″(x0)")
            if ode_order >= 3
            else None
        )

        try:
            steps_value = float(steps_input)
        except (TypeError, ValueError, OverflowError) as error:
            raise ValueError("The number of steps must be an integer.") from error
        if not math.isfinite(steps_value) or not steps_value.is_integer():
            raise ValueError("The number of steps must be an integer.")
        steps = int(steps_value)

        try:
            levels_value = float(convergence_levels_input)
        except (TypeError, ValueError, OverflowError) as error:
            raise ValueError("Convergence levels must be an integer.") from error
        if not math.isfinite(levels_value) or not levels_value.is_integer():
            raise ValueError("Convergence levels must be an integer.")
        convergence_levels = int(levels_value)

        if not MIN_STEPS <= steps <= MAX_STEPS:
            raise ValueError(
                f"The number of steps must be between {MIN_STEPS} and {MAX_STEPS}."
            )
        if not MIN_CONVERGENCE_LEVELS <= convergence_levels <= MAX_CONVERGENCE_LEVELS:
            raise ValueError(
                "Convergence levels must be between "
                f"{MIN_CONVERGENCE_LEVELS} and {MAX_CONVERGENCE_LEVELS}."
            )

        refined_steps = steps * (2 ** (convergence_levels - 1))
        if refined_steps > MAX_REFINED_STEPS:
            raise ValueError(
                "The requested convergence analysis is too large. "
                "Reduce the number of steps or convergence levels."
            )
        if abs(x_end - x0) <= ZERO_TOLERANCE:
            raise ValueError("The final x value must be different from x0.")

        input_signature = create_input_signature(
            ode_order=ode_order,
            ode_text=ode_text,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            steps=steps,
            exact_solution_text=exact_solution_text,
            convergence_levels=convergence_levels,
        )

        ode_expression, ode_function = parse_ode_expression(ode_text, ode_order)
        exact_expression, _ = parse_exact_solution(exact_solution_text)

        warnings: list[str] = []
        if exact_expression is not None:
            verify_exact_solution(
                ode_order=ode_order,
                ode_expression=ode_expression,
                exact_expression=exact_expression,
                x0=x0,
                y0=y0,
                yp0=yp0,
                ypp0=ypp0,
                x_end=x_end,
            )

        final_state, iterations = run_rk4_grid(
            ode_function=ode_function,
            ode_order=ode_order,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            steps=steps,
            exact_expression=exact_expression,
            store_history=True,
        )
        final_y, final_yp, final_ypp = final_state

        exact_final_y: float | None = None
        exact_final_yp: float | None = None
        exact_final_ypp: float | None = None
        signed_final_error_y: float | None = None
        signed_final_error_yp: float | None = None
        signed_final_error_ypp: float | None = None
        absolute_final_error_y: float | None = None
        absolute_final_error_yp: float | None = None
        absolute_final_error_ypp: float | None = None
        relative_final_error_y_percent: float | None = None
        relative_final_error_yp_percent: float | None = None
        relative_final_error_ypp_percent: float | None = None
        maximum_absolute_error_y: float | None = None
        rmse_y: float | None = None

        exact_functions = build_exact_functions(exact_expression)
        if exact_expression is not None:
            exact_final_state = exact_state_at(exact_functions, ode_order, x_end)
            numerical_final_state = state_to_array(
                ode_order, final_y, final_yp, final_ypp
            )
            signed_final_errors = numerical_final_state - exact_final_state
            absolute_final_errors = np.abs(signed_final_errors)

            exact_final_y = float(exact_final_state[0])
            signed_final_error_y = float(signed_final_errors[0])
            absolute_final_error_y = float(absolute_final_errors[0])
            relative_final_error_y_percent = safe_relative_percent(
                absolute_final_error_y, exact_final_y
            )

            if ode_order >= 2:
                exact_final_yp = float(exact_final_state[1])
                signed_final_error_yp = float(signed_final_errors[1])
                absolute_final_error_yp = float(absolute_final_errors[1])
                relative_final_error_yp_percent = safe_relative_percent(
                    absolute_final_error_yp, exact_final_yp
                )

            if ode_order >= 3:
                exact_final_ypp = float(exact_final_state[2])
                signed_final_error_ypp = float(signed_final_errors[2])
                absolute_final_error_ypp = float(absolute_final_errors[2])
                relative_final_error_ypp_percent = safe_relative_percent(
                    absolute_final_error_ypp, exact_final_ypp
                )

            y_errors = np.asarray(
                [
                    record.signed_error_y
                    for record in iterations
                    if record.signed_error_y is not None
                ],
                dtype=float,
            )
            if y_errors.size:
                maximum_absolute_error_y = float(np.max(np.abs(y_errors)))
                rmse_y = float(np.sqrt(np.mean(y_errors**2)))

        convergence_records = build_convergence_analysis(
            ode_function=ode_function,
            ode_order=ode_order,
            exact_expression=exact_expression,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            base_steps=steps,
            levels=convergence_levels,
        )
        observed_orders = [
            record.observed_order
            for record in convergence_records
            if record.observed_order is not None and math.isfinite(record.observed_order)
        ]
        latest_observed_order = observed_orders[-1] if observed_orders else None

        accepted_magnitudes: list[float] = []
        stage_magnitudes: list[float] = []
        for record in iterations:
            for value in (record.y_next, record.yp_next, record.ypp_next):
                if value is not None:
                    accepted_magnitudes.append(abs(value))
            for value in (
                record.stage2_y,
                record.stage2_yp,
                record.stage2_ypp,
                record.stage3_y,
                record.stage3_yp,
                record.stage3_ypp,
                record.stage4_y,
                record.stage4_yp,
                record.stage4_ypp,
            ):
                if value is not None:
                    stage_magnitudes.append(abs(value))

        if accepted_magnitudes and max(accepted_magnitudes) >= VALUE_MAGNITUDE_WARNING:
            warnings.append(
                "The accepted numerical state reached a very large magnitude. "
                "Classical RK4 may be unstable for this step size or problem."
            )
        if stage_magnitudes and max(stage_magnitudes) >= VALUE_MAGNITUDE_WARNING:
            warnings.append(
                "An intermediate RK4 stage reached a very large magnitude. "
                "Reduce the step size and inspect stability."
            )
        if latest_observed_order is not None and latest_observed_order < 3.0:
            warnings.append(
                "The observed convergence is weaker than the expected fourth-order "
                "behavior. Reduce the step size and inspect stability, smoothness, "
                "or floating-point round-off."
            )
        if exact_expression is None:
            warnings.append(
                "No exact solution was supplied. True global errors are unavailable; "
                "successive final-value differences are used for convergence."
            )

        direction = "Forward integration" if x_end > x0 else "Backward integration"
        step_size = (x_end - x0) / steps

        return RK4Result(
            status="success",
            success=True,
            method=METHOD_NAME,
            ode_order=ode_order,
            message="Execution completed successfully.",
            stopping_reason=(
                "The requested number of uniform classical RK4 steps was completed "
                "and the final grid point was reached."
            ),
            ode_text=str(ode_text).strip(),
            ode_expression=ode_expression,
            exact_solution_text=str(exact_solution_text).strip(),
            exact_solution_expression=exact_expression,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            steps=steps,
            step_size=step_size,
            direction=direction,
            iterations=iterations,
            final_x=x_end,
            final_y=final_y,
            final_yp=final_yp,
            final_ypp=final_ypp,
            exact_final_y=exact_final_y,
            exact_final_yp=exact_final_yp,
            exact_final_ypp=exact_final_ypp,
            signed_final_error_y=signed_final_error_y,
            signed_final_error_yp=signed_final_error_yp,
            signed_final_error_ypp=signed_final_error_ypp,
            absolute_final_error_y=absolute_final_error_y,
            absolute_final_error_yp=absolute_final_error_yp,
            absolute_final_error_ypp=absolute_final_error_ypp,
            relative_final_error_y_percent=relative_final_error_y_percent,
            relative_final_error_yp_percent=relative_final_error_yp_percent,
            relative_final_error_ypp_percent=relative_final_error_ypp_percent,
            maximum_absolute_error_y=maximum_absolute_error_y,
            rmse_y=rmse_y,
            convergence_records=convergence_records,
            latest_observed_order=latest_observed_order,
            warnings=tuple(warnings),
            input_signature=input_signature,
            execution_datetime=current_report_datetime(),
        )

    except (ValueError, TypeError, ArithmeticError, OverflowError) as error:
        return error_result(
            message=str(error),
            ode_order=ode_order,
            input_signature=input_signature,
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def iterations_dataframe(
    result: RK4Result,
) -> pd.DataFrame:
    """Return the complete vector RK4 stage and update table."""

    rows: list[dict[str, Any]] = []
    for record in result.iterations:
        row: dict[str, Any] = {
            "Iteration": record.iteration,
            "x_n": record.x_n,
            "y_n": record.y_n,
            "h": record.step_size,
            "K1_y": record.k1_y,
            "Stage 2 x": record.stage2_x,
            "Stage 2 y": record.stage2_y,
            "K2_y": record.k2_y,
            "Stage 3 x": record.stage3_x,
            "Stage 3 y": record.stage3_y,
            "K3_y": record.k3_y,
            "Stage 4 x": record.stage4_x,
            "Stage 4 y": record.stage4_y,
            "K4_y": record.k4_y,
            "Weighted Slope y": record.weighted_slope_y,
            "x_(n+1)": record.x_next,
            "y_(n+1)": record.y_next,
            "Exact y_(n+1)": record.exact_y_next,
            "Signed Global Error y": record.signed_error_y,
            "Absolute Global Error y": record.absolute_error_y,
            "Relative Error y (%)": record.relative_error_y_percent,
            "Local Truncation Error y": record.local_truncation_error_y,
            "Stage Operations": record.stage_operation,
            "Accepted Update": record.update_operation,
            "Status": record.status,
        }

        if result.ode_order >= 2:
            row.update(
                {
                    "yp_n": record.yp_n,
                    "K1_yp": record.k1_yp,
                    "Stage 2 yp": record.stage2_yp,
                    "K2_yp": record.k2_yp,
                    "Stage 3 yp": record.stage3_yp,
                    "K3_yp": record.k3_yp,
                    "Stage 4 yp": record.stage4_yp,
                    "K4_yp": record.k4_yp,
                    "Weighted Slope yp": record.weighted_slope_yp,
                    "yp_(n+1)": record.yp_next,
                    "Exact yp_(n+1)": record.exact_yp_next,
                    "Signed Global Error yp": record.signed_error_yp,
                    "Absolute Global Error yp": record.absolute_error_yp,
                    "Relative Error yp (%)": record.relative_error_yp_percent,
                    "Local Truncation Error yp": record.local_truncation_error_yp,
                }
            )

        if result.ode_order >= 3:
            row.update(
                {
                    "ypp_n": record.ypp_n,
                    "K1_ypp": record.k1_ypp,
                    "Stage 2 ypp": record.stage2_ypp,
                    "K2_ypp": record.k2_ypp,
                    "Stage 3 ypp": record.stage3_ypp,
                    "K3_ypp": record.k3_ypp,
                    "Stage 4 ypp": record.stage4_ypp,
                    "K4_ypp": record.k4_ypp,
                    "Weighted Slope ypp": record.weighted_slope_ypp,
                    "ypp_(n+1)": record.ypp_next,
                    "Exact ypp_(n+1)": record.exact_ypp_next,
                    "Signed Global Error ypp": record.signed_error_ypp,
                    "Absolute Global Error ypp": record.absolute_error_ypp,
                    "Relative Error ypp (%)": record.relative_error_ypp_percent,
                    "Local Truncation Error ypp": record.local_truncation_error_ypp,
                }
            )

        rows.append(row)

    return pd.DataFrame(rows)


def solution_values_dataframe(
    result: RK4Result,
) -> pd.DataFrame:
    """Return all state values including the initial condition."""

    exact_functions = build_exact_functions(
        result.exact_solution_expression
    )

    exact_initial_state = (
        exact_state_at(
            exact_functions,
            result.ode_order,
            float(result.x0),
        )
        if result.exact_solution_expression is not None
        else None
    )

    initial_row: dict[str, Any] = {
        "Point": 0,
        "x": result.x0,
        "RK4 y": result.y0,
        "Exact y": (
            float(exact_initial_state[0])
            if exact_initial_state is not None
            else None
        ),
        "Absolute Error y": (
            abs(float(result.y0) - float(exact_initial_state[0]))
            if exact_initial_state is not None
            else None
        ),
    }

    if result.ode_order >= 2:
        initial_row.update(
            {
                "RK4 yp": result.yp0,
                "Exact yp": (
                    float(exact_initial_state[1])
                    if exact_initial_state is not None
                    else None
                ),
                "Absolute Error yp": (
                    abs(
                        float(result.yp0)
                        - float(exact_initial_state[1])
                    )
                    if exact_initial_state is not None
                    else None
                ),
            }
        )

    if result.ode_order >= 3:
        initial_row.update(
            {
                "RK4 ypp": result.ypp0,
                "Exact ypp": (
                    float(exact_initial_state[2])
                    if exact_initial_state is not None
                    else None
                ),
                "Absolute Error ypp": (
                    abs(
                        float(result.ypp0)
                        - float(exact_initial_state[2])
                    )
                    if exact_initial_state is not None
                    else None
                ),
            }
        )

    rows = [initial_row]

    for record in result.iterations:
        row: dict[str, Any] = {
            "Point": record.iteration,
            "x": record.x_next,
            "RK4 y": record.y_next,
            "Exact y": record.exact_y_next,
            "Absolute Error y": record.absolute_error_y,
        }

        if result.ode_order >= 2:
            row.update(
                {
                    "RK4 yp": record.yp_next,
                    "Exact yp": record.exact_yp_next,
                    "Absolute Error yp": record.absolute_error_yp,
                }
            )

        if result.ode_order >= 3:
            row.update(
                {
                    "RK4 ypp": record.ypp_next,
                    "Exact ypp": record.exact_ypp_next,
                    "Absolute Error ypp": record.absolute_error_ypp,
                }
            )

        rows.append(row)

    return pd.DataFrame(rows)


def error_analysis_dataframe(
    result: RK4Result,
) -> pd.DataFrame:
    """Return pointwise state errors when an exact solution is available."""

    if result.exact_solution_expression is None:
        return pd.DataFrame(
            {
                "Message": [
                    "No exact solution was supplied; true global "
                    "errors are unavailable."
                ]
            }
        )

    rows: list[dict[str, Any]] = []

    for record in result.iterations:
        row: dict[str, Any] = {
            "Iteration": record.iteration,
            "x": record.x_next,
            "RK4 y": record.y_next,
            "Exact y": record.exact_y_next,
            "Signed Error y": record.signed_error_y,
            "Absolute Error y": record.absolute_error_y,
            "Relative Error y (%)": record.relative_error_y_percent,
            "Local Truncation Error y": (
                record.local_truncation_error_y
            ),
        }

        if result.ode_order >= 2:
            row.update(
                {
                    "RK4 yp": record.yp_next,
                    "Exact yp": record.exact_yp_next,
                    "Signed Error yp": record.signed_error_yp,
                    "Absolute Error yp": record.absolute_error_yp,
                    "Relative Error yp (%)": (
                        record.relative_error_yp_percent
                    ),
                    "Local Truncation Error yp": (
                        record.local_truncation_error_yp
                    ),
                }
            )

        if result.ode_order >= 3:
            row.update(
                {
                    "RK4 ypp": record.ypp_next,
                    "Exact ypp": record.exact_ypp_next,
                    "Signed Error ypp": record.signed_error_ypp,
                    "Absolute Error ypp": record.absolute_error_ypp,
                    "Relative Error ypp (%)": (
                        record.relative_error_ypp_percent
                    ),
                    "Local Truncation Error ypp": (
                        record.local_truncation_error_ypp
                    ),
                }
            )

        rows.append(row)

    return pd.DataFrame(rows)


def convergence_dataframe(
    result: RK4Result,
) -> pd.DataFrame:
    """Return grid-refinement convergence results."""

    rows: list[dict[str, Any]] = []

    for record in result.convergence_records:
        row: dict[str, Any] = {
            "Level": record.level,
            "Steps": record.steps,
            "h": record.step_size,
            "Final y": record.final_y,
            "Exact Final y": record.exact_final_y,
            "Absolute Error y": record.absolute_error_y,
            "Successive Difference y": record.successive_difference_y,
            "Observed Order": record.observed_order,
        }

        if result.ode_order >= 2:
            row.update(
                {
                    "Final yp": record.final_yp,
                    "Exact Final yp": record.exact_final_yp,
                    "Absolute Error yp": record.absolute_error_yp,
                }
            )

        if result.ode_order >= 3:
            row.update(
                {
                    "Final ypp": record.final_ypp,
                    "Exact Final ypp": record.exact_final_ypp,
                    "Absolute Error ypp": record.absolute_error_ypp,
                }
            )

        rows.append(row)

    return pd.DataFrame(rows)


def method_formula_dataframe(
    result: RK4Result,
) -> pd.DataFrame:
    """Return the classical RK4 and equivalent-system formulas."""

    formulas: list[tuple[str, str]] = [
        (
            "Highest-Order IVP",
            {
                1: "y' = F(x,y), y(x0)=y0",
                2: "y'' = F(x,y,y'), y(x0)=y0, y'(x0)=yp0",
                3: (
                    "y‴ = F(x,y,y',y''), y(x0)=y0, "
                    "y'(x0)=yp0, y''(x0)=ypp0"
                ),
            }[result.ode_order],
        ),
        ("Uniform Step", "h = (x_end - x0) / N"),
        ("System Form", "U' = G(x,U)"),
        ("First Stage", "K1 = G(x_n,U_n)"),
        ("Second Stage", "K2 = G(x_n+h/2,U_n+h*K1/2)"),
        ("Third Stage", "K3 = G(x_n+h/2,U_n+h*K2/2)"),
        ("Fourth Stage", "K4 = G(x_n+h,U_n+h*K3)"),
        (
            "Weighted Slope",
            "K = (K1 + 2*K2 + 2*K3 + K4)/6",
        ),
        ("Accepted Update", "U_(n+1) = U_n + h*K"),
        ("Expected Global Order", "O(h^4)"),
        ("Expected Local Truncation Error", "O(h^5) per step"),
    ]

    if result.ode_order == 2:
        formulas.append(
            (
                "Equivalent System",
                "u1=y, u2=y'; u1'=u2, u2'=F(x,u1,u2)",
            )
        )
    if result.ode_order == 3:
        formulas.append(
            (
                "Equivalent System",
                (
                    "u1=y, u2=y', u3=y''; "
                    "u1'=u2, u2'=u3, u3'=F(x,u1,u2,u3)"
                ),
            )
        )

    return pd.DataFrame(formulas, columns=["Item", "Formula"])


# =============================================================================
# Scientific plots
# =============================================================================
def create_solution_figure(
    result: RK4Result,
) -> Figure:
    """Plot y, RK4 stage states, and the optional exact y solution."""

    dataframe = solution_values_dataframe(result)
    figure, axis = plt.subplots(figsize=(10, 6))

    axis.plot(
        dataframe["x"],
        dataframe["RK4 y"],
        marker="o",
        linewidth=2,
        label="RK4 y",
    )

    stage_x: list[float] = []
    stage_y: list[float] = []
    for record in result.iterations:
        stage_x.extend([record.stage2_x, record.stage3_x, record.stage4_x])
        stage_y.extend([record.stage2_y, record.stage3_y, record.stage4_y])

    if stage_x:
        axis.scatter(
            stage_x,
            stage_y,
            marker="x",
            s=32,
            label="RK4 stage states",
            zorder=4,
        )

    if result.exact_solution_expression is not None:
        exact_function = sp.lambdify(
            X_SYMBOL,
            result.exact_solution_expression,
            modules=["numpy"],
        )
        x_values = np.linspace(float(result.x0), float(result.x_end), 500)
        y_values = np.asarray(
            [
                evaluate_exact_function(exact_function, float(value), "Exact y")
                for value in x_values
            ],
            dtype=float,
        )
        axis.plot(x_values, y_values, linewidth=2, label="Exact y")

    axis.scatter(
        [result.x0],
        [result.y0],
        s=90,
        marker="s",
        label="Initial condition",
        zorder=5,
    )
    axis.scatter(
        [result.final_x],
        [result.final_y],
        s=110,
        marker="*",
        label=f"Final RK4 y = {float(result.final_y):.6g}",
        zorder=6,
    )

    axis.axhline(0.0, linewidth=1)
    axis.axvline(0.0, linewidth=1)
    axis.set_title(
        f"Classical RK4 Numerical Solution — {order_label(result.ode_order)}"
    )
    axis.set_xlabel("x")
    axis.set_ylabel("y")
    axis.grid(True)
    axis.legend()
    figure.tight_layout()
    return figure


def create_state_figure(
    result: RK4Result,
) -> Figure | None:
    """Plot numerical derivative-state components for higher-order ODEs."""

    if result.ode_order == 1:
        return None

    dataframe = solution_values_dataframe(result)
    figure, axis = plt.subplots(figsize=(10, 6))

    axis.plot(
        dataframe["x"],
        dataframe["RK4 yp"],
        marker="o",
        linewidth=2,
        label="RK4 y′",
    )

    if result.exact_solution_expression is not None:
        axis.plot(
            dataframe["x"],
            dataframe["Exact yp"],
            linewidth=2,
            linestyle="--",
            label="Exact y′",
        )

    if result.ode_order == 3:
        axis.plot(
            dataframe["x"],
            dataframe["RK4 ypp"],
            marker="o",
            linewidth=2,
            label="RK4 y″",
        )

        if result.exact_solution_expression is not None:
            axis.plot(
                dataframe["x"],
                dataframe["Exact ypp"],
                linewidth=2,
                linestyle="--",
                label="Exact y″",
            )

    axis.axhline(0.0, linewidth=1)
    axis.axvline(0.0, linewidth=1)
    axis.set_title("RK4 Method Derivative-State Components")
    axis.set_xlabel("x")
    axis.set_ylabel("State value")
    axis.grid(True)
    axis.legend()
    figure.tight_layout()

    return figure


def create_error_figure(
    result: RK4Result,
) -> Figure | None:
    """Plot pointwise absolute global error in y."""

    if result.exact_solution_expression is None:
        return None

    dataframe = error_analysis_dataframe(result)
    errors = np.maximum(
        dataframe[
            "Absolute Error y"
        ].to_numpy(dtype=float),
        np.finfo(float).tiny,
    )

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.semilogy(
        dataframe["x"],
        errors,
        marker="o",
        linewidth=2,
    )
    axis.set_title("RK4 Method Global Error in y")
    axis.set_xlabel("x")
    axis.set_ylabel("Absolute Error in y (Log Scale)")
    axis.grid(True, which="both")
    figure.tight_layout()

    return figure


def create_convergence_figure(
    result: RK4Result,
) -> Figure:
    """Plot grid-refinement convergence against |h|."""

    dataframe = convergence_dataframe(result)
    h_values = np.abs(
        dataframe["h"].to_numpy(dtype=float)
    )

    if result.exact_solution_expression is not None:
        metric = dataframe[
            "Absolute Error y"
        ].to_numpy(dtype=float)
        label = "Absolute Final Error in y"
    else:
        metric = dataframe[
            "Successive Difference y"
        ].to_numpy(dtype=float)
        label = "Successive Difference in Final y"

    valid = (
        np.isfinite(metric)
        & (metric > 0.0)
        & np.isfinite(h_values)
        & (h_values > 0.0)
    )

    figure, axis = plt.subplots(figsize=(10, 6))

    if np.any(valid):
        axis.loglog(
            h_values[valid],
            metric[valid],
            marker="o",
            linewidth=2,
            label=label,
        )

        reference_h = h_values[valid]
        reference_metric = metric[valid]
        reference_line = (
            reference_metric[-1]
            * (reference_h / reference_h[-1]) ** 4
        )
        axis.loglog(
            reference_h,
            reference_line,
            linestyle="--",
            label="Fourth-Order Reference O(h⁴)",
        )
    else:
        axis.text(
            0.5,
            0.5,
            (
                "Insufficient positive nonzero data "
                "for a log-log convergence plot."
            ),
            ha="center",
            va="center",
            transform=axis.transAxes,
        )

    axis.set_title(
        "RK4 Method Grid-Refinement Convergence"
    )
    axis.set_xlabel("|h|")
    axis.set_ylabel("Error Indicator")
    axis.grid(True, which="both")

    if np.any(valid):
        axis.legend()

    figure.tight_layout()
    return figure


def figure_to_png_bytes(
    figure: Figure,
) -> bytes:
    """Serialize a matplotlib figure as PNG bytes."""

    buffer = BytesIO()
    figure.savefig(
        buffer,
        format="png",
        dpi=180,
        bbox_inches="tight",
    )
    buffer.seek(0)

    return buffer.getvalue()


# =============================================================================
# Excel report
# =============================================================================
def format_workbook(
    workbook: Any,
) -> None:
    """Apply readable formatting to all workbook sheets."""

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )
    header_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        if (
            worksheet.max_row >= 1
            and worksheet.max_column >= 1
        ):
            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

                if isinstance(cell.value, float):
                    cell.number_format = (
                        "0.000000000000E+00"
                    )

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )
            max_length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                60,
            )


def add_excel_image(
    worksheet: Any,
    image_bytes: bytes,
    anchor: str,
    width: int = 780,
    height: int = 500,
) -> None:
    """Insert one PNG image into an Excel sheet."""

    image_stream = BytesIO(image_bytes)
    image = ExcelImage(image_stream)
    image.width = width
    image.height = height
    worksheet.add_image(image, anchor)


def final_state_dataframe(
    result: RK4Result,
) -> pd.DataFrame:
    """Create a compact final-state verification table."""

    rows: list[dict[str, Any]] = [
        {
            "Component": "y",
            "RK4 Value": result.final_y,
            "Exact Value": result.exact_final_y,
            "Signed Error": result.signed_final_error_y,
            "Absolute Error": result.absolute_final_error_y,
            "Relative Error (%)": (
                result.relative_final_error_y_percent
            ),
        }
    ]

    if result.ode_order >= 2:
        rows.append(
            {
                "Component": "y′",
                "RK4 Value": result.final_yp,
                "Exact Value": result.exact_final_yp,
                "Signed Error": result.signed_final_error_yp,
                "Absolute Error": result.absolute_final_error_yp,
                "Relative Error (%)": (
                    result.relative_final_error_yp_percent
                ),
            }
        )

    if result.ode_order >= 3:
        rows.append(
            {
                "Component": "y″",
                "RK4 Value": result.final_ypp,
                "Exact Value": result.exact_final_ypp,
                "Signed Error": result.signed_final_error_ypp,
                "Absolute Error": result.absolute_final_error_ypp,
                "Relative Error (%)": (
                    result.relative_final_error_ypp_percent
                ),
            }
        )

    return pd.DataFrame(rows)


def generate_excel_report(
    result: RK4Result,
) -> bytes:
    """Create a complete in-memory XLSX report."""

    if not result.success:
        raise ValueError(
            "Only successful RK4 results can be exported."
        )

    initial_condition_parts = [
        f"y({float(result.x0):.15g}) = "
        f"{float(result.y0):.15g}"
    ]

    if result.ode_order >= 2:
        initial_condition_parts.append(
            f"y'({float(result.x0):.15g}) = "
            f"{float(result.yp0):.15g}"
        )

    if result.ode_order >= 3:
        initial_condition_parts.append(
            f"y''({float(result.x0):.15g}) = "
            f"{float(result.ypp0):.15g}"
        )

    summary_rows = [
        ("Method", result.method),
        ("Status", result.status),
        ("Message", result.message),
        ("ODE Order", result.ode_order),
        (
            "Differential Equation",
            (
                f"{highest_derivative_label(result.ode_order)} "
                f"= {result.ode_text}"
            ),
        ),
        (
            "Initial Conditions",
            "; ".join(initial_condition_parts),
        ),
        ("Final x", result.x_end),
        ("Direction", result.direction),
        ("Number of Steps", result.steps),
        ("Step Size h", result.step_size),
        ("Final RK4 y", result.final_y),
        ("Final RK4 y′", result.final_yp),
        ("Final RK4 y″", result.final_ypp),
        (
            "Exact Solution",
            result.exact_solution_text or "Not supplied",
        ),
        ("Exact Final y", result.exact_final_y),
        ("Exact Final y′", result.exact_final_yp),
        ("Exact Final y″", result.exact_final_ypp),
        (
            "Absolute Final Error y",
            result.absolute_final_error_y,
        ),
        (
            "Relative Final Error y (%)",
            result.relative_final_error_y_percent,
        ),
        (
            "Maximum Absolute Global Error y",
            result.maximum_absolute_error_y,
        ),
        ("RMSE y", result.rmse_y),
        ("Expected Global Order", 4),
        ("Expected Local Error Order", 5),
        (
            "Latest Observed Order",
            result.latest_observed_order,
        ),
        (
            "Warnings",
            (
                " | ".join(result.warnings)
                if result.warnings
                else "None"
            ),
        ),
        ("Stopping Reason", result.stopping_reason),
        (
            "Execution Date",
            result.execution_datetime.strftime(
                "%Y-%m-%d %H:%M:%S %Z"
            ),
        ),
    ]
    summary_df = pd.DataFrame(
        summary_rows,
        columns=["Property", "Value"],
    )

    iteration_df = iterations_dataframe(result)
    solution_df = solution_values_dataframe(result)
    error_df = error_analysis_dataframe(result)
    convergence_df = convergence_dataframe(result)
    formula_df = method_formula_dataframe(result)
    final_state_df = final_state_dataframe(result)

    figures: list[tuple[str, bytes]] = []

    solution_figure = create_solution_figure(result)
    figures.append(
        (
            "Numerical Solution",
            figure_to_png_bytes(solution_figure),
        )
    )
    plt.close(solution_figure)

    state_figure = create_state_figure(result)
    if state_figure is not None:
        figures.append(
            (
                "Derivative States",
                figure_to_png_bytes(state_figure),
            )
        )
        plt.close(state_figure)

    error_figure = create_error_figure(result)
    if error_figure is not None:
        figures.append(
            (
                "Global Error",
                figure_to_png_bytes(error_figure),
            )
        )
        plt.close(error_figure)

    convergence_figure = create_convergence_figure(
        result
    )
    figures.append(
        (
            "Convergence",
            figure_to_png_bytes(convergence_figure),
        )
    )
    plt.close(convergence_figure)

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        summary_df.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )
        formula_df.to_excel(
            writer,
            sheet_name="Method Formula",
            index=False,
        )
        iteration_df.to_excel(
            writer,
            sheet_name="Iteration Results",
            index=False,
        )
        solution_df.to_excel(
            writer,
            sheet_name="Solution Values",
            index=False,
        )
        final_state_df.to_excel(
            writer,
            sheet_name="Final State",
            index=False,
        )
        error_df.to_excel(
            writer,
            sheet_name="Error Analysis",
            index=False,
        )
        convergence_df.to_excel(
            writer,
            sheet_name="Convergence Analysis",
            index=False,
        )

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = (
            "Classical RK4 Method Report Plots"
        )
        plots_sheet["A1"].font = Font(
            bold=True,
            size=14,
        )

        for image_index, (
            image_name,
            image_bytes,
        ) in enumerate(figures):
            anchor_row = 3 + image_index * 28
            plots_sheet[
                f"A{anchor_row - 1}"
            ] = image_name
            plots_sheet[
                f"A{anchor_row - 1}"
            ].font = Font(bold=True)
            add_excel_image(
                plots_sheet,
                image_bytes,
                f"A{anchor_row}",
            )

        if not solution_df.empty:
            summary_sheet = workbook["Summary"]
            solution_sheet = workbook[
                "Solution Values"
            ]
            max_row = len(solution_df) + 1

            chart = LineChart()
            chart.title = "RK4 Numerical Solution y"
            chart.y_axis.title = "y"
            chart.x_axis.title = "x"
            chart.height = 8
            chart.width = 15

            x_column = (
                solution_df.columns.get_loc("x") + 1
            )
            y_column = (
                solution_df.columns.get_loc("RK4 y") + 1
            )

            categories = Reference(
                solution_sheet,
                min_col=x_column,
                min_row=2,
                max_row=max_row,
            )
            data = Reference(
                solution_sheet,
                min_col=y_column,
                min_row=1,
                max_row=max_row,
            )
            chart.add_data(
                data,
                titles_from_data=True,
            )
            chart.set_categories(categories)

            if (
                result.exact_solution_expression
                is not None
            ):
                exact_column = (
                    solution_df.columns.get_loc(
                        "Exact y"
                    )
                    + 1
                )
                exact_data = Reference(
                    solution_sheet,
                    min_col=exact_column,
                    min_row=1,
                    max_row=max_row,
                )
                chart.add_data(
                    exact_data,
                    titles_from_data=True,
                )

            summary_sheet.add_chart(chart, "D2")

        if len(convergence_df) > 1:
            summary_sheet = workbook["Summary"]
            convergence_sheet = workbook[
                "Convergence Analysis"
            ]
            max_row = len(convergence_df) + 1

            chart = LineChart()
            chart.title = (
                "RK4 Grid-Refinement Convergence"
            )
            chart.x_axis.title = "Steps"
            chart.y_axis.title = (
                "Absolute Error y"
                if result.exact_solution_expression
                is not None
                else "Successive Difference y"
            )
            chart.height = 8
            chart.width = 15

            steps_column = (
                convergence_df.columns.get_loc(
                    "Steps"
                )
                + 1
            )
            metric_name = (
                "Absolute Error y"
                if result.exact_solution_expression
                is not None
                else "Successive Difference y"
            )
            metric_column = (
                convergence_df.columns.get_loc(
                    metric_name
                )
                + 1
            )

            categories = Reference(
                convergence_sheet,
                min_col=steps_column,
                min_row=2,
                max_row=max_row,
            )
            data = Reference(
                convergence_sheet,
                min_col=metric_column,
                min_row=1,
                max_row=max_row,
            )
            chart.add_data(
                data,
                titles_from_data=True,
            )
            chart.set_categories(categories)
            summary_sheet.add_chart(chart, "D20")

        format_workbook(workbook)
        workbook.active = workbook.sheetnames.index(
            "Summary"
        )

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit renderers
# =============================================================================
def render_final_result(
    result: RK4Result,
) -> None:
    """Render the final corrected state and error metrics."""

    if not result.success:
        st.error(result.message)
        return

    st.success(result.message)

    metric_columns = st.columns(2)
    metric_columns[0].metric(
        "Final x",
        format_number(result.final_x),
    )
    metric_columns[1].metric(
        "RK4 y",
        format_number(result.final_y),
    )

    st.markdown(
        f"**ODE order:** {order_label(result.ode_order)}"
    )
    st.markdown(
        f"**Differential Equation:** "
        f"`{highest_derivative_label(result.ode_order)} "
        f"= {result.ode_text}`"
    )
    st.markdown(
        f"**Initial y:** "
        f"`y({format_number(result.x0)}) = "
        f"{format_number(result.y0)}`"
    )

    if result.ode_order >= 2:
        st.markdown(
            f"**Initial y′:** "
            f"`y′({format_number(result.x0)}) = "
            f"{format_number(result.yp0)}`"
        )
        st.markdown(
            f"**Final y′:** "
            f"{format_number(result.final_yp)}"
        )

    if result.ode_order >= 3:
        st.markdown(
            f"**Initial y″:** "
            f"`y″({format_number(result.x0)}) = "
            f"{format_number(result.ypp0)}`"
        )
        st.markdown(
            f"**Final y″:** "
            f"{format_number(result.final_ypp)}"
        )

    st.markdown(f"**Number of Steps:** {result.steps}")
    st.markdown(
        f"**Step Size:** {format_number(result.step_size)}"
    )
    st.markdown(f"**Direction:** {result.direction}")

    if result.exact_final_y is not None:
        st.markdown(
            f"**Exact Final y:** "
            f"{format_number(result.exact_final_y)}"
        )
        st.markdown(
            f"**Absolute Final Error y:** "
            f"{format_number(result.absolute_final_error_y)}"
        )
        st.markdown(
            f"**Relative Final Error y:** "
            f"{format_number(result.relative_final_error_y_percent)}%"
        )

    st.markdown(
        f"**Stopping Reason:** {result.stopping_reason}"
    )

    for warning in result.warnings:
        st.warning(warning)


def render_method_summary(
    result: RK4Result,
) -> None:
    """Render vector classical RK4 formulas and problem setup."""

    st.subheader("Method Formula and Problem Setup")
    column_1, column_2 = st.columns(2)

    with column_1:
        st.latex(r"\mathbf{K}_1=\mathbf{G}(x_n,\mathbf{U}_n)")
        st.latex(
            r"\mathbf{K}_2=\mathbf{G}\left(x_n+\frac{h}{2},"
            r"\mathbf{U}_n+\frac{h}{2}\mathbf{K}_1\right)"
        )
        st.latex(
            r"\mathbf{K}_3=\mathbf{G}\left(x_n+\frac{h}{2},"
            r"\mathbf{U}_n+\frac{h}{2}\mathbf{K}_2\right)"
        )

    with column_2:
        st.latex(
            r"\mathbf{K}_4=\mathbf{G}(x_n+h,"
            r"\mathbf{U}_n+h\mathbf{K}_3)"
        )
        st.latex(
            r"\mathbf{U}_{n+1}=\mathbf{U}_n+\frac{h}{6}"
            r"(\mathbf{K}_1+2\mathbf{K}_2+2\mathbf{K}_3+\mathbf{K}_4)"
        )
        st.latex(r"\text{Global error }O(h^4),\quad \text{local error }O(h^5)")

    st.dataframe(
        round_numeric_dataframe(method_formula_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )


def render_iteration_table(
    result: RK4Result,
) -> None:
    """Render every simultaneous RK4 stage and accepted update."""

    st.subheader("RK4 Iteration Table")
    dataframe = iterations_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(dataframe),
        use_container_width=True,
        hide_index=True,
    )

    st.caption(
        "All state components are advanced simultaneously at each of the four "
        "RK4 stages. Local truncation error is calculated by starting one RK4 "
        "step from the exact state when an exact solution is supplied."
    )

    with st.expander("Detailed RK4 Stage Operations"):
        for record in result.iterations:
            st.markdown(f"**Iteration {record.iteration}**")
            st.code(
                "Stage calculations:\n"
                + record.stage_operation
                + "\n\nAccepted weighted update:\n"
                + record.update_operation,
                language=None,
            )


def render_solution_graphs(
    result: RK4Result,
) -> None:
    """Render solution and derivative-state graphs."""

    st.subheader("Numerical Solution Graph")

    try:
        figure = create_solution_figure(result)
        st.pyplot(
            figure,
            use_container_width=True,
        )
        plt.close(figure)
    except (
        ValueError,
        TypeError,
        ArithmeticError,
        OverflowError,
    ) as error:
        st.warning(
            f"The solution graph could not be displayed: {error}"
        )

    state_figure = create_state_figure(result)

    if state_figure is not None:
        st.subheader("Derivative-State Graph")
        st.pyplot(
            state_figure,
            use_container_width=True,
        )
        plt.close(state_figure)


def render_error_analysis(
    result: RK4Result,
) -> None:
    """Render pointwise and final state errors."""

    st.subheader("Error Analysis")

    if result.exact_solution_expression is None:
        st.info(
            "Enter an exact solution y(x) to calculate true "
            "global error, local truncation error, relative "
            "error, maximum absolute error, and RMSE."
        )
        return

    final_state_df = final_state_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(final_state_df),
        use_container_width=True,
        hide_index=True,
    )

    metric_columns = st.columns(4)
    metric_columns[0].metric(
        "Final Absolute Error y",
        format_number(
            result.absolute_final_error_y
        ),
    )
    metric_columns[1].metric(
        "Final Relative Error y",
        (
            format_number(
                result.relative_final_error_y_percent
            )
            + "%"
        ),
    )
    metric_columns[2].metric(
        "Maximum Absolute Error y",
        format_number(
            result.maximum_absolute_error_y
        ),
    )
    metric_columns[3].metric(
        "RMSE y",
        format_number(result.rmse_y),
    )

    dataframe = error_analysis_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(dataframe),
        use_container_width=True,
        hide_index=True,
    )

    figure = create_error_figure(result)

    if figure is not None:
        st.pyplot(
            figure,
            use_container_width=True,
        )
        plt.close(figure)


def render_convergence_analysis(
    result: RK4Result,
) -> None:
    """Render fourth-order grid-refinement convergence analysis."""

    st.subheader("Convergence Analysis")
    st.markdown(
        "The classical RK4 method has **fourth-order global convergence** "
        "under the usual smoothness and stability conditions. Halving the step "
        "size should reduce the asymptotic global error by approximately a "
        "factor of sixteen."
    )

    dataframe = convergence_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(dataframe),
        use_container_width=True,
        hide_index=True,
    )

    if result.latest_observed_order is not None:
        st.metric(
            "Latest Observed Order",
            format_number(result.latest_observed_order),
        )

    figure = create_convergence_figure(result)
    st.pyplot(figure, use_container_width=True)
    plt.close(figure)


def render_excel_download(
    result: RK4Result,
) -> None:
    """Generate and render the Excel report button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature

    if (
        st.session_state.get(
            "rk4_excel_signature"
        )
        != report_signature
        or "rk4_excel_report"
        not in st.session_state
    ):
        try:
            st.session_state.rk4_excel_report = (
                generate_excel_report(result)
            )
            st.session_state.rk4_excel_signature = (
                report_signature
            )
        except (
            ValueError,
            OSError,
            TypeError,
            ArithmeticError,
        ) as error:
            st.error(
                "The Excel report could not be generated: "
                f"{error}"
            )
            return

    report_bytes = st.session_state.get(
        "rk4_excel_report"
    )

    if not report_bytes:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime(
        "%Y%m%d_%H%M%S"
    )
    filename = (
        f"rk4_method_report_{timestamp}.xlsx"
    )

    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=filename,
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="rk4_download_button",
    )



# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete classical RK4 Streamlit page."""

    st.set_page_config(
        page_title="Fourth-Order Runge–Kutta Method Solver | Numerical Methods",
        page_icon="∫",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">ORDINARY DIFFERENTIAL EQUATIONS TOOL</div>
                <h1>Fourth-Order Runge–Kutta Method Solver</h1>
                <p>
                    Solve first-, second-, or third-order initial-value problems
                    using the classical fourth-order Runge–Kutta method. Every
                    stage advances the complete state vector simultaneously, and
                    higher-order equations are converted to first-order systems.
                </p>

                <div class="method-actions">
                    <a href="/Runge_Kutta_4" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Runge_Kutta_4_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns(
        [0.035, 0.93, 0.035]
    )

    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Write the Function")
                st.markdown(
                    """
                    Enter only the right-hand side of the highest derivative.

                    - First order: **y′ = F(x, y)** — use `x` and `y`.
                    - Second order: **y″ = F(x, y, y′)** — write `yp` for **y′**.
                    - Third order: **y‴ = F(x, y, y′, y″)** — write `yp` and `ypp`.
                    - Powers: write `x**2`, not `x^2`.
                    - Multiplication: write `2*x`, not `2x`.
                    - Enter no equals sign.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")
                st.markdown(
                    """
                    - First order requires **y(x₀)**.
                    - Second order also requires **y′(x₀)**.
                    - Third order also requires **y″(x₀)**.
                    - Classical RK4 evaluates four simultaneous state-vector stages per step.
                    - The number of steps must be a positive integer.
                    - A large step may cause inaccuracy or explicit-method instability.
                    """
                )

        input_column, result_column = st.columns(
            [1.35, 1.0]
        )

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">'
                    'Differential-equation order</div>',
                    unsafe_allow_html=True,
                )
                ode_order = st.selectbox(
                    "Differential-equation order",
                    options=[1, 2, 3],
                    index=DEFAULT_ORDER - 1,
                    format_func=lambda value: order_label(value),
                    label_visibility="collapsed",
                    key="rk4_ode_order",
                )

                st.markdown(
                    f'<div class="input-label-ui">'
                    f'Right-hand side of '
                    f'{highest_derivative_label(ode_order)} = '
                    f'F({rhs_variables_text(ode_order)})</div>',
                    unsafe_allow_html=True,
                )
                ode_text = st.text_input(
                    "Right-hand side",
                    value=DEFAULT_ODE,
                    placeholder={
                        1: "Example: x + y",
                        2: "Example: -y or -0.2*yp - y",
                        3: "Example: 6 or y - 0.2*ypp",
                    }[ode_order],
                    label_visibility="collapsed",
                    key="rk4_ode_function",
                )

                first_row = st.columns(3)

                with first_row[0]:
                    st.markdown(
                        '<div class="input-label-ui">'
                        'Initial x₀</div>',
                        unsafe_allow_html=True,
                    )
                    x0 = st.number_input(
                        "Initial x0",
                        value=DEFAULT_X0,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="rk4_x0",
                    )

                with first_row[1]:
                    st.markdown(
                        '<div class="input-label-ui">'
                        'Initial y(x₀)</div>',
                        unsafe_allow_html=True,
                    )
                    y0 = st.number_input(
                        "Initial y",
                        value=DEFAULT_Y0,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="rk4_y0",
                    )

                with first_row[2]:
                    st.markdown(
                        '<div class="input-label-ui">'
                        'Final x</div>',
                        unsafe_allow_html=True,
                    )
                    x_end = st.number_input(
                        "Final x",
                        value=DEFAULT_X_END,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="rk4_x_end",
                    )

                yp0: float | None = None
                ypp0: float | None = None

                if ode_order >= 2:
                    derivative_columns = st.columns(
                        2 if ode_order == 3 else 1
                    )

                    with derivative_columns[0]:
                        st.markdown(
                            '<div class="input-label-ui">'
                            'Initial y′(x₀)</div>',
                            unsafe_allow_html=True,
                        )
                        yp0 = st.number_input(
                            "Initial y prime",
                            value=DEFAULT_YP0,
                            format="%.12g",
                            label_visibility="collapsed",
                            key="rk4_yp0",
                        )

                    if ode_order == 3:
                        with derivative_columns[1]:
                            st.markdown(
                                '<div class="input-label-ui">'
                                'Initial y″(x₀)</div>',
                                unsafe_allow_html=True,
                            )
                            ypp0 = st.number_input(
                                "Initial y double prime",
                                value=DEFAULT_YPP0,
                                format="%.12g",
                                label_visibility="collapsed",
                                key="rk4_ypp0",
                            )

                second_row = st.columns(2)

                with second_row[0]:
                    st.markdown(
                        '<div class="input-label-ui">'
                        'Number of steps N</div>',
                        unsafe_allow_html=True,
                    )
                    steps = st.number_input(
                        "Number of steps",
                        min_value=MIN_STEPS,
                        max_value=MAX_STEPS,
                        value=DEFAULT_STEPS,
                        step=1,
                        label_visibility="collapsed",
                        key="rk4_steps",
                    )

                with second_row[1]:
                    st.markdown(
                        '<div class="input-label-ui">'
                        'Convergence levels</div>',
                        unsafe_allow_html=True,
                    )
                    convergence_levels = st.number_input(
                        "Convergence levels",
                        min_value=MIN_CONVERGENCE_LEVELS,
                        max_value=MAX_CONVERGENCE_LEVELS,
                        value=DEFAULT_CONVERGENCE_LEVELS,
                        step=1,
                        label_visibility="collapsed",
                        key="rk4_convergence_levels",
                    )

                st.markdown(
                    '<div class="input-label-ui">'
                    'Exact solution y(x) — optional</div>',
                    unsafe_allow_html=True,
                )
                exact_solution_text = st.text_input(
                    "Exact solution y(x)",
                    value="",
                    placeholder=(
                        "Example: exp(x), sin(x), or x**3"
                    ),
                    label_visibility="collapsed",
                    key="rk4_exact_solution",
                )

                if float(x_end) != float(x0):
                    preview_h = (
                        (float(x_end) - float(x0))
                        / int(steps)
                    )
                    st.caption(
                        "Computed uniform step size: "
                        f"h = {preview_h:.12g}"
                    )

                current_signature = create_input_signature(
                    ode_order=ode_order,
                    ode_text=ode_text,
                    x0=x0,
                    y0=y0,
                    yp0=yp0,
                    ypp0=ypp0,
                    x_end=x_end,
                    steps=steps,
                    exact_solution_text=(
                        exact_solution_text
                    ),
                    convergence_levels=(
                        convergence_levels
                    ),
                )

                solve_clicked = st.button(
                    "Solve with RK4",
                    type="primary",
                    use_container_width=True,
                    key="rk4_solve_button",
                )

                if solve_clicked:
                    result = solve_rk4_method(
                        ode_order_input=ode_order,
                        ode_text=ode_text,
                        x0_input=x0,
                        y0_input=y0,
                        yp0_input=yp0,
                        ypp0_input=ypp0,
                        x_end_input=x_end,
                        steps_input=steps,
                        exact_solution_text=(
                            exact_solution_text
                        ),
                        convergence_levels_input=(
                            convergence_levels
                        ),
                    )
                    st.session_state.rk4_result = result
                    st.session_state[
                        "rk4_result_signature"
                    ] = current_signature
                    st.session_state.pop(
                        "rk4_excel_report",
                        None,
                    )
                    st.session_state.pop(
                        "rk4_excel_signature",
                        None,
                    )

                with st.expander("Example Inputs"):
                    st.code(
                        "First order\n"
                        "y' = x + y\n"
                        "x0 = 0, y0 = 1, x_end = 1\n"
                        "Exact y = 2*exp(x) - x - 1",
                        language=None,
                    )
                    st.code(
                        "Second order\n"
                        "y'' = -y\n"
                        "x0 = 0, y0 = 0, yp0 = 1\n"
                        "x_end = 1\n"
                        "Exact y = sin(x)",
                        language=None,
                    )
                    st.code(
                        "Third order\n"
                        "y‴ = y\n"
                        "x0 = 0, y0 = 1, yp0 = 1, ypp0 = 1\n"
                        "x_end = 1\n"
                        "Exact y = exp(x)",
                        language=None,
                    )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">'
                    'Final Result</h3>',
                    unsafe_allow_html=True,
                )

                saved_result = st.session_state.get(
                    "rk4_result"
                )
                saved_signature = st.session_state.get(
                    "rk4_result_signature"
                )

                if saved_result is None:
                    st.info(
                        "Enter the IVP data and select "
                        "Solve with RK4."
                    )
                elif saved_signature != current_signature:
                    st.warning(
                        "The inputs have changed. Select "
                        "Solve with RK4 to update the result."
                    )
                else:
                    render_final_result(saved_result)

        saved_result = st.session_state.get(
            "rk4_result"
        )
        saved_signature = st.session_state.get(
            "rk4_result_signature"
        )

        if (
            saved_result is not None
            and saved_signature == current_signature
            and saved_result.success
        ):
            st.divider()
            render_method_summary(saved_result)
            st.divider()
            render_iteration_table(saved_result)
            st.divider()
            render_solution_graphs(saved_result)
            st.divider()
            render_error_analysis(saved_result)
            st.divider()
            render_convergence_analysis(saved_result)
            st.divider()
            render_excel_download(saved_result)

        st.markdown(
            "</main>",
            unsafe_allow_html=True,
        )

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Fourth-Order Runge–Kutta • Ordinary Differential Equations</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()