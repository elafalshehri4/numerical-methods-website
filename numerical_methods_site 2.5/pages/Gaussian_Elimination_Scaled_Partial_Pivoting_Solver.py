from __future__ import annotations

import hashlib
import math
from io import BytesIO
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from components.navigation import navbar
from utilities.ui import load_css


# =============================================================================
# Method configuration
# =============================================================================
CONFIG = {
    "method_id": "scaled_pivoting",
    "title": "Scaled Partial Pivoting Solver",
    "label": "SCALED PARTIAL PIVOTING TOOL",
    "description": (
        "Solve a square linear system using scale-aware pivot selection, "
        "forward elimination, and back substitution."
    ),
    "lesson": "Gaussian_Elimination_Scaled_Partial_Pivoting",
    "quiz": "Gaussian_Elimination_Scaled_Partial_Pivoting_Quiz",
    "footer": "Scaled Partial Pivoting Solver • Linear Systems",
    "conditions": [
        "The coefficient matrix must be square.",
        "The matrix and right-hand-side dimensions must agree.",
        "Each row must contain at least one nonzero coefficient.",
        "A unique solution requires a valid scaled pivot at every stage.",
    ],
    "formula": (
        r"s_i=\max_j|a_{ij}|,\qquad "
        r"r_i=\frac{|a_{ik}|}{s_i},\qquad "
        r"r_p=\max_{i\geq k}r_i"
    ),
    "defaults": {
        "A": [[0.001, 1.0], [1.0, 1.0]],
        "b": [1.001, 2.0],
    },
}

DISPLAY_DECIMALS = 3
MIN_SYSTEM_SIZE = 2
MAX_SYSTEM_SIZE = 6

CONDITION_NUMBER_WARNING = 1.0e12
RELATIVE_RESIDUAL_WARNING = 1.0e-10

EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# =============================================================================
# Numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Format a number as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable

    number = float(value)
    if not math.isfinite(number):
        return str(number)

    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0**exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)

    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Use fixed notation unless scientific notation is clearer."""

    if value is None:
        return unavailable

    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6
    ):
        return format_scientific_power(
            number,
            decimals=decimals,
            unavailable=unavailable,
        )

    return f"{number:.{decimals}f}"


def format_number(value: float | int | None) -> str:
    """Format a final solver value."""

    return format_display_number(
        value,
        unavailable="Not available",
    )


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only for interface display."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(
        include=[np.number]
    ).columns

    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[
            numeric_columns
        ].round(decimals)

    return rounded


# =============================================================================
# Input helpers
# =============================================================================
def create_default_table(size: int) -> pd.DataFrame:
    """Create the editable coefficient table."""

    default_A = np.eye(size, dtype=float) * 2.0
    default_b = np.ones(size, dtype=float)

    stored_A = np.asarray(
        CONFIG["defaults"]["A"],
        dtype=float,
    )
    stored_b = np.asarray(
        CONFIG["defaults"]["b"],
        dtype=float,
    )

    copied_rows = min(size, stored_A.shape[0])
    copied_columns = min(size, stored_A.shape[1])

    default_A[:copied_rows, :copied_columns] = stored_A[
        :copied_rows,
        :copied_columns,
    ]
    default_b[:copied_rows] = stored_b[:copied_rows]

    data = {
        f"x{column + 1}": default_A[:, column]
        for column in range(size)
    }
    data["b"] = default_b

    return pd.DataFrame(data)


def create_input_signature(
    input_table: pd.DataFrame,
    size: int,
) -> str:
    """Create a stable signature to prevent stale results."""

    normalized = input_table.copy()
    normalized.columns = [
        str(column) for column in normalized.columns
    ]

    payload = (
        f"{size}|"
        + normalized.to_csv(
            index=False,
            float_format="%.17g",
        )
    )

    return hashlib.sha256(
        payload.encode("utf-8")
    ).hexdigest()


def extract_system_from_table(
    input_table: pd.DataFrame,
    size: int,
) -> tuple[np.ndarray, np.ndarray]:
    """Convert and validate table values."""

    required_columns = [
        f"x{index + 1}" for index in range(size)
    ] + ["b"]

    missing_columns = [
        column
        for column in required_columns
        if column not in input_table.columns
    ]

    if missing_columns:
        raise ValueError(
            "The input table is missing required column(s): "
            + ", ".join(missing_columns)
        )

    try:
        numeric_table = input_table[
            required_columns
        ].apply(
            pd.to_numeric,
            errors="raise",
        )
    except (TypeError, ValueError) as error:
        raise ValueError(
            "Every matrix and vector entry must be a valid number."
        ) from error

    if numeric_table.shape != (size, size + 1):
        raise ValueError(
            f"The table must contain exactly {size} equations."
        )

    matrix = numeric_table.iloc[
        :, :size
    ].to_numpy(dtype=float)
    vector = numeric_table["b"].to_numpy(dtype=float)

    if not np.all(np.isfinite(matrix)):
        raise ValueError(
            "All coefficient-matrix entries must be finite."
        )

    if not np.all(np.isfinite(vector)):
        raise ValueError(
            "All right-hand-side entries must be finite."
        )

    return matrix, vector


# =============================================================================
# General linear-system helpers
# =============================================================================
def augmented_text(
    matrix: np.ndarray,
    vector: np.ndarray,
) -> str:
    """Return a readable augmented matrix."""

    augmented = np.column_stack((matrix, vector))

    return np.array2string(
        augmented,
        precision=8,
        suppress_small=True,
        max_line_width=160,
    )


def calculate_diagnostic_tolerance(
    matrix: np.ndarray,
) -> float:
    """Return a scale-aware rank and pivot tolerance."""

    matrix_norm = float(
        np.linalg.norm(matrix, ord=np.inf)
    )
    scale = max(
        matrix_norm,
        np.finfo(float).tiny,
    )

    return (
        100.0
        * np.finfo(float).eps
        * max(matrix.shape)
        * scale
    )


def calculate_row_threshold(
    row_scale: float,
    matrix_size: int,
) -> float:
    """Return a pivot threshold relative to one row's scale."""

    return (
        100.0
        * np.finfo(float).eps
        * max(matrix_size, 1)
        * max(abs(float(row_scale)), np.finfo(float).tiny)
    )


def clean_small_values(
    values: np.ndarray,
    tolerance: float,
) -> np.ndarray:
    """Replace insignificant arithmetic noise by zero."""

    cleaned = np.asarray(values, dtype=float).copy()
    cleaned[np.abs(cleaned) <= tolerance] = 0.0
    return cleaned


def classify_system(
    matrix: np.ndarray,
    vector: np.ndarray,
    tolerance: float,
) -> dict[str, Any]:
    """Classify the original system using matrix ranks."""

    augmented = np.column_stack((matrix, vector))

    rank_A = int(
        np.linalg.matrix_rank(
            matrix,
            tol=tolerance,
        )
    )
    rank_augmented = int(
        np.linalg.matrix_rank(
            augmented,
            tol=tolerance,
        )
    )

    variable_count = matrix.shape[1]

    if rank_A < rank_augmented:
        classification = "No solution"
    elif rank_A < variable_count:
        classification = "Infinitely many solutions"
    else:
        classification = "Unique solution"

    return {
        "classification": classification,
        "rank_A": rank_A,
        "rank_augmented": rank_augmented,
    }


# =============================================================================
# Scaled Partial Pivoting algorithm
# =============================================================================
def solve_linear_system(
    matrix: Any,
    vector: Any,
    input_signature: str = "",
) -> dict[str, Any]:
    """Solve Ax=b using Gaussian elimination with scaled partial pivoting.

    The scale factors s_i=max_j|a_ij| are computed once from the original
    coefficient rows. Whenever rows are interchanged, the associated scale
    factors and original-row labels are interchanged with them.
    """

    try:
        A = np.asarray(
            matrix,
            dtype=float,
        ).copy()
        b = np.asarray(
            vector,
            dtype=float,
        ).reshape(-1).copy()
    except (TypeError, ValueError) as error:
        return {
            "status": "error",
            "classification": "Invalid input",
            "message": (
                "The matrix and vector must contain numerical values."
            ),
            "input_signature": input_signature,
        }

    if A.ndim != 2 or A.shape[0] != A.shape[1]:
        return {
            "status": "error",
            "classification": "Invalid input",
            "message": "A must be a square matrix.",
            "input_signature": input_signature,
        }

    size = A.shape[0]

    if b.shape != (size,):
        return {
            "status": "error",
            "classification": "Invalid input",
            "message": (
                "Vector b must contain exactly one value per equation."
            ),
            "input_signature": input_signature,
        }

    if not np.all(np.isfinite(A)) or not np.all(np.isfinite(b)):
        return {
            "status": "error",
            "classification": "Invalid input",
            "message": "All entries must be finite real numbers.",
            "input_signature": input_signature,
        }

    original_A = A.copy()
    original_b = b.copy()

    diagnostic_tolerance = calculate_diagnostic_tolerance(
        original_A
    )
    diagnostics = classify_system(
        original_A,
        original_b,
        diagnostic_tolerance,
    )

    scale_factors = np.max(
        np.abs(A),
        axis=1,
    ).astype(float)

    zero_scale_rows = np.where(
        scale_factors <= diagnostic_tolerance
    )[0]

    initial_scale_table = pd.DataFrame(
        {
            "Current Row": np.arange(1, size + 1),
            "Original Row": np.arange(1, size + 1),
            "Scale Factor": scale_factors,
        }
    )

    if zero_scale_rows.size > 0:
        rows_text = ", ".join(
            str(index + 1) for index in zero_scale_rows
        )

        if diagnostics["classification"] == "No solution":
            message = (
                "The system is inconsistent and has no solution. "
                f"Zero-coefficient row(s): {rows_text}."
            )
        else:
            message = (
                "The coefficient matrix is singular and does not have "
                "a unique solution. "
                f"Zero-coefficient row(s): {rows_text}."
            )

        return {
            "status": "error",
            "classification": diagnostics["classification"],
            "message": message,
            "history": [],
            "pivot_candidates": [],
            "back_substitution": [],
            "original_A": original_A,
            "original_b": original_b,
            "rank_A": diagnostics["rank_A"],
            "rank_augmented": diagnostics["rank_augmented"],
            "scale_factors_initial": scale_factors,
            "initial_scale_table": initial_scale_table,
            "input_signature": input_signature,
        }

    if diagnostics["classification"] != "Unique solution":
        return {
            "status": "error",
            "classification": diagnostics["classification"],
            "message": (
                "Scaled Gaussian elimination with back substitution "
                "requires a unique solution. The original system is "
                f"classified as: {diagnostics['classification']}."
            ),
            "history": [],
            "pivot_candidates": [],
            "back_substitution": [],
            "original_A": original_A,
            "original_b": original_b,
            "rank_A": diagnostics["rank_A"],
            "rank_augmented": diagnostics["rank_augmented"],
            "scale_factors_initial": scale_factors,
            "initial_scale_table": initial_scale_table,
            "input_signature": input_signature,
        }

    history: list[dict[str, Any]] = []
    pivot_candidates: list[dict[str, Any]] = []
    back_substitution: list[dict[str, Any]] = []

    row_order = np.arange(1, size + 1, dtype=int)
    maximum_original_entry = max(
        float(np.max(np.abs(original_A))),
        np.finfo(float).tiny,
    )
    maximum_elimination_entry = maximum_original_entry

    cleanup_tolerance = diagnostic_tolerance

    try:
        for pivot_column in range(size - 1):
            candidate_rows = np.arange(
                pivot_column,
                size,
            )

            candidate_scales = scale_factors[
                candidate_rows
            ]
            candidate_values = np.abs(
                A[candidate_rows, pivot_column]
            )
            candidate_ratios = (
                candidate_values / candidate_scales
            )

            if not np.all(np.isfinite(candidate_ratios)):
                raise ArithmeticError(
                    "A non-finite scaled pivot ratio was produced."
                )

            selected_offset = int(
                np.argmax(candidate_ratios)
            )
            pivot_row = int(
                candidate_rows[selected_offset]
            )
            selected_ratio = float(
                candidate_ratios[selected_offset]
            )
            selected_pivot = float(
                A[pivot_row, pivot_column]
            )
            selected_scale = float(
                scale_factors[pivot_row]
            )

            for candidate_row, value, scale, ratio in zip(
                candidate_rows,
                candidate_values,
                candidate_scales,
                candidate_ratios,
            ):
                pivot_candidates.append(
                    {
                        "Pivot Stage": pivot_column + 1,
                        "Current Row": int(candidate_row + 1),
                        "Original Row": int(
                            row_order[candidate_row]
                        ),
                        "Column Value": float(
                            A[candidate_row, pivot_column]
                        ),
                        "Absolute Column Value": float(value),
                        "Scale Factor": float(scale),
                        "Scaled Ratio": float(ratio),
                        "Selected": (
                            "Yes"
                            if candidate_row == pivot_row
                            else "No"
                        ),
                    }
                )

            pivot_threshold = calculate_row_threshold(
                selected_scale,
                size,
            )

            if (
                abs(selected_pivot) <= pivot_threshold
                or selected_ratio
                <= 100.0 * np.finfo(float).eps * size
            ):
                raise ArithmeticError(
                    f"No numerically usable scaled pivot exists "
                    f"in column {pivot_column + 1}."
                )

            row_order_before = row_order.copy()
            matrix_before_selection = A.copy()
            vector_before_selection = b.copy()
            scales_before_selection = scale_factors.copy()

            if pivot_row != pivot_column:
                A[[pivot_column, pivot_row], :] = A[
                    [pivot_row, pivot_column],
                    :,
                ]
                b[[pivot_column, pivot_row]] = b[
                    [pivot_row, pivot_column]
                ]
                scale_factors[
                    [pivot_column, pivot_row]
                ] = scale_factors[
                    [pivot_row, pivot_column]
                ]
                row_order[
                    [pivot_column, pivot_row]
                ] = row_order[
                    [pivot_row, pivot_column]
                ]

                swap_operation = (
                    f"R{pivot_column + 1} ↔ R{pivot_row + 1}"
                )
            else:
                swap_operation = "No row interchange required"

            history.append(
                {
                    "Step": len(history) + 1,
                    "Stage": "Scaled pivot selection",
                    "Pivot Column": pivot_column + 1,
                    "Operation": swap_operation,
                    "Selected Original Row": int(
                        row_order[pivot_column]
                    ),
                    "Pivot Value": float(
                        A[pivot_column, pivot_column]
                    ),
                    "Scale Factor": float(
                        scale_factors[pivot_column]
                    ),
                    "Scaled Ratio": selected_ratio,
                    "Multiplier": None,
                    "Row Order Before": str(
                        row_order_before.tolist()
                    ),
                    "Row Order After": str(
                        row_order.tolist()
                    ),
                    "Scale Factors Before": str(
                        scales_before_selection.tolist()
                    ),
                    "Scale Factors After": str(
                        scale_factors.tolist()
                    ),
                    "Augmented Matrix Before": augmented_text(
                        matrix_before_selection,
                        vector_before_selection,
                    ),
                    "Augmented Matrix After": augmented_text(
                        A,
                        b,
                    ),
                }
            )

            pivot = float(
                A[pivot_column, pivot_column]
            )
            pivot_threshold = calculate_row_threshold(
                scale_factors[pivot_column],
                size,
            )

            if abs(pivot) <= pivot_threshold:
                raise ArithmeticError(
                    f"The selected pivot at stage "
                    f"{pivot_column + 1} is numerically unusable."
                )

            for target_row in range(
                pivot_column + 1,
                size,
            ):
                matrix_before = A.copy()
                vector_before = b.copy()

                eliminated_value = float(
                    A[target_row, pivot_column]
                )
                multiplier = eliminated_value / pivot

                if not math.isfinite(multiplier):
                    raise ArithmeticError(
                        "A non-finite elimination multiplier "
                        "was produced."
                    )

                A[
                    target_row,
                    pivot_column:,
                ] -= (
                    multiplier
                    * A[
                        pivot_column,
                        pivot_column:,
                    ]
                )
                b[target_row] -= (
                    multiplier * b[pivot_column]
                )
                A[target_row, pivot_column] = 0.0

                if (
                    not np.all(np.isfinite(A))
                    or not np.all(np.isfinite(b))
                ):
                    raise ArithmeticError(
                        "Forward elimination produced NaN or infinity."
                    )

                A = clean_small_values(
                    A,
                    cleanup_tolerance,
                )
                b = clean_small_values(
                    b,
                    cleanup_tolerance,
                )

                maximum_elimination_entry = max(
                    maximum_elimination_entry,
                    float(np.max(np.abs(A))),
                )

                history.append(
                    {
                        "Step": len(history) + 1,
                        "Stage": "Forward elimination",
                        "Pivot Column": pivot_column + 1,
                        "Operation": (
                            f"R{target_row + 1} ← "
                            f"R{target_row + 1} − "
                            f"({multiplier:.12g})"
                            f"R{pivot_column + 1}"
                        ),
                        "Selected Original Row": int(
                            row_order[pivot_column]
                        ),
                        "Pivot Value": pivot,
                        "Scale Factor": float(
                            scale_factors[pivot_column]
                        ),
                        "Scaled Ratio": selected_ratio,
                        "Multiplier": float(multiplier),
                        "Row Order Before": str(
                            row_order.tolist()
                        ),
                        "Row Order After": str(
                            row_order.tolist()
                        ),
                        "Scale Factors Before": str(
                            scale_factors.tolist()
                        ),
                        "Scale Factors After": str(
                            scale_factors.tolist()
                        ),
                        "Augmented Matrix Before": augmented_text(
                            matrix_before,
                            vector_before,
                        ),
                        "Augmented Matrix After": augmented_text(
                            A,
                            b,
                        ),
                    }
                )

        final_threshold = calculate_row_threshold(
            scale_factors[-1],
            size,
        )

        if abs(A[-1, -1]) <= final_threshold:
            raise ArithmeticError(
                "The final diagonal pivot is zero or "
                "numerically unusable."
            )

        upper_matrix = A.copy()
        transformed_vector = b.copy()

        solution = np.zeros(size, dtype=float)

        for row in range(size - 1, -1, -1):
            diagonal = float(A[row, row])
            diagonal_threshold = calculate_row_threshold(
                scale_factors[row],
                size,
            )

            if abs(diagonal) <= diagonal_threshold:
                raise ArithmeticError(
                    f"Back substitution encountered an unusable "
                    f"diagonal coefficient in row {row + 1}."
                )

            known_sum = float(
                np.dot(
                    A[row, row + 1 :],
                    solution[row + 1 :],
                )
            )
            numerator = float(
                b[row] - known_sum
            )
            value = numerator / diagonal

            if not math.isfinite(value):
                raise ArithmeticError(
                    "Back substitution produced a non-finite value."
                )

            solution[row] = value

            back_substitution.append(
                {
                    "Step": len(back_substitution) + 1,
                    "Row": row + 1,
                    "Variable": f"x{row + 1}",
                    "Right-Hand Side": float(b[row]),
                    "Known-Term Sum": known_sum,
                    "Numerator": numerator,
                    "Diagonal Coefficient": diagonal,
                    "Calculated Value": float(value),
                    "Formula": (
                        f"x{row + 1} = "
                        f"({b[row]:.12g} − {known_sum:.12g}) "
                        f"/ {diagonal:.12g}"
                    ),
                }
            )

        computed_b = original_A @ solution
        residual_vector = computed_b - original_b

        residual_norm_2 = float(
            np.linalg.norm(
                residual_vector,
                ord=2,
            )
        )
        residual_norm_inf = float(
            np.linalg.norm(
                residual_vector,
                ord=np.inf,
            )
        )
        maximum_absolute_residual = float(
            np.max(np.abs(residual_vector))
        )

        denominator = float(
            np.linalg.norm(original_A, ord=np.inf)
            * np.linalg.norm(solution, ord=np.inf)
            + np.linalg.norm(original_b, ord=np.inf)
        )

        relative_residual = (
            residual_norm_inf / denominator
            if denominator > 0.0
            else residual_norm_inf
        )

        determinant = float(
            np.linalg.det(original_A)
        )

        try:
            condition_number = float(
                np.linalg.cond(
                    original_A,
                    p=np.inf,
                )
            )
        except np.linalg.LinAlgError:
            condition_number = math.inf

        growth_factor = (
            maximum_elimination_entry
            / maximum_original_entry
        )

        warnings: list[str] = []

        if (
            not math.isfinite(condition_number)
            or condition_number
            >= CONDITION_NUMBER_WARNING
        ):
            warnings.append(
                "The coefficient matrix is ill-conditioned. "
                "The solution can be highly sensitive to input "
                "and rounding errors."
            )

        if relative_residual > RELATIVE_RESIDUAL_WARNING:
            warnings.append(
                "The scale-aware relative residual is larger "
                "than expected. Inspect conditioning and row operations."
            )

        if growth_factor > 1.0e6:
            warnings.append(
                "Large element growth occurred during elimination. "
                "This can amplify floating-point round-off."
            )

        return {
            "status": "success",
            "classification": "Unique solution",
            "message": (
                "The linear system was solved successfully using "
                "Gaussian Elimination with Scaled Partial Pivoting."
            ),
            "solution": solution,
            "history": history,
            "pivot_candidates": pivot_candidates,
            "back_substitution": back_substitution,
            "original_A": original_A,
            "original_b": original_b,
            "upper_matrix": upper_matrix,
            "transformed_vector": transformed_vector,
            "final_matrix": upper_matrix,
            "final_vector": transformed_vector,
            "final_form_name": "Upper-Triangular Form",
            "scale_factors_initial": initial_scale_table[
                "Scale Factor"
            ].to_numpy(dtype=float),
            "scale_factors_final": scale_factors,
            "row_order_final": row_order,
            "initial_scale_table": initial_scale_table,
            "computed_b": computed_b,
            "residual_vector": residual_vector,
            "residual_norm_2": residual_norm_2,
            "residual_norm": residual_norm_inf,
            "residual_norm_inf": residual_norm_inf,
            "maximum_absolute_residual": (
                maximum_absolute_residual
            ),
            "relative_residual": relative_residual,
            "determinant": determinant,
            "condition_number": condition_number,
            "growth_factor": growth_factor,
            "rank_A": diagnostics["rank_A"],
            "rank_augmented": diagnostics[
                "rank_augmented"
            ],
            "operations": len(history),
            "row_swaps": sum(
                1
                for item in history
                if item["Stage"] == "Scaled pivot selection"
                and item["Operation"]
                != "No row interchange required"
            ),
            "warnings": tuple(warnings),
            "diagnostic_tolerance": (
                diagnostic_tolerance
            ),
            "input_signature": input_signature,
        }

    except (
        ArithmeticError,
        FloatingPointError,
        OverflowError,
    ) as error:
        diagnostics = classify_system(
            original_A,
            original_b,
            diagnostic_tolerance,
        )

        if diagnostics["classification"] == "No solution":
            message = (
                "The system is inconsistent and has no solution."
            )
        elif (
            diagnostics["classification"]
            == "Infinitely many solutions"
        ):
            message = (
                "The coefficient matrix is singular and the system "
                "has infinitely many solutions."
            )
        else:
            message = str(error)

        return {
            "status": "error",
            "classification": diagnostics[
                "classification"
            ],
            "message": message,
            "details": str(error),
            "history": history,
            "pivot_candidates": pivot_candidates,
            "back_substitution": back_substitution,
            "original_A": original_A,
            "original_b": original_b,
            "final_matrix": A,
            "final_vector": b,
            "rank_A": diagnostics["rank_A"],
            "rank_augmented": diagnostics[
                "rank_augmented"
            ],
            "scale_factors_initial": (
                initial_scale_table[
                    "Scale Factor"
                ].to_numpy(dtype=float)
            ),
            "scale_factors_final": scale_factors,
            "row_order_final": row_order,
            "initial_scale_table": initial_scale_table,
            "input_signature": input_signature,
        }


# =============================================================================
# DataFrame builders
# =============================================================================
def solution_dataframe(
    result: dict[str, Any],
) -> pd.DataFrame:
    """Build the final solution table."""

    solution = np.asarray(
        result["solution"],
        dtype=float,
    )

    return pd.DataFrame(
        {
            "Variable": [
                f"x{index + 1}"
                for index in range(len(solution))
            ],
            "Calculated Value": solution,
        }
    )


def scale_factor_dataframe(
    result: dict[str, Any],
) -> pd.DataFrame:
    """Build the initial scale-factor table."""

    initial = result["initial_scale_table"].copy()

    return initial


def pivot_candidate_dataframe(
    result: dict[str, Any],
) -> pd.DataFrame:
    """Build all scaled pivot-candidate comparisons."""

    records = result.get(
        "pivot_candidates",
        [],
    )

    if not records:
        return pd.DataFrame(
            {
                "Message": [
                    "No pivot-candidate analysis was completed."
                ]
            }
        )

    return pd.DataFrame(records)


def operation_dataframe(
    result: dict[str, Any],
) -> pd.DataFrame:
    """Build the full pivot and elimination history."""

    history = result.get("history", [])

    if not history:
        return pd.DataFrame(
            [
                {
                    "Step": 0,
                    "Stage": "Not completed",
                    "Operation": (
                        "No pivot or elimination operation was completed."
                    ),
                }
            ]
        )

    return pd.DataFrame(history)


def back_substitution_dataframe(
    result: dict[str, Any],
) -> pd.DataFrame:
    """Build the back-substitution calculation table."""

    records = result.get(
        "back_substitution",
        [],
    )

    if not records:
        return pd.DataFrame(
            {
                "Message": [
                    "Back substitution was not completed."
                ]
            }
        )

    return pd.DataFrame(records)


def residual_dataframe(
    result: dict[str, Any],
) -> pd.DataFrame:
    """Build equation-by-equation residual verification."""

    residual = np.asarray(
        result["residual_vector"],
        dtype=float,
    )

    return pd.DataFrame(
        {
            "Equation": [
                f"Equation {index + 1}"
                for index in range(len(residual))
            ],
            "Computed Ax": result["computed_b"],
            "Original b": result["original_b"],
            "Residual Ax − b": residual,
            "Absolute Residual": np.abs(residual),
        }
    )


# =============================================================================
# Plot builders
# =============================================================================
def create_matrix_figure(
    result: dict[str, Any],
) -> plt.Figure:
    """Plot the original coefficient matrix."""

    matrix = np.asarray(
        result["original_A"],
        dtype=float,
    )

    figure, axis = plt.subplots(
        figsize=(7.0, 5.0)
    )
    image = axis.imshow(
        matrix,
        aspect="auto",
    )
    figure.colorbar(
        image,
        ax=axis,
        label="Coefficient value",
    )

    axis.set_title("Coefficient Matrix A")
    axis.set_xlabel("Variable Column")
    axis.set_ylabel("Equation Row")
    axis.set_xticks(
        range(matrix.shape[1]),
        [
            f"x{index + 1}"
            for index in range(matrix.shape[1])
        ],
    )
    axis.set_yticks(
        range(matrix.shape[0]),
        [
            f"Eq. {index + 1}"
            for index in range(matrix.shape[0])
        ],
    )

    for row in range(matrix.shape[0]):
        for column in range(matrix.shape[1]):
            axis.text(
                column,
                row,
                f"{matrix[row, column]:.3g}",
                ha="center",
                va="center",
            )

    figure.tight_layout()
    return figure


def create_solution_figure(
    result: dict[str, Any],
) -> plt.Figure:
    """Plot the final solution values."""

    solution = np.asarray(
        result["solution"],
        dtype=float,
    )
    positions = np.arange(len(solution))

    figure, axis = plt.subplots(
        figsize=(7.0, 5.0)
    )
    axis.bar(positions, solution)
    axis.axhline(0.0, linewidth=1)
    axis.set_xticks(
        positions,
        [
            f"x{index + 1}"
            for index in positions
        ],
    )
    axis.set_title(
        "Scaled Partial Pivoting Solution"
    )
    axis.set_xlabel("Variable")
    axis.set_ylabel("Value")
    axis.grid(True, axis="y", alpha=0.3)

    figure.tight_layout()
    return figure


def create_residual_figure(
    result: dict[str, Any],
) -> plt.Figure:
    """Plot absolute residual components."""

    residual = np.abs(
        np.asarray(
            result["residual_vector"],
            dtype=float,
        )
    )
    positions = np.arange(len(residual))

    figure, axis = plt.subplots(
        figsize=(7.0, 5.0)
    )
    axis.bar(positions, residual)
    axis.set_xticks(
        positions,
        [
            f"Eq. {index + 1}"
            for index in positions
        ],
    )
    axis.set_title(
        "Absolute Residual Components"
    )
    axis.set_xlabel("Equation")
    axis.set_ylabel("|Ax − b|")
    axis.grid(True, axis="y", alpha=0.3)

    figure.tight_layout()
    return figure


def figure_to_png_bytes(
    figure: plt.Figure,
) -> bytes:
    """Serialize a matplotlib figure as PNG."""

    buffer = BytesIO()
    figure.savefig(
        buffer,
        format="png",
        dpi=180,
        bbox_inches="tight",
    )
    buffer.seek(0)

    return buffer.getvalue()


# =============================================================================
# Excel export
# =============================================================================
def style_excel_workbook(
    workbook: Any,
) -> None:
    """Apply consistent formatting to every worksheet."""

    header_fill = PatternFill(
        "solid",
        fgColor="0D3151",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for worksheet in workbook.worksheets:
        if (
            worksheet.max_row >= 1
            and worksheet.max_column >= 1
        ):
            worksheet.freeze_panes = "A2"

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = (
                    worksheet.dimensions
                )

        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):
            column_letter = get_column_letter(
                column_index
            )
            maximum_length = 0

            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

                maximum_length = max(
                    maximum_length,
                    len(str(cell.value))
                    if cell.value is not None
                    else 0,
                )

                if isinstance(cell.value, float):
                    cell.number_format = (
                        "0.000000000000E+00"
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(maximum_length + 2, 12),
                55,
            )


def add_excel_image(
    worksheet: Any,
    image_bytes: bytes,
    anchor: str,
    width: int = 760,
    height: int = 500,
) -> None:
    """Insert PNG bytes into an Excel worksheet."""

    image_stream = BytesIO(image_bytes)
    excel_image = ExcelImage(image_stream)
    excel_image.width = width
    excel_image.height = height
    worksheet.add_image(excel_image, anchor)


def create_excel_report(
    result: dict[str, Any],
) -> bytes:
    """Create a complete in-memory XLSX report."""

    if result.get("status") != "success":
        raise ValueError(
            "Only a successful calculation can be exported."
        )

    solution = np.asarray(
        result["solution"],
        dtype=float,
    )
    residual = np.asarray(
        result["residual_vector"],
        dtype=float,
    )
    size = len(solution)

    summary = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "System Classification",
                "Matrix Size",
                "Final Matrix Form",
                "Rank(A)",
                "Rank([A|b])",
                "Pivot Stages",
                "Row Swaps",
                "Recorded Operations",
                "Diagnostic Tolerance",
                "Determinant",
                "Condition Number (∞-norm)",
                "Element Growth Factor",
                "Residual 2-Norm",
                "Residual Infinity Norm",
                "Maximum Absolute Residual",
                "Relative Residual",
                "Final Row Order",
                "Final Scale Factors",
                "Solution Vector",
                "Warnings",
                "Message",
            ],
            "Value": [
                CONFIG["title"],
                result["status"].title(),
                result["classification"],
                f"{size} × {size}",
                result["final_form_name"],
                result["rank_A"],
                result["rank_augmented"],
                size - 1,
                result["row_swaps"],
                result["operations"],
                result["diagnostic_tolerance"],
                result["determinant"],
                result["condition_number"],
                result["growth_factor"],
                result["residual_norm_2"],
                result["residual_norm_inf"],
                result[
                    "maximum_absolute_residual"
                ],
                result["relative_residual"],
                str(
                    result["row_order_final"].tolist()
                ),
                str(
                    result[
                        "scale_factors_final"
                    ].tolist()
                ),
                ", ".join(
                    f"x{index + 1} = "
                    f"{value:.15g}"
                    for index, value
                    in enumerate(solution)
                ),
                (
                    "\n".join(
                        f"• {warning}"
                        for warning
                        in result["warnings"]
                    )
                    if result["warnings"]
                    else "None"
                ),
                result["message"],
            ],
        }
    )

    input_system = pd.DataFrame(
        np.column_stack(
            (
                result["original_A"],
                result["original_b"],
            )
        ),
        columns=[
            f"x{index + 1}"
            for index in range(size)
        ]
        + ["b"],
        index=[
            f"Equation {index + 1}"
            for index in range(size)
        ],
    ).reset_index(names="Equation")

    upper_system = pd.DataFrame(
        np.column_stack(
            (
                result["upper_matrix"],
                result["transformed_vector"],
            )
        ),
        columns=[
            f"x{index + 1}"
            for index in range(size)
        ]
        + ["Transformed RHS"],
        index=[
            f"Row {index + 1}"
            for index in range(size)
        ],
    ).reset_index(names="Row")

    scale_table = scale_factor_dataframe(result)
    pivot_table = pivot_candidate_dataframe(result)
    operations_table = operation_dataframe(result)
    back_table = back_substitution_dataframe(result)
    solution_table = solution_dataframe(result)
    residual_table = residual_dataframe(result)

    verification_table = pd.DataFrame(
        {
            "Metric": [
                "||Ax − b||₂",
                "||Ax − b||∞",
                "max |Ax − b|",
                (
                    "||Ax − b||∞ / "
                    "(||A||∞||x||∞ + ||b||∞)"
                ),
                "κ∞(A)",
                "Element Growth Factor",
                "det(A)",
            ],
            "Value": [
                result["residual_norm_2"],
                result["residual_norm_inf"],
                result[
                    "maximum_absolute_residual"
                ],
                result["relative_residual"],
                result["condition_number"],
                result["growth_factor"],
                result["determinant"],
            ],
            "Interpretation": [
                "Euclidean residual norm",
                "Largest absolute residual component",
                "Maximum equation mismatch",
                "Scale-aware backward-error indicator",
                "Sensitivity indicator",
                "Largest elimination entry / largest original entry",
                "Nonzero for a nonsingular square matrix",
            ],
        }
    )

    chart_data = pd.DataFrame(
        {
            "Variable": [
                f"x{index + 1}"
                for index in range(size)
            ],
            "Solution": solution,
            "Equation": [
                f"Equation {index + 1}"
                for index in range(size)
            ],
            "Absolute Residual": np.abs(residual),
        }
    )

    matrix_figure = create_matrix_figure(result)
    solution_figure = create_solution_figure(result)
    residual_figure = create_residual_figure(result)

    matrix_png = figure_to_png_bytes(
        matrix_figure
    )
    solution_png = figure_to_png_bytes(
        solution_figure
    )
    residual_png = figure_to_png_bytes(
        residual_figure
    )

    plt.close(matrix_figure)
    plt.close(solution_figure)
    plt.close(residual_figure)

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        summary.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )
        input_system.to_excel(
            writer,
            sheet_name="Input System",
            index=False,
        )
        scale_table.to_excel(
            writer,
            sheet_name="Scale Factors",
            index=False,
        )
        pivot_table.to_excel(
            writer,
            sheet_name="Pivot Selection",
            index=False,
        )
        operations_table.to_excel(
            writer,
            sheet_name="Forward Elimination",
            index=False,
        )
        upper_system.to_excel(
            writer,
            sheet_name="Upper Triangular",
            index=False,
        )
        back_table.to_excel(
            writer,
            sheet_name="Back Substitution",
            index=False,
        )
        solution_table.to_excel(
            writer,
            sheet_name="Solution",
            index=False,
        )
        residual_table.to_excel(
            writer,
            sheet_name="Residual Analysis",
            index=False,
        )
        verification_table.to_excel(
            writer,
            sheet_name="Verification",
            index=False,
        )
        chart_data.to_excel(
            writer,
            sheet_name="Chart Data",
            index=False,
        )

        workbook = writer.book
        plots_sheet = workbook.create_sheet(
            "Plots"
        )
        plots_sheet["A1"] = (
            "Scaled Partial Pivoting Report Plots"
        )
        plots_sheet["A1"].font = Font(
            bold=True,
            size=14,
        )

        add_excel_image(
            plots_sheet,
            matrix_png,
            "A3",
        )
        add_excel_image(
            plots_sheet,
            solution_png,
            "A31",
        )
        add_excel_image(
            plots_sheet,
            residual_png,
            "A59",
        )

        data_sheet = workbook["Chart Data"]
        summary_sheet = workbook["Summary"]

        solution_chart = BarChart()
        solution_chart.title = (
            "Final Solution Values"
        )
        solution_chart.y_axis.title = "Value"
        solution_chart.x_axis.title = "Variable"
        solution_chart.height = 8
        solution_chart.width = 15
        solution_chart.add_data(
            Reference(
                data_sheet,
                min_col=2,
                min_row=1,
                max_row=size + 1,
            ),
            titles_from_data=True,
        )
        solution_chart.set_categories(
            Reference(
                data_sheet,
                min_col=1,
                min_row=2,
                max_row=size + 1,
            )
        )
        summary_sheet.add_chart(
            solution_chart,
            "D2",
        )

        residual_chart = BarChart()
        residual_chart.title = (
            "Absolute Residual by Equation"
        )
        residual_chart.y_axis.title = (
            "Absolute Residual"
        )
        residual_chart.x_axis.title = "Equation"
        residual_chart.height = 8
        residual_chart.width = 15
        residual_chart.add_data(
            Reference(
                data_sheet,
                min_col=4,
                min_row=1,
                max_row=size + 1,
            ),
            titles_from_data=True,
        )
        residual_chart.set_categories(
            Reference(
                data_sheet,
                min_col=3,
                min_row=2,
                max_row=size + 1,
            )
        )
        summary_sheet.add_chart(
            residual_chart,
            "D20",
        )

        style_excel_workbook(workbook)
        workbook.active = workbook.sheetnames.index(
            "Summary"
        )

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete scaled-pivoting solver page."""

    st.set_page_config(
        page_title=(
            f"{CONFIG['title']} | Numerical Methods"
        ),
        page_icon="📘",
        layout="wide",
    )

    load_css()
    navbar(active_page="solver")

    st.html(
        f"""
        <section class="solver-hero">
            <div>
                <div class="page-label">
                    {CONFIG['label']}
                </div>
                <h1>{CONFIG['title']}</h1>
                <p>{CONFIG['description']}</p>
                <div class="method-actions">
                    <a
                        href="/{CONFIG['lesson']}"
                        target="_self"
                        class="btn-outline-ui"
                    >
                        Review Lesson →
                    </a>
                    <a
                        href="/{CONFIG['quiz']}"
                        target="_self"
                        class="btn-primary-ui"
                    >
                        Take Quiz →
                    </a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns(
        [0.035, 0.93, 0.035]
    )

    with main_area:
        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader(
                    "How to Enter the System"
                )
                st.markdown(
                    """
                    Choose the number of equations, then enter the
                    coefficients of **A** and the right-hand side
                    **b** in the table.

                    Each row represents one equation and each `x`
                    column represents one unknown coefficient.
                    """
                )
                st.markdown("**Method formula**")
                st.latex(CONFIG["formula"])

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")

                for condition in CONFIG["conditions"]:
                    st.markdown(f"- {condition}")

                st.info(
                    "The scale factors are calculated once from "
                    "the original coefficient rows. They move with "
                    "their rows whenever a row interchange occurs."
                )

        input_column, result_column = st.columns(2)

        with input_column:
            with st.container(border=True):
                st.subheader("Input")

                size = int(
                    st.number_input(
                        "Number of equations",
                        min_value=MIN_SYSTEM_SIZE,
                        max_value=MAX_SYSTEM_SIZE,
                        value=2,
                        step=1,
                        key=(
                            f"{CONFIG['method_id']}_size"
                        ),
                    )
                )

                input_table = st.data_editor(
                    create_default_table(size),
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    key=(
                        f"{CONFIG['method_id']}"
                        f"_table_{size}"
                    ),
                )

                st.caption(
                    "Columns x1 … xn form A; the final "
                    "column is b."
                )

                current_signature = create_input_signature(
                    input_table=input_table,
                    size=size,
                )

                solve_column, reset_column = st.columns(2)

                with solve_column:
                    solve_button = st.button(
                        "Solve",
                        type="primary",
                        use_container_width=True,
                        key=(
                            f"{CONFIG['method_id']}_solve"
                        ),
                    )

                with reset_column:
                    reset_button = st.button(
                        "Reset Result",
                        use_container_width=True,
                        key=(
                            f"{CONFIG['method_id']}_reset"
                        ),
                    )

                if reset_button:
                    st.session_state.pop(
                        f"{CONFIG['method_id']}_result",
                        None,
                    )
                    st.session_state.pop(
                        f"{CONFIG['method_id']}"
                        "_excel_report",
                        None,
                    )
                    st.session_state.pop(
                        f"{CONFIG['method_id']}"
                        "_excel_signature",
                        None,
                    )
                    st.rerun()

                if solve_button:
                    try:
                        matrix, vector = (
                            extract_system_from_table(
                                input_table,
                                size,
                            )
                        )

                        result = solve_linear_system(
                            matrix,
                            vector,
                            input_signature=(
                                current_signature
                            ),
                        )

                    except ValueError as error:
                        result = {
                            "status": "error",
                            "classification": (
                                "Invalid input"
                            ),
                            "message": str(error),
                            "input_signature": (
                                current_signature
                            ),
                        }

                    st.session_state[
                        f"{CONFIG['method_id']}_result"
                    ] = result
                    st.session_state.pop(
                        f"{CONFIG['method_id']}"
                        "_excel_report",
                        None,
                    )
                    st.session_state.pop(
                        f"{CONFIG['method_id']}"
                        "_excel_signature",
                        None,
                    )
                    st.rerun()

        with result_column:
            with st.container(border=True):
                st.subheader("Final Result")

                result = st.session_state.get(
                    f"{CONFIG['method_id']}_result"
                )

                if result is None:
                    st.info(
                        "Enter the system and click Solve."
                    )

                elif (
                    result.get("input_signature")
                    != current_signature
                ):
                    st.info(
                        "The matrix or system size changed. "
                        "Click Solve to calculate a new result."
                    )

                elif result["status"] == "error":
                    st.error(result["message"])
                    st.markdown(
                        "**System classification:** "
                        f"{result.get('classification', 'Not available')}"
                    )

                    if "rank_A" in result:
                        rank_columns = st.columns(2)
                        rank_columns[0].metric(
                            "Rank(A)",
                            result["rank_A"],
                        )
                        rank_columns[1].metric(
                            "Rank([A|b])",
                            result["rank_augmented"],
                        )

                    if "final_matrix" in result:
                        st.markdown(
                            "**Last available transformed system**"
                        )
                        st.code(
                            augmented_text(
                                result["final_matrix"],
                                result["final_vector"],
                            ),
                            language="text",
                        )

                else:
                    st.success(result["message"])
                    st.markdown(
                        "**System classification:** "
                        f"{result['classification']}"
                    )

                    first_metrics = st.columns(2)
                    first_metrics[0].metric(
                        "Residual Norm",
                        format_number(
                            result["residual_norm_inf"]
                        ),
                    )
                    first_metrics[1].metric(
                        "Relative Residual",
                        format_number(
                            result["relative_residual"]
                        ),
                    )

                    second_metrics = st.columns(2)
                    second_metrics[0].metric(
                        "Row Swaps",
                        result["row_swaps"],
                    )
                    second_metrics[1].metric(
                        "Condition Number",
                        format_number(
                            result["condition_number"]
                        ),
                    )

                    st.markdown("**Solution Vector**")
                    st.dataframe(
                        round_numeric_dataframe(
                            solution_dataframe(result)
                        ),
                        use_container_width=True,
                        hide_index=True,
                    )

                    for warning in result["warnings"]:
                        st.warning(warning)

        result = st.session_state.get(
            f"{CONFIG['method_id']}_result"
        )

        result_is_current = (
            result is not None
            and result.get("input_signature")
            == current_signature
        )

        if (
            result_is_current
            and result["status"] == "success"
        ):
            with st.container(border=True):
                st.subheader(
                    "Initial Scale Factors"
                )
                st.dataframe(
                    round_numeric_dataframe(
                        scale_factor_dataframe(result)
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                st.caption(
                    "Each scale factor is sᵢ = maxⱼ|aᵢⱼ| "
                    "and remains attached to its row."
                )

            with st.container(border=True):
                st.subheader(
                    "Scaled Pivot Selection"
                )
                st.dataframe(
                    round_numeric_dataframe(
                        pivot_candidate_dataframe(result)
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                st.caption(
                    "At each stage, the selected row has the "
                    "largest ratio |aᵢₖ| / sᵢ among active rows."
                )

            with st.container(border=True):
                st.subheader(
                    "Forward Elimination History"
                )

                operation_table = operation_dataframe(
                    result
                )

                compact_columns = [
                    "Step",
                    "Stage",
                    "Pivot Column",
                    "Operation",
                    "Selected Original Row",
                    "Pivot Value",
                    "Scale Factor",
                    "Scaled Ratio",
                    "Multiplier",
                ]

                st.dataframe(
                    round_numeric_dataframe(
                        operation_table[
                            [
                                column
                                for column
                                in compact_columns
                                if column
                                in operation_table.columns
                            ]
                        ]
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                with st.expander(
                    "View Intermediate Matrices",
                    expanded=False,
                ):
                    for item in result["history"]:
                        st.markdown(
                            f"**Step {item['Step']}: "
                            f"{item['Operation']}**"
                        )
                        st.code(
                            item[
                                "Augmented Matrix After"
                            ],
                            language="text",
                        )

            with st.container(border=True):
                st.subheader(
                    result["final_form_name"]
                )
                st.code(
                    augmented_text(
                        result["upper_matrix"],
                        result[
                            "transformed_vector"
                        ],
                    ),
                    language="text",
                )

                detail_columns = st.columns(4)
                detail_columns[0].metric(
                    "Rank(A)",
                    result["rank_A"],
                )
                detail_columns[1].metric(
                    "Rank([A|b])",
                    result["rank_augmented"],
                )
                detail_columns[2].metric(
                    "Determinant",
                    format_number(
                        result["determinant"]
                    ),
                )
                detail_columns[3].metric(
                    "Growth Factor",
                    format_number(
                        result["growth_factor"]
                    ),
                )

            with st.container(border=True):
                st.subheader("Back Substitution")
                st.dataframe(
                    round_numeric_dataframe(
                        back_substitution_dataframe(
                            result
                        )
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

            with st.container(border=True):
                st.subheader(
                    "Residual Verification"
                )
                st.dataframe(
                    round_numeric_dataframe(
                        residual_dataframe(result)
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                verification_columns = st.columns(3)
                verification_columns[0].metric(
                    "Residual 2-Norm",
                    format_number(
                        result["residual_norm_2"]
                    ),
                )
                verification_columns[1].metric(
                    "Residual ∞-Norm",
                    format_number(
                        result["residual_norm_inf"]
                    ),
                )
                verification_columns[2].metric(
                    "Maximum |Residual|",
                    format_number(
                        result[
                            "maximum_absolute_residual"
                        ]
                    ),
                )

            with st.container(border=True):
                st.subheader("Excel Report")

                report_key = (
                    f"{CONFIG['method_id']}"
                    "_excel_report"
                )
                signature_key = (
                    f"{CONFIG['method_id']}"
                    "_excel_signature"
                )

                try:
                    if (
                        st.session_state.get(
                            signature_key
                        )
                        != current_signature
                        or report_key
                        not in st.session_state
                    ):
                        st.session_state[
                            report_key
                        ] = create_excel_report(
                            result
                        )
                        st.session_state[
                            signature_key
                        ] = current_signature

                    st.download_button(
                        label=(
                            "Download Complete "
                            "Excel Report"
                        ),
                        data=st.session_state[
                            report_key
                        ],
                        file_name=(
                            "scaled_partial_pivoting"
                            "_complete_report.xlsx"
                        ),
                        mime=EXCEL_MIME_TYPE,
                        use_container_width=True,
                        key=(
                            f"{CONFIG['method_id']}"
                            "_excel_download"
                        ),
                    )

                except (
                    ValueError,
                    OSError,
                    TypeError,
                    ArithmeticError,
                ) as error:
                    st.error(
                        "The Excel report could not "
                        f"be generated. Details: {error}"
                    )

            matrix_column, solution_column, residual_column = (
                st.columns(3)
            )

            with matrix_column:
                with st.container(border=True):
                    st.subheader(
                        "Coefficient Matrix"
                    )
                    figure = create_matrix_figure(
                        result
                    )
                    st.pyplot(
                        figure,
                        use_container_width=True,
                    )
                    plt.close(figure)

            with solution_column:
                with st.container(border=True):
                    st.subheader(
                        "Solution Values"
                    )
                    figure = create_solution_figure(
                        result
                    )
                    st.pyplot(
                        figure,
                        use_container_width=True,
                    )
                    plt.close(figure)

            with residual_column:
                with st.container(border=True):
                    st.subheader(
                        "Residual Components"
                    )
                    figure = create_residual_figure(
                        result
                    )
                    st.pyplot(
                        figure,
                        use_container_width=True,
                    )
                    plt.close(figure)

            with st.container(border=True):
                st.subheader("Continue Learning")
                navigation_column1, navigation_column2 = (
                    st.columns(2)
                )

                with navigation_column1:
                    if st.button(
                        "Review Lesson",
                        use_container_width=True,
                        key=(
                            f"{CONFIG['method_id']}"
                            "_lesson_button"
                        ),
                    ):
                        st.switch_page(
                            f"pages/{CONFIG['lesson']}.py"
                        )

                with navigation_column2:
                    if st.button(
                        "Back to Solver Menu",
                        use_container_width=True,
                        key=(
                            f"{CONFIG['method_id']}"
                            "_menu_button"
                        ),
                    ):
                        st.switch_page(
                            "pages/Numerical_Solver.py"
                        )

    st.html(
        f"""
        <footer class="footer-ui">
            <div>
                NM • © 2026 Numerical Methods
            </div>
            <div>{CONFIG['footer']}</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()

            if reset_button:
                st.session_state.pop(
                    f"{CONFIG['method_id']}_result",
                    None,
                )
                st.rerun()

            if solve_button:
                try:
                    matrix = input_table[
                        [f"x{i + 1}" for i in range(int(size))]
                    ].to_numpy(dtype=float)
                    vector = input_table["b"].to_numpy(dtype=float)

                    st.session_state[
                        f"{CONFIG['method_id']}_result"
                    ] = solve_linear_system(matrix, vector)

                except Exception as error:
                    st.session_state[
                        f"{CONFIG['method_id']}_result"
                    ] = {
                        "status": "error",
                        "message": f"Invalid table values: {error}",
                    }

                st.rerun()

    with result_column:
        with st.container(border=True):
            st.subheader("Final Result")

            result = st.session_state.get(
                f"{CONFIG['method_id']}_result"
            )

            if result is None:
                st.info("Enter the system and click Solve to display the result.")

            elif result["status"] == "error":
                st.error(result["message"])

            else:
                st.success(result["message"])

                metric_column1, metric_column2 = st.columns(2)

                with metric_column1:
                    st.metric(
                        "Residual Norm",
                        format_number(result["residual_norm"]),
                    )

                with metric_column2:
                    st.metric(
                        "Row Operations",
                        result["operations"],
                    )

                metric_column3, metric_column4 = st.columns(2)

                with metric_column3:
                    st.metric(
                        "System Size",
                        len(result["solution"]),
                    )

                with metric_column4:
                    st.metric(
                        "Determinant",
                        format_number(result["determinant"]),
                    )

                solution_frame = pd.DataFrame(
                    {
                        "Variable": [
                            f"x{i + 1}"
                            for i in range(len(result["solution"]))
                        ],
                        "Value": result["solution"],
                    }
                )

                st.markdown("**Solution Vector**")
                st.dataframe(
                    solution_frame.round(3),
                    use_container_width=True,
                    hide_index=True,
                )


    result = st.session_state.get(f"{CONFIG['method_id']}_result")

    if result and result["status"] == "success":
        with st.container(border=True):
            st.subheader(result["final_form_name"])
            st.code(
                augmented_text(
                    result["final_matrix"],
                    result["final_vector"],
                ),
                language="text",
            )

        with st.container(border=True):
            st.subheader("Operation History")

            if not result["history"]:
                st.info("No row operations were required.")
            else:
                history_frame = pd.DataFrame(
                    [
                        {
                            "Step": item["Step"],
                            "Operation": item["Operation"],
                            "Pivot": item["Pivot"],
                        }
                        for item in result["history"]
                    ]
                )

                st.dataframe(
                    history_frame,
                    use_container_width=True,
                    hide_index=True,
                )

                with st.expander("View Intermediate Matrices", expanded=False):
                    for item in result["history"]:
                        st.markdown(
                            f"**Step {item['Step']}: {item['Operation']}**"
                        )
                        st.code(
                            item["Augmented Matrix"],
                            language="text",
                        )

                solution_download = pd.DataFrame(
                    {
                        "Variable": [
                            f"x{i + 1}"
                            for i in range(len(result["solution"]))
                        ],
                        "Value": result["solution"],
                    }
                )
                st.subheader("Excel Report")
                excel_report = create_excel_report(result)
                st.download_button(
                    "Download Complete Excel Report",
                    data=excel_report,
                    file_name=f"{CONFIG['method_id']}_complete_report.xlsx",
                    mime=EXCEL_MIME_TYPE,
                    use_container_width=True,
                    key=f"{CONFIG['method_id']}_excel_report",
                )

        matrix_column, solution_column, residual_column = st.columns(3)

        with matrix_column:
            with st.container(border=True):
                st.subheader("Coefficient Matrix")
                figure, axis = plt.subplots(figsize=(4.2, 3.0))
                image = axis.imshow(result["original_A"], aspect="auto")
                axis.set_xlabel("Column")
                axis.set_ylabel("Row")
                axis.set_xticks(range(result["original_A"].shape[1]))
                axis.set_yticks(range(result["original_A"].shape[0]))
                st.pyplot(figure, use_container_width=True)
                plt.close(figure)

        with solution_column:
            with st.container(border=True):
                st.subheader("Solution Values")
                figure, axis = plt.subplots(figsize=(4.2, 3.0))
                positions = np.arange(len(result["solution"]))
                axis.bar(positions, result["solution"])
                axis.set_xticks(
                    positions,
                    [f"x{i + 1}" for i in positions],
                )
                axis.set_ylabel("Value")
                axis.grid(True, axis="y")
                st.pyplot(figure, use_container_width=True)
                plt.close(figure)

        with residual_column:
            with st.container(border=True):
                st.subheader("Residual Components")
                figure, axis = plt.subplots(figsize=(4.2, 3.0))
                positions = np.arange(len(result["residual_vector"]))
                axis.bar(positions, np.abs(result["residual_vector"]))
                axis.set_xticks(
                    positions,
                    [f"Eq. {i + 1}" for i in positions],
                )
                axis.set_ylabel("|Ax − b|")
                axis.grid(True, axis="y")
                st.pyplot(figure, use_container_width=True)
                plt.close(figure)

        with st.container(border=True):
            st.subheader("Continue Learning")
            navigation_column1, navigation_column2 = st.columns(2)

            with navigation_column1:
                if st.button(
                    "Review Lesson",
                    use_container_width=True,
                    key=f"{CONFIG['method_id']}_lesson_button",
                ):
                    st.switch_page(f"pages/{CONFIG['lesson']}.py")

            with navigation_column2:
                if st.button(
                    "Back to Solver Menu",
                    use_container_width=True,
                    key=f"{CONFIG['method_id']}_menu_button",
                ):
                    st.switch_page("pages/Numerical_Solver.py")


st.html(
    f"""
    <footer class="footer-ui">
        <div>NM • © 2026 Numerical Methods</div>
        <div>{CONFIG['footer']}</div>
    </footer>
    """
)
