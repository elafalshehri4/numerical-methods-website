from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Sequence
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css


# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0**exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Show fixed notation or ×10ⁿ notation when that is clearer."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6
    ):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


def format_number(
    value: float | int | None,
    decimals: int = 3,
) -> str:
    """Format a displayed numerical value."""

    return format_display_number(value, decimals)


def scientific_number(value: float | int | None) -> str:
    """Format a value in scientific notation."""

    return format_scientific_power(value)


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Backward Difference Method"
DISPLAY_DECIMALS = 3
DEFAULT_FUNCTION = "x**3 - 2*x + 1"
DEFAULT_X_VALUE = 1.0
DEFAULT_STEP_SIZE = 0.1
DEFAULT_REFINEMENT_LEVELS = 5
MIN_REFINEMENT_LEVELS = 2
MAX_REFINEMENT_LEVELS = 10
ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DERIVATIVE_OPTIONS = {
    "First derivative f′(x₀)": 1,
    "Second derivative f″(x₀)": 2,
    "Third derivative f‴(x₀)": 3,
}

FORMULA_OPTIONS = {
    "Standard backward formula — first-order O(h)": "standard",
    "Higher-accuracy backward formula — second-order O(h²)": "high_accuracy",
}

DERIVATIVE_SYMBOLS = {
    1: "f′(x₀)",
    2: "f″(x₀)",
    3: "f‴(x₀)",
}

ALLOWED_FUNCTION_NAMES = {
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class FormulaMetadata:
    """Definition of one backward finite-difference formula."""

    derivative_order: int
    formula_key: str
    formula_text: str
    theoretical_order: int
    denominator_factor: float
    coefficients: tuple[tuple[int, float], ...]

    @property
    def maximum_backward_offset(self) -> int:
        return max(-offset for offset, _ in self.coefficients)


@dataclass(frozen=True)
class TaylorSeriesRow:
    """Taylor expansion information for one backward sample point."""

    offset_number: int
    point: float
    actual_value: float
    expansion_text: str
    third_degree_taylor_value: float | None
    third_degree_remainder: float | None


@dataclass(frozen=True)
class DifferenceEvaluation:
    """One backward-difference calculation at a specific step size."""

    level: int
    step_size: float
    x_value: float
    derivative_order: int
    sample_points: tuple[float, ...]
    sample_values: tuple[float, ...]
    numerator: float
    denominator: float
    derivative_approximation: float
    exact_derivative: float | None
    absolute_error: float | None
    relative_error_percent: float | None
    successive_difference: float | None
    observed_order: float | None
    formula_text: str
    substitution_text: str


@dataclass(frozen=True)
class BackwardDifferenceResult:
    """Complete solver result shared by the page and Excel export."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str
    function_text: str
    function_expression: sp.Expr | None
    derivative_order: int
    derivative_label: str
    derivative_symbol: str
    derivative_expression: sp.Expr | None
    derivative_expression_text: str
    formula_key: str
    formula_name: str
    formula_text: str
    theoretical_order: int
    maximum_backward_offset: int
    x_value: float | None
    initial_step_size: float | None
    refinement_levels: int
    evaluations: tuple[DifferenceEvaluation, ...]
    taylor_series_rows: tuple[TaylorSeriesRow, ...]
    primary_approximation: float | None
    primary_exact_derivative: float | None
    primary_absolute_error: float | None
    primary_relative_error_percent: float | None
    finest_step_size: float | None
    finest_approximation: float | None
    finest_absolute_error: float | None
    observed_order_latest: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only for display."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def create_input_signature(
    function_text: str,
    x_value: Any,
    step_size: Any,
    derivative_name: str,
    formula_name: str,
    refinement_levels: Any,
) -> str:
    """Create a stable signature used to prevent stale Streamlit results."""

    payload = repr(
        (
            str(function_text).strip(),
            str(x_value),
            str(step_size),
            str(derivative_name),
            str(formula_name),
            str(refinement_levels),
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def safe_float(raw_value: Any, value_name: str) -> float:
    """Convert one user input to a finite real floating-point value."""

    if raw_value is None:
        raise ValueError(f"{value_name} is required.")
    if isinstance(raw_value, str) and not raw_value.strip():
        raise ValueError(f"{value_name} is required.")
    try:
        value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be a valid numerical value.") from error
    if not math.isfinite(value):
        raise ValueError(
            f"{value_name} must be finite; NaN and infinity are not allowed."
        )
    return value


def human_readable_expression(expression: sp.Expr | None) -> str:
    """Return a compact readable mathematical expression."""

    if expression is None:
        return "Not available"
    return str(sp.simplify(expression))


# =============================================================================
# Function parsing and safe evaluation
# =============================================================================
def parse_function(function_text: str) -> tuple[sp.Expr, sp.Symbol]:
    """Parse and validate a real single-variable mathematical function."""

    if not isinstance(function_text, str) or not function_text.strip():
        raise ValueError("The function input cannot be empty.")

    x_symbol = sp.Symbol("x", real=True)
    local_dictionary = dict(ALLOWED_FUNCTION_NAMES)
    local_dictionary["x"] = x_symbol

    try:
        expression = sp.sympify(
            function_text.strip(),
            locals=local_dictionary,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The function could not be parsed. Use valid SymPy syntax, for "
            "example: sin(x), exp(x), log(x), or x**3 - 2*x + 1."
        ) from error

    if isinstance(expression, (sp.MatrixBase, Relational)):
        raise ValueError(
            "The input must be a scalar function of x, not a matrix or relation."
        )

    unexpected_symbols = expression.free_symbols.difference({x_symbol})
    if unexpected_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unexpected_symbols))
        raise ValueError(
            "The function may contain only the variable x. "
            f"Unexpected symbol(s): {names}."
        )

    if expression.has(sp.zoo, sp.nan, sp.oo, -sp.oo):
        raise ValueError("The function contains an undefined or non-finite constant.")

    if expression.atoms(AppliedUndef):
        raise ValueError("The function contains an unsupported undefined function.")

    return sp.simplify(expression), x_symbol


def create_numeric_function(
    expression: sp.Expr,
    x_symbol: sp.Symbol,
) -> Callable[[Any], Any]:
    """Create a NumPy-compatible numerical function."""

    try:
        return sp.lambdify(x_symbol, expression, modules=["numpy"])
    except (TypeError, ValueError) as error:
        raise ValueError("The function could not be converted to numerical form.") from error


def evaluate_real_scalar(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    description: str,
) -> float:
    """Evaluate a numerical function and require a finite real scalar."""

    try:
        with np.errstate(all="ignore"):
            raw_value = numeric_function(x_value)
    except (ArithmeticError, TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The function could not be evaluated at x = {x_value:.12g} "
            f"while calculating {description}."
        ) from error

    array = np.asarray(raw_value)
    if array.size != 1:
        raise ValueError(f"{description} did not produce a scalar value.")

    scalar = array.reshape(-1)[0]
    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)
        if abs(complex_value.imag) > ZERO_TOLERANCE:
            raise ValueError(
                f"{description} is complex at x = {x_value:.12g}; "
                "this solver accepts real-valued functions only."
            )
        scalar = complex_value.real

    try:
        numeric_value = float(scalar)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{description} could not be converted to a real number.") from error

    if not math.isfinite(numeric_value):
        raise ValueError(
            f"{description} is undefined, NaN, or infinite at x = {x_value:.12g}."
        )
    return numeric_value


def evaluate_real_array(
    numeric_function: Callable[[Any], Any],
    x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate a function over an array for plotting, returning NaN if invalid."""

    with np.errstate(all="ignore"):
        try:
            raw_values = numeric_function(x_values)
        except Exception:
            return np.full_like(x_values, np.nan, dtype=float)

    array = np.asarray(raw_values)
    if array.ndim == 0:
        array = np.full_like(
            x_values,
            array,
            dtype=complex if np.iscomplexobj(array) else float,
        )
    try:
        array = np.broadcast_to(array, x_values.shape)
    except ValueError:
        return np.full_like(x_values, np.nan, dtype=float)

    if np.iscomplexobj(array):
        valid_imaginary = np.abs(np.imag(array)) <= ZERO_TOLERANCE
        real_values = np.real(array).astype(float)
        real_values[~valid_imaginary] = np.nan
    else:
        try:
            real_values = array.astype(float)
        except (TypeError, ValueError):
            return np.full_like(x_values, np.nan, dtype=float)

    real_values[~np.isfinite(real_values)] = np.nan
    return real_values


# =============================================================================
# Backward finite-difference formulas
# =============================================================================
def formula_metadata(
    derivative_order: int,
    formula_key: str,
) -> FormulaMetadata:
    """Return the textbook backward formula for derivatives one through three."""

    formulas: dict[tuple[int, str], FormulaMetadata] = {
        (1, "standard"): FormulaMetadata(
            derivative_order=1,
            formula_key="standard",
            formula_text="f′(x₀) ≈ [f(x₀) − f(x₀−h)] / h",
            theoretical_order=1,
            denominator_factor=1.0,
            coefficients=((0, 1.0), (-1, -1.0)),
        ),
        (1, "high_accuracy"): FormulaMetadata(
            derivative_order=1,
            formula_key="high_accuracy",
            formula_text=(
                "f′(x₀) ≈ [3f(x₀) − 4f(x₀−h) + f(x₀−2h)] / (2h)"
            ),
            theoretical_order=2,
            denominator_factor=2.0,
            coefficients=((0, 3.0), (-1, -4.0), (-2, 1.0)),
        ),
        (2, "standard"): FormulaMetadata(
            derivative_order=2,
            formula_key="standard",
            formula_text=(
                "f″(x₀) ≈ [f(x₀) − 2f(x₀−h) + f(x₀−2h)] / h²"
            ),
            theoretical_order=1,
            denominator_factor=1.0,
            coefficients=((0, 1.0), (-1, -2.0), (-2, 1.0)),
        ),
        (2, "high_accuracy"): FormulaMetadata(
            derivative_order=2,
            formula_key="high_accuracy",
            formula_text=(
                "f″(x₀) ≈ [2f(x₀) − 5f(x₀−h) + 4f(x₀−2h) "
                "− f(x₀−3h)] / h²"
            ),
            theoretical_order=2,
            denominator_factor=1.0,
            coefficients=((0, 2.0), (-1, -5.0), (-2, 4.0), (-3, -1.0)),
        ),
        (3, "standard"): FormulaMetadata(
            derivative_order=3,
            formula_key="standard",
            formula_text=(
                "f‴(x₀) ≈ [f(x₀) − 3f(x₀−h) + 3f(x₀−2h) "
                "− f(x₀−3h)] / h³"
            ),
            theoretical_order=1,
            denominator_factor=1.0,
            coefficients=((0, 1.0), (-1, -3.0), (-2, 3.0), (-3, -1.0)),
        ),
        (3, "high_accuracy"): FormulaMetadata(
            derivative_order=3,
            formula_key="high_accuracy",
            formula_text=(
                "f‴(x₀) ≈ [5f(x₀) − 18f(x₀−h) + 24f(x₀−2h) "
                "− 14f(x₀−3h) + 3f(x₀−4h)] / (2h³)"
            ),
            theoretical_order=2,
            denominator_factor=2.0,
            coefficients=(
                (0, 5.0),
                (-1, -18.0),
                (-2, 24.0),
                (-3, -14.0),
                (-4, 3.0),
            ),
        ),
    }

    metadata = formulas.get((derivative_order, formula_key))
    if metadata is None:
        raise ValueError("Unsupported Backward Difference formula selection.")
    return metadata


def coefficient_term_text(coefficient: float, value: float) -> str:
    """Format one coefficient-times-function-value term for substitution."""

    if coefficient == 1.0:
        return f"({format_number(value, 10)})"
    if coefficient == -1.0:
        return f"− ({format_number(value, 10)})"
    sign = "+" if coefficient > 0 else "−"
    return f"{sign} {abs(coefficient):g}({format_number(value, 10)})"


def calculate_backward_difference(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    step_size: float,
    metadata: FormulaMetadata,
) -> dict[str, Any]:
    """Calculate one backward finite-difference approximation manually."""

    sample_points: list[float] = []
    sample_values: list[float] = []

    for offset, _ in metadata.coefficients:
        point = x_value + offset * step_size
        if not math.isfinite(point):
            raise ValueError(
                "A required backward point is not finite. Reduce the magnitude of x₀ or h."
            )
        description = "f(x₀)" if offset == 0 else f"f(x₀ − {-offset}h)"
        value = evaluate_real_scalar(numeric_function, point, description)
        sample_points.append(point)
        sample_values.append(value)

    numerator = sum(
        coefficient * value
        for (_, coefficient), value in zip(metadata.coefficients, sample_values)
    )
    denominator = metadata.denominator_factor * step_size ** metadata.derivative_order
    derivative = numerator / denominator

    if not all(
        math.isfinite(float(value))
        for value in (numerator, denominator, derivative)
    ):
        raise ValueError(
            "The Backward Difference calculation produced a non-finite value."
        )

    substitution_terms: list[str] = []
    for index, ((_, coefficient), value) in enumerate(
        zip(metadata.coefficients, sample_values)
    ):
        term = coefficient_term_text(coefficient, value)
        if index == 0 and term.startswith("+ "):
            term = term[2:]
        substitution_terms.append(term)

    denominator_text = format_number(denominator, 10)
    substitution = f"[{' '.join(substitution_terms)}] / ({denominator_text})"

    return {
        "sample_points": tuple(sample_points),
        "sample_values": tuple(sample_values),
        "numerator": float(numerator),
        "denominator": float(denominator),
        "derivative": float(derivative),
        "substitution": substitution,
    }


def safe_relative_error_percent(
    absolute_error: float,
    exact_value: float,
) -> float | None:
    """Calculate the absolute true relative percent error safely."""

    if abs(exact_value) <= RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
        return None
    return absolute_error / abs(exact_value) * 100.0


def calculate_observed_order(
    previous_error: float | None,
    current_error: float | None,
    refinement_ratio: float = 2.0,
) -> float | None:
    """Estimate convergence order from two consecutive exact errors."""

    if previous_error is None or current_error is None:
        return None
    if previous_error <= 0.0 or current_error <= 0.0:
        return None
    try:
        order = math.log(previous_error / current_error) / math.log(
            refinement_ratio
        )
    except (ValueError, ZeroDivisionError):
        return None
    return order if math.isfinite(order) else None


def calculate_successive_order(
    difference_older: float | None,
    difference_newer: float | None,
    refinement_ratio: float = 2.0,
) -> float | None:
    """Estimate order from three successive numerical approximations."""

    if difference_older is None or difference_newer is None:
        return None
    if difference_older <= 0.0 or difference_newer <= 0.0:
        return None
    try:
        order = math.log(difference_older / difference_newer) / math.log(
            refinement_ratio
        )
    except (ValueError, ZeroDivisionError):
        return None
    return order if math.isfinite(order) else None


def taylor_expansion_text(offset_number: int) -> str:
    """Return the backward Taylor expansion through the third derivative."""

    k = offset_number
    return (
        f"f(x₀−{k}h) = f(x₀) − {k}h f′(x₀) + "
        f"({k**2}h²/2) f″(x₀) − ({k**3}h³/6) f‴(x₀) + O(h⁴)"
    )


def build_taylor_series_rows(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    step_size: float,
    maximum_offset: int,
    derivative_values: dict[int, float | None],
) -> tuple[TaylorSeriesRow, ...]:
    """Build Taylor-series rows through third degree for required points."""

    f_x = evaluate_real_scalar(numeric_function, x_value, "f(x₀)")
    rows: list[TaylorSeriesRow] = []

    for k in range(1, maximum_offset + 1):
        point = x_value - k * step_size
        actual = evaluate_real_scalar(numeric_function, point, f"f(x₀ − {k}h)")

        d1 = derivative_values.get(1)
        d2 = derivative_values.get(2)
        d3 = derivative_values.get(3)
        if d1 is None or d2 is None or d3 is None:
            taylor_value = None
            remainder = None
        else:
            taylor_value = (
                f_x
                - k * step_size * d1
                + (k * step_size) ** 2 * d2 / 2.0
                - (k * step_size) ** 3 * d3 / 6.0
            )
            remainder = actual - taylor_value

        rows.append(
            TaylorSeriesRow(
                offset_number=k,
                point=point,
                actual_value=actual,
                expansion_text=taylor_expansion_text(k),
                third_degree_taylor_value=taylor_value,
                third_degree_remainder=remainder,
            )
        )

    return tuple(rows)


def empty_error_result(
    *,
    message: str,
    function_text: str,
    derivative_order: int,
    derivative_label: str,
    formula_key: str,
    formula_name: str,
    input_signature: str,
    execution_datetime: datetime,
) -> BackwardDifferenceResult:
    """Create a consistent unsuccessful result."""

    try:
        metadata = formula_metadata(derivative_order, formula_key)
    except ValueError:
        metadata = formula_metadata(1, "standard")

    return BackwardDifferenceResult(
        status="error",
        success=False,
        method=METHOD_NAME,
        message=message,
        stopping_reason=(
            "The calculation stopped during input validation or function evaluation."
        ),
        function_text=str(function_text).strip(),
        function_expression=None,
        derivative_order=derivative_order,
        derivative_label=derivative_label,
        derivative_symbol=DERIVATIVE_SYMBOLS.get(derivative_order, "f′(x₀)"),
        derivative_expression=None,
        derivative_expression_text="Not available",
        formula_key=metadata.formula_key,
        formula_name=formula_name,
        formula_text=metadata.formula_text,
        theoretical_order=metadata.theoretical_order,
        maximum_backward_offset=metadata.maximum_backward_offset,
        x_value=None,
        initial_step_size=None,
        refinement_levels=0,
        evaluations=tuple(),
        taylor_series_rows=tuple(),
        primary_approximation=None,
        primary_exact_derivative=None,
        primary_absolute_error=None,
        primary_relative_error_percent=None,
        finest_step_size=None,
        finest_approximation=None,
        finest_absolute_error=None,
        observed_order_latest=None,
        warnings=tuple(),
        input_signature=input_signature,
        execution_datetime=execution_datetime,
    )


def solve_backward_difference(
    function_text: str,
    raw_x_value: Any,
    raw_step_size: Any,
    derivative_name: str,
    formula_name: str,
    raw_refinement_levels: Any,
    input_signature: str,
) -> BackwardDifferenceResult:
    """Validate inputs and perform the complete Backward Difference workflow."""

    execution_datetime = current_report_datetime()
    derivative_order = DERIVATIVE_OPTIONS.get(derivative_name, 1)
    formula_key = FORMULA_OPTIONS.get(formula_name, "standard")

    try:
        if derivative_name not in DERIVATIVE_OPTIONS:
            raise ValueError("Select a valid derivative order.")
        if formula_name not in FORMULA_OPTIONS:
            raise ValueError("Select a valid Backward Difference formula.")

        x_value = safe_float(raw_x_value, "Evaluation point x₀")
        initial_step_size = safe_float(raw_step_size, "Step size h")
        if initial_step_size <= 0.0:
            raise ValueError("Step size h must be greater than zero.")
        if initial_step_size <= ZERO_TOLERANCE:
            raise ValueError(
                "Step size h is too close to machine precision. Choose a larger value."
            )

        try:
            refinement_levels = int(raw_refinement_levels)
        except (TypeError, ValueError) as error:
            raise ValueError(
                "The number of refinement levels must be an integer."
            ) from error
        if not MIN_REFINEMENT_LEVELS <= refinement_levels <= MAX_REFINEMENT_LEVELS:
            raise ValueError(
                f"Refinement levels must be between {MIN_REFINEMENT_LEVELS} "
                f"and {MAX_REFINEMENT_LEVELS}."
            )

        metadata = formula_metadata(derivative_order, formula_key)
        expression, x_symbol = parse_function(function_text)
        numeric_function = create_numeric_function(expression, x_symbol)

        derivative_expressions: dict[int, sp.Expr | None] = {}
        derivative_values: dict[int, float | None] = {}
        derivative_warnings: list[str] = []

        for order in range(1, 4):
            try:
                derivative_expression = sp.simplify(
                    sp.diff(expression, x_symbol, order)
                )
                if derivative_expression.has(sp.zoo, sp.nan, sp.oo, -sp.oo):
                    raise ValueError("The symbolic derivative is non-finite.")
                derivative_numeric = create_numeric_function(
                    derivative_expression,
                    x_symbol,
                )
                derivative_value = evaluate_real_scalar(
                    derivative_numeric,
                    x_value,
                    f"the exact derivative of order {order} at x₀",
                )
                derivative_expressions[order] = derivative_expression
                derivative_values[order] = derivative_value
            except (ValueError, TypeError, NotImplementedError) as error:
                derivative_expressions[order] = None
                derivative_values[order] = None
                derivative_warnings.append(
                    f"The exact derivative of order {order} could not be evaluated "
                    f"at x₀. Details: {error}"
                )

        selected_derivative_expression = derivative_expressions[derivative_order]
        exact_derivative = derivative_values[derivative_order]

        evaluations: list[DifferenceEvaluation] = []
        approximations: list[float] = []
        exact_errors: list[float | None] = []
        successive_differences: list[float | None] = []

        for level in range(refinement_levels):
            step_size = initial_step_size / (2.0**level)
            if step_size <= ZERO_TOLERANCE or not math.isfinite(step_size):
                raise ValueError(
                    "Step-size refinement became numerically unusable. "
                    "Reduce the number of refinement levels."
                )

            calculation = calculate_backward_difference(
                numeric_function=numeric_function,
                x_value=x_value,
                step_size=step_size,
                metadata=metadata,
            )
            approximation = float(calculation["derivative"])

            if exact_derivative is not None:
                absolute_error = abs(approximation - exact_derivative)
                relative_error_percent = safe_relative_error_percent(
                    absolute_error,
                    exact_derivative,
                )
            else:
                absolute_error = None
                relative_error_percent = None

            successive_difference = (
                abs(approximation - approximations[-1])
                if approximations
                else None
            )

            if exact_derivative is not None and exact_errors:
                observed_order = calculate_observed_order(
                    exact_errors[-1],
                    absolute_error,
                )
            elif len(successive_differences) >= 1:
                observed_order = calculate_successive_order(
                    successive_differences[-1],
                    successive_difference,
                )
            else:
                observed_order = None

            evaluations.append(
                DifferenceEvaluation(
                    level=level,
                    step_size=step_size,
                    x_value=x_value,
                    derivative_order=derivative_order,
                    sample_points=tuple(calculation["sample_points"]),
                    sample_values=tuple(calculation["sample_values"]),
                    numerator=float(calculation["numerator"]),
                    denominator=float(calculation["denominator"]),
                    derivative_approximation=approximation,
                    exact_derivative=exact_derivative,
                    absolute_error=absolute_error,
                    relative_error_percent=relative_error_percent,
                    successive_difference=successive_difference,
                    observed_order=observed_order,
                    formula_text=metadata.formula_text,
                    substitution_text=str(calculation["substitution"]),
                )
            )
            approximations.append(approximation)
            exact_errors.append(absolute_error)
            successive_differences.append(successive_difference)

        taylor_rows = build_taylor_series_rows(
            numeric_function=numeric_function,
            x_value=x_value,
            step_size=initial_step_size,
            maximum_offset=metadata.maximum_backward_offset,
            derivative_values=derivative_values,
        )

        warnings: list[str] = list(derivative_warnings)
        primary = evaluations[0]
        finest = evaluations[-1]

        if max(primary.sample_values) - min(primary.sample_values) <= ZERO_TOLERANCE:
            warnings.append(
                "All required sampled function values are numerically identical. "
                "The derivative approximation may be zero because of local behavior "
                "or floating-point cancellation."
            )

        if finest.absolute_error is not None and primary.absolute_error is not None:
            if finest.absolute_error > primary.absolute_error:
                warnings.append(
                    "The exact error increased after step refinement. Round-off or "
                    "subtractive cancellation may be affecting the calculation."
                )

        roundoff_threshold = np.finfo(float).eps ** (
            1.0 / (derivative_order + metadata.theoretical_order)
        ) * max(1.0, abs(x_value))
        if finest.step_size < roundoff_threshold:
            warnings.append(
                "The finest step is very small relative to x₀. Backward differences "
                "may suffer from subtractive cancellation at very small h."
            )

        observed_orders = [
            item.observed_order
            for item in evaluations
            if item.observed_order is not None
            and math.isfinite(item.observed_order)
        ]
        latest_order = observed_orders[-1] if observed_orders else None

        return BackwardDifferenceResult(
            status="success",
            success=True,
            method=METHOD_NAME,
            message="Execution completed successfully.",
            stopping_reason=(
                "The requested Backward Difference approximation and step-size "
                "refinement analysis were completed."
            ),
            function_text=function_text.strip(),
            function_expression=expression,
            derivative_order=derivative_order,
            derivative_label=derivative_name,
            derivative_symbol=DERIVATIVE_SYMBOLS[derivative_order],
            derivative_expression=selected_derivative_expression,
            derivative_expression_text=human_readable_expression(
                selected_derivative_expression
            ),
            formula_key=formula_key,
            formula_name=formula_name,
            formula_text=metadata.formula_text,
            theoretical_order=metadata.theoretical_order,
            maximum_backward_offset=metadata.maximum_backward_offset,
            x_value=x_value,
            initial_step_size=initial_step_size,
            refinement_levels=refinement_levels,
            evaluations=tuple(evaluations),
            taylor_series_rows=taylor_rows,
            primary_approximation=primary.derivative_approximation,
            primary_exact_derivative=primary.exact_derivative,
            primary_absolute_error=primary.absolute_error,
            primary_relative_error_percent=primary.relative_error_percent,
            finest_step_size=finest.step_size,
            finest_approximation=finest.derivative_approximation,
            finest_absolute_error=finest.absolute_error,
            observed_order_latest=latest_order,
            warnings=tuple(warnings),
            input_signature=input_signature,
            execution_datetime=execution_datetime,
        )

    except ValueError as error:
        return empty_error_result(
            message=str(error),
            function_text=function_text,
            derivative_order=derivative_order,
            derivative_label=derivative_name,
            formula_key=formula_key,
            formula_name=formula_name,
            input_signature=input_signature,
            execution_datetime=execution_datetime,
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def sample_point_columns(item: DifferenceEvaluation) -> dict[str, Any]:
    """Convert sampled backward points into consistent table columns."""

    columns: dict[str, Any] = {}
    for offset_number in range(0, 5):
        point_name = "x0" if offset_number == 0 else f"x0 - {offset_number}h"
        value_name = "f(x0)" if offset_number == 0 else f"f(x0 - {offset_number}h)"

        if offset_number < len(item.sample_points):
            columns[point_name] = item.sample_points[offset_number]
            columns[value_name] = item.sample_values[offset_number]
        else:
            columns[point_name] = None
            columns[value_name] = None
    return columns


def create_evaluation_dataframe(result: BackwardDifferenceResult) -> pd.DataFrame:
    """Build the complete refinement and error-analysis table."""

    rows: list[dict[str, Any]] = []
    for item in result.evaluations:
        row: dict[str, Any] = {
            "Level": item.level,
            "Derivative Order": item.derivative_order,
            "h": item.step_size,
        }
        row.update(sample_point_columns(item))
        row.update(
            {
                "Numerator": item.numerator,
                "Denominator": item.denominator,
                "Approximate Derivative": item.derivative_approximation,
                "Exact Derivative": item.exact_derivative,
                "Absolute Error": item.absolute_error,
                "Absolute Relative Error (%)": item.relative_error_percent,
                "Successive Difference": item.successive_difference,
                "Observed Order": item.observed_order,
            }
        )
        rows.append(row)
    return pd.DataFrame(rows)


def create_primary_calculation_dataframe(
    result: BackwardDifferenceResult,
) -> pd.DataFrame:
    """Build a vertical table describing the calculation at the entered h."""

    if not result.evaluations:
        return pd.DataFrame()

    item = result.evaluations[0]
    rows: list[tuple[str, Any]] = [
        ("Function", result.function_text),
        ("Requested Derivative", result.derivative_label),
        ("Exact Derivative Expression", result.derivative_expression_text),
        ("Formula", result.formula_text),
        ("x₀", item.x_value),
        ("h", item.step_size),
    ]

    for offset_number, (point, value) in enumerate(
        zip(item.sample_points, item.sample_values)
    ):
        if offset_number == 0:
            rows.extend([("x₀", point), ("f(x₀)", value)])
        else:
            rows.extend(
                [
                    (f"x₀ − {offset_number}h", point),
                    (f"f(x₀ − {offset_number}h)", value),
                ]
            )

    rows.extend(
        [
            ("Numerator", item.numerator),
            ("Denominator", item.denominator),
            ("Approximate Derivative", item.derivative_approximation),
            ("Exact Derivative", item.exact_derivative),
            ("Absolute Error", item.absolute_error),
            ("Absolute Relative Error (%)", item.relative_error_percent),
        ]
    )
    return pd.DataFrame(rows, columns=["Property", "Value"])


def create_error_dataframe(result: BackwardDifferenceResult) -> pd.DataFrame:
    """Build a focused error and convergence DataFrame."""

    evaluation_dataframe = create_evaluation_dataframe(result)
    selected_columns = [
        "Level",
        "Derivative Order",
        "h",
        "Approximate Derivative",
        "Exact Derivative",
        "Absolute Error",
        "Absolute Relative Error (%)",
        "Successive Difference",
        "Observed Order",
    ]
    return evaluation_dataframe[selected_columns].copy()


def create_taylor_series_dataframe(result: BackwardDifferenceResult) -> pd.DataFrame:
    """Build the Taylor-series table through the third derivative."""

    return pd.DataFrame(
        [
            {
                "Backward Point": f"x0 - {row.offset_number}h",
                "x Value": row.point,
                "Actual f(x)": row.actual_value,
                "Taylor Expansion Through Third Degree": row.expansion_text,
                "Third-Degree Taylor Value": row.third_degree_taylor_value,
                "Remainder (Actual - Taylor)": row.third_degree_remainder,
            }
            for row in result.taylor_series_rows
        ]
    )


# =============================================================================
# Plot builders
# =============================================================================
def create_function_plot(result: BackwardDifferenceResult) -> plt.Figure:
    """Plot the function and all points used by the selected formula."""

    if not result.success or result.function_expression is None:
        raise ValueError("A successful result is required to create the function graph.")

    x_symbol = sp.Symbol("x", real=True)
    numeric_function = create_numeric_function(result.function_expression, x_symbol)
    primary = result.evaluations[0]

    required_span = result.maximum_backward_offset * primary.step_size
    half_width = max(
        2.5 * required_span,
        1.0,
        0.25 * max(1.0, abs(primary.x_value)),
    )
    x_min = primary.x_value - max(half_width, required_span * 1.2)
    x_max = primary.x_value + half_width
    x_values = np.linspace(x_min, x_max, 600)
    y_values = evaluate_real_array(numeric_function, x_values)

    if np.count_nonzero(np.isfinite(y_values)) < 2:
        raise ValueError("The function could not be plotted near x₀.")

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.plot(
        x_values,
        y_values,
        linewidth=2,
        label=f"f(x) = {result.function_text}",
    )
    axis.axhline(0.0, linewidth=1)
    axis.axvline(primary.x_value, linestyle=":", linewidth=1.5, label="x = x₀")
    axis.scatter(
        primary.sample_points,
        primary.sample_values,
        s=75,
        label="Backward formula sample points",
        zorder=5,
    )

    if result.derivative_order == 1:
        f_x0 = primary.sample_values[0]
        tangent_values = f_x0 + primary.derivative_approximation * (
            x_values - primary.x_value
        )
        axis.plot(
            x_values,
            tangent_values,
            linestyle="--",
            linewidth=2,
            label=(
                "Approximate tangent: slope = "
                f"{primary.derivative_approximation:.6g}"
            ),
        )

    axis.set_title(
        "Backward Difference — Function and Required Sample Points\n"
        f"{result.derivative_symbol} ≈ {primary.derivative_approximation:.10g}"
    )
    axis.set_xlabel("x")
    axis.set_ylabel("f(x)")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_refinement_plot(result: BackwardDifferenceResult) -> plt.Figure:
    """Plot the derivative approximation at every refinement level."""

    if not result.evaluations:
        raise ValueError("No refinement history is available.")

    levels = np.asarray([item.level for item in result.evaluations], dtype=int)
    approximations = np.asarray(
        [item.derivative_approximation for item in result.evaluations],
        dtype=float,
    )

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.plot(
        levels,
        approximations,
        marker="o",
        linewidth=2,
        label="Numerical approximation",
    )
    if result.primary_exact_derivative is not None:
        axis.axhline(
            result.primary_exact_derivative,
            linestyle="--",
            linewidth=1.5,
            label="Exact derivative",
        )
    axis.set_title("Backward Difference — Refinement Results")
    axis.set_xlabel("Refinement Level")
    axis.set_ylabel(result.derivative_symbol)
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_convergence_plot(result: BackwardDifferenceResult) -> plt.Figure:
    """Create a log-log plot of exact error or successive differences."""

    if not result.evaluations:
        raise ValueError("No refinement history is available.")

    step_sizes = np.asarray(
        [item.step_size for item in result.evaluations],
        dtype=float,
    )
    exact_errors = np.asarray(
        [
            np.nan if item.absolute_error is None else item.absolute_error
            for item in result.evaluations
        ],
        dtype=float,
    )
    successive = np.asarray(
        [
            np.nan
            if item.successive_difference is None
            else item.successive_difference
            for item in result.evaluations
        ],
        dtype=float,
    )

    use_exact_error = np.count_nonzero(np.isfinite(exact_errors)) >= 2
    values = exact_errors if use_exact_error else successive
    label = "Absolute Error" if use_exact_error else "Successive Difference"

    valid_mask = np.isfinite(values) & (values > 0.0) & np.isfinite(step_sizes)
    if np.count_nonzero(valid_mask) < 2:
        raise ValueError(
            "At least two positive error values are required for the convergence plot."
        )

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.loglog(
        step_sizes[valid_mask],
        values[valid_mask],
        marker="o",
        linewidth=2,
        label=label,
    )
    axis.invert_xaxis()
    axis.set_title("Backward Difference — Convergence Analysis")
    axis.set_xlabel("Step Size h (Log Scale)")
    axis.set_ylabel(f"{label} (Log Scale)")
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def figure_to_png_bytes(figure: plt.Figure) -> bytes:
    """Serialize a Matplotlib figure as PNG bytes."""

    output = BytesIO()
    figure.savefig(output, format="png", dpi=160, bbox_inches="tight")
    output.seek(0)
    return output.getvalue()


# =============================================================================
# Excel export
# =============================================================================
def serialize_warnings(warnings: Sequence[str]) -> str:
    """Serialize warnings for an Excel cell."""

    return "None" if not warnings else "\n".join(f"• {warning}" for warning in warnings)


def apply_excel_style(workbook: Any) -> None:
    """Apply consistent professional formatting to workbook data sheets."""

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        if worksheet.title == "Plots":
            worksheet.sheet_view.showGridLines = False
            continue

        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            for column_index, column_cells in enumerate(
                worksheet.columns,
                start=1,
            ):
                maximum_length = 0
                for cell in column_cells:
                    value_length = len(str(cell.value)) if cell.value is not None else 0
                    maximum_length = max(maximum_length, value_length)
                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width = min(max(maximum_length + 2, 12), 65)


def create_excel_report(result: BackwardDifferenceResult) -> bytes:
    """Create a complete XLSX report with tables, charts, and plot images."""

    if not result.success:
        raise ValueError("Only successful results can be exported.")

    summary_dataframe = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "Function",
                "Requested Derivative",
                "Derivative Order",
                "Exact Derivative Expression",
                "Formula",
                "Theoretical Accuracy Order",
                "Evaluation Point x0",
                "Entered Step Size h",
                "Refinement Levels",
                "Primary Approximation",
                "Exact Derivative at x0",
                "Primary Absolute Error",
                "Primary Absolute Relative Error (%)",
                "Finest Step Size",
                "Finest Approximation",
                "Finest Absolute Error",
                "Latest Observed Order",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                result.function_text,
                result.derivative_label,
                result.derivative_order,
                result.derivative_expression_text,
                result.formula_text,
                result.theoretical_order,
                result.x_value,
                result.initial_step_size,
                result.refinement_levels,
                result.primary_approximation,
                result.primary_exact_derivative,
                result.primary_absolute_error,
                result.primary_relative_error_percent,
                result.finest_step_size,
                result.finest_approximation,
                result.finest_absolute_error,
                result.observed_order_latest,
                serialize_warnings(result.warnings),
                result.stopping_reason,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )

    evaluation_dataframe = create_evaluation_dataframe(result)
    primary_dataframe = create_primary_calculation_dataframe(result)
    error_dataframe = create_error_dataframe(result)
    taylor_dataframe = create_taylor_series_dataframe(result)
    formula_dataframe = pd.DataFrame(
        [
            {
                "Level": item.level,
                "Derivative Order": item.derivative_order,
                "h": item.step_size,
                "Formula": item.formula_text,
                "Substitution": item.substitution_text,
                "Numerator": item.numerator,
                "Denominator": item.denominator,
                "Approximate Derivative": item.derivative_approximation,
            }
            for item in result.evaluations
        ]
    )

    function_figure = create_function_plot(result)
    refinement_figure = create_refinement_plot(result)
    function_png = figure_to_png_bytes(function_figure)
    refinement_png = figure_to_png_bytes(refinement_figure)
    plt.close(function_figure)
    plt.close(refinement_figure)

    try:
        convergence_figure = create_convergence_plot(result)
    except ValueError:
        convergence_png = None
    else:
        convergence_png = figure_to_png_bytes(convergence_figure)
        plt.close(convergence_figure)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_dataframe.to_excel(writer, sheet_name="Summary", index=False)
        primary_dataframe.to_excel(
            writer,
            sheet_name="Primary Calculation",
            index=False,
        )
        evaluation_dataframe.to_excel(
            writer,
            sheet_name="Difference Results",
            index=False,
        )
        formula_dataframe.to_excel(
            writer,
            sheet_name="Formula Steps",
            index=False,
        )
        taylor_dataframe.to_excel(
            writer,
            sheet_name="Taylor Series",
            index=False,
        )
        error_dataframe.to_excel(
            writer,
            sheet_name="Error Analysis",
            index=False,
        )
        error_dataframe.to_excel(
            writer,
            sheet_name="Convergence Analysis",
            index=False,
        )

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Backward Difference Plots"
        plots_sheet["A1"].font = Font(bold=True, size=14)

        function_image_stream = BytesIO(function_png)
        function_image = OpenpyxlImage(function_image_stream)
        function_image.width = 900
        function_image.height = 540
        plots_sheet.add_image(function_image, "A3")

        refinement_image_stream = BytesIO(refinement_png)
        refinement_image = OpenpyxlImage(refinement_image_stream)
        refinement_image.width = 900
        refinement_image.height = 540
        plots_sheet.add_image(refinement_image, "A32")

        if convergence_png is not None:
            convergence_image_stream = BytesIO(convergence_png)
            convergence_image = OpenpyxlImage(convergence_image_stream)
            convergence_image.width = 900
            convergence_image.height = 540
            plots_sheet.add_image(convergence_image, "A61")

        convergence_sheet = workbook["Convergence Analysis"]
        row_count = len(error_dataframe) + 1
        if row_count >= 3:
            chart = ScatterChart()
            chart.title = "Approximation versus Step Size"
            chart.x_axis.title = "Step Size h"
            chart.y_axis.title = "Approximate Derivative"
            chart.height = 8
            chart.width = 15

            x_reference = Reference(
                convergence_sheet,
                min_col=3,
                min_row=2,
                max_row=row_count,
            )
            y_reference = Reference(
                convergence_sheet,
                min_col=4,
                min_row=2,
                max_row=row_count,
            )
            series = Series(
                y_reference,
                x_reference,
                title="Approximate Derivative",
            )
            chart.series.append(series)
            convergence_sheet.add_chart(chart, "K2")

            if any(item.absolute_error is not None for item in result.evaluations):
                error_chart = LineChart()
                error_chart.title = "Absolute Error by Refinement Level"
                error_chart.x_axis.title = "Refinement Level"
                error_chart.y_axis.title = "Absolute Error"
                error_chart.height = 8
                error_chart.width = 15
                data_reference = Reference(
                    convergence_sheet,
                    min_col=6,
                    min_row=1,
                    max_row=row_count,
                )
                category_reference = Reference(
                    convergence_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=row_count,
                )
                error_chart.add_data(data_reference, titles_from_data=True)
                error_chart.set_categories(category_reference)
                convergence_sheet.add_chart(error_chart, "K20")

        apply_excel_style(workbook)
        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit result rendering
# =============================================================================
def render_final_result(result: BackwardDifferenceResult) -> None:
    """Render the compact final-result card."""

    if not result.success:
        st.error(result.message)
        st.caption(result.stopping_reason)
        return

    st.success(result.message)
    st.markdown(f"**Function:** `{result.function_text}`")
    st.markdown(f"**Requested derivative:** {result.derivative_label}")
    st.markdown(f"**Formula:** {result.formula_text}")

    metric_columns = st.columns(2)
    metric_columns[0].metric(
        f"Approximate {result.derivative_symbol}",
        format_number(result.primary_approximation),
    )
    metric_columns[1].metric(
        f"Exact {result.derivative_symbol}",
        format_number(result.primary_exact_derivative),
    )

    detail_columns = st.columns(2)
    detail_columns[0].metric("x₀", format_number(result.x_value))
    detail_columns[1].metric("h", format_number(result.initial_step_size))

    error_columns = st.columns(2)
    error_columns[0].metric(
        "Absolute Error",
        scientific_number(result.primary_absolute_error),
    )
    error_columns[1].metric(
        "Relative Error (%)",
        format_number(result.primary_relative_error_percent),
    )

    if result.warnings:
        for warning in result.warnings:
            st.warning(warning)


def render_primary_calculation(result: BackwardDifferenceResult) -> None:
    """Render the full primary calculation and formula substitution."""

    st.subheader("Primary Backward Difference Calculation")
    st.dataframe(
        round_numeric_dataframe(create_primary_calculation_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    primary = result.evaluations[0]
    with st.container(border=True):
        st.markdown("#### Formula substitution")
        st.code(primary.substitution_text, language=None)
        st.markdown(
            f"**Approximate {result.derivative_symbol}:** "
            f"{format_number(primary.derivative_approximation, 10)}"
        )
        if primary.exact_derivative is not None:
            st.markdown(
                f"**Exact {result.derivative_symbol}:** "
                f"{format_number(primary.exact_derivative, 10)}"
            )
            st.markdown(
                f"**Absolute error:** {scientific_number(primary.absolute_error)}"
            )


def render_taylor_series(result: BackwardDifferenceResult) -> None:
    """Render Taylor expansions through the third derivative."""

    st.subheader("Taylor Series Through the Third Derivative")
    st.dataframe(
        round_numeric_dataframe(create_taylor_series_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )
    st.caption(
        "The listed backward-point expansions are centered at x₀ and retain "
        "terms through f‴(x₀). The finite-difference coefficients combine these "
        "expansions to isolate the selected derivative."
    )


def render_refinement_analysis(result: BackwardDifferenceResult) -> None:
    """Render the step-size refinement table and convergence interpretation."""

    st.subheader("Step-Size Refinement and Error Analysis")
    error_dataframe = create_error_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(error_dataframe),
        use_container_width=True,
        hide_index=True,
    )

    st.info(
        f"The selected formula has theoretical truncation error "
        f"O(h^{result.theoretical_order}). When the function is sufficiently "
        "smooth and round-off error is not dominant, halving h should reduce "
        f"the truncation error by approximately 2^{result.theoretical_order}."
    )

    if result.observed_order_latest is not None:
        st.markdown(
            "**Latest observed convergence order:** "
            f"{format_number(result.observed_order_latest, 4)}"
        )

    refinement_figure = create_refinement_plot(result)
    st.pyplot(refinement_figure, use_container_width=True)
    plt.close(refinement_figure)

    try:
        convergence_figure = create_convergence_plot(result)
    except ValueError as error:
        st.info(str(error))
    else:
        st.pyplot(convergence_figure, use_container_width=True)
        plt.close(convergence_figure)


def render_function_graph(result: BackwardDifferenceResult) -> None:
    """Render the function and sampled backward points safely."""

    st.subheader("Function Graph")
    try:
        figure = create_function_plot(result)
    except ValueError as error:
        st.warning(f"The graph could not be displayed. {error}")
    else:
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)


def render_excel_download(result: BackwardDifferenceResult) -> None:
    """Create and render the Excel report download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature
    cached_signature = st.session_state.get("backward_difference_excel_signature")

    if cached_signature != report_signature:
        try:
            report_bytes = create_excel_report(result)
        except (ValueError, OSError, RuntimeError) as error:
            st.error(f"The Excel report could not be generated. {error}")
            return
        st.session_state.backward_difference_excel_report = report_bytes
        st.session_state.backward_difference_excel_signature = report_signature

    report_bytes = st.session_state.get("backward_difference_excel_report")
    if report_bytes is None:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=f"backward_difference_report_{timestamp}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="backward_difference_download_button",
    )


# =============================================================================
# Streamlit page — the existing project template is preserved
# =============================================================================
def render_page() -> None:
    """Render the complete Backward Difference Streamlit solver page."""

    st.set_page_config(
        page_title="Backward Difference Solver | Numerical Methods",
        page_icon="📉",
        layout="wide",
    )
    load_css()

    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">NUMERICAL DIFFERENTIATION TOOL</div>
                <h1>Backward Difference Solver</h1>
                <p>
                    Enter a function, evaluation point, and step size to compute
                    the first, second, or third derivative using standard backward
                    finite-difference formulas, then review the complete calculation,
                    Taylor series, refinement, error analysis, graph, and Excel report.
                </p>

                <div class="method-actions">
                    <a href="/Backward_Difference" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Backward_Difference_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Write the Function")
                st.markdown(
                    """
                Enter only the mathematical expression, without `f(x) =` or an equals sign.

                - Use only **x** as the variable.
                - Powers: write `x**2`, not **x^2**.
                - Multiplication: write `2*x`, not `2x`.
                - Use lowercase functions such as **sin(x)**, **cos(x)**, **exp(x)**, **sqrt(x)**, and **log(x)**.
                - Use parentheses whenever the order of operations could be unclear.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")
                st.markdown(
                    """
                - The step size **h** must be positive.
                - Select the derivative order: first, second, or third.
                - The function must be finite at every required backward point.
                - Depending on the selected formula, points from **x₀ − h** through **x₀ − 4h** may be required.
                - Reducing **h** usually reduces truncation error, but an extremely small **h** may increase round-off error.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">Function f(x)</div>',
                    unsafe_allow_html=True,
                )
                function_text = st.text_input(
                    "Function f(x)",
                    value=DEFAULT_FUNCTION,
                    placeholder="Example: sin(x), exp(x), or x**3 - 2*x + 1",
                    label_visibility="collapsed",
                    key="backward_difference_function",
                )

                value_columns = st.columns(2)
                with value_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Evaluation point x₀</div>',
                        unsafe_allow_html=True,
                    )
                    x_value = st.number_input(
                        "Evaluation point x0",
                        value=DEFAULT_X_VALUE,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="backward_difference_x_value",
                    )

                with value_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Step size h</div>',
                        unsafe_allow_html=True,
                    )
                    step_size = st.number_input(
                        "Step size h",
                        value=DEFAULT_STEP_SIZE,
                        min_value=0.0,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="backward_difference_step_size",
                    )

                st.markdown(
                    '<div class="input-label-ui">Derivative order</div>',
                    unsafe_allow_html=True,
                )
                derivative_name = st.selectbox(
                    "Derivative order",
                    options=list(DERIVATIVE_OPTIONS.keys()),
                    index=0,
                    label_visibility="collapsed",
                    key="backward_difference_derivative_order",
                )

                st.markdown(
                    '<div class="input-label-ui">Backward formula</div>',
                    unsafe_allow_html=True,
                )
                formula_name = st.selectbox(
                    "Backward formula",
                    options=list(FORMULA_OPTIONS.keys()),
                    index=0,
                    label_visibility="collapsed",
                    key="backward_difference_formula",
                )

                st.markdown(
                    '<div class="input-label-ui">Step-size refinement levels</div>',
                    unsafe_allow_html=True,
                )
                refinement_levels = st.slider(
                    "Step-size refinement levels",
                    min_value=MIN_REFINEMENT_LEVELS,
                    max_value=MAX_REFINEMENT_LEVELS,
                    value=DEFAULT_REFINEMENT_LEVELS,
                    step=1,
                    label_visibility="collapsed",
                    key="backward_difference_refinement_levels",
                )

                st.caption(
                    "The main result uses the entered h. Additional levels use "
                    "h/2, h/4, and so on for convergence analysis. The formulas "
                    "are evaluated directly from backward function values."
                )

                solve_button_clicked = st.button(
                    "Solve",
                    use_container_width=True,
                    key="backward_difference_solve_button",
                )

        current_input_signature = create_input_signature(
            function_text=function_text,
            x_value=x_value,
            step_size=step_size,
            derivative_name=derivative_name,
            formula_name=formula_name,
            refinement_levels=refinement_levels,
        )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                stored_result = st.session_state.get("backward_difference_result")
                if stored_result is None:
                    st.info("Enter the function and parameters, then click Solve.")
                elif stored_result.input_signature != current_input_signature:
                    st.info(
                        "The function or numerical parameters have changed. "
                        "Click Solve to calculate a new result."
                    )
                else:
                    render_final_result(stored_result)

        if solve_button_clicked:
            st.session_state.backward_difference_result = solve_backward_difference(
                function_text=function_text,
                raw_x_value=x_value,
                raw_step_size=step_size,
                derivative_name=derivative_name,
                formula_name=formula_name,
                raw_refinement_levels=refinement_levels,
                input_signature=current_input_signature,
            )
            st.session_state.pop("backward_difference_excel_report", None)
            st.session_state.pop("backward_difference_excel_signature", None)
            st.rerun()

        active_result = st.session_state.get("backward_difference_result")
        if (
            active_result is not None
            and active_result.input_signature == current_input_signature
        ):
            if active_result.success:
                st.divider()
                render_primary_calculation(active_result)

                st.divider()
                render_taylor_series(active_result)

                st.divider()
                render_refinement_analysis(active_result)

                st.divider()
                render_function_graph(active_result)

                st.divider()
                render_excel_download(active_result)

                st.divider()
                navigation_left_column, navigation_right_column = st.columns(2)

                with navigation_left_column:
                    if st.button(
                        "Review Backward Difference Lesson",
                        use_container_width=True,
                        key="review_backward_difference_lesson",
                    ):
                        st.switch_page("pages/Backward_Difference.py")

                with navigation_right_column:
                    if st.button(
                        "Back to Solver Menu",
                        use_container_width=True,
                        key="back_to_solver_menu_backward_difference",
                    ):
                        st.switch_page("pages/Numerical_Solver.py")

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Numerical Differentiation • Backward Difference</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()