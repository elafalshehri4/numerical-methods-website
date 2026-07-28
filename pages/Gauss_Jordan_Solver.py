from __future__ import annotations

import hashlib
import math
from io import BytesIO
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from components.navigation import navbar
from utilities.ui import load_css


# =============================================================================
# Page and method configuration
# =============================================================================
CONFIG = {
    "method_id": "gauss_jordan",
    "title": "Gauss-Jordan Solver",
    "label": "GAUSS-JORDAN METHOD TOOL",
    "description": (
        "Reduce an augmented matrix to reduced row echelon form and read "
        "the linear-system solution directly."
    ),
    "lesson": "Gauss_Jordan_Method",
    "quiz": "Gauss_Jordan_Quiz",
    "footer": "Gauss-Jordan Solver • Linear Systems",
    "conditions": [
        "The coefficient matrix must be square.",
        "The matrix and right-hand-side dimensions must agree.",
        "A unique solution requires a pivot in every variable column.",
        "Partial pivoting is used to avoid zero or numerically weak pivots.",
    ],
    "formula": (
        r"R_k\leftarrow\frac{R_k}{a_{kk}},"
        r"\qquad R_i\leftarrow R_i-a_{ik}R_k"
    ),
    "defaults": {
        "A": [[2.0, 1.0], [1.0, -1.0]],
        "b": [5.0, 1.0],
    },
}

DISPLAY_DECIMALS = 3
MIN_SYSTEM_SIZE = 2
MAX_SYSTEM_SIZE = 6
CONDITION_NUMBER_WARNING = 1.0e12
RELATIVE_RESIDUAL_WARNING = 1.0e-10
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable

    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0**exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Show fixed notation or scientific notation when appropriate."""

    if value is None:
        return unavailable

    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6
    ):
        return format_scientific_power(number, decimals, unavailable)

    return f"{number:.{decimals}f}"


def format_number(value: float | int | None) -> str:
    """Format final solver values consistently."""

    return format_display_number(value, unavailable="Not available")


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only for display."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns

    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)

    return rounded


# =============================================================================
# Input helpers
# =============================================================================
def create_default_table(size: int) -> pd.DataFrame:
    """Build the editable coefficient table for the selected system size."""

    default_A = np.eye(size, dtype=float) * 2.0
    default_b = np.ones(size, dtype=float)

    stored_A = np.asarray(CONFIG["defaults"]["A"], dtype=float)
    stored_b = np.asarray(CONFIG["defaults"]["b"], dtype=float)

    rows = min(size, stored_A.shape[0])
    columns = min(size, stored_A.shape[1])

    default_A[:rows, :columns] = stored_A[:rows, :columns]
    default_b[:rows] = stored_b[:rows]

    data = {
        f"x{column + 1}": default_A[:, column]
        for column in range(size)
    }
    data["b"] = default_b

    return pd.DataFrame(data)


def create_input_signature(
    input_table: pd.DataFrame,
    size: int,
) -> str:
    """Create a stable signature to prevent stale Streamlit results."""

    normalized = input_table.copy()
    normalized.columns = [str(column) for column in normalized.columns]

    payload = (
        f"{size}|"
        + normalized.to_csv(index=False, float_format="%.17g")
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def extract_system_from_table(
    input_table: pd.DataFrame,
    size: int,
) -> tuple[np.ndarray, np.ndarray]:
    """Extract and validate A and b from the editable table."""

    required_columns = [f"x{i + 1}" for i in range(size)] + ["b"]
    missing_columns = [
        column for column in required_columns if column not in input_table.columns
    ]

    if missing_columns:
        raise ValueError(
            "The input table is missing required column(s): "
            + ", ".join(missing_columns)
        )

    try:
        numeric_table = input_table[required_columns].apply(
            pd.to_numeric,
            errors="raise",
        )
    except (TypeError, ValueError) as error:
        raise ValueError("Every table entry must be a valid number.") from error

    if numeric_table.shape != (size, size + 1):
        raise ValueError(
            f"The table must contain exactly {size} equations and {size + 1} columns."
        )

    matrix = numeric_table.iloc[:, :size].to_numpy(dtype=float)
    vector = numeric_table["b"].to_numpy(dtype=float)

    if not np.all(np.isfinite(matrix)) or not np.all(np.isfinite(vector)):
        raise ValueError("All matrix and vector entries must be finite numbers.")

    return matrix, vector


# =============================================================================
# Gauss-Jordan numerical algorithm
# =============================================================================
def augmented_text(
    matrix: np.ndarray,
    vector: np.ndarray,
) -> str:
    """Return a readable augmented matrix string."""

    augmented = np.column_stack((matrix, vector))
    return np.array2string(
        augmented,
        precision=8,
        suppress_small=True,
        max_line_width=160,
    )


def matrix_text(augmented: np.ndarray) -> str:
    """Return a readable augmented-matrix string."""

    return np.array2string(
        augmented,
        precision=8,
        suppress_small=True,
        max_line_width=160,
    )


def calculate_pivot_tolerance(matrix: np.ndarray) -> float:
    """Return a scale-aware pivot tolerance for floating-point elimination."""

    scale = float(np.linalg.norm(matrix, ord=np.inf))
    scale = max(scale, 1.0)

    return (
        100.0
        * np.finfo(float).eps
        * max(matrix.shape)
        * scale
    )


def clean_small_values(
    values: np.ndarray,
    tolerance: float,
) -> np.ndarray:
    """Replace insignificant floating-point noise by exact zero."""

    cleaned = np.asarray(values, dtype=float).copy()
    cleaned[np.abs(cleaned) <= tolerance] = 0.0
    return cleaned


def solve_linear_system(
    matrix: Any,
    vector: Any,
    input_signature: str = "",
) -> dict[str, Any]:
    """Solve Ax=b manually by Gauss-Jordan elimination with partial pivoting.

    The algorithm supports square systems and classifies a singular system as
    inconsistent or underdetermined instead of reporting every singular case
    as the same error.
    """

    try:
        A = np.asarray(matrix, dtype=float).copy()
        b = np.asarray(vector, dtype=float).reshape(-1).copy()
    except (TypeError, ValueError) as error:
        return {
            "status": "error",
            "classification": "Invalid input",
            "message": "The matrix and vector must contain numerical values.",
            "input_signature": input_signature,
        }

    if A.ndim != 2 or A.shape[0] != A.shape[1]:
        return {
            "status": "error",
            "classification": "Invalid input",
            "message": "A must be a square matrix.",
            "input_signature": input_signature,
        }

    n = A.shape[0]

    if b.shape != (n,):
        return {
            "status": "error",
            "classification": "Invalid input",
            "message": (
                "The right-hand side must contain exactly one value "
                "for each equation."
            ),
            "input_signature": input_signature,
        }

    if not np.all(np.isfinite(A)) or not np.all(np.isfinite(b)):
        return {
            "status": "error",
            "classification": "Invalid input",
            "message": "All entries must be finite real numbers.",
            "input_signature": input_signature,
        }

    original_A = A.copy()
    original_b = b.copy()
    augmented = np.column_stack((A, b))

    pivot_tolerance = calculate_pivot_tolerance(A)
    cleanup_tolerance = max(
        pivot_tolerance,
        100.0 * np.finfo(float).eps,
    )

    history: list[dict[str, Any]] = []
    pivot_columns: list[int] = []
    active_row = 0

    try:
        for column in range(n):
            if active_row >= n:
                break

            candidate_rows = np.arange(active_row, n)
            candidate_values = np.abs(augmented[active_row:, column])
            relative_index = int(np.argmax(candidate_values))
            selected_row = active_row + relative_index
            selected_pivot = float(augmented[selected_row, column])

            # No usable pivot in this variable column: skip it and classify later.
            if abs(selected_pivot) <= pivot_tolerance:
                continue

            if selected_row != active_row:
                augmented[[active_row, selected_row]] = augmented[
                    [selected_row, active_row]
                ]

                history.append(
                    {
                        "Step": len(history) + 1,
                        "Stage": "Row interchange",
                        "Pivot Column": column + 1,
                        "Operation": (
                            f"Swap R{active_row + 1} with R{selected_row + 1}"
                        ),
                        "Pivot Before Operation": selected_pivot,
                        "Factor": None,
                        "Augmented Matrix": matrix_text(augmented),
                    }
                )

            pivot = float(augmented[active_row, column])

            if abs(pivot) <= pivot_tolerance:
                raise ArithmeticError(
                    f"No reliable pivot exists in column {column + 1}."
                )

            augmented[active_row, :] = augmented[active_row, :] / pivot
            augmented = clean_small_values(augmented, cleanup_tolerance)

            history.append(
                {
                    "Step": len(history) + 1,
                    "Stage": "Pivot normalization",
                    "Pivot Column": column + 1,
                    "Operation": (
                        f"R{active_row + 1} ← "
                        f"R{active_row + 1} / ({pivot:.12g})"
                    ),
                    "Pivot Before Operation": pivot,
                    "Factor": 1.0 / pivot,
                    "Augmented Matrix": matrix_text(augmented),
                }
            )

            for row in range(n):
                if row == active_row:
                    continue

                factor = float(augmented[row, column])

                if abs(factor) <= cleanup_tolerance:
                    augmented[row, column] = 0.0
                    continue

                augmented[row, :] = (
                    augmented[row, :]
                    - factor * augmented[active_row, :]
                )
                augmented = clean_small_values(
                    augmented,
                    cleanup_tolerance,
                )

                history.append(
                    {
                        "Step": len(history) + 1,
                        "Stage": "Column elimination",
                        "Pivot Column": column + 1,
                        "Operation": (
                            f"R{row + 1} ← R{row + 1} − "
                            f"({factor:.12g})R{active_row + 1}"
                        ),
                        "Pivot Before Operation": 1.0,
                        "Factor": factor,
                        "Augmented Matrix": matrix_text(augmented),
                    }
                )

            pivot_columns.append(column)
            active_row += 1

        augmented = clean_small_values(augmented, cleanup_tolerance)
        coefficient_rref = augmented[:, :n]
        right_hand_side_rref = augmented[:, -1]

        zero_coefficient_rows = np.all(
            np.abs(coefficient_rref) <= cleanup_tolerance,
            axis=1,
        )
        inconsistent_rows = zero_coefficient_rows & (
            np.abs(right_hand_side_rref) > cleanup_tolerance
        )

        rank_A = len(pivot_columns)
        rank_augmented = rank_A + int(np.any(inconsistent_rows))

        determinant = float(np.linalg.det(original_A))
        if not math.isfinite(determinant):
            determinant = math.nan

        if np.any(inconsistent_rows):
            return {
                "status": "error",
                "classification": "No solution",
                "message": (
                    "The system is inconsistent: the reduced matrix contains "
                    "a row of the form [0 … 0 | c] with c ≠ 0."
                ),
                "history": history,
                "original_A": original_A,
                "original_b": original_b,
                "final_matrix": coefficient_rref,
                "final_vector": right_hand_side_rref,
                "final_form_name": "Reduced Row Echelon Form",
                "rank_A": rank_A,
                "rank_augmented": rank_augmented,
                "pivot_columns": pivot_columns,
                "pivot_tolerance": pivot_tolerance,
                "determinant": determinant,
                "input_signature": input_signature,
            }

        if rank_A < n:
            free_variables = [
                f"x{column + 1}"
                for column in range(n)
                if column not in pivot_columns
            ]

            return {
                "status": "error",
                "classification": "Infinitely many solutions",
                "message": (
                    "The system is underdetermined and has infinitely many "
                    "solutions. Free variable(s): "
                    + ", ".join(free_variables)
                    + "."
                ),
                "history": history,
                "original_A": original_A,
                "original_b": original_b,
                "final_matrix": coefficient_rref,
                "final_vector": right_hand_side_rref,
                "final_form_name": "Reduced Row Echelon Form",
                "rank_A": rank_A,
                "rank_augmented": rank_A,
                "pivot_columns": pivot_columns,
                "free_variables": free_variables,
                "pivot_tolerance": pivot_tolerance,
                "determinant": determinant,
                "input_signature": input_signature,
            }

        # Every variable column is a pivot column, so the left block is I.
        solution = right_hand_side_rref.copy()
        residual_vector = original_A @ solution - original_b
        computed_b = original_A @ solution

        residual_norm = float(
            np.linalg.norm(residual_vector, ord=np.inf)
        )
        matrix_norm = float(np.linalg.norm(original_A, ord=np.inf))
        solution_norm = float(np.linalg.norm(solution, ord=np.inf))
        vector_norm = float(np.linalg.norm(original_b, ord=np.inf))

        residual_denominator = (
            matrix_norm * solution_norm + vector_norm
        )
        relative_residual = (
            residual_norm / residual_denominator
            if residual_denominator > 0.0
            else residual_norm
        )

        identity_error = float(
            np.linalg.norm(
                coefficient_rref - np.eye(n),
                ord=np.inf,
            )
        )

        try:
            condition_number = float(
                np.linalg.cond(original_A, p=np.inf)
            )
        except np.linalg.LinAlgError:
            condition_number = math.inf

        warnings: list[str] = []

        if (
            not math.isfinite(condition_number)
            or condition_number >= CONDITION_NUMBER_WARNING
        ):
            warnings.append(
                "The coefficient matrix is ill-conditioned. Small changes in "
                "the input or floating-point round-off may cause large changes "
                "in the solution."
            )

        if relative_residual > RELATIVE_RESIDUAL_WARNING:
            warnings.append(
                "The relative residual is larger than expected. Inspect the "
                "input scale, conditioning, and displayed row operations."
            )

        if identity_error > 1.0e-9:
            warnings.append(
                "The reduced coefficient matrix differs noticeably from the "
                "identity matrix because of floating-point effects."
            )

        return {
            "status": "success",
            "classification": "Unique solution",
            "message": "The linear system was solved successfully.",
            "solution": solution,
            "history": history,
            "original_A": original_A,
            "original_b": original_b,
            "final_matrix": coefficient_rref,
            "final_vector": right_hand_side_rref,
            "final_form_name": "Reduced Row Echelon Form",
            "residual_vector": residual_vector,
            "computed_b": computed_b,
            "residual_norm": residual_norm,
            "relative_residual": relative_residual,
            "identity_error": identity_error,
            "determinant": determinant,
            "condition_number": condition_number,
            "rank_A": rank_A,
            "rank_augmented": rank_A,
            "pivot_columns": pivot_columns,
            "pivot_tolerance": pivot_tolerance,
            "operations": len(history),
            "warnings": tuple(warnings),
            "input_signature": input_signature,
        }

    except (ArithmeticError, FloatingPointError, OverflowError) as error:
        return {
            "status": "error",
            "classification": "Numerical failure",
            "message": str(error),
            "history": history,
            "input_signature": input_signature,
        }


# =============================================================================
# Data tables
# =============================================================================
def solution_dataframe(result: dict[str, Any]) -> pd.DataFrame:
    """Build the solution-vector table."""

    solution = np.asarray(result["solution"], dtype=float)

    return pd.DataFrame(
        {
            "Variable": [
                f"x{index + 1}" for index in range(len(solution))
            ],
            "Calculated Value": solution,
        }
    )


def residual_dataframe(result: dict[str, Any]) -> pd.DataFrame:
    """Build equation-by-equation residual verification."""

    solution = np.asarray(result["solution"], dtype=float)
    residual = np.asarray(result["residual_vector"], dtype=float)
    computed_b = np.asarray(result["computed_b"], dtype=float)

    return pd.DataFrame(
        {
            "Equation": [
                f"Equation {index + 1}"
                for index in range(len(solution))
            ],
            "Computed Ax": computed_b,
            "Original b": result["original_b"],
            "Residual Ax − b": residual,
            "Absolute Residual": np.abs(residual),
        }
    )


def operations_dataframe(result: dict[str, Any]) -> pd.DataFrame:
    """Build the row-operation history table."""

    history = result.get("history", [])

    if not history:
        return pd.DataFrame(
            [
                {
                    "Step": 0,
                    "Stage": "Completed",
                    "Pivot Column": None,
                    "Operation": "No row operation was required.",
                    "Pivot Before Operation": None,
                    "Factor": None,
                }
            ]
        )

    return pd.DataFrame(
        [
            {
                "Step": item["Step"],
                "Stage": item["Stage"],
                "Pivot Column": item["Pivot Column"],
                "Operation": item["Operation"],
                "Pivot Before Operation": item[
                    "Pivot Before Operation"
                ],
                "Factor": item["Factor"],
                "Augmented Matrix": item["Augmented Matrix"],
            }
            for item in history
        ]
    )


# =============================================================================
# Plot builders
# =============================================================================
def create_matrix_figure(result: dict[str, Any]) -> plt.Figure:
    """Create a heatmap of the original coefficient matrix."""

    matrix = np.asarray(result["original_A"], dtype=float)

    figure, axis = plt.subplots(figsize=(7.0, 5.0))
    image = axis.imshow(matrix, aspect="auto")
    figure.colorbar(image, ax=axis, label="Coefficient value")

    axis.set_title("Coefficient Matrix A")
    axis.set_xlabel("Variable Column")
    axis.set_ylabel("Equation Row")
    axis.set_xticks(
        range(matrix.shape[1]),
        [f"x{index + 1}" for index in range(matrix.shape[1])],
    )
    axis.set_yticks(
        range(matrix.shape[0]),
        [f"Eq. {index + 1}" for index in range(matrix.shape[0])],
    )

    for row in range(matrix.shape[0]):
        for column in range(matrix.shape[1]):
            axis.text(
                column,
                row,
                f"{matrix[row, column]:.3g}",
                ha="center",
                va="center",
            )

    figure.tight_layout()
    return figure


def create_solution_figure(result: dict[str, Any]) -> plt.Figure:
    """Create a bar chart of the solution vector."""

    solution = np.asarray(result["solution"], dtype=float)
    positions = np.arange(len(solution))

    figure, axis = plt.subplots(figsize=(7.0, 5.0))
    axis.bar(positions, solution)
    axis.axhline(0.0, linewidth=1)
    axis.set_xticks(
        positions,
        [f"x{index + 1}" for index in positions],
    )
    axis.set_title("Gauss-Jordan Solution Values")
    axis.set_xlabel("Variable")
    axis.set_ylabel("Value")
    axis.grid(True, axis="y", alpha=0.3)

    figure.tight_layout()
    return figure


def create_residual_figure(result: dict[str, Any]) -> plt.Figure:
    """Create a bar chart of absolute residual components."""

    residual = np.abs(
        np.asarray(result["residual_vector"], dtype=float)
    )
    positions = np.arange(len(residual))

    figure, axis = plt.subplots(figsize=(7.0, 5.0))
    axis.bar(positions, residual)
    axis.set_xticks(
        positions,
        [f"Eq. {index + 1}" for index in positions],
    )
    axis.set_title("Absolute Residual Components")
    axis.set_xlabel("Equation")
    axis.set_ylabel("|Ax − b|")
    axis.grid(True, axis="y", alpha=0.3)

    figure.tight_layout()
    return figure


def figure_to_png_bytes(figure: plt.Figure) -> bytes:
    """Serialize a matplotlib figure as PNG bytes."""

    buffer = BytesIO()
    figure.savefig(
        buffer,
        format="png",
        dpi=180,
        bbox_inches="tight",
    )
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# Excel report
# =============================================================================
def style_excel_workbook(workbook: Any) -> None:
    """Apply consistent formatting to every worksheet."""

    header_fill = PatternFill("solid", fgColor="0D3151")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.freeze_panes = "A2"

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions

        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            maximum_length = 0

            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

                maximum_length = max(
                    maximum_length,
                    len(str(cell.value))
                    if cell.value is not None
                    else 0,
                )

                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

            worksheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 12),
                55,
            )


def add_excel_image(
    worksheet: Any,
    image_bytes: bytes,
    anchor: str,
    width: int = 760,
    height: int = 500,
) -> None:
    """Insert PNG bytes into an Excel worksheet."""

    image_stream = BytesIO(image_bytes)
    excel_image = ExcelImage(image_stream)
    excel_image.width = width
    excel_image.height = height
    worksheet.add_image(excel_image, anchor)


def create_excel_report(result: dict[str, Any]) -> bytes:
    """Create a complete Gauss-Jordan XLSX report in memory."""

    if result.get("status") != "success":
        raise ValueError("Only a successful calculation can be exported.")

    solution = np.asarray(result["solution"], dtype=float)
    residual = np.asarray(result["residual_vector"], dtype=float)
    size = len(solution)

    summary = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "System Classification",
                "Matrix Size",
                "Final Matrix Form",
                "Rank(A)",
                "Rank([A|b])",
                "Row Operations",
                "Pivot Tolerance",
                "Determinant",
                "Condition Number (∞-norm)",
                "Residual Infinity Norm",
                "Relative Residual",
                "RREF Identity Error",
                "Solution Vector",
                "Warnings",
                "Message",
            ],
            "Value": [
                CONFIG["title"],
                result["status"].title(),
                result["classification"],
                f"{size} × {size}",
                result["final_form_name"],
                result["rank_A"],
                result["rank_augmented"],
                result["operations"],
                result["pivot_tolerance"],
                result["determinant"],
                result["condition_number"],
                result["residual_norm"],
                result["relative_residual"],
                result["identity_error"],
                ", ".join(
                    f"x{index + 1} = {value:.15g}"
                    for index, value in enumerate(solution)
                ),
                (
                    "\n".join(f"• {warning}" for warning in result["warnings"])
                    if result["warnings"]
                    else "None"
                ),
                result["message"],
            ],
        }
    )

    input_system = pd.DataFrame(
        np.column_stack(
            (result["original_A"], result["original_b"])
        ),
        columns=[
            f"x{index + 1}" for index in range(size)
        ]
        + ["b"],
        index=[
            f"Equation {index + 1}" for index in range(size)
        ],
    ).reset_index(names="Equation")

    final_system = pd.DataFrame(
        np.column_stack(
            (result["final_matrix"], result["final_vector"])
        ),
        columns=[
            f"x{index + 1}" for index in range(size)
        ]
        + ["Right-Hand Side"],
        index=[f"Row {index + 1}" for index in range(size)],
    ).reset_index(names="Row")

    solution_table = solution_dataframe(result)
    residual_table = residual_dataframe(result)
    operations_table = operations_dataframe(result)

    verification_table = pd.DataFrame(
        {
            "Metric": [
                "||Ax − b||∞",
                "||Ax − b||∞ / (||A||∞||x||∞ + ||b||∞)",
                "||RREF(A) − I||∞",
                "Condition Number κ∞(A)",
                "Determinant det(A)",
            ],
            "Value": [
                result["residual_norm"],
                result["relative_residual"],
                result["identity_error"],
                result["condition_number"],
                result["determinant"],
            ],
            "Interpretation": [
                "Absolute residual infinity norm",
                "Scale-aware backward-error indicator",
                "Numerical deviation of the reduced left block from I",
                "Sensitivity indicator for the linear system",
                "Nonzero for a nonsingular square matrix",
            ],
        }
    )

    chart_data = pd.DataFrame(
        {
            "Variable": [
                f"x{index + 1}" for index in range(size)
            ],
            "Solution": solution,
            "Equation": [
                f"Equation {index + 1}" for index in range(size)
            ],
            "Absolute Residual": np.abs(residual),
        }
    )

    matrix_figure = create_matrix_figure(result)
    solution_figure = create_solution_figure(result)
    residual_figure = create_residual_figure(result)

    matrix_png = figure_to_png_bytes(matrix_figure)
    solution_png = figure_to_png_bytes(solution_figure)
    residual_png = figure_to_png_bytes(residual_figure)

    plt.close(matrix_figure)
    plt.close(solution_figure)
    plt.close(residual_figure)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        input_system.to_excel(
            writer,
            sheet_name="Input System",
            index=False,
        )
        final_system.to_excel(
            writer,
            sheet_name="RREF",
            index=False,
        )
        solution_table.to_excel(
            writer,
            sheet_name="Solution",
            index=False,
        )
        residual_table.to_excel(
            writer,
            sheet_name="Residual Analysis",
            index=False,
        )
        verification_table.to_excel(
            writer,
            sheet_name="Verification",
            index=False,
        )
        operations_table.to_excel(
            writer,
            sheet_name="Row Operations",
            index=False,
        )
        chart_data.to_excel(
            writer,
            sheet_name="Chart Data",
            index=False,
        )

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Gauss-Jordan Report Plots"
        plots_sheet["A1"].font = Font(bold=True, size=14)

        add_excel_image(plots_sheet, matrix_png, "A3")
        add_excel_image(plots_sheet, solution_png, "A31")
        add_excel_image(plots_sheet, residual_png, "A59")

        data_sheet = workbook["Chart Data"]

        solution_chart = BarChart()
        solution_chart.title = "Final Solution Values"
        solution_chart.y_axis.title = "Value"
        solution_chart.x_axis.title = "Variable"
        solution_chart.height = 8
        solution_chart.width = 15
        solution_chart.add_data(
            Reference(
                data_sheet,
                min_col=2,
                min_row=1,
                max_row=size + 1,
            ),
            titles_from_data=True,
        )
        solution_chart.set_categories(
            Reference(
                data_sheet,
                min_col=1,
                min_row=2,
                max_row=size + 1,
            )
        )

        residual_chart = BarChart()
        residual_chart.title = "Absolute Residual by Equation"
        residual_chart.y_axis.title = "Absolute Residual"
        residual_chart.x_axis.title = "Equation"
        residual_chart.height = 8
        residual_chart.width = 15
        residual_chart.add_data(
            Reference(
                data_sheet,
                min_col=4,
                min_row=1,
                max_row=size + 1,
            ),
            titles_from_data=True,
        )
        residual_chart.set_categories(
            Reference(
                data_sheet,
                min_col=3,
                min_row=2,
                max_row=size + 1,
            )
        )

        summary_sheet = workbook["Summary"]
        summary_sheet.add_chart(solution_chart, "D2")
        summary_sheet.add_chart(residual_chart, "D20")

        style_excel_workbook(workbook)
        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Gauss-Jordan solver page."""

    st.set_page_config(
        page_title=f"{CONFIG['title']} | Numerical Methods",
        page_icon="📘",
        layout="wide",
    )

    load_css()
    navbar(active_page="solver")

    st.html(
        f"""
        <section class="solver-hero">
            <div>
                <div class="page-label">{CONFIG['label']}</div>
                <h1>{CONFIG['title']}</h1>
                <p>{CONFIG['description']}</p>
                <div class="method-actions">
                    <a href="/{CONFIG['lesson']}" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/{CONFIG['quiz']}" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns(
        [0.035, 0.93, 0.035]
    )

    with main_area:
        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Enter the System")
                st.markdown(
                    """
                    Choose the number of equations, then enter the
                    coefficients of **A** and the right-hand side **b**
                    in the table.

                    Each row represents one equation and each `x` column
                    represents one unknown coefficient.
                    """
                )
                st.markdown("**Method formula**")
                st.latex(CONFIG["formula"])

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")

                for condition in CONFIG["conditions"]:
                    st.markdown(f"- {condition}")

                st.info(
                    "Inspect the RREF, residual, condition number, and "
                    "operation history rather than relying only on rounded "
                    "solution values."
                )

        input_column, result_column = st.columns(2)

        with input_column:
            with st.container(border=True):
                st.subheader("Input")

                size = int(
                    st.number_input(
                        "Number of equations",
                        min_value=MIN_SYSTEM_SIZE,
                        max_value=MAX_SYSTEM_SIZE,
                        value=2,
                        step=1,
                        key=f"{CONFIG['method_id']}_size",
                    )
                )

                input_table = st.data_editor(
                    create_default_table(size),
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    key=f"{CONFIG['method_id']}_table_{size}",
                )

                st.caption(
                    "Columns x1 … xn form the coefficient matrix; "
                    "the final column is b."
                )

                current_signature = create_input_signature(
                    input_table=input_table,
                    size=size,
                )

                solve_column, reset_column = st.columns(2)

                with solve_column:
                    solve_button = st.button(
                        "Solve",
                        type="primary",
                        use_container_width=True,
                        key=f"{CONFIG['method_id']}_solve",
                    )

                with reset_column:
                    reset_button = st.button(
                        "Reset Result",
                        use_container_width=True,
                        key=f"{CONFIG['method_id']}_reset",
                    )

                if reset_button:
                    st.session_state.pop(
                        f"{CONFIG['method_id']}_result",
                        None,
                    )
                    st.session_state.pop(
                        f"{CONFIG['method_id']}_excel_report",
                        None,
                    )
                    st.session_state.pop(
                        f"{CONFIG['method_id']}_excel_signature",
                        None,
                    )
                    st.rerun()

                if solve_button:
                    try:
                        matrix, vector = extract_system_from_table(
                            input_table,
                            size,
                        )
                        result = solve_linear_system(
                            matrix,
                            vector,
                            input_signature=current_signature,
                        )
                    except ValueError as error:
                        result = {
                            "status": "error",
                            "classification": "Invalid input",
                            "message": str(error),
                            "input_signature": current_signature,
                        }

                    st.session_state[
                        f"{CONFIG['method_id']}_result"
                    ] = result
                    st.session_state.pop(
                        f"{CONFIG['method_id']}_excel_report",
                        None,
                    )
                    st.session_state.pop(
                        f"{CONFIG['method_id']}_excel_signature",
                        None,
                    )
                    st.rerun()

        with result_column:
            with st.container(border=True):
                st.subheader("Final Result")

                result = st.session_state.get(
                    f"{CONFIG['method_id']}_result"
                )

                if result is None:
                    st.info(
                        "Enter the system and click Solve to display "
                        "the result."
                    )

                elif result.get("input_signature") != current_signature:
                    st.info(
                        "The matrix entries or system size have changed. "
                        "Click Solve to calculate a new result."
                    )

                elif result["status"] == "error":
                    st.error(result["message"])
                    st.markdown(
                        f"**System classification:** "
                        f"{result.get('classification', 'Not available')}"
                    )

                    if "rank_A" in result:
                        rank_columns = st.columns(2)
                        rank_columns[0].metric(
                            "Rank(A)",
                            result["rank_A"],
                        )
                        rank_columns[1].metric(
                            "Rank([A|b])",
                            result["rank_augmented"],
                        )

                    if "final_matrix" in result:
                        st.markdown("**Reduced augmented matrix**")
                        st.code(
                            augmented_text(
                                result["final_matrix"],
                                result["final_vector"],
                            ),
                            language="text",
                        )

                else:
                    st.success(result["message"])
                    st.markdown(
                        f"**System classification:** "
                        f"{result['classification']}"
                    )

                    metric_column1, metric_column2 = st.columns(2)

                    metric_column1.metric(
                        "Residual Norm",
                        format_number(result["residual_norm"]),
                    )
                    metric_column2.metric(
                        "Relative Residual",
                        format_number(result["relative_residual"]),
                    )

                    metric_column3, metric_column4 = st.columns(2)

                    metric_column3.metric(
                        "Row Operations",
                        result["operations"],
                    )
                    metric_column4.metric(
                        "Condition Number",
                        format_number(result["condition_number"]),
                    )

                    st.markdown("**Solution Vector**")
                    st.dataframe(
                        round_numeric_dataframe(
                            solution_dataframe(result)
                        ),
                        use_container_width=True,
                        hide_index=True,
                    )

                    for warning in result["warnings"]:
                        st.warning(warning)

        result = st.session_state.get(
            f"{CONFIG['method_id']}_result"
        )

        result_is_current = (
            result is not None
            and result.get("input_signature") == current_signature
        )

        if result_is_current and result["status"] == "success":
            with st.container(border=True):
                st.subheader(result["final_form_name"])
                st.code(
                    augmented_text(
                        result["final_matrix"],
                        result["final_vector"],
                    ),
                    language="text",
                )

                verification_columns = st.columns(4)
                verification_columns[0].metric(
                    "Rank(A)",
                    result["rank_A"],
                )
                verification_columns[1].metric(
                    "Rank([A|b])",
                    result["rank_augmented"],
                )
                verification_columns[2].metric(
                    "Determinant",
                    format_number(result["determinant"]),
                )
                verification_columns[3].metric(
                    "RREF Identity Error",
                    format_number(result["identity_error"]),
                )

            with st.container(border=True):
                st.subheader("Residual Verification")
                st.dataframe(
                    round_numeric_dataframe(
                        residual_dataframe(result)
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

            with st.container(border=True):
                st.subheader("Operation History")

                history_frame = operations_dataframe(result)

                st.dataframe(
                    round_numeric_dataframe(
                        history_frame.drop(
                            columns=["Augmented Matrix"],
                            errors="ignore",
                        )
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                with st.expander(
                    "View Intermediate Matrices",
                    expanded=False,
                ):
                    if not result["history"]:
                        st.info("No row operations were required.")
                    else:
                        for item in result["history"]:
                            st.markdown(
                                f"**Step {item['Step']}: "
                                f"{item['Operation']}**"
                            )
                            st.code(
                                item["Augmented Matrix"],
                                language="text",
                            )

                st.subheader("Excel Report")

                report_cache_key = (
                    f"{CONFIG['method_id']}_excel_report"
                )
                report_signature_key = (
                    f"{CONFIG['method_id']}_excel_signature"
                )

                try:
                    if (
                        st.session_state.get(report_signature_key)
                        != current_signature
                        or report_cache_key not in st.session_state
                    ):
                        st.session_state[report_cache_key] = (
                            create_excel_report(result)
                        )
                        st.session_state[report_signature_key] = (
                            current_signature
                        )

                    st.download_button(
                        "Download Complete Excel Report",
                        data=st.session_state[report_cache_key],
                        file_name=(
                            f"{CONFIG['method_id']}_complete_report.xlsx"
                        ),
                        mime=EXCEL_MIME_TYPE,
                        use_container_width=True,
                        key=f"{CONFIG['method_id']}_excel_download",
                    )

                except (
                    ValueError,
                    OSError,
                    TypeError,
                    ArithmeticError,
                ) as error:
                    st.error(
                        "The Excel report could not be generated. "
                        f"Details: {error}"
                    )

            matrix_column, solution_column, residual_column = st.columns(3)

            with matrix_column:
                with st.container(border=True):
                    st.subheader("Coefficient Matrix")
                    matrix_figure = create_matrix_figure(result)
                    st.pyplot(
                        matrix_figure,
                        use_container_width=True,
                    )
                    plt.close(matrix_figure)

            with solution_column:
                with st.container(border=True):
                    st.subheader("Solution Values")
                    solution_figure = create_solution_figure(result)
                    st.pyplot(
                        solution_figure,
                        use_container_width=True,
                    )
                    plt.close(solution_figure)

            with residual_column:
                with st.container(border=True):
                    st.subheader("Residual Components")
                    residual_figure = create_residual_figure(result)
                    st.pyplot(
                        residual_figure,
                        use_container_width=True,
                    )
                    plt.close(residual_figure)

            with st.container(border=True):
                st.subheader("Continue Learning")
                navigation_column1, navigation_column2 = st.columns(2)

                with navigation_column1:
                    if st.button(
                        "Review Lesson",
                        use_container_width=True,
                        key=f"{CONFIG['method_id']}_lesson_button",
                    ):
                        st.switch_page(
                            f"pages/{CONFIG['lesson']}.py"
                        )

                with navigation_column2:
                    if st.button(
                        "Back to Solver Menu",
                        use_container_width=True,
                        key=f"{CONFIG['method_id']}_menu_button",
                    ):
                        st.switch_page("pages/Numerical_Solver.py")

    st.html(
        f"""
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>{CONFIG['footer']}</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()