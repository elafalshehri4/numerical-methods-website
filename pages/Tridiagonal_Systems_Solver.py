from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from components.navigation import navbar
from utilities.ui import load_css


# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Show three decimals and use × 10ⁿ when fixed notation loses meaning."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"

from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series

# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Thomas Algorithm for Tridiagonal Systems"
SUPPORTED_SIZES = (2, 3, 4, 5)
DEFAULT_SIZE = 4
DISPLAY_DECIMALS = 3
STRUCTURE_TOLERANCE = 1.0e-12
PIVOT_TOLERANCE = 1.0e-12
CONDITION_NUMBER_WARNING = 1.0e12
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DEFAULT_SYSTEMS: dict[int, tuple[np.ndarray, np.ndarray]] = {
    2: (
        np.array([[2.0, -1.0], [-1.0, 2.0]], dtype=float),
        np.array([1.0, 0.0], dtype=float),
    ),
    3: (
        np.array(
            [
                [2.0, -1.0, 0.0],
                [-1.0, 2.0, -1.0],
                [0.0, -1.0, 2.0],
            ],
            dtype=float,
        ),
        np.array([1.0, 0.0, 1.0], dtype=float),
    ),
    4: (
        np.array(
            [
                [2.0, -1.0, 0.0, 0.0],
                [-1.0, 2.0, -1.0, 0.0],
                [0.0, -1.0, 2.0, -1.0],
                [0.0, 0.0, -1.0, 2.0],
            ],
            dtype=float,
        ),
        np.array([1.0, 0.0, 0.0, 1.0], dtype=float),
    ),
    5: (
        np.array(
            [
                [4.0, -1.0, 0.0, 0.0, 0.0],
                [-1.0, 4.0, -1.0, 0.0, 0.0],
                [0.0, -1.0, 4.0, -1.0, 0.0],
                [0.0, 0.0, -1.0, 4.0, -1.0],
                [0.0, 0.0, 0.0, -1.0, 4.0],
            ],
            dtype=float,
        ),
        np.array([2.0, 4.0, 6.0, 8.0, 10.0], dtype=float),
    ),
}


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class SystemDiagnostics:
    """Diagnostic information about the original linear system."""

    rank_a: int
    rank_augmented: int
    classification: str
    condition_number: float
    diagnostic_tolerance: float


@dataclass(frozen=True)
class ForwardSweepRecord:
    """One row of the Thomas Algorithm forward sweep."""

    step: int
    row: int
    lower_coefficient: float
    main_coefficient: float
    upper_coefficient: float
    right_hand_side: float
    previous_modified_upper: float | None
    previous_modified_rhs: float | None
    denominator: float
    modified_upper: float
    modified_rhs: float
    denominator_formula: str
    modified_upper_formula: str
    modified_rhs_formula: str
    status: str
    notes: str


@dataclass(frozen=True)
class BackSubstitutionRecord:
    """One row of the Thomas Algorithm back-substitution phase."""

    step: int
    row: int
    variable: str
    modified_upper: float
    modified_rhs: float
    known_next_value: float | None
    calculated_value: float
    formula: str
    status: str


@dataclass(frozen=True)
class TridiagonalResult:
    """Complete solver result used by Streamlit and the Excel exporter."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str
    original_matrix: np.ndarray
    original_vector: np.ndarray
    lower_diagonal: np.ndarray
    main_diagonal: np.ndarray
    upper_diagonal: np.ndarray
    modified_upper: np.ndarray
    modified_rhs: np.ndarray
    solution: np.ndarray | None
    forward_history: tuple[ForwardSweepRecord, ...]
    back_history: tuple[BackSubstitutionRecord, ...]
    computed_right_hand_side: np.ndarray | None
    residual_vector: np.ndarray | None
    residual_norm_2: float | None
    residual_norm_inf: float | None
    maximum_absolute_residual: float | None
    relative_residual: float | None
    diagnostics: SystemDiagnostics | None
    warnings: tuple[str, ...]
    forward_steps: int
    back_substitution_steps: int
    pivot_tolerance: float
    structure_tolerance: float
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def format_number(
    value: float | int | None,
    decimals: int = 3,
) -> str:
    """Format displayed values with three decimals and × 10ⁿ notation."""

    return format_display_number(value, decimals)



def variable_names(size: int) -> list[str]:
    """Return user-facing variable labels."""

    return [f"x{i + 1}" for i in range(size)]


def equation_names(size: int) -> list[str]:
    """Return user-facing equation labels."""

    return [f"Equation {i + 1}" for i in range(size)]


def create_input_signature(
    raw_matrix: Sequence[Sequence[Any]],
    raw_vector: Sequence[Any],
    pivot_tolerance: float,
    structure_tolerance: float,
) -> str:
    """Create a stable signature used to detect stale Streamlit results."""

    serialized = repr(
        (
            [[str(value).strip() for value in row] for row in raw_matrix],
            [str(value).strip() for value in raw_vector],
            float(pivot_tolerance),
            float(structure_tolerance),
        )
    )
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric DataFrame columns for display without altering results."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def matrix_dataframe(matrix: np.ndarray) -> pd.DataFrame:
    """Create a labeled DataFrame for a coefficient matrix."""

    size = matrix.shape[0]
    return pd.DataFrame(
        matrix,
        index=equation_names(size),
        columns=variable_names(size),
    )


def vector_dataframe(vector: np.ndarray, column_name: str = "b") -> pd.DataFrame:
    """Create a labeled DataFrame for a vector."""

    return pd.DataFrame(
        {column_name: vector},
        index=equation_names(len(vector)),
    )


def augmented_dataframe(matrix: np.ndarray, vector: np.ndarray) -> pd.DataFrame:
    """Create a labeled DataFrame for the augmented matrix [A | b]."""

    dataframe = matrix_dataframe(matrix)
    dataframe["b"] = vector
    return dataframe


def diagonal_dataframe(result: TridiagonalResult) -> pd.DataFrame:
    """Create a table containing the three diagonals and right-hand side."""

    size = len(result.main_diagonal)
    return pd.DataFrame(
        {
            "Row": np.arange(1, size + 1),
            "Lower Diagonal aᵢ": result.lower_diagonal,
            "Main Diagonal bᵢ": result.main_diagonal,
            "Upper Diagonal cᵢ": result.upper_diagonal,
            "Right-Hand Side dᵢ": result.original_vector,
        }
    )


def serialize_array(values: np.ndarray, decimals: int = 12) -> str:
    """Serialize an array as readable text for Excel cells."""

    return np.array2string(
        np.asarray(values, dtype=float),
        precision=decimals,
        suppress_small=False,
        separator=", ",
        max_line_width=10_000,
    )


# =============================================================================
# Input conversion and validation
# =============================================================================
def parse_finite_real(value: Any, field_name: str) -> float:
    """Convert one input to a finite real floating-point value."""

    if value is None:
        raise ValueError(f"{field_name} is empty.")

    if isinstance(value, str):
        cleaned_value = value.strip()
        if not cleaned_value:
            raise ValueError(f"{field_name} is empty.")
    else:
        cleaned_value = value

    try:
        numeric_value = float(cleaned_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{field_name} must contain a valid real number.") from error

    if not math.isfinite(numeric_value):
        raise ValueError(f"{field_name} must be finite; NaN and infinity are not allowed.")

    return numeric_value


def convert_and_validate_inputs(
    raw_matrix: Sequence[Sequence[Any]],
    raw_vector: Sequence[Any],
    pivot_tolerance: float,
    structure_tolerance: float,
) -> tuple[np.ndarray, np.ndarray]:
    """Convert raw values and validate dimensions, finiteness, and structure."""

    if not math.isfinite(float(pivot_tolerance)) or pivot_tolerance <= 0:
        raise ValueError("Pivot tolerance must be a positive finite number.")

    if not math.isfinite(float(structure_tolerance)) or structure_tolerance < 0:
        raise ValueError("Structure tolerance must be a finite non-negative number.")

    if not isinstance(raw_matrix, Sequence) or isinstance(raw_matrix, (str, bytes)):
        raise ValueError("Matrix A must be a two-dimensional collection of values.")

    if not isinstance(raw_vector, Sequence) or isinstance(raw_vector, (str, bytes)):
        raise ValueError("Vector b must be a one-dimensional collection of values.")

    size = len(raw_matrix)
    if size not in SUPPORTED_SIZES:
        supported_text = ", ".join(str(value) for value in SUPPORTED_SIZES)
        raise ValueError(f"The system size must be one of: {supported_text}.")

    if len(raw_vector) != size:
        raise ValueError("Vector b must contain one value for every matrix row.")

    matrix = np.empty((size, size), dtype=float)
    vector = np.empty(size, dtype=float)

    for row_index, raw_row in enumerate(raw_matrix):
        if not isinstance(raw_row, Sequence) or isinstance(raw_row, (str, bytes)):
            raise ValueError(f"Row {row_index + 1} of matrix A is invalid.")
        if len(raw_row) != size:
            raise ValueError(
                f"Row {row_index + 1} must contain exactly {size} coefficients."
            )
        for column_index, raw_value in enumerate(raw_row):
            matrix[row_index, column_index] = parse_finite_real(
                raw_value,
                f"A[{row_index + 1}, {column_index + 1}]",
            )

    for row_index, raw_value in enumerate(raw_vector):
        vector[row_index] = parse_finite_real(raw_value, f"b[{row_index + 1}]")

    if matrix.ndim != 2 or matrix.shape[0] != matrix.shape[1]:
        raise ValueError("Matrix A must be square.")

    if not np.all(np.isfinite(matrix)) or not np.all(np.isfinite(vector)):
        raise ValueError("The system contains NaN or infinity.")

    off_band_entries: list[str] = []
    for row_index in range(size):
        for column_index in range(size):
            if abs(row_index - column_index) > 1:
                value = matrix[row_index, column_index]
                if abs(value) > structure_tolerance:
                    off_band_entries.append(
                        f"A[{row_index + 1}, {column_index + 1}] = "
                        f"{format_number(value)}"
                    )

    if off_band_entries:
        preview = "; ".join(off_band_entries[:5])
        extra_count = len(off_band_entries) - 5
        if extra_count > 0:
            preview += f"; and {extra_count} additional off-band value(s)"
        raise ValueError(
            "Matrix A is not tridiagonal. Every coefficient outside the main, "
            "lower, and upper diagonals must be zero. Detected: " + preview + "."
        )

    # Remove harmless floating-point noise outside the tridiagonal band.
    for row_index in range(size):
        for column_index in range(size):
            if abs(row_index - column_index) > 1:
                matrix[row_index, column_index] = 0.0

    return matrix, vector


# =============================================================================
# Diagnostics and numerical analysis
# =============================================================================
def analyze_linear_system(
    matrix: np.ndarray,
    vector: np.ndarray,
    pivot_tolerance: float,
) -> SystemDiagnostics:
    """Classify the system diagnostically without calculating the solution."""

    size = matrix.shape[0]
    matrix_scale = max(float(np.linalg.norm(matrix, ord=np.inf)), 1.0)
    diagnostic_tolerance = max(
        pivot_tolerance,
        np.finfo(float).eps * size * matrix_scale,
    )

    rank_a = int(np.linalg.matrix_rank(matrix, tol=diagnostic_tolerance))
    augmented = np.column_stack((matrix, vector))
    rank_augmented = int(
        np.linalg.matrix_rank(augmented, tol=diagnostic_tolerance)
    )

    if rank_a < rank_augmented:
        classification = "Inconsistent system: no solution exists."
    elif rank_a < size:
        classification = "Singular system: infinitely many solutions may exist."
    else:
        classification = "Nonsingular system with a unique solution."

    try:
        condition_number = float(np.linalg.cond(matrix))
    except np.linalg.LinAlgError:
        condition_number = float("inf")

    return SystemDiagnostics(
        rank_a=rank_a,
        rank_augmented=rank_augmented,
        classification=classification,
        condition_number=condition_number,
        diagnostic_tolerance=diagnostic_tolerance,
    )


def extract_tridiagonal_coefficients(
    matrix: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Extract lower, main, and upper diagonal vectors from matrix A."""

    size = matrix.shape[0]
    lower = np.zeros(size, dtype=float)
    main = np.diag(matrix).astype(float).copy()
    upper = np.zeros(size, dtype=float)

    if size > 1:
        lower[1:] = np.diag(matrix, k=-1)
        upper[:-1] = np.diag(matrix, k=1)

    return lower, main, upper


def calculate_residual_analysis(
    matrix: np.ndarray,
    vector: np.ndarray,
    solution: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, float, float, float, float]:
    """Calculate residual vector and scale-aware residual measures."""

    computed_rhs = matrix @ solution
    residual = computed_rhs - vector
    residual_norm_2 = float(np.linalg.norm(residual, ord=2))
    residual_norm_inf = float(np.linalg.norm(residual, ord=np.inf))
    maximum_absolute_residual = float(np.max(np.abs(residual)))

    denominator = (
        float(np.linalg.norm(matrix, ord=2))
        * float(np.linalg.norm(solution, ord=2))
        + float(np.linalg.norm(vector, ord=2))
    )
    if denominator > np.finfo(float).tiny:
        relative_residual = residual_norm_2 / denominator
    else:
        relative_residual = 0.0 if residual_norm_2 == 0.0 else float("inf")

    return (
        computed_rhs,
        residual,
        residual_norm_2,
        residual_norm_inf,
        maximum_absolute_residual,
        relative_residual,
    )


def build_error_result(
    *,
    message: str,
    stopping_reason: str,
    input_signature: str,
    original_matrix: np.ndarray | None = None,
    original_vector: np.ndarray | None = None,
    lower_diagonal: np.ndarray | None = None,
    main_diagonal: np.ndarray | None = None,
    upper_diagonal: np.ndarray | None = None,
    modified_upper: np.ndarray | None = None,
    modified_rhs: np.ndarray | None = None,
    forward_history: Sequence[ForwardSweepRecord] = (),
    back_history: Sequence[BackSubstitutionRecord] = (),
    diagnostics: SystemDiagnostics | None = None,
    warnings: Sequence[str] = (),
    pivot_tolerance: float = PIVOT_TOLERANCE,
    structure_tolerance: float = STRUCTURE_TOLERANCE,
) -> TridiagonalResult:
    """Construct a consistent non-crashing error result."""

    empty_matrix = np.empty((0, 0), dtype=float)
    empty_vector = np.empty(0, dtype=float)
    return TridiagonalResult(
        status="Error",
        success=False,
        method=METHOD_NAME,
        message=message,
        stopping_reason=stopping_reason,
        original_matrix=(
            original_matrix.copy() if original_matrix is not None else empty_matrix
        ),
        original_vector=(
            original_vector.copy() if original_vector is not None else empty_vector
        ),
        lower_diagonal=(
            lower_diagonal.copy() if lower_diagonal is not None else empty_vector
        ),
        main_diagonal=(
            main_diagonal.copy() if main_diagonal is not None else empty_vector
        ),
        upper_diagonal=(
            upper_diagonal.copy() if upper_diagonal is not None else empty_vector
        ),
        modified_upper=(
            modified_upper.copy() if modified_upper is not None else empty_vector
        ),
        modified_rhs=(
            modified_rhs.copy() if modified_rhs is not None else empty_vector
        ),
        solution=None,
        forward_history=tuple(forward_history),
        back_history=tuple(back_history),
        computed_right_hand_side=None,
        residual_vector=None,
        residual_norm_2=None,
        residual_norm_inf=None,
        maximum_absolute_residual=None,
        relative_residual=None,
        diagnostics=diagnostics,
        warnings=tuple(warnings),
        forward_steps=len(forward_history),
        back_substitution_steps=len(back_history),
        pivot_tolerance=float(pivot_tolerance),
        structure_tolerance=float(structure_tolerance),
        input_signature=input_signature,
        execution_datetime=datetime.now(ZoneInfo(REPORT_TIME_ZONE)),
    )


# =============================================================================
# Thomas Algorithm
# =============================================================================
def solve_tridiagonal_system(
    raw_matrix: Sequence[Sequence[Any]],
    raw_vector: Sequence[Any],
    input_signature: str,
    pivot_tolerance: float = PIVOT_TOLERANCE,
    structure_tolerance: float = STRUCTURE_TOLERANCE,
) -> TridiagonalResult:
    """Solve a tridiagonal system with the standard Thomas Algorithm.

    The method performs the modified-coefficient forward sweep:

        c'_0 = c_0 / b_0
        d'_0 = d_0 / b_0

        denominator_i = b_i - a_i c'_(i-1)
        c'_i = c_i / denominator_i
        d'_i = (d_i - a_i d'_(i-1)) / denominator_i

    followed by back substitution:

        x_(n-1) = d'_(n-1)
        x_i = d'_i - c'_i x_(i+1)

    No row pivoting is performed because the standard Thomas Algorithm assumes
    every modified pivot is nonzero. A zero or near-zero denominator stops the
    method safely and reports that pivoting or another solver may be required.
    """

    try:
        matrix, vector = convert_and_validate_inputs(
            raw_matrix,
            raw_vector,
            pivot_tolerance,
            structure_tolerance,
        )
    except ValueError as error:
        return build_error_result(
            message=str(error),
            stopping_reason="Input validation failed.",
            input_signature=input_signature,
            pivot_tolerance=pivot_tolerance,
            structure_tolerance=structure_tolerance,
        )

    original_matrix = matrix.copy()
    original_vector = vector.copy()
    diagnostics = analyze_linear_system(matrix, vector, pivot_tolerance)
    lower, main, upper = extract_tridiagonal_coefficients(matrix)
    warnings: list[str] = []

    if not math.isfinite(diagnostics.condition_number):
        warnings.append(
            "The original coefficient matrix has a non-finite condition number "
            "and is singular or numerically unusable."
        )
    elif diagnostics.condition_number >= CONDITION_NUMBER_WARNING:
        warnings.append(
            "The coefficient matrix is highly ill-conditioned. The computed "
            "solution may be sensitive to small input or rounding changes."
        )

    size = matrix.shape[0]
    if diagnostics.rank_a < diagnostics.rank_augmented:
        return build_error_result(
            message="The tridiagonal system is inconsistent and has no solution.",
            stopping_reason=diagnostics.classification,
            input_signature=input_signature,
            original_matrix=original_matrix,
            original_vector=original_vector,
            lower_diagonal=lower,
            main_diagonal=main,
            upper_diagonal=upper,
            diagnostics=diagnostics,
            warnings=warnings,
            pivot_tolerance=pivot_tolerance,
            structure_tolerance=structure_tolerance,
        )

    if diagnostics.rank_a < size:
        return build_error_result(
            message=(
                "The coefficient matrix is singular. The Thomas Algorithm "
                "requires a unique solution and cannot continue."
            ),
            stopping_reason=diagnostics.classification,
            input_signature=input_signature,
            original_matrix=original_matrix,
            original_vector=original_vector,
            lower_diagonal=lower,
            main_diagonal=main,
            upper_diagonal=upper,
            diagnostics=diagnostics,
            warnings=warnings,
            pivot_tolerance=pivot_tolerance,
            structure_tolerance=structure_tolerance,
        )

    modified_upper = np.zeros(size, dtype=float)
    modified_rhs = np.zeros(size, dtype=float)
    forward_history: list[ForwardSweepRecord] = []
    back_history: list[BackSubstitutionRecord] = []

    # -------------------------------------------------------------------------
    # Forward sweep: first row
    # -------------------------------------------------------------------------
    first_denominator = float(main[0])
    if abs(first_denominator) <= pivot_tolerance:
        return build_error_result(
            message=(
                "The Thomas Algorithm cannot start because the first main-"
                "diagonal pivot is zero or numerically near zero. This system "
                "may require row pivoting or a different linear-system method."
            ),
            stopping_reason=(
                f"First modified pivot |b1| = {format_scientific_power(abs(first_denominator))} "
                f"is not greater than {format_scientific_power(pivot_tolerance)}."
            ),
            input_signature=input_signature,
            original_matrix=original_matrix,
            original_vector=original_vector,
            lower_diagonal=lower,
            main_diagonal=main,
            upper_diagonal=upper,
            modified_upper=modified_upper,
            modified_rhs=modified_rhs,
            diagnostics=diagnostics,
            warnings=warnings,
            pivot_tolerance=pivot_tolerance,
            structure_tolerance=structure_tolerance,
        )

    modified_upper[0] = upper[0] / first_denominator if size > 1 else 0.0
    modified_rhs[0] = vector[0] / first_denominator

    if not np.isfinite(modified_upper[0]) or not np.isfinite(modified_rhs[0]):
        return build_error_result(
            message="The first forward-sweep calculation produced a non-finite value.",
            stopping_reason="Non-finite arithmetic occurred during the forward sweep.",
            input_signature=input_signature,
            original_matrix=original_matrix,
            original_vector=original_vector,
            lower_diagonal=lower,
            main_diagonal=main,
            upper_diagonal=upper,
            modified_upper=modified_upper,
            modified_rhs=modified_rhs,
            diagnostics=diagnostics,
            warnings=warnings,
            pivot_tolerance=pivot_tolerance,
            structure_tolerance=structure_tolerance,
        )

    forward_history.append(
        ForwardSweepRecord(
            step=1,
            row=1,
            lower_coefficient=lower[0],
            main_coefficient=main[0],
            upper_coefficient=upper[0],
            right_hand_side=vector[0],
            previous_modified_upper=None,
            previous_modified_rhs=None,
            denominator=first_denominator,
            modified_upper=modified_upper[0],
            modified_rhs=modified_rhs[0],
            denominator_formula="p₁ = b₁",
            modified_upper_formula=(
                "c′₁ = c₁ / p₁" if size > 1 else "c′₁ = 0 (last row)"
            ),
            modified_rhs_formula="d′₁ = d₁ / p₁",
            status="Completed",
            notes="Initial modified coefficients calculated.",
        )
    )

    # -------------------------------------------------------------------------
    # Forward sweep: rows 2 through n
    # -------------------------------------------------------------------------
    for row_index in range(1, size):
        denominator = main[row_index] - lower[row_index] * modified_upper[row_index - 1]

        if not math.isfinite(float(denominator)):
            return build_error_result(
                message=(
                    f"The modified pivot at row {row_index + 1} is non-finite. "
                    "The Thomas Algorithm cannot continue."
                ),
                stopping_reason="Non-finite modified pivot encountered.",
                input_signature=input_signature,
                original_matrix=original_matrix,
                original_vector=original_vector,
                lower_diagonal=lower,
                main_diagonal=main,
                upper_diagonal=upper,
                modified_upper=modified_upper,
                modified_rhs=modified_rhs,
                forward_history=forward_history,
                diagnostics=diagnostics,
                warnings=warnings,
                pivot_tolerance=pivot_tolerance,
                structure_tolerance=structure_tolerance,
            )

        if abs(denominator) <= pivot_tolerance:
            return build_error_result(
                message=(
                    f"The Thomas Algorithm stopped at row {row_index + 1} because "
                    "the modified pivot is zero or numerically near zero. The "
                    "system may still have a unique solution, but the standard "
                    "Thomas Algorithm without pivoting is not applicable."
                ),
                stopping_reason=(
                    f"Modified pivot |p{row_index + 1}| = {format_scientific_power(abs(denominator))} "
                    f"is not greater than {format_scientific_power(pivot_tolerance)}."
                ),
                input_signature=input_signature,
                original_matrix=original_matrix,
                original_vector=original_vector,
                lower_diagonal=lower,
                main_diagonal=main,
                upper_diagonal=upper,
                modified_upper=modified_upper,
                modified_rhs=modified_rhs,
                forward_history=forward_history,
                diagnostics=diagnostics,
                warnings=warnings,
                pivot_tolerance=pivot_tolerance,
                structure_tolerance=structure_tolerance,
            )

        if row_index < size - 1:
            modified_upper[row_index] = upper[row_index] / denominator
        else:
            modified_upper[row_index] = 0.0

        modified_rhs[row_index] = (
            vector[row_index]
            - lower[row_index] * modified_rhs[row_index - 1]
        ) / denominator

        if not np.isfinite(modified_upper[row_index]) or not np.isfinite(
            modified_rhs[row_index]
        ):
            return build_error_result(
                message=(
                    f"Forward-sweep row {row_index + 1} produced a non-finite "
                    "modified coefficient."
                ),
                stopping_reason="Non-finite arithmetic occurred during the forward sweep.",
                input_signature=input_signature,
                original_matrix=original_matrix,
                original_vector=original_vector,
                lower_diagonal=lower,
                main_diagonal=main,
                upper_diagonal=upper,
                modified_upper=modified_upper,
                modified_rhs=modified_rhs,
                forward_history=forward_history,
                diagnostics=diagnostics,
                warnings=warnings,
                pivot_tolerance=pivot_tolerance,
                structure_tolerance=structure_tolerance,
            )

        forward_history.append(
            ForwardSweepRecord(
                step=row_index + 1,
                row=row_index + 1,
                lower_coefficient=lower[row_index],
                main_coefficient=main[row_index],
                upper_coefficient=upper[row_index],
                right_hand_side=vector[row_index],
                previous_modified_upper=modified_upper[row_index - 1],
                previous_modified_rhs=modified_rhs[row_index - 1],
                denominator=denominator,
                modified_upper=modified_upper[row_index],
                modified_rhs=modified_rhs[row_index],
                denominator_formula=(
                    f"p{row_index + 1} = b{row_index + 1} − "
                    f"a{row_index + 1}c′{row_index}"
                ),
                modified_upper_formula=(
                    f"c′{row_index + 1} = c{row_index + 1} / p{row_index + 1}"
                    if row_index < size - 1
                    else f"c′{row_index + 1} = 0 (last row)"
                ),
                modified_rhs_formula=(
                    f"d′{row_index + 1} = (d{row_index + 1} − "
                    f"a{row_index + 1}d′{row_index}) / p{row_index + 1}"
                ),
                status="Completed",
                notes="Modified coefficients calculated.",
            )
        )

    # -------------------------------------------------------------------------
    # Back substitution
    # -------------------------------------------------------------------------
    solution = np.zeros(size, dtype=float)
    solution[-1] = modified_rhs[-1]
    if not math.isfinite(float(solution[-1])):
        return build_error_result(
            message="The final variable became non-finite during back substitution.",
            stopping_reason="Non-finite arithmetic occurred during back substitution.",
            input_signature=input_signature,
            original_matrix=original_matrix,
            original_vector=original_vector,
            lower_diagonal=lower,
            main_diagonal=main,
            upper_diagonal=upper,
            modified_upper=modified_upper,
            modified_rhs=modified_rhs,
            forward_history=forward_history,
            diagnostics=diagnostics,
            warnings=warnings,
            pivot_tolerance=pivot_tolerance,
            structure_tolerance=structure_tolerance,
        )

    back_history.append(
        BackSubstitutionRecord(
            step=1,
            row=size,
            variable=f"x{size}",
            modified_upper=modified_upper[-1],
            modified_rhs=modified_rhs[-1],
            known_next_value=None,
            calculated_value=solution[-1],
            formula=f"x{size} = d′{size}",
            status="Completed",
        )
    )

    back_step = 2
    for row_index in range(size - 2, -1, -1):
        solution[row_index] = (
            modified_rhs[row_index]
            - modified_upper[row_index] * solution[row_index + 1]
        )

        if not math.isfinite(float(solution[row_index])):
            return build_error_result(
                message=(
                    f"Back substitution produced a non-finite value for "
                    f"x{row_index + 1}."
                ),
                stopping_reason="Non-finite arithmetic occurred during back substitution.",
                input_signature=input_signature,
                original_matrix=original_matrix,
                original_vector=original_vector,
                lower_diagonal=lower,
                main_diagonal=main,
                upper_diagonal=upper,
                modified_upper=modified_upper,
                modified_rhs=modified_rhs,
                forward_history=forward_history,
                back_history=back_history,
                diagnostics=diagnostics,
                warnings=warnings,
                pivot_tolerance=pivot_tolerance,
                structure_tolerance=structure_tolerance,
            )

        back_history.append(
            BackSubstitutionRecord(
                step=back_step,
                row=row_index + 1,
                variable=f"x{row_index + 1}",
                modified_upper=modified_upper[row_index],
                modified_rhs=modified_rhs[row_index],
                known_next_value=solution[row_index + 1],
                calculated_value=solution[row_index],
                formula=(
                    f"x{row_index + 1} = d′{row_index + 1} − "
                    f"c′{row_index + 1}x{row_index + 2}"
                ),
                status="Completed",
            )
        )
        back_step += 1

    (
        computed_rhs,
        residual,
        residual_norm_2,
        residual_norm_inf,
        maximum_absolute_residual,
        relative_residual,
    ) = calculate_residual_analysis(original_matrix, original_vector, solution)

    return TridiagonalResult(
        status="Success",
        success=True,
        method=METHOD_NAME,
        message="Execution completed successfully.",
        stopping_reason=(
            "The forward sweep and back substitution were completed with all "
            "modified pivots greater than the numerical pivot tolerance."
        ),
        original_matrix=original_matrix,
        original_vector=original_vector,
        lower_diagonal=lower,
        main_diagonal=main,
        upper_diagonal=upper,
        modified_upper=modified_upper,
        modified_rhs=modified_rhs,
        solution=solution,
        forward_history=tuple(forward_history),
        back_history=tuple(back_history),
        computed_right_hand_side=computed_rhs,
        residual_vector=residual,
        residual_norm_2=residual_norm_2,
        residual_norm_inf=residual_norm_inf,
        maximum_absolute_residual=maximum_absolute_residual,
        relative_residual=relative_residual,
        diagnostics=diagnostics,
        warnings=tuple(warnings),
        forward_steps=len(forward_history),
        back_substitution_steps=len(back_history),
        pivot_tolerance=float(pivot_tolerance),
        structure_tolerance=float(structure_tolerance),
        input_signature=input_signature,
        execution_datetime=datetime.now(ZoneInfo(REPORT_TIME_ZONE)),
    )


# =============================================================================
# DataFrame builders
# =============================================================================
def build_forward_dataframe(result: TridiagonalResult) -> pd.DataFrame:
    """Create the concise forward-sweep display table."""

    rows = []
    for record in result.forward_history:
        rows.append(
            {
                "Step": record.step,
                "Row": record.row,
                "aᵢ": record.lower_coefficient,
                "bᵢ": record.main_coefficient,
                "cᵢ": record.upper_coefficient,
                "dᵢ": record.right_hand_side,
                "Modified Pivot pᵢ": record.denominator,
                "Modified Upper c′ᵢ": record.modified_upper,
                "Modified RHS d′ᵢ": record.modified_rhs,
                "Status": record.status,
            }
        )
    return pd.DataFrame(rows)


def build_forward_export_dataframe(result: TridiagonalResult) -> pd.DataFrame:
    """Create the detailed forward-sweep table for Excel."""

    rows = []
    for record in result.forward_history:
        rows.append(
            {
                "Step": record.step,
                "Row": record.row,
                "Lower Coefficient a_i": record.lower_coefficient,
                "Main Coefficient b_i": record.main_coefficient,
                "Upper Coefficient c_i": record.upper_coefficient,
                "Right-Hand Side d_i": record.right_hand_side,
                "Previous Modified Upper": record.previous_modified_upper,
                "Previous Modified RHS": record.previous_modified_rhs,
                "Modified Pivot": record.denominator,
                "Modified Upper": record.modified_upper,
                "Modified RHS": record.modified_rhs,
                "Pivot Formula": record.denominator_formula,
                "Modified Upper Formula": record.modified_upper_formula,
                "Modified RHS Formula": record.modified_rhs_formula,
                "Status": record.status,
                "Notes": record.notes,
            }
        )
    return pd.DataFrame(rows)


def build_back_dataframe(result: TridiagonalResult) -> pd.DataFrame:
    """Create the back-substitution table."""

    rows = []
    for record in result.back_history:
        rows.append(
            {
                "Step": record.step,
                "Row": record.row,
                "Variable": record.variable,
                "Modified Upper c′ᵢ": record.modified_upper,
                "Modified RHS d′ᵢ": record.modified_rhs,
                "Known Next Value": record.known_next_value,
                "Calculated Value": record.calculated_value,
                "Formula": record.formula,
                "Status": record.status,
            }
        )
    return pd.DataFrame(rows)


def build_solution_dataframe(result: TridiagonalResult) -> pd.DataFrame:
    """Create the final solution table."""

    if result.solution is None:
        return pd.DataFrame()
    return pd.DataFrame(
        {
            "Variable": variable_names(len(result.solution)),
            "Calculated Value": result.solution,
        }
    )


def build_residual_dataframe(result: TridiagonalResult) -> pd.DataFrame:
    """Create the equation-by-equation residual table."""

    if (
        result.solution is None
        or result.computed_right_hand_side is None
        or result.residual_vector is None
    ):
        return pd.DataFrame()

    return pd.DataFrame(
        {
            "Equation": equation_names(len(result.solution)),
            "Computed Ax": result.computed_right_hand_side,
            "Original b": result.original_vector,
            "Residual Ax − b": result.residual_vector,
            "Absolute Residual": np.abs(result.residual_vector),
        }
    )


def build_modified_coefficients_dataframe(result: TridiagonalResult) -> pd.DataFrame:
    """Create the final modified-coefficient table."""

    size = len(result.main_diagonal)
    return pd.DataFrame(
        {
            "Row": np.arange(1, size + 1),
            "Original Main bᵢ": result.main_diagonal,
            "Modified Upper c′ᵢ": result.modified_upper,
            "Modified RHS d′ᵢ": result.modified_rhs,
        }
    )


# =============================================================================
# Excel export
# =============================================================================
def style_excel_worksheet(worksheet: Any) -> None:
    """Apply consistent professional formatting to one worksheet."""

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="17324D")

    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.000000000000E+00"

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for cell in worksheet[column_letter]:
            value_length = len(str(cell.value)) if cell.value is not None else 0
            max_length = max(max_length, value_length)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)


def create_excel_report(result: TridiagonalResult) -> bytes:
    """Create a complete XLSX report in memory."""

    if not result.success or result.solution is None:
        raise ValueError("Only a successful calculation can be exported.")

    diagnostics = result.diagnostics
    summary_df = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "Matrix Size",
                "System Classification",
                "Rank of A",
                "Rank of [A | b]",
                "Condition Number",
                "Pivot Tolerance",
                "Structure Tolerance",
                "Forward Sweep Steps",
                "Back Substitution Steps",
                "Solution Vector",
                "Residual 2-Norm",
                "Residual Infinity Norm",
                "Maximum Absolute Residual",
                "Relative Residual",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                f"{result.original_matrix.shape[0]} × {result.original_matrix.shape[1]}",
                diagnostics.classification if diagnostics else "Not available",
                diagnostics.rank_a if diagnostics else "Not available",
                diagnostics.rank_augmented if diagnostics else "Not available",
                diagnostics.condition_number if diagnostics else "Not available",
                result.pivot_tolerance,
                result.structure_tolerance,
                result.forward_steps,
                result.back_substitution_steps,
                serialize_array(result.solution),
                result.residual_norm_2,
                result.residual_norm_inf,
                result.maximum_absolute_residual,
                result.relative_residual,
                " | ".join(result.warnings) if result.warnings else "None",
                result.stopping_reason,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )

    original_matrix_df = matrix_dataframe(result.original_matrix).reset_index(
        names="Equation"
    )
    original_vector_df = vector_dataframe(result.original_vector).reset_index(
        names="Equation"
    )
    original_augmented_df = augmented_dataframe(
        result.original_matrix, result.original_vector
    ).reset_index(names="Equation")
    diagonals_df = diagonal_dataframe(result)
    forward_df = build_forward_export_dataframe(result)
    modified_df = build_modified_coefficients_dataframe(result)
    back_df = build_back_dataframe(result)
    solution_df = build_solution_dataframe(result)
    residual_df = build_residual_dataframe(result)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        original_matrix_df.to_excel(writer, sheet_name="Original Matrix", index=False)
        original_vector_df.to_excel(writer, sheet_name="Original Vector", index=False)
        original_augmented_df.to_excel(writer, sheet_name="Original Augmented", index=False)
        diagonals_df.to_excel(writer, sheet_name="Tridiagonal Coeffs", index=False)
        forward_df.to_excel(writer, sheet_name="Forward Sweep", index=False)
        modified_df.to_excel(writer, sheet_name="Modified Coeffs", index=False)
        back_df.to_excel(writer, sheet_name="Back Substitution", index=False)
        solution_df.to_excel(writer, sheet_name="Solution", index=False)
        residual_df.to_excel(writer, sheet_name="Residual Analysis", index=False)

        workbook = writer.book
        chart_data_df = pd.DataFrame(
            {
                "Variable": variable_names(len(result.solution)),
                "Solution": result.solution,
                "Equation": equation_names(len(result.solution)),
                "Absolute Residual": np.abs(result.residual_vector),
            }
        )
        chart_data_df.to_excel(writer, sheet_name="Chart Data", index=False)
        data_sheet = workbook["Chart Data"]
        data_sheet["F1"] = "Graphs based on the report data"
        data_sheet["F1"].font = Font(bold=True, size=14)
        data_rows = len(result.solution) + 1

        solution_chart = BarChart()
        solution_chart.title = "Thomas Algorithm Solution"
        solution_chart.x_axis.title = "Variable"
        solution_chart.y_axis.title = "Value"
        solution_chart.height = 8
        solution_chart.width = 15
        solution_chart.add_data(
            Reference(data_sheet, min_col=2, min_row=1, max_row=data_rows),
            titles_from_data=True,
        )
        solution_chart.set_categories(
            Reference(data_sheet, min_col=1, min_row=2, max_row=data_rows)
        )
        data_sheet.add_chart(solution_chart, "F3")

        residual_chart = BarChart()
        residual_chart.title = "Absolute Residual by Equation"
        residual_chart.x_axis.title = "Equation"
        residual_chart.y_axis.title = "Absolute residual"
        residual_chart.height = 8
        residual_chart.width = 15
        residual_chart.add_data(
            Reference(data_sheet, min_col=4, min_row=1, max_row=data_rows),
            titles_from_data=True,
        )
        residual_chart.set_categories(
            Reference(data_sheet, min_col=3, min_row=2, max_row=data_rows)
        )
        data_sheet.add_chart(residual_chart, "F20")

        for worksheet in workbook.worksheets:
            style_excel_worksheet(worksheet)

    output.seek(0)
    return finalize_excel_report_with_visible_charts(output.getvalue())


# =============================================================================
# Streamlit input state and widgets
# =============================================================================
def initialize_input_state(size: int) -> None:
    """Initialize matrix and vector widget values for the selected size."""

    default_matrix, default_vector = DEFAULT_SYSTEMS[size]
    state_size_key = "tridiagonal_initialized_size"

    if st.session_state.get(state_size_key) == size:
        return

    for row_index in range(size):
        for column_index in range(size):
            st.session_state[
                f"tridiagonal_a_{row_index}_{column_index}"
            ] = format_number(default_matrix[row_index, column_index])
        st.session_state[f"tridiagonal_b_{row_index}"] = format_number(
            default_vector[row_index]
        )

    st.session_state[state_size_key] = size
    st.session_state.pop("tridiagonal_result", None)
    st.session_state.pop("tridiagonal_excel_report", None)
    st.session_state.pop("tridiagonal_excel_signature", None)


def collect_raw_inputs(size: int) -> tuple[list[list[str]], list[str]]:
    """Read the current matrix and vector values from session state."""

    raw_matrix = [
        [
            str(st.session_state.get(f"tridiagonal_a_{row}_{column}", ""))
            for column in range(size)
        ]
        for row in range(size)
    ]
    raw_vector = [
        str(st.session_state.get(f"tridiagonal_b_{row}", ""))
        for row in range(size)
    ]
    return raw_matrix, raw_vector



def finalize_excel_report_with_visible_charts(report_bytes: bytes) -> bytes:
    """Place existing workbook charts on Summary so they are immediately visible.

    The report data and worksheets are preserved. If Excel chart post-processing is
    unavailable for any reason, the original report is returned unchanged rather
    than risking a damaged workbook.
    """
    try:
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_workbook

        workbook = _load_workbook(_BytesIO(report_bytes))
        if "Summary" not in workbook.sheetnames:
            return report_bytes

        summary_sheet = workbook["Summary"]
        collected_charts = list(summary_sheet._charts)
        summary_sheet._charts = []

        for worksheet in workbook.worksheets:
            if worksheet is summary_sheet:
                continue
            if worksheet._charts:
                collected_charts.extend(list(worksheet._charts))
                worksheet._charts = []

        # Put every chart in the existing Summary worksheet, below one another.
        # D2 keeps the first chart visible beside the main results when the report opens.
        for chart_index, chart in enumerate(collected_charts):
            anchor_row = 2 + chart_index * 19
            summary_sheet.add_chart(chart, f"D{anchor_row}")

        workbook.active = workbook.sheetnames.index("Summary")
        output = _BytesIO()
        workbook.save(output)
        return output.getvalue()
    except Exception:
        return report_bytes


def render_matrix_input_grid(size: int) -> None:
    """Render a dynamic matrix A and vector b input grid."""

    initialize_input_state(size)
    header_columns = st.columns([*[1.0] * size, 0.22, 1.0])
    for column_index in range(size):
        header_columns[column_index].markdown(
            f"<div style='text-align:center;font-weight:700;'>x{column_index + 1}</div>",
            unsafe_allow_html=True,
        )
    header_columns[size].markdown(
        "<div style='text-align:center;font-weight:700;'>|</div>",
        unsafe_allow_html=True,
    )
    header_columns[size + 1].markdown(
        "<div style='text-align:center;font-weight:700;'>b</div>",
        unsafe_allow_html=True,
    )

    for row_index in range(size):
        row_columns = st.columns([*[1.0] * size, 0.22, 1.0])
        for column_index in range(size):
            row_columns[column_index].text_input(
                f"A[{row_index + 1},{column_index + 1}]",
                key=f"tridiagonal_a_{row_index}_{column_index}",
                label_visibility="collapsed",
            )
        row_columns[size].markdown(
            "<div style='text-align:center;padding-top:0.55rem;font-weight:700;'>|</div>",
            unsafe_allow_html=True,
        )
        row_columns[size + 1].text_input(
            f"b[{row_index + 1}]",
            key=f"tridiagonal_b_{row_index}",
            label_visibility="collapsed",
        )


# =============================================================================
# Streamlit result rendering
# =============================================================================
def render_final_result(result: TridiagonalResult) -> None:
    """Render the compact final-result card."""

    if not result.success:
        st.error(result.message)
        st.caption(result.stopping_reason)
        if result.diagnostics is not None:
            st.write(f"**System classification:** {result.diagnostics.classification}")
        return

    st.success(result.message)
    if result.solution is None:
        return

    solution_text = "  ·  ".join(
        f"x{index + 1} = {format_number(value)}"
        for index, value in enumerate(result.solution)
    )
    st.markdown(f"### {solution_text}")

    first_row = st.columns(3)
    first_row[0].metric("Forward Steps", result.forward_steps)
    first_row[1].metric("Back Steps", result.back_substitution_steps)
    first_row[2].metric(
        "Residual Norm",
        format_scientific_power(result.residual_norm_2),
    )

    second_row = st.columns(2)
    second_row[0].metric(
        "Maximum |Residual|",
        (
            format_scientific_power(result.maximum_absolute_residual)
        ),
    )
    second_row[1].metric("Status", result.status)

    st.caption(result.stopping_reason)
    for warning in result.warnings:
        st.warning(warning)


def render_original_system(result: TridiagonalResult) -> None:
    """Render the original system and extracted diagonal vectors."""

    st.subheader("Original Tridiagonal System")
    matrix_column, vector_column = st.columns([2.2, 1.0])
    with matrix_column:
        st.markdown("**Coefficient Matrix A**")
        st.dataframe(
            round_numeric_dataframe(matrix_dataframe(result.original_matrix)),
            use_container_width=True,
        )
    with vector_column:
        st.markdown("**Right-Hand Side b**")
        st.dataframe(
            round_numeric_dataframe(vector_dataframe(result.original_vector)),
            use_container_width=True,
        )

    st.markdown("**Original Augmented Matrix [A | b]**")
    st.dataframe(
        round_numeric_dataframe(
            augmented_dataframe(result.original_matrix, result.original_vector)
        ),
        use_container_width=True,
    )

    if result.main_diagonal.size > 0:
        st.markdown("**Extracted Tridiagonal Coefficients**")
        st.dataframe(
            round_numeric_dataframe(diagonal_dataframe(result)),
            use_container_width=True,
            hide_index=True,
        )


def render_forward_sweep(result: TridiagonalResult) -> None:
    """Render all Thomas Algorithm forward-sweep calculations."""

    st.subheader("Forward Sweep")
    st.caption(
        "The method calculates modified upper-diagonal coefficients c′ and "
        "modified right-hand-side coefficients d′ without forming a full "
        "upper-triangular matrix."
    )

    forward_df = build_forward_dataframe(result)
    if not forward_df.empty:
        st.dataframe(
            round_numeric_dataframe(forward_df),
            use_container_width=True,
            hide_index=True,
        )

    for record in result.forward_history:
        with st.expander(
            f"Forward Step {record.step} — Row {record.row}",
            expanded=False,
        ):
            metric_columns = st.columns(3)
            metric_columns[0].metric("Modified Pivot", format_number(record.denominator))
            metric_columns[1].metric("Modified Upper", format_number(record.modified_upper))
            metric_columns[2].metric("Modified RHS", format_number(record.modified_rhs))

            st.markdown(f"**Pivot formula:** `{record.denominator_formula}`")
            st.markdown(f"**Upper formula:** `{record.modified_upper_formula}`")
            st.markdown(f"**RHS formula:** `{record.modified_rhs_formula}`")

            details_df = pd.DataFrame(
                {
                    "Quantity": [
                        "aᵢ",
                        "bᵢ",
                        "cᵢ",
                        "dᵢ",
                        "Previous c′",
                        "Previous d′",
                        "Modified pivot",
                        "New c′",
                        "New d′",
                    ],
                    "Value": [
                        record.lower_coefficient,
                        record.main_coefficient,
                        record.upper_coefficient,
                        record.right_hand_side,
                        record.previous_modified_upper,
                        record.previous_modified_rhs,
                        record.denominator,
                        record.modified_upper,
                        record.modified_rhs,
                    ],
                }
            )
            st.dataframe(
                round_numeric_dataframe(details_df),
                use_container_width=True,
                hide_index=True,
            )
            st.caption(record.notes)


def render_modified_coefficients(result: TridiagonalResult) -> None:
    """Render the completed modified coefficient vectors."""

    st.subheader("Modified Coefficients")
    st.dataframe(
        round_numeric_dataframe(build_modified_coefficients_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )


def render_back_substitution(result: TridiagonalResult) -> None:
    """Render the complete back-substitution table."""

    st.subheader("Back Substitution")
    back_df = build_back_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(back_df),
        use_container_width=True,
        hide_index=True,
    )

    for record in result.back_history:
        with st.expander(
            f"Back Step {record.step} — Calculate {record.variable}",
            expanded=False,
        ):
            st.markdown(f"**Formula:** `{record.formula}`")
            st.metric(record.variable, format_number(record.calculated_value))
            if record.known_next_value is not None:
                st.caption(
                    f"Known next value = {format_number(record.known_next_value)}"
                )


def render_solution_and_residual(result: TridiagonalResult) -> None:
    """Render solution and residual analysis."""

    st.subheader("Final Solution")
    st.dataframe(
        round_numeric_dataframe(build_solution_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Residual Analysis")
    st.latex(r"r = Ax - b")
    st.dataframe(
        round_numeric_dataframe(build_residual_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    metrics = st.columns(4)
    metrics[0].metric(
        "Residual 2-Norm",
        format_scientific_power(result.residual_norm_2),
    )
    metrics[1].metric(
        "Residual ∞-Norm",
        format_scientific_power(result.residual_norm_inf),
    )
    metrics[2].metric(
        "Maximum |Residual|",
        (
            format_scientific_power(result.maximum_absolute_residual)
        ),
    )
    metrics[3].metric(
        "Relative Residual",
        format_scientific_power(result.relative_residual),
    )

    if result.diagnostics is not None:
        st.caption(
            f"System classification: {result.diagnostics.classification} "
            f"Condition number: {format_scientific_power(result.diagnostics.condition_number)}."
        )


def render_excel_download(result: TridiagonalResult) -> None:
    """Generate and render the Excel download button."""

    st.subheader("Excel Report")
    signature = result.input_signature

    if st.session_state.get("tridiagonal_excel_signature") != signature:
        try:
            st.session_state.tridiagonal_excel_report = create_excel_report(result)
            st.session_state.tridiagonal_excel_signature = signature
        except (ValueError, OSError, RuntimeError) as error:
            st.error(f"The Excel report could not be generated. Details: {error}")
            return

    report_bytes = st.session_state.get("tridiagonal_excel_report")
    if report_bytes is None:
        st.error("The Excel report is unavailable.")
        return

    date_text = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=f"tridiagonal_system_thomas_report_{date_text}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="tridiagonal_download_excel",
    )


# =============================================================================
# Streamlit page layout
# =============================================================================
def render_page() -> None:
    """Render the existing website layout and connect the solver logic."""

    st.set_page_config(
        page_title="Tridiagonal Systems Solver | Numerical Methods",
        page_icon="📘",
        layout="wide",
    )
    load_css()

    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">LINEAR SYSTEM TOOL</div>
                <h1>Tridiagonal Systems Solver</h1>
                <p>
                    Enter a tridiagonal linear system and follow every modified
                    coefficient in the Thomas Algorithm forward sweep and every
                    variable calculation in back substitution.
                </p>

                <div class="method-actions">
                    <a href="/Tridiagonal_Systems" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Tridiagonal_Systems_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Match the centered Bisection solver content width.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )
    
        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader('How to Enter the System')
                st.markdown(
                    """
                Enter the coefficients of the tridiagonal matrix and the right-hand side vector.

                - The main diagonal contains **bᵢ**.
                - The lower diagonal contains **aᵢ** below the main diagonal.
                - The upper diagonal contains **cᵢ** above the main diagonal.
                - The right-hand side contains **dᵢ**.
                - Entries outside these three diagonals must be zero.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader('Before Solving')
                st.markdown(
                    """
                - The coefficient matrix must be square and genuinely tridiagonal.
                - The Thomas Algorithm requires nonzero effective pivots during the forward sweep.
                - A zero or very small pivot may indicate that pivoting or another linear solver is required.
                - Inspect the residual **Ax − d** after solving to verify the result.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])
    
        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    '<div class="input-label-ui">System size</div>',
                    unsafe_allow_html=True,
                )
                system_size = st.selectbox(
                    "System size",
                    options=SUPPORTED_SIZES,
                    index=SUPPORTED_SIZES.index(DEFAULT_SIZE),
                    format_func=lambda value: f"{value} × {value}",
                    label_visibility="collapsed",
                    key="tridiagonal_system_size",
                )
    
                st.markdown(
                    '<div class="input-label-ui">Matrix A and vector b</div>',
                    unsafe_allow_html=True,
                )
                render_matrix_input_grid(system_size)
    
                st.caption(
                    "All coefficients outside the main, lower, and upper diagonals "
                    "must be zero. Decimal, negative, and scientific-notation values "
                    "are supported."
                )
                st.info(
                    "The standard Thomas Algorithm does not perform row pivoting. "
                    "If a modified pivot becomes zero or near zero, the solver stops "
                    "safely and recommends another method."
                )
    
                solve_button_clicked = st.button(
                    "Solve",
                    use_container_width=True,
                    key="tridiagonal_solve_button",
                )
    
        raw_matrix_values, raw_vector_values = collect_raw_inputs(system_size)
        current_input_signature = create_input_signature(
            raw_matrix_values,
            raw_vector_values,
            PIVOT_TOLERANCE,
            STRUCTURE_TOLERANCE,
        )
    
        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )
    
                stored_result = st.session_state.get("tridiagonal_result")
                if stored_result is None:
                    st.info("Enter the system values and click Solve to see the result.")
                elif stored_result.input_signature != current_input_signature:
                    st.info(
                        "The matrix or vector has changed. Click Solve to calculate "
                        "a new result."
                    )
                else:
                    render_final_result(stored_result)
    
        if solve_button_clicked:
            st.session_state.tridiagonal_result = solve_tridiagonal_system(
                raw_matrix=raw_matrix_values,
                raw_vector=raw_vector_values,
                input_signature=current_input_signature,
                pivot_tolerance=PIVOT_TOLERANCE,
                structure_tolerance=STRUCTURE_TOLERANCE,
            )
            st.session_state.pop("tridiagonal_excel_report", None)
            st.session_state.pop("tridiagonal_excel_signature", None)
            st.rerun()
    
        active_result = st.session_state.get("tridiagonal_result")
        if (
            active_result is not None
            and active_result.input_signature == current_input_signature
        ):
            if active_result.original_matrix.size > 0:
                st.divider()
                render_original_system(active_result)
    
            if active_result.forward_history:
                st.divider()
                render_forward_sweep(active_result)
    
            if active_result.modified_upper.size > 0:
                st.divider()
                render_modified_coefficients(active_result)
    
            if active_result.back_history:
                st.divider()
                render_back_substitution(active_result)
    
            if active_result.success:
                st.divider()
                render_solution_and_residual(active_result)
    
                st.divider()
                render_excel_download(active_result)
    
                st.divider()
                navigation_left_column, navigation_right_column = st.columns(2)
    
                with navigation_left_column:
                    if st.button(
                        "Review Tridiagonal Systems Lesson",
                        use_container_width=True,
                        key="review_tridiagonal_lesson",
                    ):
                        st.switch_page("pages/Tridiagonal_Systems.py")
    
                with navigation_right_column:
                    if st.button(
                        "Back to Solver Menu",
                        use_container_width=True,
                        key="back_to_solver_menu_tridiagonal",
                    ):
                        st.switch_page("pages/Numerical_Solver.py")
    
        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Tridiagonal Systems • Linear Systems</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()
