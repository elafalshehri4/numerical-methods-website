from __future__ import annotations

import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any, Callable, Sequence
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css
from utilities.safe_math import safe_sympify


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Richardson Extrapolation"
DISPLAY_DECIMALS = 3
DEFAULT_FUNCTION = "sin(x)"
DEFAULT_X_VALUE = 1.0
DEFAULT_STEP_SIZE = 0.2
DEFAULT_LEVELS = 5
DEFAULT_DERIVATIVE_ORDER = 1

MIN_LEVELS = 2
MAX_LEVELS = 9
SUPPORTED_DERIVATIVE_ORDERS = (1, 2, 3)

ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
REFINEMENT_RATIO = 2.0
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

BASE_METHOD_OPTIONS = {
    "Central difference — base error O(h²)": "central",
    "Forward difference — base error O(h)": "forward",
}

DERIVATIVE_LABELS = {
    1: "f′(x₀)",
    2: "f″(x₀)",
    3: "f‴(x₀)",
}

ALLOWED_FUNCTION_NAMES = {
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}


# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable

    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0**exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Use fixed or scientific notation according to numerical magnitude."""

    if value is None:
        return unavailable

    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6
    ):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


def format_number(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
) -> str:
    """Format one numerical value for the website."""

    return format_display_number(value, decimals, unavailable="Not available")


def scientific_number(value: float | int | None) -> str:
    """Format one numerical value in scientific notation."""

    return format_scientific_power(value)


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only in a display copy."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class FormulaSpecification:
    """One finite-difference base formula used by Richardson extrapolation."""

    method_key: str
    derivative_order: int
    offsets: tuple[int, ...]
    coefficients: tuple[float, ...]
    denominator_factor: float
    base_error_order: int
    formula_text: str


@dataclass(frozen=True)
class BaseApproximation:
    """One base finite-difference estimate at one step size."""

    level: int
    step_size: float
    derivative_order: int
    offsets: tuple[int, ...]
    sample_x: tuple[float, ...]
    sample_values: tuple[float, ...]
    coefficients: tuple[float, ...]
    weighted_numerator: float
    denominator: float
    approximation: float
    exact_derivative: float | None
    absolute_error: float | None
    relative_error_percent: float | None
    successive_difference: float | None
    observed_order: float | None
    substitution_text: str


@dataclass(frozen=True)
class ExtrapolationStep:
    """One computed cell in the triangular Richardson table."""

    row: int
    column: int
    step_size: float
    lower_order_estimate: float
    previous_row_estimate: float
    refinement_ratio: float
    base_order: int
    exponent: int
    denominator_factor: float
    correction: float
    extrapolated_estimate: float
    expected_order: int
    exact_derivative: float | None
    absolute_error: float | None
    relative_error_percent: float | None
    formula_text: str
    substitution_text: str


@dataclass(frozen=True)
class RichardsonResult:
    """Complete Richardson result shared by Streamlit and Excel renderers."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str

    function_text: str
    function_expression: sp.Expr | None
    derivative_order: int
    derivative_label: str
    derivative_expression: sp.Expr | None
    derivative_expression_text: str

    base_method_key: str
    base_method_name: str
    base_formula_text: str
    base_order: int

    x_value: float | None
    initial_step_size: float | None
    refinement_ratio: float
    levels: int

    base_approximations: tuple[BaseApproximation, ...]
    extrapolation_steps: tuple[ExtrapolationStep, ...]
    richardson_table: tuple[tuple[float | None, ...], ...]

    exact_derivative: float | None
    primary_base_estimate: float | None
    final_estimate: float | None
    final_expected_order: int | None
    primary_absolute_error: float | None
    final_absolute_error: float | None
    final_relative_error_percent: float | None
    error_improvement_factor: float | None
    latest_observed_order: float | None
    best_diagonal_estimate: float | None
    best_diagonal_error: float | None

    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers and validation
# =============================================================================
def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def create_input_signature(
    function_text: str,
    x_value: Any,
    step_size: Any,
    derivative_order: Any,
    base_method_name: str,
    levels: Any,
) -> str:
    """Create a stable signature to prevent stale Streamlit output."""

    payload = repr(
        (
            str(function_text).strip(),
            repr(x_value),
            repr(step_size),
            repr(derivative_order),
            str(base_method_name),
            repr(levels),
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def safe_float(raw_value: Any, value_name: str) -> float:
    """Convert one input to a finite real float."""

    if raw_value is None:
        raise ValueError(f"{value_name} is required.")
    if isinstance(raw_value, str) and not raw_value.strip():
        raise ValueError(f"{value_name} is required.")

    try:
        complex_value = complex(raw_value)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(f"{value_name} must be a valid real number.") from error

    if abs(complex_value.imag) > ZERO_TOLERANCE:
        raise ValueError(f"{value_name} must be real, not complex.")

    value = float(complex_value.real)
    if not math.isfinite(value):
        raise ValueError(f"{value_name} must be finite.")
    return value


def safe_integer(
    raw_value: Any,
    value_name: str,
    minimum: int,
    maximum: int,
) -> int:
    """Validate an integer without silently truncating decimal values."""

    try:
        numeric_value = float(raw_value)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(f"{value_name} must be an integer.") from error

    if not math.isfinite(numeric_value) or not numeric_value.is_integer():
        raise ValueError(f"{value_name} must be an integer.")

    value = int(numeric_value)
    if value < minimum or value > maximum:
        raise ValueError(
            f"{value_name} must be between {minimum} and {maximum}."
        )
    return value


def safe_relative_error_percent(
    absolute_error: float,
    exact_value: float,
) -> float | None:
    """Return percentage relative error when the exact derivative is nonzero."""

    if abs(exact_value) <= RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
        return None
    return absolute_error / abs(exact_value) * 100.0


def human_readable_expression(expression: sp.Expr | None) -> str:
    """Return a compact expression for reports."""

    if expression is None:
        return "Not available"
    return str(sp.simplify(expression))


# =============================================================================
# Safe function parsing and evaluation
# =============================================================================
def parse_function(function_text: str) -> tuple[sp.Expr, sp.Symbol]:
    """Parse a safe real single-variable expression."""

    if not isinstance(function_text, str) or not function_text.strip():
        raise ValueError("Function f(x) is required.")

    text = function_text.strip().replace("^", "**")
    if "=" in text:
        raise ValueError("Enter only f(x), without an equals sign.")

    x_symbol = sp.Symbol("x", real=True)
    local_dictionary = {"x": x_symbol, **ALLOWED_FUNCTION_NAMES}

    try:
        expression = safe_sympify(
            text,
            locals=local_dictionary,
            evaluate=True,
        )
    except (sp.SympifyError, SyntaxError, TypeError, ValueError) as error:
        raise ValueError(
            "The function format is invalid. Use expressions such as sin(x), "
            "exp(x), or x**3 - 2*x + 1."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError("The function could not be interpreted as a scalar expression.")
    if isinstance(expression, Relational) or expression.has(Relational):
        raise ValueError("Enter a function expression, not an equation or inequality.")
    if expression.is_Matrix:
        raise ValueError("A scalar function is required, not a matrix.")
    if expression.has(
        AppliedUndef,
        sp.Derivative,
        sp.Integral,
        sp.Sum,
        sp.Product,
        sp.Limit,
    ):
        raise ValueError("The function contains unsupported symbolic operations.")
    if expression.has(sp.I, sp.zoo, sp.oo, -sp.oo, sp.nan):
        raise ValueError("The function contains a complex or non-finite constant.")

    unknown_symbols = expression.free_symbols.difference({x_symbol})
    if unknown_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unknown_symbols))
        raise ValueError(f"Only x is supported as a variable. Unsupported: {names}.")

    return sp.simplify(expression), x_symbol


def create_numeric_function(
    expression: sp.Expr,
    x_symbol: sp.Symbol,
) -> Callable[[Any], Any]:
    """Convert one SymPy expression to a NumPy-compatible function."""

    try:
        return sp.lambdify(x_symbol, expression, modules=["numpy"])
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The function could not be converted to numerical form."
        ) from error


def evaluate_real_scalar(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    value_name: str,
) -> float:
    """Evaluate one finite real scalar safely."""

    try:
        with np.errstate(all="raise"):
            raw_value = numeric_function(float(x_value))
        array = np.asarray(raw_value)
    except (
        ArithmeticError,
        TypeError,
        ValueError,
        OverflowError,
        FloatingPointError,
        ZeroDivisionError,
    ) as error:
        raise ValueError(
            f"{value_name} is undefined at x = {x_value:.12g}. Reason: {error}"
        ) from error

    if array.size != 1:
        raise ValueError(f"{value_name} did not return a scalar value.")

    scalar = array.reshape(-1)[0]
    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)
        if abs(complex_value.imag) > 1.0e-12:
            raise ValueError(
                f"{value_name} is complex at x = {x_value:.12g}."
            )
        scalar = complex_value.real

    return safe_float(scalar, value_name)


def evaluate_real_array(
    numeric_function: Callable[[Any], Any],
    x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate a function for plotting and replace invalid values by NaN."""

    try:
        with np.errstate(all="ignore"):
            raw_values = numeric_function(x_values)
        values = np.asarray(raw_values)
    except Exception:
        return np.full_like(x_values, np.nan, dtype=float)

    if values.ndim == 0:
        values = np.full_like(
            x_values,
            values,
            dtype=complex if np.iscomplexobj(values) else float,
        )
    else:
        try:
            values = np.broadcast_to(values, x_values.shape)
        except ValueError:
            return np.full_like(x_values, np.nan, dtype=float)

    if np.iscomplexobj(values):
        imaginary = np.abs(np.imag(values))
        result = np.real(values).astype(float)
        result[imaginary > 1.0e-12] = np.nan
    else:
        try:
            result = values.astype(float)
        except (TypeError, ValueError):
            return np.full_like(x_values, np.nan, dtype=float)

    result[~np.isfinite(result)] = np.nan
    return result


# =============================================================================
# Finite-difference specifications for derivatives 1 through 3
# =============================================================================
def formula_specification(
    method_key: str,
    derivative_order: int,
) -> FormulaSpecification:
    """Return a textbook finite-difference formula and its error order."""

    if derivative_order not in SUPPORTED_DERIVATIVE_ORDERS:
        raise ValueError("Derivative order must be 1, 2, or 3.")

    specifications: dict[tuple[str, int], FormulaSpecification] = {
        (
            "central",
            1,
        ): FormulaSpecification(
            method_key="central",
            derivative_order=1,
            offsets=(-1, 1),
            coefficients=(-1.0, 1.0),
            denominator_factor=2.0,
            base_error_order=2,
            formula_text="[f(x₀+h) − f(x₀−h)] / (2h)",
        ),
        (
            "central",
            2,
        ): FormulaSpecification(
            method_key="central",
            derivative_order=2,
            offsets=(-1, 0, 1),
            coefficients=(1.0, -2.0, 1.0),
            denominator_factor=1.0,
            base_error_order=2,
            formula_text="[f(x₀+h) − 2f(x₀) + f(x₀−h)] / h²",
        ),
        (
            "central",
            3,
        ): FormulaSpecification(
            method_key="central",
            derivative_order=3,
            offsets=(-2, -1, 1, 2),
            coefficients=(-1.0, 2.0, -2.0, 1.0),
            denominator_factor=2.0,
            base_error_order=2,
            formula_text=(
                "[f(x₀+2h) − 2f(x₀+h) + 2f(x₀−h) − f(x₀−2h)] / (2h³)"
            ),
        ),
        (
            "forward",
            1,
        ): FormulaSpecification(
            method_key="forward",
            derivative_order=1,
            offsets=(0, 1),
            coefficients=(-1.0, 1.0),
            denominator_factor=1.0,
            base_error_order=1,
            formula_text="[f(x₀+h) − f(x₀)] / h",
        ),
        (
            "forward",
            2,
        ): FormulaSpecification(
            method_key="forward",
            derivative_order=2,
            offsets=(0, 1, 2),
            coefficients=(1.0, -2.0, 1.0),
            denominator_factor=1.0,
            base_error_order=1,
            formula_text="[f(x₀+2h) − 2f(x₀+h) + f(x₀)] / h²",
        ),
        (
            "forward",
            3,
        ): FormulaSpecification(
            method_key="forward",
            derivative_order=3,
            offsets=(0, 1, 2, 3),
            coefficients=(-1.0, 3.0, -3.0, 1.0),
            denominator_factor=1.0,
            base_error_order=1,
            formula_text=(
                "[f(x₀+3h) − 3f(x₀+2h) + 3f(x₀+h) − f(x₀)] / h³"
            ),
        ),
    }

    specification = specifications.get((method_key, derivative_order))
    if specification is None:
        raise ValueError("Unsupported finite-difference method or derivative order.")
    return specification


def offset_label(offset: int) -> str:
    """Return a readable x-coordinate label for one integer offset."""

    if offset == 0:
        return "x₀"
    if offset == 1:
        return "x₀+h"
    if offset == -1:
        return "x₀−h"
    if offset > 0:
        return f"x₀+{offset}h"
    return f"x₀−{abs(offset)}h"


def calculate_base_approximation(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    step_size: float,
    specification: FormulaSpecification,
) -> dict[str, Any]:
    """Calculate one finite-difference base estimate manually."""

    sample_x: list[float] = []
    sample_values: list[float] = []

    for offset in specification.offsets:
        sample_point = x_value + offset * step_size
        if not math.isfinite(sample_point):
            raise ValueError(
                f"{offset_label(offset)} is non-finite. Reduce |x₀| or h."
            )
        if offset != 0 and sample_point == x_value:
            raise ValueError(
                "The refined step is too small to create distinct floating-point "
                "sample coordinates. Use a larger initial h or fewer levels."
            )

        sample_value = evaluate_real_scalar(
            numeric_function,
            sample_point,
            f"f({offset_label(offset)})",
        )
        sample_x.append(float(sample_point))
        sample_values.append(float(sample_value))

    weighted_terms = [
        coefficient * sample_value
        for coefficient, sample_value in zip(
            specification.coefficients,
            sample_values,
        )
    ]
    numerator = math.fsum(weighted_terms)
    denominator = (
        specification.denominator_factor
        * step_size ** specification.derivative_order
    )

    if denominator == 0.0 or not math.isfinite(denominator):
        raise ValueError("The finite-difference denominator is numerically unusable.")

    approximation = numerator / denominator
    if not math.isfinite(approximation):
        raise ValueError("The finite-difference estimate is NaN or infinity.")

    numerator_parts = [
        f"({coefficient:.12g})({sample_value:.12g})"
        for coefficient, sample_value in zip(
            specification.coefficients,
            sample_values,
        )
    ]
    substitution_text = (
        "[" + " + ".join(numerator_parts) + "] / "
        f"[{specification.denominator_factor:.12g} × "
        f"({step_size:.12g})^{specification.derivative_order}] "
        f"= {approximation:.15g}"
    )

    return {
        "sample_x": tuple(sample_x),
        "sample_values": tuple(sample_values),
        "numerator": float(numerator),
        "denominator": float(denominator),
        "approximation": float(approximation),
        "substitution_text": substitution_text,
    }


# =============================================================================
# Richardson algorithm and convergence analysis
# =============================================================================
def calculate_observed_order(
    previous_error: float | None,
    current_error: float | None,
    refinement_ratio: float = REFINEMENT_RATIO,
) -> float | None:
    """Estimate order from consecutive positive errors at the same formal order."""

    if previous_error is None or current_error is None:
        return None
    if previous_error <= ZERO_TOLERANCE or current_error <= ZERO_TOLERANCE:
        return None

    try:
        value = math.log(previous_error / current_error) / math.log(
            refinement_ratio
        )
    except (ValueError, ZeroDivisionError):
        return None
    return value if math.isfinite(value) else None


def calculate_successive_order(
    older_difference: float | None,
    newer_difference: float | None,
    refinement_ratio: float = REFINEMENT_RATIO,
) -> float | None:
    """Estimate order from two successive differences within one table column."""

    if older_difference is None or newer_difference is None:
        return None
    if older_difference <= ZERO_TOLERANCE or newer_difference <= ZERO_TOLERANCE:
        return None

    try:
        value = math.log(older_difference / newer_difference) / math.log(
            refinement_ratio
        )
    except (ValueError, ZeroDivisionError):
        return None
    return value if math.isfinite(value) else None


def build_richardson_table(
    base_values: Sequence[float],
    step_sizes: Sequence[float],
    base_order: int,
    exact_derivative: float | None,
    refinement_ratio: float = REFINEMENT_RATIO,
) -> tuple[list[list[float | None]], list[ExtrapolationStep]]:
    """Build the complete triangular Richardson table manually.

    For forward formulas, the base expansion begins with h¹, h², h³, ... .
    For centered formulas, it begins with h², h⁴, h⁶, ... . Therefore the
    denominator in column j is r^(p*j)-1, where p is 1 or 2.
    """

    levels = len(base_values)
    table: list[list[float | None]] = [
        [None for _ in range(levels)] for _ in range(levels)
    ]
    history: list[ExtrapolationStep] = []

    for row, value in enumerate(base_values):
        table[row][0] = float(value)

    for row in range(1, levels):
        for column in range(1, row + 1):
            current_lower = table[row][column - 1]
            previous_lower = table[row - 1][column - 1]
            if current_lower is None or previous_lower is None:
                raise ValueError("The Richardson table is missing a prerequisite value.")

            exponent = base_order * column
            denominator_factor = refinement_ratio**exponent - 1.0
            if abs(denominator_factor) <= ZERO_TOLERANCE:
                raise ValueError("A Richardson denominator became zero.")

            correction = (current_lower - previous_lower) / denominator_factor
            extrapolated = current_lower + correction

            if not all(
                math.isfinite(value)
                for value in (denominator_factor, correction, extrapolated)
            ):
                raise ValueError("Richardson extrapolation produced a non-finite value.")

            table[row][column] = float(extrapolated)
            absolute_error = (
                None
                if exact_derivative is None
                else abs(extrapolated - exact_derivative)
            )
            relative_percent = (
                None
                if absolute_error is None or exact_derivative is None
                else safe_relative_error_percent(absolute_error, exact_derivative)
            )
            expected_order = base_order * (column + 1)

            formula_text = (
                f"R[{row},{column}] = R[{row},{column - 1}] + "
                f"(R[{row},{column - 1}] − R[{row - 1},{column - 1}]) / "
                f"({refinement_ratio:g}^{exponent} − 1)"
            )
            substitution_text = (
                f"{current_lower:.15g} + ({current_lower:.15g} − "
                f"{previous_lower:.15g}) / ({refinement_ratio:g}^{exponent} − 1) "
                f"= {extrapolated:.15g}"
            )

            history.append(
                ExtrapolationStep(
                    row=row,
                    column=column,
                    step_size=float(step_sizes[row]),
                    lower_order_estimate=float(current_lower),
                    previous_row_estimate=float(previous_lower),
                    refinement_ratio=float(refinement_ratio),
                    base_order=base_order,
                    exponent=exponent,
                    denominator_factor=float(denominator_factor),
                    correction=float(correction),
                    extrapolated_estimate=float(extrapolated),
                    expected_order=expected_order,
                    exact_derivative=exact_derivative,
                    absolute_error=absolute_error,
                    relative_error_percent=relative_percent,
                    formula_text=formula_text,
                    substitution_text=substitution_text,
                )
            )

    return table, history


def table_column_convergence_rows(
    table: Sequence[Sequence[float | None]],
    step_sizes: Sequence[float],
    base_order: int,
    exact_derivative: float | None,
    refinement_ratio: float,
) -> list[dict[str, Any]]:
    """Analyze convergence down each fixed Richardson column.

    Comparing diagonal cells is not a valid order estimate because every diagonal
    cell has a different formal order. The observed order is therefore calculated
    only between entries in the same column.
    """

    levels = len(table)
    rows: list[dict[str, Any]] = []

    for column in range(levels):
        previous_error: float | None = None
        previous_estimate: float | None = None
        previous_difference: float | None = None

        for row in range(column, levels):
            estimate = table[row][column]
            if estimate is None:
                continue

            estimate_value = float(estimate)
            absolute_error = (
                None
                if exact_derivative is None
                else abs(estimate_value - exact_derivative)
            )
            relative_percent = (
                None
                if absolute_error is None or exact_derivative is None
                else safe_relative_error_percent(absolute_error, exact_derivative)
            )
            successive_difference = (
                None
                if previous_estimate is None
                else abs(estimate_value - previous_estimate)
            )

            if exact_derivative is not None:
                observed_order = calculate_observed_order(
                    previous_error,
                    absolute_error,
                    refinement_ratio,
                )
            else:
                observed_order = calculate_successive_order(
                    previous_difference,
                    successive_difference,
                    refinement_ratio,
                )

            rows.append(
                {
                    "Column": column,
                    "Row": row,
                    "h": float(step_sizes[row]),
                    "Estimate": estimate_value,
                    "Expected Order": base_order * (column + 1),
                    "Exact Derivative": exact_derivative,
                    "Absolute Error": absolute_error,
                    "Relative Error (%)": relative_percent,
                    "Successive Difference": successive_difference,
                    "Observed Order (same column)": observed_order,
                }
            )

            previous_estimate = estimate_value
            previous_error = absolute_error
            previous_difference = successive_difference

    return rows


def solve_richardson_extrapolation(
    function_text: str,
    raw_x_value: Any,
    raw_step_size: Any,
    raw_derivative_order: Any,
    base_method_name: str,
    raw_levels: Any,
    input_signature: str,
) -> RichardsonResult:
    """Validate inputs and run the complete Richardson workflow."""

    execution_datetime = current_report_datetime()

    try:
        x_value = safe_float(raw_x_value, "Evaluation point x₀")
        initial_step_size = safe_float(raw_step_size, "Initial step size h")
        derivative_order = safe_integer(
            raw_derivative_order,
            "Derivative order",
            min(SUPPORTED_DERIVATIVE_ORDERS),
            max(SUPPORTED_DERIVATIVE_ORDERS),
        )
        levels = safe_integer(
            raw_levels,
            "Richardson levels",
            MIN_LEVELS,
            MAX_LEVELS,
        )

        if initial_step_size <= 0.0:
            raise ValueError("Initial step size h must be greater than zero.")

        base_method_key = BASE_METHOD_OPTIONS.get(base_method_name)
        if base_method_key is None:
            raise ValueError("Select a valid base finite-difference method.")

        specification = formula_specification(
            base_method_key,
            derivative_order,
        )

        expression, x_symbol = parse_function(function_text)
        numeric_function = create_numeric_function(expression, x_symbol)

        derivative_expression: sp.Expr | None
        exact_derivative: float | None
        derivative_warning: str | None = None

        try:
            derivative_expression = sp.simplify(
                sp.diff(expression, x_symbol, derivative_order)
            )
            derivative_numeric = create_numeric_function(
                derivative_expression,
                x_symbol,
            )
            exact_derivative = evaluate_real_scalar(
                derivative_numeric,
                x_value,
                f"exact {DERIVATIVE_LABELS[derivative_order]}",
            )
        except (ValueError, TypeError, NotImplementedError) as error:
            derivative_expression = None
            exact_derivative = None
            derivative_warning = (
                "The symbolic derivative could not be evaluated at x₀. "
                "The Richardson estimate remains available, but true-error metrics "
                f"are omitted. Details: {error}"
            )

        base_approximations: list[BaseApproximation] = []
        base_values: list[float] = []
        step_sizes: list[float] = []
        previous_approximation: float | None = None
        previous_error: float | None = None
        previous_difference: float | None = None

        for level in range(levels):
            step_size = initial_step_size / (REFINEMENT_RATIO**level)
            if step_size <= 0.0 or not math.isfinite(step_size):
                raise ValueError(
                    "Step-size refinement became numerically unusable. "
                    "Use fewer levels or a larger initial h."
                )

            calculation = calculate_base_approximation(
                numeric_function=numeric_function,
                x_value=x_value,
                step_size=step_size,
                specification=specification,
            )
            approximation = float(calculation["approximation"])
            absolute_error = (
                None
                if exact_derivative is None
                else abs(approximation - exact_derivative)
            )
            relative_percent = (
                None
                if absolute_error is None or exact_derivative is None
                else safe_relative_error_percent(absolute_error, exact_derivative)
            )
            successive_difference = (
                None
                if previous_approximation is None
                else abs(approximation - previous_approximation)
            )

            if exact_derivative is not None:
                observed_order = calculate_observed_order(
                    previous_error,
                    absolute_error,
                    REFINEMENT_RATIO,
                )
            else:
                observed_order = calculate_successive_order(
                    previous_difference,
                    successive_difference,
                    REFINEMENT_RATIO,
                )

            base_approximations.append(
                BaseApproximation(
                    level=level,
                    step_size=float(step_size),
                    derivative_order=derivative_order,
                    offsets=specification.offsets,
                    sample_x=tuple(calculation["sample_x"]),
                    sample_values=tuple(calculation["sample_values"]),
                    coefficients=specification.coefficients,
                    weighted_numerator=float(calculation["numerator"]),
                    denominator=float(calculation["denominator"]),
                    approximation=approximation,
                    exact_derivative=exact_derivative,
                    absolute_error=absolute_error,
                    relative_error_percent=relative_percent,
                    successive_difference=successive_difference,
                    observed_order=observed_order,
                    substitution_text=str(calculation["substitution_text"]),
                )
            )
            base_values.append(approximation)
            step_sizes.append(step_size)

            previous_approximation = approximation
            previous_error = absolute_error
            previous_difference = successive_difference

        table, extrapolation_steps = build_richardson_table(
            base_values=base_values,
            step_sizes=step_sizes,
            base_order=specification.base_error_order,
            exact_derivative=exact_derivative,
            refinement_ratio=REFINEMENT_RATIO,
        )

        final_estimate = table[-1][-1]
        if final_estimate is None:
            raise ValueError("The final Richardson estimate could not be created.")

        primary_absolute_error = base_approximations[0].absolute_error
        final_absolute_error = (
            None
            if exact_derivative is None
            else abs(float(final_estimate) - exact_derivative)
        )
        final_relative_percent = (
            None
            if final_absolute_error is None or exact_derivative is None
            else safe_relative_error_percent(final_absolute_error, exact_derivative)
        )

        if (
            primary_absolute_error is not None
            and final_absolute_error is not None
            and final_absolute_error > 0.0
        ):
            error_improvement_factor = primary_absolute_error / final_absolute_error
        elif primary_absolute_error is not None and final_absolute_error == 0.0:
            error_improvement_factor = math.inf
        else:
            error_improvement_factor = None

        column_rows = table_column_convergence_rows(
            table=table,
            step_sizes=step_sizes,
            base_order=specification.base_error_order,
            exact_derivative=exact_derivative,
            refinement_ratio=REFINEMENT_RATIO,
        )
        order_candidates = [
            row
            for row in column_rows
            if row["Observed Order (same column)"] is not None
            and math.isfinite(float(row["Observed Order (same column)"]))
        ]
        reliable_order_candidates = [
            row
            for row in order_candidates
            if 0.70 * float(row["Expected Order"])
            <= float(row["Observed Order (same column)"])
            <= 1.30 * float(row["Expected Order"])
        ]
        selected_order_candidates = (
            reliable_order_candidates
            if reliable_order_candidates
            else [
                row
                for row in order_candidates
                if float(row["Observed Order (same column)"]) > 0.0
            ]
        )
        selected_order_candidates.sort(
            key=lambda row: (row["Column"], row["Row"])
        )
        latest_observed_order = (
            float(
                selected_order_candidates[-1][
                    "Observed Order (same column)"
                ]
            )
            if selected_order_candidates
            else None
        )

        diagonal_values = [float(table[row][row]) for row in range(levels)]
        diagonal_errors = (
            [abs(value - exact_derivative) for value in diagonal_values]
            if exact_derivative is not None
            else []
        )
        if diagonal_errors:
            best_index = int(np.argmin(np.asarray(diagonal_errors, dtype=float)))
            best_diagonal_estimate = diagonal_values[best_index]
            best_diagonal_error = diagonal_errors[best_index]
        else:
            best_diagonal_estimate = None
            best_diagonal_error = None

        warnings: list[str] = []
        if derivative_warning:
            warnings.append(derivative_warning)

        finest_h = step_sizes[-1]
        floating_scale = max(1.0, abs(x_value))
        if finest_h <= 100.0 * np.finfo(float).eps * floating_scale:
            warnings.append(
                "The finest h is close to the coordinate resolution of floating-point "
                "arithmetic. Sample points may collapse and round-off can dominate."
            )
        elif derivative_order >= 2 and finest_h <= np.sqrt(np.finfo(float).eps) * floating_scale:
            warnings.append(
                "The finest h is very small for a higher derivative. Division by h^m "
                "can strongly amplify round-off and subtractive cancellation."
            )

        if (
            primary_absolute_error is not None
            and final_absolute_error is not None
            and final_absolute_error > primary_absolute_error
        ):
            warnings.append(
                "The final Richardson value is less accurate than the initial base "
                "estimate. The table may have entered a round-off-dominated range."
            )

        if diagonal_errors and len(diagonal_errors) >= 3:
            if diagonal_errors[-1] > diagonal_errors[-2]:
                warnings.append(
                    "The last diagonal extrapolation is less accurate than the previous "
                    "diagonal value. Consider fewer levels or a larger initial h."
                )

        base_orders = [
            item.observed_order
            for item in base_approximations
            if item.observed_order is not None and math.isfinite(item.observed_order)
        ]
        if base_orders and base_orders[-1] < 0.5 * specification.base_error_order:
            warnings.append(
                "The base finite-difference sequence is not showing the expected "
                f"O(h^{specification.base_error_order}) behavior at the finest levels."
            )

        return RichardsonResult(
            status="success",
            success=True,
            method=METHOD_NAME,
            message="Execution completed successfully.",
            stopping_reason=(
                "All requested finite-difference estimates and Richardson table "
                "entries were calculated."
            ),
            function_text=function_text.strip(),
            function_expression=expression,
            derivative_order=derivative_order,
            derivative_label=DERIVATIVE_LABELS[derivative_order],
            derivative_expression=derivative_expression,
            derivative_expression_text=human_readable_expression(
                derivative_expression
            ),
            base_method_key=base_method_key,
            base_method_name=base_method_name,
            base_formula_text=specification.formula_text,
            base_order=specification.base_error_order,
            x_value=x_value,
            initial_step_size=initial_step_size,
            refinement_ratio=REFINEMENT_RATIO,
            levels=levels,
            base_approximations=tuple(base_approximations),
            extrapolation_steps=tuple(extrapolation_steps),
            richardson_table=tuple(tuple(row) for row in table),
            exact_derivative=exact_derivative,
            primary_base_estimate=base_values[0],
            final_estimate=float(final_estimate),
            final_expected_order=specification.base_error_order * levels,
            primary_absolute_error=primary_absolute_error,
            final_absolute_error=final_absolute_error,
            final_relative_error_percent=final_relative_percent,
            error_improvement_factor=error_improvement_factor,
            latest_observed_order=latest_observed_order,
            best_diagonal_estimate=best_diagonal_estimate,
            best_diagonal_error=best_diagonal_error,
            warnings=tuple(dict.fromkeys(warnings)),
            input_signature=input_signature,
            execution_datetime=execution_datetime,
        )

    except (ValueError, TypeError, ArithmeticError, OverflowError) as error:
        method_key = BASE_METHOD_OPTIONS.get(base_method_name, "central")
        try:
            derivative_order = safe_integer(
                raw_derivative_order,
                "Derivative order",
                1,
                3,
            )
        except ValueError:
            derivative_order = DEFAULT_DERIVATIVE_ORDER
        specification = formula_specification(method_key, derivative_order)

        return RichardsonResult(
            status="error",
            success=False,
            method=METHOD_NAME,
            message=str(error),
            stopping_reason=(
                "The calculation stopped during input validation, function "
                "evaluation, or Richardson table construction."
            ),
            function_text=str(function_text).strip(),
            function_expression=None,
            derivative_order=derivative_order,
            derivative_label=DERIVATIVE_LABELS[derivative_order],
            derivative_expression=None,
            derivative_expression_text="Not available",
            base_method_key=method_key,
            base_method_name=base_method_name,
            base_formula_text=specification.formula_text,
            base_order=specification.base_error_order,
            x_value=None,
            initial_step_size=None,
            refinement_ratio=REFINEMENT_RATIO,
            levels=0,
            base_approximations=(),
            extrapolation_steps=(),
            richardson_table=(),
            exact_derivative=None,
            primary_base_estimate=None,
            final_estimate=None,
            final_expected_order=None,
            primary_absolute_error=None,
            final_absolute_error=None,
            final_relative_error_percent=None,
            error_improvement_factor=None,
            latest_observed_order=None,
            best_diagonal_estimate=None,
            best_diagonal_error=None,
            warnings=(),
            input_signature=input_signature,
            execution_datetime=execution_datetime,
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def create_function_samples_dataframe(result: RichardsonResult) -> pd.DataFrame:
    """Return every function sample used by every base estimate."""

    rows: list[dict[str, Any]] = []
    for item in result.base_approximations:
        for offset, x_sample, f_sample, coefficient in zip(
            item.offsets,
            item.sample_x,
            item.sample_values,
            item.coefficients,
        ):
            rows.append(
                {
                    "Level": item.level,
                    "h": item.step_size,
                    "Offset k": offset,
                    "Coordinate": offset_label(offset),
                    "Sample x = x0 + k*h": x_sample,
                    "f(Sample x)": f_sample,
                    "Formula Coefficient": coefficient,
                    "Weighted Contribution": coefficient * f_sample,
                }
            )
    return pd.DataFrame(rows)


def create_base_approximation_dataframe(result: RichardsonResult) -> pd.DataFrame:
    """Build the base finite-difference approximation table."""

    return pd.DataFrame(
        [
            {
                "Level": item.level,
                "h": item.step_size,
                "Derivative Order": item.derivative_order,
                "Weighted Numerator": item.weighted_numerator,
                "Denominator": item.denominator,
                "Base Approximation R(i,0)": item.approximation,
                "Exact Derivative": item.exact_derivative,
                "Absolute Error": item.absolute_error,
                "Relative Error (%)": item.relative_error_percent,
                "Successive Difference": item.successive_difference,
                "Observed Base Order": item.observed_order,
                "Substitution": item.substitution_text,
            }
            for item in result.base_approximations
        ]
    )


def create_richardson_table_dataframe(result: RichardsonResult) -> pd.DataFrame:
    """Create a rectangular representation of the triangular Richardson table."""

    columns = [
        "h",
        *[
            f"R(i,{column}) — O(h^{result.base_order * (column + 1)})"
            for column in range(result.levels)
        ],
    ]

    rows: list[list[Any]] = []
    for row_index, table_row in enumerate(result.richardson_table):
        rows.append([result.base_approximations[row_index].step_size, *table_row])
    return pd.DataFrame(rows, columns=columns)


def create_extrapolation_steps_dataframe(result: RichardsonResult) -> pd.DataFrame:
    """Return every Richardson recursion calculation."""

    return pd.DataFrame(
        [
            {
                "Row i": item.row,
                "Column j": item.column,
                "h_i": item.step_size,
                "R(i,j-1)": item.lower_order_estimate,
                "R(i-1,j-1)": item.previous_row_estimate,
                "Refinement Ratio": item.refinement_ratio,
                "Exponent p*j": item.exponent,
                "Denominator r^(p*j)-1": item.denominator_factor,
                "Correction": item.correction,
                "R(i,j)": item.extrapolated_estimate,
                "Expected Order": item.expected_order,
                "Exact Derivative": item.exact_derivative,
                "Absolute Error": item.absolute_error,
                "Relative Error (%)": item.relative_error_percent,
                "Formula": item.formula_text,
                "Substitution": item.substitution_text,
            }
            for item in result.extrapolation_steps
        ]
    )


def create_column_convergence_dataframe(result: RichardsonResult) -> pd.DataFrame:
    """Return correct observed-order analysis down fixed table columns."""

    rows = table_column_convergence_rows(
        table=result.richardson_table,
        step_sizes=[item.step_size for item in result.base_approximations],
        base_order=result.base_order,
        exact_derivative=result.exact_derivative,
        refinement_ratio=result.refinement_ratio,
    )
    return pd.DataFrame(rows)


def create_diagonal_analysis_dataframe(result: RichardsonResult) -> pd.DataFrame:
    """Analyze the improving diagonal sequence without mislabeling its order."""

    rows: list[dict[str, Any]] = []
    previous_estimate: float | None = None
    previous_error: float | None = None

    for row in range(result.levels):
        estimate = result.richardson_table[row][row]
        if estimate is None:
            continue

        estimate_value = float(estimate)
        absolute_error = (
            None
            if result.exact_derivative is None
            else abs(estimate_value - result.exact_derivative)
        )
        relative_percent = (
            None
            if absolute_error is None or result.exact_derivative is None
            else safe_relative_error_percent(absolute_error, result.exact_derivative)
        )
        successive_difference = (
            None
            if previous_estimate is None
            else abs(estimate_value - previous_estimate)
        )
        improvement_factor = (
            None
            if previous_error is None
            or absolute_error is None
            or absolute_error <= 0.0
            else previous_error / absolute_error
        )

        rows.append(
            {
                "Level": row,
                "h": result.base_approximations[row].step_size,
                "Diagonal Estimate R(i,i)": estimate_value,
                "Formal Order of This Cell": result.base_order * (row + 1),
                "Exact Derivative": result.exact_derivative,
                "Absolute Error": absolute_error,
                "Relative Error (%)": relative_percent,
                "Successive Diagonal Difference": successive_difference,
                "Error Improvement from Previous Diagonal": improvement_factor,
            }
        )

        previous_estimate = estimate_value
        previous_error = absolute_error

    return pd.DataFrame(rows)


def create_summary_dataframe(result: RichardsonResult) -> pd.DataFrame:
    """Create the Excel property-value summary."""

    warnings = " | ".join(result.warnings) if result.warnings else "None"

    return pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "Function",
                "Symbolic Function",
                "Derivative Order",
                "Derivative Label",
                "Symbolic Derivative",
                "Base Method",
                "Base Formula",
                "Base Error Order",
                "Evaluation Point x0",
                "Initial Step Size h",
                "Refinement Ratio",
                "Richardson Levels",
                "Initial Base Estimate",
                "Final Richardson Estimate",
                "Exact Derivative",
                "Expected Final Formal Order",
                "Initial Absolute Error",
                "Final Absolute Error",
                "Final Relative Error (%)",
                "Error Improvement Factor",
                "Latest Observed Order (Fixed Column)",
                "Best Diagonal Estimate",
                "Best Diagonal Error",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                result.function_text,
                human_readable_expression(result.function_expression),
                result.derivative_order,
                result.derivative_label,
                result.derivative_expression_text,
                result.base_method_name,
                result.base_formula_text,
                f"O(h^{result.base_order})",
                result.x_value,
                result.initial_step_size,
                result.refinement_ratio,
                result.levels,
                result.primary_base_estimate,
                result.final_estimate,
                result.exact_derivative,
                result.final_expected_order,
                result.primary_absolute_error,
                result.final_absolute_error,
                result.final_relative_error_percent,
                result.error_improvement_factor,
                result.latest_observed_order,
                result.best_diagonal_estimate,
                result.best_diagonal_error,
                warnings,
                result.stopping_reason,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )


def create_method_formulas_dataframe(result: RichardsonResult) -> pd.DataFrame:
    """Return the formulas used in the calculation."""

    return pd.DataFrame(
        {
            "Item": [
                "Requested derivative",
                "Base formula",
                "Base error model",
                "Step refinement",
                "Richardson recursion",
                "Column expected order",
                "Observed-order rule",
                "Relative error",
            ],
            "Expression / Meaning": [
                result.derivative_label,
                result.base_formula_text,
                f"D(h) = D + c1*h^{result.base_order} + higher powers",
                "h_i = h_0 / 2^i",
                (
                    "R(i,j)=R(i,j-1)+[R(i,j-1)-R(i-1,j-1)]/"
                    "[2^(p*j)-1]"
                ),
                "O(h^[p*(j+1)])",
                (
                    "Observed order is calculated only down a fixed Richardson "
                    "column, because diagonal cells have different formal orders"
                ),
                "|approx-exact| / |exact| × 100%",
            ],
        }
    )


# =============================================================================
# Scientific plots
# =============================================================================
def create_function_plot(result: RichardsonResult) -> Figure:
    """Plot the function and all sample points used by the first base estimate."""

    if not result.success or result.function_expression is None or result.x_value is None:
        raise ValueError("A successful result is required for plotting.")

    x_symbol = sp.Symbol("x", real=True)
    numeric_function = create_numeric_function(result.function_expression, x_symbol)

    largest_offset = max(abs(offset) for offset in result.base_approximations[0].offsets)
    largest_h = result.initial_step_size or 1.0
    span = max(1.25 * largest_offset * largest_h, 1.0, 0.25 * abs(result.x_value))
    x_values = np.linspace(result.x_value - span, result.x_value + span, 700)
    y_values = evaluate_real_array(numeric_function, x_values)
    valid = np.isfinite(y_values)

    if np.count_nonzero(valid) < 10:
        raise ValueError("The function is not finite over enough of the plotting interval.")

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.plot(
        x_values[valid],
        y_values[valid],
        linewidth=2,
        label=f"f(x) = {result.function_text}",
    )
    axis.axhline(0.0, linewidth=1)
    axis.axvline(result.x_value, linestyle="--", linewidth=1, label="x₀")

    primary = result.base_approximations[0]
    axis.scatter(
        primary.sample_x,
        primary.sample_values,
        s=65,
        zorder=5,
        label="Base-formula sample points",
    )

    point_y = evaluate_real_scalar(numeric_function, result.x_value, "f(x₀)")
    axis.scatter(
        [result.x_value],
        [point_y],
        s=85,
        marker="s",
        zorder=6,
        label="Evaluation point",
    )

    if result.derivative_order == 1 and result.final_estimate is not None:
        tangent_x = np.linspace(
            result.x_value - 0.45 * span,
            result.x_value + 0.45 * span,
            180,
        )
        tangent_y = point_y + result.final_estimate * (tangent_x - result.x_value)
        axis.plot(
            tangent_x,
            tangent_y,
            linestyle="--",
            linewidth=2,
            label=f"Tangent slope ≈ {format_number(result.final_estimate, 8)}",
        )

    axis.set_title(
        f"Function Samples for Richardson Approximation of {result.derivative_label}"
    )
    axis.set_xlabel("x")
    axis.set_ylabel("f(x)")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_convergence_plot(result: RichardsonResult) -> Figure:
    """Plot errors down fixed Richardson columns on log-log axes."""

    dataframe = create_column_convergence_dataframe(result)
    figure, axis = plt.subplots(figsize=(10, 6))
    plotted = False

    for column in sorted(dataframe["Column"].unique()):
        subset = dataframe[dataframe["Column"] == column]
        h_values = subset["h"].to_numpy(dtype=float)

        if result.exact_derivative is not None:
            metric = subset["Absolute Error"].to_numpy(dtype=float)
            metric_name = "Absolute error"
        else:
            metric = subset["Successive Difference"].to_numpy(dtype=float)
            metric_name = "Successive difference"

        valid = (
            np.isfinite(h_values)
            & (h_values > 0.0)
            & np.isfinite(metric)
            & (metric > 0.0)
        )
        if not np.any(valid):
            continue

        expected_order = result.base_order * (int(column) + 1)
        axis.loglog(
            h_values[valid],
            metric[valid],
            marker="o",
            linewidth=2,
            label=f"Column {column}: expected O(h^{expected_order})",
        )
        plotted = True

    if not plotted:
        raise ValueError("There are not enough positive convergence indicators to plot.")

    axis.invert_xaxis()
    axis.set_title("Richardson Convergence by Fixed Table Column")
    axis.set_xlabel("Step size h")
    axis.set_ylabel(metric_name)
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def figure_to_png_bytes(figure: Figure) -> bytes:
    """Serialize a matplotlib figure to PNG bytes."""

    buffer = BytesIO()
    figure.savefig(buffer, format="png", dpi=180, bbox_inches="tight")
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# Excel export
# =============================================================================
def apply_excel_style(workbook: Any) -> None:
    """Apply readable professional formatting to all worksheets."""

    header_fill = PatternFill("solid", fgColor="0D3151")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False

        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            maximum_length = max(
                (
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                ),
                default=0,
            )
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(maximum_length + 2, 12),
                58,
            )


def add_excel_image(
    worksheet: Any,
    image_bytes: bytes,
    anchor: str,
    width: int = 780,
    height: int = 480,
) -> None:
    """Insert PNG bytes into an Excel worksheet."""

    image_stream = BytesIO(image_bytes)
    image = ExcelImage(image_stream)
    image.width = width
    image.height = height
    worksheet.add_image(image, anchor)


def create_excel_report(result: RichardsonResult) -> bytes:
    """Generate a complete formatted XLSX report in memory."""

    if not result.success:
        raise ValueError("Only a successful result can be exported.")

    summary_df = create_summary_dataframe(result)
    formulas_df = create_method_formulas_dataframe(result)
    samples_df = create_function_samples_dataframe(result)
    base_df = create_base_approximation_dataframe(result)
    table_df = create_richardson_table_dataframe(result)
    steps_df = create_extrapolation_steps_dataframe(result)
    column_df = create_column_convergence_dataframe(result)
    diagonal_df = create_diagonal_analysis_dataframe(result)

    function_figure = create_function_plot(result)
    function_png = figure_to_png_bytes(function_figure)
    plt.close(function_figure)

    convergence_png: bytes | None = None
    try:
        convergence_figure = create_convergence_plot(result)
    except ValueError:
        convergence_figure = None
    if convergence_figure is not None:
        convergence_png = figure_to_png_bytes(convergence_figure)
        plt.close(convergence_figure)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        formulas_df.to_excel(writer, sheet_name="Method Formulas", index=False)
        samples_df.to_excel(writer, sheet_name="Function Samples", index=False)
        base_df.to_excel(writer, sheet_name="Base Approximations", index=False)
        table_df.to_excel(writer, sheet_name="Richardson Table", index=False)
        steps_df.to_excel(writer, sheet_name="Extrapolation Steps", index=False)
        column_df.to_excel(writer, sheet_name="Column Convergence", index=False)
        diagonal_df.to_excel(writer, sheet_name="Diagonal Analysis", index=False)

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Richardson Extrapolation Plots"
        plots_sheet["A1"].font = Font(bold=True, size=14)
        add_excel_image(plots_sheet, function_png, "A3")
        if convergence_png is not None:
            add_excel_image(plots_sheet, convergence_png, "A31")

        summary_sheet = workbook["Summary"]
        column_sheet = workbook["Column Convergence"]

        if not column_df.empty:
            row_count = len(column_df) + 1

            estimate_chart = LineChart()
            estimate_chart.title = "Richardson Estimates by Table Cell"
            estimate_chart.x_axis.title = "Table Record"
            estimate_chart.y_axis.title = result.derivative_label
            estimate_chart.height = 8
            estimate_chart.width = 16
            estimate_chart.add_data(
                Reference(column_sheet, min_col=4, min_row=1, max_row=row_count),
                titles_from_data=True,
            )
            estimate_chart.set_categories(
                Reference(column_sheet, min_col=2, min_row=2, max_row=row_count)
            )
            summary_sheet.add_chart(estimate_chart, "D2")

            metric_column = 7 if result.exact_derivative is not None else 9
            convergence_chart = LineChart()
            convergence_chart.title = "Richardson Error Indicator"
            convergence_chart.x_axis.title = "Table Record"
            convergence_chart.y_axis.title = (
                "Absolute Error"
                if result.exact_derivative is not None
                else "Successive Difference"
            )
            convergence_chart.height = 8
            convergence_chart.width = 16
            convergence_chart.add_data(
                Reference(
                    column_sheet,
                    min_col=metric_column,
                    min_row=1,
                    max_row=row_count,
                ),
                titles_from_data=True,
            )
            convergence_chart.set_categories(
                Reference(column_sheet, min_col=2, min_row=2, max_row=row_count)
            )
            summary_sheet.add_chart(convergence_chart, "D20")

        apply_excel_style(workbook)
        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit rendering helpers
# =============================================================================
def render_final_result(result: RichardsonResult) -> None:
    """Render the compact final result card."""

    if not result.success:
        st.error(result.message)
        st.caption(result.stopping_reason)
        return

    st.success(result.message)
    st.markdown(f"**Function:** `{result.function_text}`")
    st.markdown(f"**Requested derivative:** {result.derivative_label}")
    st.markdown(f"**Base method:** {result.base_method_name}")

    metric_columns = st.columns(2)
    metric_columns[0].metric(
        "Final Richardson Estimate",
        format_number(result.final_estimate, 10),
    )
    metric_columns[1].metric(
        f"Exact {result.derivative_label}",
        format_number(result.exact_derivative, 10),
    )

    detail_columns = st.columns(2)
    detail_columns[0].metric("Levels", result.levels)
    detail_columns[1].metric(
        "Final Formal Order",
        f"O(h^{result.final_expected_order})",
    )

    if result.final_absolute_error is not None:
        error_columns = st.columns(2)
        error_columns[0].metric(
            "Final Absolute Error",
            scientific_number(result.final_absolute_error),
        )
        error_columns[1].metric(
            "Final Relative Error",
            (
                "Undefined"
                if result.final_relative_error_percent is None
                else f"{format_number(result.final_relative_error_percent, 6)}%"
            ),
        )

    if result.latest_observed_order is not None:
        st.metric(
            "Latest Observed Order (Same Column)",
            format_number(result.latest_observed_order, 4),
        )

    for warning in result.warnings:
        st.warning(warning)


def render_base_calculations(result: RichardsonResult) -> None:
    """Render base formulas, samples, and estimates."""

    st.subheader("Base Finite-Difference Approximations")
    st.markdown(f"**Formula:** {result.base_formula_text}")
    st.markdown(
        f"**Theoretical base truncation order:** O(h^{result.base_order})"
    )

    st.dataframe(
        round_numeric_dataframe(create_base_approximation_dataframe(result), 10),
        use_container_width=True,
        hide_index=True,
    )

    for item in result.base_approximations:
        with st.expander(
            f"Level {item.level}: h = {format_number(item.step_size, 10)}",
            expanded=(item.level == 0),
        ):
            sample_df = pd.DataFrame(
                {
                    "Offset": item.offsets,
                    "Coordinate": [offset_label(offset) for offset in item.offsets],
                    "Sample x": item.sample_x,
                    "f(Sample x)": item.sample_values,
                    "Coefficient": item.coefficients,
                }
            )
            st.dataframe(
                round_numeric_dataframe(sample_df, 10),
                use_container_width=True,
                hide_index=True,
            )
            st.code(item.substitution_text, language=None)
            st.markdown(
                f"**Base estimate:** {format_number(item.approximation, 12)}"
            )
            if item.absolute_error is not None:
                st.markdown(
                    f"**Absolute error:** {scientific_number(item.absolute_error)}"
                )


def render_richardson_table(result: RichardsonResult) -> None:
    """Render the table and every extrapolation step."""

    st.subheader("Richardson Extrapolation Table")
    st.dataframe(
        round_numeric_dataframe(create_richardson_table_dataframe(result), 10),
        use_container_width=True,
        hide_index=True,
    )

    st.info(
        "Each column removes the next term in the assumed truncation-error "
        f"expansion. The selected base method begins at O(h^{result.base_order}), "
        "so column j uses 2^(p·j)−1."
    )

    st.markdown("#### Extrapolation Steps")
    for step in result.extrapolation_steps:
        with st.expander(
            f"R({step.row},{step.column}) — formal O(h^{step.expected_order})"
        ):
            st.markdown(f"**Formula:** `{step.formula_text}`")
            st.code(step.substitution_text, language=None)
            detail_df = pd.DataFrame(
                {
                    "Property": [
                        "Step size",
                        "Current lower-order estimate",
                        "Previous-row estimate",
                        "Exponent",
                        "Denominator factor",
                        "Correction",
                        "Extrapolated estimate",
                        "Absolute error",
                    ],
                    "Value": [
                        step.step_size,
                        step.lower_order_estimate,
                        step.previous_row_estimate,
                        step.exponent,
                        step.denominator_factor,
                        step.correction,
                        step.extrapolated_estimate,
                        step.absolute_error,
                    ],
                }
            )
            st.dataframe(
                round_numeric_dataframe(detail_df, 10),
                use_container_width=True,
                hide_index=True,
            )


def render_error_analysis(result: RichardsonResult) -> None:
    """Render correct fixed-column convergence and diagonal improvement tables."""

    st.subheader("Error and Convergence Analysis")
    st.markdown("#### Fixed-Column Convergence")
    st.dataframe(
        round_numeric_dataframe(create_column_convergence_dataframe(result), 10),
        use_container_width=True,
        hide_index=True,
    )
    st.caption(
        "Observed order is calculated only between entries in the same Richardson "
        "column. Diagonal entries are not compared for order because each diagonal "
        "entry has a different formal accuracy."
    )

    st.markdown("#### Diagonal Improvement Sequence")
    st.dataframe(
        round_numeric_dataframe(create_diagonal_analysis_dataframe(result), 10),
        use_container_width=True,
        hide_index=True,
    )

    try:
        figure = create_convergence_plot(result)
    except ValueError as error:
        st.info(str(error))
    else:
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)


def render_function_graph(result: RichardsonResult) -> None:
    """Render function samples and an optional tangent for the first derivative."""

    st.subheader("Function Graph")
    try:
        figure = create_function_plot(result)
    except ValueError as error:
        st.warning(f"The graph could not be displayed: {error}")
    else:
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)


def render_excel_download(result: RichardsonResult) -> None:
    """Generate and render the Excel report button."""

    st.subheader("Excel Report")
    signature_key = "richardson_excel_signature"
    report_key = "richardson_excel_report"

    if (
        st.session_state.get(signature_key) != result.input_signature
        or report_key not in st.session_state
    ):
        try:
            st.session_state[report_key] = create_excel_report(result)
            st.session_state[signature_key] = result.input_signature
        except (ValueError, OSError, RuntimeError, TypeError) as error:
            st.error(f"The Excel report could not be generated: {error}")
            return

    report_bytes = st.session_state.get(report_key)
    if not report_bytes:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=f"richardson_extrapolation_report_{timestamp}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="richardson_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Richardson Extrapolation Streamlit page."""

    st.set_page_config(
        page_title="Richardson Extrapolation Solver | Numerical Methods",
        page_icon="📈",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">NUMERICAL DIFFERENTIATION TOOL</div>
                <h1>Richardson Extrapolation Solver</h1>
                <p>
                    Improve first-, second-, or third-derivative finite-difference
                    estimates by combining successively refined step sizes. Review
                    every sampled function value, the complete Richardson table,
                    fixed-column convergence, graphs, and Excel report.
                </p>

                <div class="method-actions">
                    <a href="/Richardson_Extrapolation" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Richardson_Extrapolation_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])

    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Write the Function")
                st.markdown(
                    """
                    Enter only the mathematical expression, without `f(x)=`.

                    - Use only **x** as the variable.
                    - Powers: write `x**2`, not `x^2`.
                    - Multiplication: write `2*x`, not `2x`.
                    - Functions: `sin(x)`, `cos(x)`, `exp(x)`, `sqrt(x)`, and `log(x)`.
                    - Select derivative order **1**, **2**, or **3**.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")
                st.markdown(
                    """
                    - Choose a positive initial **h** and at least two levels.
                    - The function must be finite at every sample required by the formula.
                    - Central formulas have base error **O(h²)**; forward formulas have **O(h)**.
                    - Richardson assumes a regular truncation-error expansion.
                    - Excessively small **h** amplifies round-off, especially for higher derivatives.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">Function f(x)</div>',
                    unsafe_allow_html=True,
                )
                function_text = st.text_input(
                    "Function f(x)",
                    value=DEFAULT_FUNCTION,
                    placeholder="Example: sin(x), exp(x), or x**5 - 2*x",
                    label_visibility="collapsed",
                    key="richardson_function",
                )

                first_row = st.columns(3)
                with first_row[0]:
                    st.markdown(
                        '<div class="input-label-ui">Evaluation point x₀</div>',
                        unsafe_allow_html=True,
                    )
                    x_value = st.number_input(
                        "Evaluation point x0",
                        value=DEFAULT_X_VALUE,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="richardson_x_value",
                    )

                with first_row[1]:
                    st.markdown(
                        '<div class="input-label-ui">Initial step size h</div>',
                        unsafe_allow_html=True,
                    )
                    step_size = st.number_input(
                        "Initial step size h",
                        value=DEFAULT_STEP_SIZE,
                        min_value=0.0,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="richardson_step_size",
                    )

                with first_row[2]:
                    st.markdown(
                        '<div class="input-label-ui">Derivative order</div>',
                        unsafe_allow_html=True,
                    )
                    derivative_order = st.selectbox(
                        "Derivative order",
                        options=SUPPORTED_DERIVATIVE_ORDERS,
                        index=SUPPORTED_DERIVATIVE_ORDERS.index(
                            DEFAULT_DERIVATIVE_ORDER
                        ),
                        format_func=lambda value: (
                            f"Order {value} — {DERIVATIVE_LABELS[value]}"
                        ),
                        label_visibility="collapsed",
                        key="richardson_derivative_order",
                    )

                second_row = st.columns(2)
                with second_row[0]:
                    st.markdown(
                        '<div class="input-label-ui">Base difference method</div>',
                        unsafe_allow_html=True,
                    )
                    base_method_name = st.selectbox(
                        "Base difference method",
                        options=list(BASE_METHOD_OPTIONS.keys()),
                        index=0,
                        label_visibility="collapsed",
                        key="richardson_base_method",
                    )

                with second_row[1]:
                    st.markdown(
                        '<div class="input-label-ui">Richardson levels</div>',
                        unsafe_allow_html=True,
                    )
                    levels = st.slider(
                        "Richardson levels",
                        min_value=MIN_LEVELS,
                        max_value=MAX_LEVELS,
                        value=DEFAULT_LEVELS,
                        step=1,
                        label_visibility="collapsed",
                        key="richardson_levels",
                    )

                selected_method_key = BASE_METHOD_OPTIONS[base_method_name]
                selected_specification = formula_specification(
                    selected_method_key,
                    int(derivative_order),
                )
                st.markdown("**Selected base formula**")
                st.code(selected_specification.formula_text, language=None)
                st.caption(
                    "The solver uses h, h/2, h/4, and so on. More levels do not "
                    "always mean greater accuracy because round-off eventually grows."
                )

                current_signature = create_input_signature(
                    function_text=function_text,
                    x_value=x_value,
                    step_size=step_size,
                    derivative_order=derivative_order,
                    base_method_name=base_method_name,
                    levels=levels,
                )

                solve_button_clicked = st.button(
                    "Solve",
                    type="primary",
                    use_container_width=True,
                    key="richardson_solve_button",
                )

                if solve_button_clicked:
                    st.session_state.richardson_result = solve_richardson_extrapolation(
                        function_text=function_text,
                        raw_x_value=x_value,
                        raw_step_size=step_size,
                        raw_derivative_order=derivative_order,
                        base_method_name=base_method_name,
                        raw_levels=levels,
                        input_signature=current_signature,
                    )
                    st.session_state.pop("richardson_excel_report", None)
                    st.session_state.pop("richardson_excel_signature", None)
                    st.rerun()

                with st.expander("Example Inputs"):
                    st.code(
                        "Function: sin(x)\n"
                        "x0 = 1\n"
                        "h = 0.2\n"
                        "Derivative order = 1, 2, or 3\n"
                        "Base method = Central difference\n"
                        "Levels = 5",
                        language=None,
                    )
                    st.code(
                        "Function: exp(x)\n"
                        "x0 = 0\n"
                        "h = 0.1\n"
                        "Derivative order = 3\n"
                        "Base method = Forward difference\n"
                        "Levels = 5",
                        language=None,
                    )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                stored_result = st.session_state.get("richardson_result")
                if stored_result is None:
                    st.info("Enter the function and parameters, then click Solve.")
                elif stored_result.input_signature != current_signature:
                    st.info(
                        "The function or numerical parameters have changed. "
                        "Click Solve to calculate a new result."
                    )
                else:
                    render_final_result(stored_result)

        active_result = st.session_state.get("richardson_result")
        if (
            active_result is not None
            and active_result.input_signature == current_signature
            and active_result.success
        ):
            st.divider()
            render_base_calculations(active_result)

            st.divider()
            render_richardson_table(active_result)

            st.divider()
            render_error_analysis(active_result)

            st.divider()
            render_function_graph(active_result)

            st.divider()
            render_excel_download(active_result)

            st.divider()
            navigation_left_column, navigation_right_column = st.columns(2)

            with navigation_left_column:
                if st.button(
                    "Review Richardson Extrapolation Lesson",
                    use_container_width=True,
                    key="review_richardson_lesson",
                ):
                    st.switch_page("pages/Richardson_Extrapolation.py")

            with navigation_right_column:
                if st.button(
                    "Back to Solver Menu",
                    use_container_width=True,
                    key="back_to_solver_menu_richardson",
                ):
                    st.switch_page("pages/Numerical_Solver.py")

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Numerical Differentiation • Richardson Extrapolation</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()