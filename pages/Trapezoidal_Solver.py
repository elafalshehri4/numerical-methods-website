from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css
from utilities.safe_math import safe_sympify

# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
        value: float | int | None,
        decimals: int = 3,
        unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
        value: float | int | None,
        decimals: int = 3,
        unavailable: str = "—",
) -> str:
    """Show three decimals and use × 10ⁿ when fixed notation loses meaning."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Trapezoidal Rule"
DISPLAY_DECIMALS = 3
DEFAULT_FUNCTION = "sin(x)"
DEFAULT_LOWER_LIMIT = 0.0
DEFAULT_UPPER_LIMIT = float(np.pi)
DEFAULT_SUBINTERVALS = 8
MIN_SUBINTERVALS = 1
MAX_SUBINTERVALS = 500
DEFAULT_CONVERGENCE_LEVELS = 5
MIN_CONVERGENCE_LEVELS = 3
MAX_CONVERGENCE_LEVELS = 7
ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
CONTINUITY_SAMPLE_COUNT = 2001
SMOOTHNESS_SAMPLE_COUNT = 1201
MAX_CONVERGENCE_SUBINTERVALS = 100_000
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

RULE_OPTIONS = {
    "Composite Trapezoidal Rule": "composite",
    "Simple Trapezoidal Rule": "simple",
}

ALLOWED_FUNCTION_NAMES = {
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class FunctionSample:
    """One function value used by the Trapezoidal Rule."""

    index: int
    x_value: float
    function_value: float
    weight: int
    weighted_value: float
    point_type: str


@dataclass(frozen=True)
class TrapezoidContribution:
    """One subinterval and its trapezoidal-area contribution."""

    interval_index: int
    left_index: int
    right_index: int
    x_left: float
    x_right: float
    interval_width: float
    f_left: float
    f_right: float
    average_height: float
    signed_area: float
    absolute_geometric_area: float
    cumulative_signed_area: float
    formula_text: str
    substitution_text: str


@dataclass(frozen=True)
class ConvergenceRecord:
    """One refined composite Trapezoidal Rule calculation."""

    level: int
    subintervals: int
    step_size: float
    approximation: float
    exact_integral: float | None
    absolute_error: float | None
    relative_error: float | None
    successive_difference: float | None
    estimated_error_from_refinement: float | None
    richardson_extrapolated: float | None
    observed_order: float | None


@dataclass(frozen=True)
class TrapezoidalResult:
    """Complete result shared by Streamlit and Excel renderers."""

    status: str
    success: bool
    method: str
    rule_key: str
    rule_name: str
    message: str
    stopping_reason: str
    function_text: str
    function_expression: sp.Expr | None
    antiderivative_expression: sp.Expr | None
    antiderivative_text: str
    reference_integral_source: str
    lower_limit: float | None
    upper_limit: float | None
    orientation: str
    subintervals: int
    step_size: float | None
    function_samples: tuple[FunctionSample, ...]
    interval_contributions: tuple[TrapezoidContribution, ...]
    endpoint_sum: float | None
    interior_sum: float | None
    weighted_sum: float | None
    approximate_integral: float | None
    exact_integral: float | None
    absolute_error: float | None
    relative_error: float | None
    percentage_error: float | None
    theoretical_order: int
    second_derivative_expression: sp.Expr | None
    second_derivative_text: str
    sampled_max_second_derivative: float | None
    theoretical_error_bound: float | None
    polynomial_degree: int | None
    exact_for_linear: bool
    interval_continuity_status: str
    convergence_rule_name: str
    convergence_records: tuple[ConvergenceRecord, ...]
    latest_observed_order: float | None
    final_refinement_error_estimate: float | None
    final_richardson_estimate: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def format_number(
        value: float | int | None,
        decimals: int = 3,
) -> str:
    """Format displayed values with three decimals and × 10ⁿ notation."""

    return format_display_number(value, decimals)


def scientific_number(value: float | int | None) -> str:
    """Format a value as a three-decimal coefficient multiplied by 10ⁿ."""

    return format_scientific_power(value)


def round_numeric_dataframe(
        dataframe: pd.DataFrame,
        decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only for display."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def create_input_signature(
        function_text: str,
        lower_limit: Any,
        upper_limit: Any,
        rule_name: str,
        subintervals: Any,
        convergence_levels: Any,
) -> str:
    """Create a stable signature used to prevent stale Streamlit results."""

    payload = repr(
        (
            str(function_text).strip(),
            str(lower_limit),
            str(upper_limit),
            str(rule_name),
            str(subintervals),
            str(convergence_levels),
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def safe_float(raw_value: Any, value_name: str) -> float:
    """Convert one user input to a finite real floating-point value."""

    if raw_value is None:
        raise ValueError(f"{value_name} is required.")
    if isinstance(raw_value, str) and not raw_value.strip():
        raise ValueError(f"{value_name} is required.")

    try:
        value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be a valid real number.") from error

    if not math.isfinite(value):
        raise ValueError(f"{value_name} must be finite; NaN and infinity are not allowed.")
    return value


def safe_positive_integer(
        raw_value: Any,
        value_name: str,
        minimum: int,
        maximum: int,
) -> int:
    """Convert one user input to a bounded positive integer."""

    if isinstance(raw_value, bool):
        raise ValueError(f"{value_name} must be an integer.")

    try:
        numeric_value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be an integer.") from error

    if not math.isfinite(numeric_value) or not numeric_value.is_integer():
        raise ValueError(f"{value_name} must be an integer.")

    integer_value = int(numeric_value)
    if integer_value < minimum or integer_value > maximum:
        raise ValueError(
            f"{value_name} must be between {minimum} and {maximum}."
        )
    return integer_value


def empty_result(
        *,
        message: str,
        stopping_reason: str,
        input_signature: str,
        rule_key: str = "composite",
        rule_name: str = "Composite Trapezoidal Rule",
) -> TrapezoidalResult:
    """Create a consistent structured error result."""

    return TrapezoidalResult(
        status="error",
        success=False,
        method=METHOD_NAME,
        rule_key=rule_key,
        rule_name=rule_name,
        message=message,
        stopping_reason=stopping_reason,
        function_text="",
        function_expression=None,
        antiderivative_expression=None,
        antiderivative_text="Not available",
        reference_integral_source="Not available",
        lower_limit=None,
        upper_limit=None,
        orientation="Not available",
        subintervals=0,
        step_size=None,
        function_samples=(),
        interval_contributions=(),
        endpoint_sum=None,
        interior_sum=None,
        weighted_sum=None,
        approximate_integral=None,
        exact_integral=None,
        absolute_error=None,
        relative_error=None,
        percentage_error=None,
        theoretical_order=2,
        second_derivative_expression=None,
        second_derivative_text="Not available",
        sampled_max_second_derivative=None,
        theoretical_error_bound=None,
        polynomial_degree=None,
        exact_for_linear=False,
        interval_continuity_status="Not checked",
        convergence_rule_name="Not available",
        convergence_records=(),
        latest_observed_order=None,
        final_refinement_error_estimate=None,
        final_richardson_estimate=None,
        warnings=(),
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


# =============================================================================
# Function parsing and safe evaluation
# =============================================================================
def parse_function(function_text: str) -> tuple[sp.Symbol, sp.Expr]:
    """Parse and validate a real single-variable mathematical function."""

    if not isinstance(function_text, str) or not function_text.strip():
        raise ValueError("Function f(x) is required.")

    cleaned_text = function_text.strip().replace("^", "**")
    if "=" in cleaned_text:
        raise ValueError(
            "Enter only the function expression, not an equation. "
            "For example, enter sin(x) instead of y = sin(x)."
        )

    x_symbol = sp.Symbol("x", real=True)
    local_dictionary = {"x": x_symbol, **ALLOWED_FUNCTION_NAMES}

    try:
        expression = safe_sympify(cleaned_text, locals=local_dictionary)
    except (sp.SympifyError, SyntaxError, TypeError, ValueError) as error:
        raise ValueError(
            "The function format is invalid. Use Python/SymPy syntax such as "
            "sin(x), exp(x), sqrt(x), or x**3 - 2*x + 1."
        ) from error

    if isinstance(expression, Relational):
        raise ValueError("The input must be a function expression, not a relation.")

    unknown_symbols = expression.free_symbols - {x_symbol}
    if unknown_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unknown_symbols))
        raise ValueError(
            "Only the independent variable x is allowed. "
            f"Unexpected symbol(s): {names}."
        )

    undefined_functions = expression.atoms(AppliedUndef)
    if undefined_functions:
        names = ", ".join(sorted(str(item.func) for item in undefined_functions))
        raise ValueError(f"Unsupported function name(s): {names}.")

    if expression.has(sp.zoo, sp.oo, -sp.oo, sp.nan):
        raise ValueError("The function contains an undefined or infinite quantity.")

    return x_symbol, sp.simplify(expression)


def build_numeric_function(
        x_symbol: sp.Symbol,
        expression: sp.Expr,
) -> Callable[[Any], Any]:
    """Convert a symbolic expression to a NumPy-compatible function."""

    try:
        return sp.lambdify(x_symbol, expression, modules=["numpy"])
    except (TypeError, ValueError, NotImplementedError) as error:
        raise ValueError(
            "The function could not be converted to a numerical function."
        ) from error


def evaluate_real_scalar(
        numeric_function: Callable[[Any], Any],
        x_value: float,
        value_name: str,
) -> float:
    """Evaluate a function safely and return one finite real scalar."""

    try:
        with np.errstate(all="ignore"):
            raw_value = numeric_function(float(x_value))
    except (ArithmeticError, TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The function could not be evaluated at x = {x_value:.12g}."
        ) from error

    array = np.asarray(raw_value)
    if array.size != 1:
        raise ValueError(f"{value_name} did not produce a scalar value.")

    scalar = array.reshape(-1)[0]
    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)
        if abs(complex_value.imag) > 1.0e-12:
            raise ValueError(
                f"The function is complex at x = {x_value:.12g}; "
                "this solver requires real-valued integrands."
            )
        scalar = complex_value.real

    try:
        value = float(scalar)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The function value at x = {x_value:.12g} is not a valid real number."
        ) from error

    if not math.isfinite(value):
        raise ValueError(
            f"The function is undefined or non-finite at x = {x_value:.12g}."
        )
    return value


def evaluate_real_array(
        numeric_function: Callable[[Any], Any],
        x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate a function on an array, replacing invalid values with NaN."""

    try:
        with np.errstate(all="ignore"):
            raw_values = numeric_function(x_values)
    except (ArithmeticError, TypeError, ValueError, OverflowError):
        return np.full_like(x_values, np.nan, dtype=float)

    array = np.asarray(raw_values)
    if array.ndim == 0:
        array = np.full_like(
            x_values,
            array,
            dtype=np.complex128 if np.iscomplexobj(array) else float,
        )
    else:
        try:
            array = np.broadcast_to(array, x_values.shape)
        except ValueError:
            return np.full_like(x_values, np.nan, dtype=float)

    if np.iscomplexobj(array):
        imaginary = np.abs(np.imag(array))
        values = np.real(array).astype(float)
        values[imaginary > 1.0e-12] = np.nan
    else:
        try:
            values = np.asarray(array, dtype=float).copy()
        except (TypeError, ValueError, OverflowError):
            return np.full_like(x_values, np.nan, dtype=float)

    values[~np.isfinite(values)] = np.nan
    return values


def analyze_interval_and_smoothness(
        expression: sp.Expr,
        x_symbol: sp.Symbol,
        numeric_function: Callable[[Any], Any],
        lower_limit: float,
        upper_limit: float,
        step_size: float,
) -> tuple[
    str,
    sp.Expr | None,
    float | None,
    float | None,
    int | None,
    bool,
    tuple[str, ...],
]:
    """Check interval continuity and estimate the classical second-derivative bound."""

    warnings: list[str] = []
    interval_min = min(lower_limit, upper_limit)
    interval_max = max(lower_limit, upper_limit)
    requested_interval = sp.Interval(interval_min, interval_max)

    try:
        continuous_part = sp.calculus.util.continuous_domain(
            expression,
            x_symbol,
            requested_interval,
        )
    except NotImplementedError:
        warnings.append(
            "Symbolic continuity-domain analysis was unavailable; dense numerical "
            "interval checks were used instead."
        )
    except (TypeError, ValueError, NotImplementedError):
        warnings.append(
            "Symbolic continuity-domain analysis could not be completed; dense "
            "numerical interval checks were used instead."
        )
    else:
        if continuous_part != requested_interval:
            missing_part = requested_interval - continuous_part
            raise ValueError(
                "The integrand is not continuous over the complete integration "
                f"interval. Excluded point(s) or region(s): {missing_part}."
            )

    dense_x = np.linspace(interval_min, interval_max, CONTINUITY_SAMPLE_COUNT)
    dense_y = evaluate_real_array(numeric_function, dense_x)
    if np.any(~np.isfinite(dense_y)):
        bad_index = int(np.flatnonzero(~np.isfinite(dense_y))[0])
        raise ValueError(
            "The integrand is not finite throughout the integration interval. "
            f"A non-finite or complex value was detected near x = "
            f"{dense_x[bad_index]:.12g}."
        )

    continuity_status = (
        f"Passed symbolic/dense checks on [{interval_min:.12g}, {interval_max:.12g}]"
    )

    polynomial_degree: int | None = None
    exact_for_linear = False
    try:
        polynomial = sp.Poly(expression, x_symbol)
        polynomial_degree = int(polynomial.degree())
        exact_for_linear = polynomial_degree <= 1
    except (sp.PolynomialError, TypeError, ValueError):
        polynomial_degree = None

    second_derivative: sp.Expr | None = None
    sampled_max_second: float | None = None
    theoretical_error_bound: float | None = None

    try:
        second_derivative = sp.simplify(sp.diff(expression, x_symbol, 2))
        second_numeric = build_numeric_function(x_symbol, second_derivative)
        smooth_x = np.linspace(interval_min, interval_max, SMOOTHNESS_SAMPLE_COUNT)
        second_values = evaluate_real_array(second_numeric, smooth_x)
        if np.all(np.isfinite(second_values)):
            sampled_max_second = float(np.max(np.abs(second_values)))
            theoretical_error_bound = float(
                abs(upper_limit - lower_limit)
                * abs(step_size) ** 2
                * sampled_max_second
                / 12.0
            )
        else:
            warnings.append(
                "The second derivative is not finite across the interval, so the "
                "classical O(h^2) error bound is not available."
            )
    except (ValueError, TypeError, NotImplementedError, ArithmeticError, OverflowError):
        warnings.append(
            "The second derivative could not be evaluated reliably; the classical "
            "smooth-function error bound is omitted."
        )

    if exact_for_linear:
        warnings.append(
            "The integrand is a polynomial of degree at most one, so the "
            "Trapezoidal Rule is theoretically exact apart from floating-point round-off."
        )

    if sampled_max_second is not None and sampled_max_second == 0.0:
        exact_for_linear = True

    return (
        continuity_status,
        second_derivative,
        sampled_max_second,
        theoretical_error_bound,
        polynomial_degree,
        exact_for_linear,
        tuple(warnings),
    )


def calculate_reference_integral(
        expression: sp.Expr,
        x_symbol: sp.Symbol,
        numeric_function: Callable[[Any], Any],
        lower_limit: float,
        upper_limit: float,
) -> tuple[sp.Expr | None, float | None, str, str | None]:
    """Obtain a symbolic reference when possible, otherwise numerical quadrature."""

    antiderivative: sp.Expr | None = None
    symbolic_warning: str | None = None

    try:
        candidate = sp.integrate(expression, x_symbol)
        if not isinstance(candidate, sp.Integral) and not candidate.has(sp.Integral):
            antiderivative = candidate
            lower_symbolic = sp.Float(lower_limit, 40)
            upper_symbolic = sp.Float(upper_limit, 40)
            exact_expression = candidate.subs(
                x_symbol, upper_symbolic
            ) - candidate.subs(x_symbol, lower_symbolic)
            exact_numeric = sp.N(exact_expression, 40)
            if (
                    not exact_numeric.has(sp.zoo, sp.oo, -sp.oo, sp.nan)
                    and exact_numeric.is_real is not False
            ):
                exact_value = float(exact_numeric)
                if math.isfinite(exact_value):
                    return (
                        antiderivative,
                        exact_value,
                        "Closed-form symbolic antiderivative",
                        None,
                    )
    except (TypeError, ValueError, ArithmeticError, NotImplementedError, OverflowError):
        symbolic_warning = "A closed-form symbolic reference integral was not available."

    # High-resolution composite Simpson reference after continuity checks.
    try:
        reference_n = 8192
        x_values = np.linspace(lower_limit, upper_limit, reference_n + 1)
        y_values = evaluate_real_array(numeric_function, x_values)
        if np.any(~np.isfinite(y_values)):
            raise ValueError("non-finite values occurred in numerical reference sampling")
        h = (upper_limit - lower_limit) / reference_n
        reference_value = float(
            h
            / 3.0
            * (
                    y_values[0]
                    + y_values[-1]
                    + 4.0 * np.sum(y_values[1:-1:2])
                    + 2.0 * np.sum(y_values[2:-1:2])
            )
        )
        if not math.isfinite(reference_value):
            raise ValueError("the numerical reference was non-finite")
        warning = (
            "No closed-form antiderivative was found. Error metrics use a "
            "high-resolution numerical Simpson reference and should be interpreted "
            "as estimated rather than exact."
        )
        if symbolic_warning:
            warning = symbolic_warning + " " + warning
        return antiderivative, reference_value, "High-resolution numerical reference", warning
    except (ValueError, TypeError, ArithmeticError, OverflowError):
        return (
            antiderivative,
            None,
            "Not available",
            "Neither a reliable symbolic nor numerical reference integral could be obtained.",
        )


# =============================================================================
# Core Trapezoidal Rule calculations
# =============================================================================
def calculate_trapezoidal_approximation(
        numeric_function: Callable[[Any], Any],
        lower_limit: float,
        upper_limit: float,
        subintervals: int,
) -> tuple[
    tuple[FunctionSample, ...],
    tuple[TrapezoidContribution, ...],
    float,
    float,
    float,
    float,
    float,
]:
    """Apply the simple or composite Trapezoidal Rule manually."""

    if subintervals < 1:
        raise ValueError("The number of subintervals must be at least one.")

    step_size = (upper_limit - lower_limit) / subintervals
    if step_size == 0.0 or not math.isfinite(step_size):
        raise ValueError("The calculated step size is zero or non-finite.")

    x_values = [lower_limit + index * step_size for index in range(subintervals + 1)]
    x_values[-1] = upper_limit
    if any(not math.isfinite(value) for value in x_values):
        raise ValueError("At least one integration node is non-finite.")
    if any(x_values[index] == x_values[index + 1] for index in range(subintervals)):
        raise ValueError(
            "The integration nodes are numerically indistinguishable. Reduce n or "
            "rescale the interval."
        )

    function_values = [
        evaluate_real_scalar(
            numeric_function,
            x_value,
            f"f(x_{index})",
        )
        for index, x_value in enumerate(x_values)
    ]

    samples: list[FunctionSample] = []
    for index, (x_value, function_value) in enumerate(zip(x_values, function_values)):
        endpoint = index in (0, subintervals)
        weight = 1 if endpoint else 2
        samples.append(
            FunctionSample(
                index=index,
                x_value=float(x_value),
                function_value=float(function_value),
                weight=weight,
                weighted_value=float(weight * function_value),
                point_type="Endpoint" if endpoint else "Interior point",
            )
        )

    endpoint_sum = function_values[0] + function_values[-1]
    interior_sum = float(sum(function_values[1:-1]))
    weighted_sum = endpoint_sum + 2.0 * interior_sum
    approximate_integral = 0.5 * step_size * weighted_sum

    if not all(
            math.isfinite(value)
            for value in (endpoint_sum, interior_sum, weighted_sum, approximate_integral)
    ):
        raise ValueError("The Trapezoidal Rule produced a non-finite result.")

    contributions: list[TrapezoidContribution] = []
    cumulative_area = 0.0
    for index in range(subintervals):
        x_left = x_values[index]
        x_right = x_values[index + 1]
        f_left = function_values[index]
        f_right = function_values[index + 1]
        interval_width = x_right - x_left
        average_height = 0.5 * (f_left + f_right)
        signed_area = interval_width * average_height
        cumulative_area += signed_area

        formula_text = (
            f"A_{index + 1} = (x_{index + 1} - x_{index}) "
            f"[f(x_{index}) + f(x_{index + 1})] / 2"
        )
        substitution_text = (
            f"A_{index + 1} = ({x_right:.12g} - {x_left:.12g}) "
            f"[({f_left:.12g}) + ({f_right:.12g})] / 2 "
            f"= {signed_area:.12g}"
        )

        contributions.append(
            TrapezoidContribution(
                interval_index=index + 1,
                left_index=index,
                right_index=index + 1,
                x_left=float(x_left),
                x_right=float(x_right),
                interval_width=float(interval_width),
                f_left=float(f_left),
                f_right=float(f_right),
                average_height=float(average_height),
                signed_area=float(signed_area),
                absolute_geometric_area=float(abs(signed_area)),
                cumulative_signed_area=float(cumulative_area),
                formula_text=formula_text,
                substitution_text=substitution_text,
            )
        )

    panel_sum = float(sum(item.signed_area for item in contributions))
    consistency_scale = max(1.0, abs(panel_sum), abs(approximate_integral))
    if abs(panel_sum - approximate_integral) > 1.0e-11 * consistency_scale:
        raise ValueError("An internal consistency check failed for the trapezoid areas.")

    return (
        tuple(samples),
        tuple(contributions),
        float(step_size),
        float(endpoint_sum),
        float(interior_sum),
        float(weighted_sum),
        float(approximate_integral),
    )


def calculate_error_metrics(
        approximation: float,
        exact_value: float | None,
) -> tuple[float | None, float | None, float | None]:
    """Calculate absolute, relative, and percentage integration errors."""

    if exact_value is None:
        return None, None, None

    absolute_error = abs(exact_value - approximation)
    if abs(exact_value) <= RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
        return float(absolute_error), None, None

    relative_error = absolute_error / abs(exact_value)
    return float(absolute_error), float(relative_error), float(100.0 * relative_error)


def calculate_convergence_analysis(
        numeric_function: Callable[[Any], Any],
        lower_limit: float,
        upper_limit: float,
        initial_subintervals: int,
        levels: int,
        exact_integral: float | None,
) -> tuple[tuple[ConvergenceRecord, ...], float | None, tuple[str, ...]]:
    """Refine n by factors of two and estimate second-order convergence."""

    records: list[ConvergenceRecord] = []
    warnings: list[str] = []
    previous_approximation: float | None = None
    previous_absolute_error: float | None = None
    previous_successive_difference: float | None = None

    for level in range(levels):
        subintervals = initial_subintervals * (2 ** level)
        if subintervals > MAX_CONVERGENCE_SUBINTERVALS:
            warnings.append(
                "Convergence refinement stopped early because the number of "
                f"subintervals would exceed {MAX_CONVERGENCE_SUBINTERVALS:,}."
            )
            break

        (
            _samples,
            _contributions,
            step_size,
            _endpoint_sum,
            _interior_sum,
            _weighted_sum,
            approximation,
        ) = calculate_trapezoidal_approximation(
            numeric_function,
            lower_limit,
            upper_limit,
            subintervals,
        )

        absolute_error, relative_error, _percentage_error = calculate_error_metrics(
            approximation,
            exact_integral,
        )
        successive_difference = (
            None
            if previous_approximation is None
            else abs(approximation - previous_approximation)
        )
        estimated_error = (
            None if successive_difference is None else successive_difference / 3.0
        )
        richardson_extrapolated = (
            None
            if previous_approximation is None
            else approximation + (approximation - previous_approximation) / 3.0
        )

        observed_order: float | None = None
        if (
                exact_integral is not None
                and previous_absolute_error is not None
                and absolute_error is not None
                and previous_absolute_error > ZERO_TOLERANCE
                and absolute_error > ZERO_TOLERANCE
        ):
            observed_order = math.log(previous_absolute_error / absolute_error, 2.0)
        elif (
                previous_successive_difference is not None
                and successive_difference is not None
                and previous_successive_difference > ZERO_TOLERANCE
                and successive_difference > ZERO_TOLERANCE
        ):
            observed_order = math.log(
                previous_successive_difference / successive_difference,
                2.0,
            )

        records.append(
            ConvergenceRecord(
                level=level,
                subintervals=subintervals,
                step_size=float(step_size),
                approximation=float(approximation),
                exact_integral=exact_integral,
                absolute_error=absolute_error,
                relative_error=relative_error,
                successive_difference=successive_difference,
                estimated_error_from_refinement=estimated_error,
                richardson_extrapolated=richardson_extrapolated,
                observed_order=observed_order,
            )
        )

        if (
                previous_absolute_error is not None
                and absolute_error is not None
                and absolute_error > previous_absolute_error
        ):
            warnings.append(
                "A refined grid increased the reference error. Round-off, "
                "insufficient smoothness, or non-asymptotic behavior may be dominating."
            )
        elif (
                previous_successive_difference is not None
                and successive_difference is not None
                and successive_difference > previous_successive_difference
        ):
            warnings.append(
                "Successive refined approximations stopped improving monotonically. "
                "The finest levels may be affected by round-off."
            )

        previous_approximation = approximation
        previous_absolute_error = absolute_error
        previous_successive_difference = successive_difference

    latest_order = next(
        (
            record.observed_order
            for record in reversed(records)
            if record.observed_order is not None and math.isfinite(record.observed_order)
        ),
        None,
    )
    return tuple(records), latest_order, tuple(dict.fromkeys(warnings))


def solve_trapezoidal_rule(
        function_text: str,
        lower_limit_input: Any,
        upper_limit_input: Any,
        rule_name: str,
        subintervals_input: Any,
        convergence_levels_input: Any,
) -> TrapezoidalResult:
    """Validate all inputs and execute the complete Trapezoidal Rule workflow."""

    input_signature = create_input_signature(
        function_text,
        lower_limit_input,
        upper_limit_input,
        rule_name,
        subintervals_input,
        convergence_levels_input,
    )

    rule_key = RULE_OPTIONS.get(rule_name)
    if rule_key is None:
        return empty_result(
            message="The selected Trapezoidal Rule option is invalid.",
            stopping_reason="Input validation failed.",
            input_signature=input_signature,
        )

    try:
        lower_limit = safe_float(lower_limit_input, "Lower integration limit a")
        upper_limit = safe_float(upper_limit_input, "Upper integration limit b")
        convergence_levels = safe_positive_integer(
            convergence_levels_input,
            "Convergence levels",
            MIN_CONVERGENCE_LEVELS,
            MAX_CONVERGENCE_LEVELS,
        )
        if lower_limit == upper_limit:
            raise ValueError("The lower and upper integration limits must be different.")

        if rule_key == "simple":
            subintervals = 1
        else:
            subintervals = safe_positive_integer(
                subintervals_input,
                "Number of subintervals n",
                MIN_SUBINTERVALS,
                MAX_SUBINTERVALS,
            )

        x_symbol, function_expression = parse_function(function_text)
        numeric_function = build_numeric_function(x_symbol, function_expression)
        preview_step_size = (upper_limit - lower_limit) / subintervals

        (
            continuity_status,
            second_derivative,
            sampled_max_second,
            theoretical_error_bound,
            polynomial_degree,
            exact_for_linear,
            smoothness_warnings,
        ) = analyze_interval_and_smoothness(
            function_expression,
            x_symbol,
            numeric_function,
            lower_limit,
            upper_limit,
            preview_step_size,
        )

        (
            function_samples,
            interval_contributions,
            step_size,
            endpoint_sum,
            interior_sum,
            weighted_sum,
            approximate_integral,
        ) = calculate_trapezoidal_approximation(
            numeric_function,
            lower_limit,
            upper_limit,
            subintervals,
        )

        (
            antiderivative_expression,
            exact_integral,
            reference_source,
            reference_warning,
        ) = calculate_reference_integral(
            function_expression,
            x_symbol,
            numeric_function,
            lower_limit,
            upper_limit,
        )

        absolute_error, relative_error, percentage_error = calculate_error_metrics(
            approximate_integral,
            exact_integral,
        )

        convergence_records, latest_order, convergence_warnings = (
            calculate_convergence_analysis(
                numeric_function,
                lower_limit,
                upper_limit,
                subintervals,
                convergence_levels,
                exact_integral,
            )
        )
    except ValueError as error:
        return empty_result(
            message=str(error),
            stopping_reason=(
                "The Trapezoidal Rule was not executed because input, domain, "
                "or numerical validation failed."
            ),
            input_signature=input_signature,
            rule_key=rule_key,
            rule_name=rule_name,
        )

    warnings: list[str] = [*smoothness_warnings, *convergence_warnings]
    if reference_warning:
        warnings.append(reference_warning)
    if upper_limit < lower_limit:
        warnings.append(
            "The limits are reversed. The reported approximation is the signed integral."
        )
    if rule_key == "simple":
        warnings.append(
            "The simple Trapezoidal Rule uses only the two endpoints. The convergence "
            "study refines it using the composite Trapezoidal Rule with n = 1, 2, 4, ... ."
        )
    if latest_order is not None and latest_order < 1.4:
        warnings.append(
            "The latest observed order is below the theoretical second order. "
            "The integrand may be insufficiently smooth, the grid may not yet be "
            "in the asymptotic range, or floating-point effects may be significant."
        )
    if (
            absolute_error is not None
            and theoretical_error_bound is not None
            and absolute_error > 1.05 * theoretical_error_bound
    ):
        warnings.append(
            "The sampled second-derivative bound is smaller than the measured "
            "reference error. Increase the smoothness sampling density or inspect "
            "the integrand for rapid variation."
        )

    final_refinement_error_estimate = (
        convergence_records[-1].estimated_error_from_refinement
        if convergence_records
        else None
    )
    final_richardson_estimate = (
        convergence_records[-1].richardson_extrapolated
        if convergence_records
        else None
    )
    convergence_rule_name = (
        "Composite Trapezoidal refinement beginning with n = 1"
        if rule_key == "simple"
        else "Composite Trapezoidal refinement"
    )

    return TrapezoidalResult(
        status="success",
        success=True,
        method=METHOD_NAME,
        rule_key=rule_key,
        rule_name=rule_name,
        message="Execution completed successfully.",
        stopping_reason=(
            "The definite integral was approximated after all trapezoidal "
            "subinterval contributions and diagnostics were evaluated."
        ),
        function_text=function_text.strip().replace("^", "**"),
        function_expression=function_expression,
        antiderivative_expression=antiderivative_expression,
        antiderivative_text=(
            str(antiderivative_expression)
            if antiderivative_expression is not None
            else "Not available"
        ),
        reference_integral_source=reference_source,
        lower_limit=lower_limit,
        upper_limit=upper_limit,
        orientation="Forward" if upper_limit > lower_limit else "Reversed",
        subintervals=subintervals,
        step_size=step_size,
        function_samples=function_samples,
        interval_contributions=interval_contributions,
        endpoint_sum=endpoint_sum,
        interior_sum=interior_sum,
        weighted_sum=weighted_sum,
        approximate_integral=approximate_integral,
        exact_integral=exact_integral,
        absolute_error=absolute_error,
        relative_error=relative_error,
        percentage_error=percentage_error,
        theoretical_order=2,
        second_derivative_expression=second_derivative,
        second_derivative_text=(
            str(second_derivative) if second_derivative is not None else "Not available"
        ),
        sampled_max_second_derivative=sampled_max_second,
        theoretical_error_bound=theoretical_error_bound,
        polynomial_degree=polynomial_degree,
        exact_for_linear=exact_for_linear,
        interval_continuity_status=continuity_status,
        convergence_rule_name=convergence_rule_name,
        convergence_records=convergence_records,
        latest_observed_order=latest_order,
        final_refinement_error_estimate=final_refinement_error_estimate,
        final_richardson_estimate=final_richardson_estimate,
        warnings=tuple(dict.fromkeys(warnings)),
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


# =============================================================================
# DataFrame builders
# =============================================================================
def create_function_values_dataframe(result: TrapezoidalResult) -> pd.DataFrame:
    """Create the complete node and weight table."""

    return pd.DataFrame(
        {
            "i": [sample.index for sample in result.function_samples],
            "x_i": [sample.x_value for sample in result.function_samples],
            "f(x_i)": [sample.function_value for sample in result.function_samples],
            "Weight": [sample.weight for sample in result.function_samples],
            "Weighted Value": [sample.weighted_value for sample in result.function_samples],
            "Point Type": [sample.point_type for sample in result.function_samples],
        }
    )


def create_interval_dataframe(result: TrapezoidalResult) -> pd.DataFrame:
    """Create the table of individual trapezoid contributions."""

    return pd.DataFrame(
        {
            "Interval": [item.interval_index for item in result.interval_contributions],
            "Left Node": [item.left_index for item in result.interval_contributions],
            "Right Node": [item.right_index for item in result.interval_contributions],
            "x_left": [item.x_left for item in result.interval_contributions],
            "x_right": [item.x_right for item in result.interval_contributions],
            "Width": [item.interval_width for item in result.interval_contributions],
            "f(x_left)": [item.f_left for item in result.interval_contributions],
            "f(x_right)": [item.f_right for item in result.interval_contributions],
            "Average Height": [item.average_height for item in result.interval_contributions],
            "Signed Area": [item.signed_area for item in result.interval_contributions],
            "Absolute Geometric Area": [
                item.absolute_geometric_area for item in result.interval_contributions
            ],
            "Cumulative Signed Area": [
                item.cumulative_signed_area for item in result.interval_contributions
            ],
            "Operation": [item.substitution_text for item in result.interval_contributions],
        }
    )


def create_convergence_dataframe(result: TrapezoidalResult) -> pd.DataFrame:
    """Create the refinement and convergence table."""

    return pd.DataFrame(
        {
            "Level": [record.level for record in result.convergence_records],
            "Subintervals n": [record.subintervals for record in result.convergence_records],
            "Step Size h": [record.step_size for record in result.convergence_records],
            "Approximate Integral": [record.approximation for record in result.convergence_records],
            "Exact/Reference Integral": [
                record.exact_integral for record in result.convergence_records
            ],
            "Absolute Error": [record.absolute_error for record in result.convergence_records],
            "Relative Error": [record.relative_error for record in result.convergence_records],
            "Successive Difference": [
                record.successive_difference for record in result.convergence_records
            ],
            "Estimated Refined Error |T_2n-T_n|/3": [
                record.estimated_error_from_refinement
                for record in result.convergence_records
            ],
            "Richardson Extrapolated (4T_2n-T_n)/3": [
                record.richardson_extrapolated for record in result.convergence_records
            ],
            "Observed Order": [record.observed_order for record in result.convergence_records],
        }
    )


def create_error_dataframe(result: TrapezoidalResult) -> pd.DataFrame:
    """Create a compact error-analysis table."""

    return pd.DataFrame(
        {
            "Metric": [
                "Approximate Integral",
                "Exact/Reference Integral",
                "Reference Source",
                "Absolute Error",
                "Relative Error",
                "Percentage Error",
                "Theoretical Order",
                "Sampled Error Bound",
                "Final Refinement Error Estimate",
                "Final Richardson Extrapolated Value",
                "Latest Observed Order",
            ],
            "Value": [
                result.approximate_integral,
                result.exact_integral,
                result.reference_integral_source,
                result.absolute_error,
                result.relative_error,
                result.percentage_error,
                result.theoretical_order,
                result.theoretical_error_bound,
                result.final_refinement_error_estimate,
                result.final_richardson_estimate,
                result.latest_observed_order,
            ],
        }
    )


def create_smoothness_dataframe(result: TrapezoidalResult) -> pd.DataFrame:
    """Create continuity and second-derivative diagnostics."""

    return pd.DataFrame(
        {
            "Diagnostic": [
                "Interval Continuity Status",
                "Second Derivative f''(x)",
                "Sampled max |f''(x)|",
                "Classical Error Bound",
                "Polynomial Degree",
                "Theoretically Exact for Linear Polynomial",
            ],
            "Value": [
                result.interval_continuity_status,
                result.second_derivative_text,
                result.sampled_max_second_derivative,
                result.theoretical_error_bound,
                result.polynomial_degree,
                "Yes" if result.exact_for_linear else "No",
            ],
        }
    )


def create_summary_dataframe(result: TrapezoidalResult) -> pd.DataFrame:
    """Create the Summary worksheet table."""

    return pd.DataFrame(
        {
            "Property": [
                "Method",
                "Rule",
                "Status",
                "Function",
                "Antiderivative",
                "Reference Integral Source",
                "Lower Limit a",
                "Upper Limit b",
                "Orientation",
                "Number of Subintervals",
                "Step Size h",
                "Endpoint Sum",
                "Interior Sum",
                "Weighted Sum",
                "Approximate Integral",
                "Exact/Reference Integral",
                "Absolute Error",
                "Relative Error",
                "Percentage Error",
                "Theoretical Order",
                "Second Derivative",
                "Sampled max |f''|",
                "Theoretical Error Bound",
                "Polynomial Degree",
                "Interval Continuity Status",
                "Convergence Rule",
                "Final Refinement Error Estimate",
                "Final Richardson Estimate",
                "Latest Observed Order",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.rule_name,
                result.status,
                result.function_text,
                result.antiderivative_text,
                result.reference_integral_source,
                result.lower_limit,
                result.upper_limit,
                result.orientation,
                result.subintervals,
                result.step_size,
                result.endpoint_sum,
                result.interior_sum,
                result.weighted_sum,
                result.approximate_integral,
                result.exact_integral,
                result.absolute_error,
                result.relative_error,
                result.percentage_error,
                result.theoretical_order,
                result.second_derivative_text,
                result.sampled_max_second_derivative,
                result.theoretical_error_bound,
                result.polynomial_degree,
                result.interval_continuity_status,
                result.convergence_rule_name,
                result.final_refinement_error_estimate,
                result.final_richardson_estimate,
                result.latest_observed_order,
                " | ".join(result.warnings) if result.warnings else "None",
                result.stopping_reason,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )


def create_formula_dataframe(result: TrapezoidalResult) -> pd.DataFrame:
    """Create a workbook table containing formulas and assumptions."""

    return pd.DataFrame(
        {
            "Item": [
                "Simple Trapezoidal Rule",
                "Composite Trapezoidal Rule",
                "Step Size",
                "Composite Error Form",
                "Classical Magnitude Bound",
                "Refined-Grid Error Estimate",
                "Richardson Extrapolation",
                "Theoretical Global Order",
            ],
            "Expression": [
                "T = (b-a)[f(a)+f(b)]/2",
                "T_n = h/2 [f(x0) + 2*sum(f(x_i), i=1..n-1) + f(x_n)]",
                "h = (b-a)/n",
                "E_t = -(b-a)h^2 f''(xi)/12",
                "|E_t| <= |b-a| h^2 max|f''(x)| / 12",
                "|E_2n| approximately |T_2n - T_n| / 3",
                "I approximately (4T_2n - T_n) / 3",
                "O(h^2) for a sufficiently smooth integrand",
            ],
        }
    )


# =============================================================================
# Scientific plots
# =============================================================================
def create_trapezoid_plot(result: TrapezoidalResult) -> Figure:
    """Create a function graph with all trapezoids drawn explicitly."""

    if not result.success or result.function_expression is None:
        raise ValueError("A successful result is required for plotting.")

    x_symbol = sp.Symbol("x", real=True)
    numeric_function = build_numeric_function(x_symbol, result.function_expression)
    lower_limit = float(result.lower_limit)
    upper_limit = float(result.upper_limit)
    plotting_min = min(lower_limit, upper_limit)
    plotting_max = max(lower_limit, upper_limit)
    span = plotting_max - plotting_min
    margin = 0.08 * span if span > 0 else 1.0

    dense_x = np.linspace(plotting_min - margin, plotting_max + margin, 1200)
    dense_y = evaluate_real_array(numeric_function, dense_x)
    finite_mask = np.isfinite(dense_y)
    if np.count_nonzero(finite_mask) < 2:
        raise ValueError("The function has insufficient finite values for plotting.")

    figure, axis = plt.subplots(figsize=(11, 6.5))
    axis.plot(dense_x[finite_mask], dense_y[finite_mask], linewidth=2.2, label="f(x)")
    axis.axhline(0.0, linewidth=1.0)

    for contribution in result.interval_contributions:
        polygon_x = [
            contribution.x_left,
            contribution.x_left,
            contribution.x_right,
            contribution.x_right,
        ]
        polygon_y = [0.0, contribution.f_left, contribution.f_right, 0.0]
        axis.fill(polygon_x, polygon_y, alpha=0.22, edgecolor="black", linewidth=0.7)
        axis.plot(
            [contribution.x_left, contribution.x_right],
            [contribution.f_left, contribution.f_right],
            linewidth=1.35,
        )

    node_x = np.array([item.x_value for item in result.function_samples], dtype=float)
    node_y = np.array([item.function_value for item in result.function_samples], dtype=float)
    axis.scatter(node_x, node_y, s=38, zorder=5, label="Integration Nodes")
    axis.axvline(lower_limit, linestyle="--", linewidth=1.0, label=f"a = {lower_limit:.6g}")
    axis.axvline(upper_limit, linestyle="--", linewidth=1.0, label=f"b = {upper_limit:.6g}")

    axis.set_title(f"{result.rule_name}: {result.subintervals} Subinterval(s)")
    axis.set_xlabel("x")
    axis.set_ylabel("f(x)")
    axis.grid(True, alpha=0.28)
    axis.legend()
    figure.tight_layout()
    return figure


def create_convergence_plot(result: TrapezoidalResult) -> Figure:
    """Create a log-log convergence visualization with an O(h^2) reference."""

    if len(result.convergence_records) < 2:
        raise ValueError("At least two refinement levels are required for plotting.")

    step_sizes = np.array(
        [abs(record.step_size) for record in result.convergence_records],
        dtype=float,
    )
    if result.exact_integral is not None:
        errors = np.array(
            [
                record.absolute_error if record.absolute_error is not None else np.nan
                for record in result.convergence_records
            ],
            dtype=float,
        )
        y_label = "Absolute Error"
    else:
        errors = np.array(
            [
                record.successive_difference
                if record.successive_difference is not None
                else np.nan
                for record in result.convergence_records
            ],
            dtype=float,
        )
        y_label = "Successive Difference"

    valid = (
            np.isfinite(step_sizes)
            & np.isfinite(errors)
            & (step_sizes > 0.0)
            & (errors > 0.0)
    )
    if np.count_nonzero(valid) < 2:
        raise ValueError("There are not enough positive finite error indicators for plotting.")

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.loglog(step_sizes[valid], errors[valid], marker="o", linewidth=2.0, label=y_label)
    reference_h = step_sizes[valid]
    reference_error = errors[valid][0] * (reference_h / reference_h[0]) ** 2
    axis.loglog(
        reference_h,
        reference_error,
        linestyle="--",
        linewidth=1.4,
        label="Reference slope 2",
    )
    axis.invert_xaxis()
    axis.set_title("Trapezoidal Rule Convergence Analysis")
    axis.set_xlabel("Step Size |h|")
    axis.set_ylabel(y_label)
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def figure_to_png_bytes(figure: Figure) -> bytes:
    """Serialize one Matplotlib figure as a PNG image."""

    output = BytesIO()
    figure.savefig(output, format="png", dpi=160, bbox_inches="tight")
    output.seek(0)
    return output.getvalue()


# =============================================================================
# Excel report
# =============================================================================
def apply_excel_style(workbook: Any) -> None:
    """Apply consistent professional formatting to all worksheets."""

    header_fill = PatternFill("solid", fgColor="0D3151")
    header_font = Font(bold=True, color="FFFFFF")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            maximum_length = 0
            for cell in column_cells:
                value_length = len(str(cell.value)) if cell.value is not None else 0
                maximum_length = max(maximum_length, value_length)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(maximum_length + 2, 12),
                62,
            )


def add_excel_image(
        worksheet: Any,
        image_bytes: bytes,
        anchor: str,
        width: int = 900,
        height: int = 520,
) -> None:
    """Insert one PNG image into a worksheet."""

    image = ExcelImage(BytesIO(image_bytes))
    image.width = width
    image.height = height
    worksheet.add_image(image, anchor)


def create_plot_data_dataframe(result: TrapezoidalResult) -> pd.DataFrame:
    """Create aligned chart data for editable Excel charts."""

    interval_df = create_interval_dataframe(result)
    convergence_df = create_convergence_dataframe(result)
    row_count = max(len(interval_df), len(convergence_df), 1)
    dataframe = pd.DataFrame(index=range(row_count))
    dataframe["Interval"] = pd.Series(interval_df.get("Interval", pd.Series(dtype=float)))
    dataframe["Signed Area"] = pd.Series(interval_df.get("Signed Area", pd.Series(dtype=float)))
    dataframe["Cumulative Signed Area"] = pd.Series(
        interval_df.get("Cumulative Signed Area", pd.Series(dtype=float))
    )
    dataframe["Level"] = pd.Series(convergence_df.get("Level", pd.Series(dtype=float)))
    dataframe["Subintervals n"] = pd.Series(
        convergence_df.get("Subintervals n", pd.Series(dtype=float))
    )
    dataframe["|h|"] = pd.Series(
        np.abs(convergence_df.get("Step Size h", pd.Series(dtype=float)))
    )
    dataframe["Approximate Integral"] = pd.Series(
        convergence_df.get("Approximate Integral", pd.Series(dtype=float))
    )
    error_column = (
        "Absolute Error"
        if result.exact_integral is not None
        else "Successive Difference"
    )
    dataframe["Error Indicator"] = pd.Series(
        convergence_df.get(error_column, pd.Series(dtype=float))
    )
    dataframe["Richardson Extrapolated"] = pd.Series(
        convergence_df.get(
            "Richardson Extrapolated (4T_2n-T_n)/3",
            pd.Series(dtype=float),
        )
    )
    return dataframe


def create_excel_report(result: TrapezoidalResult) -> bytes:
    """Generate a formatted in-memory XLSX report with data, charts, and plots."""

    if not result.success:
        raise ValueError("Only a successful result can be exported.")

    summary_df = create_summary_dataframe(result)
    formulas_df = create_formula_dataframe(result)
    values_df = create_function_values_dataframe(result)
    intervals_df = create_interval_dataframe(result)
    smoothness_df = create_smoothness_dataframe(result)
    error_df = create_error_dataframe(result)
    convergence_df = create_convergence_dataframe(result)
    plot_data_df = create_plot_data_dataframe(result)

    trapezoid_figure = create_trapezoid_plot(result)
    trapezoid_png = figure_to_png_bytes(trapezoid_figure)
    plt.close(trapezoid_figure)
    try:
        convergence_figure = create_convergence_plot(result)
    except ValueError:
        convergence_png = None
    else:
        convergence_png = figure_to_png_bytes(convergence_figure)
        plt.close(convergence_figure)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        formulas_df.to_excel(writer, sheet_name="Method Formulas", index=False)
        values_df.to_excel(writer, sheet_name="Function Values", index=False)
        intervals_df.to_excel(writer, sheet_name="Interval Areas", index=False)
        smoothness_df.to_excel(writer, sheet_name="Smoothness Diagnostics", index=False)
        error_df.to_excel(writer, sheet_name="Error Analysis", index=False)
        convergence_df.to_excel(writer, sheet_name="Convergence Analysis", index=False)
        plot_data_df.to_excel(writer, sheet_name="Plot Data", index=False)

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Trapezoidal Rule Scientific Plots"
        plots_sheet["A1"].font = Font(bold=True, size=16)
        add_excel_image(plots_sheet, trapezoid_png, "A3")
        if convergence_png is not None:
            add_excel_image(plots_sheet, convergence_png, "A31")

        apply_excel_style(workbook)

        summary_sheet = workbook["Summary"]
        plot_sheet = workbook["Plot Data"]
        interval_rows = len(intervals_df) + 1
        convergence_rows = len(convergence_df) + 1

        if len(intervals_df) > 0:
            area_chart = BarChart()
            area_chart.title = "Trapezoid Signed Contributions"
            area_chart.x_axis.title = "Interval"
            area_chart.y_axis.title = "Signed Area"
            area_chart.height = 8
            area_chart.width = 15
            area_chart.add_data(
                Reference(plot_sheet, min_col=2, min_row=1, max_row=interval_rows),
                titles_from_data=True,
            )
            area_chart.set_categories(
                Reference(plot_sheet, min_col=1, min_row=2, max_row=interval_rows)
            )
            summary_sheet.add_chart(area_chart, "D2")

        if len(convergence_df) > 0:
            approximation_chart = LineChart()
            approximation_chart.title = "Approximation by Mesh Refinement"
            approximation_chart.x_axis.title = "Subintervals n"
            approximation_chart.y_axis.title = "Integral"
            approximation_chart.height = 8
            approximation_chart.width = 15
            approximation_chart.add_data(
                Reference(plot_sheet, min_col=7, min_row=1, max_row=convergence_rows),
                titles_from_data=True,
            )
            approximation_chart.set_categories(
                Reference(plot_sheet, min_col=5, min_row=2, max_row=convergence_rows)
            )
            summary_sheet.add_chart(approximation_chart, "D21")

            error_chart = LineChart()
            error_chart.title = "Convergence Error Indicator"
            error_chart.x_axis.title = "Subintervals n"
            error_chart.y_axis.title = "Magnitude"
            error_chart.height = 8
            error_chart.width = 15
            error_chart.add_data(
                Reference(plot_sheet, min_col=8, min_row=1, max_row=convergence_rows),
                titles_from_data=True,
            )
            error_chart.set_categories(
                Reference(plot_sheet, min_col=5, min_row=2, max_row=convergence_rows)
            )
            summary_sheet.add_chart(error_chart, "D40")

        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit result rendering
# =============================================================================
def render_final_result(result: TrapezoidalResult) -> None:
    """Render the final-result card."""

    if not result.success:
        st.error(result.message)
        st.caption(result.stopping_reason)
        return

    st.success(result.message)
    st.markdown(f"**Function:** `{result.function_text}`")
    st.markdown(f"**Rule:** {result.rule_name}")
    st.markdown(
        "**Interval:** "
        f"[{format_number(result.lower_limit)}, {format_number(result.upper_limit)}]"
    )

    metric_columns = st.columns(2)
    metric_columns[0].metric(
        "Approximate Integral",
        format_number(result.approximate_integral),
    )
    metric_columns[1].metric(
        "Reference Integral",
        format_number(result.exact_integral),
    )

    detail_columns = st.columns(2)
    detail_columns[0].metric("Subintervals", str(result.subintervals))
    detail_columns[1].metric("Step Size h", format_number(result.step_size))

    if result.absolute_error is not None:
        error_columns = st.columns(2)
        error_columns[0].metric("Absolute Error", scientific_number(result.absolute_error))
        error_columns[1].metric(
            "Relative Error",
            "Undefined (reference is zero)"
            if result.relative_error is None
            else f"{format_number(result.percentage_error)}%",
        )

    if result.final_refinement_error_estimate is not None:
        refinement_columns = st.columns(2)
        refinement_columns[0].metric(
            "Refined Error Estimate",
            scientific_number(result.final_refinement_error_estimate),
        )
        refinement_columns[1].metric(
            "Richardson Value",
            format_number(result.final_richardson_estimate),
        )

    st.caption(f"Reference source: {result.reference_integral_source}")
    st.caption(result.stopping_reason)
    for warning in result.warnings:
        st.warning(warning)


def render_formula_summary(result: TrapezoidalResult) -> None:
    """Render the complete numerical substitution used by the method."""

    st.subheader("Trapezoidal Formula Calculation")
    if result.subintervals == 1:
        st.latex(r"T = \frac{b-a}{2}\,[f(a)+f(b)]")
    else:
        st.latex(
            r"T_n = \frac{h}{2}\left[f(x_0)+2\sum_{i=1}^{n-1}f(x_i)+f(x_n)\right]"
        )

    formula_df = pd.DataFrame(
        {
            "Quantity": [
                "h = (b-a)/n",
                "Endpoint Sum",
                "Interior Sum",
                "Weighted Sum",
                "T_n = (h/2)(Weighted Sum)",
            ],
            "Value": [
                result.step_size,
                result.endpoint_sum,
                result.interior_sum,
                result.weighted_sum,
                result.approximate_integral,
            ],
        }
    )
    st.dataframe(
        round_numeric_dataframe(formula_df, 12),
        use_container_width=True,
        hide_index=True,
    )
    st.code(
        "T_n = "
        f"({result.step_size:.12g}/2) × "
        f"[{result.endpoint_sum:.12g} + 2({result.interior_sum:.12g})] "
        f"= {result.approximate_integral:.12g}",
        language=None,
    )


def render_smoothness_diagnostics(result: TrapezoidalResult) -> None:
    """Render interval, derivative, and theoretical-error diagnostics."""

    st.subheader("Continuity and Error-Bound Diagnostics")
    st.dataframe(
        round_numeric_dataframe(create_smoothness_dataframe(result), 12),
        use_container_width=True,
        hide_index=True,
    )
    st.latex(
        r"|E_t|\leq\frac{|b-a|h^2}{12}\max_{x\in[a,b]}|f''(x)|"
    )
    if result.theoretical_error_bound is None:
        st.info(
            "The classical second-derivative error bound is unavailable for this "
            "integrand, but the numerical refinement study is still reported."
        )


def render_function_values(result: TrapezoidalResult) -> None:
    """Render all nodes, values, and textbook weights."""

    st.subheader("Function Values and Weights")
    st.dataframe(
        round_numeric_dataframe(create_function_values_dataframe(result), 10),
        use_container_width=True,
        hide_index=True,
    )


def render_interval_steps(result: TrapezoidalResult) -> None:
    """Render every trapezoid contribution and detailed substitution."""

    st.subheader("Trapezoid Contributions")
    st.dataframe(
        round_numeric_dataframe(create_interval_dataframe(result), 10),
        use_container_width=True,
        hide_index=True,
    )

    for item in result.interval_contributions:
        with st.expander(
                f"Interval {item.interval_index}: "
                f"[{format_number(item.x_left, 8)}, {format_number(item.x_right, 8)}]",
                expanded=(item.interval_index == 1),
        ):
            st.markdown(f"**Formula:** `{item.formula_text}`")
            st.code(item.substitution_text, language=None)
            detail_df = pd.DataFrame(
                {
                    "Property": [
                        "Interval Width",
                        "Left Height",
                        "Right Height",
                        "Average Height",
                        "Signed Area",
                        "Absolute Geometric Area",
                        "Cumulative Signed Area",
                    ],
                    "Value": [
                        item.interval_width,
                        item.f_left,
                        item.f_right,
                        item.average_height,
                        item.signed_area,
                        item.absolute_geometric_area,
                        item.cumulative_signed_area,
                    ],
                }
            )
            st.dataframe(
                round_numeric_dataframe(detail_df, 12),
                use_container_width=True,
                hide_index=True,
            )


def render_error_analysis(result: TrapezoidalResult) -> None:
    """Render reference error and refinement estimates."""

    st.subheader("Error Analysis")
    st.dataframe(
        round_numeric_dataframe(create_error_dataframe(result), 12),
        use_container_width=True,
        hide_index=True,
    )

    if result.exact_integral is None:
        st.info(
            "A reliable symbolic or numerical reference integral was unavailable. "
            "Convergence is assessed using successive refined approximations."
        )
    else:
        st.markdown(f"**Reference source:** {result.reference_integral_source}")
        if result.antiderivative_expression is not None:
            st.markdown(f"**Reference antiderivative:** `{result.antiderivative_text}`")


def render_convergence_analysis(result: TrapezoidalResult) -> None:
    """Render the complete mesh-refinement study."""

    st.subheader("Convergence Analysis")
    convergence_df = create_convergence_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(convergence_df, 12),
        use_container_width=True,
        hide_index=True,
    )

    st.info(
        f"Convergence mode: {result.convergence_rule_name}. For a sufficiently "
        "smooth integrand, halving h should reduce the composite Trapezoidal "
        "Rule error by approximately four. The refined-grid error estimate is "
        "|T₂ₙ − Tₙ|/3, and (4T₂ₙ − Tₙ)/3 is the Richardson-improved value."
    )

    if result.latest_observed_order is not None:
        st.markdown(
            "**Latest observed convergence order:** "
            f"{format_number(result.latest_observed_order, 6)}"
        )

    try:
        figure = create_convergence_plot(result)
    except ValueError as error:
        st.info(str(error))
    else:
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)


def render_function_graph(result: TrapezoidalResult) -> None:
    """Render the integrand and all geometric trapezoids."""

    st.subheader("Function and Trapezoid Visualization")
    try:
        figure = create_trapezoid_plot(result)
    except ValueError as error:
        st.warning(f"The graph could not be displayed. {error}")
    else:
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)


def render_excel_download(result: TrapezoidalResult) -> None:
    """Create and render the Excel report download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature
    cached_signature = st.session_state.get("trapezoidal_excel_signature")

    if cached_signature != report_signature:
        try:
            report_bytes = create_excel_report(result)
        except (ValueError, OSError, RuntimeError, TypeError, ArithmeticError) as error:
            st.error(f"The Excel report could not be generated. {error}")
            return
        st.session_state.trapezoidal_excel_report = report_bytes
        st.session_state.trapezoidal_excel_signature = report_signature

    report_bytes = st.session_state.get("trapezoidal_excel_report")
    if report_bytes is None:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=f"trapezoidal_rule_report_{timestamp}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="trapezoidal_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Trapezoidal Rule Streamlit page."""

    st.set_page_config(
        page_title="Trapezoidal Rule Solver | Numerical Methods",
        page_icon="∫",
        layout="wide",
    )
    load_css()

    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">NUMERICAL INTEGRATION TOOL</div>
                <h1>Trapezoidal Rule Solver</h1>
                <p>
                    Approximate a definite integral using the simple or composite
                    Trapezoidal Rule. Review every function value, weighted term,
                    individual trapezoid contribution, error metric, convergence
                    result, scientific graph, and Excel report.
                </p>

                <div class="method-actions">
                    <a href="/Trapezoidal_Rule" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Trapezoidal_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Match the centered Bisection solver content width.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader('How to Write the Function')
                st.markdown(
                    """
                Enter only the mathematical expression, without `f(x) =` or an equals sign.

                - Use only **x** as the variable.
                - Powers: write `x**2`, not **x^2**.
                - Multiplication: write `2*x`, not `2x`.
                - Use lowercase functions such as **sin(x)**, **cos(x)**, **exp(x)**, **sqrt(x)**, and **log(x)**.
                - Use parentheses whenever the order of operations could be unclear.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader('Before Solving')
                st.markdown(
                    """
                - The lower and upper limits must be different.
                - The function must be continuous and finite throughout the complete integration interval, not only at the selected nodes.
                - The number of subintervals must be a positive integer.
                - The composite rule is globally second-order accurate for sufficiently smooth functions.
                - Reversing the limits correctly changes the sign of the integral.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">Function f(x)</div>',
                    unsafe_allow_html=True,
                )
                function_text = st.text_input(
                    "Function f(x)",
                    value=DEFAULT_FUNCTION,
                    placeholder="Example: sin(x), exp(x), or x**3 - 2*x + 1",
                    label_visibility="collapsed",
                    key="trapezoidal_function",
                )

                rule_name = st.selectbox(
                    "Rule Type",
                    options=list(RULE_OPTIONS.keys()),
                    index=0,
                    key="trapezoidal_rule_type",
                )

                limit_columns = st.columns(2)
                with limit_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Lower limit a</div>',
                        unsafe_allow_html=True,
                    )
                    lower_limit = st.number_input(
                        "Lower limit a",
                        value=DEFAULT_LOWER_LIMIT,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="trapezoidal_lower_limit",
                    )

                with limit_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Upper limit b</div>',
                        unsafe_allow_html=True,
                    )
                    upper_limit = st.number_input(
                        "Upper limit b",
                        value=DEFAULT_UPPER_LIMIT,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="trapezoidal_upper_limit",
                    )

                settings_columns = st.columns(2)
                with settings_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Subintervals n</div>',
                        unsafe_allow_html=True,
                    )
                    subintervals = st.number_input(
                        "Subintervals n",
                        min_value=MIN_SUBINTERVALS,
                        max_value=MAX_SUBINTERVALS,
                        value=(1 if RULE_OPTIONS[rule_name] == "simple" else DEFAULT_SUBINTERVALS),
                        step=1,
                        disabled=(RULE_OPTIONS[rule_name] == "simple"),
                        label_visibility="collapsed",
                        key="trapezoidal_subintervals",
                    )

                with settings_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Convergence levels</div>',
                        unsafe_allow_html=True,
                    )
                    convergence_levels = st.number_input(
                        "Convergence levels",
                        min_value=MIN_CONVERGENCE_LEVELS,
                        max_value=MAX_CONVERGENCE_LEVELS,
                        value=DEFAULT_CONVERGENCE_LEVELS,
                        step=1,
                        label_visibility="collapsed",
                        key="trapezoidal_convergence_levels",
                    )

                current_signature = create_input_signature(
                    function_text,
                    lower_limit,
                    upper_limit,
                    rule_name,
                    subintervals,
                    convergence_levels,
                )

                solve_clicked = st.button(
                    "Calculate Integral",
                    type="primary",
                    use_container_width=True,
                    key="trapezoidal_solve_button",
                )

                if solve_clicked:
                    result = solve_trapezoidal_rule(
                        function_text=function_text,
                        lower_limit_input=lower_limit,
                        upper_limit_input=upper_limit,
                        rule_name=rule_name,
                        subintervals_input=subintervals,
                        convergence_levels_input=convergence_levels,
                    )
                    st.session_state.trapezoidal_result = result
                    st.session_state.trapezoidal_result_signature = current_signature
                    st.session_state.pop("trapezoidal_excel_report", None)
                    st.session_state.pop("trapezoidal_excel_signature", None)

                with st.expander("Example Inputs"):
                    st.code(
                        "Function: sin(x)\n"
                        "a = 0\n"
                        "b = pi ≈ 3.141592653589793\n"
                        "n = 8\n"
                        "Expected exact integral: 2",
                        language=None,
                    )
                    st.code(
                        "Function: x**3 - 2*x + 1\n"
                        "a = -1\n"
                        "b = 2\n"
                        "n = 6",
                        language=None,
                    )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                saved_result = st.session_state.get("trapezoidal_result")
                saved_signature = st.session_state.get("trapezoidal_result_signature")

                if saved_result is None:
                    st.info("Enter the integration data and select Calculate Integral.")
                elif saved_signature != current_signature:
                    st.warning(
                        "The inputs have changed. Select Calculate Integral to update the result."
                    )
                else:
                    render_final_result(saved_result)

        saved_result = st.session_state.get("trapezoidal_result")
        saved_signature = st.session_state.get("trapezoidal_result_signature")

        if saved_result is not None and saved_signature == current_signature:
            if saved_result.success:
                st.divider()
                render_formula_summary(saved_result)
                st.divider()
                render_smoothness_diagnostics(saved_result)
                st.divider()
                render_function_values(saved_result)
                st.divider()
                render_interval_steps(saved_result)
                st.divider()
                render_function_graph(saved_result)
                st.divider()
                render_error_analysis(saved_result)
                st.divider()
                render_convergence_analysis(saved_result)
                st.divider()
                render_excel_download(saved_result)

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Trapezoidal Rule • Numerical Integration</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()