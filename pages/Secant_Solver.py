from __future__ import annotations

import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any, Callable
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css


# =============================================================================
# Method configuration
# =============================================================================
CONFIG = {
    "method_id": "secant",
    "title": "Secant Solver",
    "label": "SECANT METHOD TOOL",
    "description": (
        "Enter a nonlinear function and two starting values to approximate a "
        "root from successive secant-line x-intercepts without evaluating an "
        "analytical derivative during the iteration."
    ),
    "lesson": "Secant_Method",
    "quiz": "Secant_Quiz",
    "footer": "Secant Solver • Root Finding",
    "default_eq": "x**3 - x - 2",
    "conditions": [
        "Two numerically distinct starting values are required.",
        "The function must remain finite at every approximation.",
        "Consecutive function values must define a usable secant slope.",
        "The method is open and does not preserve a root bracket.",
    ],
    "formula": (
        r"x_{n+1}=x_n-"
        r"\frac{f(x_n)(x_n-x_{n-1})}"
        r"{f(x_n)-f(x_{n-1})}"
    ),
}

METHOD_NAME = "Secant Method"
DISPLAY_DECIMALS = 3
EXPECTED_SIMPLE_ROOT_ORDER = (1.0 + math.sqrt(5.0)) / 2.0

DEFAULT_X0 = 1.0
DEFAULT_X1 = 2.0
DEFAULT_TOLERANCE = 1.0e-8
DEFAULT_MAX_ITERATIONS = 100

MIN_ITERATIONS = 1
MAX_ITERATIONS = 1000
MIN_TOLERANCE = 1.0e-14

ZERO_TOLERANCE = 1.0e-15
MACHINE_EPSILON_FACTOR = 100.0
VALUE_MAGNITUDE_WARNING = 1.0e12
DENOMINATOR_SEPARATION_WARNING = 1.0e-10
ORDER_WARNING_THRESHOLD = 1.25
MULTIPLE_ROOT_WARNING_THRESHOLD = 1.5

REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

X_SYMBOL = sp.Symbol("x", real=True)

ALLOWED_FUNCTION_NAMES = {
    "x": X_SYMBOL,
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}


# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable

    number = float(value)
    if not math.isfinite(number):
        return str(number)

    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0**exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)

    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Use fixed or scientific notation according to numerical magnitude."""

    if value is None:
        return unavailable

    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals)
        or magnitude >= 1.0e6
    ):
        return format_scientific_power(
            number,
            decimals,
            unavailable,
        )

    return f"{number:.{decimals}f}"


def format_number(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
) -> str:
    """Format one numerical value for Streamlit display."""

    return format_display_number(
        value,
        decimals,
        unavailable="Not available",
    )


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only in a display copy."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(
        include=[np.number]
    ).columns

    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[
            numeric_columns
        ].round(decimals)

    return rounded


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class SecantIteration:
    """One secant-line root update."""

    iteration: int
    x_previous: float
    x_current: float
    function_previous: float
    function_current: float
    delta_x: float
    delta_function: float
    secant_slope: float
    denominator_relative_separation: float
    correction: float
    x_next: float
    function_next: float
    absolute_step: float
    relative_step_percent: float
    residual: float
    bracketed_before_update: bool
    exact_root_error: float | None
    exact_root_relative_error_percent: float | None
    observed_order: float | None
    secant_asymptotic_indicator: float | None
    operation: str
    status: str


@dataclass(frozen=True)
class SecantResult:
    """Complete result shared by Streamlit and Excel renderers."""

    status: str
    success: bool
    converged: bool
    method: str
    message: str
    stopping_reason: str

    function_text: str
    function_expression: sp.Expr | None
    derivative_expression: sp.Expr | None
    second_derivative_expression: sp.Expr | None

    initial_x0: float | None
    initial_x1: float | None
    tolerance: float | None
    maximum_iterations: int

    known_exact_root: float | None
    exact_root_verified: bool | None
    symbolic_multiplicity_at_exact_root: int | None

    iterations: tuple[SecantIteration, ...]
    approximate_root: float | None
    function_at_root: float | None
    final_absolute_step: float | None
    final_relative_step_percent: float | None
    final_residual: float | None
    final_exact_error: float | None
    latest_observed_order: float | None
    latest_secant_asymptotic_indicator: float | None
    minimum_denominator_relative_separation: float | None

    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General validation helpers
# =============================================================================
def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def validate_finite_real(
    raw_value: Any,
    value_name: str,
) -> float:
    """Convert one input to a finite real float."""

    try:
        complex_value = complex(raw_value)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"{value_name} must be a valid real number."
        ) from error

    if abs(complex_value.imag) > ZERO_TOLERANCE:
        raise ValueError(
            f"{value_name} must be real, not complex."
        )

    value = float(complex_value.real)

    if not math.isfinite(value):
        raise ValueError(
            f"{value_name} must be finite; NaN and infinity are not allowed."
        )

    return value


def validate_integer(
    raw_value: Any,
    value_name: str,
    minimum: int,
    maximum: int,
) -> int:
    """Validate an integer without silently truncating decimal values."""

    try:
        numeric_value = float(raw_value)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"{value_name} must be an integer."
        ) from error

    if (
        not math.isfinite(numeric_value)
        or not numeric_value.is_integer()
    ):
        raise ValueError(
            f"{value_name} must be an integer."
        )

    value = int(numeric_value)

    if value < minimum or value > maximum:
        raise ValueError(
            f"{value_name} must be between {minimum} and {maximum}."
        )

    return value


def parse_optional_exact_root(
    raw_value: Any,
) -> float | None:
    """Parse an optional known root entered as text."""

    if raw_value is None:
        return None

    text = str(raw_value).strip()
    if not text:
        return None

    try:
        expression = sp.sympify(
            text.replace("^", "**"),
            locals={
                "pi": sp.pi,
                "E": sp.E,
                "sqrt": sp.sqrt,
            },
        )
        evaluated = complex(sp.N(expression, 30))
    except (
        sp.SympifyError,
        TypeError,
        ValueError,
        OverflowError,
        SyntaxError,
    ) as error:
        raise ValueError(
            "The optional exact root must be a finite real number "
            "or a simple expression such as sqrt(2)."
        ) from error

    if (
        abs(evaluated.imag) > ZERO_TOLERANCE
        or not math.isfinite(evaluated.real)
    ):
        raise ValueError(
            "The optional exact root must evaluate to a finite real value."
        )

    return float(evaluated.real)


def create_input_signature(
    function_text: str,
    initial_x0: Any,
    initial_x1: Any,
    tolerance: Any,
    maximum_iterations: Any,
    exact_root_text: str,
) -> str:
    """Create a stable signature to prevent stale Streamlit results."""

    payload = "|".join(
        [
            str(function_text).strip(),
            repr(initial_x0),
            repr(initial_x1),
            repr(tolerance),
            repr(maximum_iterations),
            str(exact_root_text).strip(),
        ]
    )

    return hashlib.sha256(
        payload.encode("utf-8")
    ).hexdigest()


def values_are_numerically_distinct(
    first: float,
    second: float,
) -> bool:
    """Return whether two x-values are distinguishable in double precision."""

    scale = max(
        abs(first),
        abs(second),
        np.finfo(float).tiny,
    )
    threshold = (
        MACHINE_EPSILON_FACTOR
        * np.finfo(float).eps
        * scale
    )
    return abs(second - first) > threshold


# =============================================================================
# Safe symbolic parsing and numerical evaluation
# =============================================================================
def reject_unsupported_constructs(
    expression: sp.Expr,
) -> None:
    """Reject symbolic constructs inappropriate for scalar root finding."""

    unsupported_atoms = (
        AppliedUndef,
        sp.Derivative,
        sp.Integral,
        sp.Sum,
        sp.Product,
        sp.Limit,
    )

    if expression.has(*unsupported_atoms):
        raise ValueError(
            "The function contains unsupported symbolic operations."
        )

    if isinstance(expression, Relational) or expression.has(Relational):
        raise ValueError(
            "Enter a function expression, not an equation or inequality."
        )

    unexpected_symbols = expression.free_symbols.difference(
        {X_SYMBOL}
    )

    if unexpected_symbols:
        names = ", ".join(
            sorted(str(symbol) for symbol in unexpected_symbols)
        )
        raise ValueError(
            f"Only x is allowed as a variable. Unsupported: {names}."
        )

    if expression.has(
        sp.I,
        sp.zoo,
        sp.oo,
        -sp.oo,
        sp.nan,
    ):
        raise ValueError(
            "The function contains a complex or non-finite symbolic value."
        )


def parse_function(
    function_text: str,
) -> tuple[
    sp.Expr,
    sp.Expr,
    sp.Expr,
    Callable[[Any], Any],
]:
    """Parse f(x) and return optional derivative expressions for diagnostics."""

    if not isinstance(function_text, str) or not function_text.strip():
        raise ValueError(
            "Enter a function before solving."
        )

    text = function_text.strip().replace("^", "**")

    if "=" in text:
        raise ValueError(
            "Enter only f(x), without an equals sign."
        )

    try:
        expression = sp.sympify(
            text,
            locals=ALLOWED_FUNCTION_NAMES,
            evaluate=True,
        )
    except (
        sp.SympifyError,
        TypeError,
        ValueError,
        SyntaxError,
    ) as error:
        raise ValueError(
            "The function has an invalid format. Use expressions such as "
            "x**3 - x - 2, exp(-x) - x, or sin(x)."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError(
            "The function could not be interpreted."
        )

    reject_unsupported_constructs(expression)

    if expression.is_constant():
        raise ValueError(
            "A constant function does not define a usable Secant iteration."
        )

    derivative_expression = sp.diff(
        expression,
        X_SYMBOL,
    )
    second_derivative_expression = sp.diff(
        expression,
        X_SYMBOL,
        2,
    )

    try:
        function = sp.lambdify(
            X_SYMBOL,
            expression,
            modules=["numpy"],
        )
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The function could not be converted to numerical form."
        ) from error

    return (
        expression,
        derivative_expression,
        second_derivative_expression,
        function,
    )


def evaluate_real_scalar(
    numerical_function: Callable[[Any], Any],
    x_value: float,
    value_name: str = "f(x)",
) -> float:
    """Evaluate one finite real function value safely."""

    try:
        with np.errstate(all="raise"):
            raw_value = numerical_function(float(x_value))
        array = np.asarray(raw_value)
    except (
        TypeError,
        ValueError,
        OverflowError,
        ZeroDivisionError,
        FloatingPointError,
    ) as error:
        raise ValueError(
            f"{value_name} is undefined at x = {x_value:.12g}. "
            f"Reason: {error}"
        ) from error

    if array.size != 1:
        raise ValueError(
            f"{value_name} did not return a scalar at x = {x_value:.12g}."
        )

    scalar = array.reshape(-1)[0]

    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)

        if abs(complex_value.imag) > ZERO_TOLERANCE:
            raise ValueError(
                f"{value_name} is complex at x = {x_value:.12g}."
            )

        scalar = complex_value.real

    return validate_finite_real(
        scalar,
        f"{value_name} at x = {x_value:.12g}",
    )


def evaluate_real_array(
    numerical_function: Callable[[Any], Any],
    x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate a function for plotting and preserve invalid points as NaN."""

    try:
        with np.errstate(all="ignore"):
            raw_values = numerical_function(x_values)
        values = np.asarray(raw_values)
    except Exception:
        return np.full_like(
            x_values,
            np.nan,
            dtype=float,
        )

    if values.ndim == 0:
        values = np.full_like(
            x_values,
            values,
            dtype=complex if np.iscomplexobj(values) else float,
        )
    else:
        try:
            values = np.broadcast_to(
                values,
                x_values.shape,
            )
        except ValueError:
            return np.full_like(
                x_values,
                np.nan,
                dtype=float,
            )

    if np.iscomplexobj(values):
        imaginary = np.abs(np.imag(values))
        result = np.real(values).astype(float)
        result[imaginary > ZERO_TOLERANCE] = np.nan
    else:
        try:
            result = values.astype(float)
        except (TypeError, ValueError):
            return np.full_like(
                x_values,
                np.nan,
                dtype=float,
            )

    result[~np.isfinite(result)] = np.nan
    return result


# =============================================================================
# Convergence and multiplicity diagnostics
# =============================================================================
def observed_order_from_errors(
    errors: list[float],
) -> float | None:
    """Estimate convergence order from three consecutive positive indicators."""

    if len(errors) < 3:
        return None

    e_previous_previous = float(errors[-3])
    e_previous = float(errors[-2])
    e_current = float(errors[-1])

    if (
        e_previous_previous <= ZERO_TOLERANCE
        or e_previous <= ZERO_TOLERANCE
        or e_current <= ZERO_TOLERANCE
    ):
        return None

    denominator = math.log(
        e_previous / e_previous_previous
    )

    if abs(denominator) <= ZERO_TOLERANCE:
        return None

    value = math.log(
        e_current / e_previous
    ) / denominator

    return value if math.isfinite(value) else None


def secant_asymptotic_indicator(
    previous_error: float | None,
    current_error: float | None,
) -> float | None:
    """Return e_(n+1)/e_n^φ for known-root diagnostic purposes."""

    if (
        previous_error is None
        or current_error is None
        or previous_error <= ZERO_TOLERANCE
        or current_error < 0.0
    ):
        return None

    denominator = previous_error**EXPECTED_SIMPLE_ROOT_ORDER
    if denominator <= np.finfo(float).tiny:
        return None

    value = current_error / denominator
    return value if math.isfinite(value) else None


def determine_symbolic_multiplicity(
    expression: sp.Expr,
    exact_root: float,
    maximum_order: int = 10,
) -> int | None:
    """Attempt to identify root multiplicity through symbolic derivatives."""

    root_symbolic = sp.Float(
        exact_root,
        40,
    )

    for derivative_order in range(
        0,
        maximum_order + 1,
    ):
        derivative_expression = sp.diff(
            expression,
            X_SYMBOL,
            derivative_order,
        )

        try:
            value = complex(
                sp.N(
                    derivative_expression.subs(
                        X_SYMBOL,
                        root_symbolic,
                    ),
                    50,
                )
            )
        except (
            TypeError,
            ValueError,
            OverflowError,
        ):
            return None

        if (
            abs(value.imag) > ZERO_TOLERANCE
            or not math.isfinite(value.real)
        ):
            return None

        if abs(value.real) > 1.0e-10:
            return derivative_order

    return None


# =============================================================================
# Secant algorithm
# =============================================================================
def error_result(
    message: str,
    input_signature: str = "",
) -> SecantResult:
    """Create a complete failed result."""

    return SecantResult(
        status="error",
        success=False,
        converged=False,
        method=METHOD_NAME,
        message=message,
        stopping_reason=(
            "Execution stopped during input validation or numerical evaluation."
        ),
        function_text="",
        function_expression=None,
        derivative_expression=None,
        second_derivative_expression=None,
        initial_x0=None,
        initial_x1=None,
        tolerance=None,
        maximum_iterations=0,
        known_exact_root=None,
        exact_root_verified=None,
        symbolic_multiplicity_at_exact_root=None,
        iterations=(),
        approximate_root=None,
        function_at_root=None,
        final_absolute_step=None,
        final_relative_step_percent=None,
        final_residual=None,
        final_exact_error=None,
        latest_observed_order=None,
        latest_secant_asymptotic_indicator=None,
        minimum_denominator_relative_separation=None,
        warnings=(),
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


def solve_secant_method(
    function_text: str,
    initial_x0_input: Any,
    initial_x1_input: Any,
    tolerance_input: Any,
    maximum_iterations_input: Any,
    exact_root_text: str = "",
) -> SecantResult:
    """Solve f(x)=0 using successive secant-line x-intercepts."""

    input_signature = create_input_signature(
        function_text=function_text,
        initial_x0=initial_x0_input,
        initial_x1=initial_x1_input,
        tolerance=tolerance_input,
        maximum_iterations=maximum_iterations_input,
        exact_root_text=exact_root_text,
    )

    try:
        initial_x0 = validate_finite_real(
            initial_x0_input,
            "Initial approximation x0",
        )
        initial_x1 = validate_finite_real(
            initial_x1_input,
            "Initial approximation x1",
        )
        tolerance = validate_finite_real(
            tolerance_input,
            "Tolerance",
        )
        maximum_iterations = validate_integer(
            maximum_iterations_input,
            "Maximum iterations",
            MIN_ITERATIONS,
            MAX_ITERATIONS,
        )
        known_exact_root = parse_optional_exact_root(
            exact_root_text
        )

        if tolerance <= 0.0:
            raise ValueError(
                "Tolerance must be greater than zero."
            )

        if not values_are_numerically_distinct(
            initial_x0,
            initial_x1,
        ):
            raise ValueError(
                "The two starting values must be numerically distinct."
            )

        (
            expression,
            derivative_expression,
            second_derivative_expression,
            function,
        ) = parse_function(function_text)

        previous_x = initial_x0
        current_x = initial_x1
        previous_function = evaluate_real_scalar(
            function,
            previous_x,
            "f(x)",
        )
        current_function = evaluate_real_scalar(
            function,
            current_x,
            "f(x)",
        )

        function_scale = max(
            abs(previous_function),
            abs(current_function),
            np.finfo(float).tiny,
        )
        residual_tolerance = tolerance * function_scale

        warnings: list[str] = []
        exact_root_verified: bool | None = None
        symbolic_multiplicity: int | None = None

        if known_exact_root is not None:
            exact_function_value = evaluate_real_scalar(
                function,
                known_exact_root,
                "f(x)",
            )

            # Verify the optional root in a way that is insensitive to a
            # simple rescaling of f.  The derivative is used only for this
            # diagnostic check; the Secant iteration itself remains
            # derivative-free.
            try:
                derivative_numeric = sp.lambdify(
                    X_SYMBOL,
                    derivative_expression,
                    modules=["numpy"],
                )
                derivative_at_exact = evaluate_real_scalar(
                    derivative_numeric,
                    known_exact_root,
                    "f'(x)",
                )
            except (
                ValueError,
                TypeError,
                OverflowError,
                FloatingPointError,
            ):
                derivative_at_exact = 0.0

            verification_scale = max(
                1.0,
                abs(derivative_at_exact)
                * max(1.0, abs(known_exact_root)),
            )
            verification_tolerance = max(
                1.0e-12,
                tolerance,
            ) * verification_scale
            exact_root_verified = (
                abs(exact_function_value)
                <= verification_tolerance
            )

            if not exact_root_verified:
                raise ValueError(
                    "The supplied exact root does not satisfy f(x)=0 "
                    "within the numerical verification tolerance."
                )

            symbolic_multiplicity = determine_symbolic_multiplicity(
                expression,
                known_exact_root,
            )

            if (
                symbolic_multiplicity is not None
                and symbolic_multiplicity > 1
            ):
                warnings.append(
                    "The known root is multiple. The ordinary Secant Method "
                    "generally loses its simple-root superlinear convergence; "
                    "the Multiple Roots Solver is more appropriate."
                )

        history: list[SecantIteration] = []
        convergence_indicators: list[float] = []

        # Include both initial errors for true-order estimation when available.
        if known_exact_root is not None:
            convergence_indicators.extend(
                [
                    abs(previous_x - known_exact_root),
                    abs(current_x - known_exact_root),
                ]
            )

        converged = False
        stopping_reason = (
            "Maximum iterations reached before the complete convergence "
            "criteria were satisfied."
        )
        accepted_initial: float | None = None

        if previous_function == 0.0:
            accepted_initial = previous_x
            converged = True
            stopping_reason = (
                "The first starting value evaluates to an exact "
                "floating-point root."
            )
        elif current_function == 0.0:
            accepted_initial = current_x
            converged = True
            stopping_reason = (
                "The second starting value evaluates to an exact "
                "floating-point root."
            )
        elif abs(previous_function) <= residual_tolerance:
            accepted_initial = previous_x
            converged = True
            stopping_reason = (
                "The first starting value satisfies the scaled residual tolerance."
            )
        elif abs(current_function) <= residual_tolerance:
            accepted_initial = current_x
            converged = True
            stopping_reason = (
                "The second starting value satisfies the scaled residual tolerance."
            )

        minimum_denominator_separation: float | None = None

        for iteration in range(
            1,
            maximum_iterations + 1,
        ):
            if converged:
                break

            delta_x = current_x - previous_x
            delta_function = current_function - previous_function

            if not values_are_numerically_distinct(
                previous_x,
                current_x,
            ):
                raise ValueError(
                    "The two most recent x-values became numerically "
                    "indistinguishable before convergence."
                )

            function_pair_scale = max(
                abs(previous_function),
                abs(current_function),
                np.finfo(float).tiny,
            )
            denominator_threshold = (
                MACHINE_EPSILON_FACTOR
                * np.finfo(float).eps
                * function_pair_scale
            )
            denominator_relative_separation = (
                abs(delta_function) / function_pair_scale
            )

            if (
                minimum_denominator_separation is None
                or denominator_relative_separation
                < minimum_denominator_separation
            ):
                minimum_denominator_separation = (
                    denominator_relative_separation
                )

            if abs(delta_function) <= denominator_threshold:
                raise ValueError(
                    "The Secant denominator f(x_n)-f(x_(n-1)) became "
                    "zero or numerically too small relative to the function "
                    f"values at iteration {iteration}. Choose different "
                    "starting values or use a safeguarded method."
                )

            secant_slope = delta_function / delta_x
            correction = current_function * delta_x / delta_function
            next_x = current_x - correction

            if not all(
                math.isfinite(value)
                for value in (
                    secant_slope,
                    correction,
                    next_x,
                )
            ):
                raise ValueError(
                    "The Secant update produced NaN or infinity."
                )

            next_function = evaluate_real_scalar(
                function,
                next_x,
                "f(x)",
            )

            absolute_step = abs(next_x - current_x)
            relative_step_percent = (
                absolute_step
                / max(1.0, abs(next_x))
                * 100.0
            )
            residual = abs(next_function)
            bracketed_before_update = (
                previous_function * current_function < 0.0
            )

            exact_root_error: float | None = None
            exact_root_relative_error_percent: float | None = None

            if known_exact_root is not None:
                exact_root_error = abs(
                    next_x - known_exact_root
                )
                exact_root_relative_error_percent = (
                    exact_root_error
                    / max(1.0, abs(known_exact_root))
                    * 100.0
                )
                convergence_indicators.append(
                    exact_root_error
                )
            else:
                convergence_indicators.append(
                    absolute_step
                )

            observed_order = observed_order_from_errors(
                convergence_indicators
            )

            previous_true_error = (
                abs(current_x - known_exact_root)
                if known_exact_root is not None
                else None
            )
            asymptotic_indicator = secant_asymptotic_indicator(
                previous_true_error,
                exact_root_error,
            )

            operation = (
                f"x_{iteration + 1} = {current_x:.15g} - "
                f"({current_function:.15g})"
                f"({current_x:.15g} - {previous_x:.15g}) / "
                f"({current_function:.15g} - "
                f"{previous_function:.15g}) = {next_x:.15g}"
            )

            history.append(
                SecantIteration(
                    iteration=iteration,
                    x_previous=float(previous_x),
                    x_current=float(current_x),
                    function_previous=float(previous_function),
                    function_current=float(current_function),
                    delta_x=float(delta_x),
                    delta_function=float(delta_function),
                    secant_slope=float(secant_slope),
                    denominator_relative_separation=float(
                        denominator_relative_separation
                    ),
                    correction=float(correction),
                    x_next=float(next_x),
                    function_next=float(next_function),
                    absolute_step=float(absolute_step),
                    relative_step_percent=float(
                        relative_step_percent
                    ),
                    residual=float(residual),
                    bracketed_before_update=bool(
                        bracketed_before_update
                    ),
                    exact_root_error=exact_root_error,
                    exact_root_relative_error_percent=(
                        exact_root_relative_error_percent
                    ),
                    observed_order=observed_order,
                    secant_asymptotic_indicator=(
                        asymptotic_indicator
                    ),
                    operation=operation,
                    status="Completed",
                )
            )

            step_tolerance = tolerance * max(
                1.0,
                abs(next_x),
            )
            step_condition = (
                absolute_step <= step_tolerance
            )
            residual_condition = (
                residual <= residual_tolerance
            )
            exact_condition = (
                known_exact_root is not None
                and exact_root_error is not None
                and exact_root_error
                <= tolerance
                * max(1.0, abs(known_exact_root))
            )

            if next_function == 0.0:
                converged = True
                stopping_reason = (
                    "The function evaluated to an exact floating-point "
                    "zero at the new approximation."
                )
                current_x = next_x
                current_function = next_function
                break

            if step_condition and residual_condition:
                converged = True
                stopping_reason = (
                    "Both the scaled step-size tolerance and the scaled "
                    "residual tolerance were satisfied."
                )
                current_x = next_x
                current_function = next_function
                break

            if exact_condition and residual_condition:
                converged = True
                stopping_reason = (
                    "The known-root error and scaled residual tolerance "
                    "were satisfied."
                )
                current_x = next_x
                current_function = next_function
                break

            if next_x == current_x and not residual_condition:
                raise ValueError(
                    "The iteration stagnated because floating-point "
                    "arithmetic could not change x while the residual "
                    "remained above tolerance."
                )

            if (
                abs(next_x) >= VALUE_MAGNITUDE_WARNING
                or abs(next_function) >= VALUE_MAGNITUDE_WARNING
            ):
                warnings.append(
                    "The iteration reached a very large magnitude. "
                    "The Secant Method may be diverging for these "
                    "starting values."
                )

            previous_x = float(current_x)
            previous_function = float(current_function)
            current_x = float(next_x)
            current_function = float(next_function)

        if accepted_initial is not None:
            approximate_root = accepted_initial
        else:
            approximate_root = current_x

        function_at_root = evaluate_real_scalar(
            function,
            approximate_root,
            "f(x)",
        )

        if history:
            final_absolute_step = history[-1].absolute_step
            final_relative_step_percent = (
                history[-1].relative_step_percent
            )
            latest_observed_order = next(
                (
                    item.observed_order
                    for item in reversed(history)
                    if item.observed_order is not None
                    and math.isfinite(item.observed_order)
                ),
                None,
            )
            latest_asymptotic_indicator = next(
                (
                    item.secant_asymptotic_indicator
                    for item in reversed(history)
                    if item.secant_asymptotic_indicator
                    is not None
                    and math.isfinite(
                        item.secant_asymptotic_indicator
                    )
                ),
                None,
            )
        else:
            final_absolute_step = 0.0
            final_relative_step_percent = 0.0
            latest_observed_order = None
            latest_asymptotic_indicator = None

        final_residual = abs(function_at_root)
        final_exact_error = (
            abs(approximate_root - known_exact_root)
            if known_exact_root is not None
            else None
        )

        if (
            minimum_denominator_separation is not None
            and minimum_denominator_separation
            <= DENOMINATOR_SEPARATION_WARNING
        ):
            warnings.append(
                "At least one pair of consecutive function values was "
                "very close relative to its scale. The secant slope was "
                "sensitive to floating-point rounding."
            )

        if (
            latest_observed_order is not None
            and converged
            and latest_observed_order < ORDER_WARNING_THRESHOLD
            and (
                symbolic_multiplicity is None
                or symbolic_multiplicity <= 1
            )
        ):
            warnings.append(
                "The observed convergence was weaker than the expected "
                f"simple-root Secant order φ ≈ "
                f"{EXPECTED_SIMPLE_ROOT_ORDER:.6f}. Poor initial values, "
                "a nonsmooth function, or round-off may be responsible."
            )

        if not converged:
            warnings.append(
                "The final approximation is reported, but the solver did "
                "not satisfy the complete convergence criteria."
            )

        if known_exact_root is None:
            warnings.append(
                "No exact root was supplied. Observed order is estimated "
                "from successive step sizes rather than true root errors."
            )

        return SecantResult(
            status="success",
            success=True,
            converged=converged,
            method=METHOD_NAME,
            message=(
                "Root refined successfully with the Secant Method."
                if converged
                else (
                    "Maximum iterations reached; the final "
                    "approximation is shown."
                )
            ),
            stopping_reason=stopping_reason,
            function_text=function_text.strip(),
            function_expression=expression,
            derivative_expression=derivative_expression,
            second_derivative_expression=(
                second_derivative_expression
            ),
            initial_x0=initial_x0,
            initial_x1=initial_x1,
            tolerance=tolerance,
            maximum_iterations=maximum_iterations,
            known_exact_root=known_exact_root,
            exact_root_verified=exact_root_verified,
            symbolic_multiplicity_at_exact_root=(
                symbolic_multiplicity
            ),
            iterations=tuple(history),
            approximate_root=float(approximate_root),
            function_at_root=float(function_at_root),
            final_absolute_step=final_absolute_step,
            final_relative_step_percent=(
                final_relative_step_percent
            ),
            final_residual=final_residual,
            final_exact_error=final_exact_error,
            latest_observed_order=latest_observed_order,
            latest_secant_asymptotic_indicator=(
                latest_asymptotic_indicator
            ),
            minimum_denominator_relative_separation=(
                minimum_denominator_separation
            ),
            warnings=tuple(dict.fromkeys(warnings)),
            input_signature=input_signature,
            execution_datetime=current_report_datetime(),
        )

    except (
        ValueError,
        TypeError,
        ArithmeticError,
        OverflowError,
    ) as error:
        return error_result(
            message=str(error),
            input_signature=input_signature,
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def iterations_dataframe(
    result: SecantResult,
) -> pd.DataFrame:
    """Return the complete Secant iteration table."""

    return pd.DataFrame(
        [
            {
                "Iteration": item.iteration,
                "x_(n-1)": item.x_previous,
                "x_n": item.x_current,
                "f(x_(n-1))": item.function_previous,
                "f(x_n)": item.function_current,
                "Δx": item.delta_x,
                "Δf": item.delta_function,
                "Secant Slope Δf/Δx": item.secant_slope,
                "Relative Denominator Separation": (
                    item.denominator_relative_separation
                ),
                "Correction f_n*Δx/Δf": item.correction,
                "x_(n+1)": item.x_next,
                "f(x_(n+1))": item.function_next,
                "Absolute Step": item.absolute_step,
                "Scaled Relative Step (%)": (
                    item.relative_step_percent
                ),
                "Residual |f(x_(n+1))|": item.residual,
                "Sign Change Before Update": (
                    "Yes" if item.bracketed_before_update else "No"
                ),
                "Exact Root Error": item.exact_root_error,
                "Exact Root Relative Error (%)": (
                    item.exact_root_relative_error_percent
                ),
                "Observed Order": item.observed_order,
                "Secant Asymptotic Indicator": (
                    item.secant_asymptotic_indicator
                ),
                "Operation": item.operation,
                "Status": item.status,
            }
            for item in result.iterations
        ]
    )


def method_formula_dataframe(
    result: SecantResult,
) -> pd.DataFrame:
    """Return formulas and interpretations for the report."""

    return pd.DataFrame(
        {
            "Item": [
                "Problem",
                "Secant Slope",
                "Secant Update",
                "Absolute Step",
                "Scaled Relative Step",
                "Residual",
                "Expected Simple-Root Order",
                "Bracketing Requirement",
                "Complete Stopping Test",
            ],
            "Formula / Meaning": [
                "Solve f(x)=0 using two recent approximations",
                "[f(x_n)-f(x_(n-1))]/[x_n-x_(n-1)]",
                (
                    "x_(n+1)=x_n-f(x_n)[x_n-x_(n-1)]/"
                    "[f(x_n)-f(x_(n-1))]"
                ),
                "|x_(n+1)-x_n|",
                (
                    "|x_(n+1)-x_n| / max(1,|x_(n+1)|) × 100%"
                ),
                "|f(x_(n+1))|",
                f"φ=(1+sqrt(5))/2≈{EXPECTED_SIMPLE_ROOT_ORDER:.12g}",
                "Not required; Secant is an open method",
                (
                    "Require both a small scaled step and a small "
                    "scaled residual"
                ),
            ],
        }
    )


def convergence_diagnostics_dataframe(
    result: SecantResult,
) -> pd.DataFrame:
    """Return convergence and denominator diagnostics."""

    rows: list[dict[str, Any]] = [
        {
            "Diagnostic": "Expected simple-root order",
            "Value": EXPECTED_SIMPLE_ROOT_ORDER,
            "Interpretation": "Superlinear Secant convergence near a smooth simple root",
        },
        {
            "Diagnostic": "Latest observed order",
            "Value": result.latest_observed_order,
            "Interpretation": (
                "Uses true errors when a known root is supplied; otherwise uses step sizes"
            ),
        },
        {
            "Diagnostic": "Latest Secant asymptotic indicator",
            "Value": result.latest_secant_asymptotic_indicator,
            "Interpretation": "e_(n+1)/e_n^φ; available only with a known root",
        },
        {
            "Diagnostic": "Minimum relative denominator separation",
            "Value": result.minimum_denominator_relative_separation,
            "Interpretation": (
                "Small values indicate f(x_n)-f(x_(n-1)) is sensitive to rounding"
            ),
        },
        {
            "Diagnostic": "Symbolic multiplicity at known root",
            "Value": result.symbolic_multiplicity_at_exact_root,
            "Interpretation": "Available only when a known exact root is supplied",
        },
        {
            "Diagnostic": "Final known-root error",
            "Value": result.final_exact_error,
            "Interpretation": "Available only when a known exact root is supplied",
        },
    ]

    for item in result.iterations:
        rows.append(
            {
                "Diagnostic": f"Observed order at iteration {item.iteration}",
                "Value": item.observed_order,
                "Interpretation": (
                    "Local empirical order estimate"
                ),
            }
        )

    return pd.DataFrame(rows)


def summary_dataframe(
    result: SecantResult,
) -> pd.DataFrame:
    """Create the report summary."""

    return pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "Converged",
                "Function",
                "Initial x0",
                "Initial x1",
                "Tolerance",
                "Maximum Iterations",
                "Iterations Used",
                "Approximate Root",
                "f(Approximate Root)",
                "Final Absolute Step",
                "Final Relative Step (%)",
                "Final Residual",
                "Known Exact Root",
                "Exact Root Verified",
                "Detected Multiplicity at Known Root",
                "Final Exact-Root Error",
                "Expected Simple-Root Order",
                "Latest Observed Order",
                "Minimum Relative Denominator Separation",
                "Stopping Reason",
                "Warnings",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                "Yes" if result.converged else "No",
                str(result.function_expression),
                result.initial_x0,
                result.initial_x1,
                result.tolerance,
                result.maximum_iterations,
                len(result.iterations),
                result.approximate_root,
                result.function_at_root,
                result.final_absolute_step,
                result.final_relative_step_percent,
                result.final_residual,
                result.known_exact_root,
                result.exact_root_verified,
                result.symbolic_multiplicity_at_exact_root,
                result.final_exact_error,
                EXPECTED_SIMPLE_ROOT_ORDER,
                result.latest_observed_order,
                result.minimum_denominator_relative_separation,
                result.stopping_reason,
                (
                    " | ".join(result.warnings)
                    if result.warnings
                    else "None"
                ),
                result.execution_datetime.strftime(
                    "%Y-%m-%d %H:%M:%S %Z"
                ),
            ],
        }
    )


def build_plot_dataframe(
    result: SecantResult,
    sample_count: int = 700,
) -> pd.DataFrame:
    """Build function and iteration data used by charts."""

    if (
        result.function_expression is None
        or result.approximate_root is None
        or result.initial_x0 is None
        or result.initial_x1 is None
    ):
        return pd.DataFrame()

    function = sp.lambdify(
        X_SYMBOL,
        result.function_expression,
        modules=["numpy"],
    )

    points = [
        result.initial_x0,
        result.initial_x1,
        result.approximate_root,
    ]
    points.extend(
        item.x_next for item in result.iterations
    )

    if result.known_exact_root is not None:
        points.append(result.known_exact_root)

    minimum_x = min(points)
    maximum_x = max(points)
    span = maximum_x - minimum_x
    padding = (
        1.0
        if span <= 0.0
        else max(0.25 * span, 0.5)
    )

    graph_x = np.linspace(
        minimum_x - padding,
        maximum_x + padding,
        sample_count,
    )
    graph_y = evaluate_real_array(
        function,
        graph_x,
    )

    row_count = max(
        sample_count,
        len(result.iterations),
        1,
    )
    dataframe = pd.DataFrame(index=range(row_count))
    dataframe["Function x"] = pd.Series(graph_x)
    dataframe["f(x)"] = pd.Series(graph_y)
    dataframe["Iteration"] = pd.Series(
        [item.iteration for item in result.iterations]
    )
    dataframe["Approximation"] = pd.Series(
        [item.x_next for item in result.iterations]
    )
    dataframe["Absolute Step"] = pd.Series(
        [item.absolute_step for item in result.iterations]
    )
    dataframe["Residual"] = pd.Series(
        [item.residual for item in result.iterations]
    )
    dataframe["Exact Root Error"] = pd.Series(
        [item.exact_root_error for item in result.iterations]
    )
    dataframe["Observed Order"] = pd.Series(
        [item.observed_order for item in result.iterations]
    )

    return dataframe


# =============================================================================
# Scientific plots
# =============================================================================
def create_function_figure(
    result: SecantResult,
) -> Figure:
    """Plot f(x), iteration points, and representative secant lines."""

    if (
        result.function_expression is None
        or result.approximate_root is None
    ):
        raise ValueError(
            "A successful result is required for plotting."
        )

    plot_data = build_plot_dataframe(result)
    if plot_data.empty:
        raise ValueError(
            "Plot data are unavailable."
        )

    valid = np.isfinite(
        plot_data["Function x"].to_numpy(dtype=float)
    ) & np.isfinite(
        plot_data["f(x)"].to_numpy(dtype=float)
    )

    if np.count_nonzero(valid) < 2:
        raise ValueError(
            "The function has insufficient finite values for plotting."
        )

    figure, axis = plt.subplots(figsize=(9.8, 5.9))
    axis.plot(
        plot_data.loc[valid, "Function x"],
        plot_data.loc[valid, "f(x)"],
        linewidth=2.0,
        label="f(x)",
    )
    axis.axhline(0.0, linewidth=1.0)

    if result.iterations:
        iteration_x = np.asarray(
            [item.x_next for item in result.iterations],
            dtype=float,
        )
        iteration_y = np.asarray(
            [item.function_next for item in result.iterations],
            dtype=float,
        )
        axis.scatter(
            iteration_x,
            iteration_y,
            s=48,
            label="Secant approximations",
            zorder=5,
        )

        selected_indices = np.unique(
            np.linspace(
                0,
                len(result.iterations) - 1,
                min(6, len(result.iterations)),
                dtype=int,
            )
        )
        for selected_index in selected_indices:
            item = result.iterations[int(selected_index)]
            line_x = np.asarray(
                [
                    item.x_previous,
                    item.x_current,
                    item.x_next,
                ],
                dtype=float,
            )
            line_y = item.function_previous + item.secant_slope * (
                line_x - item.x_previous
            )
            axis.plot(
                line_x,
                line_y,
                linestyle="--",
                linewidth=0.9,
                alpha=0.7,
            )

    axis.scatter(
        [result.approximate_root],
        [result.function_at_root],
        s=115,
        marker="*",
        label=(
            f"Approximate root = "
            f"{result.approximate_root:.8g}"
        ),
        zorder=6,
    )
    axis.axvline(
        result.approximate_root,
        linestyle="--",
        linewidth=1.1,
    )

    if result.known_exact_root is not None:
        axis.axvline(
            result.known_exact_root,
            linestyle=":",
            linewidth=1.3,
            label=(
                f"Known root = "
                f"{result.known_exact_root:.8g}"
            ),
        )

    axis.set_title("Function and Secant-Line Root Updates")
    axis.set_xlabel("x")
    axis.set_ylabel("f(x)")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()

    return figure


def create_approximation_figure(
    result: SecantResult,
) -> Figure:
    """Plot root approximations by iteration."""

    figure, axis = plt.subplots(figsize=(9.5, 5.8))

    if result.iterations:
        iterations = np.asarray(
            [item.iteration for item in result.iterations],
            dtype=int,
        )
        approximations = np.asarray(
            [item.x_next for item in result.iterations],
            dtype=float,
        )
        axis.plot(
            iterations,
            approximations,
            marker="o",
            linewidth=2.0,
            label="Secant approximation",
        )
        axis.axhline(
            result.approximate_root,
            linestyle="--",
            linewidth=1.0,
            label="Final approximation",
        )

        if result.known_exact_root is not None:
            axis.axhline(
                result.known_exact_root,
                linestyle=":",
                linewidth=1.2,
                label="Known exact root",
            )
    else:
        axis.scatter(
            [0],
            [result.approximate_root],
            s=90,
            label="Starting value accepted",
        )

    axis.set_title("Secant Root Approximation by Iteration")
    axis.set_xlabel("Iteration")
    axis.set_ylabel("x")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()

    return figure


def create_convergence_figure(
    result: SecantResult,
) -> Figure:
    """Plot step, residual, and optional true-root error."""

    figure, axis = plt.subplots(figsize=(9.5, 5.8))

    if not result.iterations:
        axis.text(
            0.5,
            0.5,
            "No Secant iteration was required.",
            ha="center",
            va="center",
            transform=axis.transAxes,
        )
    else:
        iterations = np.asarray(
            [item.iteration for item in result.iterations],
            dtype=int,
        )
        steps = np.maximum(
            np.asarray(
                [item.absolute_step for item in result.iterations],
                dtype=float,
            ),
            np.finfo(float).tiny,
        )
        residuals = np.maximum(
            np.asarray(
                [item.residual for item in result.iterations],
                dtype=float,
            ),
            np.finfo(float).tiny,
        )

        axis.semilogy(
            iterations,
            steps,
            marker="o",
            linewidth=2.0,
            label="Absolute step",
        )
        axis.semilogy(
            iterations,
            residuals,
            marker="o",
            linewidth=2.0,
            label="Residual",
        )

        if result.known_exact_root is not None:
            true_errors = np.asarray(
                [
                    (
                        item.exact_root_error
                        if item.exact_root_error is not None
                        else np.nan
                    )
                    for item in result.iterations
                ],
                dtype=float,
            )
            valid = np.isfinite(true_errors)
            if np.any(valid):
                axis.semilogy(
                    iterations[valid],
                    np.maximum(
                        true_errors[valid],
                        np.finfo(float).tiny,
                    ),
                    marker="o",
                    linewidth=2.0,
                    label="Known-root error",
                )

    axis.set_title("Secant Convergence Indicators")
    axis.set_xlabel("Iteration")
    axis.set_ylabel("Magnitude (log scale)")
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    figure.tight_layout()

    return figure


def figure_to_png_bytes(
    figure: Figure,
) -> bytes:
    """Serialize a matplotlib figure to PNG bytes."""

    buffer = BytesIO()
    figure.savefig(
        buffer,
        format="png",
        dpi=180,
        bbox_inches="tight",
    )
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# Excel report
# =============================================================================
def style_excel_workbook(
    workbook: Any,
) -> None:
    """Apply professional formatting to every worksheet."""

    header_fill = PatternFill(
        "solid",
        fgColor="0D3151",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.freeze_panes = "A2"

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions

        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            maximum_length = 0

            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
                maximum_length = max(
                    maximum_length,
                    len(str(cell.value))
                    if cell.value is not None
                    else 0,
                )

                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(maximum_length + 2, 12),
                58,
            )


def add_excel_image(
    worksheet: Any,
    image_bytes: bytes,
    anchor: str,
    width: int = 760,
    height: int = 470,
) -> None:
    """Insert PNG bytes into an Excel worksheet."""

    image_stream = BytesIO(image_bytes)
    excel_image = ExcelImage(image_stream)
    excel_image.width = width
    excel_image.height = height
    worksheet.add_image(excel_image, anchor)


def create_excel_report(
    result: SecantResult,
) -> bytes:
    """Create a complete in-memory XLSX report."""

    if not result.success:
        raise ValueError(
            "Only a completed calculation can be exported."
        )

    summary = summary_dataframe(result)
    inputs = pd.DataFrame(
        {
            "Input": [
                "Function",
                "Initial approximation x0",
                "Initial approximation x1",
                "Tolerance",
                "Maximum iterations",
                "Known exact root",
            ],
            "Value": [
                result.function_text,
                result.initial_x0,
                result.initial_x1,
                result.tolerance,
                result.maximum_iterations,
                result.known_exact_root,
            ],
        }
    )
    formulas = method_formula_dataframe(result)
    iterations = iterations_dataframe(result)
    diagnostics = convergence_diagnostics_dataframe(result)
    plot_data = build_plot_dataframe(result)

    if iterations.empty:
        iterations = pd.DataFrame(
            [
                {
                    "Iteration": 0,
                    "x_(n-1)": result.initial_x0,
                    "x_n": result.initial_x1,
                    "x_(n+1)": result.approximate_root,
                    "Residual |f(x_(n+1))|": result.final_residual,
                    "Absolute Step": 0.0,
                    "Status": "Starting value accepted",
                }
            ]
        )

    function_figure = create_function_figure(result)
    approximation_figure = create_approximation_figure(result)
    convergence_figure = create_convergence_figure(result)

    function_png = figure_to_png_bytes(function_figure)
    approximation_png = figure_to_png_bytes(approximation_figure)
    convergence_png = figure_to_png_bytes(convergence_figure)

    plt.close(function_figure)
    plt.close(approximation_figure)
    plt.close(convergence_figure)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        inputs.to_excel(writer, sheet_name="Inputs", index=False)
        formulas.to_excel(writer, sheet_name="Method Formula", index=False)
        iterations.to_excel(writer, sheet_name="Iterations", index=False)
        diagnostics.to_excel(
            writer,
            sheet_name="Convergence Diagnostics",
            index=False,
        )
        plot_data.to_excel(writer, sheet_name="Plot Data", index=False)

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Secant Solver Plots"
        plots_sheet["A1"].font = Font(bold=True, size=14)

        add_excel_image(plots_sheet, function_png, "A3")
        add_excel_image(plots_sheet, approximation_png, "A30")
        add_excel_image(plots_sheet, convergence_png, "A57")

        summary_sheet = workbook["Summary"]
        plot_sheet = workbook["Plot Data"]

        if len(plot_data) > 1:
            function_chart = ScatterChart()
            function_chart.title = "Function Near the Estimated Root"
            function_chart.x_axis.title = "x"
            function_chart.y_axis.title = "f(x)"
            function_chart.height = 8
            function_chart.width = 15

            function_x = Reference(
                plot_sheet,
                min_col=1,
                min_row=2,
                max_row=len(plot_data) + 1,
            )
            function_y = Reference(
                plot_sheet,
                min_col=2,
                min_row=2,
                max_row=len(plot_data) + 1,
            )
            function_series = Series(
                function_y,
                function_x,
                title="f(x)",
            )
            function_chart.series.append(function_series)
            summary_sheet.add_chart(function_chart, "D2")

        if result.iterations:
            row_count = len(result.iterations) + 1

            approximation_chart = LineChart()
            approximation_chart.title = "Root Approximation by Iteration"
            approximation_chart.x_axis.title = "Iteration"
            approximation_chart.y_axis.title = "Approximation"
            approximation_chart.height = 8
            approximation_chart.width = 15
            approximation_chart.add_data(
                Reference(
                    plot_sheet,
                    min_col=4,
                    min_row=1,
                    max_row=row_count,
                ),
                titles_from_data=True,
            )
            approximation_chart.set_categories(
                Reference(
                    plot_sheet,
                    min_col=3,
                    min_row=2,
                    max_row=row_count,
                )
            )
            summary_sheet.add_chart(approximation_chart, "D20")

            convergence_chart = LineChart()
            convergence_chart.title = "Step, Residual, and Root Error"
            convergence_chart.x_axis.title = "Iteration"
            convergence_chart.y_axis.title = "Magnitude"
            convergence_chart.height = 8
            convergence_chart.width = 15
            convergence_chart.add_data(
                Reference(
                    plot_sheet,
                    min_col=5,
                    max_col=(
                        7
                        if result.known_exact_root is not None
                        else 6
                    ),
                    min_row=1,
                    max_row=row_count,
                ),
                titles_from_data=True,
            )
            convergence_chart.set_categories(
                Reference(
                    plot_sheet,
                    min_col=3,
                    min_row=2,
                    max_row=row_count,
                )
            )
            summary_sheet.add_chart(convergence_chart, "D38")

        style_excel_workbook(workbook)
        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit renderers
# =============================================================================
def render_final_result(
    result: SecantResult,
) -> None:
    """Render the compact final-result card."""

    if not result.success:
        st.error(result.message)
        return

    if result.converged:
        st.success(result.message)
    else:
        st.warning(result.message)

    first_metrics = st.columns(2)
    first_metrics[0].metric(
        "Approximate Root",
        format_number(result.approximate_root),
    )
    first_metrics[1].metric(
        "Iterations",
        len(result.iterations),
    )

    second_metrics = st.columns(2)
    second_metrics[0].metric(
        "Final Residual",
        format_number(result.final_residual),
    )
    second_metrics[1].metric(
        "Final Absolute Step",
        format_number(result.final_absolute_step),
    )

    st.markdown(
        f"**Stopping reason:** {result.stopping_reason}"
    )

    if result.known_exact_root is not None:
        exact_metrics = st.columns(2)
        exact_metrics[0].metric(
            "Known Exact Root",
            format_number(result.known_exact_root),
        )
        exact_metrics[1].metric(
            "Final Exact-Root Error",
            format_number(result.final_exact_error),
        )

    if result.latest_observed_order is not None:
        st.metric(
            "Latest Observed Order",
            format_number(result.latest_observed_order),
        )

    for warning in result.warnings:
        st.warning(warning)


def render_method_summary(
    result: SecantResult,
) -> None:
    """Render formulas and problem setup."""

    st.subheader("Method Formula and Problem Setup")

    formula_column, interpretation_column = st.columns(2)

    with formula_column:
        st.latex(CONFIG["formula"])
        st.latex(
            r"m_n=\frac{f(x_n)-f(x_{n-1})}"
            r"{x_n-x_{n-1}}"
        )

    with interpretation_column:
        st.latex(
            r"p=\varphi=\frac{1+\sqrt{5}}{2}"
            r"\approx1.618"
        )
        st.markdown(
            "The method uses the **x-intercept of the secant line**. "
            "No analytical derivative is evaluated in the iteration."
        )

    summary = pd.DataFrame(
        {
            "Property": [
                "Function f(x)",
                "Initial x0",
                "Initial x1",
                "Tolerance",
                "Maximum iterations",
                "Expected simple-root order",
                "Bracket required",
            ],
            "Value": [
                str(result.function_expression),
                result.initial_x0,
                result.initial_x1,
                result.tolerance,
                result.maximum_iterations,
                EXPECTED_SIMPLE_ROOT_ORDER,
                "No",
            ],
        }
    )

    st.dataframe(
        round_numeric_dataframe(summary),
        use_container_width=True,
        hide_index=True,
    )


def render_iteration_table(
    result: SecantResult,
) -> None:
    """Render all Secant calculations."""

    st.subheader("Iteration Table")
    dataframe = iterations_dataframe(result)

    if dataframe.empty:
        st.info(
            "No iteration was required because one starting value "
            "already satisfied the root criterion."
        )
        return

    display_columns = [
        "Iteration",
        "x_(n-1)",
        "x_n",
        "f(x_(n-1))",
        "f(x_n)",
        "Δx",
        "Δf",
        "Secant Slope Δf/Δx",
        "Correction f_n*Δx/Δf",
        "x_(n+1)",
        "f(x_(n+1))",
        "Absolute Step",
        "Scaled Relative Step (%)",
        "Residual |f(x_(n+1))|",
        "Sign Change Before Update",
        "Exact Root Error",
        "Observed Order",
    ]

    st.dataframe(
        round_numeric_dataframe(
            dataframe[
                [
                    column
                    for column in display_columns
                    if column in dataframe.columns
                ]
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )

    st.caption(
        "The solver requires a small scaled step and a small residual. "
        "The sign-change column is diagnostic only; the Secant Method "
        "does not require or preserve a bracket."
    )

    with st.expander(
        "Detailed Secant Operations",
        expanded=False,
    ):
        for item in result.iterations:
            st.markdown(f"**Iteration {item.iteration}:**")
            st.code(item.operation, language=None)
            st.caption(
                f"Secant slope = {item.secant_slope:.12g}; "
                f"absolute step = {item.absolute_step:.12g}; "
                f"residual = {item.residual:.12g}."
            )


def render_convergence_diagnostics(
    result: SecantResult,
) -> None:
    """Render order and denominator diagnostics."""

    st.subheader("Convergence Diagnostics")
    st.markdown(
        "Near a sufficiently smooth **simple root**, the Secant Method "
        f"has expected order **φ ≈ {EXPECTED_SIMPLE_ROOT_ORDER:.6f}**. "
        "Multiple roots generally reduce the convergence rate."
    )

    st.dataframe(
        round_numeric_dataframe(
            convergence_diagnostics_dataframe(result)
        ),
        use_container_width=True,
        hide_index=True,
    )


def render_graphs(
    result: SecantResult,
) -> None:
    """Render function, approximation, and convergence plots."""

    graph_column, approximation_column, error_column = st.columns(3)

    with graph_column:
        with st.container(border=True):
            st.subheader("Function Graph")
            try:
                figure = create_function_figure(result)
                st.pyplot(figure, use_container_width=True)
                plt.close(figure)
            except (
                ValueError,
                TypeError,
                ArithmeticError,
            ) as error:
                st.warning(
                    f"The function graph could not be displayed: {error}"
                )

    with approximation_column:
        with st.container(border=True):
            st.subheader("Root Approximation")
            figure = create_approximation_figure(result)
            st.pyplot(figure, use_container_width=True)
            plt.close(figure)

    with error_column:
        with st.container(border=True):
            st.subheader("Error Analysis")
            figure = create_convergence_figure(result)
            st.pyplot(figure, use_container_width=True)
            plt.close(figure)


def render_excel_download(
    result: SecantResult,
) -> None:
    """Generate and render the Excel report button."""

    st.subheader("Excel Report")

    signature_key = "secant_excel_signature"
    report_key = "secant_excel_report"

    if (
        st.session_state.get(signature_key)
        != result.input_signature
        or report_key not in st.session_state
    ):
        try:
            st.session_state[report_key] = create_excel_report(result)
            st.session_state[signature_key] = result.input_signature
        except (
            ValueError,
            TypeError,
            OSError,
            ArithmeticError,
        ) as error:
            st.error(
                "The Excel report could not be generated. "
                f"Details: {error}"
            )
            return

    report_bytes = st.session_state.get(report_key)
    if not report_bytes:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")

    st.download_button(
        label="Download Complete Excel Report",
        data=report_bytes,
        file_name=f"secant_method_report_{timestamp}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="secant_excel_download",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Secant Method solver page."""

    st.set_page_config(
        page_title=f"{CONFIG['title']} | Numerical Methods",
        page_icon="📘",
        layout="wide",
    )

    load_css()
    navbar(active_page="solver")

    st.html(
        f"""
        <section class="solver-hero">
            <div>
                <div class="page-label">{CONFIG['label']}</div>
                <h1>{CONFIG['title']}</h1>
                <p>{CONFIG['description']}</p>
                <div class="method-actions">
                    <a
                        href="/{CONFIG['lesson']}"
                        target="_self"
                        class="btn-outline-ui"
                    >
                        Review Lesson →
                    </a>
                    <a
                        href="/{CONFIG['quiz']}"
                        target="_self"
                        class="btn-primary-ui"
                    >
                        Take Quiz →
                    </a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns(
        [0.035, 0.93, 0.035]
    )

    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Write the Function")
                st.markdown(
                    """
                    Enter **f(x)** without an equals sign and use only
                    **x** as the variable.

                    - Powers: write `x**2`, not `x^2`.
                    - Multiplication: write `2*x`, not `2x`.
                    - Functions: `sin(x)`, `cos(x)`, `sqrt(x)`,
                      `exp(x)`, and `log(x)`.
                    - Example: `x**3 - x - 2`.
                    """
                )
                st.markdown("**Method formula**")
                st.latex(CONFIG["formula"])

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")

                for condition in CONFIG["conditions"]:
                    st.markdown(f"- {condition}")

                st.info(
                    "A small approximation change alone does not prove that "
                    "the result is a root. This solver checks both the step "
                    "size and the function residual."
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">Function f(x)</div>',
                    unsafe_allow_html=True,
                )
                function_text = st.text_input(
                    "Function f(x)",
                    value=CONFIG["default_eq"],
                    placeholder="Example: x**3 - x - 2",
                    label_visibility="collapsed",
                    key="secant_function",
                )

                first_row = st.columns(2)

                with first_row[0]:
                    st.markdown(
                        '<div class="input-label-ui">Initial approximation x₀</div>',
                        unsafe_allow_html=True,
                    )
                    initial_x0 = st.number_input(
                        "Initial approximation x0",
                        value=DEFAULT_X0,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="secant_x0",
                    )

                with first_row[1]:
                    st.markdown(
                        '<div class="input-label-ui">Initial approximation x₁</div>',
                        unsafe_allow_html=True,
                    )
                    initial_x1 = st.number_input(
                        "Initial approximation x1",
                        value=DEFAULT_X1,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="secant_x1",
                    )

                second_row = st.columns(2)

                with second_row[0]:
                    st.markdown(
                        '<div class="input-label-ui">Tolerance</div>',
                        unsafe_allow_html=True,
                    )
                    tolerance = st.number_input(
                        "Tolerance",
                        min_value=MIN_TOLERANCE,
                        value=DEFAULT_TOLERANCE,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="secant_tolerance",
                    )

                with second_row[1]:
                    st.markdown(
                        '<div class="input-label-ui">Maximum iterations</div>',
                        unsafe_allow_html=True,
                    )
                    maximum_iterations = st.number_input(
                        "Maximum iterations",
                        min_value=MIN_ITERATIONS,
                        max_value=MAX_ITERATIONS,
                        value=DEFAULT_MAX_ITERATIONS,
                        step=1,
                        label_visibility="collapsed",
                        key="secant_max_iterations",
                    )

                st.markdown(
                    '<div class="input-label-ui">Known exact root — optional</div>',
                    unsafe_allow_html=True,
                )
                exact_root_text = st.text_input(
                    "Known exact root",
                    value="",
                    placeholder="Example: 1.5213797068 or sqrt(2)",
                    label_visibility="collapsed",
                    key="secant_exact_root",
                )

                st.caption(
                    "The optional exact root is used only to verify the "
                    "answer and estimate the empirical convergence order."
                )

                st.markdown("**Equation Preview**")
                try:
                    preview_expression, _, _, _ = parse_function(function_text)
                    st.latex(r"f(x)=" + sp.latex(preview_expression))
                except ValueError as error:
                    st.warning(str(error))

                current_signature = create_input_signature(
                    function_text=function_text,
                    initial_x0=initial_x0,
                    initial_x1=initial_x1,
                    tolerance=tolerance,
                    maximum_iterations=maximum_iterations,
                    exact_root_text=exact_root_text,
                )

                solve_column, reset_column = st.columns(2)

                with solve_column:
                    solve_button = st.button(
                        "Solve",
                        type="primary",
                        use_container_width=True,
                        key="secant_solve",
                    )

                with reset_column:
                    reset_button = st.button(
                        "Reset Result",
                        use_container_width=True,
                        key="secant_reset",
                    )

                if reset_button:
                    st.session_state.pop("secant_result", None)
                    st.session_state.pop("secant_excel_report", None)
                    st.session_state.pop("secant_excel_signature", None)
                    st.rerun()

                if solve_button:
                    st.session_state["secant_result"] = solve_secant_method(
                        function_text=function_text,
                        initial_x0_input=initial_x0,
                        initial_x1_input=initial_x1,
                        tolerance_input=tolerance,
                        maximum_iterations_input=maximum_iterations,
                        exact_root_text=exact_root_text,
                    )
                    st.session_state.pop("secant_excel_report", None)
                    st.session_state.pop("secant_excel_signature", None)
                    st.rerun()

                with st.expander("Example Inputs"):
                    st.code(
                        "Function: x**3 - x - 2\n"
                        "x0 = 1\n"
                        "x1 = 2\n"
                        "Tolerance = 1e-8\n"
                        "Known exact root = 1.52137970680457",
                        language=None,
                    )
                    st.code(
                        "Function: exp(-x) - x\n"
                        "x0 = 0\n"
                        "x1 = 1\n"
                        "Tolerance = 1e-10\n"
                        "Known exact root = 0.567143290409784",
                        language=None,
                    )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                result = st.session_state.get("secant_result")

                if result is None:
                    st.info(
                        "Enter the function and starting values, then click Solve."
                    )
                elif result.input_signature != current_signature:
                    st.info(
                        "The function or numerical settings have changed. "
                        "Click Solve to calculate a new result."
                    )
                else:
                    render_final_result(result)

        result = st.session_state.get("secant_result")
        result_is_current = (
            result is not None
            and result.input_signature == current_signature
        )

        if result_is_current and result.success:
            st.divider()
            render_method_summary(result)

            st.divider()
            render_iteration_table(result)

            st.divider()
            render_convergence_diagnostics(result)

            st.divider()
            render_graphs(result)

            st.divider()
            render_excel_download(result)

            st.divider()
            navigation_left, navigation_right = st.columns(2)

            with navigation_left:
                if st.button(
                    "Review Secant Method Lesson",
                    use_container_width=True,
                    key="secant_lesson_button",
                ):
                    st.switch_page("pages/Secant_Method.py")

            with navigation_right:
                if st.button(
                    "Back to Solver Menu",
                    use_container_width=True,
                    key="secant_menu_button",
                ):
                    st.switch_page("pages/Numerical_Solver.py")

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        f"""
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>{CONFIG['footer']}</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()