from __future__ import annotations

import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any, Callable
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css
from utilities.safe_math import safe_sympify


# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0**exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Show fixed notation, using scientific notation when needed."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6
    ):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


def format_number(
    value: float | int | None,
    decimals: int = 3,
) -> str:
    """Format a displayed numerical value."""

    return format_display_number(value, decimals)


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Explicit Euler Method"
DISPLAY_DECIMALS = 3
DEFAULT_ORDER = 1
DEFAULT_ODE = "x + y"
DEFAULT_X0 = 0.0
DEFAULT_Y0 = 1.0
DEFAULT_YP0 = 0.0
DEFAULT_YPP0 = 0.0
DEFAULT_X_END = 1.0
DEFAULT_STEPS = 10
MIN_STEPS = 1
MAX_STEPS = 10000
DEFAULT_CONVERGENCE_LEVELS = 5
MIN_CONVERGENCE_LEVELS = 3
MAX_CONVERGENCE_LEVELS = 7
MAX_REFINED_STEPS = 200000
ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
VALUE_MAGNITUDE_WARNING = 1.0e12
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

X_SYMBOL = sp.Symbol("x", real=True)
Y_SYMBOL = sp.Symbol("y", real=True)
YP_SYMBOL = sp.Symbol("yp", real=True)
YPP_SYMBOL = sp.Symbol("ypp", real=True)

ALLOWED_FUNCTION_NAMES = {
    "x": X_SYMBOL,
    "y": Y_SYMBOL,
    "yp": YP_SYMBOL,
    "v": YP_SYMBOL,
    "ypp": YPP_SYMBOL,
    "a": YPP_SYMBOL,
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}

EXACT_SOLUTION_NAMES = {
    key: value
    for key, value in ALLOWED_FUNCTION_NAMES.items()
    if key not in {"y", "yp", "v", "ypp", "a"}
}


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class EulerIteration:
    """One explicit Euler step for an ODE of order one, two, or three."""

    iteration: int
    x_n: float
    y_n: float
    yp_n: float | None
    ypp_n: float | None
    highest_derivative: float
    step_size: float
    x_next: float
    y_next: float
    yp_next: float | None
    ypp_next: float | None
    exact_y_next: float | None
    exact_yp_next: float | None
    exact_ypp_next: float | None
    signed_error_y: float | None
    absolute_error_y: float | None
    relative_error_y_percent: float | None
    local_truncation_error_y: float | None
    local_truncation_error_yp: float | None
    local_truncation_error_ypp: float | None
    operation_y: str
    operation_yp: str | None
    operation_ypp: str | None
    status: str


@dataclass(frozen=True)
class ConvergenceRecord:
    """One grid-refinement result for Euler's method."""

    level: int
    steps: int
    step_size: float
    final_y: float
    exact_final_y: float | None
    absolute_error_y: float | None
    successive_difference_y: float | None
    observed_order: float | None


@dataclass(frozen=True)
class EulerResult:
    """Complete Euler result shared by Streamlit and Excel renderers."""

    status: str
    success: bool
    method: str
    ode_order: int
    message: str
    stopping_reason: str
    ode_text: str
    ode_expression: sp.Expr | None
    exact_solution_text: str
    exact_solution_expression: sp.Expr | None
    x0: float | None
    y0: float | None
    yp0: float | None
    ypp0: float | None
    x_end: float | None
    steps: int
    step_size: float | None
    direction: str
    iterations: tuple[EulerIteration, ...]
    final_x: float | None
    final_y: float | None
    final_yp: float | None
    final_ypp: float | None
    exact_final_y: float | None
    exact_final_yp: float | None
    exact_final_ypp: float | None
    signed_final_error_y: float | None
    absolute_final_error_y: float | None
    relative_final_error_y_percent: float | None
    maximum_absolute_error_y: float | None
    rmse_y: float | None
    convergence_records: tuple[ConvergenceRecord, ...]
    latest_observed_order: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def current_report_datetime() -> datetime:
    """Return the current date and time in the report time zone."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def create_input_signature(
    ode_order: int,
    ode_text: str,
    x0: float,
    y0: float,
    yp0: float | None,
    ypp0: float | None,
    x_end: float,
    steps: int,
    exact_solution_text: str,
    convergence_levels: int,
) -> str:
    """Create a stable signature used to detect stale Streamlit results."""

    payload = "|".join(
        [
            str(int(ode_order)),
            str(ode_text).strip(),
            f"{float(x0):.17g}",
            f"{float(y0):.17g}",
            "" if yp0 is None else f"{float(yp0):.17g}",
            "" if ypp0 is None else f"{float(ypp0):.17g}",
            f"{float(x_end):.17g}",
            str(int(steps)),
            str(exact_solution_text).strip(),
            str(int(convergence_levels)),
        ]
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def validate_finite_real(value: Any, value_name: str) -> float:
    """Convert a scalar to a finite real float or raise ValueError."""

    try:
        complex_value = complex(value)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(f"{value_name} is not a valid numerical value.") from error

    if abs(complex_value.imag) > ZERO_TOLERANCE:
        raise ValueError(f"{value_name} is complex; a real value is required.")

    real_value = float(complex_value.real)
    if not math.isfinite(real_value):
        raise ValueError(f"{value_name} is NaN or infinity.")
    return real_value


def round_numeric_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Return a display copy with numerical columns rounded."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(DISPLAY_DECIMALS)
    return rounded


def order_label(order: int) -> str:
    """Return a readable differential-equation order label."""

    return {
        1: "First-order ODE",
        2: "Second-order ODE",
        3: "Third-order ODE",
    }[order]


def highest_derivative_label(order: int) -> str:
    """Return the highest derivative label."""

    return {1: "y′", 2: "y″", 3: "y‴"}[order]


def rhs_variables_text(order: int) -> str:
    """Return the variables accepted in the right-hand side."""

    return {
        1: "x, y",
        2: "x, y, yp",
        3: "x, y, yp, ypp",
    }[order]


# =============================================================================
# Safe symbolic parsing and evaluation
# =============================================================================
def _reject_unsupported_symbolic_constructs(
    expression: sp.Expr,
    allowed_symbols: set[sp.Symbol],
    expression_name: str,
) -> None:
    """Reject symbolic constructs inappropriate for numerical evaluation."""

    unsupported_atoms = (
        AppliedUndef,
        sp.Derivative,
        sp.Integral,
        sp.Sum,
        sp.Product,
        sp.Limit,
    )

    if expression.has(*unsupported_atoms):
        raise ValueError(
            f"{expression_name} contains unsupported symbolic operations."
        )

    if isinstance(expression, Relational) or expression.has(Relational):
        raise ValueError(
            f"{expression_name} must be an expression, not an equation or inequality."
        )

    unexpected_symbols = expression.free_symbols.difference(allowed_symbols)
    if unexpected_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unexpected_symbols))
        raise ValueError(
            f"{expression_name} contains unsupported variable(s): {names}."
        )

    if expression.has(sp.zoo, sp.oo, -sp.oo, sp.nan):
        raise ValueError(f"{expression_name} contains a non-finite symbolic value.")


def parse_ode_expression(
    ode_text: str,
    ode_order: int,
) -> tuple[sp.Expr, Callable[..., Any]]:
    """Parse the highest derivative equation for order one, two, or three."""

    if ode_order not in {1, 2, 3}:
        raise ValueError("The ODE order must be 1, 2, or 3.")

    if not isinstance(ode_text, str) or not ode_text.strip():
        raise ValueError("The differential-equation expression cannot be empty.")

    text = ode_text.strip().replace("^", "**")
    if "=" in text:
        raise ValueError(
            "Enter only the right-hand side F(...), without an equals sign."
        )

    try:
        expression = safe_sympify(
            text,
            locals=ALLOWED_FUNCTION_NAMES,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The differential equation has an invalid format. "
            "Examples: x + y, -y, -yp, or x + y + yp + ypp."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError("The differential equation could not be interpreted.")

    allowed_symbols = {
        1: {X_SYMBOL, Y_SYMBOL},
        2: {X_SYMBOL, Y_SYMBOL, YP_SYMBOL},
        3: {X_SYMBOL, Y_SYMBOL, YP_SYMBOL, YPP_SYMBOL},
    }[ode_order]

    _reject_unsupported_symbolic_constructs(
        expression,
        allowed_symbols=allowed_symbols,
        expression_name="The differential equation",
    )

    arguments = {
        1: (X_SYMBOL, Y_SYMBOL),
        2: (X_SYMBOL, Y_SYMBOL, YP_SYMBOL),
        3: (X_SYMBOL, Y_SYMBOL, YP_SYMBOL, YPP_SYMBOL),
    }[ode_order]

    try:
        numerical_function = sp.lambdify(
            arguments,
            expression,
            modules=["numpy"],
        )
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The differential equation could not be converted to a numerical function."
        ) from error

    return expression, numerical_function


def parse_exact_solution(
    exact_solution_text: str,
) -> tuple[sp.Expr | None, Callable[[float], Any] | None]:
    """Parse an optional exact solution y(x)."""

    if not str(exact_solution_text).strip():
        return None, None

    text = str(exact_solution_text).strip().replace("^", "**")
    if "=" in text:
        left, right = text.split("=", maxsplit=1)
        if left.strip().lower() not in {"y", "y(x)"}:
            raise ValueError(
                "Enter the exact solution as an expression or as y = expression."
            )
        text = right.strip()

    try:
        expression = safe_sympify(
            text,
            locals=EXACT_SOLUTION_NAMES,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The exact solution has an invalid format. "
            "Example: 2*exp(x) - x - 1 or cos(x)."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError("The exact solution could not be interpreted.")

    _reject_unsupported_symbolic_constructs(
        expression,
        allowed_symbols={X_SYMBOL},
        expression_name="The exact solution",
    )

    try:
        numerical_function = sp.lambdify(
            X_SYMBOL,
            expression,
            modules=["numpy"],
        )
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The exact solution could not be converted to a numerical function."
        ) from error

    return expression, numerical_function


def evaluate_exact_function(
    numerical_function: Callable[[float], Any],
    x_value: float,
    description: str,
) -> float:
    """Evaluate an exact solution or derivative safely."""

    try:
        with np.errstate(all="raise"):
            raw_value = numerical_function(x_value)
    except (
        ValueError,
        TypeError,
        OverflowError,
        ZeroDivisionError,
        FloatingPointError,
    ) as error:
        raise ValueError(
            f"{description} is undefined at x = {x_value:.12g}. Reason: {error}"
        ) from error

    return validate_finite_real(raw_value, description)


def evaluate_highest_derivative(
    numerical_function: Callable[..., Any],
    ode_order: int,
    x_value: float,
    y_value: float,
    yp_value: float | None,
    ypp_value: float | None,
) -> float:
    """Evaluate F at one state of the converted first-order system."""

    try:
        with np.errstate(all="raise"):
            if ode_order == 1:
                raw_value = numerical_function(x_value, y_value)
            elif ode_order == 2:
                if yp_value is None:
                    raise ValueError("The state y′ is missing.")
                raw_value = numerical_function(x_value, y_value, yp_value)
            else:
                if yp_value is None or ypp_value is None:
                    raise ValueError("The states y′ and y″ are required.")
                raw_value = numerical_function(
                    x_value,
                    y_value,
                    yp_value,
                    ypp_value,
                )
    except (
        ValueError,
        TypeError,
        OverflowError,
        ZeroDivisionError,
        FloatingPointError,
    ) as error:
        raise ValueError(
            f"The differential equation is undefined at x = {x_value:.12g}. "
            f"Reason: {error}"
        ) from error

    return validate_finite_real(raw_value, f"F at x = {x_value:.12g}")


def build_exact_functions(
    exact_expression: sp.Expr | None,
) -> tuple[
    Callable[[float], Any] | None,
    Callable[[float], Any] | None,
    Callable[[float], Any] | None,
    Callable[[float], Any] | None,
]:
    """Create exact y, y′, y″, and y‴ functions."""

    if exact_expression is None:
        return None, None, None, None

    expressions = [
        exact_expression,
        sp.diff(exact_expression, X_SYMBOL, 1),
        sp.diff(exact_expression, X_SYMBOL, 2),
        sp.diff(exact_expression, X_SYMBOL, 3),
    ]
    functions = tuple(
        sp.lambdify(X_SYMBOL, expression, modules=["numpy"])
        for expression in expressions
    )
    return functions  # type: ignore[return-value]


def verify_exact_solution(
    ode_order: int,
    ode_expression: sp.Expr,
    exact_expression: sp.Expr,
    x0: float,
    y0: float,
    yp0: float | None,
    ypp0: float | None,
    x_end: float,
) -> None:
    """Verify all required initial conditions and the ODE residual."""

    exact_functions = build_exact_functions(exact_expression)
    exact_y_function = exact_functions[0]
    exact_yp_function = exact_functions[1]
    exact_ypp_function = exact_functions[2]
    assert exact_y_function is not None

    exact_y0 = evaluate_exact_function(exact_y_function, x0, "Exact y(x0)")
    if abs(exact_y0 - y0) > 1.0e-8 * max(1.0, abs(exact_y0), abs(y0)):
        raise ValueError(
            "The exact solution does not satisfy the supplied y(x0) value."
        )

    exact_yp_expression = sp.diff(exact_expression, X_SYMBOL, 1)
    exact_ypp_expression = sp.diff(exact_expression, X_SYMBOL, 2)

    if ode_order >= 2:
        if yp0 is None or exact_yp_function is None:
            raise ValueError("The initial value y′(x0) is required.")
        exact_yp0 = evaluate_exact_function(
            exact_yp_function,
            x0,
            "Exact y′(x0)",
        )
        if abs(exact_yp0 - yp0) > 1.0e-8 * max(
            1.0,
            abs(exact_yp0),
            abs(yp0),
        ):
            raise ValueError(
                "The exact solution does not satisfy the supplied y′(x0) value."
            )

    if ode_order >= 3:
        if ypp0 is None or exact_ypp_function is None:
            raise ValueError("The initial value y″(x0) is required.")
        exact_ypp0 = evaluate_exact_function(
            exact_ypp_function,
            x0,
            "Exact y″(x0)",
        )
        if abs(exact_ypp0 - ypp0) > 1.0e-8 * max(
            1.0,
            abs(exact_ypp0),
            abs(ypp0),
        ):
            raise ValueError(
                "The exact solution does not satisfy the supplied y″(x0) value."
            )

    exact_highest_expression = sp.diff(
        exact_expression,
        X_SYMBOL,
        ode_order,
    )
    substituted_rhs = ode_expression.subs(
        {
            Y_SYMBOL: exact_expression,
            YP_SYMBOL: exact_yp_expression,
            YPP_SYMBOL: exact_ypp_expression,
        }
    )
    residual_expression = sp.simplify(
        exact_highest_expression - substituted_rhs
    )

    if residual_expression == 0:
        return

    residual_function = sp.lambdify(
        X_SYMBOL,
        residual_expression,
        modules=["numpy"],
    )
    residual_values = [
        abs(
            evaluate_exact_function(
                residual_function,
                float(point),
                "Exact-solution residual",
            )
        )
        for point in np.linspace(x0, x_end, 7)
    ]

    if max(residual_values) > 1.0e-7:
        raise ValueError(
            "The supplied exact solution does not satisfy the differential equation."
        )


# =============================================================================
# Euler algorithm
# =============================================================================
def run_euler_grid(
    ode_function: Callable[..., Any],
    ode_order: int,
    x0: float,
    y0: float,
    yp0: float | None,
    ypp0: float | None,
    x_end: float,
    steps: int,
    exact_expression: sp.Expr | None = None,
    store_history: bool = True,
) -> tuple[
    tuple[float, float | None, float | None],
    tuple[EulerIteration, ...],
]:
    """Run explicit Euler on the equivalent first-order system."""

    if steps < 1:
        raise ValueError("The number of steps must be at least 1.")

    step_size = (x_end - x0) / steps
    if not math.isfinite(step_size) or abs(step_size) <= ZERO_TOLERANCE:
        raise ValueError("The computed step size is zero or numerically unusable.")

    current_x = float(x0)
    current_y = float(y0)
    current_yp = None if ode_order == 1 else float(yp0)
    current_ypp = None if ode_order < 3 else float(ypp0)

    exact_y_function, exact_yp_function, exact_ypp_function, _ = (
        build_exact_functions(exact_expression)
    )
    history: list[EulerIteration] = []

    for iteration in range(1, steps + 1):
        highest_derivative = evaluate_highest_derivative(
            numerical_function=ode_function,
            ode_order=ode_order,
            x_value=current_x,
            y_value=current_y,
            yp_value=current_yp,
            ypp_value=current_ypp,
        )

        next_x = x0 + iteration * step_size

        if ode_order == 1:
            next_y = current_y + step_size * highest_derivative
            next_yp = None
            next_ypp = None
            operation_y = (
                f"y_{iteration} = {current_y:.12g} + "
                f"({step_size:.12g})({highest_derivative:.12g}) "
                f"= {next_y:.12g}"
            )
            operation_yp = None
            operation_ypp = None

        elif ode_order == 2:
            assert current_yp is not None
            next_y = current_y + step_size * current_yp
            next_yp = current_yp + step_size * highest_derivative
            next_ypp = None
            operation_y = (
                f"y_{iteration} = {current_y:.12g} + "
                f"({step_size:.12g})({current_yp:.12g}) = {next_y:.12g}"
            )
            operation_yp = (
                f"yp_{iteration} = {current_yp:.12g} + "
                f"({step_size:.12g})({highest_derivative:.12g}) "
                f"= {next_yp:.12g}"
            )
            operation_ypp = None

        else:
            assert current_yp is not None
            assert current_ypp is not None
            next_y = current_y + step_size * current_yp
            next_yp = current_yp + step_size * current_ypp
            next_ypp = current_ypp + step_size * highest_derivative
            operation_y = (
                f"y_{iteration} = {current_y:.12g} + "
                f"({step_size:.12g})({current_yp:.12g}) = {next_y:.12g}"
            )
            operation_yp = (
                f"yp_{iteration} = {current_yp:.12g} + "
                f"({step_size:.12g})({current_ypp:.12g}) = {next_yp:.12g}"
            )
            operation_ypp = (
                f"ypp_{iteration} = {current_ypp:.12g} + "
                f"({step_size:.12g})({highest_derivative:.12g}) "
                f"= {next_ypp:.12g}"
            )

        numerical_values = [next_x, next_y]
        if next_yp is not None:
            numerical_values.append(next_yp)
        if next_ypp is not None:
            numerical_values.append(next_ypp)

        if not all(math.isfinite(float(value)) for value in numerical_values):
            raise ValueError(
                f"Euler's method produced a non-finite value at iteration {iteration}."
            )

        exact_y_next: float | None = None
        exact_yp_next: float | None = None
        exact_ypp_next: float | None = None
        signed_error_y: float | None = None
        absolute_error_y: float | None = None
        relative_error_y_percent: float | None = None
        local_truncation_error_y: float | None = None
        local_truncation_error_yp: float | None = None
        local_truncation_error_ypp: float | None = None

        if exact_y_function is not None:
            exact_y_current = evaluate_exact_function(
                exact_y_function,
                current_x,
                "Exact y",
            )
            exact_y_next = evaluate_exact_function(
                exact_y_function,
                next_x,
                "Exact y",
            )
            signed_error_y = next_y - exact_y_next
            absolute_error_y = abs(signed_error_y)

            if abs(exact_y_next) > RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
                relative_error_y_percent = (
                    absolute_error_y / abs(exact_y_next)
                ) * 100.0

            if ode_order == 1:
                exact_state_slope_y = evaluate_highest_derivative(
                    ode_function,
                    1,
                    current_x,
                    exact_y_current,
                    None,
                    None,
                )
                exact_one_step_y = (
                    exact_y_current + step_size * exact_state_slope_y
                )
                local_truncation_error_y = exact_y_next - exact_one_step_y

            else:
                assert exact_yp_function is not None
                exact_yp_current = evaluate_exact_function(
                    exact_yp_function,
                    current_x,
                    "Exact y′",
                )
                exact_yp_next = evaluate_exact_function(
                    exact_yp_function,
                    next_x,
                    "Exact y′",
                )
                local_truncation_error_y = exact_y_next - (
                    exact_y_current + step_size * exact_yp_current
                )

                if ode_order == 2:
                    exact_highest_current = evaluate_highest_derivative(
                        ode_function,
                        2,
                        current_x,
                        exact_y_current,
                        exact_yp_current,
                        None,
                    )
                    local_truncation_error_yp = exact_yp_next - (
                        exact_yp_current + step_size * exact_highest_current
                    )

                else:
                    assert exact_ypp_function is not None
                    exact_ypp_current = evaluate_exact_function(
                        exact_ypp_function,
                        current_x,
                        "Exact y″",
                    )
                    exact_ypp_next = evaluate_exact_function(
                        exact_ypp_function,
                        next_x,
                        "Exact y″",
                    )
                    local_truncation_error_yp = exact_yp_next - (
                        exact_yp_current + step_size * exact_ypp_current
                    )
                    exact_highest_current = evaluate_highest_derivative(
                        ode_function,
                        3,
                        current_x,
                        exact_y_current,
                        exact_yp_current,
                        exact_ypp_current,
                    )
                    local_truncation_error_ypp = exact_ypp_next - (
                        exact_ypp_current + step_size * exact_highest_current
                    )

        if store_history:
            history.append(
                EulerIteration(
                    iteration=iteration,
                    x_n=float(current_x),
                    y_n=float(current_y),
                    yp_n=current_yp,
                    ypp_n=current_ypp,
                    highest_derivative=float(highest_derivative),
                    step_size=float(step_size),
                    x_next=float(next_x),
                    y_next=float(next_y),
                    yp_next=next_yp,
                    ypp_next=next_ypp,
                    exact_y_next=exact_y_next,
                    exact_yp_next=exact_yp_next,
                    exact_ypp_next=exact_ypp_next,
                    signed_error_y=signed_error_y,
                    absolute_error_y=absolute_error_y,
                    relative_error_y_percent=relative_error_y_percent,
                    local_truncation_error_y=local_truncation_error_y,
                    local_truncation_error_yp=local_truncation_error_yp,
                    local_truncation_error_ypp=local_truncation_error_ypp,
                    operation_y=operation_y,
                    operation_yp=operation_yp,
                    operation_ypp=operation_ypp,
                    status="Completed",
                )
            )

        current_x = float(next_x)
        current_y = float(next_y)
        current_yp = None if next_yp is None else float(next_yp)
        current_ypp = None if next_ypp is None else float(next_ypp)

    return (
        current_y,
        current_yp,
        current_ypp,
    ), tuple(history)


def build_convergence_analysis(
    ode_function: Callable[..., Any],
    ode_order: int,
    exact_expression: sp.Expr | None,
    x0: float,
    y0: float,
    yp0: float | None,
    ypp0: float | None,
    x_end: float,
    base_steps: int,
    levels: int,
) -> tuple[ConvergenceRecord, ...]:
    """Refine the grid and estimate Euler's first-order convergence."""

    raw_records: list[dict[str, float | int | None]] = []
    exact_y_function, _, _, _ = build_exact_functions(exact_expression)
    exact_final_y = (
        evaluate_exact_function(exact_y_function, x_end, "Exact final y")
        if exact_y_function is not None
        else None
    )

    for level in range(1, levels + 1):
        refined_steps = base_steps * (2 ** (level - 1))
        final_state, _ = run_euler_grid(
            ode_function=ode_function,
            ode_order=ode_order,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            steps=refined_steps,
            exact_expression=None,
            store_history=False,
        )
        final_y = final_state[0]
        absolute_error_y = (
            abs(final_y - exact_final_y)
            if exact_final_y is not None
            else None
        )

        raw_records.append(
            {
                "level": level,
                "steps": refined_steps,
                "step_size": (x_end - x0) / refined_steps,
                "final_y": final_y,
                "exact_final_y": exact_final_y,
                "absolute_error_y": absolute_error_y,
                "successive_difference_y": None,
                "observed_order": None,
            }
        )

    for index in range(1, len(raw_records)):
        raw_records[index]["successive_difference_y"] = abs(
            float(raw_records[index]["final_y"])
            - float(raw_records[index - 1]["final_y"])
        )

    if exact_final_y is not None:
        for index in range(1, len(raw_records)):
            previous_error = raw_records[index - 1]["absolute_error_y"]
            current_error = raw_records[index]["absolute_error_y"]
            if (
                previous_error is not None
                and current_error is not None
                and float(previous_error) > ZERO_TOLERANCE
                and float(current_error) > ZERO_TOLERANCE
            ):
                raw_records[index]["observed_order"] = math.log(
                    float(previous_error) / float(current_error),
                    2.0,
                )
    else:
        for index in range(2, len(raw_records)):
            previous_difference = raw_records[index - 1][
                "successive_difference_y"
            ]
            current_difference = raw_records[index]["successive_difference_y"]
            if (
                previous_difference is not None
                and current_difference is not None
                and float(previous_difference) > ZERO_TOLERANCE
                and float(current_difference) > ZERO_TOLERANCE
            ):
                raw_records[index]["observed_order"] = math.log(
                    float(previous_difference) / float(current_difference),
                    2.0,
                )

    return tuple(ConvergenceRecord(**record) for record in raw_records)


def error_result(
    message: str,
    ode_order: int = 1,
    input_signature: str = "",
) -> EulerResult:
    """Create a consistent failed result."""

    return EulerResult(
        status="error",
        success=False,
        method=METHOD_NAME,
        ode_order=ode_order,
        message=message,
        stopping_reason=(
            "Execution stopped because input validation or numerical evaluation failed."
        ),
        ode_text="",
        ode_expression=None,
        exact_solution_text="",
        exact_solution_expression=None,
        x0=None,
        y0=None,
        yp0=None,
        ypp0=None,
        x_end=None,
        steps=0,
        step_size=None,
        direction="Not available",
        iterations=(),
        final_x=None,
        final_y=None,
        final_yp=None,
        final_ypp=None,
        exact_final_y=None,
        exact_final_yp=None,
        exact_final_ypp=None,
        signed_final_error_y=None,
        absolute_final_error_y=None,
        relative_final_error_y_percent=None,
        maximum_absolute_error_y=None,
        rmse_y=None,
        convergence_records=(),
        latest_observed_order=None,
        warnings=(),
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


def solve_euler_method(
    ode_order_input: Any,
    ode_text: str,
    x0_input: Any,
    y0_input: Any,
    yp0_input: Any = None,
    ypp0_input: Any = None,
    x_end_input: Any = DEFAULT_X_END,
    steps_input: Any = DEFAULT_STEPS,
    exact_solution_text: str = "",
    convergence_levels_input: Any = DEFAULT_CONVERGENCE_LEVELS,
) -> EulerResult:
    """Validate inputs, solve the IVP, and analyze errors and convergence."""

    ode_order = 1

    try:
        ode_order = int(ode_order_input)
        if ode_order not in {1, 2, 3}:
            raise ValueError("The ODE order must be 1, 2, or 3.")

        x0 = validate_finite_real(x0_input, "Initial x value x0")
        y0 = validate_finite_real(y0_input, "Initial value y(x0)")
        x_end = validate_finite_real(x_end_input, "Final x value")

        yp0 = (
            validate_finite_real(yp0_input, "Initial value y′(x0)")
            if ode_order >= 2
            else None
        )
        ypp0 = (
            validate_finite_real(ypp0_input, "Initial value y″(x0)")
            if ode_order >= 3
            else None
        )

        try:
            steps = int(steps_input)
        except (TypeError, ValueError, OverflowError) as error:
            raise ValueError("The number of steps must be an integer.") from error

        try:
            convergence_levels = int(convergence_levels_input)
        except (TypeError, ValueError, OverflowError) as error:
            raise ValueError("Convergence levels must be an integer.") from error

        if not MIN_STEPS <= steps <= MAX_STEPS:
            raise ValueError(
                f"The number of steps must be between {MIN_STEPS} and {MAX_STEPS}."
            )

        if not MIN_CONVERGENCE_LEVELS <= convergence_levels <= MAX_CONVERGENCE_LEVELS:
            raise ValueError(
                "Convergence levels must be between "
                f"{MIN_CONVERGENCE_LEVELS} and {MAX_CONVERGENCE_LEVELS}."
            )

        refined_steps = steps * (2 ** (convergence_levels - 1))
        if refined_steps > MAX_REFINED_STEPS:
            raise ValueError(
                "The requested convergence analysis is too large. "
                "Reduce the number of steps or convergence levels."
            )

        if abs(x_end - x0) <= ZERO_TOLERANCE:
            raise ValueError("The final x value must be different from x0.")

        input_signature = create_input_signature(
            ode_order=ode_order,
            ode_text=ode_text,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            steps=steps,
            exact_solution_text=exact_solution_text,
            convergence_levels=convergence_levels,
        )

        ode_expression, ode_function = parse_ode_expression(
            ode_text,
            ode_order,
        )
        exact_expression, _ = parse_exact_solution(exact_solution_text)

        warnings: list[str] = []

        if exact_expression is not None:
            verify_exact_solution(
                ode_order=ode_order,
                ode_expression=ode_expression,
                exact_expression=exact_expression,
                x0=x0,
                y0=y0,
                yp0=yp0,
                ypp0=ypp0,
                x_end=x_end,
            )

        final_state, iterations = run_euler_grid(
            ode_function=ode_function,
            ode_order=ode_order,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            steps=steps,
            exact_expression=exact_expression,
            store_history=True,
        )
        final_y, final_yp, final_ypp = final_state

        exact_final_y: float | None = None
        exact_final_yp: float | None = None
        exact_final_ypp: float | None = None
        signed_final_error_y: float | None = None
        absolute_final_error_y: float | None = None
        relative_final_error_y_percent: float | None = None
        maximum_absolute_error_y: float | None = None
        rmse_y: float | None = None

        exact_y_function, exact_yp_function, exact_ypp_function, _ = (
            build_exact_functions(exact_expression)
        )

        if exact_y_function is not None:
            exact_final_y = evaluate_exact_function(
                exact_y_function,
                x_end,
                "Exact final y",
            )
            signed_final_error_y = final_y - exact_final_y
            absolute_final_error_y = abs(signed_final_error_y)

            if abs(exact_final_y) > RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
                relative_final_error_y_percent = (
                    absolute_final_error_y / abs(exact_final_y)
                ) * 100.0

            if ode_order >= 2 and exact_yp_function is not None:
                exact_final_yp = evaluate_exact_function(
                    exact_yp_function,
                    x_end,
                    "Exact final y′",
                )

            if ode_order >= 3 and exact_ypp_function is not None:
                exact_final_ypp = evaluate_exact_function(
                    exact_ypp_function,
                    x_end,
                    "Exact final y″",
                )

            global_errors = np.asarray(
                [
                    record.signed_error_y
                    for record in iterations
                    if record.signed_error_y is not None
                ],
                dtype=float,
            )
            if global_errors.size:
                maximum_absolute_error_y = float(np.max(np.abs(global_errors)))
                rmse_y = float(np.sqrt(np.mean(global_errors**2)))

        convergence_records = build_convergence_analysis(
            ode_function=ode_function,
            ode_order=ode_order,
            exact_expression=exact_expression,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            base_steps=steps,
            levels=convergence_levels,
        )

        observed_orders = [
            record.observed_order
            for record in convergence_records
            if record.observed_order is not None
            and math.isfinite(record.observed_order)
        ]
        latest_observed_order = (
            observed_orders[-1] if observed_orders else None
        )

        state_magnitudes: list[float] = []
        for record in iterations:
            state_magnitudes.append(abs(record.y_next))
            if record.yp_next is not None:
                state_magnitudes.append(abs(record.yp_next))
            if record.ypp_next is not None:
                state_magnitudes.append(abs(record.ypp_next))

        if state_magnitudes and max(state_magnitudes) >= VALUE_MAGNITUDE_WARNING:
            warnings.append(
                "The numerical state reached a very large magnitude. "
                "Explicit Euler may be unstable for this step size or problem."
            )

        if latest_observed_order is not None and latest_observed_order < 0.5:
            warnings.append(
                "The observed convergence is weaker than the expected first-order "
                "behavior. Reduce the step size and inspect stability."
            )

        if exact_expression is None:
            warnings.append(
                "No exact solution was supplied. True global errors are unavailable; "
                "successive final-value differences are used for convergence."
            )

        direction = (
            "Forward integration" if x_end > x0 else "Backward integration"
        )
        step_size = (x_end - x0) / steps

        return EulerResult(
            status="success",
            success=True,
            method=METHOD_NAME,
            ode_order=ode_order,
            message="Execution completed successfully.",
            stopping_reason=(
                "The requested number of uniform Euler steps was completed "
                "and the final grid point was reached."
            ),
            ode_text=str(ode_text).strip(),
            ode_expression=ode_expression,
            exact_solution_text=str(exact_solution_text).strip(),
            exact_solution_expression=exact_expression,
            x0=x0,
            y0=y0,
            yp0=yp0,
            ypp0=ypp0,
            x_end=x_end,
            steps=steps,
            step_size=step_size,
            direction=direction,
            iterations=iterations,
            final_x=x_end,
            final_y=final_y,
            final_yp=final_yp,
            final_ypp=final_ypp,
            exact_final_y=exact_final_y,
            exact_final_yp=exact_final_yp,
            exact_final_ypp=exact_final_ypp,
            signed_final_error_y=signed_final_error_y,
            absolute_final_error_y=absolute_final_error_y,
            relative_final_error_y_percent=relative_final_error_y_percent,
            maximum_absolute_error_y=maximum_absolute_error_y,
            rmse_y=rmse_y,
            convergence_records=convergence_records,
            latest_observed_order=latest_observed_order,
            warnings=tuple(warnings),
            input_signature=input_signature,
            execution_datetime=current_report_datetime(),
        )

    except (ValueError, TypeError, ArithmeticError, OverflowError) as error:
        return error_result(
            message=str(error),
            ode_order=ode_order,
            input_signature="",
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def iterations_dataframe(result: EulerResult) -> pd.DataFrame:
    """Return the complete Euler iteration table."""

    rows: list[dict[str, Any]] = []

    for record in result.iterations:
        row: dict[str, Any] = {
            "Iteration": record.iteration,
            "x_n": record.x_n,
            "y_n": record.y_n,
            "h": record.step_size,
            f"{highest_derivative_label(result.ode_order)}(state_n)": (
                record.highest_derivative
            ),
            "x_(n+1)": record.x_next,
            "y_(n+1)": record.y_next,
            "Exact y_(n+1)": record.exact_y_next,
            "Signed Global Error y": record.signed_error_y,
            "Absolute Global Error y": record.absolute_error_y,
            "Relative Error y (%)": record.relative_error_y_percent,
            "Local Truncation Error y": record.local_truncation_error_y,
            "y Update": record.operation_y,
            "Status": record.status,
        }

        if result.ode_order >= 2:
            row.update(
                {
                    "yp_n": record.yp_n,
                    "yp_(n+1)": record.yp_next,
                    "Exact yp_(n+1)": record.exact_yp_next,
                    "Local Truncation Error yp": (
                        record.local_truncation_error_yp
                    ),
                    "yp Update": record.operation_yp,
                }
            )

        if result.ode_order >= 3:
            row.update(
                {
                    "ypp_n": record.ypp_n,
                    "ypp_(n+1)": record.ypp_next,
                    "Exact ypp_(n+1)": record.exact_ypp_next,
                    "Local Truncation Error ypp": (
                        record.local_truncation_error_ypp
                    ),
                    "ypp Update": record.operation_ypp,
                }
            )

        rows.append(row)

    return pd.DataFrame(rows)


def solution_values_dataframe(result: EulerResult) -> pd.DataFrame:
    """Return every numerical solution point including the initial state."""

    exact_y_function, exact_yp_function, exact_ypp_function, _ = (
        build_exact_functions(result.exact_solution_expression)
    )

    initial_exact_y = (
        evaluate_exact_function(exact_y_function, float(result.x0), "Exact y")
        if exact_y_function is not None and result.x0 is not None
        else None
    )
    initial_exact_yp = (
        evaluate_exact_function(exact_yp_function, float(result.x0), "Exact y′")
        if exact_yp_function is not None
        and result.x0 is not None
        and result.ode_order >= 2
        else None
    )
    initial_exact_ypp = (
        evaluate_exact_function(exact_ypp_function, float(result.x0), "Exact y″")
        if exact_ypp_function is not None
        and result.x0 is not None
        and result.ode_order >= 3
        else None
    )

    rows: list[dict[str, Any]] = [
        {
            "Point": 0,
            "x": result.x0,
            "Euler y": result.y0,
            "Euler yp": result.yp0,
            "Euler ypp": result.ypp0,
            "Exact y": initial_exact_y,
            "Exact yp": initial_exact_yp,
            "Exact ypp": initial_exact_ypp,
            "Absolute Error y": (
                0.0 if result.exact_solution_expression is not None else None
            ),
        }
    ]

    for record in result.iterations:
        rows.append(
            {
                "Point": record.iteration,
                "x": record.x_next,
                "Euler y": record.y_next,
                "Euler yp": record.yp_next,
                "Euler ypp": record.ypp_next,
                "Exact y": record.exact_y_next,
                "Exact yp": record.exact_yp_next,
                "Exact ypp": record.exact_ypp_next,
                "Absolute Error y": record.absolute_error_y,
            }
        )

    dataframe = pd.DataFrame(rows)
    columns_to_remove: list[str] = []
    if result.ode_order == 1:
        columns_to_remove.extend(["Euler yp", "Euler ypp", "Exact yp", "Exact ypp"])
    elif result.ode_order == 2:
        columns_to_remove.extend(["Euler ypp", "Exact ypp"])
    return dataframe.drop(columns=columns_to_remove)


def error_analysis_dataframe(result: EulerResult) -> pd.DataFrame:
    """Return global and local error information."""

    if result.exact_solution_expression is None:
        return pd.DataFrame(
            {
                "Message": [
                    "No exact solution was supplied; true global and local errors are unavailable."
                ]
            }
        )

    rows: list[dict[str, Any]] = []
    for record in result.iterations:
        row: dict[str, Any] = {
            "Iteration": record.iteration,
            "x": record.x_next,
            "Euler y": record.y_next,
            "Exact y": record.exact_y_next,
            "Signed Global Error y": record.signed_error_y,
            "Absolute Global Error y": record.absolute_error_y,
            "Relative Error y (%)": record.relative_error_y_percent,
            "Local Truncation Error y": record.local_truncation_error_y,
        }
        if result.ode_order >= 2:
            row["Local Truncation Error yp"] = record.local_truncation_error_yp
        if result.ode_order >= 3:
            row["Local Truncation Error ypp"] = record.local_truncation_error_ypp
        rows.append(row)
    return pd.DataFrame(rows)


def convergence_dataframe(result: EulerResult) -> pd.DataFrame:
    """Return grid-refinement convergence results."""

    return pd.DataFrame(
        [
            {
                "Level": record.level,
                "Steps": record.steps,
                "h": record.step_size,
                "Final y": record.final_y,
                "Exact Final y": record.exact_final_y,
                "Absolute Error y": record.absolute_error_y,
                "Successive Difference y": record.successive_difference_y,
                "Observed Order": record.observed_order,
            }
            for record in result.convergence_records
        ]
    )


def method_formula_dataframe(result: EulerResult) -> pd.DataFrame:
    """Return the converted system and Euler formulas."""

    rows: list[tuple[str, str]] = [
        (
            "Original IVP",
            f"{highest_derivative_label(result.ode_order)} = {result.ode_text}",
        ),
        ("Uniform Step Size", "h = (x_end - x0) / N"),
        ("Expected Global Error", "O(h)"),
        ("Local Truncation Error per Step", "O(h^2)"),
    ]

    if result.ode_order == 1:
        rows.extend(
            [
                ("System Equation", "u1 = y; u1' = F(x, u1)"),
                ("Euler Update", "y_(n+1) = y_n + h*F(x_n, y_n)"),
            ]
        )
    elif result.ode_order == 2:
        rows.extend(
            [
                ("State Definitions", "u1 = y; u2 = y'"),
                ("System Equation 1", "u1' = u2"),
                ("System Equation 2", "u2' = F(x, u1, u2)"),
                ("Euler y Update", "y_(n+1) = y_n + h*yp_n"),
                ("Euler yp Update", "yp_(n+1) = yp_n + h*F_n"),
            ]
        )
    else:
        rows.extend(
            [
                ("State Definitions", "u1 = y; u2 = y'; u3 = y''"),
                ("System Equation 1", "u1' = u2"),
                ("System Equation 2", "u2' = u3"),
                ("System Equation 3", "u3' = F(x, u1, u2, u3)"),
                ("Euler y Update", "y_(n+1) = y_n + h*yp_n"),
                ("Euler yp Update", "yp_(n+1) = yp_n + h*ypp_n"),
                ("Euler ypp Update", "ypp_(n+1) = ypp_n + h*F_n"),
            ]
        )

    return pd.DataFrame(rows, columns=["Item", "Formula"])


# =============================================================================
# Scientific plots
# =============================================================================
def create_solution_figure(result: EulerResult) -> Figure:
    """Plot y and, where applicable, its state derivatives."""

    dataframe = solution_values_dataframe(result)
    figure, axis = plt.subplots(figsize=(10, 6))

    axis.plot(
        dataframe["x"],
        dataframe["Euler y"],
        marker="o",
        linewidth=2,
        label="Euler y",
    )

    if result.exact_solution_expression is not None:
        exact_y_function, _, _, _ = build_exact_functions(
            result.exact_solution_expression
        )
        assert exact_y_function is not None
        x_values = np.linspace(float(result.x0), float(result.x_end), 500)
        y_values = np.asarray(
            [
                evaluate_exact_function(exact_y_function, value, "Exact y")
                for value in x_values
            ],
            dtype=float,
        )
        axis.plot(x_values, y_values, linewidth=2, label="Exact y")

    axis.scatter(
        [result.x0],
        [result.y0],
        s=90,
        marker="s",
        label="Initial condition",
        zorder=5,
    )
    axis.scatter(
        [result.final_x],
        [result.final_y],
        s=110,
        marker="*",
        label=f"Final Euler y = {result.final_y:.6g}",
        zorder=6,
    )
    axis.axhline(0.0, linewidth=1)
    axis.axvline(0.0, linewidth=1)
    axis.set_title(f"Euler Numerical Solution — {order_label(result.ode_order)}")
    axis.set_xlabel("x")
    axis.set_ylabel("y")
    axis.grid(True)
    axis.legend()
    figure.tight_layout()
    return figure


def create_state_figure(result: EulerResult) -> Figure | None:
    """Plot the converted state variables for second- and third-order ODEs."""

    if result.ode_order == 1:
        return None

    dataframe = solution_values_dataframe(result)
    figure, axis = plt.subplots(figsize=(10, 6))
    axis.plot(dataframe["x"], dataframe["Euler y"], marker="o", label="y")
    axis.plot(dataframe["x"], dataframe["Euler yp"], marker="o", label="y′")
    if result.ode_order == 3:
        axis.plot(
            dataframe["x"],
            dataframe["Euler ypp"],
            marker="o",
            label="y″",
        )
    axis.set_title("Equivalent First-Order System States")
    axis.set_xlabel("x")
    axis.set_ylabel("State value")
    axis.grid(True)
    axis.legend()
    figure.tight_layout()
    return figure


def create_error_figure(result: EulerResult) -> Figure | None:
    """Plot pointwise absolute global error in y."""

    if result.exact_solution_expression is None:
        return None

    dataframe = error_analysis_dataframe(result)
    errors = np.maximum(
        dataframe["Absolute Global Error y"].to_numpy(dtype=float),
        np.finfo(float).tiny,
    )

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.semilogy(
        dataframe["x"],
        errors,
        marker="o",
        linewidth=2,
    )
    axis.set_title("Euler Global Error in y")
    axis.set_xlabel("x")
    axis.set_ylabel("Absolute global error (log scale)")
    axis.grid(True, which="both")
    figure.tight_layout()
    return figure


def create_convergence_figure(result: EulerResult) -> Figure:
    """Plot Euler convergence against absolute step size."""

    dataframe = convergence_dataframe(result)
    h_values = np.abs(dataframe["h"].to_numpy(dtype=float))

    if result.exact_solution_expression is not None:
        metric = dataframe["Absolute Error y"].to_numpy(dtype=float)
        label = "Absolute final error in y"
    else:
        metric = dataframe["Successive Difference y"].to_numpy(dtype=float)
        label = "Successive final-y difference"

    valid = (
        np.isfinite(metric)
        & (metric > 0.0)
        & np.isfinite(h_values)
        & (h_values > 0.0)
    )

    figure, axis = plt.subplots(figsize=(10, 6))
    if np.any(valid):
        valid_h = h_values[valid]
        valid_metric = metric[valid]
        axis.loglog(
            valid_h,
            valid_metric,
            marker="o",
            linewidth=2,
            label=label,
        )
        reference_line = valid_metric[-1] * (valid_h / valid_h[-1])
        axis.loglog(
            valid_h,
            reference_line,
            linestyle="--",
            label="First-order reference O(h)",
        )
    else:
        axis.text(
            0.5,
            0.5,
            "Insufficient nonzero data for a log-log convergence plot.",
            ha="center",
            va="center",
            transform=axis.transAxes,
        )

    axis.set_title("Euler Grid-Refinement Convergence")
    axis.set_xlabel("|h|")
    axis.set_ylabel("Error indicator")
    axis.grid(True, which="both")
    axis.legend()
    figure.tight_layout()
    return figure


def figure_to_png_bytes(figure: Figure) -> bytes:
    """Serialize a Matplotlib figure as PNG bytes."""

    output = BytesIO()
    figure.savefig(output, format="png", dpi=180, bbox_inches="tight")
    output.seek(0)
    return output.getvalue()


# =============================================================================
# Excel report
# =============================================================================
def format_workbook(workbook: Any) -> None:
    """Apply readable formatting to all workbook sheets."""

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            maximum_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 12),
                55,
            )


def generate_excel_report(result: EulerResult) -> bytes:
    """Create a complete formatted XLSX report with tables, charts, and images."""

    if not result.success:
        raise ValueError("Only successful Euler results can be exported.")

    summary_rows: list[tuple[str, Any]] = [
        ("Method", result.method),
        ("ODE Order", result.ode_order),
        ("Problem Type", order_label(result.ode_order)),
        ("Differential Equation", f"{highest_derivative_label(result.ode_order)} = {result.ode_text}"),
        ("Initial x", result.x0),
        ("Initial y", result.y0),
        ("Initial y'", result.yp0),
        ("Initial y''", result.ypp0),
        ("Final x", result.x_end),
        ("Direction", result.direction),
        ("Number of Steps", result.steps),
        ("Step Size h", result.step_size),
        ("Final Euler y", result.final_y),
        ("Final Euler y'", result.final_yp),
        ("Final Euler y''", result.final_ypp),
        ("Exact Solution", result.exact_solution_text or "Not supplied"),
        ("Exact Final y", result.exact_final_y),
        ("Exact Final y'", result.exact_final_yp),
        ("Exact Final y''", result.exact_final_ypp),
        ("Signed Final Error y", result.signed_final_error_y),
        ("Absolute Final Error y", result.absolute_final_error_y),
        ("Relative Final Error y (%)", result.relative_final_error_y_percent),
        ("Maximum Absolute Error y", result.maximum_absolute_error_y),
        ("RMSE y", result.rmse_y),
        ("Expected Global Order", 1),
        ("Expected Local Truncation Order", 2),
        ("Latest Observed Order", result.latest_observed_order),
        ("Warnings", " | ".join(result.warnings) if result.warnings else "None"),
        ("Stopping Reason", result.stopping_reason),
        (
            "Execution Date",
            result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
        ),
    ]
    summary_dataframe = pd.DataFrame(
        summary_rows,
        columns=["Property", "Value"],
    )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_dataframe.to_excel(writer, sheet_name="Summary", index=False)
        method_formula_dataframe(result).to_excel(
            writer,
            sheet_name="System Formulation",
            index=False,
        )
        iterations_dataframe(result).to_excel(
            writer,
            sheet_name="Iteration Results",
            index=False,
        )
        solution_values_dataframe(result).to_excel(
            writer,
            sheet_name="Solution Values",
            index=False,
        )
        error_analysis_dataframe(result).to_excel(
            writer,
            sheet_name="Error Analysis",
            index=False,
        )
        convergence_dataframe(result).to_excel(
            writer,
            sheet_name="Convergence Analysis",
            index=False,
        )
        pd.DataFrame(
            {"Description": ["Matplotlib plots embedded below."]}
        ).to_excel(writer, sheet_name="Plots", index=False)

        workbook = writer.book
        format_workbook(workbook)

        solution_sheet = workbook["Solution Values"]
        solution_dataframe = solution_values_dataframe(result)
        solution_chart = LineChart()
        solution_chart.title = "Euler Numerical Solution y(x)"
        solution_chart.y_axis.title = "y"
        solution_chart.x_axis.title = "x"
        solution_chart.height = 8
        solution_chart.width = 15
        x_reference = Reference(
            solution_sheet,
            min_col=2,
            min_row=2,
            max_row=len(solution_dataframe) + 1,
        )
        y_reference = Reference(
            solution_sheet,
            min_col=3,
            min_row=1,
            max_row=len(solution_dataframe) + 1,
        )
        solution_chart.add_data(y_reference, titles_from_data=True)
        solution_chart.set_categories(x_reference)
        solution_sheet.add_chart(solution_chart, "K2")

        convergence_sheet = workbook["Convergence Analysis"]
        convergence_data = convergence_dataframe(result)
        if len(convergence_data) > 1:
            convergence_chart = LineChart()
            convergence_chart.title = "Grid-Refinement Convergence"
            convergence_chart.y_axis.title = (
                "Absolute Error y"
                if result.exact_solution_expression is not None
                else "Successive Difference y"
            )
            convergence_chart.x_axis.title = "Steps"
            convergence_chart.height = 8
            convergence_chart.width = 15
            category_reference = Reference(
                convergence_sheet,
                min_col=2,
                min_row=2,
                max_row=len(convergence_data) + 1,
            )
            metric_column = (
                6 if result.exact_solution_expression is not None else 7
            )
            metric_reference = Reference(
                convergence_sheet,
                min_col=metric_column,
                min_row=1,
                max_row=len(convergence_data) + 1,
            )
            convergence_chart.add_data(
                metric_reference,
                titles_from_data=True,
            )
            convergence_chart.set_categories(category_reference)
            convergence_sheet.add_chart(convergence_chart, "J2")

        plots_sheet = workbook["Plots"]
        plot_figures: list[tuple[str, Figure | None]] = [
            ("Solution", create_solution_figure(result)),
            ("System States", create_state_figure(result)),
            ("Global Error", create_error_figure(result)),
            ("Convergence", create_convergence_figure(result)),
        ]
        anchor_rows = [3, 32, 61, 90]

        for (title, figure), anchor_row in zip(plot_figures, anchor_rows):
            if figure is None:
                continue
            plots_sheet.cell(anchor_row - 1, 1, title).font = Font(
                bold=True,
                size=13,
            )
            image_bytes = BytesIO(figure_to_png_bytes(figure))
            image = ExcelImage(image_bytes)
            image.width = 840
            image.height = 500
            plots_sheet.add_image(image, f"A{anchor_row}")
            plt.close(figure)

        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit renderers
# =============================================================================
def render_final_result(result: EulerResult) -> None:
    """Render the final result card."""

    if not result.success:
        st.error(result.message)
        st.caption(result.stopping_reason)
        return

    st.success(result.message)
    st.markdown(f"**Problem:** {order_label(result.ode_order)}")
    st.markdown(
        f"**Differential Equation:** "
        f"`{highest_derivative_label(result.ode_order)} = {result.ode_text}`"
    )

    metric_columns = st.columns(2)
    metric_columns[0].metric("Final x", format_number(result.final_x))
    metric_columns[1].metric("Euler y", format_number(result.final_y))

    if result.ode_order >= 2:
        st.metric("Euler y′", format_number(result.final_yp))
    if result.ode_order >= 3:
        st.metric("Euler y″", format_number(result.final_ypp))

    initial_conditions = [
        f"y({format_number(result.x0)}) = {format_number(result.y0)}"
    ]
    if result.ode_order >= 2:
        initial_conditions.append(
            f"y′({format_number(result.x0)}) = {format_number(result.yp0)}"
        )
    if result.ode_order >= 3:
        initial_conditions.append(
            f"y″({format_number(result.x0)}) = {format_number(result.ypp0)}"
        )
    st.markdown("**Initial Conditions:** `" + "; ".join(initial_conditions) + "`")
    st.markdown(f"**Number of Steps:** {result.steps}")
    st.markdown(f"**Step Size:** {format_number(result.step_size)}")
    st.markdown(f"**Direction:** {result.direction}")

    if result.exact_final_y is not None:
        st.markdown(f"**Exact Final y:** {format_number(result.exact_final_y)}")
        st.markdown(
            "**Absolute Final Error y:** "
            f"{format_number(result.absolute_final_error_y)}"
        )
        st.markdown(
            "**Relative Final Error y (%):** "
            f"{format_number(result.relative_final_error_y_percent)}"
        )

    st.markdown(f"**Stopping Reason:** {result.stopping_reason}")

    for warning in result.warnings:
        st.warning(warning)


def render_method_summary(result: EulerResult) -> None:
    """Render the original problem and converted first-order system."""

    st.subheader("Method Formula and Problem Setup")

    if result.ode_order == 1:
        st.latex(r"y'=F(x,y)")
        st.latex(r"y_{n+1}=y_n+hF(x_n,y_n)")
    elif result.ode_order == 2:
        st.latex(r"u_1=y,\quad u_2=y'")
        st.latex(r"u_1'=u_2,\quad u_2'=F(x,u_1,u_2)")
        st.latex(
            r"y_{n+1}=y_n+h y'_n,\quad "
            r"y'_{n+1}=y'_n+hF_n"
        )
    else:
        st.latex(r"u_1=y,\quad u_2=y',\quad u_3=y''")
        st.latex(
            r"u_1'=u_2,\quad u_2'=u_3,\quad "
            r"u_3'=F(x,u_1,u_2,u_3)"
        )
        st.latex(
            r"y_{n+1}=y_n+h y'_n,\quad "
            r"y'_{n+1}=y'_n+h y''_n,\quad "
            r"y''_{n+1}=y''_n+hF_n"
        )

    st.dataframe(
        round_numeric_dataframe(method_formula_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )
    st.info(
        "Explicit Euler remains first-order accurate after converting a "
        "higher-order equation to a first-order system."
    )


def render_iteration_table(result: EulerResult) -> None:
    """Render every Euler step and its detailed operations."""

    st.subheader("Euler Iteration Table")
    dataframe = iterations_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(dataframe),
        use_container_width=True,
        hide_index=True,
    )

    st.caption(
        "Global error compares the accumulated numerical solution with the exact "
        "solution. Local truncation error measures one Euler step started from the "
        "exact state."
    )

    with st.expander("Detailed Euler Operations"):
        for record in result.iterations:
            st.markdown(f"**Iteration {record.iteration}**")
            st.code(record.operation_y, language=None)
            if record.operation_yp is not None:
                st.code(record.operation_yp, language=None)
            if record.operation_ypp is not None:
                st.code(record.operation_ypp, language=None)


def render_solution_graphs(result: EulerResult) -> None:
    """Render solution and converted-system graphs."""

    st.subheader("Numerical Solution Graph")
    try:
        solution_figure = create_solution_figure(result)
        st.pyplot(solution_figure, use_container_width=True)
        plt.close(solution_figure)
    except (ValueError, TypeError, ArithmeticError, OverflowError) as error:
        st.warning(f"The solution graph could not be displayed: {error}")

    state_figure = create_state_figure(result)
    if state_figure is not None:
        st.subheader("Converted System State Graph")
        st.pyplot(state_figure, use_container_width=True)
        plt.close(state_figure)


def render_error_analysis(result: EulerResult) -> None:
    """Render pointwise and summary error analysis."""

    st.subheader("Error Analysis")
    if result.exact_solution_expression is None:
        st.info(
            "Enter an exact solution y(x) to calculate global errors, local "
            "truncation errors, relative error, maximum error, and RMSE."
        )
        return

    metric_columns = st.columns(4)
    metric_columns[0].metric(
        "Final Absolute Error y",
        format_number(result.absolute_final_error_y),
    )
    metric_columns[1].metric(
        "Final Relative Error y (%)",
        format_number(result.relative_final_error_y_percent),
    )
    metric_columns[2].metric(
        "Maximum Absolute Error y",
        format_number(result.maximum_absolute_error_y),
    )
    metric_columns[3].metric("RMSE y", format_number(result.rmse_y))

    st.dataframe(
        round_numeric_dataframe(error_analysis_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    error_figure = create_error_figure(result)
    if error_figure is not None:
        st.pyplot(error_figure, use_container_width=True)
        plt.close(error_figure)


def render_convergence_analysis(result: EulerResult) -> None:
    """Render grid-refinement convergence analysis."""

    st.subheader("Convergence Analysis")
    st.markdown(
        "Euler's method has **first-order global convergence**. Halving the step "
        "size should approximately halve the global error after the asymptotic "
        "range is reached."
    )

    st.dataframe(
        round_numeric_dataframe(convergence_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    if result.latest_observed_order is not None:
        st.metric(
            "Latest Observed Order",
            format_number(result.latest_observed_order),
        )

    convergence_figure = create_convergence_figure(result)
    st.pyplot(convergence_figure, use_container_width=True)
    plt.close(convergence_figure)


def render_excel_download(result: EulerResult) -> None:
    """Generate and render the Excel download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature

    if (
        st.session_state.get("euler_excel_signature") != report_signature
        or "euler_excel_report" not in st.session_state
    ):
        try:
            st.session_state.euler_excel_report = generate_excel_report(result)
            st.session_state.euler_excel_signature = report_signature
        except (ValueError, OSError, TypeError, ArithmeticError) as error:
            st.error(f"The Excel report could not be generated: {error}")
            return

    report_bytes = st.session_state.get("euler_excel_report")
    if not report_bytes:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    filename = f"euler_method_order_{result.ode_order}_report_{timestamp}.xlsx"
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=filename,
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="euler_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Euler Method Streamlit page."""

    st.set_page_config(
        page_title="Euler Method Solver | Numerical Methods",
        page_icon="∫",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">ORDINARY DIFFERENTIAL EQUATIONS TOOL</div>
                <h1>Euler Method Solver</h1>
                <p>
                    Solve first-, second-, or third-order initial-value problems
                    using the explicit Euler Method. Higher-order equations are
                    converted internally to equivalent systems of first-order ODEs.
                </p>

                <div class="method-actions">
                    <a href="/Euler_Method" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Euler_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Write the Function")
                st.markdown(
                    """
                    Enter only the right-hand side of the highest derivative.

                    - First order: **y′ = F(x, y)** — use `x` and `y`.
                    - Second order: **y″ = F(x, y, y′)** — write `yp` for **y′**.
                    - Third order: **y‴ = F(x, y, y′, y″)** — write `yp` and `ypp`.
                    - Powers: write `x**2`, not `x^2`.
                    - Multiplication: write `2*x`, not `2x`.
                    - Enter no equals sign.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")
                st.markdown(
                    """
                    - First order requires **y(x₀)**.
                    - Second order also requires **y′(x₀)**.
                    - Third order also requires **y″(x₀)**.
                    - The number of steps must be a positive integer.
                    - Explicit Euler is globally first-order accurate.
                    - A very large step can make the solution inaccurate or unstable.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">Differential-equation order</div>',
                    unsafe_allow_html=True,
                )
                ode_order = st.selectbox(
                    "Differential-equation order",
                    options=[1, 2, 3],
                    index=DEFAULT_ORDER - 1,
                    format_func=lambda value: order_label(value),
                    label_visibility="collapsed",
                    key="euler_ode_order",
                )

                st.markdown(
                    f'<div class="input-label-ui">Right-hand side of '
                    f'{highest_derivative_label(ode_order)} = F({rhs_variables_text(ode_order)})</div>',
                    unsafe_allow_html=True,
                )
                ode_text = st.text_input(
                    "Right-hand side",
                    value=DEFAULT_ODE,
                    placeholder={
                        1: "Example: x + y",
                        2: "Example: -y or -0.2*yp - y",
                        3: "Example: 6 or -y - 0.2*ypp",
                    }[ode_order],
                    label_visibility="collapsed",
                    key="euler_ode_function",
                )

                first_row = st.columns(3)
                with first_row[0]:
                    st.markdown(
                        '<div class="input-label-ui">Initial x₀</div>',
                        unsafe_allow_html=True,
                    )
                    x0 = st.number_input(
                        "Initial x0",
                        value=DEFAULT_X0,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="euler_x0",
                    )

                with first_row[1]:
                    st.markdown(
                        '<div class="input-label-ui">Initial y(x₀)</div>',
                        unsafe_allow_html=True,
                    )
                    y0 = st.number_input(
                        "Initial y",
                        value=DEFAULT_Y0,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="euler_y0",
                    )

                with first_row[2]:
                    st.markdown(
                        '<div class="input-label-ui">Final x</div>',
                        unsafe_allow_html=True,
                    )
                    x_end = st.number_input(
                        "Final x",
                        value=DEFAULT_X_END,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="euler_x_end",
                    )

                yp0: float | None = None
                ypp0: float | None = None

                if ode_order >= 2:
                    derivative_columns = st.columns(2 if ode_order == 3 else 1)
                    with derivative_columns[0]:
                        st.markdown(
                            '<div class="input-label-ui">Initial y′(x₀)</div>',
                            unsafe_allow_html=True,
                        )
                        yp0 = st.number_input(
                            "Initial y prime",
                            value=DEFAULT_YP0,
                            format="%.12g",
                            label_visibility="collapsed",
                            key="euler_yp0",
                        )

                    if ode_order == 3:
                        with derivative_columns[1]:
                            st.markdown(
                                '<div class="input-label-ui">Initial y″(x₀)</div>',
                                unsafe_allow_html=True,
                            )
                            ypp0 = st.number_input(
                                "Initial y double prime",
                                value=DEFAULT_YPP0,
                                format="%.12g",
                                label_visibility="collapsed",
                                key="euler_ypp0",
                            )

                second_row = st.columns(2)
                with second_row[0]:
                    st.markdown(
                        '<div class="input-label-ui">Number of steps N</div>',
                        unsafe_allow_html=True,
                    )
                    steps = st.number_input(
                        "Number of steps",
                        min_value=MIN_STEPS,
                        max_value=MAX_STEPS,
                        value=DEFAULT_STEPS,
                        step=1,
                        label_visibility="collapsed",
                        key="euler_steps",
                    )

                with second_row[1]:
                    st.markdown(
                        '<div class="input-label-ui">Convergence levels</div>',
                        unsafe_allow_html=True,
                    )
                    convergence_levels = st.number_input(
                        "Convergence levels",
                        min_value=MIN_CONVERGENCE_LEVELS,
                        max_value=MAX_CONVERGENCE_LEVELS,
                        value=DEFAULT_CONVERGENCE_LEVELS,
                        step=1,
                        label_visibility="collapsed",
                        key="euler_convergence_levels",
                    )

                st.markdown(
                    '<div class="input-label-ui">Exact solution y(x) — optional</div>',
                    unsafe_allow_html=True,
                )
                exact_solution_text = st.text_input(
                    "Exact solution y(x)",
                    value="",
                    placeholder="Example: exp(x), sin(x), or x**3",
                    label_visibility="collapsed",
                    key="euler_exact_solution",
                )

                if float(x_end) != float(x0):
                    preview_h = (float(x_end) - float(x0)) / int(steps)
                    st.caption(f"Computed uniform step size: h = {preview_h:.12g}")

                current_signature = create_input_signature(
                    ode_order=ode_order,
                    ode_text=ode_text,
                    x0=x0,
                    y0=y0,
                    yp0=yp0,
                    ypp0=ypp0,
                    x_end=x_end,
                    steps=steps,
                    exact_solution_text=exact_solution_text,
                    convergence_levels=convergence_levels,
                )

                solve_clicked = st.button(
                    "Solve with Euler Method",
                    type="primary",
                    use_container_width=True,
                    key="euler_solve_button",
                )

                if solve_clicked:
                    result = solve_euler_method(
                        ode_order_input=ode_order,
                        ode_text=ode_text,
                        x0_input=x0,
                        y0_input=y0,
                        yp0_input=yp0,
                        ypp0_input=ypp0,
                        x_end_input=x_end,
                        steps_input=steps,
                        exact_solution_text=exact_solution_text,
                        convergence_levels_input=convergence_levels,
                    )
                    st.session_state.euler_result = result
                    st.session_state.euler_result_signature = current_signature
                    st.session_state.pop("euler_excel_report", None)
                    st.session_state.pop("euler_excel_signature", None)
                    st.rerun()

                with st.expander("Example Inputs"):
                    st.code(
                        "FIRST ORDER\n"
                        "Order = 1\n"
                        "y' = y\n"
                        "x0 = 0, y0 = 1, x_end = 1\n"
                        "Exact solution: exp(x)",
                        language=None,
                    )
                    st.code(
                        "SECOND ORDER\n"
                        "Order = 2\n"
                        "y'' = -y\n"
                        "x0 = 0, y0 = 0, y'(0) = 1\n"
                        "Exact solution: sin(x)",
                        language=None,
                    )
                    st.code(
                        "THIRD ORDER\n"
                        "Order = 3\n"
                        "y''' = 6\n"
                        "x0 = 0, y0 = 0, y'(0) = 0, y''(0) = 0\n"
                        "Exact solution: x**3",
                        language=None,
                    )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                saved_result = st.session_state.get("euler_result")
                saved_signature = st.session_state.get("euler_result_signature")

                if saved_result is None:
                    st.info("Enter the IVP data and select Solve with Euler Method.")
                elif saved_signature != current_signature:
                    st.warning(
                        "The inputs have changed. Select Solve with Euler Method "
                        "to update the result."
                    )
                else:
                    render_final_result(saved_result)

        saved_result = st.session_state.get("euler_result")
        saved_signature = st.session_state.get("euler_result_signature")

        if saved_result is not None and saved_signature == current_signature:
            if saved_result.success:
                st.divider()
                render_method_summary(saved_result)
                st.divider()
                render_iteration_table(saved_result)
                st.divider()
                render_solution_graphs(saved_result)
                st.divider()
                render_error_analysis(saved_result)
                st.divider()
                render_convergence_analysis(saved_result)
                st.divider()
                render_excel_download(saved_result)

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Euler Method • Ordinary Differential Equations</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()