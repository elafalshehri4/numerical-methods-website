from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Sequence
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Forward Difference Method"
DISPLAY_DECIMALS = 6
DEFAULT_FUNCTION = "x**3 - 2*x + 1"
DEFAULT_X_VALUE = 1.0
DEFAULT_STEP_SIZE = 0.1
DEFAULT_REFINEMENT_LEVELS = 5
MIN_REFINEMENT_LEVELS = 2
MAX_REFINEMENT_LEVELS = 10
ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DERIVATIVE_OPTIONS = {
    "First derivative f′(x)": 1,
    "Second derivative f″(x)": 2,
    "Third derivative f‴(x)": 3,
}

ACCURACY_OPTIONS = {
    "Standard Taylor formula — O(h)": "standard",
    "Higher-accuracy Taylor formula — O(h²)": "high_accuracy",
}

ALLOWED_FUNCTION_NAMES = {
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}


# =============================================================================
# Data models
# =============================================================================
@dataclass(frozen=True)
class FormulaSpec:
    """Finite-difference formula metadata."""

    derivative_order: int
    accuracy_key: str
    offsets: tuple[int, ...]
    coefficients: tuple[float, ...]
    denominator_factor: float
    theoretical_order: int
    formula_text: str


@dataclass(frozen=True)
class DifferenceEvaluation:
    """One calculation at a specific step size."""

    level: int
    step_size: float
    derivative_order: int
    offsets: tuple[int, ...]
    coefficients: tuple[float, ...]
    sample_points: tuple[float, ...]
    function_values: tuple[float, ...]
    numerator: float
    denominator: float
    derivative_approximation: float
    exact_derivative: float | None
    absolute_error: float | None
    relative_error: float | None
    successive_difference: float | None
    observed_order: float | None
    formula_text: str
    substitution_text: str


@dataclass(frozen=True)
class ForwardDifferenceResult:
    """Complete result shared by Streamlit and Excel export."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str
    function_text: str
    function_expression: sp.Expr | None
    derivative_expression: sp.Expr | None
    derivative_expression_text: str
    derivative_order: int
    derivative_name: str
    accuracy_key: str
    accuracy_name: str
    formula_spec: FormulaSpec
    x_value: float | None
    initial_step_size: float | None
    refinement_levels: int
    evaluations: tuple[DifferenceEvaluation, ...]
    primary_approximation: float | None
    primary_exact_derivative: float | None
    primary_absolute_error: float | None
    primary_relative_error: float | None
    finest_step_size: float | None
    finest_approximation: float | None
    finest_absolute_error: float | None
    observed_order_latest: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a local image as Base64 text."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def format_number(value: float | None, decimals: int = DISPLAY_DECIMALS) -> str:
    """Format one scalar for display."""

    if value is None:
        return "—"
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if abs(number) < 10 ** (-(decimals + 2)):
        number = 0.0
    return f"{number:.{decimals}f}"


def scientific_number(value: float | None) -> str:
    """Format one scalar in scientific notation."""

    if value is None:
        return "—"
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    return f"{number:.6e}"


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns for display only."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def safe_float(raw_value: Any, value_name: str) -> float:
    """Convert user input to a finite real number."""

    if raw_value is None or (
        isinstance(raw_value, str) and not raw_value.strip()
    ):
        raise ValueError(f"{value_name} is required.")
    try:
        value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be numerical.") from error
    if not math.isfinite(value):
        raise ValueError(f"{value_name} must be finite.")
    return value


def create_input_signature(
    function_text: str,
    x_value: Any,
    step_size: Any,
    derivative_name: str,
    accuracy_name: str,
    refinement_levels: Any,
) -> str:
    """Create a stable signature to prevent stale displayed results."""

    payload = repr(
        (
            str(function_text).strip(),
            str(x_value),
            str(step_size),
            derivative_name,
            accuracy_name,
            str(refinement_levels),
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def human_readable_expression(expression: sp.Expr | None) -> str:
    """Return a compact readable expression."""

    if expression is None:
        return "Not available"
    return str(sp.simplify(expression))


# =============================================================================
# Function parsing and evaluation
# =============================================================================
def parse_function(function_text: str) -> tuple[sp.Expr, sp.Symbol]:
    """Parse a real scalar function containing only x."""

    if not isinstance(function_text, str) or not function_text.strip():
        raise ValueError("The function input cannot be empty.")

    x_symbol = sp.Symbol("x", real=True)
    local_dictionary = dict(ALLOWED_FUNCTION_NAMES)
    local_dictionary["x"] = x_symbol

    try:
        expression = sp.sympify(
            function_text.strip(),
            locals=local_dictionary,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The function could not be parsed. Use syntax such as "
            "sin(x), exp(x), log(x), or x**3 - 2*x + 1."
        ) from error

    if isinstance(expression, (sp.MatrixBase, Relational)):
        raise ValueError("The input must be a scalar function of x.")

    unexpected_symbols = expression.free_symbols.difference({x_symbol})
    if unexpected_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unexpected_symbols))
        raise ValueError(
            "Only x is allowed as a variable. "
            f"Unexpected symbol(s): {names}."
        )

    if expression.has(sp.zoo, sp.nan, sp.oo, -sp.oo):
        raise ValueError("The function contains an undefined value.")

    if expression.atoms(AppliedUndef):
        raise ValueError("The function contains an unsupported function.")

    return sp.simplify(expression), x_symbol


def create_numeric_function(
    expression: sp.Expr,
    x_symbol: sp.Symbol,
) -> Callable[[Any], Any]:
    """Convert a SymPy expression to a NumPy-compatible function."""

    try:
        return sp.lambdify(x_symbol, expression, modules=["numpy"])
    except (TypeError, ValueError) as error:
        raise ValueError("The function could not be converted numerically.") from error


def evaluate_real_scalar(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    description: str,
) -> float:
    """Evaluate a function and require one finite real value."""

    try:
        with np.errstate(all="ignore"):
            raw_value = numeric_function(x_value)
    except (ArithmeticError, TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The function could not be evaluated at x = {x_value:.12g} "
            f"while calculating {description}."
        ) from error

    array = np.asarray(raw_value)
    if array.size != 1:
        raise ValueError(f"{description} did not produce a scalar.")

    scalar = array.reshape(-1)[0]
    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)
        if abs(complex_value.imag) > ZERO_TOLERANCE:
            raise ValueError(
                f"{description} is complex at x = {x_value:.12g}."
            )
        scalar = complex_value.real

    try:
        numeric_value = float(scalar)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{description} is not a real number.") from error

    if not math.isfinite(numeric_value):
        raise ValueError(
            f"{description} is undefined, NaN, or infinite at "
            f"x = {x_value:.12g}."
        )
    return numeric_value


def evaluate_real_array(
    numeric_function: Callable[[Any], Any],
    x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate a function over an array for plotting."""

    with np.errstate(all="ignore"):
        try:
            raw_values = numeric_function(x_values)
        except Exception:
            return np.full_like(x_values, np.nan, dtype=float)

    array = np.asarray(raw_values)
    if array.ndim == 0:
        dtype = complex if np.iscomplexobj(array) else float
        array = np.full_like(x_values, array, dtype=dtype)
    try:
        array = np.broadcast_to(array, x_values.shape)
    except ValueError:
        return np.full_like(x_values, np.nan, dtype=float)

    if np.iscomplexobj(array):
        real_values = np.real(array).astype(float)
        real_values[np.abs(np.imag(array)) > ZERO_TOLERANCE] = np.nan
    else:
        try:
            real_values = array.astype(float)
        except (TypeError, ValueError):
            return np.full_like(x_values, np.nan, dtype=float)

    real_values[~np.isfinite(real_values)] = np.nan
    return real_values


# =============================================================================
# Textbook forward-difference formulas derived from Taylor series
# =============================================================================
def get_formula_spec(
    derivative_order: int,
    accuracy_key: str,
) -> FormulaSpec:
    """Return the standard Chapra forward-difference formula."""

    specifications: dict[tuple[int, str], FormulaSpec] = {
        (1, "standard"): FormulaSpec(
            1,
            "standard",
            (0, 1),
            (-1.0, 1.0),
            1.0,
            1,
            "f′(x₀) ≈ [f(x₀+h) − f(x₀)] / h",
        ),
        (1, "high_accuracy"): FormulaSpec(
            1,
            "high_accuracy",
            (0, 1, 2),
            (-3.0, 4.0, -1.0),
            2.0,
            2,
            "f′(x₀) ≈ [−3f(x₀)+4f(x₀+h)−f(x₀+2h)] / (2h)",
        ),
        (2, "standard"): FormulaSpec(
            2,
            "standard",
            (0, 1, 2),
            (1.0, -2.0, 1.0),
            1.0,
            1,
            "f″(x₀) ≈ [f(x₀)−2f(x₀+h)+f(x₀+2h)] / h²",
        ),
        (2, "high_accuracy"): FormulaSpec(
            2,
            "high_accuracy",
            (0, 1, 2, 3),
            (2.0, -5.0, 4.0, -1.0),
            1.0,
            2,
            "f″(x₀) ≈ [2f(x₀)−5f(x₀+h)+4f(x₀+2h)−f(x₀+3h)] / h²",
        ),
        (3, "standard"): FormulaSpec(
            3,
            "standard",
            (0, 1, 2, 3),
            (-1.0, 3.0, -3.0, 1.0),
            1.0,
            1,
            "f‴(x₀) ≈ [−f(x₀)+3f(x₀+h)−3f(x₀+2h)+f(x₀+3h)] / h³",
        ),
        (3, "high_accuracy"): FormulaSpec(
            3,
            "high_accuracy",
            (0, 1, 2, 3, 4),
            (-5.0, 18.0, -24.0, 14.0, -3.0),
            2.0,
            2,
            "f‴(x₀) ≈ [−5f(x₀)+18f(x₀+h)−24f(x₀+2h)"
            "+14f(x₀+3h)−3f(x₀+4h)] / (2h³)",
        ),
    }

    try:
        return specifications[(derivative_order, accuracy_key)]
    except KeyError as error:
        raise ValueError("Unsupported derivative order or formula accuracy.") from error


def calculate_forward_difference(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    step_size: float,
    formula_spec: FormulaSpec,
) -> dict[str, Any]:
    """Calculate one forward finite-difference approximation manually."""

    sample_points: list[float] = []
    function_values: list[float] = []

    for offset in formula_spec.offsets:
        sample_point = x_value + offset * step_size
        if not math.isfinite(sample_point):
            raise ValueError("A required forward sample point is not finite.")
        function_value = evaluate_real_scalar(
            numeric_function,
            sample_point,
            f"f(x₀ + {offset}h)",
        )
        sample_points.append(sample_point)
        function_values.append(function_value)

    numerator = float(
        sum(
            coefficient * value
            for coefficient, value in zip(
                formula_spec.coefficients,
                function_values,
            )
        )
    )
    denominator = float(
        formula_spec.denominator_factor
        * step_size ** formula_spec.derivative_order
    )

    if abs(denominator) <= ZERO_TOLERANCE:
        raise ValueError("The formula denominator is numerically zero.")

    approximation = numerator / denominator
    if not all(math.isfinite(value) for value in (numerator, denominator, approximation)):
        raise ValueError("The calculation produced a non-finite value.")

    terms = []
    for coefficient, offset, value in zip(
        formula_spec.coefficients,
        formula_spec.offsets,
        function_values,
    ):
        terms.append(
            f"({coefficient:g})f(x₀+{offset}h)="
            f"({coefficient:g})({format_number(value, 10)})"
        )
    substitution = (
        "[" + " + ".join(terms) + "] / "
        f"[{formula_spec.denominator_factor:g}"
        f"({format_number(step_size, 10)})^{formula_spec.derivative_order}]"
    )

    return {
        "sample_points": tuple(sample_points),
        "function_values": tuple(function_values),
        "numerator": numerator,
        "denominator": denominator,
        "approximation": approximation,
        "substitution": substitution,
    }


def safe_relative_error(
    absolute_error: float,
    exact_value: float,
) -> float | None:
    """Calculate the absolute relative true error as a percentage."""

    if abs(exact_value) <= RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
        return None
    return (absolute_error / abs(exact_value)) * 100.0


def calculate_observed_order(
    previous_error: float | None,
    current_error: float | None,
    refinement_ratio: float = 2.0,
) -> float | None:
    """Estimate convergence order from consecutive exact errors."""

    if previous_error is None or current_error is None:
        return None
    if previous_error <= 0.0 or current_error <= 0.0:
        return None
    try:
        order = math.log(previous_error / current_error) / math.log(
            refinement_ratio
        )
    except (ValueError, ZeroDivisionError):
        return None
    return order if math.isfinite(order) else None


def calculate_successive_order(
    older_difference: float | None,
    newer_difference: float | None,
    refinement_ratio: float = 2.0,
) -> float | None:
    """Estimate convergence order without an exact derivative."""

    if older_difference is None or newer_difference is None:
        return None
    if older_difference <= 0.0 or newer_difference <= 0.0:
        return None
    try:
        order = math.log(older_difference / newer_difference) / math.log(
            refinement_ratio
        )
    except (ValueError, ZeroDivisionError):
        return None
    return order if math.isfinite(order) else None


def solve_forward_difference(
    function_text: str,
    raw_x_value: Any,
    raw_step_size: Any,
    derivative_name: str,
    accuracy_name: str,
    raw_refinement_levels: Any,
    input_signature: str,
) -> ForwardDifferenceResult:
    """Perform the complete validated Forward Difference workflow."""

    execution_datetime = current_report_datetime()
    derivative_order = DERIVATIVE_OPTIONS.get(derivative_name, 1)
    accuracy_key = ACCURACY_OPTIONS.get(accuracy_name, "standard")
    formula_spec = get_formula_spec(derivative_order, accuracy_key)

    try:
        x_value = safe_float(raw_x_value, "Evaluation point x₀")
        initial_step_size = safe_float(raw_step_size, "Step size h")
        if initial_step_size <= 0.0:
            raise ValueError("Step size h must be greater than zero.")
        if initial_step_size <= ZERO_TOLERANCE:
            raise ValueError("Step size h is too close to machine precision.")

        try:
            refinement_levels = int(raw_refinement_levels)
        except (TypeError, ValueError) as error:
            raise ValueError("Refinement levels must be an integer.") from error
        if not MIN_REFINEMENT_LEVELS <= refinement_levels <= MAX_REFINEMENT_LEVELS:
            raise ValueError(
                f"Refinement levels must be between {MIN_REFINEMENT_LEVELS} "
                f"and {MAX_REFINEMENT_LEVELS}."
            )

        expression, x_symbol = parse_function(function_text)
        numeric_function = create_numeric_function(expression, x_symbol)

        derivative_expression: sp.Expr | None
        exact_derivative: float | None
        derivative_warning: str | None = None
        try:
            derivative_expression = sp.simplify(
                sp.diff(expression, x_symbol, derivative_order)
            )
            derivative_numeric = create_numeric_function(
                derivative_expression,
                x_symbol,
            )
            exact_derivative = evaluate_real_scalar(
                derivative_numeric,
                x_value,
                f"the exact derivative of order {derivative_order}",
            )
        except (ValueError, TypeError, NotImplementedError) as error:
            derivative_expression = None
            exact_derivative = None
            derivative_warning = (
                "The exact symbolic derivative could not be evaluated. "
                "The numerical approximation is still available. "
                f"Details: {error}"
            )

        evaluations: list[DifferenceEvaluation] = []
        approximations: list[float] = []
        exact_errors: list[float | None] = []
        successive_differences: list[float | None] = []

        for level in range(refinement_levels):
            step_size = initial_step_size / (2.0**level)
            if step_size <= ZERO_TOLERANCE or not math.isfinite(step_size):
                raise ValueError(
                    "Step-size refinement became numerically unusable."
                )

            calculation = calculate_forward_difference(
                numeric_function,
                x_value,
                step_size,
                formula_spec,
            )
            approximation = float(calculation["approximation"])

            if exact_derivative is not None:
                absolute_error = abs(approximation - exact_derivative)
                relative_error = safe_relative_error(
                    absolute_error,
                    exact_derivative,
                )
            else:
                absolute_error = None
                relative_error = None

            successive_difference = (
                abs(approximation - approximations[-1])
                if approximations
                else None
            )

            if exact_derivative is not None and exact_errors:
                observed_order = calculate_observed_order(
                    exact_errors[-1],
                    absolute_error,
                )
            elif successive_differences:
                observed_order = calculate_successive_order(
                    successive_differences[-1],
                    successive_difference,
                )
            else:
                observed_order = None

            evaluations.append(
                DifferenceEvaluation(
                    level=level,
                    step_size=step_size,
                    derivative_order=derivative_order,
                    offsets=formula_spec.offsets,
                    coefficients=formula_spec.coefficients,
                    sample_points=calculation["sample_points"],
                    function_values=calculation["function_values"],
                    numerator=float(calculation["numerator"]),
                    denominator=float(calculation["denominator"]),
                    derivative_approximation=approximation,
                    exact_derivative=exact_derivative,
                    absolute_error=absolute_error,
                    relative_error=relative_error,
                    successive_difference=successive_difference,
                    observed_order=observed_order,
                    formula_text=formula_spec.formula_text,
                    substitution_text=str(calculation["substitution"]),
                )
            )
            approximations.append(approximation)
            exact_errors.append(absolute_error)
            successive_differences.append(successive_difference)

        warnings: list[str] = []
        if derivative_warning:
            warnings.append(derivative_warning)

        primary = evaluations[0]
        finest = evaluations[-1]

        if finest.absolute_error is not None and primary.absolute_error is not None:
            if finest.absolute_error > primary.absolute_error:
                warnings.append(
                    "The exact error increased after refinement. Round-off or "
                    "subtractive cancellation may be dominant."
                )

        if finest.step_size < math.sqrt(np.finfo(float).eps) * max(
            1.0,
            abs(x_value),
        ):
            warnings.append(
                "The finest h is very small relative to x₀. Cancellation may "
                "reduce numerical accuracy."
            )

        observed_orders = [
            item.observed_order
            for item in evaluations
            if item.observed_order is not None
            and math.isfinite(item.observed_order)
        ]
        latest_order = observed_orders[-1] if observed_orders else None

        return ForwardDifferenceResult(
            status="success",
            success=True,
            method=METHOD_NAME,
            message="Execution completed successfully.",
            stopping_reason=(
                "The requested derivative and step-size refinement analysis "
                "were completed."
            ),
            function_text=function_text.strip(),
            function_expression=expression,
            derivative_expression=derivative_expression,
            derivative_expression_text=human_readable_expression(
                derivative_expression
            ),
            derivative_order=derivative_order,
            derivative_name=derivative_name,
            accuracy_key=accuracy_key,
            accuracy_name=accuracy_name,
            formula_spec=formula_spec,
            x_value=x_value,
            initial_step_size=initial_step_size,
            refinement_levels=refinement_levels,
            evaluations=tuple(evaluations),
            primary_approximation=primary.derivative_approximation,
            primary_exact_derivative=primary.exact_derivative,
            primary_absolute_error=primary.absolute_error,
            primary_relative_error=primary.relative_error,
            finest_step_size=finest.step_size,
            finest_approximation=finest.derivative_approximation,
            finest_absolute_error=finest.absolute_error,
            observed_order_latest=latest_order,
            warnings=tuple(warnings),
            input_signature=input_signature,
            execution_datetime=execution_datetime,
        )

    except ValueError as error:
        return ForwardDifferenceResult(
            status="error",
            success=False,
            method=METHOD_NAME,
            message=str(error),
            stopping_reason=(
                "The calculation stopped during validation or function evaluation."
            ),
            function_text=str(function_text).strip(),
            function_expression=None,
            derivative_expression=None,
            derivative_expression_text="Not available",
            derivative_order=derivative_order,
            derivative_name=derivative_name,
            accuracy_key=accuracy_key,
            accuracy_name=accuracy_name,
            formula_spec=formula_spec,
            x_value=None,
            initial_step_size=None,
            refinement_levels=0,
            evaluations=tuple(),
            primary_approximation=None,
            primary_exact_derivative=None,
            primary_absolute_error=None,
            primary_relative_error=None,
            finest_step_size=None,
            finest_approximation=None,
            finest_absolute_error=None,
            observed_order_latest=None,
            warnings=tuple(),
            input_signature=input_signature,
            execution_datetime=execution_datetime,
        )


# =============================================================================
# DataFrames
# =============================================================================
def create_evaluation_dataframe(
    result: ForwardDifferenceResult,
) -> pd.DataFrame:
    """Build the complete refinement table."""

    rows: list[dict[str, Any]] = []
    for item in result.evaluations:
        row: dict[str, Any] = {
            "Level": item.level,
            "Derivative Order": item.derivative_order,
            "h": item.step_size,
            "Numerator": item.numerator,
            "Denominator": item.denominator,
            "Approximate Derivative": item.derivative_approximation,
            "Exact Derivative": item.exact_derivative,
            "Absolute Error": item.absolute_error,
            "Relative Error (%)": item.relative_error,
            "Successive Difference": item.successive_difference,
            "Observed Order": item.observed_order,
        }
        for offset, point, value in zip(
            item.offsets,
            item.sample_points,
            item.function_values,
        ):
            row[f"x0 + {offset}h"] = point
            row[f"f(x0 + {offset}h)"] = value
        rows.append(row)
    return pd.DataFrame(rows)


def create_primary_calculation_dataframe(
    result: ForwardDifferenceResult,
) -> pd.DataFrame:
    """Build a vertical table for the entered step size."""

    if not result.evaluations:
        return pd.DataFrame()
    item = result.evaluations[0]
    rows: list[tuple[str, Any]] = [
        ("Function", result.function_text),
        ("Derivative", result.derivative_name),
        ("Formula Accuracy", result.accuracy_name),
        ("Formula", result.formula_spec.formula_text),
        ("x₀", result.x_value),
        ("h", item.step_size),
    ]
    for offset, coefficient, point, value in zip(
        item.offsets,
        item.coefficients,
        item.sample_points,
        item.function_values,
    ):
        rows.extend(
            [
                (f"Offset k={offset}", offset),
                (f"Coefficient at k={offset}", coefficient),
                (f"x₀ + {offset}h", point),
                (f"f(x₀ + {offset}h)", value),
            ]
        )
    rows.extend(
        [
            ("Numerator", item.numerator),
            ("Denominator", item.denominator),
            ("Approximate Derivative", item.derivative_approximation),
            ("Exact Derivative", item.exact_derivative),
            ("Absolute Error", item.absolute_error),
            ("Relative Error (%)", item.relative_error),
        ]
    )
    return pd.DataFrame(rows, columns=["Property", "Value"])


def create_sample_values_dataframe(
    result: ForwardDifferenceResult,
) -> pd.DataFrame:
    """Create one row per sampled point for every refinement level."""

    rows = []
    for item in result.evaluations:
        for offset, coefficient, point, value in zip(
            item.offsets,
            item.coefficients,
            item.sample_points,
            item.function_values,
        ):
            rows.append(
                {
                    "Level": item.level,
                    "h": item.step_size,
                    "Offset k": offset,
                    "Coefficient": coefficient,
                    "Sample Point": point,
                    "Function Value": value,
                    "Weighted Contribution": coefficient * value,
                }
            )
    return pd.DataFrame(rows)


def create_error_dataframe(
    result: ForwardDifferenceResult,
) -> pd.DataFrame:
    """Build a focused error and convergence table."""

    evaluation_dataframe = create_evaluation_dataframe(result)
    selected = [
        "Level",
        "h",
        "Approximate Derivative",
        "Exact Derivative",
        "Absolute Error",
        "Relative Error (%)",
        "Successive Difference",
        "Observed Order",
    ]
    return evaluation_dataframe[selected].copy()


def create_taylor_series_dataframe(
    result: ForwardDifferenceResult,
) -> pd.DataFrame:
    """Show Taylor expansions through the third derivative."""

    if not result.evaluations:
        return pd.DataFrame()
    h_value = result.evaluations[0].step_size
    rows = []
    for offset in result.formula_spec.offsets:
        rows.append(
            {
                "Offset k": offset,
                "Sample": f"f(x0 + {offset}h)",
                "f(x0) coefficient": 1.0,
                "f'(x0) coefficient": offset * h_value,
                "f''(x0) coefficient": (offset * h_value) ** 2 / 2.0,
                "f'''(x0) coefficient": (offset * h_value) ** 3 / 6.0,
                "Taylor Expansion to Third Degree": (
                    f"f(x0+{offset}h) = f(x0) + ({offset}h)f'(x0) + "
                    f"({offset}h)^2/2! f''(x0) + "
                    f"({offset}h)^3/3! f'''(x0) + R4"
                ),
            }
        )
    return pd.DataFrame(rows)


# =============================================================================
# Plots
# =============================================================================
def create_function_plot(result: ForwardDifferenceResult) -> plt.Figure:
    """Plot the function and all primary forward sample points."""

    if not result.success or result.function_expression is None:
        raise ValueError("A successful result is required for plotting.")

    x_symbol = sp.Symbol("x", real=True)
    numeric_function = create_numeric_function(
        result.function_expression,
        x_symbol,
    )
    primary = result.evaluations[0]
    max_offset = max(primary.offsets)
    required_span = max_offset * primary.step_size
    half_width = max(
        5.0 * max(required_span, primary.step_size),
        1.0,
        0.25 * max(1.0, abs(primary.sample_points[0])),
    )
    x_min = primary.sample_points[0] - half_width
    x_max = primary.sample_points[0] + half_width
    x_values = np.linspace(x_min, x_max, 600)
    y_values = evaluate_real_array(numeric_function, x_values)

    if np.count_nonzero(np.isfinite(y_values)) < 2:
        raise ValueError("The function is not plottable near x₀.")

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.plot(
        x_values,
        y_values,
        linewidth=2,
        label=f"f(x) = {result.function_text}",
    )
    axis.axhline(0.0, linewidth=1)
    axis.axvline(
        result.x_value,
        linestyle=":",
        linewidth=1.5,
        label="x = x₀",
    )
    axis.scatter(
        primary.sample_points,
        primary.function_values,
        s=70,
        label="Forward sample points",
        zorder=5,
    )

    if result.derivative_order == 1:
        tangent_values = primary.function_values[0] + (
            primary.derivative_approximation
            * (x_values - primary.sample_points[0])
        )
        axis.plot(
            x_values,
            tangent_values,
            linestyle="--",
            linewidth=2,
            label=(
                "Approximate tangent: slope = "
                f"{primary.derivative_approximation:.6g}"
            ),
        )

    axis.set_title(
        f"Forward Difference — Derivative Order {result.derivative_order}"
    )
    axis.set_xlabel("x")
    axis.set_ylabel("f(x)")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_convergence_plot(result: ForwardDifferenceResult) -> plt.Figure:
    """Create a log-log error-convergence plot."""

    if not result.evaluations:
        raise ValueError("No refinement history is available.")

    step_sizes = np.asarray(
        [item.step_size for item in result.evaluations],
        dtype=float,
    )
    exact_errors = np.asarray(
        [
            np.nan if item.absolute_error is None else item.absolute_error
            for item in result.evaluations
        ],
        dtype=float,
    )
    successive = np.asarray(
        [
            np.nan
            if item.successive_difference is None
            else item.successive_difference
            for item in result.evaluations
        ],
        dtype=float,
    )

    use_exact = np.count_nonzero(np.isfinite(exact_errors)) >= 2
    values = exact_errors if use_exact else successive
    label = "Absolute Error" if use_exact else "Successive Difference"
    valid = np.isfinite(values) & (values > 0.0) & np.isfinite(step_sizes)

    if np.count_nonzero(valid) < 2:
        raise ValueError("At least two positive error values are required.")

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.loglog(
        step_sizes[valid],
        values[valid],
        marker="o",
        linewidth=2,
        label=label,
    )
    axis.invert_xaxis()
    axis.set_title("Forward Difference — Convergence Analysis")
    axis.set_xlabel("Step Size h (Log Scale)")
    axis.set_ylabel(f"{label} (Log Scale)")
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def figure_to_png_bytes(figure: plt.Figure) -> bytes:
    """Convert a Matplotlib figure to PNG bytes."""

    buffer = BytesIO()
    figure.savefig(buffer, format="png", dpi=180, bbox_inches="tight")
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# Excel export
# =============================================================================
def serialize_warnings(warnings: Sequence[str]) -> str:
    """Serialize warning messages for one Excel cell."""

    return "None" if not warnings else "\n".join(
        f"• {warning}" for warning in warnings
    )


def apply_excel_style(workbook: Any) -> None:
    """Apply consistent formatting to workbook sheets."""

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

        for column_index, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):
            maximum_length = max(
                (
                    len(str(cell.value))
                    if cell.value is not None
                    else 0
                )
                for cell in column_cells
            )
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = min(max(maximum_length + 2, 12), 58)


def create_excel_report(result: ForwardDifferenceResult) -> bytes:
    """Create an XLSX report containing all tables, charts, and images."""

    if not result.success:
        raise ValueError("Only successful results can be exported.")

    summary_dataframe = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "Function",
                "Derivative Order",
                "Derivative Expression",
                "Formula Accuracy",
                "Formula",
                "Theoretical Error Order",
                "Evaluation Point x0",
                "Entered Step Size h",
                "Refinement Levels",
                "Primary Approximation",
                "Exact Derivative",
                "Primary Absolute Error",
                "Primary Relative Error (%)",
                "Finest Step Size",
                "Finest Approximation",
                "Finest Absolute Error",
                "Latest Observed Order",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                result.function_text,
                result.derivative_order,
                result.derivative_expression_text,
                result.accuracy_name,
                result.formula_spec.formula_text,
                f"O(h^{result.formula_spec.theoretical_order})",
                result.x_value,
                result.initial_step_size,
                result.refinement_levels,
                result.primary_approximation,
                result.primary_exact_derivative,
                result.primary_absolute_error,
                result.primary_relative_error,
                result.finest_step_size,
                result.finest_approximation,
                result.finest_absolute_error,
                result.observed_order_latest,
                serialize_warnings(result.warnings),
                result.stopping_reason,
                result.execution_datetime.strftime(
                    "%Y-%m-%d %H:%M:%S %Z"
                ),
            ],
        }
    )

    primary_dataframe = create_primary_calculation_dataframe(result)
    evaluation_dataframe = create_evaluation_dataframe(result)
    sample_dataframe = create_sample_values_dataframe(result)
    taylor_dataframe = create_taylor_series_dataframe(result)
    error_dataframe = create_error_dataframe(result)
    formula_dataframe = pd.DataFrame(
        [
            {
                "Level": item.level,
                "h": item.step_size,
                "Formula": item.formula_text,
                "Substitution": item.substitution_text,
                "Approximate Derivative": item.derivative_approximation,
            }
            for item in result.evaluations
        ]
    )

    function_figure = create_function_plot(result)
    function_png = figure_to_png_bytes(function_figure)
    plt.close(function_figure)

    convergence_png: bytes | None = None
    try:
        convergence_figure = create_convergence_plot(result)
    except ValueError:
        convergence_figure = None
    if convergence_figure is not None:
        convergence_png = figure_to_png_bytes(convergence_figure)
        plt.close(convergence_figure)

    output = BytesIO()
    image_buffers: list[BytesIO] = []

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_dataframe.to_excel(writer, sheet_name="Summary", index=False)
        primary_dataframe.to_excel(
            writer,
            sheet_name="Primary Calculation",
            index=False,
        )
        evaluation_dataframe.to_excel(
            writer,
            sheet_name="Difference Results",
            index=False,
        )
        sample_dataframe.to_excel(
            writer,
            sheet_name="Sample Values",
            index=False,
        )
        taylor_dataframe.to_excel(
            writer,
            sheet_name="Taylor Series",
            index=False,
        )
        formula_dataframe.to_excel(
            writer,
            sheet_name="Formula Steps",
            index=False,
        )
        error_dataframe.to_excel(
            writer,
            sheet_name="Error Analysis",
            index=False,
        )
        error_dataframe.to_excel(
            writer,
            sheet_name="Convergence Analysis",
            index=False,
        )
        pd.DataFrame(
            {"Embedded Charts": ["Function graph", "Convergence graph"]}
        ).to_excel(writer, sheet_name="Charts", index=False)

        workbook = writer.book
        apply_excel_style(workbook)

        convergence_sheet = workbook["Convergence Analysis"]
        row_count = len(error_dataframe) + 1
        if row_count >= 3:
            approximation_chart = ScatterChart()
            approximation_chart.title = "Approximation versus Step Size"
            approximation_chart.x_axis.title = "Step Size h"
            approximation_chart.y_axis.title = "Approximate Derivative"
            approximation_chart.height = 8
            approximation_chart.width = 15
            x_reference = Reference(
                convergence_sheet,
                min_col=2,
                min_row=2,
                max_row=row_count,
            )
            y_reference = Reference(
                convergence_sheet,
                min_col=3,
                min_row=2,
                max_row=row_count,
            )
            approximation_chart.series.append(
                Series(
                    y_reference,
                    x_reference,
                    title="Approximate Derivative",
                )
            )
            convergence_sheet.add_chart(approximation_chart, "J2")

            if any(
                item.absolute_error is not None
                for item in result.evaluations
            ):
                error_chart = LineChart()
                error_chart.title = "Absolute Error by Refinement Level"
                error_chart.x_axis.title = "Refinement Level"
                error_chart.y_axis.title = "Absolute Error"
                error_chart.height = 8
                error_chart.width = 15
                data_reference = Reference(
                    convergence_sheet,
                    min_col=5,
                    min_row=1,
                    max_row=row_count,
                )
                category_reference = Reference(
                    convergence_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=row_count,
                )
                error_chart.add_data(
                    data_reference,
                    titles_from_data=True,
                )
                error_chart.set_categories(category_reference)
                convergence_sheet.add_chart(error_chart, "J20")

        charts_sheet = workbook["Charts"]
        function_buffer = BytesIO(function_png)
        function_buffer.seek(0)
        image_buffers.append(function_buffer)
        function_image = ExcelImage(function_buffer)
        function_image.width = 760
        function_image.height = 450
        charts_sheet.add_image(function_image, "A3")

        if convergence_png is not None:
            convergence_buffer = BytesIO(convergence_png)
            convergence_buffer.seek(0)
            image_buffers.append(convergence_buffer)
            convergence_image = ExcelImage(convergence_buffer)
            convergence_image.width = 760
            convergence_image.height = 450
            charts_sheet.add_image(convergence_image, "A28")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit rendering
# =============================================================================
def render_final_result(result: ForwardDifferenceResult) -> None:
    """Render the compact final result card."""

    if not result.success:
        st.error(result.message)
        st.caption(result.stopping_reason)
        return

    st.success(result.message)
    st.markdown(f"**Function:** `{result.function_text}`")
    st.markdown(f"**Derivative:** {result.derivative_name}")
    st.markdown(f"**Formula:** {result.formula_spec.formula_text}")

    metrics = st.columns(2)
    metrics[0].metric(
        "Approximate Derivative",
        format_number(result.primary_approximation),
    )
    metrics[1].metric(
        "Exact Derivative",
        format_number(result.primary_exact_derivative),
    )

    details = st.columns(2)
    details[0].metric("x₀", format_number(result.x_value))
    details[1].metric("h", format_number(result.initial_step_size))

    if result.primary_absolute_error is not None:
        error_columns = st.columns(2)
        error_columns[0].metric(
            "Absolute Error",
            scientific_number(result.primary_absolute_error),
        )
        error_columns[1].metric(
            "Relative Error (%)",
            format_number(result.primary_relative_error),
        )

    for warning in result.warnings:
        st.warning(warning)


def render_primary_calculation(result: ForwardDifferenceResult) -> None:
    """Render the entered-h calculation and Taylor series table."""

    st.subheader("Primary Forward Difference Calculation")
    st.dataframe(
        round_numeric_dataframe(
            create_primary_calculation_dataframe(result)
        ),
        use_container_width=True,
        hide_index=True,
    )

    primary = result.evaluations[0]
    with st.container(border=True):
        st.markdown("#### Formula substitution")
        st.code(primary.substitution_text, language=None)
        st.markdown(
            "**Approximate derivative:** "
            f"{format_number(primary.derivative_approximation, 10)}"
        )
        if primary.exact_derivative is not None:
            st.markdown(
                "**Exact derivative:** "
                f"{format_number(primary.exact_derivative, 10)}"
            )
            st.markdown(
                "**Absolute error:** "
                f"{scientific_number(primary.absolute_error)}"
            )

    st.subheader("Taylor Series Through the Third Derivative")
    st.dataframe(
        round_numeric_dataframe(create_taylor_series_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )
    st.info(
        "The forward formulas are obtained by combining Taylor expansions "
        "so unwanted derivative terms cancel. The table displays the "
        "expansions through f‴(x₀), as requested."
    )


def render_refinement_analysis(result: ForwardDifferenceResult) -> None:
    """Render refinement and convergence analysis."""

    st.subheader("Step-Size Refinement and Error Analysis")
    st.dataframe(
        round_numeric_dataframe(create_error_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    order = result.formula_spec.theoretical_order
    st.info(
        f"The selected formula has truncation error O(h^{order}). "
        f"When h is halved, the truncation error should decrease by "
        f"approximately 2^{order}, provided round-off is not dominant."
    )

    if result.observed_order_latest is not None:
        st.markdown(
            "**Latest observed convergence order:** "
            f"{format_number(result.observed_order_latest, 4)}"
        )

    try:
        convergence_figure = create_convergence_plot(result)
    except ValueError as error:
        st.info(str(error))
    else:
        st.pyplot(convergence_figure, use_container_width=True)
        plt.close(convergence_figure)


def render_function_graph(result: ForwardDifferenceResult) -> None:
    """Render the function graph safely."""

    st.subheader("Function Graph")
    try:
        figure = create_function_plot(result)
    except ValueError as error:
        st.warning(f"The graph could not be displayed. {error}")
    else:
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)


def render_excel_download(result: ForwardDifferenceResult) -> None:
    """Render a cached Excel report download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature
    cached_signature = st.session_state.get(
        "forward_difference_excel_signature"
    )

    if cached_signature != report_signature:
        try:
            report_bytes = create_excel_report(result)
        except (ValueError, OSError, RuntimeError) as error:
            st.error(f"The Excel report could not be generated. {error}")
            return
        st.session_state.forward_difference_excel_report = report_bytes
        st.session_state.forward_difference_excel_signature = report_signature

    report_bytes = st.session_state.get("forward_difference_excel_report")
    if report_bytes is None:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=f"forward_difference_report_{timestamp}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="forward_difference_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Forward Difference Streamlit solver page."""

    st.set_page_config(
        page_title="Forward Difference Solver | Numerical Methods",
        page_icon="📉",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">NUMERICAL DIFFERENTIATION TOOL</div>
                <h1>Forward Difference Solver</h1>
                <p>
                    Enter a function, evaluation point, and step size to compute
                    the first, second, or third derivative using textbook forward
                    finite-difference formulas. Review the complete calculation,
                    Taylor-series basis, refinement, error analysis, graphs, and
                    Excel report.
                </p>

                <div class="method-actions">
                    <a href="/Forward_Difference" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Forward_Difference_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Preserve the final centered solver layout used by the project.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])

    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Write the Function")
                st.markdown(
                    """
                    Enter only the mathematical expression, without `f(x) =`
                    or an equals sign.

                    - Use only **x** as the variable.
                    - Powers: write `x**2`, not **x^2**.
                    - Multiplication: write `2*x`, not `2x`.
                    - Use lowercase functions such as **sin(x)**, **cos(x)**,
                      **exp(x)**, **sqrt(x)**, and **log(x)**.
                    - Use parentheses whenever the order of operations could
                      be unclear.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")
                st.markdown(
                    """
                    - The step size **h** must be positive.
                    - Select the required derivative: **f′**, **f″**, or **f‴**.
                    - The selected formula may require values from **x₀** through
                      **x₀ + 4h**.
                    - Every required forward point must lie in the real domain
                      of the function.
                    - A very small **h** can increase round-off and cancellation
                      errors.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">Function f(x)</div>',
                    unsafe_allow_html=True,
                )
                function_text = st.text_input(
                    "Function f(x)",
                    value=DEFAULT_FUNCTION,
                    placeholder="Example: sin(x), exp(x), or x**3 - 2*x + 1",
                    label_visibility="collapsed",
                    key="forward_difference_function",
                )

                value_columns = st.columns(2)
                with value_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Evaluation point x₀</div>',
                        unsafe_allow_html=True,
                    )
                    x_value = st.number_input(
                        "Evaluation point x0",
                        value=DEFAULT_X_VALUE,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="forward_difference_x_value",
                    )

                with value_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Step size h</div>',
                        unsafe_allow_html=True,
                    )
                    step_size = st.number_input(
                        "Step size h",
                        value=DEFAULT_STEP_SIZE,
                        min_value=0.0,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="forward_difference_step_size",
                    )

                selector_columns = st.columns(2)
                with selector_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Derivative order</div>',
                        unsafe_allow_html=True,
                    )
                    derivative_name = st.selectbox(
                        "Derivative order",
                        options=list(DERIVATIVE_OPTIONS.keys()),
                        index=0,
                        label_visibility="collapsed",
                        key="forward_difference_derivative_order",
                    )

                with selector_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Formula accuracy</div>',
                        unsafe_allow_html=True,
                    )
                    accuracy_name = st.selectbox(
                        "Formula accuracy",
                        options=list(ACCURACY_OPTIONS.keys()),
                        index=0,
                        label_visibility="collapsed",
                        key="forward_difference_accuracy",
                    )

                st.markdown(
                    '<div class="input-label-ui">Step-size refinement levels</div>',
                    unsafe_allow_html=True,
                )
                refinement_levels = st.slider(
                    "Step-size refinement levels",
                    min_value=MIN_REFINEMENT_LEVELS,
                    max_value=MAX_REFINEMENT_LEVELS,
                    value=DEFAULT_REFINEMENT_LEVELS,
                    step=1,
                    label_visibility="collapsed",
                    key="forward_difference_refinement_levels",
                )

                st.caption(
                    "The main result uses the entered h. Additional levels use "
                    "h/2, h/4, and so on for convergence analysis. Calculations "
                    "use full precision; displayed tables are rounded only for readability."
                )

                solve_button_clicked = st.button(
                    "Solve",
                    use_container_width=True,
                    key="forward_difference_solve_button",
                )

        current_input_signature = create_input_signature(
            function_text=function_text,
            x_value=x_value,
            step_size=step_size,
            derivative_name=derivative_name,
            accuracy_name=accuracy_name,
            refinement_levels=refinement_levels,
        )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                stored_result = st.session_state.get("forward_difference_result")

                if stored_result is None:
                    st.info("Enter the function and parameters, then click Solve.")
                elif stored_result.input_signature != current_input_signature:
                    st.info(
                        "The function or numerical parameters have changed. "
                        "Click Solve to calculate a new result."
                    )
                else:
                    render_final_result(stored_result)

        if solve_button_clicked:
            st.session_state.forward_difference_result = solve_forward_difference(
                function_text=function_text,
                raw_x_value=x_value,
                raw_step_size=step_size,
                derivative_name=derivative_name,
                accuracy_name=accuracy_name,
                raw_refinement_levels=refinement_levels,
                input_signature=current_input_signature,
            )
            st.session_state.pop("forward_difference_excel_report", None)
            st.session_state.pop("forward_difference_excel_signature", None)
            st.rerun()

        active_result = st.session_state.get("forward_difference_result")

        if (
            active_result is not None
            and active_result.input_signature == current_input_signature
            and active_result.success
        ):
            st.divider()
            render_primary_calculation(active_result)

            st.divider()
            render_refinement_analysis(active_result)

            st.divider()
            render_function_graph(active_result)

            st.divider()
            render_excel_download(active_result)

            st.divider()
            navigation_left_column, navigation_right_column = st.columns(2)

            with navigation_left_column:
                if st.button(
                    "Review Forward Difference Lesson",
                    use_container_width=True,
                    key="review_forward_difference_lesson",
                ):
                    st.switch_page("pages/Forward_Difference.py")

            with navigation_right_column:
                if st.button(
                    "Back to Solver Menu",
                    use_container_width=True,
                    key="back_to_solver_menu_forward_difference",
                ):
                    st.switch_page("pages/Numerical_Solver.py")

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Numerical Differentiation • Forward Difference</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()