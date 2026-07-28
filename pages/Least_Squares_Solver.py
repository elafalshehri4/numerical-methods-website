from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from components.navigation import navbar
from utilities.ui import load_css

# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
        value: float | int | None,
        decimals: int = 3,
        unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
        value: float | int | None,
        decimals: int = 3,
        unavailable: str = "—",
) -> str:
    """Show three decimals and use × 10ⁿ when fixed notation loses meaning."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Polynomial Least Squares Method"
SUPPORTED_POINT_COUNTS = tuple(range(3, 13))
SUPPORTED_DEGREES = (1, 2, 3)
DEFAULT_POINT_COUNT = 6
DEFAULT_DEGREE = 1
DISPLAY_DECIMALS = 3
MACHINE_EPSILON = np.finfo(float).eps
PIVOT_TOLERANCE = 100.0 * MACHINE_EPSILON
CONDITION_NUMBER_WARNING = 1.0e12
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DEFAULT_X_VALUES = np.array(
    [-2.0, -1.0, 0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0],
    dtype=float,
)
DEFAULT_Y_VALUES = np.array(
    [-2.2, -0.1, 1.4, 3.2, 5.1, 7.0, 8.7, 10.8, 12.6, 14.5, 16.1, 18.3],
    dtype=float,
)


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class LinearSolveStep:
    """One row-elimination operation used to solve the normal equations."""

    step: int
    pivot_stage: int
    pivot_row: int
    target_row: int
    pivot_value: float
    scaled_pivot_ratio: float
    multiplier: float
    operation: str
    matrix_before: np.ndarray
    vector_before: np.ndarray
    matrix_after: np.ndarray
    vector_after: np.ndarray


@dataclass(frozen=True)
class BackSubstitutionStep:
    """One coefficient calculation during back substitution."""

    step: int
    row: int
    coefficient: str
    diagonal_value: float
    right_hand_side: float
    known_sum: float
    numerator: float
    calculated_value: float
    formula: str


@dataclass(frozen=True)
class LeastSquaresResult:
    """Complete result shared by Streamlit renderers and Excel export."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str
    degree: int
    point_count: int
    parameter_count: int
    original_x: np.ndarray
    original_y: np.ndarray
    x_center: float
    x_scale: float
    scaled_x: np.ndarray
    design_matrix: np.ndarray
    normal_matrix: np.ndarray
    normal_rhs: np.ndarray
    coefficients: np.ndarray | None
    scaled_coefficients: np.ndarray | None
    equation_text: str
    fitted_values: np.ndarray | None
    residuals: np.ndarray | None
    absolute_residuals: np.ndarray | None
    squared_residuals: np.ndarray | None
    sse: float | None
    mse: float | None
    rmse: float | None
    mae: float | None
    maximum_absolute_residual: float | None
    residual_standard_error: float | None
    r_squared: float | None
    adjusted_r_squared: float | None
    total_sum_of_squares: float | None
    design_rank: int
    design_condition_number: float
    normal_condition_number: float
    raw_design_condition_number: float
    raw_normal_condition_number: float
    coefficient_history: tuple[LinearSolveStep, ...]
    back_substitution_history: tuple[BackSubstitutionStep, ...]
    warnings: tuple[str, ...]
    pivot_tolerance: float
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def format_number(
        value: float | int | None,
        decimals: int = 3,
) -> str:
    """Format displayed values with three decimals and × 10ⁿ notation."""

    return format_display_number(value, decimals)


def round_numeric_dataframe(
        dataframe: pd.DataFrame,
        decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns for display without altering stored precision."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def coefficient_names(degree: int) -> list[str]:
    """Return coefficient labels in ascending-power order."""

    return [f"a{i}" for i in range(degree + 1)]


def design_column_names(degree: int) -> list[str]:
    """Return readable design-matrix column labels."""

    labels = ["1"]
    for power in range(1, degree + 1):
        labels.append("x" if power == 1 else f"x^{power}")
    return labels


def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def create_input_signature(
        data: pd.DataFrame,
        degree: int,
        pivot_tolerance: float,
) -> str:
    """Create a stable signature used to prevent stale Streamlit results."""

    serialized_rows = []
    for _, row in data.iterrows():
        serialized_rows.append((str(row.get("x", "")), str(row.get("y", ""))))
    payload = repr((serialized_rows, int(degree), float(pivot_tolerance)))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def serialize_matrix(matrix: np.ndarray) -> str:
    """Serialize a matrix clearly for Excel history cells."""

    return np.array2string(
        np.asarray(matrix, dtype=float),
        precision=12,
        suppress_small=False,
        separator=", ",
        max_line_width=1000,
    )


def serialize_vector(vector: np.ndarray) -> str:
    """Serialize a vector clearly for Excel history cells."""

    return np.array2string(
        np.asarray(vector, dtype=float).reshape(-1),
        precision=12,
        suppress_small=False,
        separator=", ",
        max_line_width=1000,
    )


def safe_numeric_preview(values: Sequence[Any]) -> np.ndarray:
    """Return only safely convertible finite values for an error result preview."""

    preview: list[float] = []
    for raw_value in values:
        if raw_value is None:
            continue
        if isinstance(raw_value, str) and not raw_value.strip():
            continue
        try:
            numeric_value = float(raw_value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(numeric_value):
            preview.append(numeric_value)
    return np.asarray(preview, dtype=float)


# =============================================================================
# Input data and validation
# =============================================================================
def default_points_dataframe(point_count: int) -> pd.DataFrame:
    """Return deterministic educational example data."""

    return pd.DataFrame(
        {
            "x": DEFAULT_X_VALUES[:point_count],
            "y": DEFAULT_Y_VALUES[:point_count],
        }
    )


def coerce_real_finite_vector(
        values: Sequence[Any],
        vector_name: str,
) -> np.ndarray:
    """Convert values to a finite real one-dimensional NumPy vector."""

    if values is None:
        raise ValueError(f"{vector_name} is missing.")

    converted: list[float] = []
    for index, raw_value in enumerate(values, start=1):
        if raw_value is None:
            raise ValueError(f"{vector_name} contains an empty value at row {index}.")
        if isinstance(raw_value, str) and not raw_value.strip():
            raise ValueError(f"{vector_name} contains an empty value at row {index}.")
        try:
            numeric_value = float(raw_value)
        except (TypeError, ValueError) as error:
            raise ValueError(
                f"{vector_name} contains a non-numeric value at row {index}."
            ) from error
        if not math.isfinite(numeric_value):
            raise ValueError(
                f"{vector_name} contains NaN or infinity at row {index}."
            )
        converted.append(numeric_value)

    return np.asarray(converted, dtype=float)


def validate_and_prepare_points(
        raw_data: pd.DataFrame,
        degree: int,
) -> tuple[np.ndarray, np.ndarray, tuple[str, ...]]:
    """Validate point data and return clean x and y vectors."""

    if not isinstance(raw_data, pd.DataFrame):
        raise ValueError("The data table could not be read.")
    if "x" not in raw_data.columns or "y" not in raw_data.columns:
        raise ValueError("The data table must contain x and y columns.")
    if degree not in SUPPORTED_DEGREES:
        raise ValueError("Polynomial degree must be 1, 2, or 3.")
    if len(raw_data) not in SUPPORTED_POINT_COUNTS:
        raise ValueError(
            "The number of data points must be between "
            f"{min(SUPPORTED_POINT_COUNTS)} and {max(SUPPORTED_POINT_COUNTS)}."
        )

    x_values = coerce_real_finite_vector(raw_data["x"].tolist(), "x values")
    y_values = coerce_real_finite_vector(raw_data["y"].tolist(), "y values")

    if x_values.size != y_values.size:
        raise ValueError("The number of x values must match the number of y values.")
    if x_values.size < degree + 1:
        raise ValueError(
            f"A degree-{degree} polynomial requires at least {degree + 1} data points."
        )

    unique_x_count = int(np.unique(x_values).size)
    if unique_x_count < degree + 1:
        raise ValueError(
            f"At least {degree + 1} distinct x values are required for a "
            f"degree-{degree} least-squares fit."
        )

    warnings: list[str] = []
    if x_values.size == degree + 1:
        warnings.append(
            "The number of data points equals the number of coefficients. "
            "The polynomial may interpolate the points exactly rather than "
            "provide an overdetermined least-squares fit."
        )
    if unique_x_count < x_values.size:
        warnings.append(
            "Repeated x values were detected. They are allowed because each row "
            "represents an observation, but repeated observations receive repeated weight."
        )

    return x_values, y_values, tuple(warnings)


# =============================================================================
# Least-squares mathematics
# =============================================================================
def scale_x_values(x_values: np.ndarray) -> tuple[float, float, np.ndarray]:
    """Center and scale x to improve the conditioning of polynomial regression."""

    x = np.asarray(x_values, dtype=float).reshape(-1)
    center = float(np.mean(x))
    scale = float(np.max(np.abs(x - center)))

    if not math.isfinite(center) or not math.isfinite(scale):
        raise ValueError("The x-value scaling parameters are not finite.")
    if scale <= np.finfo(float).tiny:
        raise ValueError("The x values do not span a usable numerical interval.")

    scaled = (x - center) / scale
    if not np.all(np.isfinite(scaled)):
        raise ValueError("Centering and scaling the x values produced non-finite values.")

    return center, scale, scaled


def convert_scaled_coefficients_to_original(
        scaled_coefficients: np.ndarray,
        center: float,
        scale: float,
) -> np.ndarray:
    """Convert p(z), z=(x-center)/scale, to coefficients in powers of x."""

    b = np.asarray(scaled_coefficients, dtype=float).reshape(-1)
    degree = b.size - 1
    a = np.zeros_like(b)

    for k in range(degree + 1):
        factor = b[k] / (scale ** k)
        for j in range(k + 1):
            a[j] += (
                    factor
                    * math.comb(k, j)
                    * ((-center) ** (k - j))
            )

    if not np.all(np.isfinite(a)):
        raise ValueError("Converting the scaled polynomial to the x basis failed.")
    return a


def numerical_rank(matrix: np.ndarray) -> tuple[int, float]:
    """Return SVD rank and the scale-aware rank tolerance used."""

    singular_values = np.linalg.svd(np.asarray(matrix, dtype=float), compute_uv=False)
    if singular_values.size == 0:
        return 0, 0.0
    tolerance = (
            MACHINE_EPSILON
            * max(matrix.shape)
            * float(singular_values[0])
    )
    rank = int(np.count_nonzero(singular_values > tolerance))
    return rank, tolerance


def build_design_matrix(x_values: np.ndarray, degree: int) -> np.ndarray:
    """Build X = [1, x, x^2, ..., x^degree] manually."""

    point_count = x_values.size
    design_matrix = np.empty((point_count, degree + 1), dtype=float)
    design_matrix[:, 0] = 1.0
    for power in range(1, degree + 1):
        design_matrix[:, power] = x_values ** power

    if not np.all(np.isfinite(design_matrix)):
        raise ValueError(
            "The polynomial design matrix contains non-finite values. "
            "Use smaller x values or a lower polynomial degree."
        )
    return design_matrix


def calculate_normal_equations(
        design_matrix: np.ndarray,
        y_values: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    """Calculate (X^T X)a = X^T y using direct matrix products."""

    normal_matrix = design_matrix.T @ design_matrix
    normal_rhs = design_matrix.T @ y_values

    if not np.all(np.isfinite(normal_matrix)) or not np.all(np.isfinite(normal_rhs)):
        raise ValueError(
            "The normal equations contain NaN or infinity. "
            "Use smaller input magnitudes or a lower polynomial degree."
        )
    return normal_matrix.astype(float), normal_rhs.astype(float)


def effective_pivot_threshold(
        scale: float,
        pivot_tolerance: float,
        matrix_size: int = 1,
) -> float:
    """Return a scale-aware threshold without rejecting valid small systems."""

    row_scale = abs(float(scale))
    if not math.isfinite(row_scale):
        return math.inf
    return (
            max(float(pivot_tolerance), MACHINE_EPSILON)
            * max(int(matrix_size), 1)
            * max(row_scale, np.finfo(float).tiny)
    )


def solve_normal_equations_manually(
        normal_matrix: np.ndarray,
        normal_rhs: np.ndarray,
        pivot_tolerance: float,
) -> tuple[
    np.ndarray,
    tuple[LinearSolveStep, ...],
    tuple[BackSubstitutionStep, ...],
]:
    """Solve the scaled normal equations by scaled partial pivoting."""

    matrix = np.asarray(normal_matrix, dtype=float).copy()
    rhs = np.asarray(normal_rhs, dtype=float).reshape(-1).copy()
    size = rhs.size

    if matrix.shape != (size, size):
        raise ValueError("The normal-equation matrix must be square.")
    if not np.all(np.isfinite(matrix)) or not np.all(np.isfinite(rhs)):
        raise ValueError("The normal equations contain non-finite values.")

    scales = np.max(np.abs(matrix), axis=1)
    for row_index, row_scale in enumerate(scales):
        threshold = effective_pivot_threshold(row_scale, pivot_tolerance, size)
        if row_scale <= threshold:
            raise ValueError(
                "The normal equations are rank-deficient and cannot determine "
                f"a unique coefficient vector (row {row_index + 1})."
            )

    elimination_history: list[LinearSolveStep] = []
    step_number = 0

    for pivot_column in range(size - 1):
        candidate_rows = np.arange(pivot_column, size)
        ratios = np.abs(matrix[candidate_rows, pivot_column]) / scales[candidate_rows]
        selected_offset = int(np.argmax(ratios))
        selected_row = pivot_column + selected_offset
        selected_ratio = float(ratios[selected_offset])
        selected_pivot = float(matrix[selected_row, pivot_column])
        threshold = effective_pivot_threshold(
            scales[selected_row], pivot_tolerance, size
        )

        if abs(selected_pivot) <= threshold:
            raise ValueError(
                "The normal equations are singular or numerically rank-deficient. "
                "Try a lower polynomial degree or more independent x values."
            )

        if selected_row != pivot_column:
            matrix_before = matrix.copy()
            vector_before = rhs.copy()
            matrix[[pivot_column, selected_row], :] = matrix[
                [selected_row, pivot_column], :
            ]
            rhs[[pivot_column, selected_row]] = rhs[[selected_row, pivot_column]]
            scales[[pivot_column, selected_row]] = scales[[selected_row, pivot_column]]
            step_number += 1
            elimination_history.append(
                LinearSolveStep(
                    step=step_number,
                    pivot_stage=pivot_column + 1,
                    pivot_row=pivot_column + 1,
                    target_row=selected_row + 1,
                    pivot_value=float(matrix[pivot_column, pivot_column]),
                    scaled_pivot_ratio=selected_ratio,
                    multiplier=float("nan"),
                    operation=(
                        f"Swap R{pivot_column + 1} with R{selected_row + 1}"
                    ),
                    matrix_before=matrix_before,
                    vector_before=vector_before,
                    matrix_after=matrix.copy(),
                    vector_after=rhs.copy(),
                )
            )

        pivot_value = float(matrix[pivot_column, pivot_column])
        pivot_threshold = effective_pivot_threshold(
            scales[pivot_column], pivot_tolerance, size
        )
        if abs(pivot_value) <= pivot_threshold:
            raise ValueError(
                "A zero or near-zero pivot occurred while solving the normal equations."
            )

        for target_row in range(pivot_column + 1, size):
            matrix_before = matrix.copy()
            vector_before = rhs.copy()
            multiplier = float(matrix[target_row, pivot_column] / pivot_value)

            matrix[target_row, pivot_column:] -= (
                    multiplier * matrix[pivot_column, pivot_column:]
            )
            rhs[target_row] -= multiplier * rhs[pivot_column]
            matrix[target_row, pivot_column] = 0.0

            cleanup = effective_pivot_threshold(
                scales[target_row], pivot_tolerance, size
            )
            matrix[target_row, np.abs(matrix[target_row]) <= cleanup] = 0.0
            if abs(rhs[target_row]) <= cleanup:
                rhs[target_row] = 0.0

            if not np.all(np.isfinite(matrix)) or not np.all(np.isfinite(rhs)):
                raise ValueError(
                    "Non-finite arithmetic occurred while solving the normal equations."
                )

            step_number += 1
            elimination_history.append(
                LinearSolveStep(
                    step=step_number,
                    pivot_stage=pivot_column + 1,
                    pivot_row=pivot_column + 1,
                    target_row=target_row + 1,
                    pivot_value=pivot_value,
                    scaled_pivot_ratio=selected_ratio,
                    multiplier=multiplier,
                    operation=(
                        f"R{target_row + 1} = R{target_row + 1} - "
                        f"({format_number(multiplier)}) × R{pivot_column + 1}"
                    ),
                    matrix_before=matrix_before,
                    vector_before=vector_before,
                    matrix_after=matrix.copy(),
                    vector_after=rhs.copy(),
                )
            )

    last_threshold = effective_pivot_threshold(
        scales[-1], pivot_tolerance, size
    )
    if abs(float(matrix[-1, -1])) <= last_threshold:
        raise ValueError(
            "The final normal-equation pivot is zero or near zero. "
            "The polynomial coefficients are not uniquely determined."
        )

    coefficients = np.zeros(size, dtype=float)
    back_history: list[BackSubstitutionStep] = []
    back_step = 0

    for row in range(size - 1, -1, -1):
        diagonal = float(matrix[row, row])
        threshold = effective_pivot_threshold(scales[row], pivot_tolerance, size)
        if abs(diagonal) <= threshold:
            raise ValueError(
                "Back substitution encountered a zero or near-zero diagonal value."
            )

        known_sum = float(np.dot(matrix[row, row + 1:], coefficients[row + 1:]))
        numerator = float(rhs[row] - known_sum)
        coefficient_value = numerator / diagonal
        if not math.isfinite(coefficient_value):
            raise ValueError(
                "A non-finite coefficient was produced during back substitution."
            )
        coefficients[row] = coefficient_value
        back_step += 1

        back_history.append(
            BackSubstitutionStep(
                step=back_step,
                row=row + 1,
                coefficient=f"b{row}",
                diagonal_value=diagonal,
                right_hand_side=float(rhs[row]),
                known_sum=known_sum,
                numerator=numerator,
                calculated_value=coefficient_value,
                formula=(
                    f"b{row} = ({format_number(rhs[row])} - "
                    f"{format_number(known_sum)}) / {format_number(diagonal)}"
                ),
            )
        )

    return coefficients, tuple(elimination_history), tuple(back_history)


def evaluate_polynomial(
        coefficients: np.ndarray,
        x_values: float | np.ndarray,
) -> float | np.ndarray:
    """Evaluate ascending-power coefficients with Horner's method."""

    coefficients_array = np.asarray(coefficients, dtype=float).reshape(-1)
    x_array = np.asarray(x_values, dtype=float)
    result = np.zeros_like(x_array, dtype=float)
    for coefficient in coefficients_array[::-1]:
        result = result * x_array + float(coefficient)

    if np.ndim(x_values) == 0:
        return float(np.asarray(result).reshape(-1)[0])
    return result


def polynomial_equation_text(coefficients: np.ndarray | None) -> str:
    """Build a readable polynomial in descending-power visual order."""

    if coefficients is None or coefficients.size == 0:
        return "Unavailable"

    coefficients_array = np.asarray(coefficients, dtype=float).reshape(-1)
    scale = max(float(np.max(np.abs(coefficients_array))), 1.0)
    zero_threshold = 100.0 * MACHINE_EPSILON * scale
    terms: list[str] = []

    for power in range(coefficients_array.size - 1, -1, -1):
        value = float(coefficients_array[power])
        if abs(value) <= zero_threshold:
            continue

        magnitude = abs(value)
        coefficient_text = f"{magnitude:.12g}"
        if power == 0:
            body = coefficient_text
        elif power == 1:
            body = "x" if math.isclose(magnitude, 1.0) else f"{coefficient_text}x"
        else:
            body = (
                f"x^{power}"
                if math.isclose(magnitude, 1.0)
                else f"{coefficient_text}x^{power}"
            )

        if not terms:
            terms.append(f"-{body}" if value < 0 else body)
        else:
            terms.append((" - " if value < 0 else " + ") + body)

    return "ŷ = " + ("".join(terms) if terms else "0")


def calculate_fit_metrics(
        y_values: np.ndarray,
        fitted_values: np.ndarray,
        parameter_count: int,
) -> dict[str, float | None | np.ndarray | int]:
    """Calculate residual, fit, and residual-variance statistics."""

    residuals = y_values - fitted_values
    absolute_residuals = np.abs(residuals)
    squared_residuals = residuals ** 2
    sse = float(math.fsum(float(value) for value in squared_residuals))
    point_count = int(y_values.size)
    mse = float(sse / point_count)
    rmse = float(math.sqrt(mse))
    mae = float(math.fsum(float(value) for value in absolute_residuals) / point_count)
    maximum_absolute_residual = float(np.max(absolute_residuals))

    degrees_of_freedom = point_count - parameter_count
    residual_mean_square = (
        float(sse / degrees_of_freedom)
        if degrees_of_freedom > 0
        else None
    )
    residual_standard_error = (
        float(math.sqrt(residual_mean_square))
        if residual_mean_square is not None
        else None
    )

    mean_y = float(np.mean(y_values))
    centered_y = y_values - mean_y
    total_sum_of_squares = float(
        math.fsum(float(value * value) for value in centered_y)
    )

    sst_threshold = 100.0 * MACHINE_EPSILON * max(
        1.0,
        float(np.max(np.abs(y_values))) ** 2 * point_count,
    )
    if total_sum_of_squares <= sst_threshold:
        r_squared = None
        adjusted_r_squared = None
    else:
        raw_r_squared = 1.0 - sse / total_sum_of_squares
        if -1.0e-12 <= raw_r_squared <= 0.0:
            raw_r_squared = 0.0
        if 1.0 <= raw_r_squared <= 1.0 + 1.0e-12:
            raw_r_squared = 1.0
        r_squared = float(raw_r_squared)
        adjusted_r_squared = (
            float(
                1.0
                - (1.0 - r_squared)
                * (point_count - 1)
                / degrees_of_freedom
            )
            if degrees_of_freedom > 0
            else None
        )

    return {
        "residuals": residuals,
        "absolute_residuals": absolute_residuals,
        "squared_residuals": squared_residuals,
        "sse": sse,
        "mse": mse,
        "rmse": rmse,
        "mae": mae,
        "maximum_absolute_residual": maximum_absolute_residual,
        "degrees_of_freedom": degrees_of_freedom,
        "residual_mean_square": residual_mean_square,
        "residual_standard_error": residual_standard_error,
        "r_squared": r_squared,
        "adjusted_r_squared": adjusted_r_squared,
        "total_sum_of_squares": total_sum_of_squares,
    }


def empty_error_result(
        degree: int,
        raw_x: np.ndarray,
        raw_y: np.ndarray,
        input_signature: str,
        message: str,
        warnings: tuple[str, ...] = (),
) -> LeastSquaresResult:
    """Create a structured failure result without crashing Streamlit."""

    point_count = min(raw_x.size, raw_y.size)
    return LeastSquaresResult(
        status="Error",
        success=False,
        method=METHOD_NAME,
        message=message,
        stopping_reason=message,
        degree=degree,
        point_count=point_count,
        parameter_count=degree + 1,
        original_x=raw_x.copy(),
        original_y=raw_y.copy(),
        x_center=0.0,
        x_scale=1.0,
        scaled_x=np.empty(0, dtype=float),
        design_matrix=np.empty((0, 0), dtype=float),
        normal_matrix=np.empty((0, 0), dtype=float),
        normal_rhs=np.empty(0, dtype=float),
        coefficients=None,
        scaled_coefficients=None,
        equation_text="Unavailable",
        fitted_values=None,
        residuals=None,
        absolute_residuals=None,
        squared_residuals=None,
        sse=None,
        mse=None,
        rmse=None,
        mae=None,
        maximum_absolute_residual=None,
        residual_standard_error=None,
        r_squared=None,
        adjusted_r_squared=None,
        total_sum_of_squares=None,
        design_rank=0,
        design_condition_number=float("inf"),
        normal_condition_number=float("inf"),
        raw_design_condition_number=float("inf"),
        raw_normal_condition_number=float("inf"),
        coefficient_history=(),
        back_substitution_history=(),
        warnings=warnings,
        pivot_tolerance=PIVOT_TOLERANCE,
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


def solve_least_squares(
        raw_data: pd.DataFrame,
        degree: int,
        input_signature: str,
        pivot_tolerance: float = PIVOT_TOLERANCE,
) -> LeastSquaresResult:
    """Run a complete polynomial least-squares workflow safely."""

    raw_x = np.asarray(
        raw_data.get("x", pd.Series(dtype=float)).tolist(),
        dtype=object,
    )
    raw_y = np.asarray(
        raw_data.get("y", pd.Series(dtype=float)).tolist(),
        dtype=object,
    )

    try:
        x_values, y_values, validation_warnings = validate_and_prepare_points(
            raw_data,
            degree,
        )
        parameter_count = degree + 1

        raw_design_matrix = build_design_matrix(x_values, degree)
        raw_design_condition_number = float(np.linalg.cond(raw_design_matrix))
        raw_normal_matrix, _ = calculate_normal_equations(
            raw_design_matrix,
            y_values,
        )
        raw_normal_condition_number = float(np.linalg.cond(raw_normal_matrix))

        x_center, x_scale, scaled_x = scale_x_values(x_values)
        design_matrix = build_design_matrix(scaled_x, degree)
        design_rank, rank_tolerance = numerical_rank(design_matrix)

        if design_rank < parameter_count:
            raise ValueError(
                "The design matrix is rank-deficient. The requested polynomial "
                "coefficients cannot be determined uniquely."
            )

        design_condition_number = float(np.linalg.cond(design_matrix))
        normal_matrix, normal_rhs = calculate_normal_equations(
            design_matrix,
            y_values,
        )
        normal_condition_number = float(np.linalg.cond(normal_matrix))

        warnings = list(validation_warnings)
        if (
                not math.isfinite(raw_design_condition_number)
                or raw_design_condition_number >= CONDITION_NUMBER_WARNING
        ):
            warnings.append(
                "The unscaled monomial design matrix is ill-conditioned. "
                "The solver centers and scales x before forming the normal equations."
            )
        if (
                not math.isfinite(raw_normal_condition_number)
                or raw_normal_condition_number >= CONDITION_NUMBER_WARNING
        ):
            warnings.append(
                "The unscaled normal equations are strongly ill-conditioned, "
                "which is expected because XᵀX approximately squares the condition number."
            )
        if (
                not math.isfinite(normal_condition_number)
                or normal_condition_number >= CONDITION_NUMBER_WARNING
        ):
            warnings.append(
                "Even after centering and scaling, the normal equations remain "
                "ill-conditioned. Interpret the coefficients cautiously."
            )

        scaled_coefficients, elimination_history, back_history = (
            solve_normal_equations_manually(
                normal_matrix,
                normal_rhs,
                pivot_tolerance,
            )
        )
        coefficients = convert_scaled_coefficients_to_original(
            scaled_coefficients,
            x_center,
            x_scale,
        )

        fitted_values = np.asarray(
            evaluate_polynomial(scaled_coefficients, scaled_x),
            dtype=float,
        )
        if not np.all(np.isfinite(fitted_values)):
            raise ValueError("The fitted values contain NaN or infinity.")

        # Check that the converted x-basis polynomial represents the same curve.
        converted_values = np.asarray(
            evaluate_polynomial(coefficients, x_values),
            dtype=float,
        )
        conversion_difference = float(
            np.max(np.abs(converted_values - fitted_values))
        )
        conversion_scale = max(1.0, float(np.max(np.abs(fitted_values))))
        if conversion_difference > 1.0e-8 * conversion_scale:
            warnings.append(
                "The expanded x-basis coefficients lose some precision because "
                "the original x values are strongly offset or scaled. The fitted "
                "values and graphs use the numerically safer scaled polynomial."
            )

        metrics = calculate_fit_metrics(
            y_values,
            fitted_values,
            parameter_count,
        )

        return LeastSquaresResult(
            status="Success",
            success=True,
            method=METHOD_NAME,
            message="Execution completed successfully.",
            stopping_reason=(
                "The centered-and-scaled normal equations were solved and all "
                "fitted values and residual statistics were calculated successfully."
            ),
            degree=degree,
            point_count=x_values.size,
            parameter_count=parameter_count,
            original_x=x_values.copy(),
            original_y=y_values.copy(),
            x_center=x_center,
            x_scale=x_scale,
            scaled_x=scaled_x.copy(),
            design_matrix=design_matrix.copy(),
            normal_matrix=normal_matrix.copy(),
            normal_rhs=normal_rhs.copy(),
            coefficients=coefficients.copy(),
            scaled_coefficients=scaled_coefficients.copy(),
            equation_text=polynomial_equation_text(coefficients),
            fitted_values=fitted_values.copy(),
            residuals=np.asarray(metrics["residuals"], dtype=float),
            absolute_residuals=np.asarray(metrics["absolute_residuals"], dtype=float),
            squared_residuals=np.asarray(metrics["squared_residuals"], dtype=float),
            sse=float(metrics["sse"]),
            mse=float(metrics["mse"]),
            rmse=float(metrics["rmse"]),
            mae=float(metrics["mae"]),
            maximum_absolute_residual=float(metrics["maximum_absolute_residual"]),
            residual_standard_error=(
                None
                if metrics["residual_standard_error"] is None
                else float(metrics["residual_standard_error"])
            ),
            r_squared=(
                None if metrics["r_squared"] is None else float(metrics["r_squared"])
            ),
            adjusted_r_squared=(
                None
                if metrics["adjusted_r_squared"] is None
                else float(metrics["adjusted_r_squared"])
            ),
            total_sum_of_squares=float(metrics["total_sum_of_squares"]),
            design_rank=design_rank,
            design_condition_number=design_condition_number,
            normal_condition_number=normal_condition_number,
            raw_design_condition_number=raw_design_condition_number,
            raw_normal_condition_number=raw_normal_condition_number,
            coefficient_history=elimination_history,
            back_substitution_history=back_history,
            warnings=tuple(warnings),
            pivot_tolerance=pivot_tolerance,
            input_signature=input_signature,
            execution_datetime=current_report_datetime(),
        )

    except (ValueError, FloatingPointError, OverflowError, ZeroDivisionError) as error:
        clean_x = safe_numeric_preview(raw_x.tolist())
        clean_y = safe_numeric_preview(raw_y.tolist())
        return empty_error_result(
            degree=degree,
            raw_x=clean_x,
            raw_y=clean_y,
            input_signature=input_signature,
            message=str(error),
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def input_points_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return the original points and scaled coordinates."""

    return pd.DataFrame(
        {
            "Point": np.arange(1, result.point_count + 1),
            "x": result.original_x,
            "Scaled z": result.scaled_x,
            "y": result.original_y,
        }
    )


def design_matrix_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return the labeled centered-and-scaled design matrix."""

    labels = ["1"] + [
        "z" if power == 1 else f"z^{power}"
        for power in range(1, result.degree + 1)
    ]
    dataframe = pd.DataFrame(result.design_matrix, columns=labels)
    dataframe.insert(0, "Point", np.arange(1, result.point_count + 1))
    dataframe.insert(1, "x", result.original_x)
    dataframe.insert(2, "z = (x-center)/scale", result.scaled_x)
    return dataframe


def normal_matrix_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return XᵀX with coefficient labels."""

    names = coefficient_names(result.degree)
    return pd.DataFrame(result.normal_matrix, index=names, columns=names)


def normal_rhs_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return Xᵀy."""

    return pd.DataFrame(
        {
            "Coefficient Equation": coefficient_names(result.degree),
            "X^T y": result.normal_rhs,
        }
    )


def coefficient_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return original x-basis and scaled z-basis coefficients."""

    if result.coefficients is None or result.scaled_coefficients is None:
        return pd.DataFrame(
            columns=[
                "Power",
                "x-basis Coefficient",
                "x-basis Value",
                "z-basis Coefficient",
                "z-basis Value",
            ]
        )

    return pd.DataFrame(
        {
            "Power": np.arange(result.degree + 1),
            "x-basis Coefficient": coefficient_names(result.degree),
            "x-basis Value": result.coefficients,
            "z-basis Coefficient": [f"b{i}" for i in range(result.degree + 1)],
            "z-basis Value": result.scaled_coefficients,
        }
    )


def fitted_values_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return fitted values and residuals for every observation."""

    if not result.success or result.fitted_values is None:
        return pd.DataFrame()
    return pd.DataFrame(
        {
            "Point": np.arange(1, result.point_count + 1),
            "x": result.original_x,
            "Observed y": result.original_y,
            "Fitted y": result.fitted_values,
            "Residual (y - fitted y)": result.residuals,
            "Absolute Residual": result.absolute_residuals,
            "Squared Residual": result.squared_residuals,
        }
    )


def elimination_history_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return the normal-equation elimination summary."""

    rows = []
    for record in result.coefficient_history:
        rows.append(
            {
                "Step": record.step,
                "Pivot Stage": record.pivot_stage,
                "Pivot Row": record.pivot_row,
                "Target Row": record.target_row,
                "Pivot Value": record.pivot_value,
                "Scaled Pivot Ratio": record.scaled_pivot_ratio,
                "Multiplier": record.multiplier,
                "Operation": record.operation,
            }
        )
    return pd.DataFrame(rows)


def back_substitution_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return coefficient back-substitution calculations."""

    rows = []
    for record in result.back_substitution_history:
        rows.append(
            {
                "Step": record.step,
                "Row": record.row,
                "Coefficient": record.coefficient,
                "Diagonal": record.diagonal_value,
                "Right-Hand Side": record.right_hand_side,
                "Known Sum": record.known_sum,
                "Numerator": record.numerator,
                "Calculated Value": record.calculated_value,
                "Formula": record.formula,
            }
        )
    return pd.DataFrame(rows)


def residual_metrics_dataframe(result: LeastSquaresResult) -> pd.DataFrame:
    """Return scalar error and fit statistics with clear denominators."""

    residual_dof = result.point_count - result.parameter_count
    residual_mean_square = (
        result.sse / residual_dof
        if result.sse is not None and residual_dof > 0
        else None
    )

    return pd.DataFrame(
        {
            "Metric": [
                "Sum of Squared Errors (SSE)",
                "Prediction MSE = SSE / n",
                "Prediction RMSE = sqrt(SSE / n)",
                "Mean Absolute Error (MAE)",
                "Maximum Absolute Residual",
                "Residual Degrees of Freedom (n - p)",
                "Residual Mean Square = SSE / (n - p)",
                "Residual Standard Error",
                "R-squared",
                "Adjusted R-squared",
                "Total Sum of Squares (SST)",
            ],
            "Value": [
                result.sse,
                result.mse,
                result.rmse,
                result.mae,
                result.maximum_absolute_residual,
                residual_dof,
                residual_mean_square,
                result.residual_standard_error,
                result.r_squared,
                result.adjusted_r_squared,
                result.total_sum_of_squares,
            ],
        }
    )


def build_plot_dataframe(
        result: LeastSquaresResult,
        sample_count: int = 500,
) -> pd.DataFrame:
    """Build sorted smooth-curve data for charts and Excel."""

    x_min = float(np.min(result.original_x))
    x_max = float(np.max(result.original_x))
    padding = 0.05 * max(x_max - x_min, 1.0)
    smooth_x = np.linspace(x_min - padding, x_max + padding, sample_count)
    smooth_z = (smooth_x - result.x_center) / result.x_scale
    smooth_y = np.asarray(
        evaluate_polynomial(result.scaled_coefficients, smooth_z),
        dtype=float,
    )
    return pd.DataFrame({"x": smooth_x, "Fitted Polynomial": smooth_y})


def figure_to_png_bytes(figure: plt.Figure) -> bytes:
    """Serialize a matplotlib figure as PNG bytes."""

    buffer = BytesIO()
    figure.savefig(buffer, format="png", dpi=180, bbox_inches="tight")
    buffer.seek(0)
    return buffer.getvalue()


def add_excel_image(
        worksheet: Any,
        image_bytes: bytes,
        anchor: str,
        width: int = 780,
        height: int = 500,
) -> None:
    """Insert PNG bytes into an Excel worksheet."""

    stream = BytesIO(image_bytes)
    image = ExcelImage(stream)
    image.width = width
    image.height = height
    worksheet.add_image(image, anchor)


# =============================================================================
# Excel export
# =============================================================================
def style_excel_workbook(workbook: Any) -> None:
    """Apply professional workbook formatting."""

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="E8E2F4")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                text = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(text))
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000"
            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12),
                60,
            )

        if worksheet.title == "Summary":
            worksheet["A1"].fill = title_fill
            worksheet["B1"].fill = title_fill


def create_excel_report(result: LeastSquaresResult) -> bytes:
    """Create a complete in-memory XLSX report."""

    if not result.success or result.coefficients is None:
        raise ValueError("Only successful least-squares results can be exported.")

    residual_dof = result.point_count - result.parameter_count
    residual_mean_square = (
        result.sse / residual_dof
        if result.sse is not None and residual_dof > 0
        else None
    )

    summary_dataframe = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "Polynomial Degree",
                "Number of Data Points",
                "Number of Coefficients",
                "Fitted Equation in x",
                "x Center",
                "x Scale",
                "Scaled Variable",
                "x-basis Coefficients",
                "z-basis Coefficients",
                "SSE",
                "Prediction MSE = SSE/n",
                "Prediction RMSE",
                "MAE",
                "Maximum Absolute Residual",
                "Residual Degrees of Freedom",
                "Residual Mean Square = SSE/(n-p)",
                "Residual Standard Error",
                "R-squared",
                "Adjusted R-squared",
                "Design Matrix Rank",
                "Scaled Design Condition Number",
                "Scaled Normal Matrix Condition Number",
                "Raw Design Condition Number",
                "Raw Normal Matrix Condition Number",
                "Pivot Tolerance Factor",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                result.degree,
                result.point_count,
                result.parameter_count,
                result.equation_text,
                result.x_center,
                result.x_scale,
                "z = (x - center) / scale",
                ", ".join(
                    f"a{i} = {value:.15g}"
                    for i, value in enumerate(result.coefficients)
                ),
                ", ".join(
                    f"b{i} = {value:.15g}"
                    for i, value in enumerate(result.scaled_coefficients)
                ),
                result.sse,
                result.mse,
                result.rmse,
                result.mae,
                result.maximum_absolute_residual,
                residual_dof,
                residual_mean_square,
                result.residual_standard_error,
                result.r_squared,
                result.adjusted_r_squared,
                result.design_rank,
                result.design_condition_number,
                result.normal_condition_number,
                result.raw_design_condition_number,
                result.raw_normal_condition_number,
                result.pivot_tolerance,
                " | ".join(result.warnings) if result.warnings else "None",
                result.stopping_reason,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )

    normal_augmented = pd.DataFrame(
        np.column_stack((result.normal_matrix, result.normal_rhs)),
        index=[f"b{i}" for i in range(result.degree + 1)],
        columns=[f"b{i}" for i in range(result.degree + 1)] + ["X^T y"],
    )

    elimination_detail_rows = []
    for record in result.coefficient_history:
        elimination_detail_rows.append(
            {
                "Step": record.step,
                "Pivot Stage": record.pivot_stage,
                "Pivot Row": record.pivot_row,
                "Target/Swapped Row": record.target_row,
                "Pivot Value": record.pivot_value,
                "Scaled Pivot Ratio": record.scaled_pivot_ratio,
                "Multiplier": record.multiplier,
                "Operation": record.operation,
                "Matrix Before": serialize_matrix(record.matrix_before),
                "Vector Before": serialize_vector(record.vector_before),
                "Matrix After": serialize_matrix(record.matrix_after),
                "Vector After": serialize_vector(record.vector_after),
            }
        )
    elimination_details = pd.DataFrame(elimination_detail_rows)

    plot_dataframe = build_plot_dataframe(result)
    fitted_dataframe = fitted_values_dataframe(result)

    fit_figure = create_fit_figure(result)
    residual_figure = create_residual_figure(result)
    fit_png = figure_to_png_bytes(fit_figure)
    residual_png = figure_to_png_bytes(residual_figure)
    plt.close(fit_figure)
    plt.close(residual_figure)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_dataframe.to_excel(writer, sheet_name="Summary", index=False)
        input_points_dataframe(result).to_excel(writer, sheet_name="Input Data", index=False)
        design_matrix_dataframe(result).to_excel(writer, sheet_name="Scaled Design Matrix", index=False)
        normal_matrix_dataframe(result).to_excel(writer, sheet_name="Scaled Normal Matrix")
        normal_rhs_dataframe(result).to_excel(writer, sheet_name="Scaled Normal RHS", index=False)
        normal_augmented.to_excel(writer, sheet_name="Scaled Normal Equations")
        elimination_details.to_excel(writer, sheet_name="Coefficient Solve", index=False)
        back_substitution_dataframe(result).to_excel(writer, sheet_name="Back Substitution", index=False)
        coefficient_dataframe(result).to_excel(writer, sheet_name="Coefficients", index=False)
        fitted_dataframe.to_excel(writer, sheet_name="Fitted Values", index=False)
        residual_metrics_dataframe(result).to_excel(writer, sheet_name="Residual Analysis", index=False)
        plot_dataframe.to_excel(writer, sheet_name="Plot Data", index=False)

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Polynomial Least-Squares Report Plots"
        plots_sheet["A1"].font = Font(bold=True, size=14)
        add_excel_image(plots_sheet, fit_png, "A3")
        add_excel_image(plots_sheet, residual_png, "A31")

        style_excel_workbook(workbook)

        summary_sheet = workbook["Summary"]
        plot_sheet = workbook["Plot Data"]
        fitted_sheet = workbook["Fitted Values"]

        fit_chart = ScatterChart()
        fit_chart.title = "Observed Data and Least-Squares Fit"
        fit_chart.x_axis.title = "x"
        fit_chart.y_axis.title = "y"
        fit_chart.height = 9
        fit_chart.width = 16

        smooth_x_ref = Reference(plot_sheet, min_col=1, min_row=2, max_row=len(plot_dataframe) + 1)
        smooth_y_ref = Reference(plot_sheet, min_col=2, min_row=2, max_row=len(plot_dataframe) + 1)
        smooth_series = Series(smooth_y_ref, smooth_x_ref, title="Fitted Polynomial")
        fit_chart.series.append(smooth_series)

        observed_x_ref = Reference(fitted_sheet, min_col=2, min_row=2, max_row=result.point_count + 1)
        observed_y_ref = Reference(fitted_sheet, min_col=3, min_row=2, max_row=result.point_count + 1)
        observed_series = Series(observed_y_ref, observed_x_ref, title="Observed Data")
        observed_series.marker.symbol = "circle"
        observed_series.graphicalProperties.line.noFill = True
        fit_chart.series.append(observed_series)
        summary_sheet.add_chart(fit_chart, "D2")

        residual_chart = ScatterChart()
        residual_chart.title = "Residuals versus x"
        residual_chart.x_axis.title = "x"
        residual_chart.y_axis.title = "Residual"
        residual_chart.height = 9
        residual_chart.width = 16
        residual_ref = Reference(fitted_sheet, min_col=5, min_row=2, max_row=result.point_count + 1)
        residual_series = Series(residual_ref, observed_x_ref, title="Residual")
        residual_series.marker.symbol = "circle"
        residual_chart.series.append(residual_series)
        summary_sheet.add_chart(residual_chart, "D22")

        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# Matplotlib figures
# =============================================================================
def create_fit_figure(result: LeastSquaresResult) -> plt.Figure:
    """Create the observed-data and fitted-curve plot."""

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.scatter(result.original_x, result.original_y, label="Observed Data")

    plot_dataframe = build_plot_dataframe(result)
    axis.plot(
        plot_dataframe["x"],
        plot_dataframe["Fitted Polynomial"],
        linewidth=2,
        label=result.equation_text,
    )
    axis.set_title("Least-Squares Polynomial Fit")
    axis.set_xlabel("x")
    axis.set_ylabel("y")
    axis.axhline(0.0, linewidth=1)
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_residual_figure(result: LeastSquaresResult) -> plt.Figure:
    """Create a residual-versus-x diagnostic plot."""

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.scatter(result.original_x, result.residuals, label="Residual")
    axis.axhline(0.0, linewidth=1.5, label="Zero Residual")
    for x_value, residual in zip(result.original_x, result.residuals):
        axis.vlines(x_value, 0.0, residual, linewidth=1)
    axis.set_title("Residual Analysis")
    axis.set_xlabel("x")
    axis.set_ylabel("Residual (y - fitted y)")
    axis.grid(True)
    axis.legend()
    figure.tight_layout()
    return figure


# =============================================================================
# Streamlit rendering helpers
# =============================================================================

def finalize_excel_report_with_visible_charts(report_bytes: bytes) -> bytes:
    """Place existing workbook charts on Summary so they are immediately visible.

    The report data and worksheets are preserved. If Excel chart post-processing is
    unavailable for any reason, the original report is returned unchanged rather
    than risking a damaged workbook.
    """
    try:
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_workbook

        workbook = _load_workbook(_BytesIO(report_bytes))
        if "Summary" not in workbook.sheetnames:
            return report_bytes

        summary_sheet = workbook["Summary"]
        collected_charts = list(summary_sheet._charts)
        summary_sheet._charts = []

        for worksheet in workbook.worksheets:
            if worksheet is summary_sheet:
                continue
            if worksheet._charts:
                collected_charts.extend(list(worksheet._charts))
                worksheet._charts = []

        # Put every chart in the existing Summary worksheet, below one another.
        # D2 keeps the first chart visible beside the main results when the report opens.
        for chart_index, chart in enumerate(collected_charts):
            anchor_row = 2 + chart_index * 19
            summary_sheet.add_chart(chart, f"D{anchor_row}")

        workbook.active = workbook.sheetnames.index("Summary")
        output = _BytesIO()
        workbook.save(output)
        return output.getvalue()
    except Exception:
        return report_bytes


def render_final_result(result: LeastSquaresResult) -> None:
    """Render the compact final-result card."""

    if not result.success:
        st.error(result.message)
        return

    st.success(result.message)
    st.markdown(f"### {result.equation_text}")

    metric_columns = st.columns(3)
    metric_columns[0].metric("Polynomial Degree", result.degree)
    metric_columns[1].metric("Data Points", result.point_count)
    metric_columns[2].metric("RMSE", format_number(result.rmse))

    metric_columns = st.columns(3)
    metric_columns[0].metric("R²", format_number(result.r_squared))
    metric_columns[1].metric("SSE", format_number(result.sse))
    metric_columns[2].metric(
        "Maximum |Residual|",
        format_number(result.maximum_absolute_residual),
    )

    st.caption(result.stopping_reason)
    for warning in result.warnings:
        st.warning(warning)


def render_input_and_design(result: LeastSquaresResult) -> None:
    """Render original points and the polynomial design matrix."""

    st.subheader("Original Data and Design Matrix")
    left_column, right_column = st.columns(2)
    with left_column:
        st.markdown("**Input Data Points**")
        st.dataframe(
            round_numeric_dataframe(input_points_dataframe(result)),
            use_container_width=True,
            hide_index=True,
        )
    with right_column:
        st.markdown("**Design Matrix X**")
        st.dataframe(
            round_numeric_dataframe(design_matrix_dataframe(result)),
            use_container_width=True,
            hide_index=True,
        )

    st.caption(
        f"Scaled design-matrix rank: {result.design_rank} of {result.parameter_count}. "
        f"Scaled condition number: {format_number(result.design_condition_number)}. "
        f"Raw condition number: {format_number(result.raw_design_condition_number)}."
    )


def render_normal_equations(result: LeastSquaresResult) -> None:
    """Render XᵀX, Xᵀy, and the manual coefficient solution history."""

    st.subheader("Normal Equations")
    st.latex(r"z=(x-\bar{x})/s,\qquad (Z^T Z)\,b=Z^T y")

    left_column, right_column = st.columns([1.25, 0.75])
    with left_column:
        st.markdown("**Scaled Normal Matrix ZᵀZ**")
        st.dataframe(
            round_numeric_dataframe(normal_matrix_dataframe(result)),
            use_container_width=True,
        )
    with right_column:
        st.markdown("**Right-Hand Side Zᵀy**")
        st.dataframe(
            round_numeric_dataframe(normal_rhs_dataframe(result)),
            use_container_width=True,
            hide_index=True,
        )

    st.caption(
        "The centered-and-scaled coefficient system is solved manually using "
        "scaled partial-pivot Gaussian elimination. The resulting z-basis "
        "coefficients are converted to the original x basis for the final equation."
    )

    if result.coefficient_history:
        st.markdown("**Coefficient-System Elimination Summary**")
        st.dataframe(
            round_numeric_dataframe(elimination_history_dataframe(result)),
            use_container_width=True,
            hide_index=True,
        )

        for record in result.coefficient_history:
            with st.expander(
                    f"Step {record.step}: {record.operation}",
                    expanded=False,
            ):
                st.markdown(
                    f"**Pivot:** {format_number(record.pivot_value)}  "
                    f"**Multiplier:** {format_number(record.multiplier)}"
                )
                before_column, after_column = st.columns(2)
                with before_column:
                    st.markdown("**Before**")
                    before_augmented = np.column_stack(
                        (record.matrix_before, record.vector_before)
                    )
                    st.dataframe(
                        pd.DataFrame(before_augmented).round(DISPLAY_DECIMALS),
                        use_container_width=True,
                        hide_index=True,
                    )
                with after_column:
                    st.markdown("**After**")
                    after_augmented = np.column_stack(
                        (record.matrix_after, record.vector_after)
                    )
                    st.dataframe(
                        pd.DataFrame(after_augmented).round(DISPLAY_DECIMALS),
                        use_container_width=True,
                        hide_index=True,
                    )

    if result.back_substitution_history:
        st.markdown("**Coefficient Back Substitution**")
        st.dataframe(
            round_numeric_dataframe(back_substitution_dataframe(result)),
            use_container_width=True,
            hide_index=True,
        )


def render_coefficients_and_fit(result: LeastSquaresResult) -> None:
    """Render coefficients, fitted values, and the function graph."""

    st.subheader("Final Polynomial and Fitted Values")
    st.markdown(f"### {result.equation_text}")

    left_column, right_column = st.columns([0.65, 1.35])
    with left_column:
        st.markdown("**Coefficient Table**")
        st.dataframe(
            round_numeric_dataframe(coefficient_dataframe(result)),
            use_container_width=True,
            hide_index=True,
        )
    with right_column:
        st.markdown("**Observed, Fitted, and Residual Values**")
        st.dataframe(
            round_numeric_dataframe(fitted_values_dataframe(result)),
            use_container_width=True,
            hide_index=True,
        )

    st.markdown("**Function Graph**")
    try:
        fit_figure = create_fit_figure(result)
        st.pyplot(fit_figure, use_container_width=True)
        plt.close(fit_figure)
    except (ValueError, TypeError, FloatingPointError) as error:
        st.warning(f"The function graph could not be displayed: {error}")


def render_error_analysis(result: LeastSquaresResult) -> None:
    """Render residual statistics and residual diagnostics."""

    st.subheader("Error Analysis")

    st.dataframe(
        round_numeric_dataframe(residual_metrics_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    metric_columns = st.columns(4)
    metric_columns[0].metric("SSE", format_number(result.sse))
    metric_columns[1].metric("MSE", format_number(result.mse))
    metric_columns[2].metric("RMSE", format_number(result.rmse))
    metric_columns[3].metric("MAE", format_number(result.mae))

    metric_columns = st.columns(3)
    metric_columns[0].metric("R²", format_number(result.r_squared))
    metric_columns[1].metric(
        "Adjusted R²",
        format_number(result.adjusted_r_squared),
    )
    metric_columns[2].metric(
        "Residual Standard Error",
        format_number(result.residual_standard_error),
    )

    if result.r_squared is None:
        st.info(
            "R² is undefined because all observed y values are identical, so "
            "the total sum of squares is zero."
        )

    st.markdown("**Residual Plot**")
    try:
        residual_figure = create_residual_figure(result)
        st.pyplot(residual_figure, use_container_width=True)
        plt.close(residual_figure)
    except (ValueError, TypeError, FloatingPointError) as error:
        st.warning(f"The residual graph could not be displayed: {error}")

    st.caption(
        "Residuals are calculated as observed y minus fitted y. The residual "
        "plot should be inspected for patterns; random scatter around zero is "
        "generally more consistent with an adequate polynomial model."
    )


def render_excel_download(result: LeastSquaresResult) -> None:
    """Generate Excel and CSV downloads."""

    st.subheader("Reports")
    report_signature = st.session_state.get("least_squares_excel_signature")
    if report_signature != result.input_signature:
        try:
            st.session_state.least_squares_excel_report = create_excel_report(result)
            st.session_state.least_squares_excel_signature = result.input_signature
        except (ValueError, TypeError, OSError, ArithmeticError) as error:
            st.error(f"The Excel report could not be generated: {error}")
            return

    report_bytes = st.session_state.get("least_squares_excel_report")
    if report_bytes is None:
        st.error("The Excel report is unavailable.")
        return

    date_text = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    download_columns = st.columns(2)
    download_columns[0].download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=f"least_squares_method_report_{date_text}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="least_squares_download_excel",
    )

    csv_bytes = fitted_values_dataframe(result).to_csv(index=False).encode("utf-8-sig")
    download_columns[1].download_button(
        label="Download Fitted Values CSV",
        data=csv_bytes,
        file_name=f"least_squares_fitted_values_{date_text}.csv",
        mime="text/csv",
        use_container_width=True,
        key="least_squares_download_csv",
    )


# =============================================================================
# Streamlit page layout
# =============================================================================
def render_page() -> None:
    """Render the existing website layout and connect the solver logic."""

    st.set_page_config(
        page_title="Least Squares Method Solver | Numerical Methods",
        page_icon="📈",
        layout="wide",
    )
    load_css()

    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">CURVE FITTING TOOL</div>
                <h1>Least Squares Method Solver</h1>
                <p>
                    Enter experimental data, choose a polynomial degree, and
                    follow the design matrix, normal equations, coefficient
                    calculations, fitted curve, and residual analysis.
                </p>

                <div class="method-actions">
                    <a href="/Least_Squares_Method" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Least_Squares_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Match the centered Bisection solver content width.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader('How to Enter the Data')
                st.markdown(
                    """
                Enter the measured data as paired numerical values **(xᵢ, yᵢ)**.

                - Each row must contain finite values.
                - Use enough points for the selected polynomial degree.
                - Choose degree **1**, **2**, or **3** according to the expected trend.
                - The fitted curve does not usually pass through every point because it minimizes the total squared residual.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader('Before Solving')
                st.markdown(
                    """
                - The number of data points must be greater than the polynomial degree.
                - The **x** data must provide enough independent information for the normal equations.
                - A higher degree can reduce training residuals but may overfit the data.
                - Inspect residuals and the condition number, not only the final coefficients.
                - The solver centers and scales x internally before forming the normal equations.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                selection_columns = st.columns(2)
                with selection_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Number of data points</div>',
                        unsafe_allow_html=True,
                    )
                    point_count = st.selectbox(
                        "Number of data points",
                        options=SUPPORTED_POINT_COUNTS,
                        index=SUPPORTED_POINT_COUNTS.index(DEFAULT_POINT_COUNT),
                        label_visibility="collapsed",
                        key="least_squares_point_count",
                    )
                with selection_columns[1]:
                    available_degrees = tuple(
                        degree
                        for degree in SUPPORTED_DEGREES
                        if degree + 1 <= point_count
                    )
                    st.markdown(
                        '<div class="input-label-ui">Polynomial degree</div>',
                        unsafe_allow_html=True,
                    )
                    degree = st.selectbox(
                        "Polynomial degree",
                        options=available_degrees,
                        index=(
                            available_degrees.index(DEFAULT_DEGREE)
                            if DEFAULT_DEGREE in available_degrees
                            else 0
                        ),
                        format_func=lambda value: f"Degree {value}",
                        label_visibility="collapsed",
                        key="least_squares_degree",
                    )

                st.markdown(
                    '<div class="input-label-ui">Data points</div>',
                    unsafe_allow_html=True,
                )
                data_editor_key = f"least_squares_data_editor_{point_count}"
                edited_data = st.data_editor(
                    default_points_dataframe(point_count),
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    column_config={
                        "x": st.column_config.NumberColumn(
                            "x",
                            help="Independent-variable value",
                            format="%.10g",
                        ),
                        "y": st.column_config.NumberColumn(
                            "y",
                            help="Observed dependent-variable value",
                            format="%.10g",
                        ),
                    },
                    key=data_editor_key,
                )

                st.caption(
                    "Enter finite real values. At least degree + 1 distinct x values "
                    "are required. Repeated x values are allowed when the design "
                    "matrix remains full rank."
                )

                solve_button_clicked = st.button(
                    "Solve",
                    use_container_width=True,
                    key="least_squares_solve_button",
                )

        current_input_signature = create_input_signature(
            edited_data,
            degree,
            PIVOT_TOLERANCE,
        )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                stored_result = st.session_state.get("least_squares_result")
                if stored_result is None:
                    st.info("Enter the data points and click Solve to see the result.")
                elif stored_result.input_signature != current_input_signature:
                    st.info(
                        "The data or polynomial degree has changed. Click Solve to "
                        "calculate a new result."
                    )
                else:
                    render_final_result(stored_result)

        if solve_button_clicked:
            st.session_state.least_squares_result = solve_least_squares(
                raw_data=edited_data.copy(),
                degree=degree,
                input_signature=current_input_signature,
                pivot_tolerance=PIVOT_TOLERANCE,
            )
            st.session_state.pop("least_squares_excel_report", None)
            st.session_state.pop("least_squares_excel_signature", None)
            st.rerun()

        active_result = st.session_state.get("least_squares_result")
        if (
                active_result is not None
                and active_result.input_signature == current_input_signature
        ):
            if active_result.success:
                st.divider()
                render_input_and_design(active_result)

                st.divider()
                render_normal_equations(active_result)

                st.divider()
                render_coefficients_and_fit(active_result)

                st.divider()
                render_error_analysis(active_result)

                st.divider()
                render_excel_download(active_result)

                st.divider()
                navigation_left_column, navigation_right_column = st.columns(2)

                with navigation_left_column:
                    if st.button(
                            "Review Least Squares Lesson",
                            use_container_width=True,
                            key="review_least_squares_lesson",
                    ):
                        st.switch_page("pages/Least_Squares_Method.py")

                with navigation_right_column:
                    if st.button(
                            "Back to Solver Menu",
                            use_container_width=True,
                            key="back_to_solver_menu_least_squares",
                    ):
                        st.switch_page("pages/Numerical_Solver.py")

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Least Squares Method • Curve Fitting</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()