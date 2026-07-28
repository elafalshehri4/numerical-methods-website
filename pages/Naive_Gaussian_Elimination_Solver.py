from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
from matplotlib.figure import Figure
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from components.navigation import navbar
from utilities.ui import load_css



# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Naïve Gaussian Elimination"
SUPPORTED_SIZES = (2, 3, 4, 5, 6)
DEFAULT_SIZE = 2
DISPLAY_DECIMALS = 3
PIVOT_TOLERANCE = 100.0 * np.finfo(float).eps
NEAR_ZERO_WARNING_LIMIT = 1.0e-8
CONDITION_NUMBER_WARNING = 1.0e12
RELATIVE_RESIDUAL_WARNING = 1.0e-10
GROWTH_FACTOR_WARNING = 1.0e6
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DEFAULT_SYSTEMS: dict[int, tuple[np.ndarray, np.ndarray]] = {
    2: (
        np.array(
            [
                [2.0, 1.0],
                [5.0, 7.0],
            ],
            dtype=float,
        ),
        np.array([11.0, 13.0], dtype=float),
    ),
    3: (
        np.array(
            [
                [3.0, 2.0, -4.0],
                [2.0, 3.0, 3.0],
                [5.0, -3.0, 1.0],
            ],
            dtype=float,
        ),
        np.array([3.0, 15.0, 14.0], dtype=float),
    ),
    4: (
        np.array(
            [
                [10.0, 2.0, -1.0, 1.0],
                [1.0, 8.0, 2.0, -1.0],
                [2.0, -1.0, 9.0, 1.0],
                [1.0, 1.0, -1.0, 7.0],
            ],
            dtype=float,
        ),
        np.array([18.0, 12.0, -6.0, 25.0], dtype=float),
    ),
    5: (
        np.array(
            [
                [12.0, 1.0, -1.0, 2.0, 0.0],
                [1.0, 11.0, 2.0, -1.0, 1.0],
                [-1.0, 2.0, 10.0, 1.0, -1.0],
                [2.0, -1.0, 1.0, 9.0, 2.0],
                [0.0, 1.0, -1.0, 2.0, 8.0],
            ],
            dtype=float,
        ),
        np.array([10.0, -3.5, 14.5, 15.5, 22.0], dtype=float),
    ),
    6: (
        np.array(
            [
                [14.0, 1.0, -1.0, 2.0, 0.0, 1.0],
                [1.0, 13.0, 2.0, -1.0, 1.0, 0.0],
                [-1.0, 2.0, 12.0, 1.0, -1.0, 1.0],
                [2.0, -1.0, 1.0, 11.0, 2.0, -1.0],
                [0.0, 1.0, -1.0, 2.0, 10.0, 1.0],
                [1.0, 0.0, 1.0, -1.0, 1.0, 9.0],
            ],
            dtype=float,
        ),
        np.array([17.0, 16.0, 14.0, 21.0, 18.0, 11.0], dtype=float),
    ),
}


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class SystemDiagnostics:
    """Diagnostic information about the original linear system."""

    rank_a: int
    rank_augmented: int
    classification: str
    condition_number: float
    diagnostic_tolerance: float


@dataclass(frozen=True)
class EliminationRecord:
    """Complete information for one elementary row-elimination operation."""

    step: int
    pivot_stage: int
    pivot_row: int
    target_row: int
    pivot_value: float
    eliminated_element: float
    multiplier: float
    row_operation: str
    matrix_before: np.ndarray
    vector_before: np.ndarray
    matrix_after: np.ndarray
    vector_after: np.ndarray
    updated_row_before: np.ndarray
    updated_row_after: np.ndarray
    status: str
    notes: str


@dataclass(frozen=True)
class BackSubstitutionRecord:
    """Complete information for one back-substitution calculation."""

    step: int
    row_number: int
    variable_name: str
    diagonal_coefficient: float
    right_hand_side: float
    known_term_sum: float
    numerator: float
    calculated_value: float
    formula: str
    status: str


@dataclass(frozen=True)
class ForwardEliminationOutcome:
    """Structured result returned by manual forward elimination."""

    success: bool
    message: str
    stopping_reason: str
    upper_matrix: np.ndarray
    transformed_vector: np.ndarray
    history: tuple[EliminationRecord, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class BackSubstitutionOutcome:
    """Structured result returned by manual back substitution."""

    success: bool
    message: str
    stopping_reason: str
    solution: np.ndarray | None
    history: tuple[BackSubstitutionRecord, ...]


@dataclass(frozen=True)
class GaussianEliminationResult:
    """Complete result object used by the GUI and Excel exporter."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str
    original_matrix: np.ndarray
    original_vector: np.ndarray
    upper_matrix: np.ndarray
    transformed_vector: np.ndarray
    solution: np.ndarray | None
    elimination_history: tuple[EliminationRecord, ...]
    back_substitution_history: tuple[BackSubstitutionRecord, ...]
    residual_vector: np.ndarray | None
    computed_right_hand_side: np.ndarray | None
    residual_norm_2: float | None
    residual_norm_inf: float | None
    maximum_absolute_residual: float | None
    relative_residual: float | None
    condition_number: float
    diagnostics: SystemDiagnostics
    pivot_tolerance: float
    warnings: tuple[str, ...]
    execution_date: str
    input_signature: str

    @property
    def elimination_steps(self) -> int:
        """Return the number of stored row-elimination operations."""

        return len(self.elimination_history)

    @property
    def back_substitution_steps(self) -> int:
        """Return the number of completed back-substitution operations."""

        return len(self.back_substitution_history)


    @property
    def determinant(self) -> float:
        """Return det(A), used only as a diagnostic summary."""

        if self.original_matrix.size == 0:
            return math.nan
        try:
            return float(np.linalg.det(self.original_matrix))
        except np.linalg.LinAlgError:
            return math.nan

    @property
    def growth_factor(self) -> float:
        """Return max|U| / max|A| for the completed elimination."""

        if self.original_matrix.size == 0 or self.upper_matrix.size == 0:
            return math.nan
        denominator = float(np.max(np.abs(self.original_matrix)))
        numerator = float(np.max(np.abs(self.upper_matrix)))
        if denominator <= np.finfo(float).tiny:
            return math.nan
        return numerator / denominator

    @property
    def minimum_pivot_ratio(self) -> float:
        """Return the smallest |pivot| relative to its active row scale."""

        if self.upper_matrix.size == 0:
            return math.nan
        ratios: list[float] = []
        size = self.upper_matrix.shape[0]
        for index in range(size):
            row_scale = float(np.max(np.abs(self.upper_matrix[index, index:])))
            if row_scale <= np.finfo(float).tiny:
                continue
            ratios.append(abs(float(self.upper_matrix[index, index])) / row_scale)
        return min(ratios) if ratios else math.nan


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a local image as Base64, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def format_number(value: float | None, decimals: int = DISPLAY_DECIMALS) -> str:
    """Format a finite floating-point value consistently for the interface."""

    if value is None:
        return "Not available"
    if not math.isfinite(float(value)):
        return "Undefined"
    return f"{float(value):.{decimals}f}"


def variable_names(size: int) -> list[str]:
    """Return user-facing variable labels x1, x2, ..., xn."""

    return [f"x{index}" for index in range(1, size + 1)]


def equation_names(size: int) -> list[str]:
    """Return user-facing row labels Equation 1, ..., Equation n."""

    return [f"Equation {index}" for index in range(1, size + 1)]


def create_input_signature(
    matrix_values: Sequence[Sequence[str]],
    vector_values: Sequence[str],
    pivot_tolerance: float,
) -> str:
    """Create a deterministic signature used to detect stale GUI results."""

    normalized_parts = [str(pivot_tolerance)]
    normalized_parts.extend(
        str(value).strip()
        for row in matrix_values
        for value in row
    )
    normalized_parts.extend(str(value).strip() for value in vector_values)
    payload = "|".join(normalized_parts).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def array_to_text(array: np.ndarray, decimals: int = 12) -> str:
    """Serialize an array into readable multiline text for Excel worksheets."""

    return np.array2string(
        np.asarray(array, dtype=float),
        precision=decimals,
        suppress_small=False,
        separator=", ",
        max_line_width=160,
    )


def make_augmented_matrix(matrix: np.ndarray, vector: np.ndarray) -> np.ndarray:
    """Create the augmented matrix [A | b] without modifying its inputs."""

    return np.column_stack(
        (
            np.asarray(matrix, dtype=float),
            np.asarray(vector, dtype=float).reshape(-1),
        )
    )


def matrix_dataframe(matrix: np.ndarray) -> pd.DataFrame:
    """Create a labeled DataFrame for a coefficient matrix."""

    size = int(matrix.shape[0])
    dataframe = pd.DataFrame(
        np.asarray(matrix, dtype=float),
        columns=variable_names(size),
        index=equation_names(size),
    )
    dataframe.index.name = "Equation"
    return dataframe


def vector_dataframe(vector: np.ndarray, column_name: str = "b") -> pd.DataFrame:
    """Create a labeled DataFrame for a vector."""

    values = np.asarray(vector, dtype=float).reshape(-1)
    dataframe = pd.DataFrame(
        {column_name: values},
        index=equation_names(len(values)),
    )
    dataframe.index.name = "Equation"
    return dataframe


def augmented_dataframe(matrix: np.ndarray, vector: np.ndarray) -> pd.DataFrame:
    """Create a labeled DataFrame for an augmented matrix [A | b]."""

    size = int(matrix.shape[0])
    columns = variable_names(size) + ["b"]
    dataframe = pd.DataFrame(
        make_augmented_matrix(matrix, vector),
        columns=columns,
        index=equation_names(size),
    )
    dataframe.index.name = "Equation"
    return dataframe


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns for display while preserving text columns."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


# =============================================================================
# Input conversion and validation
# =============================================================================
def parse_real_scalar(raw_value: Any, field_name: str) -> float:
    """Convert one input value to a finite real float with a clear error."""

    if raw_value is None:
        raise ValueError(f"{field_name} is empty.")

    text = str(raw_value).strip()
    if not text:
        raise ValueError(f"{field_name} is empty.")

    try:
        value = float(text)
    except (TypeError, ValueError) as error:
        raise ValueError(
            f"{field_name} must contain a valid real number. "
            "Scientific notation such as 1e-6 is supported."
        ) from error

    if not math.isfinite(value):
        raise ValueError(
            f"{field_name} must be finite. NaN and infinity are not allowed."
        )

    return value


def convert_and_validate_inputs(
    raw_matrix: Sequence[Sequence[Any]],
    raw_vector: Sequence[Any],
) -> tuple[np.ndarray, np.ndarray]:
    """Validate dimensions and convert Streamlit text fields to NumPy arrays."""

    if not isinstance(raw_matrix, Sequence) or not raw_matrix:
        raise ValueError("Matrix A is empty.")

    number_of_rows = len(raw_matrix)
    if number_of_rows not in SUPPORTED_SIZES:
        raise ValueError(
            "The system size must be 2×2, 3×3, 4×4, or 5×5."
        )

    if any(not isinstance(row, Sequence) for row in raw_matrix):
        raise ValueError("Matrix A must be a two-dimensional square matrix.")

    row_lengths = [len(row) for row in raw_matrix]
    if any(length != number_of_rows for length in row_lengths):
        raise ValueError(
            "Matrix A must be square. The number of columns must equal the "
            "number of rows."
        )

    if len(raw_vector) != number_of_rows:
        raise ValueError(
            "Vector b must contain exactly one value for each row of matrix A."
        )

    matrix = np.empty((number_of_rows, number_of_rows), dtype=float)
    vector = np.empty(number_of_rows, dtype=float)

    for row_index, row in enumerate(raw_matrix):
        for column_index, raw_value in enumerate(row):
            matrix[row_index, column_index] = parse_real_scalar(
                raw_value,
                (
                    f"A[{row_index + 1}, {column_index + 1}]"
                ),
            )

    for row_index, raw_value in enumerate(raw_vector):
        vector[row_index] = parse_real_scalar(
            raw_value,
            f"b[{row_index + 1}]",
        )

    if matrix.ndim != 2 or matrix.shape[0] != matrix.shape[1]:
        raise ValueError("Matrix A must be a two-dimensional square matrix.")

    if vector.ndim != 1 or vector.shape[0] != matrix.shape[0]:
        raise ValueError(
            "The dimensions of matrix A and vector b do not match."
        )

    if not np.all(np.isfinite(matrix)) or not np.all(np.isfinite(vector)):
        raise ValueError("All matrix and vector values must be finite numbers.")

    return matrix, vector


# =============================================================================
# Diagnostic analysis
# =============================================================================
def calculate_system_diagnostics(
    matrix: np.ndarray,
    vector: np.ndarray,
    pivot_tolerance: float = PIVOT_TOLERANCE,
) -> SystemDiagnostics:
    """Classify the system for feedback without using ranks to solve it.

    Matrix-rank and condition-number calculations are diagnostic only.  The
    displayed solution is always produced by the manual textbook algorithm.
    """

    coefficient_matrix = np.asarray(matrix, dtype=float)
    right_hand_side = np.asarray(vector, dtype=float).reshape(-1)
    augmented = make_augmented_matrix(coefficient_matrix, right_hand_side)

    singular_values = np.linalg.svd(
        augmented,
        compute_uv=False,
    )
    matrix_scale = (
        float(singular_values[0])
        if singular_values.size
        else np.finfo(float).tiny
    )
    diagnostic_tolerance = (
        pivot_tolerance
        * max(augmented.shape)
        * max(matrix_scale, np.finfo(float).tiny)
    )

    rank_a = int(
        np.linalg.matrix_rank(
            coefficient_matrix,
            tol=diagnostic_tolerance,
        )
    )
    rank_augmented = int(
        np.linalg.matrix_rank(
            augmented,
            tol=diagnostic_tolerance,
        )
    )

    if rank_a < rank_augmented:
        classification = "Inconsistent system"
    elif rank_a < coefficient_matrix.shape[0]:
        classification = "Singular system with non-unique solutions"
    else:
        classification = "Unique solution detected"

    try:
        condition_number = float(np.linalg.cond(coefficient_matrix))
    except np.linalg.LinAlgError:
        condition_number = math.inf

    return SystemDiagnostics(
        rank_a=rank_a,
        rank_augmented=rank_augmented,
        classification=classification,
        condition_number=condition_number,
        diagnostic_tolerance=diagnostic_tolerance,
    )



def calculate_pivot_diagnostics(
    matrix: np.ndarray,
    pivot_index: int,
    pivot_tolerance: float,
) -> tuple[float, float, float]:
    """Return pivot value, relative pivot ratio, and scale-aware threshold.

    The ratio is scale invariant: uniformly scaling the complete system does
    not make a valid pivot fail merely because its absolute magnitude is small.
    """

    active_row = np.asarray(
        matrix[pivot_index, pivot_index:],
        dtype=float,
    )
    active_submatrix = np.asarray(
        matrix[pivot_index:, pivot_index:],
        dtype=float,
    )

    pivot_value = float(matrix[pivot_index, pivot_index])
    row_scale = float(np.max(np.abs(active_row))) if active_row.size else 0.0
    active_scale = (
        float(np.max(np.abs(active_submatrix)))
        if active_submatrix.size
        else 0.0
    )
    reference_scale = max(
        row_scale,
        active_scale,
        np.finfo(float).tiny,
    )
    relative_pivot_ratio = abs(pivot_value) / reference_scale
    threshold = pivot_tolerance * reference_scale

    return pivot_value, relative_pivot_ratio, threshold

def build_pivot_failure_message(
    diagnostics: SystemDiagnostics,
    pivot_value: float,
    pivot_stage: int,
    phase: str,
) -> tuple[str, str]:
    """Build a mathematically accurate zero-pivot failure explanation."""

    location = f"row {pivot_stage}, column {pivot_stage}"
    formatted_pivot = f"{pivot_value:.12e}"

    if diagnostics.classification == "Inconsistent system":
        return (
            "The system is inconsistent: the coefficient and augmented "
            "matrices have different ranks, so no solution satisfies all "
            "equations.",
            f"{phase} stopped after detecting an inconsistent system.",
        )

    if diagnostics.classification.startswith("Singular system"):
        return (
            "The coefficient matrix is singular or rank-deficient. The "
            "system does not have a unique solution, so Naïve Gaussian "
            "Elimination cannot complete a unique back-substitution result.",
            f"{phase} stopped because the system is singular.",
        )

    return (
        "Naïve Gaussian Elimination cannot continue because a zero or "
        f"near-zero pivot was encountered at {location}. The pivot value is "
        f"{formatted_pivot}. This solver intentionally performs no row "
        "swapping. The system may still have a unique solution, but a "
        "pivoting method is required for this row order.",
        f"{phase} stopped because pivoting is required.",
    )


# =============================================================================
# Standard Naïve Gaussian Elimination algorithm
# =============================================================================
def perform_forward_elimination(
    matrix: np.ndarray,
    vector: np.ndarray,
    diagnostics: SystemDiagnostics,
    pivot_tolerance: float = PIVOT_TOLERANCE,
) -> ForwardEliminationOutcome:
    """Perform textbook forward elimination manually and without pivoting."""

    upper_matrix = np.asarray(matrix, dtype=float).copy()
    transformed_vector = np.asarray(vector, dtype=float).reshape(-1).copy()
    size = upper_matrix.shape[0]

    history: list[EliminationRecord] = []
    warnings: list[str] = []
    operation_step = 0

    for pivot_index in range(size - 1):
        (
            pivot_value,
            pivot_ratio,
            pivot_threshold,
        ) = calculate_pivot_diagnostics(
            upper_matrix,
            pivot_index,
            pivot_tolerance,
        )

        if not math.isfinite(pivot_value):
            return ForwardEliminationOutcome(
                success=False,
                message=(
                    "Forward elimination produced a non-finite pivot value. "
                    "Check the entered matrix for invalid or excessively "
                    "large values."
                ),
                stopping_reason=(
                    "Forward elimination stopped because the pivot was not "
                    "finite."
                ),
                upper_matrix=upper_matrix.copy(),
                transformed_vector=transformed_vector.copy(),
                history=tuple(history),
                warnings=tuple(warnings),
            )

        if abs(pivot_value) <= pivot_threshold:
            message, stopping_reason = build_pivot_failure_message(
                diagnostics=diagnostics,
                pivot_value=pivot_value,
                pivot_stage=pivot_index + 1,
                phase="Forward elimination",
            )
            return ForwardEliminationOutcome(
                success=False,
                message=message,
                stopping_reason=stopping_reason,
                upper_matrix=upper_matrix.copy(),
                transformed_vector=transformed_vector.copy(),
                history=tuple(history),
                warnings=tuple(warnings),
            )

        if pivot_ratio <= NEAR_ZERO_WARNING_LIMIT:
            warning = (
                f"Pivot {pivot_index + 1} is small relative to the active "
                f"matrix scale (ratio = {pivot_ratio:.6e}). Naïve elimination "
                "may amplify rounding errors because no pivoting is used."
            )
            if warning not in warnings:
                warnings.append(warning)

        for target_index in range(pivot_index + 1, size):
            operation_step += 1

            matrix_before = upper_matrix.copy()
            vector_before = transformed_vector.copy()
            updated_row_before = make_augmented_matrix(
                upper_matrix[target_index : target_index + 1, :],
                transformed_vector[target_index : target_index + 1],
            ).reshape(-1)

            eliminated_element = float(
                upper_matrix[target_index, pivot_index]
            )
            multiplier = eliminated_element / pivot_value

            if not math.isfinite(multiplier):
                return ForwardEliminationOutcome(
                    success=False,
                    message=(
                        "A non-finite elimination multiplier was produced. "
                        "The selected pivot is numerically unusable."
                    ),
                    stopping_reason=(
                        "Forward elimination stopped because the multiplier "
                        "was not finite."
                    ),
                    upper_matrix=upper_matrix.copy(),
                    transformed_vector=transformed_vector.copy(),
                    history=tuple(history),
                    warnings=tuple(warnings),
                )

            upper_matrix[target_index, pivot_index:] = (
                upper_matrix[target_index, pivot_index:]
                - multiplier * upper_matrix[pivot_index, pivot_index:]
            )
            transformed_vector[target_index] = (
                transformed_vector[target_index]
                - multiplier * transformed_vector[pivot_index]
            )

            # The targeted entry is mathematically zero after this operation.
            # Assigning it exactly to zero improves the educational display and
            # prevents a harmless floating-point remnant from appearing.
            upper_matrix[target_index, pivot_index] = 0.0

            if not np.all(np.isfinite(upper_matrix)) or not np.all(
                np.isfinite(transformed_vector)
            ):
                return ForwardEliminationOutcome(
                    success=False,
                    message=(
                        "Forward elimination produced NaN or infinity. The "
                        "calculation cannot continue safely."
                    ),
                    stopping_reason=(
                        "Forward elimination stopped because non-finite "
                        "arithmetic was produced."
                    ),
                    upper_matrix=upper_matrix.copy(),
                    transformed_vector=transformed_vector.copy(),
                    history=tuple(history),
                    warnings=tuple(warnings),
                )

            matrix_after = upper_matrix.copy()
            vector_after = transformed_vector.copy()
            updated_row_after = make_augmented_matrix(
                upper_matrix[target_index : target_index + 1, :],
                transformed_vector[target_index : target_index + 1],
            ).reshape(-1)

            pivot_row_number = pivot_index + 1
            target_row_number = target_index + 1
            row_operation = (
                f"R{target_row_number} = R{target_row_number} - "
                f"({multiplier:.12g}) × R{pivot_row_number}"
            )

            notes = "Elimination completed normally."
            if pivot_ratio <= NEAR_ZERO_WARNING_LIMIT:
                notes = (
                    "Operation completed, but the pivot is small relative "
                    "to the active scale and may increase rounding sensitivity."
                )

            history.append(
                EliminationRecord(
                    step=operation_step,
                    pivot_stage=pivot_index + 1,
                    pivot_row=pivot_row_number,
                    target_row=target_row_number,
                    pivot_value=pivot_value,
                    eliminated_element=eliminated_element,
                    multiplier=float(multiplier),
                    row_operation=row_operation,
                    matrix_before=matrix_before,
                    vector_before=vector_before,
                    matrix_after=matrix_after,
                    vector_after=vector_after,
                    updated_row_before=updated_row_before.copy(),
                    updated_row_after=updated_row_after.copy(),
                    status="Completed",
                    notes=notes,
                )
            )

    return ForwardEliminationOutcome(
        success=True,
        message="Forward elimination completed successfully.",
        stopping_reason="Forward elimination produced an upper-triangular system.",
        upper_matrix=upper_matrix.copy(),
        transformed_vector=transformed_vector.copy(),
        history=tuple(history),
        warnings=tuple(warnings),
    )


def perform_back_substitution(
    upper_matrix: np.ndarray,
    transformed_vector: np.ndarray,
    diagnostics: SystemDiagnostics,
    pivot_tolerance: float = PIVOT_TOLERANCE,
) -> BackSubstitutionOutcome:
    """Perform textbook back substitution manually from the last row upward."""

    coefficient_matrix = np.asarray(upper_matrix, dtype=float)
    right_hand_side = np.asarray(transformed_vector, dtype=float).reshape(-1)
    size = coefficient_matrix.shape[0]
    solution = np.zeros(size, dtype=float)
    history: list[BackSubstitutionRecord] = []

    step_number = 0
    for row_index in range(size - 1, -1, -1):
        step_number += 1
        (
            diagonal_coefficient,
            _pivot_ratio,
            diagonal_threshold,
        ) = calculate_pivot_diagnostics(
            coefficient_matrix,
            row_index,
            pivot_tolerance,
        )

        if (
            not math.isfinite(diagonal_coefficient)
            or abs(diagonal_coefficient) <= diagonal_threshold
        ):
            message, stopping_reason = build_pivot_failure_message(
                diagnostics=diagnostics,
                pivot_value=diagonal_coefficient,
                pivot_stage=row_index + 1,
                phase="Back substitution",
            )
            return BackSubstitutionOutcome(
                success=False,
                message=message,
                stopping_reason=stopping_reason,
                solution=None,
                history=tuple(history),
            )

        known_term_sum = float(
            np.dot(
                coefficient_matrix[row_index, row_index + 1 :],
                solution[row_index + 1 :],
            )
        )
        numerator = float(right_hand_side[row_index] - known_term_sum)
        calculated_value = numerator / diagonal_coefficient

        if not math.isfinite(calculated_value):
            return BackSubstitutionOutcome(
                success=False,
                message=(
                    "Back substitution produced a non-finite solution value. "
                    "The triangular system is numerically unusable."
                ),
                stopping_reason=(
                    "Back substitution stopped because a non-finite value "
                    "was produced."
                ),
                solution=None,
                history=tuple(history),
            )

        solution[row_index] = calculated_value
        variable_name = f"x{row_index + 1}"
        formula = (
            f"{variable_name} = "
            f"({right_hand_side[row_index]:.12g} - "
            f"{known_term_sum:.12g}) / "
            f"{diagonal_coefficient:.12g}"
        )

        history.append(
            BackSubstitutionRecord(
                step=step_number,
                row_number=row_index + 1,
                variable_name=variable_name,
                diagonal_coefficient=diagonal_coefficient,
                right_hand_side=float(right_hand_side[row_index]),
                known_term_sum=known_term_sum,
                numerator=numerator,
                calculated_value=float(calculated_value),
                formula=formula,
                status="Completed",
            )
        )

    return BackSubstitutionOutcome(
        success=True,
        message="Back substitution completed successfully.",
        stopping_reason="All unknowns were calculated from the triangular system.",
        solution=solution.copy(),
        history=tuple(history),
    )


# =============================================================================
# Residual and solution analysis
# =============================================================================
def calculate_residual_analysis(
    original_matrix: np.ndarray,
    original_vector: np.ndarray,
    solution: np.ndarray,
) -> dict[str, Any]:
    """Calculate residual metrics using the original, unmodified system."""

    matrix = np.asarray(original_matrix, dtype=float)
    vector = np.asarray(original_vector, dtype=float).reshape(-1)
    solution_vector = np.asarray(solution, dtype=float).reshape(-1)

    computed_right_hand_side = matrix @ solution_vector
    residual_vector = computed_right_hand_side - vector

    residual_norm_2 = float(np.linalg.norm(residual_vector, ord=2))
    residual_norm_inf = float(np.linalg.norm(residual_vector, ord=np.inf))
    maximum_absolute_residual = float(np.max(np.abs(residual_vector)))

    denominator = float(
        np.linalg.norm(matrix, ord=2)
        * np.linalg.norm(solution_vector, ord=2)
        + np.linalg.norm(vector, ord=2)
    )
    if denominator > 0.0 and math.isfinite(denominator):
        relative_residual = residual_norm_2 / denominator
    elif residual_norm_2 == 0.0:
        relative_residual = 0.0
    else:
        relative_residual = math.inf

    return {
        "computed_right_hand_side": computed_right_hand_side,
        "residual_vector": residual_vector,
        "residual_norm_2": residual_norm_2,
        "residual_norm_inf": residual_norm_inf,
        "maximum_absolute_residual": maximum_absolute_residual,
        "relative_residual": float(relative_residual),
    }


# =============================================================================
# Main solver controller
# =============================================================================
def solve_naive_gaussian_elimination(
    raw_matrix: Sequence[Sequence[Any]],
    raw_vector: Sequence[Any],
    input_signature: str,
    pivot_tolerance: float = PIVOT_TOLERANCE,
) -> GaussianEliminationResult:
    """Validate, solve, analyze, and return a complete structured result."""

    execution_date = datetime.now(ZoneInfo(REPORT_TIME_ZONE)).strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    if not math.isfinite(pivot_tolerance) or pivot_tolerance <= 0.0:
        empty_matrix = np.empty((0, 0), dtype=float)
        empty_vector = np.empty(0, dtype=float)
        empty_diagnostics = SystemDiagnostics(
            rank_a=0,
            rank_augmented=0,
            classification="Not evaluated",
            condition_number=math.nan,
            diagnostic_tolerance=math.nan,
        )
        return GaussianEliminationResult(
            status="Input Error",
            success=False,
            method=METHOD_NAME,
            message="The pivot tolerance must be a positive finite number.",
            stopping_reason="Input validation failed.",
            original_matrix=empty_matrix,
            original_vector=empty_vector,
            upper_matrix=empty_matrix,
            transformed_vector=empty_vector,
            solution=None,
            elimination_history=(),
            back_substitution_history=(),
            residual_vector=None,
            computed_right_hand_side=None,
            residual_norm_2=None,
            residual_norm_inf=None,
            maximum_absolute_residual=None,
            relative_residual=None,
            condition_number=math.nan,
            diagnostics=empty_diagnostics,
            pivot_tolerance=pivot_tolerance,
            warnings=(),
            execution_date=execution_date,
            input_signature=input_signature,
        )

    try:
        original_matrix, original_vector = convert_and_validate_inputs(
            raw_matrix,
            raw_vector,
        )
    except ValueError as error:
        empty_matrix = np.empty((0, 0), dtype=float)
        empty_vector = np.empty(0, dtype=float)
        empty_diagnostics = SystemDiagnostics(
            rank_a=0,
            rank_augmented=0,
            classification="Not evaluated",
            condition_number=math.nan,
            diagnostic_tolerance=math.nan,
        )
        return GaussianEliminationResult(
            status="Input Error",
            success=False,
            method=METHOD_NAME,
            message=str(error),
            stopping_reason="Input validation failed.",
            original_matrix=empty_matrix,
            original_vector=empty_vector,
            upper_matrix=empty_matrix,
            transformed_vector=empty_vector,
            solution=None,
            elimination_history=(),
            back_substitution_history=(),
            residual_vector=None,
            computed_right_hand_side=None,
            residual_norm_2=None,
            residual_norm_inf=None,
            maximum_absolute_residual=None,
            relative_residual=None,
            condition_number=math.nan,
            diagnostics=empty_diagnostics,
            pivot_tolerance=pivot_tolerance,
            warnings=(),
            execution_date=execution_date,
            input_signature=input_signature,
        )

    diagnostics = calculate_system_diagnostics(
        original_matrix,
        original_vector,
        pivot_tolerance,
    )

    warnings: list[str] = []
    if (
        not math.isfinite(diagnostics.condition_number)
        or diagnostics.condition_number >= CONDITION_NUMBER_WARNING
    ):
        warnings.append(
            "The original coefficient matrix is singular or severely "
            "ill-conditioned. The solution may be highly sensitive to small "
            "changes or rounding in the input data."
        )

    forward_outcome = perform_forward_elimination(
        matrix=original_matrix,
        vector=original_vector,
        diagnostics=diagnostics,
        pivot_tolerance=pivot_tolerance,
    )
    warnings.extend(
        warning
        for warning in forward_outcome.warnings
        if warning not in warnings
    )

    if not forward_outcome.success:
        return GaussianEliminationResult(
            status="Failed",
            success=False,
            method=METHOD_NAME,
            message=forward_outcome.message,
            stopping_reason=forward_outcome.stopping_reason,
            original_matrix=original_matrix.copy(),
            original_vector=original_vector.copy(),
            upper_matrix=forward_outcome.upper_matrix.copy(),
            transformed_vector=forward_outcome.transformed_vector.copy(),
            solution=None,
            elimination_history=forward_outcome.history,
            back_substitution_history=(),
            residual_vector=None,
            computed_right_hand_side=None,
            residual_norm_2=None,
            residual_norm_inf=None,
            maximum_absolute_residual=None,
            relative_residual=None,
            condition_number=diagnostics.condition_number,
            diagnostics=diagnostics,
            pivot_tolerance=pivot_tolerance,
            warnings=tuple(warnings),
            execution_date=execution_date,
            input_signature=input_signature,
        )

    back_outcome = perform_back_substitution(
        upper_matrix=forward_outcome.upper_matrix,
        transformed_vector=forward_outcome.transformed_vector,
        diagnostics=diagnostics,
        pivot_tolerance=pivot_tolerance,
    )

    if not back_outcome.success or back_outcome.solution is None:
        return GaussianEliminationResult(
            status="Failed",
            success=False,
            method=METHOD_NAME,
            message=back_outcome.message,
            stopping_reason=back_outcome.stopping_reason,
            original_matrix=original_matrix.copy(),
            original_vector=original_vector.copy(),
            upper_matrix=forward_outcome.upper_matrix.copy(),
            transformed_vector=forward_outcome.transformed_vector.copy(),
            solution=None,
            elimination_history=forward_outcome.history,
            back_substitution_history=back_outcome.history,
            residual_vector=None,
            computed_right_hand_side=None,
            residual_norm_2=None,
            residual_norm_inf=None,
            maximum_absolute_residual=None,
            relative_residual=None,
            condition_number=diagnostics.condition_number,
            diagnostics=diagnostics,
            pivot_tolerance=pivot_tolerance,
            warnings=tuple(warnings),
            execution_date=execution_date,
            input_signature=input_signature,
        )

    residual_analysis = calculate_residual_analysis(
        original_matrix=original_matrix,
        original_vector=original_vector,
        solution=back_outcome.solution,
    )

    temporary_result = GaussianEliminationResult(
        status="Completed",
        success=True,
        method=METHOD_NAME,
        message="Execution completed successfully.",
        stopping_reason=(
            "Forward elimination and back substitution completed successfully."
        ),
        original_matrix=original_matrix.copy(),
        original_vector=original_vector.copy(),
        upper_matrix=forward_outcome.upper_matrix.copy(),
        transformed_vector=forward_outcome.transformed_vector.copy(),
        solution=back_outcome.solution.copy(),
        elimination_history=forward_outcome.history,
        back_substitution_history=back_outcome.history,
        residual_vector=residual_analysis["residual_vector"].copy(),
        computed_right_hand_side=residual_analysis[
            "computed_right_hand_side"
        ].copy(),
        residual_norm_2=residual_analysis["residual_norm_2"],
        residual_norm_inf=residual_analysis["residual_norm_inf"],
        maximum_absolute_residual=residual_analysis[
            "maximum_absolute_residual"
        ],
        relative_residual=residual_analysis["relative_residual"],
        condition_number=diagnostics.condition_number,
        diagnostics=diagnostics,
        pivot_tolerance=pivot_tolerance,
        warnings=(),
        execution_date=execution_date,
        input_signature=input_signature,
    )

    if temporary_result.relative_residual is not None and (
        temporary_result.relative_residual > RELATIVE_RESIDUAL_WARNING
    ):
        warnings.append(
            "The scale-aware relative residual is larger than expected. "
            "Inspect the conditioning and elimination history."
        )

    if (
        math.isfinite(temporary_result.growth_factor)
        and temporary_result.growth_factor > GROWTH_FACTOR_WARNING
    ):
        warnings.append(
            "Large element growth occurred during elimination. Because this "
            "method performs no pivoting, round-off error may be amplified."
        )

    if (
        math.isfinite(temporary_result.minimum_pivot_ratio)
        and temporary_result.minimum_pivot_ratio <= NEAR_ZERO_WARNING_LIMIT
    ):
        warnings.append(
            "At least one pivot was small relative to its active row scale. "
            "A pivoted Gaussian method is recommended for greater reliability."
        )

    return GaussianEliminationResult(
        status="Completed",
        success=True,
        method=METHOD_NAME,
        message="Execution completed successfully.",
        stopping_reason=(
            "Forward elimination and back substitution completed successfully."
        ),
        original_matrix=original_matrix.copy(),
        original_vector=original_vector.copy(),
        upper_matrix=forward_outcome.upper_matrix.copy(),
        transformed_vector=forward_outcome.transformed_vector.copy(),
        solution=back_outcome.solution.copy(),
        elimination_history=forward_outcome.history,
        back_substitution_history=back_outcome.history,
        residual_vector=residual_analysis["residual_vector"].copy(),
        computed_right_hand_side=residual_analysis[
            "computed_right_hand_side"
        ].copy(),
        residual_norm_2=residual_analysis["residual_norm_2"],
        residual_norm_inf=residual_analysis["residual_norm_inf"],
        maximum_absolute_residual=residual_analysis[
            "maximum_absolute_residual"
        ],
        relative_residual=residual_analysis["relative_residual"],
        condition_number=diagnostics.condition_number,
        diagnostics=diagnostics,
        pivot_tolerance=pivot_tolerance,
        warnings=tuple(warnings),
        execution_date=execution_date,
        input_signature=input_signature,
    )


# =============================================================================
# DataFrame builders
# =============================================================================
def build_elimination_summary_dataframe(
    result: GaussianEliminationResult,
) -> pd.DataFrame:
    """Create a concise table containing every elimination operation."""

    rows = [
        {
            "Step": record.step,
            "Pivot Stage": record.pivot_stage,
            "Pivot Row": f"R{record.pivot_row}",
            "Target Row": f"R{record.target_row}",
            "Pivot": record.pivot_value,
            "Eliminated Element": record.eliminated_element,
            "Multiplier": record.multiplier,
            "Operation": record.row_operation,
            "Status": record.status,
            "Notes": record.notes,
        }
        for record in result.elimination_history
    ]
    return pd.DataFrame(rows)


def build_elimination_steps_dataframe(
    result: GaussianEliminationResult,
) -> pd.DataFrame:
    """Create the detailed Excel table containing matrix snapshots."""

    rows = [
        {
            "Step": record.step,
            "Pivot Stage": record.pivot_stage,
            "Pivot Row": record.pivot_row,
            "Target Row": record.target_row,
            "Pivot": record.pivot_value,
            "Eliminated Element": record.eliminated_element,
            "Multiplier": record.multiplier,
            "Row Operation": record.row_operation,
            "Updated Row Before": array_to_text(record.updated_row_before),
            "Updated Row After": array_to_text(record.updated_row_after),
            "Matrix A Before": array_to_text(record.matrix_before),
            "Vector b Before": array_to_text(record.vector_before),
            "Matrix A After": array_to_text(record.matrix_after),
            "Vector b After": array_to_text(record.vector_after),
            "Status": record.status,
            "Notes": record.notes,
        }
        for record in result.elimination_history
    ]
    return pd.DataFrame(rows)


def build_back_substitution_dataframe(
    result: GaussianEliminationResult,
) -> pd.DataFrame:
    """Create the complete back-substitution history table."""

    rows = [
        {
            "Step": record.step,
            "Row": record.row_number,
            "Variable": record.variable_name,
            "Diagonal Coefficient": record.diagonal_coefficient,
            "Transformed RHS": record.right_hand_side,
            "Known-Term Sum": record.known_term_sum,
            "Numerator": record.numerator,
            "Calculated Value": record.calculated_value,
            "Formula": record.formula,
            "Status": record.status,
        }
        for record in result.back_substitution_history
    ]
    return pd.DataFrame(rows)


def build_solution_dataframe(result: GaussianEliminationResult) -> pd.DataFrame:
    """Create a variable-by-variable final solution table."""

    if result.solution is None:
        return pd.DataFrame(columns=["Variable", "Calculated Value"])

    return pd.DataFrame(
        {
            "Variable": variable_names(len(result.solution)),
            "Calculated Value": result.solution,
        }
    )


def build_residual_dataframe(result: GaussianEliminationResult) -> pd.DataFrame:
    """Create a complete equation-level residual table."""

    if (
        result.solution is None
        or result.residual_vector is None
        or result.computed_right_hand_side is None
    ):
        return pd.DataFrame(
            columns=[
                "Equation",
                "Computed Ax",
                "Original b",
                "Residual Ax - b",
                "Absolute Residual",
            ]
        )

    return pd.DataFrame(
        {
            "Equation": equation_names(len(result.original_vector)),
            "Computed Ax": result.computed_right_hand_side,
            "Original b": result.original_vector,
            "Residual Ax - b": result.residual_vector,
            "Absolute Residual": np.abs(result.residual_vector),
        }
    )



def build_summary_dataframe(
    result: GaussianEliminationResult,
) -> pd.DataFrame:
    """Create a comprehensive one-page report summary."""

    solution_text = (
        ", ".join(
            f"x{index + 1} = {value:.15g}"
            for index, value in enumerate(result.solution)
        )
        if result.solution is not None
        else "Not available"
    )

    return pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "System Classification",
                "Matrix Size",
                "Final Matrix Form",
                "Elimination Operations",
                "Back-Substitution Steps",
                "Rank(A)",
                "Rank([A|b])",
                "Determinant",
                "Condition Number",
                "Element Growth Factor",
                "Minimum Relative Pivot Ratio",
                "Residual 2-Norm",
                "Residual Infinity Norm",
                "Maximum Absolute Residual",
                "Relative Residual",
                "Solution Vector",
                "Pivot Tolerance Factor",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                result.diagnostics.classification,
                (
                    f"{result.original_matrix.shape[0]} × "
                    f"{result.original_matrix.shape[1]}"
                ),
                "Upper-Triangular Form",
                result.elimination_steps,
                result.back_substitution_steps,
                result.diagnostics.rank_a,
                result.diagnostics.rank_augmented,
                result.determinant,
                result.condition_number,
                result.growth_factor,
                result.minimum_pivot_ratio,
                result.residual_norm_2,
                result.residual_norm_inf,
                result.maximum_absolute_residual,
                result.relative_residual,
                solution_text,
                result.pivot_tolerance,
                (
                    " | ".join(result.warnings)
                    if result.warnings
                    else "None"
                ),
                result.stopping_reason,
                result.execution_date,
            ],
        }
    )



# =============================================================================
# Scientific plots
# =============================================================================
def create_matrix_figure(result: GaussianEliminationResult) -> Figure:
    """Create a heatmap of the original coefficient matrix."""

    figure, axis = plt.subplots(figsize=(7.2, 5.0))
    image = axis.imshow(result.original_matrix, aspect="auto")
    figure.colorbar(image, ax=axis, label="Coefficient value")
    axis.set_title("Original Coefficient Matrix A")
    axis.set_xlabel("Variable")
    axis.set_ylabel("Equation")
    axis.set_xticks(
        range(result.original_matrix.shape[1]),
        variable_names(result.original_matrix.shape[1]),
    )
    axis.set_yticks(
        range(result.original_matrix.shape[0]),
        equation_names(result.original_matrix.shape[0]),
    )
    for row in range(result.original_matrix.shape[0]):
        for column in range(result.original_matrix.shape[1]):
            axis.text(
                column,
                row,
                f"{result.original_matrix[row, column]:.3g}",
                ha="center",
                va="center",
            )
    figure.tight_layout()
    return figure


def create_solution_figure(result: GaussianEliminationResult) -> Figure:
    """Create a bar chart of the solution vector."""

    if result.solution is None:
        raise ValueError("A solution is required for plotting.")

    positions = np.arange(len(result.solution))
    figure, axis = plt.subplots(figsize=(7.2, 5.0))
    axis.bar(positions, result.solution)
    axis.axhline(0.0, linewidth=1.0)
    axis.set_xticks(
        positions,
        variable_names(len(result.solution)),
    )
    axis.set_title("Naive Gaussian Elimination Solution")
    axis.set_xlabel("Variable")
    axis.set_ylabel("Value")
    axis.grid(True, axis="y", alpha=0.3)
    figure.tight_layout()
    return figure


def create_residual_figure(result: GaussianEliminationResult) -> Figure:
    """Create a bar chart of absolute residual components."""

    if result.residual_vector is None:
        raise ValueError("Residual data are unavailable.")

    residual = np.abs(result.residual_vector)
    positions = np.arange(len(residual))
    figure, axis = plt.subplots(figsize=(7.2, 5.0))
    axis.bar(positions, residual)
    axis.set_xticks(
        positions,
        equation_names(len(residual)),
    )
    axis.set_title("Absolute Residual Components")
    axis.set_xlabel("Equation")
    axis.set_ylabel("|Ax − b|")
    axis.grid(True, axis="y", alpha=0.3)
    figure.tight_layout()
    return figure


def figure_to_png_bytes(figure: Figure) -> bytes:
    """Serialize one Matplotlib figure as PNG bytes."""

    buffer = BytesIO()
    figure.savefig(
        buffer,
        format="png",
        dpi=180,
        bbox_inches="tight",
    )
    buffer.seek(0)
    return buffer.getvalue()


def add_excel_image(
    worksheet: Any,
    image_bytes: bytes,
    anchor: str,
    width: int = 760,
    height: int = 500,
) -> None:
    """Insert PNG bytes into an Excel worksheet."""

    image_stream = BytesIO(image_bytes)
    image = ExcelImage(image_stream)
    image.width = width
    image.height = height
    worksheet.add_image(image, anchor)


def style_excel_worksheet(worksheet: Any) -> None:
    """Apply consistent professional formatting to one worksheet."""

    header_fill = PatternFill(fill_type="solid", fgColor="DCEEF2")
    header_font = Font(bold=True)

    if worksheet.max_row >= 1:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    worksheet.freeze_panes = "A2"
    if worksheet.max_row > 1 and worksheet.max_column > 0:
        worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_cells in enumerate(
        worksheet.iter_cols(),
        start=1,
    ):
        maximum_length = 0
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            longest_line = max(
                (len(line) for line in cell_value.splitlines()),
                default=0,
            )
            maximum_length = max(maximum_length, longest_line)

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            if isinstance(cell.value, float):
                cell.number_format = "0.000000000000E+00"

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = min(max(maximum_length + 2, 12), 55)



def create_excel_report(result: GaussianEliminationResult) -> bytes:
    """Generate a complete formatted XLSX report entirely in memory."""

    if not result.success or result.solution is None:
        raise ValueError("Only a successful solution can be exported.")

    summary_df = build_summary_dataframe(result)
    original_matrix_df = matrix_dataframe(result.original_matrix).reset_index()
    original_augmented_df = augmented_dataframe(
        result.original_matrix,
        result.original_vector,
    ).reset_index()
    elimination_summary_df = build_elimination_summary_dataframe(result)
    elimination_steps_df = build_elimination_steps_dataframe(result)
    upper_matrix_df = matrix_dataframe(result.upper_matrix).reset_index()
    final_augmented_df = augmented_dataframe(
        result.upper_matrix,
        result.transformed_vector,
    ).reset_index()
    back_substitution_df = build_back_substitution_dataframe(result)
    solution_df = build_solution_dataframe(result)
    residual_df = build_residual_dataframe(result)

    chart_data_df = pd.DataFrame(
        {
            "Variable": variable_names(len(result.solution)),
            "Solution": result.solution,
            "Equation": equation_names(len(result.residual_vector)),
            "Absolute Residual": np.abs(result.residual_vector),
        }
    )

    matrix_figure = create_matrix_figure(result)
    solution_figure = create_solution_figure(result)
    residual_figure = create_residual_figure(result)

    matrix_png = figure_to_png_bytes(matrix_figure)
    solution_png = figure_to_png_bytes(solution_figure)
    residual_png = figure_to_png_bytes(residual_figure)

    plt.close(matrix_figure)
    plt.close(solution_figure)
    plt.close(residual_figure)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        original_matrix_df.to_excel(
            writer,
            sheet_name="Original Matrix",
            index=False,
        )
        original_augmented_df.to_excel(
            writer,
            sheet_name="Original Augmented",
            index=False,
        )
        elimination_summary_df.to_excel(
            writer,
            sheet_name="Elimination Summary",
            index=False,
        )
        elimination_steps_df.to_excel(
            writer,
            sheet_name="Elimination Steps",
            index=False,
        )
        upper_matrix_df.to_excel(
            writer,
            sheet_name="Upper Triangular",
            index=False,
        )
        final_augmented_df.to_excel(
            writer,
            sheet_name="Final Augmented",
            index=False,
        )
        back_substitution_df.to_excel(
            writer,
            sheet_name="Back Substitution",
            index=False,
        )
        solution_df.to_excel(writer, sheet_name="Solution", index=False)
        residual_df.to_excel(
            writer,
            sheet_name="Residual Analysis",
            index=False,
        )
        chart_data_df.to_excel(
            writer,
            sheet_name="Chart Data",
            index=False,
        )

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Naïve Gaussian Elimination Report Plots"
        plots_sheet["A1"].font = Font(bold=True, size=14)
        add_excel_image(plots_sheet, matrix_png, "A3")
        add_excel_image(plots_sheet, solution_png, "A31")
        add_excel_image(plots_sheet, residual_png, "A59")

        data_sheet = workbook["Chart Data"]
        summary_sheet = workbook["Summary"]
        size = len(result.solution)

        solution_chart = BarChart()
        solution_chart.title = "Final Solution Values"
        solution_chart.x_axis.title = "Variable"
        solution_chart.y_axis.title = "Value"
        solution_chart.height = 8
        solution_chart.width = 15
        solution_chart.add_data(
            Reference(
                data_sheet,
                min_col=2,
                min_row=1,
                max_row=size + 1,
            ),
            titles_from_data=True,
        )
        solution_chart.set_categories(
            Reference(
                data_sheet,
                min_col=1,
                min_row=2,
                max_row=size + 1,
            )
        )
        summary_sheet.add_chart(solution_chart, "D2")

        residual_chart = BarChart()
        residual_chart.title = "Absolute Residual by Equation"
        residual_chart.x_axis.title = "Equation"
        residual_chart.y_axis.title = "Absolute Residual"
        residual_chart.height = 8
        residual_chart.width = 15
        residual_chart.add_data(
            Reference(
                data_sheet,
                min_col=4,
                min_row=1,
                max_row=size + 1,
            ),
            titles_from_data=True,
        )
        residual_chart.set_categories(
            Reference(
                data_sheet,
                min_col=3,
                min_row=2,
                max_row=size + 1,
            )
        )
        summary_sheet.add_chart(residual_chart, "D20")

        for worksheet in workbook.worksheets:
            style_excel_worksheet(worksheet)

        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


def initialize_input_state(size: int) -> None:
    """Populate a deterministic editable example only when keys are new."""

    default_matrix, default_vector = DEFAULT_SYSTEMS[size]
    for row_index in range(size):
        for column_index in range(size):
            key = f"nge_a_{size}_{row_index}_{column_index}"
            if key not in st.session_state:
                st.session_state[key] = f"{default_matrix[row_index, column_index]:g}"

        vector_key = f"nge_b_{size}_{row_index}"
        if vector_key not in st.session_state:
            st.session_state[vector_key] = f"{default_vector[row_index]:g}"


def collect_raw_inputs(size: int) -> tuple[list[list[str]], list[str]]:
    """Read all currently displayed text-input values from session state."""

    matrix_values = [
        [
            str(st.session_state.get(f"nge_a_{size}_{row}_{column}", ""))
            for column in range(size)
        ]
        for row in range(size)
    ]
    vector_values = [
        str(st.session_state.get(f"nge_b_{size}_{row}", ""))
        for row in range(size)
    ]
    return matrix_values, vector_values


def render_matrix_input_grid(size: int) -> None:
    """Render dynamic text fields for matrix A and vector b."""

    initialize_input_state(size)

    header_columns = st.columns([1.05] + [1.0] * size + [1.0])
    header_columns[0].markdown("**Row**")
    for column_index in range(size):
        header_columns[column_index + 1].markdown(
            f"**x{column_index + 1}**"
        )
    header_columns[-1].markdown("**b**")

    for row_index in range(size):
        row_columns = st.columns([1.05] + [1.0] * size + [1.0])
        row_columns[0].markdown(f"**R{row_index + 1}**")

        for column_index in range(size):
            with row_columns[column_index + 1]:
                st.text_input(
                    f"A[{row_index + 1}, {column_index + 1}]",
                    key=f"nge_a_{size}_{row_index}_{column_index}",
                    label_visibility="collapsed",
                )

        with row_columns[-1]:
            st.text_input(
                f"b[{row_index + 1}]",
                key=f"nge_b_{size}_{row_index}",
                label_visibility="collapsed",
            )


def render_final_result(result: GaussianEliminationResult) -> None:
    """Render either a complete success summary or a safe error card."""

    if not result.success:
        st.error(result.message)
        st.warning(result.stopping_reason)

        if result.original_matrix.size > 0:
            st.markdown(
                f"**System classification:** "
                f"{result.diagnostics.classification}"
            )
            st.markdown(
                f"**Completed elimination operations:** "
                f"{result.elimination_steps}"
            )
        return

    assert result.solution is not None
    st.success(result.message)

    solution_text = "<br>".join(
        f"x{index + 1} = {value:.10f}"
        for index, value in enumerate(result.solution)
    )

    st.markdown(
        f"""
        <div class="final-result-grid">
            <div class="final-result-card">
                <span>Status</span>
                <strong>{result.status}</strong>
            </div>

            <div class="final-result-card">
                <span>Elimination Steps</span>
                <strong>{result.elimination_steps}</strong>
            </div>

            <div class="final-result-card">
                <span>Back Substitution Steps</span>
                <strong>{result.back_substitution_steps}</strong>
            </div>

            <div class="final-result-card">
                <span>Residual 2-Norm</span>
                <strong>{result.residual_norm_2:.6e}</strong>
            </div>

            <div class="final-result-card">
                <span>Maximum Absolute Residual</span>
                <strong>{result.maximum_absolute_residual:.6e}</strong>
            </div>

            <div class="final-result-card">
                <span>Pivot Tolerance</span>
                <strong>{result.pivot_tolerance:.2e}</strong>
            </div>
        </div>

        <div class="final-interval-box">
            Approximate solution:<br>
            <strong>{solution_text}</strong>
        </div>

        <div class="final-interval-box">
            Stopping reason:
            <strong>{result.stopping_reason}</strong>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if result.warnings:
        for warning in result.warnings:
            st.warning(warning)


def render_original_system(result: GaussianEliminationResult) -> None:
    """Display the original coefficient, vector, and augmented matrices."""

    st.subheader("Original Matrix")
    coefficient_column, vector_column = st.columns([3, 1])

    with coefficient_column:
        st.markdown("**Coefficient Matrix A**")
        st.dataframe(
            round_numeric_dataframe(matrix_dataframe(result.original_matrix)),
            use_container_width=True,
        )

    with vector_column:
        st.markdown("**Right-Hand Side b**")
        st.dataframe(
            round_numeric_dataframe(vector_dataframe(result.original_vector)),
            use_container_width=True,
        )

    st.markdown("**Original Augmented Matrix [A | b]**")
    st.dataframe(
        round_numeric_dataframe(
            augmented_dataframe(
                result.original_matrix,
                result.original_vector,
            )
        ),
        use_container_width=True,
    )


def render_elimination_history(result: GaussianEliminationResult) -> None:
    """Display the summary and every before/after elimination snapshot."""

    st.subheader("Elimination Steps")
    summary_dataframe = build_elimination_summary_dataframe(result)

    if summary_dataframe.empty:
        st.info("No elimination operations were completed.")
        return

    st.dataframe(
        round_numeric_dataframe(summary_dataframe),
        use_container_width=True,
        hide_index=True,
    )

    for record in result.elimination_history:
        with st.expander(
            f"Step {record.step}: {record.row_operation}",
            expanded=False,
        ):
            metric_columns = st.columns(3)
            metric_columns[0].metric(
                "Pivot",
                format_number(record.pivot_value),
            )
            metric_columns[1].metric(
                "Multiplier",
                format_number(record.multiplier),
            )
            metric_columns[2].metric(
                "Eliminated Element",
                format_number(record.eliminated_element),
            )

            st.markdown(f"**Operation:** `{record.row_operation}`")
            before_column, after_column = st.columns(2)

            with before_column:
                st.markdown("**Before the operation**")
                st.dataframe(
                    round_numeric_dataframe(
                        augmented_dataframe(
                            record.matrix_before,
                            record.vector_before,
                        )
                    ),
                    use_container_width=True,
                )

            with after_column:
                st.markdown("**After the operation**")
                st.dataframe(
                    round_numeric_dataframe(
                        augmented_dataframe(
                            record.matrix_after,
                            record.vector_after,
                        )
                    ),
                    use_container_width=True,
                )

            st.caption(record.notes)


def render_upper_triangular_system(result: GaussianEliminationResult) -> None:
    """Display the final upper-triangular matrix and transformed vector."""

    st.subheader("Final Upper Triangular Matrix")
    matrix_column, vector_column = st.columns([3, 1])

    with matrix_column:
        st.dataframe(
            round_numeric_dataframe(matrix_dataframe(result.upper_matrix)),
            use_container_width=True,
        )

    with vector_column:
        st.dataframe(
            round_numeric_dataframe(
                vector_dataframe(result.transformed_vector, "Transformed b")
            ),
            use_container_width=True,
        )

    st.markdown("**Final Transformed Augmented Matrix**")
    st.dataframe(
        round_numeric_dataframe(
            augmented_dataframe(
                result.upper_matrix,
                result.transformed_vector,
            )
        ),
        use_container_width=True,
    )


def render_back_substitution(result: GaussianEliminationResult) -> None:
    """Display every manual back-substitution operation."""

    st.subheader("Back Substitution Table")
    dataframe = build_back_substitution_dataframe(result)

    if dataframe.empty:
        st.info("Back substitution was not completed.")
        return

    st.dataframe(
        round_numeric_dataframe(dataframe),
        use_container_width=True,
        hide_index=True,
    )


def render_solution_and_residual(result: GaussianEliminationResult) -> None:
    """Display the final solution and scientifically correct residual metrics."""

    st.subheader("Final Solution")
    st.dataframe(
        round_numeric_dataframe(build_solution_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Residual Analysis")
    st.latex(r"r = Ax - b")
    st.dataframe(
        round_numeric_dataframe(build_residual_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    metric_columns = st.columns(4)
    metric_columns[0].metric(
        "Residual 2-Norm",
        f"{result.residual_norm_2:.6e}",
    )
    metric_columns[1].metric(
        "Residual Infinity Norm",
        f"{result.residual_norm_inf:.6e}",
    )
    metric_columns[2].metric(
        "Maximum Absolute Residual",
        f"{result.maximum_absolute_residual:.6e}",
    )
    metric_columns[3].metric(
        "Relative Residual",
        f"{result.relative_residual:.6e}",
    )

    st.caption(
        "Residual values measure how closely the computed solution satisfies "
        "the original equations. They are not true solution errors unless an "
        "exact reference solution is known."
    )


def render_excel_download(result: GaussianEliminationResult) -> None:
    """Generate and display the required in-memory XLSX download button."""

    report_cache_key = "naive_gaussian_excel_report"
    report_signature_key = "naive_gaussian_excel_signature"

    try:
        if (
            st.session_state.get(report_signature_key)
            != result.input_signature
            or report_cache_key not in st.session_state
        ):
            st.session_state[report_cache_key] = create_excel_report(result)
            st.session_state[report_signature_key] = result.input_signature

        report_date = datetime.now(ZoneInfo(REPORT_TIME_ZONE)).strftime(
            "%Y%m%d_%H%M%S"
        )
        st.download_button(
            label="Download Excel Report",
            data=st.session_state[report_cache_key],
            file_name=(
                "naive_gaussian_elimination_report_"
                f"{report_date}.xlsx"
            ),
            mime=EXCEL_MIME_TYPE,
            use_container_width=True,
        )
    except (ValueError, ImportError, OSError) as error:
        st.error(
            "The Excel report could not be generated. Verify that openpyxl "
            f"is installed correctly. Details: {error}"
        )


# =============================================================================
# Streamlit page
# =============================================================================
def create_default_table(size: int) -> pd.DataFrame:
    """Create the editable table for the selected system size."""

    if size in DEFAULT_SYSTEMS:
        default_matrix, default_vector = DEFAULT_SYSTEMS[size]
    else:
        default_matrix = np.eye(size, dtype=float) * 2.0
        default_vector = np.ones(size, dtype=float)

    data = {
        f"x{column + 1}": default_matrix[:, column]
        for column in range(size)
    }
    data["b"] = default_vector
    return pd.DataFrame(data)


def extract_table_values(
    input_table: pd.DataFrame,
    size: int,
) -> tuple[list[list[str]], list[str]]:
    """Extract values as strings for validation and signature generation."""

    required_columns = variable_names(size) + ["b"]
    missing = [
        column for column in required_columns
        if column not in input_table.columns
    ]
    if missing:
        raise ValueError(
            "The table is missing required column(s): "
            + ", ".join(missing)
        )

    if input_table.shape != (size, size + 1):
        raise ValueError(
            f"The table must contain exactly {size} equations."
        )

    matrix_values = [
        [str(input_table.iloc[row][f"x{column + 1}"]) for column in range(size)]
        for row in range(size)
    ]
    vector_values = [
        str(input_table.iloc[row]["b"])
        for row in range(size)
    ]
    return matrix_values, vector_values


def render_diagnostics(result: GaussianEliminationResult) -> None:
    """Render numerical diagnostics for a completed solution."""

    metrics = st.columns(4)
    metrics[0].metric(
        "Condition Number",
        format_number(result.condition_number),
    )
    metrics[1].metric(
        "Determinant",
        format_number(result.determinant),
    )
    metrics[2].metric(
        "Growth Factor",
        format_number(result.growth_factor),
    )
    metrics[3].metric(
        "Minimum Pivot Ratio",
        format_number(result.minimum_pivot_ratio),
    )

    st.caption(
        f"Rank(A) = {result.diagnostics.rank_a}; "
        f"Rank([A|b]) = {result.diagnostics.rank_augmented}. "
        "Conditioning and pivot ratios are diagnostics only; the displayed "
        "solution is produced by manual elimination without row swaps."
    )


def render_page() -> None:
    """Render the complete Naïve Gaussian Elimination page."""

    st.set_page_config(
        page_title="Naive Gaussian Elimination Solver | Numerical Methods",
        page_icon="📘",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">NAIVE GAUSSIAN ELIMINATION TOOL</div>
                <h1>Naive Gaussian Elimination Solver</h1>
                <p>
                    Enter a square linear system, perform forward elimination
                    without pivoting, and inspect the upper-triangular system,
                    manual back substitution, residuals, diagnostics, graphs,
                    and Excel report.
                </p>
                <div class="method-actions">
                    <a href="/Naive_Gaussian_Elimination" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Naive_Gaussian_Elimination_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns(
        [0.035, 0.93, 0.035]
    )

    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )

        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Enter the System")
                st.markdown(
                    """
                    Choose the number of equations, then enter the
                    coefficients of **A** and the right-hand side **b**.

                    Each row represents one equation and each `x` column
                    contains the coefficient of one unknown.
                    """
                )
                st.markdown("**Method formula**")
                st.latex(
                    r"m_{ik}=\frac{a_{ik}}{a_{kk}},\qquad "
                    r"R_i\leftarrow R_i-m_{ik}R_k"
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")
                st.markdown(
                    """
                    - The coefficient matrix must be square.
                    - No row swapping is performed.
                    - Every diagonal pivot must be nonzero and usable relative
                      to the active matrix scale.
                    - A zero or poor pivot can cause method breakdown even when
                      the original system has a unique solution.
                    - Use the scaled-pivoting solver for greater numerical reliability.
                    """
                )
                st.info(
                    "The solver intentionally stops instead of secretly swapping "
                    "rows, so the page remains a true Naive Gaussian Elimination tool."
                )

        input_column, result_column = st.columns([1.35, 1.0])

        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )

                st.markdown(
                    '<div class="input-label-ui">Number of equations</div>',
                    unsafe_allow_html=True,
                )
                size = int(
                    st.number_input(
                        "Number of equations",
                        min_value=min(SUPPORTED_SIZES),
                        max_value=max(SUPPORTED_SIZES),
                        value=DEFAULT_SIZE,
                        step=1,
                        label_visibility="collapsed",
                        key="naive_gaussian_size",
                    )
                )

                st.markdown(
                    '<div class="input-label-ui">Coefficient table [A | b]</div>',
                    unsafe_allow_html=True,
                )
                input_table = st.data_editor(
                    create_default_table(size),
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    column_config={
                        **{
                            f"x{column + 1}": st.column_config.NumberColumn(
                                f"x{column + 1}",
                                format="%.12g",
                            )
                            for column in range(size)
                        },
                        "b": st.column_config.NumberColumn(
                            "b",
                            format="%.12g",
                        ),
                    },
                    key=f"naive_gaussian_table_{size}",
                )

                try:
                    raw_matrix, raw_vector = extract_table_values(
                        input_table,
                        size,
                    )
                    current_signature = create_input_signature(
                        raw_matrix,
                        raw_vector,
                        PIVOT_TOLERANCE,
                    )
                    input_error = None
                except ValueError as error:
                    raw_matrix = []
                    raw_vector = []
                    current_signature = hashlib.sha256(
                        str(error).encode("utf-8")
                    ).hexdigest()
                    input_error = str(error)

                solve_column, reset_column = st.columns(2)

                with solve_column:
                    solve_clicked = st.button(
                        "Solve",
                        type="primary",
                        use_container_width=True,
                        key="naive_gaussian_solve_button",
                    )

                with reset_column:
                    reset_clicked = st.button(
                        "Reset Result",
                        use_container_width=True,
                        key="naive_gaussian_reset_button",
                    )

                if reset_clicked:
                    st.session_state.pop(
                        "naive_gaussian_result",
                        None,
                    )
                    st.session_state.pop(
                        "naive_gaussian_excel_report",
                        None,
                    )
                    st.session_state.pop(
                        "naive_gaussian_excel_signature",
                        None,
                    )
                    st.rerun()

                if solve_clicked:
                    if input_error is not None:
                        st.session_state.naive_gaussian_result = (
                            solve_naive_gaussian_elimination(
                                raw_matrix=[],
                                raw_vector=[],
                                input_signature=current_signature,
                                pivot_tolerance=PIVOT_TOLERANCE,
                            )
                        )
                    else:
                        st.session_state.naive_gaussian_result = (
                            solve_naive_gaussian_elimination(
                                raw_matrix=raw_matrix,
                                raw_vector=raw_vector,
                                input_signature=current_signature,
                                pivot_tolerance=PIVOT_TOLERANCE,
                            )
                        )

                    st.session_state.pop(
                        "naive_gaussian_excel_report",
                        None,
                    )
                    st.session_state.pop(
                        "naive_gaussian_excel_signature",
                        None,
                    )
                    st.rerun()

                st.caption(
                    "Full precision is used for all calculations. Tables are "
                    "rounded only for display."
                )

        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )

                result = st.session_state.get(
                    "naive_gaussian_result"
                )

                if result is None:
                    st.info(
                        "Enter the system and click Solve to display the result."
                    )
                elif result.input_signature != current_signature:
                    st.info(
                        "The matrix or system size changed. Click Solve to "
                        "calculate a new result."
                    )
                else:
                    render_final_result(result)

        result = st.session_state.get(
            "naive_gaussian_result"
        )
        result_is_current = (
            result is not None
            and result.input_signature == current_signature
        )

        if result_is_current:
            if result.original_matrix.size > 0:
                st.divider()
                render_original_system(result)

            if result.elimination_history:
                st.divider()
                render_elimination_history(result)

            if result.upper_matrix.size > 0:
                st.divider()
                render_upper_triangular_system(result)

            if result.back_substitution_history:
                st.divider()
                render_back_substitution(result)

            if result.success:
                st.divider()
                render_solution_and_residual(result)

                st.divider()
                st.subheader("Numerical Diagnostics")
                render_diagnostics(result)

                st.divider()
                render_excel_download(result)

                st.divider()
                graph_column, solution_column, residual_column = st.columns(3)

                with graph_column:
                    with st.container(border=True):
                        st.subheader("Coefficient Matrix")
                        figure = create_matrix_figure(result)
                        st.pyplot(figure, use_container_width=True)
                        plt.close(figure)

                with solution_column:
                    with st.container(border=True):
                        st.subheader("Solution Values")
                        figure = create_solution_figure(result)
                        st.pyplot(figure, use_container_width=True)
                        plt.close(figure)

                with residual_column:
                    with st.container(border=True):
                        st.subheader("Residual Components")
                        figure = create_residual_figure(result)
                        st.pyplot(figure, use_container_width=True)
                        plt.close(figure)

                st.divider()
                navigation_left, navigation_right = st.columns(2)

                with navigation_left:
                    if st.button(
                        "Review Naive Gaussian Elimination Lesson",
                        use_container_width=True,
                        key="naive_gaussian_lesson_button",
                    ):
                        st.switch_page(
                            "pages/Naive_Gaussian_Elimination.py"
                        )

                with navigation_right:
                    if st.button(
                        "Back to Solver Menu",
                        use_container_width=True,
                        key="naive_gaussian_menu_button",
                    ):
                        st.switch_page(
                            "pages/Numerical_Solver.py"
                        )

        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Naive Gaussian Elimination • Linear Systems</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()