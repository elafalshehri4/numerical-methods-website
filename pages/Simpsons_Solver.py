from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css



# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Show three decimals and use × 10ⁿ when fixed notation loses meaning."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"

# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Simpson's Rules"
DISPLAY_DECIMALS = 3
DEFAULT_FUNCTION = "sin(x)"
DEFAULT_LOWER_LIMIT = 0.0
DEFAULT_UPPER_LIMIT = float(np.pi)
DEFAULT_SUBINTERVALS = 6
MIN_SUBINTERVALS = 2
MAX_SUBINTERVALS = 600
DEFAULT_CONVERGENCE_LEVELS = 5
MIN_CONVERGENCE_LEVELS = 3
MAX_CONVERGENCE_LEVELS = 7
ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

RULE_OPTIONS = {
    "Composite Simpson's 1/3 Rule": "simpson_13_composite",
    "Simple Simpson's 1/3 Rule": "simpson_13_simple",
    "Composite Simpson's 3/8 Rule": "simpson_38_composite",
    "Simple Simpson's 3/8 Rule": "simpson_38_simple",
}

ALLOWED_FUNCTION_NAMES = {
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class FunctionSample:
    """One node and its Simpson weight."""

    index: int
    x_value: float
    function_value: float
    weight: int
    weighted_value: float
    point_type: str


@dataclass(frozen=True)
class PanelContribution:
    """One Simpson panel contribution."""

    panel_index: int
    start_node: int
    end_node: int
    node_indices: tuple[int, ...]
    x_values: tuple[float, ...]
    function_values: tuple[float, ...]
    local_weights: tuple[int, ...]
    panel_width: float
    weighted_sum: float
    signed_contribution: float
    cumulative_integral: float
    operation: str


@dataclass(frozen=True)
class ConvergenceRecord:
    """One mesh-refinement calculation."""

    level: int
    subintervals: int
    step_size: float
    approximation: float
    exact_integral: float | None
    absolute_error: float | None
    relative_error: float | None
    successive_difference: float | None
    observed_order: float | None


@dataclass(frozen=True)
class SimpsonsResult:
    """Complete result shared by Streamlit and Excel renderers."""

    status: str
    success: bool
    method: str
    rule_key: str
    rule_name: str
    message: str
    stopping_reason: str
    function_text: str
    function_expression: sp.Expr | None
    antiderivative_expression: sp.Expr | None
    antiderivative_text: str
    lower_limit: float | None
    upper_limit: float | None
    orientation: str
    subintervals: int
    step_size: float | None
    panel_size: int
    panel_factor: float | None
    function_samples: tuple[FunctionSample, ...]
    panel_contributions: tuple[PanelContribution, ...]
    weighted_sum: float | None
    approximate_integral: float | None
    exact_integral: float | None
    absolute_error: float | None
    relative_error: float | None
    percentage_error: float | None
    theoretical_order: int
    convergence_records: tuple[ConvergenceRecord, ...]
    latest_observed_order: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def format_number(
    value: float | int | None,
    decimals: int = 3,
) -> str:
    """Format displayed values with three decimals and × 10ⁿ notation."""

    return format_display_number(value, decimals)



def scientific_number(value: float | int | None) -> str:
    """Format a value as a three-decimal coefficient multiplied by 10ⁿ."""

    return format_scientific_power(value)



def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only for display."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def create_input_signature(
    function_text: str,
    lower_limit: Any,
    upper_limit: Any,
    rule_name: str,
    subintervals: Any,
    convergence_levels: Any,
) -> str:
    """Create a stable signature used to prevent stale Streamlit results."""

    payload = repr(
        (
            str(function_text).strip(),
            str(lower_limit),
            str(upper_limit),
            str(rule_name),
            str(subintervals),
            str(convergence_levels),
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def safe_float(raw_value: Any, value_name: str) -> float:
    """Convert one user input to a finite real floating-point value."""

    if raw_value is None:
        raise ValueError(f"{value_name} is required.")
    if isinstance(raw_value, str) and not raw_value.strip():
        raise ValueError(f"{value_name} is required.")

    try:
        value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be a valid real number.") from error

    if not math.isfinite(value):
        raise ValueError(
            f"{value_name} must be finite; NaN and infinity are not allowed."
        )
    return value


def safe_positive_integer(
    raw_value: Any,
    value_name: str,
    minimum: int,
    maximum: int,
) -> int:
    """Convert one user input to a bounded positive integer."""

    if isinstance(raw_value, bool):
        raise ValueError(f"{value_name} must be an integer.")

    try:
        numeric_value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be an integer.") from error

    if not math.isfinite(numeric_value) or not numeric_value.is_integer():
        raise ValueError(f"{value_name} must be an integer.")

    integer_value = int(numeric_value)
    if integer_value < minimum or integer_value > maximum:
        raise ValueError(
            f"{value_name} must be between {minimum} and {maximum}."
        )
    return integer_value


def rule_configuration(rule_key: str) -> tuple[int, float, int, str]:
    """Return panel size, factor, forced n, and validity message."""

    if rule_key == "simpson_13_simple":
        return 2, 1.0 / 3.0, 2, "n must equal 2."
    if rule_key == "simpson_13_composite":
        return 2, 1.0 / 3.0, 0, "n must be a positive even integer."
    if rule_key == "simpson_38_simple":
        return 3, 3.0 / 8.0, 3, "n must equal 3."
    if rule_key == "simpson_38_composite":
        return 3, 3.0 / 8.0, 0, "n must be a positive multiple of 3."
    raise ValueError("The selected Simpson rule is invalid.")


def validate_subintervals(rule_key: str, raw_subintervals: Any) -> int:
    """Validate n against the mathematical requirement of the selected rule."""

    panel_size, _factor, forced_n, requirement = rule_configuration(rule_key)
    if forced_n:
        return forced_n

    n_value = safe_positive_integer(
        raw_subintervals,
        "Number of subintervals n",
        MIN_SUBINTERVALS,
        MAX_SUBINTERVALS,
    )
    if n_value % panel_size != 0:
        raise ValueError(
            f"Invalid number of subintervals for the selected rule: {requirement}"
        )
    return n_value


def empty_result(
    *,
    message: str,
    stopping_reason: str,
    input_signature: str,
    rule_key: str = "simpson_13_composite",
    rule_name: str = "Composite Simpson's 1/3 Rule",
) -> SimpsonsResult:
    """Create a consistent structured error result."""

    panel_size, panel_factor, _forced_n, _requirement = rule_configuration(rule_key)
    return SimpsonsResult(
        status="error",
        success=False,
        method=METHOD_NAME,
        rule_key=rule_key,
        rule_name=rule_name,
        message=message,
        stopping_reason=stopping_reason,
        function_text="",
        function_expression=None,
        antiderivative_expression=None,
        antiderivative_text="Not available",
        lower_limit=None,
        upper_limit=None,
        orientation="Not available",
        subintervals=0,
        step_size=None,
        panel_size=panel_size,
        panel_factor=panel_factor,
        function_samples=(),
        panel_contributions=(),
        weighted_sum=None,
        approximate_integral=None,
        exact_integral=None,
        absolute_error=None,
        relative_error=None,
        percentage_error=None,
        theoretical_order=4,
        convergence_records=(),
        latest_observed_order=None,
        warnings=(),
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


# =============================================================================
# Function parsing and safe evaluation
# =============================================================================
def parse_function(function_text: str) -> tuple[sp.Symbol, sp.Expr]:
    """Parse and validate a real single-variable mathematical function."""

    if not isinstance(function_text, str) or not function_text.strip():
        raise ValueError("Function f(x) is required.")

    cleaned_text = function_text.strip().replace("^", "**")
    if "=" in cleaned_text:
        raise ValueError(
            "Enter only the function expression, not an equation. "
            "For example, enter sin(x) instead of y = sin(x)."
        )

    x_symbol = sp.Symbol("x", real=True)
    local_dictionary = {"x": x_symbol, **ALLOWED_FUNCTION_NAMES}

    try:
        expression = sp.sympify(cleaned_text, locals=local_dictionary)
    except (sp.SympifyError, SyntaxError, TypeError, ValueError) as error:
        raise ValueError(
            "The function format is invalid. Use Python/SymPy syntax such as "
            "sin(x), exp(x), sqrt(x), or x**3 - 2*x + 1."
        ) from error

    if isinstance(expression, Relational):
        raise ValueError("The input must be a function expression, not a relation.")

    unknown_symbols = expression.free_symbols - {x_symbol}
    if unknown_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unknown_symbols))
        raise ValueError(
            "Only the independent variable x is allowed. "
            f"Unexpected symbol(s): {names}."
        )

    undefined_functions = expression.atoms(AppliedUndef)
    if undefined_functions:
        names = ", ".join(sorted(str(item.func) for item in undefined_functions))
        raise ValueError(f"Unsupported function name(s): {names}.")

    if expression.has(sp.zoo, sp.oo, -sp.oo, sp.nan):
        raise ValueError("The function contains an undefined or infinite quantity.")

    return x_symbol, sp.simplify(expression)


def build_numeric_function(
    x_symbol: sp.Symbol,
    expression: sp.Expr,
) -> Callable[[Any], Any]:
    """Convert a symbolic expression to a NumPy-compatible function."""

    try:
        return sp.lambdify(x_symbol, expression, modules=["numpy"])
    except (TypeError, ValueError, NotImplementedError) as error:
        raise ValueError(
            "The function could not be converted to a numerical function."
        ) from error


def evaluate_real_scalar(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    value_name: str,
) -> float:
    """Evaluate a function safely and return one finite real scalar."""

    try:
        with np.errstate(all="ignore"):
            raw_value = numeric_function(float(x_value))
    except (ArithmeticError, TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The function could not be evaluated at x = {x_value:.12g}."
        ) from error

    array = np.asarray(raw_value)
    if array.size != 1:
        raise ValueError(f"{value_name} did not produce a scalar value.")

    scalar = array.reshape(-1)[0]
    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)
        if abs(complex_value.imag) > 1.0e-12:
            raise ValueError(
                f"The function is complex at x = {x_value:.12g}; "
                "this solver requires a real-valued integrand."
            )
        scalar = complex_value.real

    try:
        value = float(scalar)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The function value at x = {x_value:.12g} is not a valid real number."
        ) from error

    if not math.isfinite(value):
        raise ValueError(
            f"The function is undefined or non-finite at x = {x_value:.12g}."
        )
    return value


def evaluate_real_array(
    numeric_function: Callable[[Any], Any],
    x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate a function on an array and return real values."""

    try:
        with np.errstate(all="ignore"):
            raw_values = numeric_function(x_values)
    except (ArithmeticError, TypeError, ValueError, OverflowError) as error:
        raise ValueError("The function could not be evaluated over the graph range.") from error

    array = np.asarray(raw_values)
    if array.ndim == 0:
        array = np.full_like(x_values, float(array), dtype=float)
    else:
        try:
            array = np.broadcast_to(array, x_values.shape)
        except ValueError as error:
            raise ValueError("The function returned an unexpected graph shape.") from error

    if np.iscomplexobj(array):
        if np.any(np.abs(np.imag(array)) > 1.0e-12):
            raise ValueError("The function produced complex values in the graph range.")
        array = np.real(array)

    try:
        return np.asarray(array, dtype=float)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError("The graph values could not be converted to real numbers.") from error


def calculate_reference_integral(
    expression: sp.Expr,
    x_symbol: sp.Symbol,
    lower_limit: float,
    upper_limit: float,
) -> tuple[sp.Expr | None, float | None, str | None]:
    """Attempt to obtain a symbolic antiderivative and definite integral."""

    try:
        antiderivative = sp.integrate(expression, x_symbol)
    except (TypeError, ValueError, NotImplementedError):
        return None, None, "A symbolic reference integral could not be obtained."

    if isinstance(antiderivative, sp.Integral) or antiderivative.has(sp.Integral):
        return None, None, "A closed-form antiderivative was not available."

    try:
        lower_symbolic = sp.Float(lower_limit, 30)
        upper_symbolic = sp.Float(upper_limit, 30)
        exact_expression = antiderivative.subs(
            x_symbol, upper_symbolic
        ) - antiderivative.subs(x_symbol, lower_symbolic)
        exact_numeric = sp.N(exact_expression, 30)
    except (TypeError, ValueError, ArithmeticError, NotImplementedError):
        return antiderivative, None, "The symbolic reference integral could not be evaluated."

    if exact_numeric.has(sp.zoo, sp.oo, -sp.oo, sp.nan):
        return antiderivative, None, "The exact integral is undefined or divergent."
    if exact_numeric.is_real is False:
        return antiderivative, None, "The exact integral is not real on the selected interval."

    try:
        exact_value = float(exact_numeric)
    except (TypeError, ValueError, OverflowError):
        return antiderivative, None, "The exact integral could not be converted to a real value."

    if not math.isfinite(exact_value):
        return antiderivative, None, "The exact integral is non-finite."
    return antiderivative, exact_value, None


# =============================================================================
# Simpson numerical algorithms
# =============================================================================
def node_weight(rule_key: str, index: int, subintervals: int) -> int:
    """Return the standard textbook Simpson weight for one node."""

    if index in (0, subintervals):
        return 1

    if rule_key in {"simpson_13_simple", "simpson_13_composite"}:
        return 4 if index % 2 == 1 else 2

    if rule_key in {"simpson_38_simple", "simpson_38_composite"}:
        return 2 if index % 3 == 0 else 3

    raise ValueError("The selected Simpson rule is invalid.")


def point_type(rule_key: str, index: int, subintervals: int) -> str:
    """Return an educational label for one Simpson node."""

    if index == 0:
        return "Left Endpoint"
    if index == subintervals:
        return "Right Endpoint"
    weight = node_weight(rule_key, index, subintervals)
    if weight == 4:
        return "Odd Interior Node"
    if weight == 2 and rule_key.startswith("simpson_13"):
        return "Even Interior Node"
    if weight == 2:
        return "Multiple-of-3 Interior Node"
    return "Non-multiple-of-3 Interior Node"


def panel_local_weights(rule_key: str) -> tuple[int, ...]:
    """Return local weights for one Simpson panel."""

    if rule_key.startswith("simpson_13"):
        return (1, 4, 1)
    if rule_key.startswith("simpson_38"):
        return (1, 3, 3, 1)
    raise ValueError("The selected Simpson rule is invalid.")


def calculate_simpson_approximation(
    numeric_function: Callable[[Any], Any],
    lower_limit: float,
    upper_limit: float,
    subintervals: int,
    rule_key: str,
) -> tuple[
    tuple[FunctionSample, ...],
    tuple[PanelContribution, ...],
    float,
    float,
    float,
]:
    """Calculate all nodes, weights, panels, and the Simpson approximation."""

    panel_size, panel_factor, _forced_n, _requirement = rule_configuration(rule_key)
    if subintervals % panel_size != 0:
        raise ValueError("The number of subintervals is incompatible with the selected rule.")

    step_size = (upper_limit - lower_limit) / subintervals
    if step_size == 0.0 or not math.isfinite(step_size):
        raise ValueError("The calculated step size h is zero or non-finite.")

    x_values = np.linspace(lower_limit, upper_limit, subintervals + 1)
    function_values = np.array(
        [
            evaluate_real_scalar(numeric_function, float(x_value), f"f(x_{index})")
            for index, x_value in enumerate(x_values)
        ],
        dtype=float,
    )

    weights = np.array(
        [node_weight(rule_key, index, subintervals) for index in range(subintervals + 1)],
        dtype=int,
    )
    weighted_values = weights * function_values

    samples = tuple(
        FunctionSample(
            index=index,
            x_value=float(x_values[index]),
            function_value=float(function_values[index]),
            weight=int(weights[index]),
            weighted_value=float(weighted_values[index]),
            point_type=point_type(rule_key, index, subintervals),
        )
        for index in range(subintervals + 1)
    )

    local_weights = panel_local_weights(rule_key)
    panels: list[PanelContribution] = []
    cumulative = 0.0

    for panel_index, start in enumerate(
        range(0, subintervals, panel_size),
        start=1,
    ):
        indices = tuple(range(start, start + panel_size + 1))
        local_x = tuple(float(x_values[index]) for index in indices)
        local_f = tuple(float(function_values[index]) for index in indices)
        local_weighted_sum = float(
            sum(weight * value for weight, value in zip(local_weights, local_f))
        )
        contribution = float(panel_factor * step_size * local_weighted_sum)
        cumulative += contribution

        weight_expression = " + ".join(
            f"{weight}f(x_{index})" for weight, index in zip(local_weights, indices)
        )
        numeric_expression = " + ".join(
            f"{weight}({value:.12g})" for weight, value in zip(local_weights, local_f)
        )
        factor_text = "h/3" if panel_size == 2 else "3h/8"
        operation = (
            f"Panel {panel_index}: {factor_text}[{weight_expression}] = "
            f"{panel_factor:.12g}({step_size:.12g})[{numeric_expression}] = "
            f"{contribution:.12g}"
        )

        panels.append(
            PanelContribution(
                panel_index=panel_index,
                start_node=start,
                end_node=start + panel_size,
                node_indices=indices,
                x_values=local_x,
                function_values=local_f,
                local_weights=local_weights,
                panel_width=float(panel_size * step_size),
                weighted_sum=local_weighted_sum,
                signed_contribution=contribution,
                cumulative_integral=float(cumulative),
                operation=operation,
            )
        )

    weighted_sum = float(np.sum(weighted_values))
    approximation = float(panel_factor * step_size * weighted_sum)

    if not math.isfinite(approximation):
        raise ValueError("The Simpson approximation became non-finite.")

    panel_sum = float(sum(panel.signed_contribution for panel in panels))
    consistency_scale = max(1.0, abs(approximation), abs(panel_sum))
    if abs(panel_sum - approximation) > 1.0e-11 * consistency_scale:
        raise ValueError("An internal consistency check failed for the Simpson panels.")

    return samples, tuple(panels), float(step_size), weighted_sum, approximation


def calculate_error_metrics(
    approximation: float,
    exact_value: float | None,
) -> tuple[float | None, float | None, float | None]:
    """Calculate absolute, relative, and percentage integration errors."""

    if exact_value is None:
        return None, None, None

    absolute_error = abs(exact_value - approximation)
    if abs(exact_value) <= RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
        return float(absolute_error), None, None

    relative_error = absolute_error / abs(exact_value)
    return float(absolute_error), float(relative_error), float(100.0 * relative_error)


def calculate_convergence_analysis(
    numeric_function: Callable[[Any], Any],
    lower_limit: float,
    upper_limit: float,
    initial_subintervals: int,
    levels: int,
    exact_integral: float | None,
    rule_key: str,
) -> tuple[tuple[ConvergenceRecord, ...], float | None, tuple[str, ...]]:
    """Refine n by factors of two and estimate convergence behavior."""

    records: list[ConvergenceRecord] = []
    warnings: list[str] = []
    previous_approximation: float | None = None
    previous_absolute_error: float | None = None
    previous_successive_difference: float | None = None

    for level in range(levels):
        subintervals = initial_subintervals * (2**level)
        if subintervals > 100_000:
            warnings.append(
                "Convergence refinement stopped because n would exceed 100,000."
            )
            break

        (
            _samples,
            _panels,
            step_size,
            _weighted_sum,
            approximation,
        ) = calculate_simpson_approximation(
            numeric_function,
            lower_limit,
            upper_limit,
            subintervals,
            rule_key,
        )

        absolute_error, relative_error, _percentage_error = calculate_error_metrics(
            approximation,
            exact_integral,
        )
        successive_difference = (
            None
            if previous_approximation is None
            else abs(approximation - previous_approximation)
        )

        observed_order: float | None = None
        if (
            exact_integral is not None
            and previous_absolute_error is not None
            and absolute_error is not None
            and previous_absolute_error > ZERO_TOLERANCE
            and absolute_error > ZERO_TOLERANCE
        ):
            observed_order = math.log(previous_absolute_error / absolute_error, 2.0)
        elif (
            previous_successive_difference is not None
            and successive_difference is not None
            and previous_successive_difference > ZERO_TOLERANCE
            and successive_difference > ZERO_TOLERANCE
        ):
            observed_order = math.log(
                previous_successive_difference / successive_difference,
                2.0,
            )

        records.append(
            ConvergenceRecord(
                level=level,
                subintervals=subintervals,
                step_size=float(step_size),
                approximation=float(approximation),
                exact_integral=exact_integral,
                absolute_error=absolute_error,
                relative_error=relative_error,
                successive_difference=successive_difference,
                observed_order=observed_order,
            )
        )

        previous_approximation = approximation
        previous_absolute_error = absolute_error
        previous_successive_difference = successive_difference

    latest_order = next(
        (
            record.observed_order
            for record in reversed(records)
            if record.observed_order is not None and math.isfinite(record.observed_order)
        ),
        None,
    )
    return tuple(records), latest_order, tuple(warnings)


def solve_simpsons_rules(
    function_text: str,
    lower_limit_input: Any,
    upper_limit_input: Any,
    rule_name: str,
    subintervals_input: Any,
    convergence_levels_input: Any,
) -> SimpsonsResult:
    """Validate all inputs and execute the complete Simpson workflow."""

    input_signature = create_input_signature(
        function_text,
        lower_limit_input,
        upper_limit_input,
        rule_name,
        subintervals_input,
        convergence_levels_input,
    )

    rule_key = RULE_OPTIONS.get(rule_name)
    if rule_key is None:
        return empty_result(
            message="The selected Simpson rule is invalid.",
            stopping_reason="Input validation failed.",
            input_signature=input_signature,
        )

    try:
        lower_limit = safe_float(lower_limit_input, "Lower integration limit a")
        upper_limit = safe_float(upper_limit_input, "Upper integration limit b")
        convergence_levels = safe_positive_integer(
            convergence_levels_input,
            "Convergence levels",
            MIN_CONVERGENCE_LEVELS,
            MAX_CONVERGENCE_LEVELS,
        )
        if lower_limit == upper_limit:
            raise ValueError("The lower and upper integration limits must be different.")

        subintervals = validate_subintervals(rule_key, subintervals_input)
        panel_size, panel_factor, _forced_n, _requirement = rule_configuration(rule_key)

        x_symbol, function_expression = parse_function(function_text)
        numeric_function = build_numeric_function(x_symbol, function_expression)

        (
            function_samples,
            panel_contributions,
            step_size,
            weighted_sum,
            approximate_integral,
        ) = calculate_simpson_approximation(
            numeric_function,
            lower_limit,
            upper_limit,
            subintervals,
            rule_key,
        )

        antiderivative_expression, exact_integral, reference_warning = (
            calculate_reference_integral(
                function_expression,
                x_symbol,
                lower_limit,
                upper_limit,
            )
        )
        absolute_error, relative_error, percentage_error = calculate_error_metrics(
            approximate_integral,
            exact_integral,
        )
        convergence_records, latest_order, convergence_warnings = (
            calculate_convergence_analysis(
                numeric_function,
                lower_limit,
                upper_limit,
                subintervals,
                convergence_levels,
                exact_integral,
                rule_key,
            )
        )

    except ValueError as error:
        return empty_result(
            message=str(error),
            stopping_reason="Simpson's Rules were not executed because validation failed.",
            input_signature=input_signature,
            rule_key=rule_key,
            rule_name=rule_name,
        )

    warnings: list[str] = list(convergence_warnings)
    if reference_warning:
        warnings.append(reference_warning)
    if upper_limit < lower_limit:
        warnings.append(
            "The limits are reversed. The reported result is a signed integral."
        )
    if rule_key.endswith("simple"):
        warnings.append(
            "The simple rule uses one Simpson panel only. Select a composite rule "
            "and increase n when improved accuracy is required."
        )
    if latest_order is not None and latest_order < 2.5:
        warnings.append(
            "The latest observed order is below the theoretical fourth order. "
            "The integrand may be insufficiently smooth, the mesh may not yet be "
            "in the asymptotic range, or floating-point effects may be significant."
        )

    return SimpsonsResult(
        status="success",
        success=True,
        method=METHOD_NAME,
        rule_key=rule_key,
        rule_name=rule_name,
        message="Execution completed successfully.",
        stopping_reason=(
            "The definite integral was approximated after all Simpson weights "
            "and panel contributions were evaluated."
        ),
        function_text=function_text.strip().replace("^", "**"),
        function_expression=function_expression,
        antiderivative_expression=antiderivative_expression,
        antiderivative_text=(
            str(antiderivative_expression)
            if antiderivative_expression is not None
            else "Not available"
        ),
        lower_limit=lower_limit,
        upper_limit=upper_limit,
        orientation="Forward" if upper_limit > lower_limit else "Reversed",
        subintervals=subintervals,
        step_size=step_size,
        panel_size=panel_size,
        panel_factor=panel_factor,
        function_samples=function_samples,
        panel_contributions=panel_contributions,
        weighted_sum=weighted_sum,
        approximate_integral=approximate_integral,
        exact_integral=exact_integral,
        absolute_error=absolute_error,
        relative_error=relative_error,
        percentage_error=percentage_error,
        theoretical_order=4,
        convergence_records=convergence_records,
        latest_observed_order=latest_order,
        warnings=tuple(dict.fromkeys(warnings)),
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


# =============================================================================
# DataFrame builders
# =============================================================================
def create_function_values_dataframe(result: SimpsonsResult) -> pd.DataFrame:
    """Create the complete Simpson node and weight table."""

    return pd.DataFrame(
        {
            "i": [sample.index for sample in result.function_samples],
            "x_i": [sample.x_value for sample in result.function_samples],
            "f(x_i)": [sample.function_value for sample in result.function_samples],
            "Weight": [sample.weight for sample in result.function_samples],
            "Weighted Value": [sample.weighted_value for sample in result.function_samples],
            "Point Type": [sample.point_type for sample in result.function_samples],
        }
    )


def create_panel_dataframe(result: SimpsonsResult) -> pd.DataFrame:
    """Create the complete table of Simpson panel contributions."""

    return pd.DataFrame(
        {
            "Panel": [panel.panel_index for panel in result.panel_contributions],
            "Start Node": [panel.start_node for panel in result.panel_contributions],
            "End Node": [panel.end_node for panel in result.panel_contributions],
            "Node Indices": [str(panel.node_indices) for panel in result.panel_contributions],
            "x Values": [str(tuple(round(value, 12) for value in panel.x_values)) for panel in result.panel_contributions],
            "f(x) Values": [str(tuple(round(value, 12) for value in panel.function_values)) for panel in result.panel_contributions],
            "Local Weights": [str(panel.local_weights) for panel in result.panel_contributions],
            "Panel Width": [panel.panel_width for panel in result.panel_contributions],
            "Weighted Sum": [panel.weighted_sum for panel in result.panel_contributions],
            "Signed Contribution": [panel.signed_contribution for panel in result.panel_contributions],
            "Cumulative Integral": [panel.cumulative_integral for panel in result.panel_contributions],
            "Operation": [panel.operation for panel in result.panel_contributions],
        }
    )


def create_convergence_dataframe(result: SimpsonsResult) -> pd.DataFrame:
    """Create the refinement and convergence table."""

    return pd.DataFrame(
        {
            "Level": [record.level for record in result.convergence_records],
            "Subintervals n": [record.subintervals for record in result.convergence_records],
            "Step Size h": [record.step_size for record in result.convergence_records],
            "Approximate Integral": [record.approximation for record in result.convergence_records],
            "Exact Integral": [record.exact_integral for record in result.convergence_records],
            "Absolute Error": [record.absolute_error for record in result.convergence_records],
            "Relative Error": [record.relative_error for record in result.convergence_records],
            "Successive Difference": [record.successive_difference for record in result.convergence_records],
            "Observed Order": [record.observed_order for record in result.convergence_records],
        }
    )


def create_error_dataframe(result: SimpsonsResult) -> pd.DataFrame:
    """Create a compact error-analysis table."""

    return pd.DataFrame(
        {
            "Metric": [
                "Approximate Integral",
                "Exact Integral",
                "Absolute Error",
                "Relative Error",
                "Percentage Error",
                "Theoretical Order",
                "Latest Observed Order",
            ],
            "Value": [
                result.approximate_integral,
                result.exact_integral,
                result.absolute_error,
                result.relative_error,
                result.percentage_error,
                result.theoretical_order,
                result.latest_observed_order,
            ],
        }
    )


def create_summary_dataframe(result: SimpsonsResult) -> pd.DataFrame:
    """Create the Summary worksheet table."""

    return pd.DataFrame(
        {
            "Property": [
                "Method",
                "Rule",
                "Status",
                "Function",
                "Antiderivative",
                "Lower Limit a",
                "Upper Limit b",
                "Orientation",
                "Number of Subintervals",
                "Panel Size",
                "Number of Panels",
                "Step Size h",
                "Global Weighted Sum",
                "Approximate Integral",
                "Exact Integral",
                "Absolute Error",
                "Relative Error",
                "Percentage Error",
                "Theoretical Order",
                "Latest Observed Order",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.rule_name,
                result.status,
                result.function_text,
                result.antiderivative_text,
                result.lower_limit,
                result.upper_limit,
                result.orientation,
                result.subintervals,
                result.panel_size,
                len(result.panel_contributions),
                result.step_size,
                result.weighted_sum,
                result.approximate_integral,
                result.exact_integral,
                result.absolute_error,
                result.relative_error,
                result.percentage_error,
                result.theoretical_order,
                result.latest_observed_order,
                " | ".join(result.warnings) if result.warnings else "None",
                result.stopping_reason,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )


def create_formula_dataframe(result: SimpsonsResult) -> pd.DataFrame:
    """Create a workbook table containing the formulas used."""

    if result.rule_key.startswith("simpson_13"):
        global_formula = "S = h/3 [f(x0) + 4Σ f(x_odd) + 2Σ f(x_even) + f(xn)]"
        panel_formula = "Panel = h/3 [f(x_i) + 4f(x_(i+1)) + f(x_(i+2))]"
        validity = "n must be even; simple form uses n = 2."
    else:
        global_formula = "S = 3h/8 [f(x0) + 3Σ f(x_i, i not multiple of 3) + 2Σ f(x_i, i multiple of 3) + f(xn)]"
        panel_formula = "Panel = 3h/8 [f(x_i) + 3f(x_(i+1)) + 3f(x_(i+2)) + f(x_(i+3))]"
        validity = "n must be a multiple of 3; simple form uses n = 3."

    return pd.DataFrame(
        {
            "Item": [
                "Selected Rule",
                "Step Size",
                "Composite Formula",
                "Panel Formula",
                "Subinterval Requirement",
                "Global Error Order",
            ],
            "Formula / Description": [
                result.rule_name,
                "h = (b - a) / n",
                global_formula,
                panel_formula,
                validity,
                "O(h^4) for sufficiently smooth integrands",
            ],
        }
    )


# =============================================================================
# Scientific plots
# =============================================================================
def lagrange_panel_values(
    x_nodes: np.ndarray,
    y_nodes: np.ndarray,
    evaluation_points: np.ndarray,
) -> np.ndarray:
    """Evaluate one local interpolation polynomial using Lagrange form."""

    values = np.zeros_like(evaluation_points, dtype=float)
    for i in range(len(x_nodes)):
        basis = np.ones_like(evaluation_points, dtype=float)
        for j in range(len(x_nodes)):
            if i == j:
                continue
            denominator = x_nodes[i] - x_nodes[j]
            if abs(denominator) <= ZERO_TOLERANCE:
                raise ValueError("Duplicate panel nodes prevented interpolation plotting.")
            basis *= (evaluation_points - x_nodes[j]) / denominator
        values += y_nodes[i] * basis
    return values


def create_simpson_plot(result: SimpsonsResult) -> Figure:
    """Create a function graph with each Simpson panel drawn explicitly."""

    if not result.success or result.function_expression is None:
        raise ValueError("A successful result is required for plotting.")

    x_symbol = sp.Symbol("x", real=True)
    numeric_function = build_numeric_function(x_symbol, result.function_expression)
    lower_limit = float(result.lower_limit)
    upper_limit = float(result.upper_limit)
    plotting_min = min(lower_limit, upper_limit)
    plotting_max = max(lower_limit, upper_limit)
    span = plotting_max - plotting_min
    margin = 0.08 * span if span > 0 else 1.0

    dense_x = np.linspace(plotting_min - margin, plotting_max + margin, 1400)
    dense_y = evaluate_real_array(numeric_function, dense_x)
    finite_mask = np.isfinite(dense_y)
    if np.count_nonzero(finite_mask) < 2:
        raise ValueError("The function has insufficient finite values for plotting.")

    figure, axis = plt.subplots(figsize=(11, 6.5))
    axis.plot(dense_x[finite_mask], dense_y[finite_mask], linewidth=2.2, label="f(x)")
    axis.axhline(0.0, linewidth=1.0)

    for panel in result.panel_contributions:
        panel_x = np.array(panel.x_values, dtype=float)
        panel_y = np.array(panel.function_values, dtype=float)
        curve_x = np.linspace(panel_x[0], panel_x[-1], 120)
        curve_y = lagrange_panel_values(panel_x, panel_y, curve_x)
        axis.fill_between(curve_x, 0.0, curve_y, alpha=0.18)
        axis.plot(curve_x, curve_y, linewidth=1.35)

    node_x = np.array([sample.x_value for sample in result.function_samples], dtype=float)
    node_y = np.array([sample.function_value for sample in result.function_samples], dtype=float)
    axis.scatter(node_x, node_y, s=38, zorder=5, label="Integration Nodes")
    axis.axvline(lower_limit, linestyle="--", linewidth=1.0, label=f"a = {lower_limit:.6g}")
    axis.axvline(upper_limit, linestyle="--", linewidth=1.0, label=f"b = {upper_limit:.6g}")

    local_degree = 2 if result.panel_size == 2 else 3
    axis.set_title(
        f"{result.rule_name}: {len(result.panel_contributions)} "
        f"Degree-{local_degree} Panel(s)"
    )
    axis.set_xlabel("x")
    axis.set_ylabel("f(x)")
    axis.grid(True, alpha=0.28)
    axis.legend()
    figure.tight_layout()
    return figure


def create_convergence_plot(result: SimpsonsResult) -> Figure:
    """Create a log-log convergence visualization."""

    if len(result.convergence_records) < 2:
        raise ValueError("At least two refinement levels are required for plotting.")

    step_sizes = np.array(
        [abs(record.step_size) for record in result.convergence_records],
        dtype=float,
    )

    if result.exact_integral is not None:
        errors = np.array(
            [
                record.absolute_error if record.absolute_error is not None else np.nan
                for record in result.convergence_records
            ],
            dtype=float,
        )
        y_label = "Absolute Error"
    else:
        errors = np.array(
            [
                record.successive_difference
                if record.successive_difference is not None
                else np.nan
                for record in result.convergence_records
            ],
            dtype=float,
        )
        y_label = "Successive Difference"

    valid = (
        np.isfinite(step_sizes)
        & np.isfinite(errors)
        & (step_sizes > 0.0)
        & (errors > 0.0)
    )
    if np.count_nonzero(valid) < 2:
        raise ValueError("There are not enough positive finite errors for plotting.")

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.loglog(step_sizes[valid], errors[valid], marker="o", linewidth=2.0)

    reference_h = step_sizes[valid]
    reference_error = errors[valid][0] * (reference_h / reference_h[0]) ** 4
    axis.loglog(
        reference_h,
        reference_error,
        linestyle="--",
        linewidth=1.3,
        label="Reference slope 4",
    )
    axis.invert_xaxis()
    axis.set_title("Simpson Convergence Analysis")
    axis.set_xlabel("Step Size |h|")
    axis.set_ylabel(y_label)
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


# =============================================================================
# Excel report
# =============================================================================
def apply_excel_style(workbook: Any) -> None:
    """Apply consistent professional formatting to every worksheet."""

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="16324F")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            maximum_length = 0
            for cell in column_cells:
                cell_text = "" if cell.value is None else str(cell.value)
                maximum_length = max(maximum_length, len(cell_text))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(maximum_length + 2, 12),
                65,
            )


def create_excel_report(result: SimpsonsResult) -> bytes:
    """Create a formatted XLSX report entirely in memory."""

    if not result.success:
        raise ValueError("Only a successful Simpson result can be exported.")

    summary_df = create_summary_dataframe(result)
    formula_df = create_formula_dataframe(result)
    function_df = create_function_values_dataframe(result)
    panel_df = create_panel_dataframe(result)
    error_df = create_error_dataframe(result)
    convergence_df = create_convergence_dataframe(result)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        formula_df.to_excel(writer, sheet_name="Method Formulas", index=False)
        function_df.to_excel(writer, sheet_name="Function Values", index=False)
        panel_df.to_excel(writer, sheet_name="Panel Contributions", index=False)
        error_df.to_excel(writer, sheet_name="Error Analysis", index=False)
        convergence_df.to_excel(writer, sheet_name="Convergence Analysis", index=False)

        workbook = writer.book
        apply_excel_style(workbook)

        if not convergence_df.empty:
            worksheet = workbook["Convergence Analysis"]
            row_count = len(convergence_df) + 1
            chart = LineChart()
            chart.title = "Simpson Approximation by Mesh Refinement"
            chart.y_axis.title = "Approximate Integral"
            chart.x_axis.title = "Subintervals n"
            data = Reference(worksheet, min_col=4, min_row=1, max_row=row_count)
            categories = Reference(worksheet, min_col=2, min_row=2, max_row=row_count)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 15
            worksheet.add_chart(chart, "K2")

            if result.exact_integral is not None:
                error_chart = LineChart()
                error_chart.title = "Absolute Error by Mesh Refinement"
                error_chart.y_axis.title = "Absolute Error"
                error_chart.x_axis.title = "Subintervals n"
                error_data = Reference(
                    worksheet,
                    min_col=6,
                    min_row=1,
                    max_row=row_count,
                )
                error_chart.add_data(error_data, titles_from_data=True)
                error_chart.set_categories(categories)
                error_chart.height = 8
                error_chart.width = 15
                worksheet.add_chart(error_chart, "K20")

    output.seek(0)
    return finalize_excel_report_with_visible_charts(output.getvalue())


# =============================================================================
# Streamlit renderers
# =============================================================================

def finalize_excel_report_with_visible_charts(report_bytes: bytes) -> bytes:
    """Place existing workbook charts on Summary so they are immediately visible.

    The report data and worksheets are preserved. If Excel chart post-processing is
    unavailable for any reason, the original report is returned unchanged rather
    than risking a damaged workbook.
    """
    try:
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_workbook

        workbook = _load_workbook(_BytesIO(report_bytes))
        if "Summary" not in workbook.sheetnames:
            return report_bytes

        summary_sheet = workbook["Summary"]
        collected_charts = list(summary_sheet._charts)
        summary_sheet._charts = []

        for worksheet in workbook.worksheets:
            if worksheet is summary_sheet:
                continue
            if worksheet._charts:
                collected_charts.extend(list(worksheet._charts))
                worksheet._charts = []

        # Put every chart in the existing Summary worksheet, below one another.
        # D2 keeps the first chart visible beside the main results when the report opens.
        for chart_index, chart in enumerate(collected_charts):
            anchor_row = 2 + chart_index * 19
            summary_sheet.add_chart(chart, f"D{anchor_row}")

        workbook.active = workbook.sheetnames.index("Summary")
        output = _BytesIO()
        workbook.save(output)
        return output.getvalue()
    except Exception:
        return report_bytes


def render_final_result(result: SimpsonsResult) -> None:
    """Render the final result card."""

    if not result.success:
        st.error(result.message)
        st.caption(result.stopping_reason)
        return

    st.success(result.message)
    st.markdown(f"**Rule:** {result.rule_name}")
    st.markdown(f"**Function:** `{result.function_text}`")

    metric_columns = st.columns(2)
    metric_columns[0].metric(
        "Approximate Integral",
        format_number(result.approximate_integral),
    )
    metric_columns[1].metric(
        "Subintervals",
        str(result.subintervals),
    )

    second_row = st.columns(2)
    second_row[0].metric("Step Size h", format_number(result.step_size))
    second_row[1].metric("Panels", str(len(result.panel_contributions)))

    if result.exact_integral is not None:
        st.metric("Exact Integral", format_number(result.exact_integral))
        st.metric("Absolute Error", scientific_number(result.absolute_error))
    else:
        st.info("A closed-form exact integral was not available for comparison.")

    st.markdown(f"**Status:** {result.status.title()}")
    st.markdown(f"**Stopping Reason:** {result.stopping_reason}")

    for warning in result.warnings:
        st.warning(warning)


def render_formula_summary(result: SimpsonsResult) -> None:
    """Render the selected formula and numerical substitution."""

    st.subheader("Method Formula and Final Substitution")
    st.dataframe(
        create_formula_dataframe(result),
        use_container_width=True,
        hide_index=True,
    )

    if result.rule_key.startswith("simpson_13"):
        st.latex(
            r"S_n=\frac{h}{3}\left[f(x_0)+4\sum_{i\,\mathrm{odd}}f(x_i)"
            r"+2\sum_{i\,\mathrm{even},\,i\ne 0,n}f(x_i)+f(x_n)\right]"
        )
        factor_text = "h / 3"
    else:
        st.latex(
            r"S_n=\frac{3h}{8}\left[f(x_0)+3\sum_{i\not\equiv0\,(3)}f(x_i)"
            r"+2\sum_{i\equiv0\,(3),\,i\ne0,n}f(x_i)+f(x_n)\right]"
        )
        factor_text = "3h / 8"

    st.code(
        f"h = ({result.upper_limit:.12g} - {result.lower_limit:.12g}) / "
        f"{result.subintervals} = {result.step_size:.12g}\n"
        f"Weighted sum = {result.weighted_sum:.12g}\n"
        f"Integral = ({factor_text}) × weighted sum = "
        f"{result.approximate_integral:.12g}",
        language=None,
    )


def render_function_values(result: SimpsonsResult) -> None:
    """Render all nodes and weights."""

    st.subheader("Function Values and Simpson Weights")
    st.dataframe(
        round_numeric_dataframe(create_function_values_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )


def render_panel_steps(result: SimpsonsResult) -> None:
    """Render every panel contribution and its substitution."""

    st.subheader("Panel Contributions")
    panel_df = create_panel_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(panel_df.drop(columns=["Operation"])),
        use_container_width=True,
        hide_index=True,
    )

    for panel in result.panel_contributions:
        with st.expander(
            f"Panel {panel.panel_index}: nodes {panel.start_node} to {panel.end_node}"
        ):
            st.code(panel.operation, language=None)
            local_df = pd.DataFrame(
                {
                    "Node": panel.node_indices,
                    "x": panel.x_values,
                    "f(x)": panel.function_values,
                    "Local Weight": panel.local_weights,
                    "Weighted Term": [
                        weight * value
                        for weight, value in zip(
                            panel.local_weights,
                            panel.function_values,
                        )
                    ],
                }
            )
            st.dataframe(
                round_numeric_dataframe(local_df),
                use_container_width=True,
                hide_index=True,
            )
            st.markdown(
                f"**Panel contribution:** {format_number(panel.signed_contribution)}  "
                f"\n**Cumulative integral:** {format_number(panel.cumulative_integral)}"
            )


def render_error_analysis(result: SimpsonsResult) -> None:
    """Render error metrics."""

    st.subheader("Error Analysis")
    st.dataframe(
        round_numeric_dataframe(create_error_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    if result.exact_integral is None:
        st.info(
            "Absolute and relative true errors require a closed-form symbolic "
            "reference integral. The convergence table still reports successive differences."
        )
    else:
        st.markdown(
            f"- **Absolute Error:** {scientific_number(result.absolute_error)}\n"
            f"- **Relative Error:** {scientific_number(result.relative_error)}\n"
            f"- **Percentage Error:** {format_number(result.percentage_error)}%"
        )


def render_convergence_analysis(result: SimpsonsResult) -> None:
    """Render the mesh-refinement table and graph."""

    st.subheader("Convergence Analysis")
    convergence_df = create_convergence_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(convergence_df),
        use_container_width=True,
        hide_index=True,
    )
    st.info(
        "For a sufficiently smooth integrand, both composite Simpson's 1/3 and "
        "3/8 Rules have global truncation error of order O(h⁴). When h is halved, "
        "the asymptotic error is therefore expected to decrease by approximately 16."
    )

    try:
        figure = create_convergence_plot(result)
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)
    except ValueError as error:
        st.warning(f"The convergence graph could not be displayed: {error}")


def render_function_graph(result: SimpsonsResult) -> None:
    """Render the integrand and local Simpson panels."""

    st.subheader("Function Graph and Simpson Panels")
    try:
        figure = create_simpson_plot(result)
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)
    except ValueError as error:
        st.warning(f"The function graph could not be displayed: {error}")


def render_excel_download(result: SimpsonsResult) -> None:
    """Build and render the Excel download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature
    cached_signature = st.session_state.get("simpsons_excel_signature")

    if cached_signature != report_signature:
        try:
            st.session_state.simpsons_excel_report = create_excel_report(result)
            st.session_state.simpsons_excel_signature = report_signature
        except (ValueError, OSError, TypeError, ArithmeticError) as error:
            st.error(f"The Excel report could not be generated: {error}")
            return

    report_bytes = st.session_state.get("simpsons_excel_report")
    if not report_bytes:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    filename = f"simpsons_rules_report_{timestamp}.xlsx"
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=filename,
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="simpsons_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Simpson's Rules Streamlit page."""

    st.set_page_config(
        page_title="Simpson's Rules Solver | Numerical Methods",
        page_icon="∫",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">NUMERICAL INTEGRATION TOOL</div>
                <h1>Simpson's Rules Solver</h1>
                <p>
                    Approximate a definite integral using Simpson's 1/3 or
                    Simpson's 3/8 Rule in simple or composite form. Review every
                    node, textbook weight, panel contribution, error metric,
                    convergence result, scientific graph, and Excel report.
                </p>

                <div class="method-actions">
                    <a href="/Simpsons_Rule" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Simpsons_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Match the centered Bisection solver content width.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )
    
        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader('How to Write the Function')
                st.markdown(
                    """
                Enter only the mathematical expression, without `f(x) =` or an equals sign.

                - Use only **x** as the variable.
                - Powers: write `x**2`, not **x^2**.
                - Multiplication: write `2*x`, not `2x`.
                - Use lowercase functions such as **sin(x)**, **cos(x)**, **exp(x)**, **sqrt(x)**, and **log(x)**.
                - Use parentheses whenever the order of operations could be unclear.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader('Before Solving')
                st.markdown(
                    """
                - The function should be continuous and finite throughout the interval.
                - Simpson’s 1/3 Rule requires an **even** number of subintervals.
                - Simpson’s 3/8 Rule requires the number of subintervals to be a **multiple of 3**.
                - Simple rules use their fixed number of subintervals; composite rules repeat the formula over the interval.
                - Reversing the limits correctly changes the sign of the integral.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])
    
        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )
    
                st.markdown(
                    '<div class="input-label-ui">Function f(x)</div>',
                    unsafe_allow_html=True,
                )
                function_text = st.text_input(
                    "Function f(x)",
                    value=DEFAULT_FUNCTION,
                    placeholder="Example: sin(x), exp(x), or x**4 - 2*x + 1",
                    label_visibility="collapsed",
                    key="simpsons_function",
                )
    
                rule_name = st.selectbox(
                    "Rule Type",
                    options=list(RULE_OPTIONS.keys()),
                    index=0,
                    key="simpsons_rule_type",
                )
                rule_key = RULE_OPTIONS[rule_name]
                _panel_size, _factor, forced_n, _requirement = rule_configuration(rule_key)
    
                limit_columns = st.columns(2)
                with limit_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Lower limit a</div>',
                        unsafe_allow_html=True,
                    )
                    lower_limit = st.number_input(
                        "Lower limit a",
                        value=DEFAULT_LOWER_LIMIT,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="simpsons_lower_limit",
                    )
    
                with limit_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Upper limit b</div>',
                        unsafe_allow_html=True,
                    )
                    upper_limit = st.number_input(
                        "Upper limit b",
                        value=DEFAULT_UPPER_LIMIT,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="simpsons_upper_limit",
                    )
    
                settings_columns = st.columns(2)
                with settings_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Subintervals n</div>',
                        unsafe_allow_html=True,
                    )
                    default_n = forced_n if forced_n else DEFAULT_SUBINTERVALS
                    subintervals = st.number_input(
                        "Subintervals n",
                        min_value=MIN_SUBINTERVALS,
                        max_value=MAX_SUBINTERVALS,
                        value=default_n,
                        step=1,
                        disabled=bool(forced_n),
                        label_visibility="collapsed",
                        key=f"simpsons_subintervals_{rule_key}",
                    )
    
                with settings_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Convergence levels</div>',
                        unsafe_allow_html=True,
                    )
                    convergence_levels = st.number_input(
                        "Convergence levels",
                        min_value=MIN_CONVERGENCE_LEVELS,
                        max_value=MAX_CONVERGENCE_LEVELS,
                        value=DEFAULT_CONVERGENCE_LEVELS,
                        step=1,
                        label_visibility="collapsed",
                        key="simpsons_convergence_levels",
                    )
    
                if rule_key.startswith("simpson_13"):
                    st.caption("Simpson's 1/3 Rule requires an even number of subintervals.")
                else:
                    st.caption("Simpson's 3/8 Rule requires n to be a multiple of 3.")
    
                current_signature = create_input_signature(
                    function_text,
                    lower_limit,
                    upper_limit,
                    rule_name,
                    subintervals,
                    convergence_levels,
                )
    
                solve_clicked = st.button(
                    "Calculate Integral",
                    type="primary",
                    use_container_width=True,
                    key="simpsons_solve_button",
                )
    
                if solve_clicked:
                    result = solve_simpsons_rules(
                        function_text=function_text,
                        lower_limit_input=lower_limit,
                        upper_limit_input=upper_limit,
                        rule_name=rule_name,
                        subintervals_input=subintervals,
                        convergence_levels_input=convergence_levels,
                    )
                    st.session_state.simpsons_result = result
                    st.session_state.simpsons_result_signature = current_signature
                    st.session_state.pop("simpsons_excel_report", None)
                    st.session_state.pop("simpsons_excel_signature", None)
    
                with st.expander("Example Inputs"):
                    st.code(
                        "Function: sin(x)\n"
                        "a = 0\n"
                        "b = pi ≈ 3.141592653589793\n"
                        "Composite Simpson's 1/3: n = 6\n"
                        "Expected exact integral: 2",
                        language=None,
                    )
                    st.code(
                        "Function: x**4 - 2*x + 1\n"
                        "a = -1\n"
                        "b = 2\n"
                        "Composite Simpson's 3/8: n = 6",
                        language=None,
                    )
    
        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )
    
                saved_result = st.session_state.get("simpsons_result")
                saved_signature = st.session_state.get("simpsons_result_signature")
    
                if saved_result is None:
                    st.info("Enter the integration data and select Calculate Integral.")
                elif saved_signature != current_signature:
                    st.warning(
                        "The inputs have changed. Select Calculate Integral to update the result."
                    )
                else:
                    render_final_result(saved_result)
    
        saved_result = st.session_state.get("simpsons_result")
        saved_signature = st.session_state.get("simpsons_result_signature")
    
        if saved_result is not None and saved_signature == current_signature:
            if saved_result.success:
                st.divider()
                render_formula_summary(saved_result)
                st.divider()
                render_function_values(saved_result)
                st.divider()
                render_panel_steps(saved_result)
                st.divider()
                render_function_graph(saved_result)
                st.divider()
                render_error_analysis(saved_result)
                st.divider()
                render_convergence_analysis(saved_result)
                st.divider()
                render_excel_download(saved_result)
    
        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Simpson's Rules • Numerical Integration</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()
