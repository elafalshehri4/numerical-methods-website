from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Sequence
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from components.navigation import navbar
from utilities.ui import load_css




# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Show three decimals and use × 10ⁿ when fixed notation loses meaning."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"

# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Lagrange Polynomial Interpolation"
SUPPORTED_POINT_COUNTS = tuple(range(2, 13))
DEFAULT_POINT_COUNT = 4
DISPLAY_DECIMALS = 3
DUPLICATE_X_TOLERANCE = 1.0e-12
NODE_RESIDUAL_TOLERANCE = 1.0e-10
CONDITION_NUMBER_WARNING = 1.0e12
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DEFAULT_X_VALUES = np.array(
    [-1.0, 0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0],
    dtype=float,
)
DEFAULT_Y_VALUES = np.array(
    [2.0, 1.0, 2.0, 5.0, 10.0, 17.0, 26.0, 37.0, 50.0, 65.0, 82.0, 101.0],
    dtype=float,
)


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class LagrangeBasisStep:
    """Complete information for one Lagrange basis polynomial."""

    basis_number: int
    node_x: float
    node_y: float
    denominator: float
    numerator_at_target: float
    basis_value_at_target: float
    contribution_at_target: float
    factor_text: str
    basis_coefficients: np.ndarray
    contribution_coefficients: np.ndarray


@dataclass(frozen=True)
class InterpolationResult:
    """Complete result shared by Streamlit renderers and Excel export."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str
    point_count: int
    polynomial_degree: int
    original_x: np.ndarray
    original_y: np.ndarray
    target_x: float
    coefficients: np.ndarray | None
    equation_text: str
    interpolated_value: float | None
    basis_history: tuple[LagrangeBasisStep, ...]
    node_predictions: np.ndarray | None
    node_residuals: np.ndarray | None
    node_absolute_residuals: np.ndarray | None
    node_residual_norm: float | None
    maximum_node_residual: float | None
    vandermonde_condition_number: float | None
    exact_function_text: str
    exact_expression_text: str
    exact_value_at_target: float | None
    absolute_error_at_target: float | None
    relative_error_at_target: float | None
    warnings: tuple[str, ...]
    duplicate_x_tolerance: float
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def format_number(
    value: float | int | None,
    decimals: int = 3,
) -> str:
    """Format displayed values with three decimals and × 10ⁿ notation."""

    return format_display_number(value, decimals)



def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns for display without changing stored precision."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def create_input_signature(
    data: pd.DataFrame,
    target_x: Any,
    exact_function_text: str,
) -> str:
    """Create a stable signature used to prevent stale Streamlit results."""

    serialized_rows = []
    for _, row in data.iterrows():
        serialized_rows.append((str(row.get("x", "")), str(row.get("y", ""))))
    payload = repr(
        (
            serialized_rows,
            str(target_x),
            str(exact_function_text).strip(),
            DUPLICATE_X_TOLERANCE,
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def serialize_vector(vector: np.ndarray) -> str:
    """Serialize a vector clearly for Excel cells."""

    return np.array2string(
        np.asarray(vector, dtype=float).reshape(-1),
        precision=12,
        suppress_small=False,
        separator=", ",
        max_line_width=1000,
    )


def safe_numeric_preview(values: Sequence[Any]) -> np.ndarray:
    """Return finite values that can safely be shown in an error result."""

    preview: list[float] = []
    for raw_value in values:
        if raw_value is None:
            continue
        if isinstance(raw_value, str) and not raw_value.strip():
            continue
        try:
            numeric_value = float(raw_value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(numeric_value):
            preview.append(numeric_value)
    return np.asarray(preview, dtype=float)


def default_points_dataframe(point_count: int) -> pd.DataFrame:
    """Return default editable interpolation points."""

    return pd.DataFrame(
        {
            "x": DEFAULT_X_VALUES[:point_count],
            "y": DEFAULT_Y_VALUES[:point_count],
        }
    )


# =============================================================================
# Input validation
# =============================================================================
def coerce_real_finite_vector(
    values: Sequence[Any],
    vector_name: str,
) -> np.ndarray:
    """Convert user values to a finite real one-dimensional NumPy vector."""

    converted: list[float] = []
    for index, raw_value in enumerate(values, start=1):
        if raw_value is None:
            raise ValueError(f"{vector_name} contains an empty value at row {index}.")
        if isinstance(raw_value, str) and not raw_value.strip():
            raise ValueError(f"{vector_name} contains an empty value at row {index}.")
        try:
            numeric_value = float(raw_value)
        except (TypeError, ValueError) as error:
            raise ValueError(
                f"{vector_name} contains a non-numeric value at row {index}."
            ) from error
        if not math.isfinite(numeric_value):
            raise ValueError(
                f"{vector_name} contains NaN or infinity at row {index}."
            )
        converted.append(numeric_value)

    vector = np.asarray(converted, dtype=float).reshape(-1)
    if vector.size == 0:
        raise ValueError(f"{vector_name} must contain at least one value.")
    return vector


def validate_distinct_x_values(x_values: np.ndarray) -> None:
    """Reject duplicate or numerically indistinguishable interpolation nodes."""

    for first_index in range(len(x_values)):
        for second_index in range(first_index + 1, len(x_values)):
            first_x = float(x_values[first_index])
            second_x = float(x_values[second_index])
            scale = max(1.0, abs(first_x), abs(second_x))
            if abs(first_x - second_x) <= DUPLICATE_X_TOLERANCE * scale:
                raise ValueError(
                    "Interpolation requires distinct x-values. "
                    f"Rows {first_index + 1} and {second_index + 1} contain "
                    "duplicate or numerically indistinguishable x-values."
                )


def validate_and_prepare_points(
    raw_data: pd.DataFrame,
) -> tuple[np.ndarray, np.ndarray]:
    """Validate the editable data table and return x and y vectors."""

    if not isinstance(raw_data, pd.DataFrame):
        raise ValueError("Data points must be supplied in a table.")
    if "x" not in raw_data.columns or "y" not in raw_data.columns:
        raise ValueError("The data table must contain x and y columns.")
    if len(raw_data) not in SUPPORTED_POINT_COUNTS:
        raise ValueError(
            "The number of interpolation points must be between "
            f"{min(SUPPORTED_POINT_COUNTS)} and {max(SUPPORTED_POINT_COUNTS)}."
        )

    x_values = coerce_real_finite_vector(raw_data["x"].tolist(), "x-values")
    y_values = coerce_real_finite_vector(raw_data["y"].tolist(), "y-values")

    if x_values.size != y_values.size:
        raise ValueError("The x and y vectors must contain the same number of values.")

    validate_distinct_x_values(x_values)
    return x_values, y_values


def validate_target_x(raw_target_x: Any) -> float:
    """Convert and validate the requested interpolation coordinate."""

    if raw_target_x is None:
        raise ValueError("The target x-value is required.")
    if isinstance(raw_target_x, str) and not raw_target_x.strip():
        raise ValueError("The target x-value is required.")
    try:
        target_x = float(raw_target_x)
    except (TypeError, ValueError) as error:
        raise ValueError("The target x-value must be numeric.") from error
    if not math.isfinite(target_x):
        raise ValueError("The target x-value must be finite.")
    return target_x


# =============================================================================
# Optional exact-function parsing
# =============================================================================
def parse_exact_function(
    function_text: str,
) -> tuple[sp.Expr | None, Callable[..., Any] | None]:
    """Parse an optional exact function of x for comparison only."""

    cleaned_text = str(function_text).strip()
    if not cleaned_text:
        return None, None
    if "=" in cleaned_text:
        raise ValueError(
            "Enter the exact function as an expression only, for example "
            "sin(x) + x**2, without 'f(x)=' or an equals sign."
        )

    x_symbol = sp.Symbol("x", real=True)
    allowed_locals = {
        "x": x_symbol,
        "sin": sp.sin,
        "cos": sp.cos,
        "tan": sp.tan,
        "asin": sp.asin,
        "acos": sp.acos,
        "atan": sp.atan,
        "sinh": sp.sinh,
        "cosh": sp.cosh,
        "tanh": sp.tanh,
        "exp": sp.exp,
        "log": sp.log,
        "ln": sp.log,
        "sqrt": sp.sqrt,
        "abs": sp.Abs,
        "pi": sp.pi,
        "E": sp.E,
    }

    normalized_text = cleaned_text.replace("^", "**")
    try:
        expression = sp.sympify(normalized_text, locals=allowed_locals)
    except Exception as error:
        raise ValueError(
            "The exact function has an invalid format. Use standard Python/SymPy "
            "syntax, such as sin(x), exp(x), or x**3 - 2*x + 1."
        ) from error

    unexpected_symbols = expression.free_symbols.difference({x_symbol})
    if unexpected_symbols:
        symbol_names = ", ".join(sorted(str(symbol) for symbol in unexpected_symbols))
        raise ValueError(
            "The exact function may contain only the variable x. "
            f"Unexpected symbol(s): {symbol_names}."
        )
    if expression.has(sp.I, sp.zoo, sp.nan, sp.oo, -sp.oo):
        raise ValueError("The exact function contains a non-real or undefined value.")

    try:
        numeric_function = sp.lambdify(x_symbol, expression, modules="numpy")
    except Exception as error:
        raise ValueError(
            "The exact function could not be converted to a numerical function."
        ) from error

    return expression, numeric_function


def evaluate_exact_function_scalar(
    numeric_function: Callable[..., Any],
    x_value: float,
) -> float:
    """Evaluate the optional exact function safely at one real point."""

    try:
        raw_value = numeric_function(float(x_value))
        array = np.asarray(raw_value)
        if np.iscomplexobj(array):
            raise ValueError("The exact function produced a complex value.")
        value = float(np.asarray(array, dtype=float).reshape(-1)[0])
    except Exception as error:
        raise ValueError(
            f"The exact function could not be evaluated at x = {x_value}. "
            f"Reason: {error}"
        ) from error
    if not math.isfinite(value):
        raise ValueError(
            f"The exact function is undefined or non-finite at x = {x_value}."
        )
    return value


def evaluate_exact_function_array(
    numeric_function: Callable[..., Any],
    x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate an exact function on a grid, preserving invalid points as NaN."""

    with np.errstate(all="ignore"):
        raw_values = numeric_function(x_values)
    array = np.asarray(raw_values)
    if np.iscomplexobj(array):
        complex_values = np.asarray(array, dtype=complex)
        real_mask = np.abs(np.imag(complex_values)) <= 1.0e-12
        array = np.where(real_mask, np.real(complex_values), np.nan)
    array = np.asarray(array, dtype=float)
    if array.ndim == 0:
        array = np.full_like(x_values, float(array), dtype=float)
    else:
        try:
            array = np.broadcast_to(array, x_values.shape).astype(float, copy=True)
        except ValueError as error:
            raise ValueError(
                "The exact function returned values with an unexpected shape."
            ) from error
    array[~np.isfinite(array)] = np.nan
    return array


# =============================================================================
# Polynomial arithmetic and Lagrange interpolation
# =============================================================================
def multiply_polynomials_ascending(
    first: np.ndarray,
    second: np.ndarray,
) -> np.ndarray:
    """Multiply two polynomials stored in ascending-power coefficient order."""

    first_coefficients = np.asarray(first, dtype=float).reshape(-1)
    second_coefficients = np.asarray(second, dtype=float).reshape(-1)
    result = np.zeros(
        first_coefficients.size + second_coefficients.size - 1,
        dtype=float,
    )
    for first_power, first_value in enumerate(first_coefficients):
        for second_power, second_value in enumerate(second_coefficients):
            result[first_power + second_power] += first_value * second_value
    return result


def evaluate_polynomial(
    coefficients: np.ndarray,
    x_values: float | np.ndarray,
) -> float | np.ndarray:
    """Evaluate ascending-power coefficients with Horner's method."""

    coefficient_vector = np.asarray(coefficients, dtype=float).reshape(-1)
    x_array = np.asarray(x_values, dtype=float)
    result = np.zeros_like(x_array, dtype=float)
    for coefficient in coefficient_vector[::-1]:
        result = result * x_array + coefficient
    if np.ndim(x_values) == 0:
        return float(np.asarray(result).reshape(-1)[0])
    return result


def build_lagrange_polynomial(
    x_values: np.ndarray,
    y_values: np.ndarray,
    target_x: float,
) -> tuple[np.ndarray, tuple[LagrangeBasisStep, ...]]:
    """Construct the expanded Lagrange polynomial and basis-step history."""

    point_count = len(x_values)
    polynomial_coefficients = np.zeros(point_count, dtype=float)
    basis_history: list[LagrangeBasisStep] = []

    for basis_index in range(point_count):
        basis_coefficients = np.array([1.0], dtype=float)
        denominator = 1.0
        numerator_at_target = 1.0
        factor_labels: list[str] = []

        for other_index in range(point_count):
            if other_index == basis_index:
                continue
            node_difference = x_values[basis_index] - x_values[other_index]
            scale = max(
                1.0,
                abs(float(x_values[basis_index])),
                abs(float(x_values[other_index])),
            )
            if abs(node_difference) <= DUPLICATE_X_TOLERANCE * scale:
                raise ValueError(
                    "The Lagrange denominator became zero or numerically "
                    "indistinguishable from zero. Check for repeated x-values."
                )

            denominator *= node_difference
            numerator_at_target *= target_x - x_values[other_index]
            basis_coefficients = multiply_polynomials_ascending(
                basis_coefficients,
                np.array([-x_values[other_index], 1.0], dtype=float),
            )
            factor_labels.append(f"(x - {format_number(x_values[other_index])})")

        if not math.isfinite(denominator) or abs(denominator) == 0.0:
            raise ValueError(
                "A non-finite or zero Lagrange denominator was produced. "
                "Check the interpolation nodes."
            )

        basis_coefficients = basis_coefficients / denominator
        basis_value_at_target = numerator_at_target / denominator
        contribution_at_target = y_values[basis_index] * basis_value_at_target
        contribution_coefficients = y_values[basis_index] * basis_coefficients
        polynomial_coefficients += contribution_coefficients

        basis_history.append(
            LagrangeBasisStep(
                basis_number=basis_index + 1,
                node_x=float(x_values[basis_index]),
                node_y=float(y_values[basis_index]),
                denominator=float(denominator),
                numerator_at_target=float(numerator_at_target),
                basis_value_at_target=float(basis_value_at_target),
                contribution_at_target=float(contribution_at_target),
                factor_text=" × ".join(factor_labels),
                basis_coefficients=basis_coefficients.copy(),
                contribution_coefficients=contribution_coefficients.copy(),
            )
        )

    if not np.all(np.isfinite(polynomial_coefficients)):
        raise ValueError(
            "Interpolation produced non-finite polynomial coefficients. "
            "The x-values may be severely ill-conditioned."
        )

    return polynomial_coefficients, tuple(basis_history)


def polynomial_equation_text(coefficients: np.ndarray | None) -> str:
    """Create a readable expanded polynomial equation."""

    if coefficients is None or len(coefficients) == 0:
        return "Not available"

    terms: list[str] = []
    for power in range(len(coefficients) - 1, -1, -1):
        coefficient = float(coefficients[power])
        if abs(coefficient) < 1.0e-14:
            continue

        magnitude = abs(coefficient)
        if power == 0:
            body = f"{magnitude:.12g}"
        elif power == 1:
            body = "x" if math.isclose(magnitude, 1.0) else f"{magnitude:.12g}x"
        else:
            body = (
                f"x^{power}"
                if math.isclose(magnitude, 1.0)
                else f"{magnitude:.12g}x^{power}"
            )

        if not terms:
            terms.append(f"-{body}" if coefficient < 0 else body)
        else:
            terms.append(f"{'-' if coefficient < 0 else '+'} {body}")

    return "P(x) = " + (" ".join(terms) if terms else "0")


def calculate_relative_error(absolute_error: float, exact_value: float) -> float:
    """Calculate a safe relative error, returning NaN when exact value is zero."""

    if abs(exact_value) <= np.finfo(float).eps:
        return float("nan")
    return float(absolute_error / abs(exact_value))


# =============================================================================
# Solver controller
# =============================================================================
def empty_error_result(
    message: str,
    raw_data: pd.DataFrame,
    raw_target_x: Any,
    exact_function_text: str,
    input_signature: str,
) -> InterpolationResult:
    """Create a structured error result without allowing the page to crash."""

    x_preview = safe_numeric_preview(
        raw_data["x"].tolist() if "x" in raw_data.columns else []
    )
    y_preview = safe_numeric_preview(
        raw_data["y"].tolist() if "y" in raw_data.columns else []
    )
    try:
        target_preview = float(raw_target_x)
        if not math.isfinite(target_preview):
            target_preview = 0.0
    except (TypeError, ValueError):
        target_preview = 0.0

    return InterpolationResult(
        status="Error",
        success=False,
        method=METHOD_NAME,
        message=message,
        stopping_reason=message,
        point_count=min(len(x_preview), len(y_preview)),
        polynomial_degree=max(min(len(x_preview), len(y_preview)) - 1, 0),
        original_x=x_preview,
        original_y=y_preview,
        target_x=target_preview,
        coefficients=None,
        equation_text="Not available",
        interpolated_value=None,
        basis_history=tuple(),
        node_predictions=None,
        node_residuals=None,
        node_absolute_residuals=None,
        node_residual_norm=None,
        maximum_node_residual=None,
        vandermonde_condition_number=None,
        exact_function_text=str(exact_function_text).strip(),
        exact_expression_text="",
        exact_value_at_target=None,
        absolute_error_at_target=None,
        relative_error_at_target=None,
        warnings=tuple(),
        duplicate_x_tolerance=DUPLICATE_X_TOLERANCE,
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


def solve_lagrange_interpolation(
    raw_data: pd.DataFrame,
    raw_target_x: Any,
    exact_function_text: str,
    input_signature: str,
) -> InterpolationResult:
    """Validate inputs and solve with standard Lagrange interpolation."""

    try:
        x_values, y_values = validate_and_prepare_points(raw_data)
        target_x = validate_target_x(raw_target_x)
        exact_expression, exact_numeric_function = parse_exact_function(
            exact_function_text
        )

        coefficients, basis_history = build_lagrange_polynomial(
            x_values,
            y_values,
            target_x,
        )
        interpolated_value = float(evaluate_polynomial(coefficients, target_x))
        node_predictions = np.asarray(
            evaluate_polynomial(coefficients, x_values),
            dtype=float,
        )
        node_residuals = node_predictions - y_values
        node_absolute_residuals = np.abs(node_residuals)
        node_residual_norm = float(np.linalg.norm(node_residuals, ord=2))
        maximum_node_residual = float(np.max(node_absolute_residuals))

        vandermonde_matrix = np.vander(
            x_values,
            N=len(x_values),
            increasing=True,
        )
        with np.errstate(all="ignore"):
            vandermonde_condition_number = float(
                np.linalg.cond(vandermonde_matrix)
            )

        warnings: list[str] = []
        if target_x < float(np.min(x_values)) or target_x > float(np.max(x_values)):
            warnings.append(
                "The requested x-value is outside the data interval. The result "
                "is polynomial extrapolation and may be less reliable."
            )
        if (
            not math.isfinite(vandermonde_condition_number)
            or vandermonde_condition_number >= CONDITION_NUMBER_WARNING
        ):
            warnings.append(
                "The interpolation nodes produce an ill-conditioned Vandermonde "
                "matrix. The expanded coefficients may be sensitive to rounding, "
                "even when the Lagrange value is evaluated correctly."
            )
        if maximum_node_residual > NODE_RESIDUAL_TOLERANCE:
            warnings.append(
                "The interpolation polynomial does not reproduce the supplied "
                "nodes within the expected floating-point tolerance."
            )

        exact_value_at_target: float | None = None
        absolute_error_at_target: float | None = None
        relative_error_at_target: float | None = None
        exact_expression_text = ""

        if exact_expression is not None and exact_numeric_function is not None:
            exact_value_at_target = evaluate_exact_function_scalar(
                exact_numeric_function,
                target_x,
            )
            absolute_error_at_target = float(
                abs(interpolated_value - exact_value_at_target)
            )
            relative_error_at_target = calculate_relative_error(
                absolute_error_at_target,
                exact_value_at_target,
            )
            exact_expression_text = str(exact_expression)

        return InterpolationResult(
            status="Success",
            success=True,
            method=METHOD_NAME,
            message="Execution completed successfully.",
            stopping_reason=(
                "The Lagrange polynomial was constructed from all distinct nodes "
                "and evaluated successfully."
            ),
            point_count=len(x_values),
            polynomial_degree=len(x_values) - 1,
            original_x=x_values.copy(),
            original_y=y_values.copy(),
            target_x=target_x,
            coefficients=coefficients.copy(),
            equation_text=polynomial_equation_text(coefficients),
            interpolated_value=interpolated_value,
            basis_history=basis_history,
            node_predictions=node_predictions,
            node_residuals=node_residuals,
            node_absolute_residuals=node_absolute_residuals,
            node_residual_norm=node_residual_norm,
            maximum_node_residual=maximum_node_residual,
            vandermonde_condition_number=vandermonde_condition_number,
            exact_function_text=str(exact_function_text).strip(),
            exact_expression_text=exact_expression_text,
            exact_value_at_target=exact_value_at_target,
            absolute_error_at_target=absolute_error_at_target,
            relative_error_at_target=relative_error_at_target,
            warnings=tuple(warnings),
            duplicate_x_tolerance=DUPLICATE_X_TOLERANCE,
            input_signature=input_signature,
            execution_datetime=current_report_datetime(),
        )

    except ValueError as error:
        return empty_error_result(
            message=str(error),
            raw_data=raw_data,
            raw_target_x=raw_target_x,
            exact_function_text=exact_function_text,
            input_signature=input_signature,
        )
    except (FloatingPointError, OverflowError, ZeroDivisionError) as error:
        return empty_error_result(
            message=f"Numerical calculation failed safely. Reason: {error}",
            raw_data=raw_data,
            raw_target_x=raw_target_x,
            exact_function_text=exact_function_text,
            input_signature=input_signature,
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def input_points_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return the original interpolation points."""

    return pd.DataFrame(
        {
            "Point": np.arange(1, result.point_count + 1),
            "x": result.original_x,
            "y": result.original_y,
        }
    )


def basis_summary_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return one summary row for every Lagrange basis polynomial."""

    rows = []
    for step in result.basis_history:
        rows.append(
            {
                "Basis": f"L{step.basis_number - 1}(x)",
                "Node x_i": step.node_x,
                "Node y_i": step.node_y,
                "Denominator": step.denominator,
                "Numerator at Target": step.numerator_at_target,
                "L_i(Target)": step.basis_value_at_target,
                "y_i L_i(Target)": step.contribution_at_target,
                "Numerator Factors": step.factor_text,
            }
        )
    return pd.DataFrame(rows)


def basis_coefficients_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return expanded basis and contribution coefficients by power."""

    rows = []
    for step in result.basis_history:
        for power in range(result.point_count):
            rows.append(
                {
                    "Basis": f"L{step.basis_number - 1}(x)",
                    "Power": power,
                    "Basis Coefficient": step.basis_coefficients[power],
                    "y_i × Basis Coefficient": (
                        step.contribution_coefficients[power]
                    ),
                }
            )
    return pd.DataFrame(rows)


def polynomial_coefficients_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return the final expanded polynomial coefficients."""

    if result.coefficients is None:
        return pd.DataFrame(columns=["Power", "Term", "Coefficient"])
    return pd.DataFrame(
        {
            "Power": np.arange(len(result.coefficients)),
            "Term": [
                "1" if power == 0 else ("x" if power == 1 else f"x^{power}")
                for power in range(len(result.coefficients))
            ],
            "Coefficient": result.coefficients,
        }
    )


def node_residuals_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return interpolation residuals at the supplied nodes."""

    if (
        result.node_predictions is None
        or result.node_residuals is None
        or result.node_absolute_residuals is None
    ):
        return pd.DataFrame()
    return pd.DataFrame(
        {
            "Point": np.arange(1, result.point_count + 1),
            "x": result.original_x,
            "Observed y": result.original_y,
            "P(x)": result.node_predictions,
            "Residual P(x) - y": result.node_residuals,
            "Absolute Residual": result.node_absolute_residuals,
        }
    )


def evaluation_dataframe(result: InterpolationResult) -> pd.DataFrame:
    """Return the target-point evaluation and optional exact comparison."""

    rows: list[dict[str, Any]] = [
        {"Property": "Target x", "Value": result.target_x},
        {"Property": "Interpolated P(x)", "Value": result.interpolated_value},
    ]
    if result.exact_value_at_target is not None:
        rows.extend(
            [
                {
                    "Property": "Exact f(x)",
                    "Value": result.exact_value_at_target,
                },
                {
                    "Property": "Absolute Error",
                    "Value": result.absolute_error_at_target,
                },
                {
                    "Property": "Relative Error",
                    "Value": result.relative_error_at_target,
                },
            ]
        )
    return pd.DataFrame(rows)


# =============================================================================
# Plot data and figures
# =============================================================================
def create_plot_interval(result: InterpolationResult) -> tuple[float, float]:
    """Create a readable interval containing the nodes and requested target."""

    minimum_value = min(float(np.min(result.original_x)), result.target_x)
    maximum_value = max(float(np.max(result.original_x)), result.target_x)
    interval_width = maximum_value - minimum_value
    if interval_width <= 0.0:
        interval_width = 1.0
    padding = 0.12 * interval_width
    return minimum_value - padding, maximum_value + padding


def build_plot_dataframe(
    result: InterpolationResult,
    sample_count: int = 500,
) -> pd.DataFrame:
    """Build polynomial and optional exact-function values for plots and Excel."""

    if result.coefficients is None:
        return pd.DataFrame()
    x_minimum, x_maximum = create_plot_interval(result)
    grid_x = np.linspace(x_minimum, x_maximum, sample_count)
    polynomial_values = np.asarray(
        evaluate_polynomial(result.coefficients, grid_x),
        dtype=float,
    )
    dataframe = pd.DataFrame({"x": grid_x, "P(x)": polynomial_values})

    if result.exact_expression_text:
        _, exact_numeric_function = parse_exact_function(result.exact_function_text)
        if exact_numeric_function is not None:
            exact_values = evaluate_exact_function_array(exact_numeric_function, grid_x)
            dataframe["Exact f(x)"] = exact_values
            dataframe["Absolute Error"] = np.abs(polynomial_values - exact_values)
    return dataframe


def create_interpolation_figure(result: InterpolationResult) -> plt.Figure:
    """Create the interpolation curve, nodes, target, and optional exact curve."""

    plot_dataframe = build_plot_dataframe(result)
    figure, axis = plt.subplots(figsize=(10, 6))
    axis.axhline(0.0, linewidth=1.0, label="x-axis")
    axis.plot(
        plot_dataframe["x"],
        plot_dataframe["P(x)"],
        linewidth=2.2,
        label=f"Interpolation polynomial (degree {result.polynomial_degree})",
    )
    if "Exact f(x)" in plot_dataframe.columns:
        exact_mask = np.isfinite(plot_dataframe["Exact f(x)"].to_numpy(dtype=float))
        if np.any(exact_mask):
            axis.plot(
                plot_dataframe.loc[exact_mask, "x"],
                plot_dataframe.loc[exact_mask, "Exact f(x)"],
                linestyle="--",
                linewidth=1.8,
                label="Exact function",
            )
    axis.scatter(
        result.original_x,
        result.original_y,
        s=65,
        zorder=4,
        label="Interpolation nodes",
    )
    axis.scatter(
        [result.target_x],
        [result.interpolated_value],
        s=95,
        marker="*",
        zorder=5,
        label=(
            f"P({result.target_x:.6g}) = {result.interpolated_value:.6g}"
        ),
    )
    axis.axvline(
        result.target_x,
        linestyle=":",
        linewidth=1.4,
        label=f"Target x = {result.target_x:.6g}",
    )
    axis.set_title("Lagrange Polynomial Interpolation")
    axis.set_xlabel("x")
    axis.set_ylabel("y")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_error_figure(result: InterpolationResult) -> plt.Figure | None:
    """Create |P(x)-f(x)| when an exact comparison function is available."""

    if not result.exact_expression_text:
        return None
    plot_dataframe = build_plot_dataframe(result)
    if "Absolute Error" not in plot_dataframe.columns:
        return None
    finite_mask = np.isfinite(plot_dataframe["Absolute Error"].to_numpy(dtype=float))
    if not np.any(finite_mask):
        return None

    figure, axis = plt.subplots(figsize=(10, 5.5))
    axis.semilogy(
        plot_dataframe.loc[finite_mask, "x"],
        np.maximum(
            plot_dataframe.loc[finite_mask, "Absolute Error"].to_numpy(dtype=float),
            np.finfo(float).tiny,
        ),
        linewidth=2.0,
    )
    axis.axvline(result.target_x, linestyle=":", linewidth=1.3)
    axis.set_title("Absolute Interpolation Error")
    axis.set_xlabel("x")
    axis.set_ylabel("|P(x) - f(x)| (log scale)")
    axis.grid(True, which="both", alpha=0.3)
    figure.tight_layout()
    return figure


# =============================================================================
# Excel report
# =============================================================================
def style_excel_workbook(workbook: Any) -> None:
    """Apply professional formatting to every worksheet."""

    header_fill = PatternFill("solid", fgColor="0D3151")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions

        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            maximum_length = 0
            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                value_length = len(str(cell.value)) if cell.value is not None else 0
                maximum_length = max(maximum_length, value_length)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"
            worksheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 12),
                55,
            )


def create_excel_report(result: InterpolationResult) -> bytes:
    """Generate a formatted in-memory XLSX interpolation report."""

    if not result.success:
        raise ValueError("Only successful interpolation results can be exported.")

    summary_rows = [
        ("Method", result.method),
        ("Status", result.status),
        ("Message", result.message),
        ("Number of Points", result.point_count),
        ("Polynomial Degree", result.polynomial_degree),
        ("Target x", result.target_x),
        ("Interpolated Value", result.interpolated_value),
        ("Expanded Polynomial", result.equation_text),
        ("Node Residual 2-Norm", result.node_residual_norm),
        ("Maximum Absolute Node Residual", result.maximum_node_residual),
        ("Vandermonde Condition Number", result.vandermonde_condition_number),
        ("Exact Function", result.exact_expression_text or "Not provided"),
        ("Exact Value at Target", result.exact_value_at_target),
        ("Absolute Error at Target", result.absolute_error_at_target),
        ("Relative Error at Target", result.relative_error_at_target),
        ("Warnings", " | ".join(result.warnings) if result.warnings else "None"),
        ("Stopping Reason", result.stopping_reason),
        (
            "Execution Date",
            result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
        ),
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["Property", "Value"])
    points_df = input_points_dataframe(result)
    basis_df = basis_summary_dataframe(result)
    basis_coefficients_df = basis_coefficients_dataframe(result)
    polynomial_coefficients_df = polynomial_coefficients_dataframe(result)
    evaluation_df = evaluation_dataframe(result)
    residuals_df = node_residuals_dataframe(result)
    plot_df = build_plot_dataframe(result)
    error_metrics_df = pd.DataFrame(
        {
            "Metric": [
                "Node Residual 2-Norm",
                "Maximum Absolute Node Residual",
                "Absolute Error at Target",
                "Relative Error at Target",
            ],
            "Value": [
                result.node_residual_norm,
                result.maximum_node_residual,
                result.absolute_error_at_target,
                result.relative_error_at_target,
            ],
        }
    )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        points_df.to_excel(writer, sheet_name="Input Points", index=False)
        basis_df.to_excel(writer, sheet_name="Lagrange Basis", index=False)
        basis_coefficients_df.to_excel(
            writer,
            sheet_name="Basis Coefficients",
            index=False,
        )
        polynomial_coefficients_df.to_excel(
            writer,
            sheet_name="Polynomial Coefficients",
            index=False,
        )
        evaluation_df.to_excel(writer, sheet_name="Evaluation", index=False)
        residuals_df.to_excel(writer, sheet_name="Node Residuals", index=False)
        plot_df.to_excel(writer, sheet_name="Plot Data", index=False)
        error_metrics_df.to_excel(writer, sheet_name="Error Analysis", index=False)

        workbook = writer.book
        style_excel_workbook(workbook)

        plot_sheet = workbook["Plot Data"]
        if len(plot_df) > 1:
            chart = ScatterChart()
            chart.title = "Interpolation Polynomial"
            chart.x_axis.title = "x"
            chart.y_axis.title = "y"
            chart.height = 8
            chart.width = 15

            x_reference = Reference(
                plot_sheet,
                min_col=1,
                min_row=2,
                max_row=len(plot_df) + 1,
            )
            polynomial_reference = Reference(
                plot_sheet,
                min_col=2,
                min_row=2,
                max_row=len(plot_df) + 1,
            )
            polynomial_series = Series(
                polynomial_reference,
                x_reference,
                title="P(x)",
            )
            chart.series.append(polynomial_series)

            if "Exact f(x)" in plot_df.columns:
                exact_reference = Reference(
                    plot_sheet,
                    min_col=3,
                    min_row=2,
                    max_row=len(plot_df) + 1,
                )
                exact_series = Series(
                    exact_reference,
                    x_reference,
                    title="Exact f(x)",
                )
                chart.series.append(exact_series)

            plot_sheet.add_chart(chart, "F2")

            if "Absolute Error" in plot_df.columns:
                error_chart = ScatterChart()
                error_chart.title = "Absolute Interpolation Error"
                error_chart.x_axis.title = "x"
                error_chart.y_axis.title = "Absolute Error"
                error_chart.height = 8
                error_chart.width = 15
                error_reference = Reference(
                    plot_sheet,
                    min_col=4,
                    min_row=2,
                    max_row=len(plot_df) + 1,
                )
                error_series = Series(
                    error_reference,
                    x_reference,
                    title="|P(x)-f(x)|",
                )
                error_chart.series.append(error_series)
                plot_sheet.add_chart(error_chart, "F20")

    output.seek(0)
    return finalize_excel_report_with_visible_charts(output.getvalue())


# =============================================================================
# Streamlit result renderers
# =============================================================================

def finalize_excel_report_with_visible_charts(report_bytes: bytes) -> bytes:
    """Place existing workbook charts on Summary so they are immediately visible.

    The report data and worksheets are preserved. If Excel chart post-processing is
    unavailable for any reason, the original report is returned unchanged rather
    than risking a damaged workbook.
    """
    try:
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_workbook

        workbook = _load_workbook(_BytesIO(report_bytes))
        if "Summary" not in workbook.sheetnames:
            return report_bytes

        summary_sheet = workbook["Summary"]
        collected_charts = list(summary_sheet._charts)
        summary_sheet._charts = []

        for worksheet in workbook.worksheets:
            if worksheet is summary_sheet:
                continue
            if worksheet._charts:
                collected_charts.extend(list(worksheet._charts))
                worksheet._charts = []

        # Put every chart in the existing Summary worksheet, below one another.
        # D2 keeps the first chart visible beside the main results when the report opens.
        for chart_index, chart in enumerate(collected_charts):
            anchor_row = 2 + chart_index * 19
            summary_sheet.add_chart(chart, f"D{anchor_row}")

        workbook.active = workbook.sheetnames.index("Summary")
        output = _BytesIO()
        workbook.save(output)
        return output.getvalue()
    except Exception:
        return report_bytes


def render_final_result(result: InterpolationResult) -> None:
    """Render the compact final-result card."""

    if not result.success:
        st.error(result.message)
        return

    st.success(result.message)
    st.markdown(f"**{result.equation_text}**")

    metric_columns = st.columns(2)
    metric_columns[0].metric("Target x", format_number(result.target_x))
    metric_columns[1].metric(
        "Interpolated Value",
        format_number(result.interpolated_value),
    )

    detail_columns = st.columns(2)
    detail_columns[0].metric("Points", result.point_count)
    detail_columns[1].metric("Polynomial Degree", result.polynomial_degree)

    if result.exact_value_at_target is not None:
        exact_columns = st.columns(2)
        exact_columns[0].metric(
            "Exact Value",
            format_number(result.exact_value_at_target),
        )
        exact_columns[1].metric(
            "Absolute Error",
            format_number(result.absolute_error_at_target),
        )

    for warning in result.warnings:
        st.warning(warning)


def render_input_points(result: InterpolationResult) -> None:
    """Render original nodes and interpolation diagnostics."""

    st.subheader("Original Interpolation Points")
    st.dataframe(
        round_numeric_dataframe(input_points_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    diagnostic_columns = st.columns(3)
    diagnostic_columns[0].metric("Number of Points", result.point_count)
    diagnostic_columns[1].metric("Polynomial Degree", result.polynomial_degree)
    diagnostic_columns[2].metric(
        "Vandermonde Condition Number",
        format_scientific_power(result.vandermonde_condition_number)
        if result.vandermonde_condition_number is not None
        else "—",
    )


def render_lagrange_steps(result: InterpolationResult) -> None:
    """Render the complete Lagrange basis history."""

    st.subheader("Lagrange Basis Calculations")
    st.latex(
        r"P(x)=\sum_{i=0}^{n} y_i L_i(x),\qquad "
        r"L_i(x)=\prod_{j\ne i}\frac{x-x_j}{x_i-x_j}"
    )
    st.dataframe(
        round_numeric_dataframe(basis_summary_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    for step in result.basis_history:
        with st.expander(
            f"Basis L{step.basis_number - 1}(x) from point "
            f"({format_number(step.node_x)}, {format_number(step.node_y)})"
        ):
            st.markdown(f"**Numerator factors:** `{step.factor_text}`")
            st.markdown(
                f"**Denominator:** {format_number(step.denominator)}"
            )
            st.markdown(
                f"**L{step.basis_number - 1}({format_number(result.target_x)}):** "
                f"{format_number(step.basis_value_at_target)}"
            )
            st.markdown(
                f"**Contribution to P({format_number(result.target_x)}):** "
                f"{format_number(step.contribution_at_target)}"
            )

            coefficient_df = pd.DataFrame(
                {
                    "Power": np.arange(result.point_count),
                    "Basis Coefficient": step.basis_coefficients,
                    "Contribution Coefficient": step.contribution_coefficients,
                }
            )
            st.dataframe(
                round_numeric_dataframe(coefficient_df),
                use_container_width=True,
                hide_index=True,
            )


def render_polynomial_and_graph(result: InterpolationResult) -> None:
    """Render final coefficients, evaluation, and interpolation graph."""

    st.subheader("Interpolating Polynomial")
    st.markdown(f"### {result.equation_text}")
    st.dataframe(
        round_numeric_dataframe(polynomial_coefficients_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Target Evaluation")
    st.dataframe(
        round_numeric_dataframe(evaluation_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Function Graph")
    interpolation_figure = create_interpolation_figure(result)
    st.pyplot(interpolation_figure, use_container_width=True)
    plt.close(interpolation_figure)


def render_error_analysis(result: InterpolationResult) -> None:
    """Render node residuals and optional true-error comparison."""

    st.subheader("Interpolation Residual Analysis")
    st.caption(
        "At the supplied nodes, an interpolation polynomial should reproduce "
        "the observed y-values up to floating-point roundoff. These residuals "
        "are not the true interpolation error between the nodes."
    )
    st.dataframe(
        round_numeric_dataframe(node_residuals_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    metric_columns = st.columns(2)
    metric_columns[0].metric(
        "Node Residual 2-Norm",
        format_scientific_power(result.node_residual_norm)
        if result.node_residual_norm is not None
        else "—",
    )
    metric_columns[1].metric(
        "Maximum Absolute Node Residual",
        format_scientific_power(result.maximum_node_residual)
        if result.maximum_node_residual is not None
        else "—",
    )

    if result.exact_value_at_target is None:
        st.info(
            "No exact function was provided. True absolute and relative errors "
            "cannot be calculated from interpolation data alone."
        )
        return

    st.subheader("Exact-Function Error Analysis")
    comparison_columns = st.columns(3)
    comparison_columns[0].metric(
        "Exact f(x)",
        format_number(result.exact_value_at_target),
    )
    comparison_columns[1].metric(
        "Absolute Error",
        format_number(result.absolute_error_at_target),
    )
    relative_error = result.relative_error_at_target
    comparison_columns[2].metric(
        "Relative Error",
        "Undefined (exact value is zero)"
        if relative_error is not None and math.isnan(relative_error)
        else (
            format_scientific_power(relative_error)
            if relative_error is not None
            else "—"
        ),
    )

    error_figure = create_error_figure(result)
    if error_figure is not None:
        st.pyplot(error_figure, use_container_width=True)
        plt.close(error_figure)
    else:
        st.warning(
            "The exact-function error graph could not be created over the "
            "selected interval because the function is not finite there."
        )


def render_excel_download(result: InterpolationResult) -> None:
    """Generate and render the Excel report download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature
    if st.session_state.get("lagrange_interpolation_excel_signature") != report_signature:
        try:
            st.session_state.lagrange_interpolation_excel_report = (
                create_excel_report(result)
            )
            st.session_state.lagrange_interpolation_excel_signature = (
                report_signature
            )
        except (ValueError, OSError, PermissionError) as error:
            st.error(f"Excel report generation failed: {error}")
            return

    report_bytes = st.session_state.get("lagrange_interpolation_excel_report")
    if report_bytes is None:
        st.error("The Excel report is not available.")
        return

    date_text = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=f"lagrange_interpolation_report_{date_text}.xlsx",
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="lagrange_interpolation_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Streamlit solver page."""

    st.set_page_config(
        page_title="Lagrange Interpolation Solver | Numerical Methods",
        page_icon="📈",
        layout="wide",
    )
    load_css()

    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">INTERPOLATION TOOL</div>
                <h1>Lagrange Interpolation Solver</h1>
                <p>
                    Enter distinct data points, choose a target x-value, and
                    follow every Lagrange basis calculation, the expanded
                    polynomial, graphical interpolation, and error analysis.
                </p>

                <div class="method-actions">
                    <a href="/Lagrange_Interpolation" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Polynomial_Interpolation_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Match the centered Bisection solver content width.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )
    
        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader('How to Enter the Data')
                st.markdown(
                    """
                Enter the interpolation points as paired numerical values **(xᵢ, yᵢ)**.

                - Every row must contain one finite **x** value and its corresponding finite **y** value.
                - All **x** values must be distinct.
                - Use at least two points.
                - Enter the target **x** value where the interpolating polynomial should be evaluated.
                - An optional exact function may be supplied only for comparison and error analysis.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader('Before Solving')
                st.markdown(
                    """
                - With **n** distinct points, the interpolating polynomial has degree at most **n − 1**.
                - Duplicate or nearly duplicate **x** values make the formula invalid or ill-conditioned.
                - A target inside the data range is interpolation; a target outside it is extrapolation and may be unreliable.
                - More points do not always improve accuracy when the data are noisy or widely spaced.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])
    
        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )
    
                st.markdown(
                    '<div class="input-label-ui">Number of data points</div>',
                    unsafe_allow_html=True,
                )
                point_count = st.selectbox(
                    "Number of data points",
                    options=SUPPORTED_POINT_COUNTS,
                    index=SUPPORTED_POINT_COUNTS.index(DEFAULT_POINT_COUNT),
                    label_visibility="collapsed",
                    key="lagrange_interpolation_point_count",
                )
    
                st.markdown(
                    '<div class="input-label-ui">Interpolation points</div>',
                    unsafe_allow_html=True,
                )
                data_editor_key = f"lagrange_interpolation_data_editor_{point_count}"
                edited_data = st.data_editor(
                    default_points_dataframe(point_count),
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    column_config={
                        "x": st.column_config.NumberColumn(
                            "x",
                            help="Distinct interpolation-node coordinate",
                            format="%.10g",
                        ),
                        "y": st.column_config.NumberColumn(
                            "y",
                            help="Function or observed value at the node",
                            format="%.10g",
                        ),
                    },
                    key=data_editor_key,
                )
    
                target_columns = st.columns(2)
                with target_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Target x-value</div>',
                        unsafe_allow_html=True,
                    )
                    target_x = st.number_input(
                        "Target x-value",
                        value=1.5,
                        format="%.10g",
                        label_visibility="collapsed",
                        key="lagrange_interpolation_target_x",
                    )
                with target_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Optional exact function</div>',
                        unsafe_allow_html=True,
                    )
                    exact_function_text = st.text_input(
                        "Optional exact function",
                        value="",
                        placeholder="Example: x**2 + 1",
                        label_visibility="collapsed",
                        key="lagrange_interpolation_exact_function",
                    )
    
                st.caption(
                    "All x-values must be distinct. The exact function is optional "
                    "and is used only to calculate true error and draw a comparison."
                )
    
                solve_button_clicked = st.button(
                    "Solve",
                    use_container_width=True,
                    key="lagrange_interpolation_solve_button",
                )
    
        current_input_signature = create_input_signature(
            edited_data,
            target_x,
            exact_function_text,
        )
    
        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )
    
                stored_result = st.session_state.get("lagrange_interpolation_result")
                if stored_result is None:
                    st.info("Enter the data points and click Solve to see the result.")
                elif stored_result.input_signature != current_input_signature:
                    st.info(
                        "The points, target value, or exact function has changed. "
                        "Click Solve to calculate a new result."
                    )
                else:
                    render_final_result(stored_result)
    
        if solve_button_clicked:
            st.session_state.lagrange_interpolation_result = (
                solve_lagrange_interpolation(
                    raw_data=edited_data.copy(),
                    raw_target_x=target_x,
                    exact_function_text=exact_function_text,
                    input_signature=current_input_signature,
                )
            )
            st.session_state.pop("lagrange_interpolation_excel_report", None)
            st.session_state.pop("lagrange_interpolation_excel_signature", None)
            st.rerun()
    
        active_result = st.session_state.get("lagrange_interpolation_result")
        if (
            active_result is not None
            and active_result.input_signature == current_input_signature
        ):
            if active_result.success:
                st.divider()
                render_input_points(active_result)
    
                st.divider()
                render_lagrange_steps(active_result)
    
                st.divider()
                render_polynomial_and_graph(active_result)
    
                st.divider()
                render_error_analysis(active_result)
    
                st.divider()
                render_excel_download(active_result)
    
                st.divider()
                navigation_left_column, navigation_right_column = st.columns(2)
    
                with navigation_left_column:
                    if st.button(
                        "Review Lagrange Interpolation Lesson",
                        use_container_width=True,
                        key="review_lagrange_interpolation_lesson",
                    ):
                        st.switch_page("pages/Lagrange_Interpolation.py")
    
                with navigation_right_column:
                    if st.button(
                        "Back to Solver Menu",
                        use_container_width=True,
                        key="back_to_solver_menu_polynomial_interpolation",
                    ):
                        st.switch_page("pages/Numerical_Solver.py")
    
        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Lagrange Interpolation • Polynomial Interpolation</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()
