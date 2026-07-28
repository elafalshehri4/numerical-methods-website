from __future__ import annotations

import hashlib
import html
import math
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Callable, Sequence
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from matplotlib.lines import Line2D
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css

from sympy.parsing.sympy_parser import (
    convert_xor,
    implicit_multiplication_application,
    parse_expr,
    standard_transformations,
)


# ============================================================
# STREAMLIT PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Nonlinear System Solver | Numerical Methods",
    page_icon="📘",
    layout="wide",
)

load_css()
navbar(active_page="solver")


# ============================================================
# CONFIGURATION CONSTANTS
# ============================================================

DEFAULT_TOLERANCE = 1e-8
DISPLAY_DECIMALS = 3
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DEFAULT_MAX_ITERATIONS = 100
MAX_SUPPORTED_EQUATIONS = 8

MACHINE_ZERO_TOLERANCE = 1e-14
STAGNATION_TOLERANCE = 1e-14
JACOBIAN_CONDITION_WARNING = 1e12
DIVERGENCE_FACTOR = 1e12
MIN_DAMPING_FACTOR = 2.0 ** -20
ARMIJO_CONSTANT = 1e-4


# ============================================================
# GENERAL DISPLAY AND SIGNATURE HELPERS
# ============================================================

_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def current_report_datetime() -> datetime:
    """Return a timezone-aware timestamp for results and reports."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def format_scientific_power(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = DISPLAY_DECIMALS,
    unavailable: str = "—",
) -> str:
    """Display fixed or scientific notation according to magnitude."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6
    ):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round only numeric columns in a display copy."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def create_input_signature(
    equation_strings: Sequence[str],
    initial_guess: Sequence[float],
    tolerance: float,
    maximum_iterations: int,
    use_damping: bool,
) -> str:
    """Create a stable signature that prevents stale Streamlit results."""

    payload = repr(
        (
            tuple(str(item).strip() for item in equation_strings),
            tuple(f"{float(value):.17g}" for value in initial_guess),
            f"{float(tolerance):.17g}",
            int(maximum_iterations),
            bool(use_damping),
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def estimate_latest_order(history: Sequence[IterationRecord]) -> float | None:
    """Estimate local order from undamped consecutive Newton step norms."""

    if len(history) < 3:
        return None

    for index in range(len(history) - 1, 1, -1):
        records = history[index - 2:index + 1]
        if any(record.damping_factor < 1.0 for record in records):
            continue

        e0, e1, e2 = (
            float(record.approximate_error) for record in records
        )
        if (
            min(e0, e1, e2) <= MACHINE_ZERO_TOLERANCE
            or not all(math.isfinite(value) for value in (e0, e1, e2))
        ):
            continue

        denominator = math.log(e1 / e0)
        if abs(denominator) <= MACHINE_ZERO_TOLERANCE:
            continue

        value = math.log(e2 / e1) / denominator
        if math.isfinite(value):
            return float(value)

    return None


SAFE_TRANSFORMATIONS = standard_transformations + (
    convert_xor,
    implicit_multiplication_application,
)

SAFE_FUNCTIONS: dict[str, object] = {
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "abs": sp.Abs,
    "Abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}

SAFE_GLOBALS: dict[str, object] = {
    "__builtins__": {},
    "Symbol": sp.Symbol,
    "Integer": sp.Integer,
    "Float": sp.Float,
    "Rational": sp.Rational,
}

BLOCKED_WORDS = {
    "import",
    "exec",
    "eval",
    "compile",
    "open",
    "input",
    "globals",
    "locals",
    "lambda",
    "class",
    "def",
    "os",
    "sys",
    "subprocess",
    "pathlib",
    "builtins",
}


# ============================================================
# CUSTOM EXCEPTIONS
# ============================================================


class NonlinearSystemError(Exception):
    """Base exception for the nonlinear-system solver."""


class InputValidationError(NonlinearSystemError):
    """Raised when user input is invalid."""


class NumericalEvaluationError(NonlinearSystemError):
    """Raised when a numerical function cannot be evaluated safely."""


# ============================================================
# DATA MODELS
# ============================================================


@dataclass(frozen=True)
class ParsedNonlinearSystem:
    """Symbolic and numerical representation of a square nonlinear system."""

    original_equations: tuple[str, ...]
    normalized_equations: tuple[str, ...]
    expressions: tuple[sp.Expr, ...]
    variables: tuple[sp.Symbol, ...]
    function_matrix: sp.Matrix
    jacobian_matrix: sp.Matrix
    function_numeric: Callable[..., object]
    jacobian_numeric: Callable[..., object]

    @property
    def size(self) -> int:
        """Return the number of equations and variables."""

        return len(self.variables)


@dataclass(frozen=True)
class IterationRecord:
    """Complete numerical information for one Newton iteration."""

    iteration: int
    current_x: np.ndarray
    current_function: np.ndarray
    correction: np.ndarray
    next_x: np.ndarray
    next_function: np.ndarray
    approximate_error: float
    relative_approximate_error: float
    residual_norm: float
    jacobian_condition_number: float
    damping_factor: float
    step_criterion_satisfied: bool
    residual_criterion_satisfied: bool
    iteration_status: str


@dataclass(frozen=True)
class NonlinearSolverResult:
    """Structured output returned by Newton's method."""

    status: str
    message: str
    converged: bool
    method: str
    system: ParsedNonlinearSystem
    initial_guess: np.ndarray
    solution: np.ndarray
    function_at_solution: np.ndarray
    iterations: int
    history: tuple[IterationRecord, ...]
    tolerance: float
    maximum_iterations: int
    final_approximate_error: float
    final_relative_approximate_error: float
    final_residual_norm: float
    stopping_reason: str
    damping_enabled: bool
    latest_estimated_order: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# ============================================================
# GENERAL VALIDATION UTILITIES
# ============================================================


def _natural_sort_key(name: str) -> list[object]:
    """Create a natural-sort key so x2 is ordered before x10."""

    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", name)
    ]


def _validate_solver_settings(
    tolerance: float,
    maximum_iterations: int,
) -> tuple[float, int]:
    """Validate and normalize the numerical solver settings."""

    try:
        normalized_tolerance = float(tolerance)
    except (TypeError, ValueError) as error:
        raise InputValidationError(
            "Tolerance must be a valid numerical value."
        ) from error

    if not np.isfinite(normalized_tolerance) or normalized_tolerance <= 0.0:
        raise InputValidationError(
            "Tolerance must be a positive finite number."
        )

    if isinstance(maximum_iterations, bool):
        raise InputValidationError(
            "Maximum iterations must be a positive integer."
        )

    try:
        normalized_iterations = int(maximum_iterations)
    except (TypeError, ValueError) as error:
        raise InputValidationError(
            "Maximum iterations must be a positive integer."
        ) from error

    if normalized_iterations <= 0:
        raise InputValidationError(
            "Maximum iterations must be greater than zero."
        )

    return normalized_tolerance, normalized_iterations


def _to_finite_real_array(
    values: object,
    expected_shape: tuple[int, ...],
    value_name: str,
) -> np.ndarray:
    """Convert values to a finite real NumPy array with a required shape."""

    try:
        array = np.asarray(values)
    except Exception as error:
        raise NumericalEvaluationError(
            f"{value_name} could not be converted to a numerical array."
        ) from error

    if np.iscomplexobj(array):
        array = np.real_if_close(array, tol=1000)

    if np.iscomplexobj(array):
        raise NumericalEvaluationError(
            f"{value_name} contains non-real values."
        )

    try:
        array = np.asarray(array, dtype=float)
    except (TypeError, ValueError) as error:
        raise NumericalEvaluationError(
            f"{value_name} contains non-numerical values."
        ) from error

    if array.size != int(np.prod(expected_shape)):
        raise NumericalEvaluationError(
            f"{value_name} has an invalid size. Expected "
            f"{int(np.prod(expected_shape))} value(s), received {array.size}."
        )

    array = array.reshape(expected_shape)

    if not np.all(np.isfinite(array)):
        raise NumericalEvaluationError(
            f"{value_name} contains NaN or infinity. The current point may "
            "be outside the mathematical domain of one or more equations."
        )

    return array


def _normalize_initial_guess(
    initial_guess: Sequence[float],
    expected_size: int,
) -> np.ndarray:
    """Validate and convert the initial guess to a finite real vector."""

    if isinstance(initial_guess, (str, bytes)):
        raise InputValidationError(
            "The initial guess must be a numerical vector, not text."
        )

    try:
        guess_array = np.asarray(initial_guess)
    except Exception as error:
        raise InputValidationError(
            "The initial guess could not be converted to a numerical vector."
        ) from error

    if np.iscomplexobj(guess_array):
        guess_array = np.real_if_close(guess_array, tol=1000)

    if np.iscomplexobj(guess_array):
        raise InputValidationError(
            "The initial guess must contain real values only."
        )

    try:
        guess_array = np.asarray(guess_array, dtype=float).reshape(-1)
    except (TypeError, ValueError) as error:
        raise InputValidationError(
            "The initial guess must contain valid numerical values."
        ) from error

    if guess_array.size != expected_size:
        raise InputValidationError(
            "The number of initial-guess values must match the number of "
            f"variables. Expected {expected_size}, received {guess_array.size}."
        )

    if not np.all(np.isfinite(guess_array)):
        raise InputValidationError(
            "The initial guess must not contain NaN or infinity."
        )

    return guess_array


# ============================================================
# SAFE EQUATION PARSING
# ============================================================


def _normalize_equation_text(equation: str, equation_number: int) -> str:
    """Normalize common mathematical characters and validate raw text."""

    if not isinstance(equation, str):
        raise InputValidationError(
            f"Equation {equation_number} must be entered as text."
        )

    normalized = (
        equation.strip()
        .replace("−", "-")
        .replace("–", "-")
        .replace("×", "*")
        .replace("÷", "/")
        .replace("^", "**")
    )

    if not normalized:
        raise InputValidationError(
            f"Equation {equation_number} is empty."
        )

    lowered = normalized.lower()

    if "__" in normalized:
        raise InputValidationError(
            f"Equation {equation_number} contains a prohibited token."
        )

    for blocked_word in BLOCKED_WORDS:
        if re.search(rf"\b{re.escape(blocked_word)}\b", lowered):
            raise InputValidationError(
                f"Equation {equation_number} contains prohibited text: "
                f"'{blocked_word}'. Enter a mathematical expression only."
            )

    if any(character in normalized for character in "[]{};:'\"\\@#$`?"):
        raise InputValidationError(
            f"Equation {equation_number} contains unsupported characters."
        )

    function_like_names = re.findall(
        r"\b([A-Za-z][A-Za-z0-9_]*)\s*\(",
        normalized,
    )
    unsupported_functions = sorted(
        {name for name in function_like_names if name not in SAFE_FUNCTIONS}
    )
    if unsupported_functions:
        raise InputValidationError(
            f"Equation {equation_number} contains unsupported function "
            f"name(s): {', '.join(unsupported_functions)}."
        )

    if normalized.count("=") > 1:
        raise InputValidationError(
            f"Equation {equation_number} may contain at most one equals sign."
        )

    if "=" in normalized:
        left_side, right_side = normalized.split("=", maxsplit=1)
        if not left_side.strip() or not right_side.strip():
            raise InputValidationError(
                f"Equation {equation_number} must have expressions on both "
                "sides of the equals sign."
            )

    return normalized


def _collect_symbol_names(equations: Sequence[str]) -> set[str]:
    """Extract user-defined symbol names from normalized equations."""

    names: set[str] = set()

    for equation in equations:
        for token in re.findall(r"\b[A-Za-z][A-Za-z0-9_]*\b", equation):
            if token not in SAFE_FUNCTIONS:
                names.add(token)

    return names


def _build_parser_dictionary(symbol_names: set[str]) -> dict[str, object]:
    """Build a restricted local dictionary for SymPy parsing."""

    local_dictionary = dict(SAFE_FUNCTIONS)

    for name in symbol_names:
        if name.startswith("_"):
            raise InputValidationError(
                f"Variable name '{name}' is not allowed."
            )
        local_dictionary[name] = sp.Symbol(name, real=True)

    return local_dictionary


def _parse_single_equation(
    normalized_equation: str,
    equation_number: int,
    local_dictionary: dict[str, object],
) -> sp.Expr:
    """Parse one equation and convert it to the form f(X) = 0."""

    try:
        if "=" in normalized_equation:
            left_text, right_text = normalized_equation.split("=", maxsplit=1)
            left_expression = parse_expr(
                left_text,
                local_dict=local_dictionary,
                global_dict=SAFE_GLOBALS,
                transformations=SAFE_TRANSFORMATIONS,
                evaluate=True,
            )
            right_expression = parse_expr(
                right_text,
                local_dict=local_dictionary,
                global_dict=SAFE_GLOBALS,
                transformations=SAFE_TRANSFORMATIONS,
                evaluate=True,
            )
            expression = left_expression - right_expression
        else:
            expression = parse_expr(
                normalized_equation,
                local_dict=local_dictionary,
                global_dict=SAFE_GLOBALS,
                transformations=SAFE_TRANSFORMATIONS,
                evaluate=True,
            )
    except Exception as error:
        raise InputValidationError(
            f"Equation {equation_number} has an invalid format. Use standard "
            "Python/SymPy notation, such as x**2 + y**2 = 4."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise InputValidationError(
            f"Equation {equation_number} did not produce a valid expression."
        )

    if isinstance(expression, Relational) or expression.has(Relational):
        raise InputValidationError(
            f"Equation {equation_number} must be an equality or scalar "
            "expression, not an inequality or relation."
        )

    if expression.is_Matrix or expression.has(AppliedUndef):
        raise InputValidationError(
            f"Equation {equation_number} contains an unsupported function "
            "or matrix expression."
        )

    prohibited_objects = (
        sp.Derivative,
        sp.Integral,
        sp.Sum,
        sp.Product,
        sp.Limit,
    )

    if expression.has(*prohibited_objects):
        raise InputValidationError(
            f"Equation {equation_number} contains an unsupported symbolic "
            "operation. Enter an explicit algebraic or transcendental equation."
        )

    if expression.has(sp.I, sp.oo, -sp.oo, sp.zoo, sp.nan):
        raise InputValidationError(
            f"Equation {equation_number} contains an undefined, infinite, or "
            "complex constant."
        )

    if not expression.free_symbols:
        if expression.equals(0):
            raise InputValidationError(
                f"Equation {equation_number} simplifies to 0 = 0 and does not "
                "provide an independent equation."
            )
        raise InputValidationError(
            f"Equation {equation_number} is a nonzero constant and cannot be "
            "satisfied by changing the variables."
        )

    return expression


def parse_nonlinear_system(
    equation_strings: Sequence[str],
) -> ParsedNonlinearSystem:
    """
    Parse a square nonlinear system and construct its symbolic Jacobian.

    Accepted examples include both forms below:

        x**2 + y**2 - 4
        x**2 + y**2 = 4
    """

    if isinstance(equation_strings, (str, bytes)):
        raise InputValidationError(
            "Equations must be provided as a sequence of equation strings."
        )

    try:
        original_equations = tuple(equation_strings)
    except TypeError as error:
        raise InputValidationError(
            "Equations must be provided as a sequence of equation strings."
        ) from error

    if not original_equations:
        raise InputValidationError(
            "At least one nonlinear equation is required."
        )

    if len(original_equations) > MAX_SUPPORTED_EQUATIONS:
        raise InputValidationError(
            f"This interface supports up to {MAX_SUPPORTED_EQUATIONS} "
            "equations at one time."
        )

    normalized_equations = tuple(
        _normalize_equation_text(equation, index)
        for index, equation in enumerate(original_equations, start=1)
    )

    symbol_names = _collect_symbol_names(normalized_equations)

    if not symbol_names:
        raise InputValidationError(
            "No symbolic variables were detected in the system."
        )

    local_dictionary = _build_parser_dictionary(symbol_names)

    expressions = tuple(
        _parse_single_equation(equation, index, local_dictionary)
        for index, equation in enumerate(normalized_equations, start=1)
    )

    all_symbols: set[sp.Symbol] = set()
    for expression in expressions:
        all_symbols.update(expression.free_symbols)

    variables = tuple(
        sorted(all_symbols, key=lambda symbol: _natural_sort_key(symbol.name))
    )

    equation_count = len(expressions)
    variable_count = len(variables)

    if equation_count != variable_count:
        raise InputValidationError(
            "Newton's method in this solver requires a square system. "
            f"Detected {equation_count} equation(s) and {variable_count} "
            "variable(s)."
        )

    function_matrix = sp.Matrix(expressions)
    jacobian_matrix = function_matrix.jacobian(variables)

    for row_index in range(equation_count):
        if all(jacobian_matrix[row_index, column_index] == 0
               for column_index in range(variable_count)):
            raise InputValidationError(
                f"Equation {row_index + 1} has a zero Jacobian row and cannot "
                "contribute to a regular Newton system."
            )

    try:
        function_numeric = sp.lambdify(
            variables,
            function_matrix,
            modules=[{"Abs": np.abs}, "numpy"],
            cse=True,
        )
        jacobian_numeric = sp.lambdify(
            variables,
            jacobian_matrix,
            modules=[{"Abs": np.abs}, "numpy"],
            cse=True,
        )
    except Exception as error:
        raise InputValidationError(
            "The equations or their Jacobian could not be converted to "
            "numerical functions."
        ) from error

    return ParsedNonlinearSystem(
        original_equations=original_equations,
        normalized_equations=normalized_equations,
        expressions=expressions,
        variables=variables,
        function_matrix=function_matrix,
        jacobian_matrix=jacobian_matrix,
        function_numeric=function_numeric,
        jacobian_numeric=jacobian_numeric,
    )


# ============================================================
# NUMERICAL EVALUATION
# ============================================================


def evaluate_system(
    system: ParsedNonlinearSystem,
    point: np.ndarray,
) -> np.ndarray:
    """Evaluate F(X) safely at a real finite point."""

    try:
        with np.errstate(all="ignore"):
            values = system.function_numeric(*point.tolist())
    except Exception as error:
        raise NumericalEvaluationError(
            f"The nonlinear system could not be evaluated at "
            f"X = {point.tolist()}. Reason: {error}"
        ) from error

    return _to_finite_real_array(
        values,
        expected_shape=(system.size,),
        value_name=f"Function vector F(X) at X = {point.tolist()}",
    )


def evaluate_jacobian(
    system: ParsedNonlinearSystem,
    point: np.ndarray,
) -> np.ndarray:
    """Evaluate J(X) safely and verify its square shape."""

    try:
        with np.errstate(all="ignore"):
            values = system.jacobian_numeric(*point.tolist())
    except Exception as error:
        raise NumericalEvaluationError(
            f"The Jacobian could not be evaluated at X = {point.tolist()}. "
            f"Reason: {error}"
        ) from error

    return _to_finite_real_array(
        values,
        expected_shape=(system.size, system.size),
        value_name=f"Jacobian matrix J(X) at X = {point.tolist()}",
    )


# ============================================================
# NEWTON METHOD
# ============================================================


def _build_solver_result(
    *,
    status: str,
    message: str,
    converged: bool,
    system: ParsedNonlinearSystem,
    initial_guess: np.ndarray,
    solution: np.ndarray,
    function_at_solution: np.ndarray,
    history: list[IterationRecord],
    tolerance: float,
    maximum_iterations: int,
    stopping_reason: str,
    damping_enabled: bool,
) -> NonlinearSolverResult:
    """Create a consistent solver result for success, warning, or failure."""

    if history:
        final_record = history[-1]
        final_error = final_record.approximate_error
        final_relative_error = final_record.relative_approximate_error
        final_residual = final_record.residual_norm
    else:
        final_error = 0.0
        final_relative_error = 0.0
        final_residual = float(np.linalg.norm(function_at_solution, ord=2))

    warnings: list[str] = []
    if any(
        record.jacobian_condition_number >= JACOBIAN_CONDITION_WARNING
        for record in history
    ):
        warnings.append(
            "The Jacobian became ill-conditioned during at least one "
            "iteration. Small perturbations may strongly affect the correction."
        )
    if any(record.damping_factor < 1.0 for record in history):
        warnings.append(
            "Backtracking shortened at least one Newton correction to obtain "
            "a finite residual-reducing step."
        )
    if not converged:
        warnings.append(
            "The reported vector is the last available approximation and must "
            "not be treated as a verified solution."
        )

    latest_order = estimate_latest_order(history)
    if (
        converged
        and latest_order is not None
        and latest_order < 1.5
        and len(history) >= 4
    ):
        warnings.append(
            "The estimated local convergence was weaker than the quadratic "
            "behavior expected near a regular solution."
        )

    return NonlinearSolverResult(
        status=status,
        message=message,
        converged=converged,
        method="Newton's Method for Systems of Nonlinear Equations",
        system=system,
        initial_guess=initial_guess.copy(),
        solution=solution.copy(),
        function_at_solution=function_at_solution.copy(),
        iterations=len(history),
        history=tuple(history),
        tolerance=tolerance,
        maximum_iterations=maximum_iterations,
        final_approximate_error=float(final_error),
        final_relative_approximate_error=float(final_relative_error),
        final_residual_norm=float(final_residual),
        stopping_reason=stopping_reason,
        damping_enabled=damping_enabled,
        latest_estimated_order=latest_order,
        warnings=tuple(dict.fromkeys(warnings)),
        input_signature=create_input_signature(
            system.original_equations,
            initial_guess,
            tolerance,
            maximum_iterations,
            damping_enabled,
        ),
        execution_datetime=current_report_datetime(),
    )


def _compute_damped_update(
    system: ParsedNonlinearSystem,
    current_x: np.ndarray,
    current_residual_norm: float,
    full_correction: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, float]:
    """Apply residual-based backtracking to a Newton correction."""

    damping_factor = 1.0
    best_candidate: np.ndarray | None = None
    best_function: np.ndarray | None = None
    best_residual = np.inf

    while damping_factor >= MIN_DAMPING_FACTOR:
        candidate_x = current_x + damping_factor * full_correction

        try:
            candidate_function = evaluate_system(system, candidate_x)
        except NumericalEvaluationError:
            damping_factor *= 0.5
            continue

        candidate_residual = float(
            np.linalg.norm(candidate_function, ord=2)
        )

        if candidate_residual < best_residual:
            best_candidate = candidate_x
            best_function = candidate_function
            best_residual = candidate_residual

        armijo_target = (
            1.0 - ARMIJO_CONSTANT * damping_factor
        ) * current_residual_norm

        if candidate_residual <= armijo_target:
            return candidate_x, candidate_function, damping_factor

        damping_factor *= 0.5

    if (
        best_candidate is not None
        and best_function is not None
        and best_residual < current_residual_norm
    ):
        return best_candidate, best_function, max(
            damping_factor * 2.0,
            MIN_DAMPING_FACTOR,
        )

    raise NumericalEvaluationError(
        "Backtracking could not find a finite Newton step that reduces the "
        "residual norm. Try a different initial guess."
    )


def solve_nonlinear_system(
    equation_strings: Sequence[str],
    initial_guess: Sequence[float],
    tolerance: float = DEFAULT_TOLERANCE,
    maximum_iterations: int = DEFAULT_MAX_ITERATIONS,
    use_damping: bool = False,
) -> NonlinearSolverResult:
    """
    Solve a square nonlinear system with multivariable Newton's method.

    The Newton correction is computed from

        J(X_n) Delta_X_n = -F(X_n)
        X_(n+1) = X_n + Delta_X_n

    The Jacobian inverse is never formed explicitly. The linear correction
    system is solved with ``numpy.linalg.solve``.
    """

    normalized_tolerance, normalized_iterations = _validate_solver_settings(
        tolerance,
        maximum_iterations,
    )

    system = parse_nonlinear_system(equation_strings)
    normalized_guess = _normalize_initial_guess(initial_guess, system.size)

    try:
        current_function = evaluate_system(system, normalized_guess)
    except NumericalEvaluationError as error:
        raise InputValidationError(
            "The initial guess is outside the valid numerical domain of the "
            f"system. {error}"
        ) from error

    current_x = normalized_guess.copy()
    initial_residual_norm = float(
        np.linalg.norm(current_function, ord=2)
    )

    if initial_residual_norm <= normalized_tolerance:
        return _build_solver_result(
            status="success",
            message="The initial guess already satisfies the system.",
            converged=True,
            system=system,
            initial_guess=normalized_guess,
            solution=current_x,
            function_at_solution=current_function,
            history=[],
            tolerance=normalized_tolerance,
            maximum_iterations=normalized_iterations,
            stopping_reason=(
                "Stopped because the residual norm at the initial guess is "
                "less than or equal to the requested tolerance."
            ),
            damping_enabled=bool(use_damping),
        )

    history: list[IterationRecord] = []

    for iteration in range(1, normalized_iterations + 1):
        try:
            jacobian = evaluate_jacobian(system, current_x)
            condition_number = float(np.linalg.cond(jacobian))
        except (NumericalEvaluationError, np.linalg.LinAlgError) as error:
            return _build_solver_result(
                status="error",
                message=str(error),
                converged=False,
                system=system,
                initial_guess=normalized_guess,
                solution=current_x,
                function_at_solution=current_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because the Jacobian could not be evaluated "
                    "reliably."
                ),
                damping_enabled=bool(use_damping),
            )

        if not np.isfinite(condition_number):
            return _build_solver_result(
                status="error",
                message=(
                    "The Jacobian is singular or numerically unusable at "
                    f"X = {current_x.tolist()}. Try a different initial guess."
                ),
                converged=False,
                system=system,
                initial_guess=normalized_guess,
                solution=current_x,
                function_at_solution=current_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because the Jacobian condition number is not "
                    "finite."
                ),
                damping_enabled=bool(use_damping),
            )

        try:
            full_correction = np.linalg.solve(jacobian, -current_function)
        except np.linalg.LinAlgError:
            return _build_solver_result(
                status="error",
                message=(
                    "The Jacobian is singular at "
                    f"X = {current_x.tolist()}. Newton's correction cannot be "
                    "computed. Try another initial guess."
                ),
                converged=False,
                system=system,
                initial_guess=normalized_guess,
                solution=current_x,
                function_at_solution=current_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because the Newton linear system is singular."
                ),
                damping_enabled=bool(use_damping),
            )

        full_correction = _to_finite_real_array(
            full_correction,
            expected_shape=(system.size,),
            value_name="Newton correction vector",
        )

        try:
            if use_damping:
                next_x, next_function, damping_factor = _compute_damped_update(
                    system,
                    current_x,
                    float(np.linalg.norm(current_function, ord=2)),
                    full_correction,
                )
            else:
                damping_factor = 1.0
                next_x = current_x + full_correction
                next_function = evaluate_system(system, next_x)
        except NumericalEvaluationError as error:
            return _build_solver_result(
                status="error",
                message=str(error),
                converged=False,
                system=system,
                initial_guess=normalized_guess,
                solution=current_x,
                function_at_solution=current_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because the next Newton approximation could not "
                    "be evaluated safely."
                ),
                damping_enabled=bool(use_damping),
            )

        applied_correction = next_x - current_x
        approximate_error = float(
            np.linalg.norm(applied_correction, ord=2)
        )
        next_x_norm = float(np.linalg.norm(next_x, ord=2))
        relative_error = approximate_error / max(next_x_norm, 1.0)
        residual_norm = float(np.linalg.norm(next_function, ord=2))

        residual_satisfied = residual_norm <= normalized_tolerance
        step_satisfied = relative_error <= normalized_tolerance

        if residual_satisfied and step_satisfied:
            iteration_status = "Step and residual tolerances satisfied"
        elif residual_satisfied:
            iteration_status = "Residual satisfied; checking step criterion"
        elif step_satisfied:
            iteration_status = "Small step; residual still above tolerance"
        elif condition_number >= JACOBIAN_CONDITION_WARNING:
            iteration_status = "Ill-conditioned Jacobian warning"
        else:
            iteration_status = "Continue"

        history.append(
            IterationRecord(
                iteration=iteration,
                current_x=current_x.copy(),
                current_function=current_function.copy(),
                correction=applied_correction.copy(),
                next_x=next_x.copy(),
                next_function=next_function.copy(),
                approximate_error=approximate_error,
                relative_approximate_error=relative_error,
                residual_norm=residual_norm,
                jacobian_condition_number=condition_number,
                damping_factor=float(damping_factor),
                step_criterion_satisfied=step_satisfied,
                residual_criterion_satisfied=residual_satisfied,
                iteration_status=iteration_status,
            )
        )

        if residual_norm <= MACHINE_ZERO_TOLERANCE:
            return _build_solver_result(
                status="success",
                message="The nonlinear system was solved successfully.",
                converged=True,
                system=system,
                initial_guess=normalized_guess,
                solution=next_x,
                function_at_solution=next_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because the residual norm is numerically zero."
                ),
                damping_enabled=bool(use_damping),
            )

        if residual_satisfied and step_satisfied:
            return _build_solver_result(
                status="success",
                message="The nonlinear system was solved successfully.",
                converged=True,
                system=system,
                initial_guess=normalized_guess,
                solution=next_x,
                function_at_solution=next_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because both the relative Newton step and the "
                    "residual norm are less than or equal to the requested "
                    "tolerance."
                ),
                damping_enabled=bool(use_damping),
            )

        scaled_stagnation_limit = STAGNATION_TOLERANCE * max(next_x_norm, 1.0)
        if approximate_error <= scaled_stagnation_limit:
            return _build_solver_result(
                status="error",
                message=(
                    "Newton's method stagnated: the correction is numerically "
                    "negligible, but the residual remains above tolerance."
                ),
                converged=False,
                system=system,
                initial_guess=normalized_guess,
                solution=next_x,
                function_at_solution=next_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because numerical stagnation was detected."
                ),
                damping_enabled=bool(use_damping),
            )

        if step_satisfied and residual_norm > normalized_tolerance:
            return _build_solver_result(
                status="error",
                message=(
                    "The relative Newton step is below tolerance, but the "
                    "residual is still too large. This indicates stagnation, "
                    "poor scaling, or convergence to a non-solution."
                ),
                converged=False,
                system=system,
                initial_guess=normalized_guess,
                solution=next_x,
                function_at_solution=next_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because the step criterion was satisfied without "
                    "satisfying the residual criterion."
                ),
                damping_enabled=bool(use_damping),
            )

        divergence_threshold = min(
            max(
                DIVERGENCE_FACTOR * max(initial_residual_norm, 1.0),
                1e12,
            ),
            1e100,
        )
        if residual_norm > divergence_threshold:
            return _build_solver_result(
                status="error",
                message=(
                    "The residual norm grew excessively. Newton's method appears "
                    "to be diverging. Try a closer initial guess or enable "
                    "backtracking damping."
                ),
                converged=False,
                system=system,
                initial_guess=normalized_guess,
                solution=next_x,
                function_at_solution=next_function,
                history=history,
                tolerance=normalized_tolerance,
                maximum_iterations=normalized_iterations,
                stopping_reason=(
                    "Stopped because numerical divergence was detected."
                ),
                damping_enabled=bool(use_damping),
            )

        current_x = next_x
        current_function = next_function

    return _build_solver_result(
        status="warning",
        message=(
            "Maximum iterations were reached before the residual tolerance was "
            "satisfied. The last approximation is shown below."
        ),
        converged=False,
        system=system,
        initial_guess=normalized_guess,
        solution=current_x,
        function_at_solution=current_function,
        history=history,
        tolerance=normalized_tolerance,
        maximum_iterations=normalized_iterations,
        stopping_reason="Maximum iterations reached.",
        damping_enabled=bool(use_damping),
    )


# ============================================================
# DATAFRAME BUILDERS
# ============================================================


def create_iteration_dataframe(
    result: NonlinearSolverResult,
) -> pd.DataFrame:
    """Create the complete dynamic Newton iteration table."""

    rows: list[dict[str, object]] = []
    variables = result.system.variables

    for record in result.history:
        row: dict[str, object] = {"Iteration": record.iteration}

        for index, variable in enumerate(variables):
            row[f"{variable}_n"] = record.current_x[index]

        for index in range(result.system.size):
            row[f"f{index + 1}(X_n)"] = record.current_function[index]

        for index, variable in enumerate(variables):
            row[f"Delta {variable}"] = record.correction[index]

        for index, variable in enumerate(variables):
            row[f"{variable}_(n+1)"] = record.next_x[index]

        for index in range(result.system.size):
            row[f"f{index + 1}(X_(n+1))"] = record.next_function[index]

        row["Approximate Error"] = record.approximate_error
        row["Relative Approximate Error"] = (
            record.relative_approximate_error
        )
        row["Residual Norm"] = record.residual_norm
        row["Jacobian Condition Number"] = (
            record.jacobian_condition_number
        )
        row["Damping Factor"] = record.damping_factor
        row["Step Criterion"] = (
            "Satisfied" if record.step_criterion_satisfied else "Not satisfied"
        )
        row["Residual Criterion"] = (
            "Satisfied"
            if record.residual_criterion_satisfied
            else "Not satisfied"
        )
        row["Iteration Status"] = record.iteration_status
        rows.append(row)

    return pd.DataFrame(rows)


def create_error_analysis_dataframe(
    result: NonlinearSolverResult,
) -> pd.DataFrame:
    """Create error, residual, and Jacobian-conditioning data."""

    return pd.DataFrame(
        [
            {
                "Iteration": record.iteration,
                "Approximate Error": record.approximate_error,
                "Relative Approximate Error": (
                    record.relative_approximate_error
                ),
                "Residual Norm": record.residual_norm,
                "Jacobian Condition Number": (
                    record.jacobian_condition_number
                ),
            }
            for record in result.history
        ]
    )


def create_convergence_analysis_dataframe(
    result: NonlinearSolverResult,
) -> pd.DataFrame:
    """Create numerical convergence indicators from successive step norms."""

    analysis = create_error_analysis_dataframe(result)

    if analysis.empty:
        return analysis

    convergence = analysis[
        ["Iteration", "Approximate Error", "Residual Norm"]
    ].copy()
    convergence["Damping Factor"] = [
        record.damping_factor for record in result.history
    ]

    errors = convergence["Approximate Error"].to_numpy(dtype=float)
    error_ratio = np.full(errors.shape, np.nan, dtype=float)
    estimated_order = np.full(errors.shape, np.nan, dtype=float)
    quadratic_indicator = np.full(errors.shape, np.nan, dtype=float)

    for index in range(1, len(errors)):
        previous_error = errors[index - 1]
        current_error = errors[index]

        if previous_error > 0.0 and np.isfinite(previous_error):
            error_ratio[index] = current_error / previous_error
            if previous_error > np.sqrt(np.finfo(float).tiny):
                quadratic_indicator[index] = current_error / previous_error**2

    for index in range(2, len(errors)):
        error_0 = errors[index - 2]
        error_1 = errors[index - 1]
        error_2 = errors[index]

        if any(
            result.history[position].damping_factor < 1.0
            for position in range(index - 2, index + 1)
        ):
            continue

        if min(error_0, error_1, error_2) <= 0.0:
            continue

        denominator = np.log(error_1 / error_0)
        numerator = np.log(error_2 / error_1)

        if np.isfinite(denominator) and abs(denominator) > MACHINE_ZERO_TOLERANCE:
            estimated_order[index] = numerator / denominator

    convergence["Error Ratio"] = error_ratio
    convergence["Estimated Order"] = estimated_order
    convergence["Quadratic Indicator"] = quadratic_indicator

    return convergence


def create_solution_dataframe(
    result: NonlinearSolverResult,
) -> pd.DataFrame:
    """Create a table containing the final variables and residual components."""

    return pd.DataFrame(
        {
            "Variable": [str(variable) for variable in result.system.variables],
            "Approximate Solution": result.solution,
            "Function Component": [
                f"f{index + 1}(X)"
                for index in range(result.system.size)
            ],
            "Function Value": result.function_at_solution,
        }
    )


# ============================================================
# MATPLOTLIB FIGURES
# ============================================================


def create_variable_approximation_figure(
    result: NonlinearSolverResult,
) -> Figure | None:
    """Plot every variable approximation against the iteration number."""

    if not result.history:
        return None

    iterations = np.array(
        [record.iteration for record in result.history],
        dtype=int,
    )

    figure, axis = plt.subplots(figsize=(10, 5.5))

    for variable_index, variable in enumerate(result.system.variables):
        approximations = np.array(
            [record.next_x[variable_index] for record in result.history],
            dtype=float,
        )
        axis.plot(
            iterations,
            approximations,
            marker="o",
            linewidth=2,
            label=f"{variable} approximation",
        )
        axis.axhline(
            result.solution[variable_index],
            linestyle="--",
            linewidth=1,
            alpha=0.65,
        )

    axis.set_title("Variable Approximations by Newton Iteration")
    axis.set_xlabel("Iteration")
    axis.set_ylabel("Variable Value")
    axis.grid(True, alpha=0.3)
    axis.legend(loc="best")
    figure.tight_layout()

    return figure


def create_error_analysis_figure(
    result: NonlinearSolverResult,
) -> Figure | None:
    """Create a logarithmic error-and-residual analysis figure."""

    analysis = create_error_analysis_dataframe(result)

    if analysis.empty:
        return None

    iterations = analysis["Iteration"].to_numpy(dtype=int)
    approximate_errors = np.maximum(
        analysis["Approximate Error"].to_numpy(dtype=float),
        np.finfo(float).tiny,
    )
    relative_errors = np.maximum(
        analysis["Relative Approximate Error"].to_numpy(dtype=float),
        np.finfo(float).tiny,
    )
    residuals = np.maximum(
        analysis["Residual Norm"].to_numpy(dtype=float),
        np.finfo(float).tiny,
    )

    figure, axis = plt.subplots(figsize=(10, 5.5))
    axis.semilogy(
        iterations,
        approximate_errors,
        marker="o",
        linewidth=2,
        label="Approximate Error ||Delta X||₂",
    )
    axis.semilogy(
        iterations,
        relative_errors,
        marker="s",
        linewidth=2,
        label="Relative Approximate Error",
    )
    axis.semilogy(
        iterations,
        residuals,
        marker="^",
        linewidth=2,
        label="Residual Norm ||F(X)||₂",
    )
    axis.axhline(
        result.tolerance,
        linestyle="--",
        linewidth=1.4,
        label=f"Tolerance = {result.tolerance:.2e}",
    )
    axis.set_title("Error and Residual Analysis")
    axis.set_xlabel("Iteration")
    axis.set_ylabel("Magnitude (Log Scale)")
    axis.grid(True, which="both", alpha=0.3)
    axis.legend(loc="best")
    figure.tight_layout()

    return figure


def create_convergence_figure(
    result: NonlinearSolverResult,
) -> Figure | None:
    """Create the semilog Newton convergence figure."""

    convergence = create_convergence_analysis_dataframe(result)

    if convergence.empty:
        return None

    iterations = convergence["Iteration"].to_numpy(dtype=int)
    step_errors = np.maximum(
        convergence["Approximate Error"].to_numpy(dtype=float),
        np.finfo(float).tiny,
    )
    residuals = np.maximum(
        convergence["Residual Norm"].to_numpy(dtype=float),
        np.finfo(float).tiny,
    )

    figure, axis = plt.subplots(figsize=(10, 5.5))
    axis.semilogy(
        iterations,
        step_errors,
        marker="o",
        linewidth=2,
        label="Approximate Error",
    )
    axis.semilogy(
        iterations,
        residuals,
        marker="s",
        linewidth=2,
        label="Residual Norm",
    )
    axis.set_title("Semilog Convergence Analysis")
    axis.set_xlabel("Iteration")
    axis.set_ylabel("Magnitude (Log Scale)")
    axis.grid(True, which="both", alpha=0.3)
    axis.legend(loc="best")
    figure.tight_layout()

    return figure


def _evaluate_expression_on_grid(
    expression: sp.Expr,
    variables: tuple[sp.Symbol, sp.Symbol],
    x_grid: np.ndarray,
    y_grid: np.ndarray,
) -> np.ndarray:
    """Evaluate one symbolic expression on a two-dimensional grid."""

    function = sp.lambdify(
        variables,
        expression,
        modules=[{"Abs": np.abs}, "numpy"],
        cse=True,
    )

    with np.errstate(all="ignore"):
        values = function(x_grid, y_grid)

    values = np.asarray(values)

    if values.ndim == 0:
        values = np.full_like(x_grid, values, dtype=float)

    if np.iscomplexobj(values):
        values = np.real_if_close(values, tol=1000)

    if np.iscomplexobj(values):
        raise NumericalEvaluationError(
            "The contour contains non-real values in the selected graph range."
        )

    values = np.asarray(values, dtype=float)

    if values.shape != x_grid.shape:
        values = np.broadcast_to(values, x_grid.shape).astype(float)

    return np.where(np.isfinite(values), values, np.nan)


def create_two_variable_contour_figure(
    result: NonlinearSolverResult,
) -> Figure | None:
    """Plot zero contours and the Newton path for a 2 x 2 system."""

    if result.system.size != 2:
        return None

    path_points = [result.initial_guess.copy()]
    path_points.extend(record.next_x.copy() for record in result.history)
    path = np.asarray(path_points, dtype=float)

    x_min = float(np.min(path[:, 0]))
    x_max = float(np.max(path[:, 0]))
    y_min = float(np.min(path[:, 1]))
    y_max = float(np.max(path[:, 1]))

    x_span = max(x_max - x_min, 1.0)
    y_span = max(y_max - y_min, 1.0)

    x_values = np.linspace(
        x_min - 0.8 * x_span,
        x_max + 0.8 * x_span,
        350,
    )
    y_values = np.linspace(
        y_min - 0.8 * y_span,
        y_max + 0.8 * y_span,
        350,
    )
    x_grid, y_grid = np.meshgrid(x_values, y_values)

    variable_pair = (
        result.system.variables[0],
        result.system.variables[1],
    )

    figure, axis = plt.subplots(figsize=(9, 7))
    contour_handles: list[Line2D] = []

    for index, expression in enumerate(result.system.expressions, start=1):
        values = _evaluate_expression_on_grid(
            expression,
            variable_pair,
            x_grid,
            y_grid,
        )

        finite_values = values[np.isfinite(values)]
        if finite_values.size == 0:
            continue

        if np.min(finite_values) <= 0.0 <= np.max(finite_values):
            contour = axis.contour(
                x_grid,
                y_grid,
                values,
                levels=[0.0],
                linewidths=2,
            )
            if contour.allsegs and contour.allsegs[0]:
                contour_handles.append(
                    Line2D(
                        [0],
                        [0],
                        linewidth=2,
                        label=f"f{index} = 0",
                    )
                )

    axis.plot(
        path[:, 0],
        path[:, 1],
        marker="o",
        linewidth=2,
        label="Newton Iteration Path",
    )
    axis.scatter(
        result.initial_guess[0],
        result.initial_guess[1],
        marker="s",
        s=90,
        zorder=5,
        label="Initial Guess",
    )
    axis.scatter(
        result.solution[0],
        result.solution[1],
        marker="*",
        s=170,
        zorder=6,
        label="Final Approximation",
    )

    axis.axhline(0.0, linewidth=0.8, alpha=0.5)
    axis.axvline(0.0, linewidth=0.8, alpha=0.5)
    axis.set_title("Zero Contours and Newton Iteration Path")
    axis.set_xlabel(str(variable_pair[0]))
    axis.set_ylabel(str(variable_pair[1]))
    axis.grid(True, alpha=0.25)

    handles, labels = axis.get_legend_handles_labels()
    axis.legend(contour_handles + handles, [
        handle.get_label() for handle in contour_handles
    ] + labels, loc="best")

    figure.tight_layout()
    return figure


# ============================================================
# EXCEL REPORT
# ============================================================


def create_method_formula_dataframe(
    result: NonlinearSolverResult,
) -> pd.DataFrame:
    """Return the Newton-system formulas and numerical criteria."""

    return pd.DataFrame(
        {
            "Item": [
                "Nonlinear System",
                "Jacobian Matrix",
                "Newton Correction",
                "State Update",
                "Absolute Step Norm",
                "Relative Step Norm",
                "Residual Norm",
                "Convergence Requirement",
                "Expected Local Behavior",
                "Damping Safeguard",
            ],
            "Formula / Meaning": [
                "F(X) = 0",
                "J_ij(X) = partial f_i / partial x_j",
                "Solve J(X_k) Delta X_k = -F(X_k)",
                "X_(k+1) = X_k + lambda_k Delta X_k",
                "||X_(k+1)-X_k||_2",
                "||X_(k+1)-X_k||_2 / max(1, ||X_(k+1)||_2)",
                "||F(X_(k+1))||_2",
                "Require both relative step and residual <= tolerance",
                "Quadratic near a regular root when J is nonsingular",
                (
                    "Optional backtracking chooses 0 < lambda_k <= 1 "
                    "to reduce the residual norm"
                ),
            ],
        }
    )


def create_jacobian_diagnostics_dataframe(
    result: NonlinearSolverResult,
) -> pd.DataFrame:
    """Create determinant, rank, singular-value, and conditioning diagnostics."""

    rows: list[dict[str, object]] = []

    for record in result.history:
        try:
            jacobian = evaluate_jacobian(result.system, record.current_x)
            singular_values = np.linalg.svd(jacobian, compute_uv=False)
            determinant = float(np.linalg.det(jacobian))
            rank = int(np.linalg.matrix_rank(jacobian))
            minimum_singular_value = float(np.min(singular_values))
            maximum_singular_value = float(np.max(singular_values))
            jacobian_text = np.array2string(
                jacobian,
                precision=12,
                separator=", ",
                max_line_width=10000,
            )
        except (NumericalEvaluationError, np.linalg.LinAlgError, ValueError):
            determinant = np.nan
            rank = 0
            minimum_singular_value = np.nan
            maximum_singular_value = np.nan
            jacobian_text = "Unavailable"

        rows.append(
            {
                "Iteration": record.iteration,
                "Jacobian Matrix": jacobian_text,
                "Determinant": determinant,
                "Rank": rank,
                "Condition Number": record.jacobian_condition_number,
                "Minimum Singular Value": minimum_singular_value,
                "Maximum Singular Value": maximum_singular_value,
                "Damping Factor": record.damping_factor,
                "Condition Assessment": (
                    "Ill-conditioned"
                    if record.jacobian_condition_number
                    >= JACOBIAN_CONDITION_WARNING
                    else "Acceptable"
                ),
            }
        )

    return pd.DataFrame(rows)


def _style_excel_workbook(workbook: object) -> None:
    """Apply readable professional formatting to every worksheet."""

    header_fill = PatternFill("solid", fgColor="0D3151")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.showGridLines = False

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions

        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            maximum_length = 0

            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                maximum_length = max(
                    maximum_length,
                    len(str(cell.value)) if cell.value is not None else 0,
                )
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

            worksheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 12),
                58,
            )


def _add_excel_line_chart(
    worksheet: object,
    *,
    title: str,
    data_columns: Sequence[int],
    category_column: int,
    maximum_row: int,
    anchor: str,
    y_axis_title: str,
    logarithmic: bool = False,
) -> None:
    """Add an editable line chart to an openpyxl worksheet."""

    if maximum_row < 2:
        return

    chart = LineChart()
    chart.title = title
    chart.x_axis.title = "Iteration"
    chart.y_axis.title = y_axis_title
    chart.height = 8
    chart.width = 15

    if logarithmic:
        chart.y_axis.scaling.logBase = 10

    categories = Reference(
        worksheet,
        min_col=category_column,
        min_row=2,
        max_row=maximum_row,
    )

    for column in data_columns:
        data = Reference(
            worksheet,
            min_col=column,
            min_row=1,
            max_row=maximum_row,
        )
        chart.add_data(data, titles_from_data=True)

    chart.set_categories(categories)
    worksheet.add_chart(chart, anchor)


def figure_to_png_bytes(figure: Figure) -> bytes:
    """Serialize a matplotlib figure as high-resolution PNG bytes."""

    output = BytesIO()
    figure.savefig(
        output,
        format="png",
        dpi=180,
        bbox_inches="tight",
    )
    output.seek(0)
    return output.getvalue()


def add_excel_image(
    worksheet: object,
    image_bytes: bytes,
    anchor: str,
    width: int = 760,
    height: int = 440,
) -> None:
    """Embed one PNG image in an Excel worksheet."""

    stream = BytesIO(image_bytes)
    image = ExcelImage(stream)
    image.width = width
    image.height = height
    worksheet.add_image(image, anchor)


def create_excel_report(result: NonlinearSolverResult) -> bytes:
    """Create a complete in-memory XLSX report for the nonlinear system."""

    variables_text = ", ".join(
        str(variable) for variable in result.system.variables
    )
    initial_guess_text = ", ".join(
        f"{variable} = {value:.15g}"
        for variable, value in zip(
            result.system.variables,
            result.initial_guess,
        )
    )
    solution_text = ", ".join(
        f"{variable} = {value:.15g}"
        for variable, value in zip(
            result.system.variables,
            result.solution,
        )
    )

    summary_dataframe = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Status",
                "Converged",
                "Number of Equations",
                "Variables",
                "Initial Guess",
                "Approximate Solution",
                "Tolerance",
                "Maximum Iterations",
                "Completed Iterations",
                "Final Step Norm",
                "Final Relative Step Norm",
                "Final Residual 2-Norm",
                "Latest Estimated Order",
                "Backtracking Damping",
                "Warnings",
                "Stopping Reason",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.status,
                "Yes" if result.converged else "No",
                result.system.size,
                variables_text,
                initial_guess_text,
                solution_text,
                result.tolerance,
                result.maximum_iterations,
                result.iterations,
                result.final_approximate_error,
                result.final_relative_approximate_error,
                result.final_residual_norm,
                result.latest_estimated_order,
                "Enabled" if result.damping_enabled else "Disabled",
                " | ".join(result.warnings) if result.warnings else "None",
                result.stopping_reason,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )

    system_rows: list[dict[str, object]] = []
    for index, expression in enumerate(result.system.expressions, start=1):
        system_rows.append(
            {
                "Item": f"Equation {index}",
                "User Input": result.system.original_equations[index - 1],
                "Normalized Form f_i(X) = 0": str(expression),
            }
        )

    system_rows.append(
        {
            "Item": "Jacobian Matrix",
            "User Input": "Automatically generated",
            "Normalized Form f_i(X) = 0": str(result.system.jacobian_matrix),
        }
    )
    system_dataframe = pd.DataFrame(system_rows)

    iteration_dataframe = create_iteration_dataframe(result)
    error_dataframe = create_error_analysis_dataframe(result)
    convergence_dataframe = create_convergence_analysis_dataframe(result)
    solution_dataframe = create_solution_dataframe(result)
    jacobian_dataframe = create_jacobian_diagnostics_dataframe(result)
    formula_dataframe = create_method_formula_dataframe(result)

    if iteration_dataframe.empty:
        iteration_dataframe = pd.DataFrame(
            {"Message": ["No Newton iterations were required."]}
        )
    if error_dataframe.empty:
        error_dataframe = pd.DataFrame(
            {"Message": ["No iterative error analysis was required."]}
        )
    if convergence_dataframe.empty:
        convergence_dataframe = pd.DataFrame(
            {"Message": ["No iterative convergence analysis was required."]}
        )
    if jacobian_dataframe.empty:
        jacobian_dataframe = pd.DataFrame(
            {"Message": ["No Jacobian diagnostics were required."]}
        )

    figures: list[tuple[str, Figure]] = []
    variable_figure = create_variable_approximation_figure(result)
    if variable_figure is not None:
        figures.append(("Variable approximations", variable_figure))

    error_figure = create_error_analysis_figure(result)
    if error_figure is not None:
        figures.append(("Error and residual analysis", error_figure))

    convergence_figure = create_convergence_figure(result)
    if convergence_figure is not None:
        figures.append(("Convergence analysis", convergence_figure))

    if result.system.size == 2:
        try:
            contour_figure = create_two_variable_contour_figure(result)
        except Exception:
            contour_figure = None
        if contour_figure is not None:
            figures.append(("Contours and Newton path", contour_figure))

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_dataframe.to_excel(writer, sheet_name="Summary", index=False)
        formula_dataframe.to_excel(
            writer,
            sheet_name="Method Formula",
            index=False,
        )
        system_dataframe.to_excel(
            writer,
            sheet_name="System Details",
            index=False,
        )
        solution_dataframe.to_excel(
            writer,
            sheet_name="Final Solution",
            index=False,
        )
        iteration_dataframe.to_excel(
            writer,
            sheet_name="Iteration Results",
            index=False,
        )
        error_dataframe.to_excel(
            writer,
            sheet_name="Error Analysis",
            index=False,
        )
        convergence_dataframe.to_excel(
            writer,
            sheet_name="Convergence Analysis",
            index=False,
        )
        jacobian_dataframe.to_excel(
            writer,
            sheet_name="Jacobian Diagnostics",
            index=False,
        )

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Nonlinear System Solver Plots"
        plots_sheet["A1"].font = Font(bold=True, size=14)

        for figure_index, (title, figure) in enumerate(figures):
            anchor_row = 3 + 27 * figure_index
            plots_sheet.cell(anchor_row - 1, 1, title)
            plots_sheet.cell(anchor_row - 1, 1).font = Font(bold=True, size=12)
            add_excel_image(
                plots_sheet,
                figure_to_png_bytes(figure),
                f"A{anchor_row}",
            )
            plt.close(figure)

        _style_excel_workbook(workbook)

        if result.history:
            summary_sheet = workbook["Summary"]
            iteration_sheet = workbook["Iteration Results"]
            error_sheet = workbook["Error Analysis"]
            jacobian_sheet = workbook["Jacobian Diagnostics"]

            iteration_columns = {
                str(cell.value): index
                for index, cell in enumerate(iteration_sheet[1], start=1)
            }
            variable_columns = [
                iteration_columns[f"{variable}_(n+1)"]
                for variable in result.system.variables
                if f"{variable}_(n+1)" in iteration_columns
            ]
            if variable_columns:
                _add_excel_line_chart(
                    summary_sheet,
                    title="Variable Approximations by Newton Iteration",
                    data_columns=variable_columns,
                    category_column=iteration_columns.get("Iteration", 1),
                    maximum_row=len(result.history) + 1,
                    anchor="D2",
                    y_axis_title="Variable Value",
                )

            error_columns = {
                str(cell.value): index
                for index, cell in enumerate(error_sheet[1], start=1)
            }
            selected_error_columns = [
                error_columns[name]
                for name in (
                    "Approximate Error",
                    "Relative Approximate Error",
                    "Residual Norm",
                )
                if name in error_columns
            ]
            if selected_error_columns:
                _add_excel_line_chart(
                    summary_sheet,
                    title="Step and Residual Convergence",
                    data_columns=selected_error_columns,
                    category_column=error_columns.get("Iteration", 1),
                    maximum_row=len(result.history) + 1,
                    anchor="D20",
                    y_axis_title="Magnitude",
                    logarithmic=True,
                )

            jacobian_columns = {
                str(cell.value): index
                for index, cell in enumerate(jacobian_sheet[1], start=1)
            }
            if "Condition Number" in jacobian_columns:
                _add_excel_line_chart(
                    summary_sheet,
                    title="Jacobian Condition Number",
                    data_columns=(jacobian_columns["Condition Number"],),
                    category_column=jacobian_columns.get("Iteration", 1),
                    maximum_row=len(result.history) + 1,
                    anchor="D38",
                    y_axis_title="Condition Number",
                    logarithmic=True,
                )

        workbook.active = workbook.sheetnames.index("Summary")

    output.seek(0)
    return output.getvalue()


# ============================================================
# STREAMLIT PAGE HELPERS
# ============================================================


def render_system_preview(system: ParsedNonlinearSystem) -> None:
    """Display detected variables and the automatically generated Jacobian."""

    variable_names = ", ".join(
        html.escape(str(variable)) for variable in system.variables
    )

    st.markdown(
        f"""
        <div class="final-interval-box">
            Detected variables: <strong>{variable_names}</strong>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("View function vector and generated Jacobian"):
        st.latex("F(X) = " + sp.latex(system.function_matrix))
        st.latex("J(X) = " + sp.latex(system.jacobian_matrix))


def render_final_result_card(result: NonlinearSolverResult) -> None:
    """Render the final nonlinear-system result card."""

    if result.status == "success":
        st.success(result.message)
    elif result.status == "warning":
        st.warning(result.message)
    else:
        st.error(result.message)

    solution_columns = st.columns(min(result.system.size, 4))
    for index, (variable, value) in enumerate(
        zip(result.system.variables, result.solution)
    ):
        solution_columns[index % len(solution_columns)].metric(
            str(variable),
            format_display_number(value, decimals=10),
        )

    first_metrics = st.columns(2)
    first_metrics[0].metric(
        "Residual 2-Norm",
        format_display_number(result.final_residual_norm, decimals=6),
    )
    first_metrics[1].metric("Iterations", result.iterations)

    second_metrics = st.columns(2)
    second_metrics[0].metric(
        "Final Step Norm",
        format_display_number(result.final_approximate_error, decimals=6),
    )
    second_metrics[1].metric(
        "Relative Step Norm",
        format_display_number(
            result.final_relative_approximate_error,
            decimals=6,
        ),
    )

    if result.latest_estimated_order is not None:
        st.metric(
            "Latest Estimated Order",
            format_display_number(result.latest_estimated_order, decimals=4),
        )

    st.markdown(
        f"**Backtracking damping:** "
        f"{'Enabled' if result.damping_enabled else 'Disabled'}"
    )
    st.markdown(f"**Stopping reason:** {result.stopping_reason}")

    for warning in result.warnings:
        st.warning(warning)


def render_iteration_results(result: NonlinearSolverResult) -> None:
    """Display the complete iteration DataFrame rounded to six decimals."""

    st.subheader("Iteration Table")
    iteration_dataframe = create_iteration_dataframe(result)

    if iteration_dataframe.empty:
        st.info(
            "No Newton iterations were required because the initial guess "
            "already satisfied the requested residual tolerance."
        )
        return

    display_dataframe = iteration_dataframe.copy()
    numeric_columns = display_dataframe.select_dtypes(
        include=[np.number]
    ).columns
    display_dataframe[numeric_columns] = display_dataframe[
        numeric_columns
    ].round(6)

    st.dataframe(
        display_dataframe,
        use_container_width=True,
        hide_index=True,
    )


def render_error_analysis_section(result: NonlinearSolverResult) -> None:
    """Display the error table and logarithmic analysis plot."""

    st.subheader("Error Analysis")
    error_dataframe = create_error_analysis_dataframe(result)

    if error_dataframe.empty:
        st.info("No iterative error analysis was required.")
        return

    display_dataframe = error_dataframe.copy()
    numeric_columns = display_dataframe.select_dtypes(
        include=[np.number]
    ).columns
    display_dataframe[numeric_columns] = display_dataframe[
        numeric_columns
    ].round(10)

    st.dataframe(
        display_dataframe,
        use_container_width=True,
        hide_index=True,
    )

    figure = create_error_analysis_figure(result)
    if figure is not None:
        st.pyplot(figure)
        plt.close(figure)


def render_convergence_analysis_section(
    result: NonlinearSolverResult,
) -> None:
    """Display convergence indicators and the semilog plot."""

    st.subheader("Convergence Analysis")
    convergence_dataframe = create_convergence_analysis_dataframe(result)

    if convergence_dataframe.empty:
        st.info("No iterative convergence analysis was required.")
        return

    display_dataframe = convergence_dataframe.copy()
    numeric_columns = display_dataframe.select_dtypes(
        include=[np.number]
    ).columns
    display_dataframe[numeric_columns] = display_dataframe[
        numeric_columns
    ].round(10)

    st.dataframe(
        display_dataframe,
        use_container_width=True,
        hide_index=True,
    )

    figure = create_convergence_figure(result)
    if figure is not None:
        st.pyplot(figure)
        plt.close(figure)

    st.info(
        "Newton's method can exhibit quadratic local convergence near a "
        "regular solution when the Jacobian is nonsingular and the initial "
        "guess is sufficiently close. The estimated order is calculated from "
        "successive approximate errors, not from a known exact solution."
    )


def render_solver_graphs(result: NonlinearSolverResult) -> None:
    """Render variable approximations and the optional 2D contour graph."""

    st.subheader("Graphs")
    graph_columns = st.columns(2)

    with graph_columns[0]:
        with st.container(border=True):
            st.markdown("#### Variable Approximations")
            variable_figure = create_variable_approximation_figure(result)

            if variable_figure is None:
                st.info("No variable-approximation graph was required.")
            else:
                st.pyplot(variable_figure)
                plt.close(variable_figure)

    with graph_columns[1]:
        with st.container(border=True):
            st.markdown("#### System Contours and Newton Path")

            if result.system.size != 2:
                st.info(
                    "This graph is available only for a two-equation, "
                    "two-variable system."
                )
            else:
                try:
                    contour_figure = create_two_variable_contour_figure(result)
                    if contour_figure is None:
                        st.warning("The contour graph could not be generated.")
                    else:
                        st.pyplot(contour_figure)
                        plt.close(contour_figure)
                except Exception as error:
                    st.warning(
                        "The contour graph could not be generated for the "
                        f"entered system. Reason: {error}"
                    )


def render_excel_download(result: NonlinearSolverResult) -> None:
    """Generate, cache, and display the Excel-report download button."""

    st.subheader("Excel Report")
    signature_key = "nonlinear_system_excel_signature"
    report_key = "nonlinear_system_excel_report"

    if (
        st.session_state.get(signature_key) != result.input_signature
        or report_key not in st.session_state
    ):
        try:
            st.session_state[report_key] = create_excel_report(result)
            st.session_state[signature_key] = result.input_signature
        except (ValueError, OSError, TypeError, ArithmeticError) as error:
            st.error(
                "The Excel report could not be generated. "
                f"Reason: {error}"
            )
            return

    report_data = st.session_state.get(report_key)
    if not report_data:
        st.error("The Excel report is unavailable.")
        return

    report_date = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    st.download_button(
        label="Download Complete Excel Report",
        data=report_data,
        file_name=(
            "nonlinear_system_newton_report_"
            f"{report_date}.xlsx"
        ),
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="nonlinear_system_excel_download",
    )


# ============================================================
# STREAMLIT PAGE
# ============================================================

st.html(
    """
    <section class="solver-hero">
        <div>
            <div class="page-label">NONLINEAR SYSTEM NEWTON TOOL</div>
            <h1>Systems of Nonlinear Equations Solver</h1>
            <p>
                Enter a square nonlinear system, generate its Jacobian
                automatically, and inspect every Newton correction, residual,
                conditioning diagnostic, convergence result, graph, and report.
            </p>
            <div class="method-actions">
                <a href="/Systems_of_Nonlinear_Equations" target="_self"
                   class="btn-outline-ui">Review Lesson →</a>
                <a href="/Systems_of_Nonlinear_Equations_Quiz" target="_self"
                   class="btn-primary-ui">Take Quiz →</a>
            </div>
        </div>
    </section>
    """
)

left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])

with main_area:
    st.markdown(
        '<main class="solver-wrapper solver-streamlit-area">',
        unsafe_allow_html=True,
    )

    guide_column, conditions_column = st.columns(2)

    with guide_column:
        with st.container(border=True):
            st.subheader("How to Write the System")
            st.markdown(
                """
                Enter one equation in each field using Python/SymPy syntax.

                - Both `f(X) = 0` and `left side = right side` are accepted.
                - Use matching symbolic variables across the system.
                - Powers: `x**2`; multiplication: `x*y`.
                - Supported functions include `sin`, `cos`, `exp`, `log`,
                  `sqrt`, and `abs`.
                - Example: `x**2 + y**2 = 4` and `x - y = 0`.
                """
            )
            st.markdown("**Newton update**")
            st.latex(r"J(X_k)\Delta X_k=-F(X_k)")
            st.latex(r"X_{k+1}=X_k+\lambda_k\Delta X_k")

    with conditions_column:
        with st.container(border=True):
            st.subheader("Before Solving")
            st.markdown(
                """
                - The number of equations must equal the number of detected variables.
                - The initial vector should be reasonably close to the desired root.
                - The Jacobian must remain nonsingular and reasonably conditioned.
                - A solution is accepted only when both the relative step and
                  residual norm satisfy the tolerance.
                - Backtracking damping can improve reliability when a full Newton
                  step leaves the valid domain or increases the residual.
                """
            )
            st.info(
                "Newton's method can converge rapidly near a regular solution, "
                "but different initial guesses may converge to different roots "
                "or fail to converge."
            )

    input_column, result_column = st.columns([1.35, 1.0])

    with input_column:
        with st.container(border=True):
            st.markdown(
                '<h3 class="solver-box-title">Input</h3>',
                unsafe_allow_html=True,
            )

            st.markdown(
                '<div class="input-label-ui">Number of equations</div>',
                unsafe_allow_html=True,
            )
            number_of_equations = int(
                st.number_input(
                    "Number of equations",
                    min_value=2,
                    max_value=MAX_SUPPORTED_EQUATIONS,
                    value=2,
                    step=1,
                    label_visibility="collapsed",
                    key="nonlinear_equation_count",
                )
            )

            default_equations = (
                "x**2 + y**2 = 4",
                "x - y = 0",
                "x + y + z**2 = 3",
                "x + y**2 + z + w = 4",
            )

            equation_strings: list[str] = []
            for equation_index in range(number_of_equations):
                st.markdown(
                    f'<div class="input-label-ui">Equation {equation_index + 1}</div>',
                    unsafe_allow_html=True,
                )
                equation_strings.append(
                    st.text_input(
                        f"Equation {equation_index + 1}",
                        value=(
                            default_equations[equation_index]
                            if equation_index < len(default_equations)
                            else ""
                        ),
                        placeholder="Example: x**2 + y**2 = 4",
                        label_visibility="collapsed",
                        key=f"nonlinear_equation_{equation_index}",
                    )
                )

            parsed_system: ParsedNonlinearSystem | None = None
            preview_error: str | None = None

            try:
                parsed_system = parse_nonlinear_system(equation_strings)
            except InputValidationError as error:
                preview_error = str(error)

            if parsed_system is not None:
                render_system_preview(parsed_system)
            elif preview_error:
                st.warning(preview_error)

            initial_guess: list[float] = []
            if parsed_system is not None:
                st.markdown(
                    '<div class="input-label-ui">Initial guess</div>',
                    unsafe_allow_html=True,
                )
                guess_columns = st.columns(min(parsed_system.size, 3))
                guess_defaults = [1.5, 1.0]

                for variable_index, variable in enumerate(parsed_system.variables):
                    with guess_columns[variable_index % len(guess_columns)]:
                        initial_guess.append(
                            float(
                                st.number_input(
                                    f"Initial {variable}",
                                    value=(
                                        guess_defaults[variable_index]
                                        if variable_index < len(guess_defaults)
                                        else 1.0
                                    ),
                                    format="%.12g",
                                    key=f"nonlinear_guess_{variable}",
                                )
                            )
                        )

            control_columns = st.columns(2)
            with control_columns[0]:
                st.markdown(
                    '<div class="input-label-ui">Tolerance</div>',
                    unsafe_allow_html=True,
                )
                tolerance = st.number_input(
                    "Tolerance",
                    value=DEFAULT_TOLERANCE,
                    min_value=1.0e-14,
                    max_value=1.0,
                    format="%.12g",
                    label_visibility="collapsed",
                    key="nonlinear_tolerance",
                )

            with control_columns[1]:
                st.markdown(
                    '<div class="input-label-ui">Maximum iterations</div>',
                    unsafe_allow_html=True,
                )
                maximum_iterations = int(
                    st.number_input(
                        "Maximum iterations",
                        value=DEFAULT_MAX_ITERATIONS,
                        min_value=1,
                        max_value=1000,
                        step=1,
                        label_visibility="collapsed",
                        key="nonlinear_maximum_iterations",
                    )
                )

            use_damping = st.checkbox(
                "Use residual-based backtracking damping",
                value=False,
                help=(
                    "Shortens the Newton correction when the full step does not "
                    "produce a finite residual-reducing iterate."
                ),
                key="nonlinear_use_damping",
            )

            current_signature: str | None = None
            if parsed_system is not None:
                current_signature = create_input_signature(
                    equation_strings,
                    initial_guess,
                    float(tolerance),
                    int(maximum_iterations),
                    bool(use_damping),
                )

            solve_column, reset_column = st.columns(2)
            with solve_column:
                solve_button_clicked = st.button(
                    "Solve System",
                    type="primary",
                    use_container_width=True,
                    disabled=parsed_system is None,
                    key="solve_nonlinear_system_button",
                )

            with reset_column:
                reset_button_clicked = st.button(
                    "Reset Result",
                    use_container_width=True,
                    key="reset_nonlinear_system_button",
                )

            if reset_button_clicked:
                for key in (
                    "nonlinear_system_result",
                    "nonlinear_system_signature",
                    "nonlinear_system_error",
                    "nonlinear_system_error_signature",
                    "nonlinear_system_excel_report",
                    "nonlinear_system_excel_signature",
                ):
                    st.session_state.pop(key, None)
                st.rerun()

            if solve_button_clicked and parsed_system is not None:
                try:
                    nonlinear_result = solve_nonlinear_system(
                        equation_strings=equation_strings,
                        initial_guess=initial_guess,
                        tolerance=float(tolerance),
                        maximum_iterations=int(maximum_iterations),
                        use_damping=bool(use_damping),
                    )
                    st.session_state.nonlinear_system_result = nonlinear_result
                    st.session_state.nonlinear_system_signature = (
                        nonlinear_result.input_signature
                    )
                    st.session_state.pop("nonlinear_system_error", None)
                    st.session_state.pop("nonlinear_system_error_signature", None)
                except NonlinearSystemError as error:
                    st.session_state.pop("nonlinear_system_result", None)
                    st.session_state.pop("nonlinear_system_signature", None)
                    st.session_state.nonlinear_system_error = str(error)
                    st.session_state.nonlinear_system_error_signature = current_signature
                except Exception as error:
                    st.session_state.pop("nonlinear_system_result", None)
                    st.session_state.pop("nonlinear_system_signature", None)
                    st.session_state.nonlinear_system_error = (
                        "An unexpected numerical error occurred. The solver "
                        f"stopped safely. Reason: {error}"
                    )
                    st.session_state.nonlinear_system_error_signature = current_signature

                st.session_state.pop("nonlinear_system_excel_report", None)
                st.session_state.pop("nonlinear_system_excel_signature", None)
                st.rerun()

            with st.expander("Example Inputs"):
                st.code(
                    "Equation 1: x**2 + y**2 = 4\n"
                    "Equation 2: x - y = 0\n"
                    "Initial guess: x = 1.5, y = 1.0",
                    language=None,
                )
                st.code(
                    "Equation 1: x**2 + x*y - 10 = 0\n"
                    "Equation 2: y + 3*x*y**2 - 57 = 0\n"
                    "Initial guess: x = 1.5, y = 3.5",
                    language=None,
                )

    with result_column:
        with st.container(border=True):
            st.markdown(
                '<h3 class="solver-box-title">Final Result</h3>',
                unsafe_allow_html=True,
            )

            stored_result = st.session_state.get("nonlinear_system_result")
            stored_signature = st.session_state.get("nonlinear_system_signature")
            stored_error = st.session_state.get("nonlinear_system_error")
            stored_error_signature = st.session_state.get(
                "nonlinear_system_error_signature"
            )

            if current_signature is None:
                st.info("Enter a valid square nonlinear system.")
            elif (
                stored_error
                and stored_error_signature == current_signature
            ):
                st.error(stored_error)
            elif not isinstance(stored_result, NonlinearSolverResult):
                st.info("Enter the system data and click Solve System.")
            elif stored_signature != current_signature:
                st.info(
                    "The equations or numerical settings changed. "
                    "Click Solve System to calculate a new result."
                )
            else:
                render_final_result_card(stored_result)

    stored_result = st.session_state.get("nonlinear_system_result")
    stored_signature = st.session_state.get("nonlinear_system_signature")

    if (
        isinstance(stored_result, NonlinearSolverResult)
        and current_signature is not None
        and stored_signature == current_signature
    ):
        st.divider()
        st.subheader("Method Formula and Problem Setup")

        formula_columns = st.columns(2)
        with formula_columns[0]:
            st.latex(r"F(X)=0")
            st.latex(r"J(X_k)\Delta X_k=-F(X_k)")
        with formula_columns[1]:
            st.latex(r"X_{k+1}=X_k+\lambda_k\Delta X_k")
            st.latex(
                r"\frac{\|X_{k+1}-X_k\|_2}"
                r"{\max(1,\|X_{k+1}\|_2)}\leq\mathrm{tol}"
            )

        render_system_preview(stored_result.system)
        setup_dataframe = pd.DataFrame(
            {
                "Property": [
                    "Number of equations",
                    "Variables",
                    "Initial guess",
                    "Tolerance",
                    "Maximum iterations",
                    "Backtracking damping",
                ],
                "Value": [
                    stored_result.system.size,
                    ", ".join(str(v) for v in stored_result.system.variables),
                    np.array2string(
                        stored_result.initial_guess,
                        precision=12,
                        separator=", ",
                    ),
                    stored_result.tolerance,
                    stored_result.maximum_iterations,
                    "Enabled" if stored_result.damping_enabled else "Disabled",
                ],
            }
        )
        st.dataframe(
            round_numeric_dataframe(setup_dataframe),
            use_container_width=True,
            hide_index=True,
        )

        st.divider()
        render_iteration_results(stored_result)

        if stored_result.history:
            with st.expander("Detailed Newton Corrections", expanded=False):
                for record in stored_result.history:
                    st.markdown(f"**Iteration {record.iteration}**")
                    st.code(
                        "Current X = "
                        + np.array2string(record.current_x, precision=12)
                        + "\nF(Current X) = "
                        + np.array2string(record.current_function, precision=12)
                        + "\nApplied correction = "
                        + np.array2string(record.correction, precision=12)
                        + "\nNext X = "
                        + np.array2string(record.next_x, precision=12)
                        + "\nF(Next X) = "
                        + np.array2string(record.next_function, precision=12),
                        language=None,
                    )
                    st.caption(
                        "Step norm = "
                        f"{record.approximate_error:.12g}; residual norm = "
                        f"{record.residual_norm:.12g}; Jacobian condition "
                        f"number = {record.jacobian_condition_number:.12g}; "
                        f"damping factor = {record.damping_factor:.12g}."
                    )

        st.divider()
        analysis_columns = st.columns(2)
        with analysis_columns[0]:
            with st.container(border=True):
                render_error_analysis_section(stored_result)
        with analysis_columns[1]:
            with st.container(border=True):
                render_convergence_analysis_section(stored_result)

        st.divider()
        st.subheader("Jacobian Diagnostics")
        st.dataframe(
            round_numeric_dataframe(
                create_jacobian_diagnostics_dataframe(stored_result),
                decimals=6,
            ),
            use_container_width=True,
            hide_index=True,
        )

        st.divider()
        render_solver_graphs(stored_result)

        st.divider()
        render_excel_download(stored_result)

        st.divider()
        navigation_left, navigation_right = st.columns(2)
        with navigation_left:
            if st.button(
                "Review Systems of Nonlinear Equations Lesson",
                use_container_width=True,
                key="nonlinear_system_lesson_button",
            ):
                st.switch_page("pages/Systems_of_Nonlinear_Equations.py")
        with navigation_right:
            if st.button(
                "Back to Solver Menu",
                use_container_width=True,
                key="nonlinear_system_menu_button",
            ):
                st.switch_page("pages/Numerical_Solver.py")

    st.markdown("</main>", unsafe_allow_html=True)

st.html(
    """
    <footer class="footer-ui">
        <div>NM • © 2026 Numerical Methods</div>
        <div>Systems of Nonlinear Equations • Newton Method</div>
    </footer>
    """
)