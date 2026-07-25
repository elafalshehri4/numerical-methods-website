"""Professional Streamlit solver for Gauss-Legendre Gaussian Quadrature.

This page approximates a definite integral using standard Gauss-Legendre
quadrature on one interval or on several equal composite panels. It validates
all inputs, maps the tabulated nodes from [-1, 1] to each physical panel,
stores every node contribution, performs symbolic-reference error analysis
when available, studies convergence under panel refinement, creates scientific
plots, and exports a formatted Excel workbook.

The numerical quadrature is implemented directly from the textbook formula.
The solver does not use SciPy integration functions, NumPy Legendre helpers,
SymPy numerical quadrature, or any built-in Gaussian quadrature routine.
SymPy is used only to parse the integrand and, when possible, obtain an exact
symbolic reference integral for educational error analysis.

The page uses the existing Numerical Methods website styles through
``utilities.ui.load_css`` and preserves the shared Navbar, Hero, layout,
theme, buttons, and footer classes.
"""

from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css




# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Show three decimals and use × 10ⁿ when fixed notation loses meaning."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"

# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Gauss-Legendre Gaussian Quadrature"
DISPLAY_DECIMALS = 3
DEFAULT_FUNCTION = "exp(-x**2)"
DEFAULT_LOWER_LIMIT = 0.0
DEFAULT_UPPER_LIMIT = 1.0
DEFAULT_GAUSS_POINTS = 3
DEFAULT_PANELS = 1
MIN_PANELS = 1
MAX_PANELS = 200
DEFAULT_CONVERGENCE_LEVELS = 5
MIN_CONVERGENCE_LEVELS = 3
MAX_CONVERGENCE_LEVELS = 7
ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

MODE_OPTIONS = {
    "Single-Interval Gauss-Legendre": "single",
    "Composite Gauss-Legendre": "composite",
}

GAUSS_RULES: dict[int, tuple[tuple[float, ...], tuple[float, ...]]] = {
    2: (
        (
            -0.5773502691896257645091488,
            0.5773502691896257645091488,
        ),
        (1.0, 1.0),
    ),
    3: (
        (
            -0.7745966692414833770358531,
            0.0,
            0.7745966692414833770358531,
        ),
        (
            0.5555555555555555555555556,
            0.8888888888888888888888889,
            0.5555555555555555555555556,
        ),
    ),
    4: (
        (
            -0.8611363115940525752239465,
            -0.3399810435848562648026658,
            0.3399810435848562648026658,
            0.8611363115940525752239465,
        ),
        (
            0.3478548451374538573730639,
            0.6521451548625461426269361,
            0.6521451548625461426269361,
            0.3478548451374538573730639,
        ),
    ),
    5: (
        (
            -0.9061798459386639927976269,
            -0.5384693101056830910363144,
            0.0,
            0.5384693101056830910363144,
            0.9061798459386639927976269,
        ),
        (
            0.2369268850561890875142640,
            0.4786286704993664680412915,
            0.5688888888888888888888889,
            0.4786286704993664680412915,
            0.2369268850561890875142640,
        ),
    ),
}

ALLOWED_FUNCTION_NAMES = {
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class StandardGaussNode:
    """One tabulated Gauss-Legendre node and weight on [-1, 1]."""

    index: int
    standard_node: float
    weight: float


@dataclass(frozen=True)
class QuadratureSample:
    """One mapped Gauss node used in a physical integration panel."""

    global_index: int
    panel_index: int
    local_node_index: int
    panel_left: float
    panel_right: float
    midpoint: float
    half_width: float
    standard_node: float
    mapped_x: float
    weight: float
    function_value: float
    weighted_function_value: float
    signed_contribution: float
    operation: str


@dataclass(frozen=True)
class PanelContribution:
    """Summary of one physical Gauss-Legendre panel."""

    panel_index: int
    left_endpoint: float
    right_endpoint: float
    midpoint: float
    half_width: float
    weighted_sum: float
    signed_contribution: float
    cumulative_integral: float
    operation: str


@dataclass(frozen=True)
class ConvergenceRecord:
    """One panel-refinement result."""

    level: int
    panels: int
    effective_panel_width: float
    approximation: float
    exact_integral: float | None
    absolute_error: float | None
    relative_error: float | None
    successive_difference: float | None
    observed_order: float | None


@dataclass(frozen=True)
class GaussianQuadratureResult:
    """Complete result shared by Streamlit and Excel renderers."""

    status: str
    success: bool
    method: str
    mode_key: str
    mode_name: str
    message: str
    stopping_reason: str
    function_text: str
    function_expression: sp.Expr | None
    antiderivative_expression: sp.Expr | None
    antiderivative_text: str
    lower_limit: float | None
    upper_limit: float | None
    orientation: str
    gauss_points: int
    polynomial_exactness_degree: int
    panels: int
    effective_panel_width: float | None
    standard_nodes: tuple[StandardGaussNode, ...]
    samples: tuple[QuadratureSample, ...]
    panel_contributions: tuple[PanelContribution, ...]
    approximate_integral: float | None
    exact_integral: float | None
    absolute_error: float | None
    relative_error: float | None
    percentage_error: float | None
    theoretical_composite_order: int
    convergence_records: tuple[ConvergenceRecord, ...]
    latest_observed_order: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def current_report_datetime() -> datetime:
    """Return a timezone-aware report timestamp."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def format_number(
    value: float | int | None,
    decimals: int = 3,
) -> str:
    """Format displayed values with three decimals and × 10ⁿ notation."""

    return format_display_number(value, decimals)



def scientific_number(value: float | int | None) -> str:
    """Format a value as a three-decimal coefficient multiplied by 10ⁿ."""

    return format_scientific_power(value)



def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only for display."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns) > 0:
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def create_input_signature(
    function_text: str,
    lower_limit: Any,
    upper_limit: Any,
    mode_name: str,
    gauss_points: Any,
    panels: Any,
    convergence_levels: Any,
) -> str:
    """Create a stable signature used to detect stale Streamlit results."""

    payload = "|".join(
        [
            str(function_text).strip(),
            repr(lower_limit),
            repr(upper_limit),
            str(mode_name),
            repr(gauss_points),
            repr(panels),
            repr(convergence_levels),
        ]
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def safe_float(raw_value: Any, value_name: str) -> float:
    """Convert an input to one finite real float."""

    try:
        value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be a valid numerical value.") from error

    if not math.isfinite(value):
        raise ValueError(f"{value_name} must be finite; NaN and infinity are not allowed.")
    return value


def safe_positive_integer(
    raw_value: Any,
    value_name: str,
    minimum: int,
    maximum: int,
) -> int:
    """Convert an input to a bounded positive integer without silent truncation."""

    try:
        numeric = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be an integer.") from error

    if not math.isfinite(numeric) or not numeric.is_integer():
        raise ValueError(f"{value_name} must be an integer.")

    value = int(numeric)
    if value < minimum or value > maximum:
        raise ValueError(
            f"{value_name} must be between {minimum} and {maximum}."
        )
    return value


def empty_result(
    *,
    message: str,
    stopping_reason: str,
    function_text: str,
    mode_key: str,
    mode_name: str,
    gauss_points: int,
    panels: int,
    input_signature: str,
    warnings: tuple[str, ...] = (),
) -> GaussianQuadratureResult:
    """Create a complete structured error result."""

    return GaussianQuadratureResult(
        status="error",
        success=False,
        method=METHOD_NAME,
        mode_key=mode_key,
        mode_name=mode_name,
        message=message,
        stopping_reason=stopping_reason,
        function_text=function_text,
        function_expression=None,
        antiderivative_expression=None,
        antiderivative_text="Not available",
        lower_limit=None,
        upper_limit=None,
        orientation="Not available",
        gauss_points=gauss_points,
        polynomial_exactness_degree=max(2 * gauss_points - 1, 0),
        panels=panels,
        effective_panel_width=None,
        standard_nodes=(),
        samples=(),
        panel_contributions=(),
        approximate_integral=None,
        exact_integral=None,
        absolute_error=None,
        relative_error=None,
        percentage_error=None,
        theoretical_composite_order=max(2 * gauss_points, 0),
        convergence_records=(),
        latest_observed_order=None,
        warnings=warnings,
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


# =============================================================================
# Safe symbolic parsing and numerical evaluation
# =============================================================================
def parse_function(function_text: str) -> tuple[sp.Symbol, sp.Expr]:
    """Parse and validate a real single-variable mathematical expression."""

    if not isinstance(function_text, str) or not function_text.strip():
        raise ValueError("The function field cannot be empty.")

    if "=" in function_text:
        raise ValueError(
            "Enter only the expression f(x), not an equation containing '='."
        )

    x_symbol = sp.Symbol("x", real=True)
    local_dictionary = dict(ALLOWED_FUNCTION_NAMES)
    local_dictionary["x"] = x_symbol

    try:
        expression = sp.sympify(
            function_text.strip(),
            locals=local_dictionary,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The function has an invalid format. Use Python/SymPy syntax, "
            "for example sin(x), exp(-x**2), or x**4 - 2*x + 1."
        ) from error

    if isinstance(expression, (Relational, sp.logic.boolalg.Boolean)):
        raise ValueError("The input must be a numerical expression, not a relation.")

    if expression.has(AppliedUndef):
        raise ValueError("The function contains an unsupported undefined function.")

    free_symbols = expression.free_symbols
    if not free_symbols.issubset({x_symbol}):
        unexpected = ", ".join(sorted(str(symbol) for symbol in free_symbols))
        raise ValueError(f"Only the variable x is allowed. Unexpected: {unexpected}.")

    if expression.has(sp.zoo, sp.oo, -sp.oo, sp.nan):
        raise ValueError("The function contains an undefined or non-finite constant.")

    return x_symbol, expression


def build_numeric_function(
    x_symbol: sp.Symbol,
    expression: sp.Expr,
) -> Callable[[Any], Any]:
    """Build a NumPy-compatible numerical function."""

    try:
        return sp.lambdify(x_symbol, expression, modules=["numpy"])
    except (TypeError, ValueError, NameError) as error:
        raise ValueError("The function could not be converted for numerical evaluation.") from error


def evaluate_real_scalar(
    numeric_function: Callable[[Any], Any],
    x_value: float,
    value_name: str,
) -> float:
    """Evaluate one real finite function value safely."""

    try:
        with np.errstate(all="ignore"):
            raw_value = numeric_function(float(x_value))
        array = np.asarray(raw_value)
    except (TypeError, ValueError, OverflowError, ZeroDivisionError, FloatingPointError) as error:
        raise ValueError(
            f"The function could not be evaluated at {value_name} = {x_value:.12g}."
        ) from error

    if array.size != 1:
        raise ValueError(
            f"The function did not return one scalar value at {value_name} = {x_value:.12g}."
        )

    scalar = array.reshape(-1)[0]
    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)
        if abs(complex_value.imag) > ZERO_TOLERANCE:
            raise ValueError(
                f"The function is complex at {value_name} = {x_value:.12g}."
            )
        scalar = complex_value.real

    try:
        value = float(scalar)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The function value at {value_name} = {x_value:.12g} is not real."
        ) from error

    if not math.isfinite(value):
        raise ValueError(
            f"The function is undefined or non-finite at {value_name} = {x_value:.12g}."
        )
    return value


def evaluate_real_array(
    numeric_function: Callable[[Any], Any],
    x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate a real finite array for plotting."""

    try:
        with np.errstate(all="ignore"):
            raw_values = numeric_function(x_values)
        values = np.asarray(raw_values)
    except (TypeError, ValueError, OverflowError, ZeroDivisionError, FloatingPointError) as error:
        raise ValueError("The function could not be evaluated over the plotting interval.") from error

    if values.ndim == 0:
        values = np.full_like(x_values, values, dtype=np.asarray(values).dtype)
    else:
        try:
            values = np.broadcast_to(values, x_values.shape)
        except ValueError as error:
            raise ValueError("The plotted function returned an unexpected array shape.") from error

    if np.iscomplexobj(values):
        if np.any(np.abs(np.imag(values)) > ZERO_TOLERANCE):
            raise ValueError("The function becomes complex over part of the plotting range.")
        values = np.real(values)

    return np.asarray(values, dtype=float)


def calculate_reference_integral(
    expression: sp.Expr,
    x_symbol: sp.Symbol,
    lower_limit: float,
    upper_limit: float,
) -> tuple[sp.Expr | None, str, float | None, tuple[str, ...]]:
    """Attempt an exact symbolic reference integral for error analysis."""

    warnings: list[str] = []
    try:
        antiderivative = sp.integrate(expression, x_symbol)
        definite = sp.integrate(expression, (x_symbol, lower_limit, upper_limit))
    except (TypeError, ValueError, NotImplementedError, PolynomialError) as error:
        warnings.append(
            "A symbolic reference integral was unavailable; true-error metrics were omitted."
        )
        return None, "Not available", None, tuple(warnings)

    antiderivative_text = str(antiderivative)
    if isinstance(definite, sp.Integral) or definite.has(sp.Integral):
        warnings.append(
            "SymPy could not obtain a closed-form definite integral; true-error metrics were omitted."
        )
        return antiderivative, antiderivative_text, None, tuple(warnings)

    try:
        evaluated = complex(sp.N(definite, 30))
    except (TypeError, ValueError, OverflowError) as error:
        warnings.append(
            "The symbolic reference integral could not be converted to a finite real number."
        )
        return antiderivative, antiderivative_text, None, tuple(warnings)

    if abs(evaluated.imag) > ZERO_TOLERANCE or not math.isfinite(evaluated.real):
        warnings.append(
            "The symbolic reference integral is not a finite real value; true-error metrics were omitted."
        )
        return antiderivative, antiderivative_text, None, tuple(warnings)

    return antiderivative, antiderivative_text, float(evaluated.real), tuple(warnings)


# =============================================================================
# Gauss-Legendre algorithm
# =============================================================================
def standard_rule(gauss_points: int) -> tuple[StandardGaussNode, ...]:
    """Return the tabulated n-point Gauss-Legendre rule on [-1, 1]."""

    if gauss_points not in GAUSS_RULES:
        raise ValueError("The solver supports 2, 3, 4, or 5 Gauss points.")

    nodes, weights = GAUSS_RULES[gauss_points]
    return tuple(
        StandardGaussNode(index=index, standard_node=float(node), weight=float(weight))
        for index, (node, weight) in enumerate(zip(nodes, weights), start=1)
    )


def calculate_gaussian_approximation(
    numeric_function: Callable[[Any], Any],
    lower_limit: float,
    upper_limit: float,
    gauss_points: int,
    panels: int,
    *,
    store_history: bool,
) -> tuple[
    float,
    tuple[StandardGaussNode, ...],
    tuple[QuadratureSample, ...],
    tuple[PanelContribution, ...],
]:
    """Apply the standard or composite Gauss-Legendre formula manually."""

    if lower_limit == upper_limit:
        raise ValueError("The lower and upper integration limits must be different.")
    if panels < 1:
        raise ValueError("The number of panels must be at least one.")

    rule = standard_rule(gauss_points)
    boundaries = np.linspace(lower_limit, upper_limit, panels + 1, dtype=float)
    samples: list[QuadratureSample] = []
    panel_history: list[PanelContribution] = []
    cumulative_integral = 0.0
    global_index = 0

    for panel_index in range(1, panels + 1):
        left = float(boundaries[panel_index - 1])
        right = float(boundaries[panel_index])
        midpoint = 0.5 * (left + right)
        half_width = 0.5 * (right - left)
        weighted_sum = 0.0
        local_operations: list[str] = []

        for node in rule:
            global_index += 1
            mapped_x = midpoint + half_width * node.standard_node
            function_value = evaluate_real_scalar(
                numeric_function,
                mapped_x,
                f"mapped node x_{global_index}",
            )
            weighted_value = node.weight * function_value
            signed_contribution = half_width * weighted_value
            weighted_sum += weighted_value
            operation = (
                f"x = ({left:.12g} + {right:.12g})/2 + "
                f"(({right:.12g} - {left:.12g})/2)({node.standard_node:.12g}) "
                f"= {mapped_x:.12g}; "
                f"term = (({right:.12g} - {left:.12g})/2)"
                f"({node.weight:.12g})f({mapped_x:.12g}) "
                f"= {signed_contribution:.12g}"
            )
            local_operations.append(
                f"w{node.index} f(x{node.index}) = "
                f"{node.weight:.12g} × {function_value:.12g}"
            )

            if store_history:
                samples.append(
                    QuadratureSample(
                        global_index=global_index,
                        panel_index=panel_index,
                        local_node_index=node.index,
                        panel_left=left,
                        panel_right=right,
                        midpoint=midpoint,
                        half_width=half_width,
                        standard_node=node.standard_node,
                        mapped_x=mapped_x,
                        weight=node.weight,
                        function_value=function_value,
                        weighted_function_value=weighted_value,
                        signed_contribution=signed_contribution,
                        operation=operation,
                    )
                )

        panel_contribution = half_width * weighted_sum
        cumulative_integral += panel_contribution
        if not math.isfinite(cumulative_integral):
            raise ValueError("Gaussian Quadrature produced a non-finite integral value.")

        if store_history:
            panel_operation = (
                f"Panel {panel_index}: I_{panel_index} = "
                f"({right:.12g} - {left:.12g})/2 × ["
                + " + ".join(local_operations)
                + f"] = {panel_contribution:.12g}"
            )
            panel_history.append(
                PanelContribution(
                    panel_index=panel_index,
                    left_endpoint=left,
                    right_endpoint=right,
                    midpoint=midpoint,
                    half_width=half_width,
                    weighted_sum=weighted_sum,
                    signed_contribution=panel_contribution,
                    cumulative_integral=cumulative_integral,
                    operation=panel_operation,
                )
            )

    return (
        float(cumulative_integral),
        rule,
        tuple(samples),
        tuple(panel_history),
    )


def calculate_error_metrics(
    approximation: float,
    exact_integral: float | None,
) -> tuple[float | None, float | None, float | None]:
    """Calculate true-error metrics when a symbolic reference is available."""

    if exact_integral is None:
        return None, None, None

    absolute_error = abs(exact_integral - approximation)
    if abs(exact_integral) <= RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
        relative_error = None
        percentage_error = None
    else:
        relative_error = absolute_error / abs(exact_integral)
        percentage_error = 100.0 * relative_error
    return absolute_error, relative_error, percentage_error


def calculate_convergence_analysis(
    numeric_function: Callable[[Any], Any],
    lower_limit: float,
    upper_limit: float,
    gauss_points: int,
    base_panels: int,
    convergence_levels: int,
    exact_integral: float | None,
) -> tuple[ConvergenceRecord, ...]:
    """Refine the composite panel count and estimate observed convergence."""

    records: list[ConvergenceRecord] = []
    previous_approximation: float | None = None
    previous_error_measure: float | None = None

    for level in range(convergence_levels):
        panels = base_panels * (2**level)
        if panels > MAX_PANELS * (2 ** (MAX_CONVERGENCE_LEVELS - 1)):
            break

        approximation, _, _, _ = calculate_gaussian_approximation(
            numeric_function,
            lower_limit,
            upper_limit,
            gauss_points,
            panels,
            store_history=False,
        )
        panel_width = abs(upper_limit - lower_limit) / panels
        absolute_error, relative_error, _ = calculate_error_metrics(
            approximation,
            exact_integral,
        )
        successive_difference = (
            None
            if previous_approximation is None
            else abs(approximation - previous_approximation)
        )
        current_error_measure = (
            absolute_error if absolute_error is not None else successive_difference
        )

        observed_order: float | None = None
        if (
            previous_error_measure is not None
            and current_error_measure is not None
            and previous_error_measure > ZERO_TOLERANCE
            and current_error_measure > ZERO_TOLERANCE
        ):
            observed_order = math.log(previous_error_measure / current_error_measure, 2.0)

        records.append(
            ConvergenceRecord(
                level=level,
                panels=panels,
                effective_panel_width=panel_width,
                approximation=approximation,
                exact_integral=exact_integral,
                absolute_error=absolute_error,
                relative_error=relative_error,
                successive_difference=successive_difference,
                observed_order=observed_order,
            )
        )
        previous_approximation = approximation
        previous_error_measure = current_error_measure

    return tuple(records)


def solve_gaussian_quadrature(
    function_text: str,
    lower_limit_input: Any,
    upper_limit_input: Any,
    mode_name: str,
    gauss_points_input: Any,
    panels_input: Any,
    convergence_levels_input: Any,
) -> GaussianQuadratureResult:
    """Validate, solve, analyze, and package one Gaussian Quadrature problem."""

    mode_key = MODE_OPTIONS.get(mode_name, "")
    raw_points = int(gauss_points_input) if str(gauss_points_input).isdigit() else 0
    raw_panels = int(panels_input) if str(panels_input).isdigit() else 0
    input_signature = create_input_signature(
        function_text,
        lower_limit_input,
        upper_limit_input,
        mode_name,
        gauss_points_input,
        panels_input,
        convergence_levels_input,
    )

    try:
        if not mode_key:
            raise ValueError("Select a supported Gaussian Quadrature mode.")
        gauss_points = safe_positive_integer(
            gauss_points_input,
            "Number of Gauss points",
            2,
            5,
        )
        if gauss_points not in GAUSS_RULES:
            raise ValueError("The number of Gauss points must be 2, 3, 4, or 5.")

        if mode_key == "single":
            panels = 1
        else:
            panels = safe_positive_integer(
                panels_input,
                "Number of composite panels",
                MIN_PANELS,
                MAX_PANELS,
            )

        convergence_levels = safe_positive_integer(
            convergence_levels_input,
            "Convergence levels",
            MIN_CONVERGENCE_LEVELS,
            MAX_CONVERGENCE_LEVELS,
        )
        lower_limit = safe_float(lower_limit_input, "Lower limit a")
        upper_limit = safe_float(upper_limit_input, "Upper limit b")
        if lower_limit == upper_limit:
            raise ValueError("The lower and upper integration limits must be different.")

        x_symbol, expression = parse_function(function_text)
        numeric_function = build_numeric_function(x_symbol, expression)

        approximation, rule, samples, panel_history = calculate_gaussian_approximation(
            numeric_function,
            lower_limit,
            upper_limit,
            gauss_points,
            panels,
            store_history=True,
        )

        antiderivative, antiderivative_text, exact_integral, reference_warnings = (
            calculate_reference_integral(
                expression,
                x_symbol,
                lower_limit,
                upper_limit,
            )
        )
        absolute_error, relative_error, percentage_error = calculate_error_metrics(
            approximation,
            exact_integral,
        )

        convergence_records = calculate_convergence_analysis(
            numeric_function,
            lower_limit,
            upper_limit,
            gauss_points,
            panels,
            convergence_levels,
            exact_integral,
        )
        observed_orders = [
            record.observed_order
            for record in convergence_records
            if record.observed_order is not None and math.isfinite(record.observed_order)
        ]
        latest_observed_order = observed_orders[-1] if observed_orders else None

        warnings = list(reference_warnings)
        if mode_key == "single":
            warnings.append(
                "Single-interval Gaussian Quadrature uses one panel. The convergence "
                "table refines that panel into a composite rule for educational comparison."
            )
        if panels * gauss_points > 500:
            warnings.append(
                "A large number of function evaluations was requested; table rendering may be slower."
            )
        if lower_limit > upper_limit:
            warnings.append(
                "The limits are reversed. The signed mapping correctly returns the negative orientation."
            )

        orientation = "Forward (a < b)" if lower_limit < upper_limit else "Reversed (a > b)"
        exactness_degree = 2 * gauss_points - 1
        message = "Execution completed successfully."
        stopping_reason = (
            f"The {gauss_points}-point Gauss-Legendre rule was applied on "
            f"{panels} panel(s) using {panels * gauss_points} mapped function evaluation(s)."
        )

        return GaussianQuadratureResult(
            status="success",
            success=True,
            method=METHOD_NAME,
            mode_key=mode_key,
            mode_name=mode_name,
            message=message,
            stopping_reason=stopping_reason,
            function_text=function_text.strip(),
            function_expression=expression,
            antiderivative_expression=antiderivative,
            antiderivative_text=antiderivative_text,
            lower_limit=lower_limit,
            upper_limit=upper_limit,
            orientation=orientation,
            gauss_points=gauss_points,
            polynomial_exactness_degree=exactness_degree,
            panels=panels,
            effective_panel_width=abs(upper_limit - lower_limit) / panels,
            standard_nodes=rule,
            samples=samples,
            panel_contributions=panel_history,
            approximate_integral=approximation,
            exact_integral=exact_integral,
            absolute_error=absolute_error,
            relative_error=relative_error,
            percentage_error=percentage_error,
            theoretical_composite_order=2 * gauss_points,
            convergence_records=convergence_records,
            latest_observed_order=latest_observed_order,
            warnings=tuple(warnings),
            input_signature=input_signature,
            execution_datetime=current_report_datetime(),
        )

    except (ValueError, TypeError, ArithmeticError, OverflowError) as error:
        return empty_result(
            message=str(error),
            stopping_reason="The calculation stopped during input validation or numerical evaluation.",
            function_text=str(function_text),
            mode_key=mode_key,
            mode_name=mode_name,
            gauss_points=raw_points,
            panels=raw_panels,
            input_signature=input_signature,
        )


# =============================================================================
# DataFrame builders
# =============================================================================
def create_standard_rule_dataframe(
    result: GaussianQuadratureResult,
) -> pd.DataFrame:
    """Create the standard nodes and weights table."""

    return pd.DataFrame(
        [
            {
                "Node Index": node.index,
                "Standard Node t_i": node.standard_node,
                "Weight w_i": node.weight,
                "Polynomial Exactness": f"Degree ≤ {result.polynomial_exactness_degree}",
            }
            for node in result.standard_nodes
        ]
    )


def create_samples_dataframe(result: GaussianQuadratureResult) -> pd.DataFrame:
    """Create all mapped node calculations."""

    return pd.DataFrame(
        [
            {
                "Global Evaluation": sample.global_index,
                "Panel": sample.panel_index,
                "Local Node": sample.local_node_index,
                "Panel Left": sample.panel_left,
                "Panel Right": sample.panel_right,
                "Midpoint": sample.midpoint,
                "Half Width": sample.half_width,
                "Standard Node t_i": sample.standard_node,
                "Mapped x_i": sample.mapped_x,
                "Weight w_i": sample.weight,
                "f(x_i)": sample.function_value,
                "w_i f(x_i)": sample.weighted_function_value,
                "Signed Contribution": sample.signed_contribution,
                "Operation": sample.operation,
            }
            for sample in result.samples
        ]
    )


def create_panel_dataframe(result: GaussianQuadratureResult) -> pd.DataFrame:
    """Create one row per physical panel."""

    return pd.DataFrame(
        [
            {
                "Panel": panel.panel_index,
                "Left Endpoint": panel.left_endpoint,
                "Right Endpoint": panel.right_endpoint,
                "Midpoint": panel.midpoint,
                "Half Width": panel.half_width,
                "Weighted Sum": panel.weighted_sum,
                "Panel Contribution": panel.signed_contribution,
                "Cumulative Integral": panel.cumulative_integral,
                "Operation": panel.operation,
            }
            for panel in result.panel_contributions
        ]
    )


def create_convergence_dataframe(result: GaussianQuadratureResult) -> pd.DataFrame:
    """Create the panel-refinement convergence table."""

    return pd.DataFrame(
        [
            {
                "Level": record.level,
                "Panels": record.panels,
                "Effective Panel Width": record.effective_panel_width,
                "Approximation": record.approximation,
                "Exact Integral": record.exact_integral,
                "Absolute Error": record.absolute_error,
                "Relative Error": record.relative_error,
                "Successive Difference": record.successive_difference,
                "Observed Order": record.observed_order,
            }
            for record in result.convergence_records
        ]
    )


def create_error_dataframe(result: GaussianQuadratureResult) -> pd.DataFrame:
    """Create the primary error metrics table."""

    return pd.DataFrame(
        {
            "Metric": [
                "Approximate Integral",
                "Exact Integral",
                "Absolute Error",
                "Relative Error",
                "Percentage Error",
                "Latest Observed Order",
                "Expected Composite Order",
            ],
            "Value": [
                result.approximate_integral,
                result.exact_integral,
                result.absolute_error,
                result.relative_error,
                result.percentage_error,
                result.latest_observed_order,
                result.theoretical_composite_order,
            ],
        }
    )


def create_summary_dataframe(result: GaussianQuadratureResult) -> pd.DataFrame:
    """Create the Excel summary sheet."""

    warnings_text = " | ".join(result.warnings) if result.warnings else "None"
    return pd.DataFrame(
        {
            "Property": [
                "Method",
                "Mode",
                "Status",
                "Function",
                "Lower Limit",
                "Upper Limit",
                "Orientation",
                "Gauss Points per Panel",
                "Polynomial Exactness Degree",
                "Panels",
                "Function Evaluations",
                "Effective Panel Width",
                "Approximate Integral",
                "Antiderivative",
                "Exact Integral",
                "Absolute Error",
                "Relative Error",
                "Percentage Error",
                "Expected Composite Order",
                "Latest Observed Order",
                "Stopping Reason",
                "Warnings",
                "Execution Date",
            ],
            "Value": [
                result.method,
                result.mode_name,
                result.status,
                result.function_text,
                result.lower_limit,
                result.upper_limit,
                result.orientation,
                result.gauss_points,
                result.polynomial_exactness_degree,
                result.panels,
                len(result.samples),
                result.effective_panel_width,
                result.approximate_integral,
                result.antiderivative_text,
                result.exact_integral,
                result.absolute_error,
                result.relative_error,
                result.percentage_error,
                result.theoretical_composite_order,
                result.latest_observed_order,
                result.stopping_reason,
                warnings_text,
                result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
            ],
        }
    )


def create_formula_dataframe(result: GaussianQuadratureResult) -> pd.DataFrame:
    """Create the textbook formulas sheet."""

    return pd.DataFrame(
        {
            "Component": [
                "Standard Interval Rule",
                "Interval Transformation",
                "Mapped Node",
                "Mapped Integral",
                "Composite Rule",
                "Polynomial Exactness",
                "Expected Composite Error Order",
            ],
            "Formula": [
                "Integral[-1,1] g(t) dt ≈ Σ w_i g(t_i)",
                "x = (a+b)/2 + (b-a)t/2",
                "x_i = midpoint + half_width × t_i",
                "Integral[a,b] f(x) dx ≈ (b-a)/2 × Σ w_i f(x_i)",
                "Sum the mapped Gauss rule over every physical panel",
                f"An n-point rule is exact for polynomials through degree {result.polynomial_exactness_degree}",
                f"O(H^{result.theoretical_composite_order}) for sufficiently smooth f",
            ],
        }
    )


# =============================================================================
# Scientific plots
# =============================================================================
def create_quadrature_plot(result: GaussianQuadratureResult) -> Figure:
    """Create a function graph showing panels, mapped nodes, and weights."""

    if not result.success or result.function_expression is None:
        raise ValueError("A successful result is required for plotting.")

    x_symbol = sp.Symbol("x", real=True)
    numeric_function = build_numeric_function(x_symbol, result.function_expression)
    lower_limit = float(result.lower_limit)
    upper_limit = float(result.upper_limit)
    plotting_min = min(lower_limit, upper_limit)
    plotting_max = max(lower_limit, upper_limit)
    span = plotting_max - plotting_min
    margin = 0.08 * span if span > 0 else 1.0

    dense_x = np.linspace(plotting_min - margin, plotting_max + margin, 1400)
    dense_y = evaluate_real_array(numeric_function, dense_x)
    finite_mask = np.isfinite(dense_y)
    if np.count_nonzero(finite_mask) < 2:
        raise ValueError("The function has insufficient finite values for plotting.")

    figure, axis = plt.subplots(figsize=(11, 6.5))
    axis.plot(dense_x[finite_mask], dense_y[finite_mask], linewidth=2.2, label="f(x)")
    axis.axhline(0.0, linewidth=1.0)

    for panel in result.panel_contributions:
        panel_left = min(panel.left_endpoint, panel.right_endpoint)
        panel_right = max(panel.left_endpoint, panel.right_endpoint)
        panel_x = np.linspace(panel_left, panel_right, 180)
        panel_y = evaluate_real_array(numeric_function, panel_x)
        finite_panel = np.isfinite(panel_y)
        axis.fill_between(
            panel_x[finite_panel],
            0.0,
            panel_y[finite_panel],
            alpha=0.10,
        )
        axis.axvline(panel.left_endpoint, linestyle=":", linewidth=0.8)

    axis.axvline(upper_limit, linestyle=":", linewidth=0.8)

    node_x = np.array([sample.mapped_x for sample in result.samples], dtype=float)
    node_y = np.array([sample.function_value for sample in result.samples], dtype=float)
    weights = np.array([sample.weight for sample in result.samples], dtype=float)
    sizes = 45.0 + 55.0 * weights / np.max(weights)
    axis.scatter(
        node_x,
        node_y,
        s=sizes,
        zorder=5,
        label="Mapped Gauss Nodes",
    )
    for sample in result.samples:
        axis.vlines(
            sample.mapped_x,
            0.0,
            sample.function_value,
            linewidth=0.75,
            alpha=0.45,
        )

    axis.axvline(lower_limit, linestyle="--", linewidth=1.0, label=f"a = {lower_limit:.6g}")
    axis.axvline(upper_limit, linestyle="--", linewidth=1.0, label=f"b = {upper_limit:.6g}")
    axis.set_title(
        f"{result.gauss_points}-Point Gauss-Legendre Quadrature on "
        f"{result.panels} Panel(s)"
    )
    axis.set_xlabel("x")
    axis.set_ylabel("f(x)")
    axis.grid(True, alpha=0.28)
    axis.legend()
    figure.tight_layout()
    return figure


def create_convergence_plot(result: GaussianQuadratureResult) -> Figure:
    """Create a log-log panel-refinement convergence visualization."""

    if len(result.convergence_records) < 2:
        raise ValueError("At least two refinement levels are required for plotting.")

    widths = np.array(
        [record.effective_panel_width for record in result.convergence_records],
        dtype=float,
    )
    if result.exact_integral is not None:
        errors = np.array(
            [
                record.absolute_error if record.absolute_error is not None else np.nan
                for record in result.convergence_records
            ],
            dtype=float,
        )
        y_label = "Absolute Error"
    else:
        errors = np.array(
            [
                record.successive_difference
                if record.successive_difference is not None
                else np.nan
                for record in result.convergence_records
            ],
            dtype=float,
        )
        y_label = "Successive Difference"

    valid = (
        np.isfinite(widths)
        & np.isfinite(errors)
        & (widths > 0.0)
        & (errors > 0.0)
    )
    if np.count_nonzero(valid) < 2:
        raise ValueError("There are not enough positive finite errors for plotting.")

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.loglog(widths[valid], errors[valid], marker="o", linewidth=2.0)

    reference_h = widths[valid]
    order = result.theoretical_composite_order
    reference_error = errors[valid][0] * (reference_h / reference_h[0]) ** order
    axis.loglog(
        reference_h,
        reference_error,
        linestyle="--",
        linewidth=1.3,
        label=f"Reference slope {order}",
    )
    axis.invert_xaxis()
    axis.set_title("Gaussian Quadrature Panel-Refinement Convergence")
    axis.set_xlabel("Effective Panel Width H")
    axis.set_ylabel(y_label)
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


# =============================================================================
# Excel report
# =============================================================================
def apply_excel_style(workbook: Any) -> None:
    """Apply consistent professional formatting to every worksheet."""

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="16324F")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            maximum_length = 0
            for cell in column_cells:
                cell_text = "" if cell.value is None else str(cell.value)
                maximum_length = max(maximum_length, len(cell_text))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(maximum_length + 2, 12),
                65,
            )


def create_excel_report(result: GaussianQuadratureResult) -> bytes:
    """Create a formatted XLSX report entirely in memory."""

    if not result.success:
        raise ValueError("Only a successful Gaussian Quadrature result can be exported.")

    summary_df = create_summary_dataframe(result)
    formula_df = create_formula_dataframe(result)
    standard_rule_df = create_standard_rule_dataframe(result)
    samples_df = create_samples_dataframe(result)
    panel_df = create_panel_dataframe(result)
    error_df = create_error_dataframe(result)
    convergence_df = create_convergence_dataframe(result)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        formula_df.to_excel(writer, sheet_name="Method Formulas", index=False)
        standard_rule_df.to_excel(writer, sheet_name="Standard Rule", index=False)
        samples_df.to_excel(writer, sheet_name="Mapped Nodes", index=False)
        panel_df.to_excel(writer, sheet_name="Panel Contributions", index=False)
        error_df.to_excel(writer, sheet_name="Error Analysis", index=False)
        convergence_df.to_excel(writer, sheet_name="Convergence Analysis", index=False)

        workbook = writer.book
        apply_excel_style(workbook)

        if not convergence_df.empty:
            worksheet = workbook["Convergence Analysis"]
            max_row = len(convergence_df) + 1
            chart = LineChart()
            chart.title = "Gaussian Quadrature Convergence"
            chart.x_axis.title = "Refinement Level"
            chart.y_axis.title = (
                "Absolute Error"
                if result.exact_integral is not None
                else "Successive Difference"
            )
            value_column = 6 if result.exact_integral is not None else 8
            data = Reference(
                worksheet,
                min_col=value_column,
                min_row=1,
                max_row=max_row,
            )
            categories = Reference(
                worksheet,
                min_col=1,
                min_row=2,
                max_row=max_row,
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 15
            worksheet.add_chart(chart, "K2")

    output.seek(0)
    return finalize_excel_report_with_visible_charts(output.getvalue())


# =============================================================================
# Streamlit result rendering
# =============================================================================

def finalize_excel_report_with_visible_charts(report_bytes: bytes) -> bytes:
    """Place existing workbook charts on Summary so they are immediately visible.

    The report data and worksheets are preserved. If Excel chart post-processing is
    unavailable for any reason, the original report is returned unchanged rather
    than risking a damaged workbook.
    """
    try:
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_workbook

        workbook = _load_workbook(_BytesIO(report_bytes))
        if "Summary" not in workbook.sheetnames:
            return report_bytes

        summary_sheet = workbook["Summary"]
        collected_charts = list(summary_sheet._charts)
        summary_sheet._charts = []

        for worksheet in workbook.worksheets:
            if worksheet is summary_sheet:
                continue
            if worksheet._charts:
                collected_charts.extend(list(worksheet._charts))
                worksheet._charts = []

        # Put every chart in the existing Summary worksheet, below one another.
        # D2 keeps the first chart visible beside the main results when the report opens.
        for chart_index, chart in enumerate(collected_charts):
            anchor_row = 2 + chart_index * 19
            summary_sheet.add_chart(chart, f"D{anchor_row}")

        workbook.active = workbook.sheetnames.index("Summary")
        output = _BytesIO()
        workbook.save(output)
        return output.getvalue()
    except Exception:
        return report_bytes


def render_final_result(result: GaussianQuadratureResult) -> None:
    """Render the compact final-result card."""

    if not result.success:
        st.error(result.message)
        st.caption(result.stopping_reason)
        return

    st.success(result.message)
    st.markdown(f"**Mode:** {result.mode_name}")
    st.markdown(f"**Function:** `{result.function_text}`")

    metric_columns = st.columns(2)
    metric_columns[0].metric(
        "Approximate Integral",
        format_number(result.approximate_integral),
    )
    metric_columns[1].metric("Gauss Points", str(result.gauss_points))

    second_row = st.columns(2)
    second_row[0].metric("Panels", str(result.panels))
    second_row[1].metric("Function Evaluations", str(len(result.samples)))

    if result.exact_integral is not None:
        st.metric("Exact Integral", format_number(result.exact_integral))
        st.metric("Absolute Error", scientific_number(result.absolute_error))
    else:
        st.info("A closed-form exact integral was not available for comparison.")

    st.markdown(f"**Polynomial Exactness:** Degree ≤ {result.polynomial_exactness_degree}")
    st.markdown(f"**Status:** {result.status.title()}")
    st.markdown(f"**Stopping Reason:** {result.stopping_reason}")

    for warning in result.warnings:
        st.warning(warning)


def render_formula_summary(result: GaussianQuadratureResult) -> None:
    """Render the transformation and quadrature formulas."""

    st.subheader("Method Formula and Interval Transformation")
    st.dataframe(
        create_formula_dataframe(result),
        use_container_width=True,
        hide_index=True,
    )
    st.latex(r"\int_{-1}^{1} g(t)\,dt \approx \sum_{i=1}^{n} w_i g(t_i)")
    st.latex(
        r"x=\frac{a+b}{2}+\frac{b-a}{2}t,\qquad "
        r"dx=\frac{b-a}{2}\,dt"
    )
    st.latex(
        r"\int_a^b f(x)\,dx \approx \frac{b-a}{2}"
        r"\sum_{i=1}^{n} w_i f\!\left("
        r"\frac{a+b}{2}+\frac{b-a}{2}t_i\right)"
    )
    st.code(
        f"Selected rule: {result.gauss_points}-point Gauss-Legendre\n"
        f"Exact for polynomials of degree ≤ {result.polynomial_exactness_degree}\n"
        f"Panels = {result.panels}\n"
        f"Total function evaluations = {len(result.samples)}\n"
        f"Approximate integral = {result.approximate_integral:.12g}",
        language=None,
    )


def render_standard_rule(result: GaussianQuadratureResult) -> None:
    """Render the tabulated nodes and weights."""

    st.subheader("Standard Gauss-Legendre Nodes and Weights")
    st.dataframe(
        round_numeric_dataframe(create_standard_rule_dataframe(result), 12),
        use_container_width=True,
        hide_index=True,
    )
    st.info(
        f"The {result.gauss_points}-point Gauss-Legendre rule integrates every "
        f"polynomial through degree {result.polynomial_exactness_degree} exactly "
        "in exact arithmetic on one panel."
    )


def render_mapped_nodes(result: GaussianQuadratureResult) -> None:
    """Render every mapped node and weighted contribution."""

    st.subheader("Mapped Nodes and Weighted Contributions")
    samples_df = create_samples_dataframe(result)
    display_columns = [
        "Global Evaluation",
        "Panel",
        "Local Node",
        "Standard Node t_i",
        "Mapped x_i",
        "Weight w_i",
        "f(x_i)",
        "w_i f(x_i)",
        "Signed Contribution",
    ]
    st.dataframe(
        round_numeric_dataframe(samples_df[display_columns]),
        use_container_width=True,
        hide_index=True,
    )

    for panel in result.panel_contributions:
        with st.expander(
            f"Panel {panel.panel_index}: [{panel.left_endpoint:.6g}, "
            f"{panel.right_endpoint:.6g}]"
        ):
            st.code(panel.operation, language=None)
            local_samples = [
                sample for sample in result.samples if sample.panel_index == panel.panel_index
            ]
            local_df = pd.DataFrame(
                [
                    {
                        "Local Node": sample.local_node_index,
                        "t_i": sample.standard_node,
                        "x_i": sample.mapped_x,
                        "w_i": sample.weight,
                        "f(x_i)": sample.function_value,
                        "w_i f(x_i)": sample.weighted_function_value,
                        "Mapped Contribution": sample.signed_contribution,
                    }
                    for sample in local_samples
                ]
            )
            st.dataframe(
                round_numeric_dataframe(local_df),
                use_container_width=True,
                hide_index=True,
            )
            st.markdown(
                f"**Panel contribution:** {format_number(panel.signed_contribution)}  "
                f"\n**Cumulative integral:** {format_number(panel.cumulative_integral)}"
            )


def render_panel_summary(result: GaussianQuadratureResult) -> None:
    """Render one summary row per composite panel."""

    st.subheader("Panel Contribution Summary")
    panel_df = create_panel_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(panel_df.drop(columns=["Operation"])),
        use_container_width=True,
        hide_index=True,
    )


def render_error_analysis(result: GaussianQuadratureResult) -> None:
    """Render error metrics."""

    st.subheader("Error Analysis")
    st.dataframe(
        round_numeric_dataframe(create_error_dataframe(result)),
        use_container_width=True,
        hide_index=True,
    )

    if result.exact_integral is None:
        st.info(
            "Absolute and relative true errors require a closed-form symbolic "
            "reference integral. The convergence table still reports successive differences."
        )
    else:
        st.markdown(
            f"- **Absolute Error:** {scientific_number(result.absolute_error)}\n"
            f"- **Relative Error:** {scientific_number(result.relative_error)}\n"
            f"- **Percentage Error:** {format_number(result.percentage_error)}%"
        )


def render_convergence_analysis(result: GaussianQuadratureResult) -> None:
    """Render panel-refinement convergence data and graph."""

    st.subheader("Convergence Analysis")
    convergence_df = create_convergence_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(convergence_df),
        use_container_width=True,
        hide_index=True,
    )
    st.info(
        f"For a sufficiently smooth integrand, a composite {result.gauss_points}-point "
        f"Gauss-Legendre rule commonly exhibits an asymptotic global error of order "
        f"O(H^{result.theoretical_composite_order}), where H is the panel width. "
        "Observed orders may differ before the asymptotic region or when round-off dominates."
    )

    try:
        figure = create_convergence_plot(result)
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)
    except ValueError as error:
        st.warning(f"The convergence graph could not be displayed: {error}")


def render_function_graph(result: GaussianQuadratureResult) -> None:
    """Render the integrand, physical panels, and mapped Gauss nodes."""

    st.subheader("Function Graph and Gaussian Nodes")
    try:
        figure = create_quadrature_plot(result)
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)
    except ValueError as error:
        st.warning(f"The function graph could not be displayed: {error}")


def render_excel_download(result: GaussianQuadratureResult) -> None:
    """Build and render the Excel download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature
    cached_signature = st.session_state.get("gaussian_quadrature_excel_signature")

    if cached_signature != report_signature:
        try:
            st.session_state.gaussian_quadrature_excel_report = create_excel_report(result)
            st.session_state.gaussian_quadrature_excel_signature = report_signature
        except (ValueError, OSError, TypeError, ArithmeticError) as error:
            st.error(f"The Excel report could not be generated: {error}")
            return

    report_bytes = st.session_state.get("gaussian_quadrature_excel_report")
    if not report_bytes:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    filename = f"gaussian_quadrature_report_{timestamp}.xlsx"
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=filename,
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="gaussian_quadrature_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Gaussian Quadrature Streamlit page."""

    st.set_page_config(
        page_title="Gaussian Quadrature Solver | Numerical Methods",
        page_icon="∫",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">NUMERICAL INTEGRATION TOOL</div>
                <h1>Gaussian Quadrature Solver</h1>
                <p>
                    Approximate a definite integral using Gauss–Legendre
                    Quadrature with two to five optimally placed nodes per panel.
                    Review the interval transformation, standard nodes and weights,
                    every mapped contribution, error metric, convergence result,
                    scientific graph, and Excel report.
                </p>

                <div class="method-actions">
                    <a href="/Gaussian_Quadrature" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Gaussian_Quadrature_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Match the centered Bisection solver content width.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )
    
        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader('How to Write the Function')
                st.markdown(
                    """
                Enter only the mathematical expression, without `f(x) =` or an equals sign.

                - Use only **x** as the variable.
                - Powers: write `x**2`, not **x^2**.
                - Multiplication: write `2*x`, not `2x`.
                - Use lowercase functions such as **sin(x)**, **cos(x)**, **exp(x)**, **sqrt(x)**, and **log(x)**.
                - Use parentheses whenever the order of operations could be unclear.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader('Before Solving')
                st.markdown(
                    """
                - The lower and upper limits must be different.
                - The function must be finite at every mapped Gaussian node.
                - Select between **2 and 5** Gauss–Legendre points.
                - Composite mode divides the interval into equal panels and applies the same Gaussian rule to each panel.
                - An **n-point** Gauss–Legendre rule integrates polynomials up to degree **2n − 1** exactly on each panel.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])
    
        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )
    
                st.markdown(
                    '<div class="input-label-ui">Function f(x)</div>',
                    unsafe_allow_html=True,
                )
                function_text = st.text_input(
                    "Function f(x)",
                    value=DEFAULT_FUNCTION,
                    placeholder="Example: exp(-x**2), sin(x), or x**5 - 2*x + 1",
                    label_visibility="collapsed",
                    key="gaussian_quadrature_function",
                )
    
                mode_name = st.selectbox(
                    "Quadrature Mode",
                    options=list(MODE_OPTIONS.keys()),
                    index=0,
                    key="gaussian_quadrature_mode",
                )
                mode_key = MODE_OPTIONS[mode_name]
    
                limit_columns = st.columns(2)
                with limit_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Lower limit a</div>',
                        unsafe_allow_html=True,
                    )
                    lower_limit = st.number_input(
                        "Lower limit a",
                        value=DEFAULT_LOWER_LIMIT,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="gaussian_quadrature_lower_limit",
                    )
    
                with limit_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Upper limit b</div>',
                        unsafe_allow_html=True,
                    )
                    upper_limit = st.number_input(
                        "Upper limit b",
                        value=DEFAULT_UPPER_LIMIT,
                        format="%.12g",
                        label_visibility="collapsed",
                        key="gaussian_quadrature_upper_limit",
                    )
    
                settings_columns = st.columns(3)
                with settings_columns[0]:
                    st.markdown(
                        '<div class="input-label-ui">Gauss points</div>',
                        unsafe_allow_html=True,
                    )
                    gauss_points = st.selectbox(
                        "Gauss points",
                        options=[2, 3, 4, 5],
                        index=[2, 3, 4, 5].index(DEFAULT_GAUSS_POINTS),
                        label_visibility="collapsed",
                        key="gaussian_quadrature_points",
                    )
    
                with settings_columns[1]:
                    st.markdown(
                        '<div class="input-label-ui">Panels</div>',
                        unsafe_allow_html=True,
                    )
                    panels = st.number_input(
                        "Panels",
                        min_value=MIN_PANELS,
                        max_value=MAX_PANELS,
                        value=1 if mode_key == "single" else DEFAULT_PANELS,
                        step=1,
                        disabled=mode_key == "single",
                        label_visibility="collapsed",
                        key=f"gaussian_quadrature_panels_{mode_key}",
                    )
    
                with settings_columns[2]:
                    st.markdown(
                        '<div class="input-label-ui">Convergence levels</div>',
                        unsafe_allow_html=True,
                    )
                    convergence_levels = st.number_input(
                        "Convergence levels",
                        min_value=MIN_CONVERGENCE_LEVELS,
                        max_value=MAX_CONVERGENCE_LEVELS,
                        value=DEFAULT_CONVERGENCE_LEVELS,
                        step=1,
                        label_visibility="collapsed",
                        key="gaussian_quadrature_convergence_levels",
                    )
    
                st.caption(
                    f"A {gauss_points}-point rule is exact for polynomials through "
                    f"degree {2 * gauss_points - 1}."
                )
    
                current_signature = create_input_signature(
                    function_text,
                    lower_limit,
                    upper_limit,
                    mode_name,
                    gauss_points,
                    panels,
                    convergence_levels,
                )
    
                solve_clicked = st.button(
                    "Calculate Integral",
                    type="primary",
                    use_container_width=True,
                    key="gaussian_quadrature_solve_button",
                )
    
                if solve_clicked:
                    result = solve_gaussian_quadrature(
                        function_text=function_text,
                        lower_limit_input=lower_limit,
                        upper_limit_input=upper_limit,
                        mode_name=mode_name,
                        gauss_points_input=gauss_points,
                        panels_input=panels,
                        convergence_levels_input=convergence_levels,
                    )
                    st.session_state.gaussian_quadrature_result = result
                    st.session_state.gaussian_quadrature_result_signature = current_signature
                    st.session_state.pop("gaussian_quadrature_excel_report", None)
                    st.session_state.pop("gaussian_quadrature_excel_signature", None)
    
                with st.expander("Example Inputs"):
                    st.code(
                        "Function: exp(-x**2)\n"
                        "a = 0\n"
                        "b = 1\n"
                        "Gauss points = 3\n"
                        "Mode: Single-Interval Gauss-Legendre",
                        language=None,
                    )
                    st.code(
                        "Function: sin(x)\n"
                        "a = 0\n"
                        "b = pi ≈ 3.141592653589793\n"
                        "Gauss points = 4\n"
                        "Mode: Composite Gauss-Legendre\n"
                        "Panels = 2",
                        language=None,
                    )
    
        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )
    
                saved_result = st.session_state.get("gaussian_quadrature_result")
                saved_signature = st.session_state.get(
                    "gaussian_quadrature_result_signature"
                )
    
                if saved_result is None:
                    st.info("Enter the integration data and select Calculate Integral.")
                elif saved_signature != current_signature:
                    st.warning(
                        "The inputs have changed. Select Calculate Integral to update the result."
                    )
                else:
                    render_final_result(saved_result)
    
        saved_result = st.session_state.get("gaussian_quadrature_result")
        saved_signature = st.session_state.get("gaussian_quadrature_result_signature")
    
        if saved_result is not None and saved_signature == current_signature:
            if saved_result.success:
                st.divider()
                render_formula_summary(saved_result)
                st.divider()
                render_standard_rule(saved_result)
                st.divider()
                render_mapped_nodes(saved_result)
                st.divider()
                render_panel_summary(saved_result)
                st.divider()
                render_function_graph(saved_result)
                st.divider()
                render_error_analysis(saved_result)
                st.divider()
                render_convergence_analysis(saved_result)
                st.divider()
                render_excel_download(saved_result)
    
        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Gaussian Quadrature • Numerical Integration</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()
