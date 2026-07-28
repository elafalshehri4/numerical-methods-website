from __future__ import annotations

import hashlib
import math
from io import BytesIO
from typing import Any, Callable

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy import lambdify, latex, symbols
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational
from sympy.calculus.util import continuous_domain

from components.navigation import navbar
from utilities.ui import load_css
from utilities.safe_math import safe_sympify


# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0**exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Use fixed notation normally and scientific notation for tiny/large values."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (
        magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6
    ):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"


# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Bisection Method"
DISPLAY_DECIMALS = 3
FUNCTION_ZERO_TOLERANCE = 1.0e-12
DENOMINATOR_TOLERANCE = 1.0e-15
CONTINUITY_SAMPLE_COUNT = 501
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

ALLOWED_FUNCTION_NAMES = {
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}


# =============================================================================
# Input and parsing helpers
# =============================================================================
def create_input_signature(
    equation_text: str,
    left_endpoint: Any,
    right_endpoint: Any,
    stopping_criterion: Any,
    max_iterations: Any,
) -> str:
    """Create a stable signature to prevent stale Streamlit results."""

    payload = repr(
        (
            str(equation_text).strip(),
            str(left_endpoint),
            str(right_endpoint),
            str(stopping_criterion),
            str(max_iterations),
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def safe_float(raw_value: Any, value_name: str) -> float:
    """Convert one input to a finite floating-point number."""

    try:
        value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{value_name} must be a valid numerical value.") from error

    if not math.isfinite(value):
        raise ValueError(f"{value_name} must be finite; NaN and infinity are not allowed.")
    return value


def create_function(
    equation_text: str,
) -> tuple[sp.Expr, sp.Symbol, Callable[[Any], Any]]:
    """Parse a real scalar function of x and create a NumPy function."""

    if not isinstance(equation_text, str) or not equation_text.strip():
        raise ValueError("Enter a function before solving.")

    if "=" in equation_text:
        raise ValueError(
            "Enter only the expression for f(x), without an equals sign."
        )

    x_symbol = symbols("x", real=True)
    local_dictionary = dict(ALLOWED_FUNCTION_NAMES)
    local_dictionary["x"] = x_symbol

    try:
        expression = safe_sympify(
            equation_text.strip(),
            locals=local_dictionary,
            evaluate=True,
        )
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "Invalid function format. Example: x**3 - x - 2"
        ) from error

    if isinstance(expression, (sp.MatrixBase, Relational)):
        raise ValueError("The input must be a scalar function of x.")

    unexpected_symbols = expression.free_symbols.difference({x_symbol})
    if unexpected_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unexpected_symbols))
        raise ValueError(
            "Only x can be used as a variable. "
            f"Unexpected symbol(s): {names}."
        )

    if expression.atoms(AppliedUndef):
        raise ValueError("The function contains an unsupported undefined function.")

    if expression.has(sp.zoo, sp.nan, sp.oo, -sp.oo):
        raise ValueError("The function contains an undefined or non-finite constant.")

    try:
        function = lambdify(x_symbol, expression, modules=["numpy"])
    except (TypeError, ValueError) as error:
        raise ValueError("The function could not be converted to numerical form.") from error

    return sp.simplify(expression), x_symbol, function


def evaluate_function(function: Callable[[Any], Any], value: float) -> float:
    """Evaluate a function and require one finite real scalar."""

    try:
        with np.errstate(all="ignore"):
            raw_result = function(value)
    except (ArithmeticError, TypeError, ValueError, OverflowError) as error:
        raise ValueError(
            f"The function could not be evaluated at x = {value:.12g}."
        ) from error

    array = np.asarray(raw_result)
    if array.size != 1:
        raise ValueError("The function did not return a scalar value.")

    scalar = array.reshape(-1)[0]
    if np.iscomplexobj(scalar):
        complex_value = complex(scalar)
        if abs(complex_value.imag) > FUNCTION_ZERO_TOLERANCE:
            raise ValueError(
                f"The function is complex at x = {value:.12g}; "
                "this solver accepts real-valued functions only."
            )
        scalar = complex_value.real

    try:
        result = float(scalar)
    except (TypeError, ValueError) as error:
        raise ValueError("The function result could not be converted to a real number.") from error

    if not math.isfinite(result):
        raise ValueError(
            f"The function is undefined, NaN, or infinite at x = {value:.12g}."
        )
    return result


def evaluate_function_array(
    function: Callable[[Any], Any],
    x_values: np.ndarray,
) -> np.ndarray:
    """Evaluate a function over an array and replace invalid points with NaN."""

    try:
        with np.errstate(all="ignore"):
            raw_values = function(x_values)
    except Exception:
        return np.full_like(x_values, np.nan, dtype=float)

    array = np.asarray(raw_values)
    if array.ndim == 0:
        array = np.full_like(
            x_values,
            array,
            dtype=complex if np.iscomplexobj(array) else float,
        )

    try:
        array = np.broadcast_to(array, x_values.shape)
    except ValueError:
        return np.full_like(x_values, np.nan, dtype=float)

    if np.iscomplexobj(array):
        valid_imaginary = np.abs(np.imag(array)) <= FUNCTION_ZERO_TOLERANCE
        values = np.real(array).astype(float)
        values[~valid_imaginary] = np.nan
    else:
        try:
            values = array.astype(float)
        except (TypeError, ValueError):
            return np.full_like(x_values, np.nan, dtype=float)

    values[~np.isfinite(values)] = np.nan
    return values


def analyze_interval_domain(
    expression: sp.Expr,
    x_symbol: sp.Symbol,
    function: Callable[[Any], Any],
    left_endpoint: float,
    right_endpoint: float,
) -> tuple[str, ...]:
    """Require a real continuous domain on the complete closed interval.

    Symbolic analysis catches poles and excluded endpoints that numerical sampling
    can miss. Sampling remains as a second check for expressions whose symbolic
    domain cannot be fully determined.
    """

    warnings: list[str] = []
    requested_interval = sp.Interval(left_endpoint, right_endpoint)

    try:
        domain = continuous_domain(expression, x_symbol, requested_interval)
        missing_domain = requested_interval - domain
        if missing_domain != sp.EmptySet and missing_domain.is_empty is not True:
            raise ValueError(
                "The function is not continuous and real-valued throughout [a, b]. "
                f"Excluded point(s) or region: {sp.sstr(missing_domain)}. "
                "Choose an interval that does not cross a singularity or invalid domain."
            )
        if missing_domain.is_empty is None:
            warnings.append(
                "The symbolic continuity check was inconclusive; the interval was "
                "also checked numerically."
            )
    except ValueError:
        raise
    except (NotImplementedError, TypeError, AttributeError, RuntimeError):
        warnings.append(
            "SymPy could not fully determine the symbolic continuous domain, so "
            "the solver used additional numerical continuity checks."
        )

    sample_x = np.linspace(
        left_endpoint,
        right_endpoint,
        CONTINUITY_SAMPLE_COUNT,
    )
    sample_y = evaluate_function_array(function, sample_x)

    if np.count_nonzero(np.isfinite(sample_y)) != sample_y.size:
        raise ValueError(
            "The function is undefined, non-finite, or non-real at one or more "
            "points inside [a, b]. Choose an interval where the function is "
            "continuous and real-valued."
        )

    differences = np.abs(np.diff(sample_y))
    finite_differences = differences[np.isfinite(differences)]
    if finite_differences.size:
        median_difference = float(np.median(finite_differences))
        maximum_difference = float(np.max(finite_differences))
        scale = max(1.0, float(np.nanmax(np.abs(sample_y))))
        if maximum_difference > max(
            1.0e6 * max(median_difference, 1.0e-15),
            1.0e6 * scale,
        ):
            warnings.append(
                "The sampled function values change extremely sharply inside the "
                "interval. Inspect the graph for a possible discontinuity."
            )

    return tuple(warnings)


def polynomial_degree(expression: sp.Expr, x_symbol: sp.Symbol) -> int | None:
    """Return the polynomial degree when the expression is polynomial in x."""

    try:
        polynomial = sp.Poly(expression, x_symbol)
    except sp.PolynomialError:
        return None
    return int(polynomial.degree())


# =============================================================================
# Bisection algorithm
# =============================================================================
def endpoint_root_result(
    *,
    root: float,
    root_value: float,
    message: str,
    equation_text: str,
    expression: sp.Expr,
    function: Callable[[Any], Any],
    initial_interval: tuple[float, float],
    function_a: float,
    function_b: float,
    degree: int | None,
    warnings: tuple[str, ...],
    input_signature: str,
) -> dict[str, Any]:
    """Build a complete result when an endpoint is already an exact root."""

    return {
        "status": "success",
        "converged": True,
        "root": root,
        "root_value": root_value,
        "residual": abs(root_value),
        "residual_tolerance": FUNCTION_ZERO_TOLERANCE,
        "iterations": 0,
        "history": [],
        "function": function,
        "expression": expression,
        "equation": equation_text.strip(),
        "polynomial_degree": degree,
        "initial_interval": initial_interval,
        "final_interval": (root, root),
        "final_absolute_error_bound": 0.0,
        "final_approximate_relative_error": 0.0,
        "function_a": function_a,
        "function_b": function_b,
        "message": message,
        "stopping_reason": "An interval endpoint satisfies f(x) = 0 numerically.",
        "warnings": warnings,
        "input_signature": input_signature,
    }


def solve_by_bisection(
    equation_text: str,
    left_endpoint: Any,
    right_endpoint: Any,
    stopping_criterion: Any,
    max_iterations: Any,
    input_signature: str = "",
) -> dict[str, Any]:
    """Solve f(x)=0 with the textbook Bisection Method.

    The stopping criterion is the approximate percent relative error εa from
    Chapra and Canale, Eq. (5.2)/(5.3). The iteration table also reports the
    absolute bisection error bound and the residual |f(c)|.
    """

    try:
        a0 = safe_float(left_endpoint, "Left endpoint a")
        b0 = safe_float(right_endpoint, "Right endpoint b")
        es = safe_float(stopping_criterion, "Stopping criterion εs")
        try:
            imax = int(max_iterations)
        except (TypeError, ValueError) as error:
            raise ValueError("Maximum iterations must be an integer.") from error

        if a0 >= b0:
            raise ValueError("The left endpoint a must be smaller than b.")
        if es <= 0.0:
            raise ValueError("Stopping criterion εs must be greater than zero.")
        if es > 100.0:
            raise ValueError("Stopping criterion εs must not exceed 100%.")
        if imax < 1:
            raise ValueError("Maximum iterations must be at least 1.")

        expression, x_symbol, function = create_function(equation_text)
        degree = polynomial_degree(expression, x_symbol)

        function_a = evaluate_function(function, a0)
        function_b = evaluate_function(function, b0)
        warnings = analyze_interval_domain(expression, x_symbol, function, a0, b0)
        initial_function_scale = max(1.0, abs(function_a), abs(function_b))
        residual_tolerance = max(
            FUNCTION_ZERO_TOLERANCE,
            initial_function_scale * max(es / 100.0, 100.0 * np.finfo(float).eps),
        )

        if abs(function_a) <= FUNCTION_ZERO_TOLERANCE:
            return endpoint_root_result(
                root=a0,
                root_value=function_a,
                message="The left endpoint is already a root.",
                equation_text=equation_text,
                expression=expression,
                function=function,
                initial_interval=(a0, b0),
                function_a=function_a,
                function_b=function_b,
                degree=degree,
                warnings=warnings,
                input_signature=input_signature,
            )

        if abs(function_b) <= FUNCTION_ZERO_TOLERANCE:
            return endpoint_root_result(
                root=b0,
                root_value=function_b,
                message="The right endpoint is already a root.",
                equation_text=equation_text,
                expression=expression,
                function=function,
                initial_interval=(a0, b0),
                function_a=function_a,
                function_b=function_b,
                degree=degree,
                warnings=warnings,
                input_signature=input_signature,
            )

        if function_a * function_b > 0.0:
            raise ValueError(
                "The interval is invalid: f(a) and f(b) must have opposite signs, "
                "unless an endpoint is already a root."
            )

        history: list[dict[str, Any]] = []
        a = a0
        b = b0
        f_a = function_a
        f_b = function_b
        midpoint = (a + b) / 2.0
        f_midpoint = evaluate_function(function, midpoint)
        final_absolute_error_bound = abs(b - a) / 2.0
        final_relative_error: float | None = None
        final_interval = (a, b)
        converged = False
        stopping_reason = "Maximum iterations reached before the stopping criterion was satisfied."

        for iteration in range(1, imax + 1):
            current_a = a
            current_b = b
            current_f_a = f_a
            current_f_b = f_b

            midpoint = (current_a + current_b) / 2.0
            f_midpoint = evaluate_function(function, midpoint)
            residual = abs(f_midpoint)

            # Exact absolute upper bound for the current midpoint estimate.
            absolute_error_bound = abs(current_b - current_a) / 2.0

            # Eq. (5.3), equivalent to Eq. (5.2) for bisection.
            if abs(midpoint) > DENOMINATOR_TOLERANCE:
                approximate_relative_error = (
                    absolute_error_bound / abs(midpoint)
                ) * 100.0
            else:
                approximate_relative_error = None

            test_product = current_f_a * f_midpoint
            if residual <= FUNCTION_ZERO_TOLERANCE:
                next_a = midpoint
                next_b = midpoint
                next_f_a = f_midpoint
                next_f_b = f_midpoint
                action = "Root found: f(c) ≈ 0"
            elif test_product < 0.0:
                next_a = current_a
                next_b = midpoint
                next_f_a = current_f_a
                next_f_b = f_midpoint
                action = "Set b = c"
            else:
                next_a = midpoint
                next_b = current_b
                next_f_a = f_midpoint
                next_f_b = current_f_b
                action = "Set a = c"

            error_satisfied = (
                approximate_relative_error is not None
                and approximate_relative_error <= es
            )
            exact_residual_satisfied = residual <= FUNCTION_ZERO_TOLERANCE
            scaled_residual_satisfied = residual <= residual_tolerance

            history.append(
                {
                    "Iteration": iteration,
                    "a": current_a,
                    "b": current_b,
                    "c": midpoint,
                    "f(a)": current_f_a,
                    "f(b)": current_f_b,
                    "f(c)": f_midpoint,
                    "Residual |f(c)|": residual,
                    "Absolute Error Bound": absolute_error_bound,
                    "Approx. Relative Error (%)": approximate_relative_error,
                    "Next a": next_a,
                    "Next b": next_b,
                    "Next Action": action,
                }
            )

            final_absolute_error_bound = absolute_error_bound
            final_relative_error = approximate_relative_error
            final_interval = (next_a, next_b)

            if exact_residual_satisfied:
                converged = True
                stopping_reason = (
                    "Stopped because the midpoint satisfies f(c) = 0 numerically."
                )
                break

            if error_satisfied and scaled_residual_satisfied:
                converged = True
                stopping_reason = (
                    "Stopped because the approximate percent relative error εa "
                    "and the scaled residual check both satisfied their tolerances."
                )
                break

            a = next_a
            b = next_b
            f_a = next_f_a
            f_b = next_f_b

        if converged:
            message = "Root found successfully."
        elif (
            final_relative_error is not None
            and final_relative_error <= es
            and abs(f_midpoint) > residual_tolerance
        ):
            message = (
                "The interval became small, but f(c) did not approach zero. "
                "This may indicate a discontinuity or an unsuitable bracket."
            )
            stopping_reason = (
                "The relative-error criterion was reached, but the scaled residual "
                "criterion was not satisfied; the result was not accepted as a root."
            )
        else:
            message = "Maximum iterations reached. The final approximation is shown."

        return {
            "status": "success",
            "converged": converged,
            "root": midpoint,
            "root_value": f_midpoint,
            "residual": abs(f_midpoint),
            "residual_tolerance": residual_tolerance,
            "iterations": len(history),
            "history": history,
            "function": function,
            "expression": expression,
            "equation": equation_text.strip(),
            "polynomial_degree": degree,
            "initial_interval": (a0, b0),
            "final_interval": final_interval,
            "final_absolute_error_bound": final_absolute_error_bound,
            "final_approximate_relative_error": final_relative_error,
            "function_a": function_a,
            "function_b": function_b,
            "message": message,
            "stopping_reason": stopping_reason,
            "warnings": warnings,
            "input_signature": input_signature,
        }

    except ValueError as error:
        return {
            "status": "error",
            "converged": False,
            "message": str(error),
            "stopping_reason": "The calculation stopped during input validation or function evaluation.",
            "input_signature": input_signature,
        }


# =============================================================================
# DataFrames and plots
# =============================================================================
def create_empty_dataframe() -> pd.DataFrame:
    """Create an empty iteration table with all expected columns."""

    return pd.DataFrame(
        columns=[
            "Iteration",
            "a",
            "b",
            "c",
            "f(a)",
            "f(b)",
            "f(c)",
            "Residual |f(c)|",
            "Absolute Error Bound",
            "Approx. Relative Error (%)",
            "Next a",
            "Next b",
            "Next Action",
        ]
    )


def round_numeric_dataframe(
    dataframe: pd.DataFrame,
    decimals: int = DISPLAY_DECIMALS,
) -> pd.DataFrame:
    """Round numeric columns only for on-screen display."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    if len(numeric_columns):
        rounded[numeric_columns] = rounded[numeric_columns].round(decimals)
    return rounded


def create_function_figure(result: dict[str, Any]) -> plt.Figure:
    """Create the function graph on the initial interval."""

    initial_a, initial_b = result["initial_interval"]
    x_values = np.linspace(initial_a, initial_b, 600)
    y_values = evaluate_function_array(result["function"], x_values)

    if np.count_nonzero(np.isfinite(y_values)) < 2:
        raise ValueError("The function could not be plotted on the initial interval.")

    figure, axis = plt.subplots(figsize=(9, 5.2))
    axis.plot(x_values, y_values, linewidth=2, label="f(x)")
    axis.axhline(0.0, linewidth=1)
    axis.axvline(initial_a, linestyle="--", linewidth=1, label="a")
    axis.axvline(initial_b, linestyle="--", linewidth=1, label="b")
    axis.scatter(
        [result["root"]],
        [result["root_value"]],
        s=70,
        zorder=5,
        label="Estimated root",
    )
    axis.set_title("Bisection Method — Function and Estimated Root")
    axis.set_xlabel("x")
    axis.set_ylabel("f(x)")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_root_convergence_figure(result: dict[str, Any]) -> plt.Figure:
    """Create a graph of midpoint approximations by iteration."""

    history = result["history"]
    if not history:
        raise ValueError("No iterations are available for a convergence graph.")

    iterations = [row["Iteration"] for row in history]
    approximations = [row["c"] for row in history]

    figure, axis = plt.subplots(figsize=(9, 5.2))
    axis.plot(iterations, approximations, marker="o", linewidth=2, label="c")
    axis.axhline(
        result["root"],
        linestyle="--",
        linewidth=1.5,
        label="Final approximation",
    )
    axis.set_title("Bisection Method — Root Approximation")
    axis.set_xlabel("Iteration")
    axis.set_ylabel("Midpoint c")
    axis.grid(True, alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def create_error_figure(
    result: dict[str, Any],
    stopping_criterion: float,
) -> plt.Figure:
    """Create a semilog graph of textbook relative error and absolute bound."""

    history = result["history"]
    if not history:
        raise ValueError("No iterations are available for an error graph.")

    iterations = np.asarray([row["Iteration"] for row in history], dtype=float)
    relative_errors = np.asarray(
        [
            np.nan
            if row["Approx. Relative Error (%)"] is None
            else row["Approx. Relative Error (%)"]
            for row in history
        ],
        dtype=float,
    )
    absolute_bounds = np.asarray(
        [row["Absolute Error Bound"] for row in history],
        dtype=float,
    )

    figure, axis = plt.subplots(figsize=(9, 5.2))
    valid_relative = np.isfinite(relative_errors) & (relative_errors > 0.0)
    if np.any(valid_relative):
        axis.semilogy(
            iterations[valid_relative],
            relative_errors[valid_relative],
            marker="o",
            linewidth=2,
            label="Approx. relative error εa (%)",
        )
        axis.axhline(
            stopping_criterion,
            linestyle="--",
            linewidth=1.5,
            label="Stopping criterion εs (%)",
        )

    valid_absolute = np.isfinite(absolute_bounds) & (absolute_bounds > 0.0)
    if np.any(valid_absolute):
        axis.semilogy(
            iterations[valid_absolute],
            absolute_bounds[valid_absolute],
            marker="s",
            linewidth=1.7,
            label="Absolute error bound",
        )

    axis.set_title("Bisection Method — Error Convergence")
    axis.set_xlabel("Iteration")
    axis.set_ylabel("Error (log scale)")
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    figure.tight_layout()
    return figure


def figure_to_png_buffer(figure: plt.Figure) -> BytesIO:
    """Serialize one Matplotlib figure as an in-memory PNG."""

    buffer = BytesIO()
    figure.savefig(buffer, format="png", dpi=160, bbox_inches="tight")
    buffer.seek(0)
    return buffer


# =============================================================================
# Excel export
# =============================================================================
def style_excel_workbook(workbook: Any) -> None:
    """Apply consistent formatting to every worksheet."""

    header_fill = PatternFill("solid", fgColor="0D3151")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions

        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            maximum_length = 0
            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                maximum_length = max(
                    maximum_length,
                    len(str(cell.value)) if cell.value is not None else 0,
                )
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000"
            worksheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 12),
                55,
            )


def create_excel_report(
    result: dict[str, Any],
    full_dataframe: pd.DataFrame,
    stopping_criterion: float,
) -> bytes:
    """Create a complete Bisection workbook with tables, charts, and images."""

    if result["status"] != "success":
        raise ValueError("Only successful results can be exported.")

    summary_dataframe = pd.DataFrame(
        {
            "Property": [
                "Method",
                "Equation",
                "Polynomial Degree",
                "Initial a",
                "Initial b",
                "f(a)",
                "f(b)",
                "Stopping Criterion εs (%)",
                "Converged",
                "Approximate Root",
                "f(Root)",
                "Residual |f(Root)|",
                "Iterations",
                "Final Approx. Relative Error (%)",
                "Final Absolute Error Bound",
                "Final Interval a",
                "Final Interval b",
                "Stopping Reason",
                "Warnings",
                "Message",
            ],
            "Value": [
                METHOD_NAME,
                result["equation"],
                (
                    "Not a polynomial"
                    if result["polynomial_degree"] is None
                    else result["polynomial_degree"]
                ),
                result["initial_interval"][0],
                result["initial_interval"][1],
                result["function_a"],
                result["function_b"],
                stopping_criterion,
                "Yes" if result["converged"] else "No",
                result["root"],
                result["root_value"],
                result["residual"],
                result["iterations"],
                result["final_approximate_relative_error"],
                result["final_absolute_error_bound"],
                result["final_interval"][0],
                result["final_interval"][1],
                result["stopping_reason"],
                "\n".join(f"• {warning}" for warning in result["warnings"]),
                result["message"],
            ],
        }
    )

    error_dataframe = full_dataframe[
        [
            "Iteration",
            "c",
            "Residual |f(c)|",
            "Absolute Error Bound",
            "Approx. Relative Error (%)",
        ]
    ].copy()

    initial_a, initial_b = result["initial_interval"]
    graph_x = np.linspace(initial_a, initial_b, 600)
    graph_y = evaluate_function_array(result["function"], graph_x)

    max_rows = max(len(graph_x), len(full_dataframe), 1)
    plot_data = pd.DataFrame(index=range(max_rows))
    plot_data["Function x"] = pd.Series(graph_x)
    plot_data["f(x)"] = pd.Series(graph_y)
    plot_data["Iteration"] = pd.Series(
        full_dataframe.get("Iteration", pd.Series(dtype=float))
    )
    plot_data["Midpoint c"] = pd.Series(
        full_dataframe.get("c", pd.Series(dtype=float))
    )
    plot_data["Absolute Error Bound"] = pd.Series(
        full_dataframe.get("Absolute Error Bound", pd.Series(dtype=float))
    )
    plot_data["Approx. Relative Error (%)"] = pd.Series(
        full_dataframe.get(
            "Approx. Relative Error (%)",
            pd.Series(dtype=float),
        )
    )

    function_figure = create_function_figure(result)
    function_buffer = figure_to_png_buffer(function_figure)
    plt.close(function_figure)

    image_buffers: list[BytesIO] = [function_buffer]
    root_buffer: BytesIO | None = None
    error_buffer: BytesIO | None = None

    if result["history"]:
        root_figure = create_root_convergence_figure(result)
        root_buffer = figure_to_png_buffer(root_figure)
        plt.close(root_figure)
        image_buffers.append(root_buffer)

        error_figure = create_error_figure(result, stopping_criterion)
        error_buffer = figure_to_png_buffer(error_figure)
        plt.close(error_figure)
        image_buffers.append(error_buffer)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_dataframe.to_excel(writer, sheet_name="Summary", index=False)
        full_dataframe.to_excel(writer, sheet_name="Iterations", index=False)
        error_dataframe.to_excel(writer, sheet_name="Error Analysis", index=False)
        plot_data.to_excel(writer, sheet_name="Plot Data", index=False)
        pd.DataFrame(
            {
                "Formula / Check": [
                    "Midpoint",
                    "Sign-change condition",
                    "Approximate percent relative error",
                    "Absolute bisection error bound",
                    "Residual check",
                ],
                "Expression": [
                    "c = (a + b) / 2",
                    "f(a) f(b) < 0",
                    "εa = |(c_new - c_old) / c_new| × 100%",
                    "|x_true - c| ≤ (b - a) / 2",
                    "Residual = |f(c)|",
                ],
            }
        ).to_excel(writer, sheet_name="Method Formulas", index=False)

        workbook = writer.book
        plots_sheet = workbook.create_sheet("Plots")
        plots_sheet["A1"] = "Bisection Method Graphs"
        plots_sheet["A1"].font = Font(bold=True, size=15)

        function_image = ExcelImage(function_buffer)
        function_image.width = 800
        function_image.height = 450
        plots_sheet.add_image(function_image, "A3")

        if root_buffer is not None:
            root_image = ExcelImage(root_buffer)
            root_image.width = 800
            root_image.height = 450
            plots_sheet.add_image(root_image, "A28")

        if error_buffer is not None:
            error_image = ExcelImage(error_buffer)
            error_image.width = 800
            error_image.height = 450
            plots_sheet.add_image(error_image, "A53")

        plot_sheet = workbook["Plot Data"]
        function_chart = ScatterChart()
        function_chart.title = "Function on the Initial Interval"
        function_chart.x_axis.title = "x"
        function_chart.y_axis.title = "f(x)"
        function_chart.height = 8
        function_chart.width = 15
        function_chart.series.append(
            Series(
                Reference(
                    plot_sheet,
                    min_col=2,
                    min_row=2,
                    max_row=len(graph_x) + 1,
                ),
                Reference(
                    plot_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=len(graph_x) + 1,
                ),
                title="f(x)",
            )
        )
        plot_sheet.add_chart(function_chart, "H3")

        if len(full_dataframe) > 0:
            row_count = len(full_dataframe) + 1

            approximation_chart = LineChart()
            approximation_chart.title = "Root Approximation"
            approximation_chart.x_axis.title = "Iteration"
            approximation_chart.y_axis.title = "Midpoint c"
            approximation_chart.height = 8
            approximation_chart.width = 15
            approximation_chart.add_data(
                Reference(
                    plot_sheet,
                    min_col=4,
                    min_row=1,
                    max_row=row_count,
                ),
                titles_from_data=True,
            )
            approximation_chart.set_categories(
                Reference(
                    plot_sheet,
                    min_col=3,
                    min_row=2,
                    max_row=row_count,
                )
            )
            plot_sheet.add_chart(approximation_chart, "H20")

            relative_error_chart = LineChart()
            relative_error_chart.title = "Approximate Relative Error"
            relative_error_chart.x_axis.title = "Iteration"
            relative_error_chart.y_axis.title = "εa (%)"
            relative_error_chart.height = 8
            relative_error_chart.width = 15
            relative_error_chart.add_data(
                Reference(
                    plot_sheet,
                    min_col=6,
                    min_row=1,
                    max_row=row_count,
                ),
                titles_from_data=True,
            )
            relative_error_chart.set_categories(
                Reference(
                    plot_sheet,
                    min_col=3,
                    min_row=2,
                    max_row=row_count,
                )
            )
            plot_sheet.add_chart(relative_error_chart, "X3")

        style_excel_workbook(workbook)
        workbook.active = workbook.sheetnames.index("Summary")

    # Keep the buffers alive until after workbook.save() has completed.
    _ = image_buffers
    output.seek(0)
    return output.getvalue()


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the reviewed Bisection Solver without changing its page template."""

    st.set_page_config(
        page_title="Bisection Solver | Numerical Methods",
        page_icon="📘",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">BISECTION METHOD TOOL</div>
                <h1>Bisection Solver</h1>
                <p>
                    Enter a continuous function and a valid interval to find
                    one approximate root using the Bisection Method. Review
                    every iteration, analyze the error, and export the results.
                </p>

                <div class="method-actions">
                    <a href="/Bisection_Method" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Bisection_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])

    with main_area:
        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader("How to Write the Function")
                st.markdown(
                    """
                    Enter the expression for **f(x)** without an equals sign.

                    Use only **x** as the variable.

                    **Writing rules**

                    - Powers: write **x\\*\\*2**, not **x^2**
                    - Functions: write **sin(x)**, not **sin x**
                    - Multiplication: write **2\\*x**, not `2x`
                    - Use parentheses when needed: **(x + 1)\\*\\*2**
                    - Mathematical functions must be lowercase, such as
                      **sin(x)**, **cos(x)**, **sqrt(x)**, and **log(x)**
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader("Before Solving")
                st.markdown(
                    """
                    Before using the **Bisection Method**:

                    1. The function must be real and continuous on **[a, b]**.
                    2. The endpoint values must have opposite signs, unless an
                       endpoint is already a root.
                    3. The method locates **one bracketed root per run**. For a
                       cubic with several roots, use a separate bracket for each root.
                    """
                )
                st.latex(r"f(a)\,f(b)<0")
                st.markdown("The midpoint is calculated using:")
                st.latex(r"c=\frac{a+b}{2}")
                st.info(
                    "The method is guaranteed to converge when the function is "
                    "continuous and the interval contains a sign change."
                )

        input_column, result_column = st.columns(2)

        with input_column:
            with st.container(border=True):
                st.subheader("Input")

                equation = st.text_input(
                    "Function f(x)",
                    value="x**3 - x - 2",
                    help=(
                        "Examples: x**3 - x - 2, sin(x) - 0.5, "
                        "sqrt(x) - 2"
                    ),
                    key="bisection_function",
                )

                endpoint_column1, endpoint_column2 = st.columns(2)
                with endpoint_column1:
                    a = st.number_input(
                        "Left endpoint a",
                        value=1.0,
                        key="bisection_left_endpoint",
                    )
                with endpoint_column2:
                    b = st.number_input(
                        "Right endpoint b",
                        value=2.0,
                        key="bisection_right_endpoint",
                    )

                settings_column1, settings_column2 = st.columns(2)
                with settings_column1:
                    stopping_criterion = st.number_input(
                        "Stopping criterion εs (%)",
                        value=0.01,
                        min_value=0.00000001,
                        max_value=100.0,
                        format="%.8f",
                        key="bisection_stopping_criterion",
                    )
                with settings_column2:
                    max_iterations = st.number_input(
                        "Maximum iterations",
                        value=100,
                        min_value=1,
                        step=1,
                        key="bisection_max_iterations",
                    )

                st.markdown("#### Endpoint Check")
                try:
                    _, _, preview_function = create_function(equation)
                    preview_f_a = evaluate_function(preview_function, float(a))
                    preview_f_b = evaluate_function(preview_function, float(b))

                    endpoint_result1, endpoint_result2 = st.columns(2)
                    endpoint_result1.metric("f(a)", format_display_number(preview_f_a))
                    endpoint_result2.metric("f(b)", format_display_number(preview_f_b))

                    if a >= b:
                        st.warning("The value of a must be smaller than b.")
                    elif abs(preview_f_a) <= FUNCTION_ZERO_TOLERANCE:
                        st.success("The left endpoint is already a root.")
                    elif abs(preview_f_b) <= FUNCTION_ZERO_TOLERANCE:
                        st.success("The right endpoint is already a root.")
                    elif preview_f_a * preview_f_b < 0.0:
                        st.success(
                            "Valid sign-change interval. The function must also be "
                            "continuous throughout [a, b]."
                        )
                    else:
                        st.warning(
                            "Invalid bracket: f(a) and f(b) do not have opposite signs."
                        )
                except Exception as error:
                    st.warning(f"The function preview is unavailable. {error}")

                solve_column, reset_column = st.columns(2)
                with solve_column:
                    solve_button = st.button(
                        "Solve",
                        type="primary",
                        use_container_width=True,
                        key="bisection_solve_button",
                    )
                with reset_column:
                    reset_button = st.button(
                        "Reset",
                        use_container_width=True,
                        key="bisection_reset_button",
                    )

        current_input_signature = create_input_signature(
            equation,
            a,
            b,
            stopping_criterion,
            max_iterations,
        )

        if reset_button:
            st.session_state.pop("bisection_result", None)
            st.session_state.pop("bisection_excel_report", None)
            st.session_state.pop("bisection_excel_signature", None)
            st.rerun()

        if solve_button:
            st.session_state.bisection_result = solve_by_bisection(
                equation_text=equation,
                left_endpoint=a,
                right_endpoint=b,
                stopping_criterion=stopping_criterion,
                max_iterations=max_iterations,
                input_signature=current_input_signature,
            )
            st.session_state.pop("bisection_excel_report", None)
            st.session_state.pop("bisection_excel_signature", None)
            st.rerun()

        with result_column:
            with st.container(border=True):
                st.subheader("Final Result")
                stored_result = st.session_state.get("bisection_result")

                if stored_result is None:
                    st.info("Enter the function and interval, then click Solve.")
                elif stored_result.get("input_signature") != current_input_signature:
                    st.info(
                        "The function or numerical parameters have changed. "
                        "Click Solve to calculate a new result."
                    )
                elif stored_result["status"] == "error":
                    st.error(stored_result["message"])
                    st.markdown(
                        """
                        **Common mistakes**

                        - Writing **x^2** instead of **x\\*\\*2**
                        - Writing `2x` instead of **2\\*x**
                        - Writing **sin x** instead of **sin(x)**
                        - Choosing **a ≥ b**
                        - Choosing endpoints without a sign change
                        - Bracketing a discontinuity rather than a root
                        """
                    )
                else:
                    result = stored_result
                    if result["converged"]:
                        st.success(result["message"])
                    else:
                        st.warning(result["message"])

                    st.latex(rf"f(x)={latex(result['expression'])}")

                    metric_column1, metric_column2 = st.columns(2)
                    metric_column1.metric(
                        "Approximate Root",
                        format_display_number(result["root"], 6),
                    )
                    metric_column2.metric("Iterations", result["iterations"])

                    metric_column3, metric_column4 = st.columns(2)
                    metric_column3.metric(
                        "Approx. Relative Error",
                        (
                            "—"
                            if result["final_approximate_relative_error"] is None
                            else f"{format_display_number(result['final_approximate_relative_error'], 6)}%"
                        ),
                    )
                    metric_column4.metric(
                        "Residual |f(root)|",
                        format_scientific_power(result["residual"]),
                    )

                    final_a, final_b = result["final_interval"]
                    st.info(
                        "Final bracket: "
                        f"[{format_display_number(final_a, 6)}, "
                        f"{format_display_number(final_b, 6)}]  •  "
                        "Absolute error bound: "
                        f"{format_scientific_power(result['final_absolute_error_bound'])}"
                    )
                    st.caption(result["stopping_reason"])

                    for warning in result["warnings"]:
                        st.warning(warning)

        active_result = st.session_state.get("bisection_result")
        active = (
            active_result is not None
            and active_result.get("input_signature") == current_input_signature
            and active_result.get("status") == "success"
        )

        if active:
            result = active_result
            history = result["history"]

            with st.container(border=True):
                st.subheader("Iteration Table")
                if history:
                    full_dataframe = pd.DataFrame(history)
                    display_dataframe = round_numeric_dataframe(full_dataframe)
                    table_html = display_dataframe.to_html(
                        index=False,
                        classes="iteration-table",
                        border=0,
                    )
                    st.markdown(
                        f'<div class="iteration-table-wrapper">{table_html}</div>',
                        unsafe_allow_html=True,
                    )
                    st.caption(
                        "Calculations and Excel export retain full precision. "
                        "Only the on-screen table is rounded for readability."
                    )
                else:
                    full_dataframe = create_empty_dataframe()
                    st.info(
                        "No iterations were required because an endpoint was already a root."
                    )

            with st.container(border=True):
                st.subheader("Error Analysis")

                error_metric1, error_metric2, error_metric3, error_metric4 = st.columns(4)
                error_metric1.metric(
                    "Initial εa",
                    (
                        "—"
                        if not history
                        or history[0]["Approx. Relative Error (%)"] is None
                        else f"{format_display_number(history[0]['Approx. Relative Error (%)'], 6)}%"
                    ),
                )
                error_metric2.metric(
                    "Final εa",
                    (
                        "—"
                        if result["final_approximate_relative_error"] is None
                        else f"{format_display_number(result['final_approximate_relative_error'], 6)}%"
                    ),
                )
                error_metric3.metric(
                    "Absolute Error Bound",
                    format_scientific_power(result["final_absolute_error_bound"]),
                )
                error_metric4.metric(
                    "Requested εs",
                    f"{format_display_number(stopping_criterion, 6)}%",
                )

                st.markdown("The textbook stopping error is:")
                st.latex(
                    r"\varepsilon_a="
                    r"\left|\frac{x_r^{\mathrm{new}}-x_r^{\mathrm{old}}}"
                    r"{x_r^{\mathrm{new}}}\right|100\%"
                )
                st.markdown("The exact bisection error bound is:")
                st.latex(
                    r"|x_{\mathrm{true}}-c_n|\leq\frac{b_n-a_n}{2}"
                )

                if result["converged"]:
                    st.success("The stopping condition was satisfied.")
                else:
                    st.warning(
                        "Maximum iterations were reached before εa satisfied εs."
                    )

            with st.container(border=True):
                st.subheader("Download Reports")
                st.write(
                    "Download the complete full-precision results, formulas, "
                    "tables, native Excel charts, and embedded graph images."
                )

                report_signature = result["input_signature"]
                if st.session_state.get("bisection_excel_signature") != report_signature:
                    try:
                        st.session_state.bisection_excel_report = create_excel_report(
                            result,
                            full_dataframe,
                            float(stopping_criterion),
                        )
                        st.session_state.bisection_excel_signature = report_signature
                    except Exception as error:
                        st.session_state.pop("bisection_excel_report", None)
                        st.error(f"The Excel report could not be generated. {error}")

                excel_data = st.session_state.get("bisection_excel_report")
                if excel_data is not None:
                    st.download_button(
                        label="Download Complete Excel Report",
                        data=excel_data,
                        file_name="bisection_complete_report.xlsx",
                        mime=EXCEL_MIME_TYPE,
                        use_container_width=True,
                        key="bisection_excel_download",
                    )

            with st.container(border=True):
                st.subheader("Graphs and Convergence Analysis")
                st.write("Examine the function, root approximations, and errors.")

            function_column, root_column, error_column = st.columns(3)

            with function_column:
                with st.container(border=True):
                    st.subheader("Function Graph")
                    try:
                        figure1 = create_function_figure(result)
                    except ValueError as error:
                        st.warning(str(error))
                    else:
                        figure1.set_size_inches(3.4, 2.6)
                        st.pyplot(figure1, use_container_width=True)
                        plt.close(figure1)

            with root_column:
                with st.container(border=True):
                    st.subheader("Root Convergence")
                    if history:
                        figure2 = create_root_convergence_figure(result)
                        figure2.set_size_inches(3.4, 2.6)
                        st.pyplot(figure2, use_container_width=True)
                        plt.close(figure2)
                    else:
                        st.info("No convergence graph was required.")

            with error_column:
                with st.container(border=True):
                    st.subheader("Error Graph")
                    if history:
                        figure3 = create_error_figure(
                            result,
                            float(stopping_criterion),
                        )
                        figure3.set_size_inches(3.4, 2.6)
                        st.pyplot(figure3, use_container_width=True)
                        plt.close(figure3)
                    else:
                        st.info("No error graph was required.")

            with st.container(border=True):
                st.subheader("Continue Learning")
                navigation_column1, navigation_column2 = st.columns(2)
                with navigation_column1:
                    if st.button(
                        "Review Bisection Lesson",
                        use_container_width=True,
                        key="review_bisection_lesson",
                    ):
                        st.switch_page("pages/Bisection_Method.py")
                with navigation_column2:
                    if st.button(
                        "Back to Solver Menu",
                        use_container_width=True,
                        key="back_to_solver_menu_bisection",
                    ):
                        st.switch_page("pages/Numerical_Solver.py")

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Bisection Solver • Root Finding</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()