from __future__ import annotations

import base64
import hashlib
import math
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
import sympy as sp
from matplotlib.figure import Figure
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sympy.core.function import AppliedUndef
from sympy.core.relational import Relational

from components.navigation import navbar
from utilities.ui import load_css




# =============================================================================
# Consistent numerical display formatting
# =============================================================================
_SUPERSCRIPT_TRANSLATION = str.maketrans(
    "0123456789-+",
    "⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺",
)


def format_scientific_power(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Format scientific notation as a coefficient multiplied by 10ⁿ."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)
    if number == 0.0:
        return f"{0.0:.{decimals}f}"

    exponent = int(math.floor(math.log10(abs(number))))
    mantissa = number / (10.0 ** exponent)
    exponent_text = str(exponent).translate(_SUPERSCRIPT_TRANSLATION)
    return f"{mantissa:.{decimals}f} × 10{exponent_text}"


def format_display_number(
    value: float | int | None,
    decimals: int = 3,
    unavailable: str = "—",
) -> str:
    """Show three decimals and use × 10ⁿ when fixed notation loses meaning."""

    if value is None:
        return unavailable
    number = float(value)
    if not math.isfinite(number):
        return str(number)

    magnitude = abs(number)
    if magnitude != 0.0 and (magnitude < 10.0 ** (-decimals) or magnitude >= 1.0e6):
        return format_scientific_power(number, decimals, unavailable)
    return f"{number:.{decimals}f}"

# =============================================================================
# Constants
# =============================================================================
METHOD_NAME = "Fourth-Order Runge–Kutta Method (Classical RK4)"
DISPLAY_DECIMALS = 3
DEFAULT_ODE = "x + y"
DEFAULT_X0 = 0.0
DEFAULT_Y0 = 1.0
DEFAULT_X_END = 1.0
DEFAULT_STEPS = 10
MIN_STEPS = 1
MAX_STEPS = 10000
DEFAULT_CONVERGENCE_LEVELS = 5
MIN_CONVERGENCE_LEVELS = 3
MAX_CONVERGENCE_LEVELS = 7
ZERO_TOLERANCE = 1.0e-15
RELATIVE_ERROR_DENOMINATOR_TOLERANCE = 1.0e-15
VALUE_MAGNITUDE_WARNING = 1.0e12
REPORT_TIME_ZONE = "Asia/Riyadh"
EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

X_SYMBOL = sp.Symbol("x", real=True)
Y_SYMBOL = sp.Symbol("y", real=True)

ALLOWED_FUNCTION_NAMES = {
    "x": X_SYMBOL,
    "y": Y_SYMBOL,
    "sin": sp.sin,
    "cos": sp.cos,
    "tan": sp.tan,
    "asin": sp.asin,
    "acos": sp.acos,
    "atan": sp.atan,
    "sinh": sp.sinh,
    "cosh": sp.cosh,
    "tanh": sp.tanh,
    "exp": sp.exp,
    "log": sp.log,
    "ln": sp.log,
    "sqrt": sp.sqrt,
    "Abs": sp.Abs,
    "abs": sp.Abs,
    "pi": sp.pi,
    "E": sp.E,
}

EXACT_SOLUTION_NAMES = {
    key: value for key, value in ALLOWED_FUNCTION_NAMES.items() if key != "y"
}


# =============================================================================
# Structured data models
# =============================================================================
@dataclass(frozen=True)
class RK4Iteration:
    """One classical fourth-order Runge-Kutta step."""

    iteration: int
    x_n: float
    y_n: float
    step_size: float
    k1: float
    k2_x: float
    k2_y: float
    k2: float
    k3_x: float
    k3_y: float
    k3: float
    k4_x: float
    k4_y: float
    k4: float
    weighted_slope: float
    increment: float
    x_next: float
    y_next: float
    step_change: float
    exact_y_next: float | None
    signed_error: float | None
    absolute_error: float | None
    relative_error: float | None
    operation: str
    status: str


@dataclass(frozen=True)
class ConvergenceRecord:
    """One grid-refinement result for the classical RK4 method."""

    level: int
    steps: int
    step_size: float
    final_approximation: float
    exact_final_value: float | None
    absolute_error: float | None
    successive_difference: float | None
    observed_order: float | None


@dataclass(frozen=True)
class RK4Result:
    """Complete RK4 result shared by Streamlit and Excel renderers."""

    status: str
    success: bool
    method: str
    message: str
    stopping_reason: str
    ode_text: str
    ode_expression: sp.Expr | None
    exact_solution_text: str
    exact_solution_expression: sp.Expr | None
    x0: float | None
    y0: float | None
    x_end: float | None
    steps: int
    step_size: float | None
    direction: str
    iterations: tuple[RK4Iteration, ...]
    final_x: float | None
    final_y: float | None
    exact_final_y: float | None
    signed_final_error: float | None
    absolute_final_error: float | None
    relative_final_error: float | None
    maximum_absolute_error: float | None
    rmse: float | None
    convergence_records: tuple[ConvergenceRecord, ...]
    latest_observed_order: float | None
    warnings: tuple[str, ...]
    input_signature: str
    execution_datetime: datetime


# =============================================================================
# General helpers
# =============================================================================
def image_to_base64(image_path: str | Path) -> str:
    """Return a file as Base64 text, or an empty string if unavailable."""

    path = Path(image_path)
    if not path.exists() or not path.is_file():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def current_report_datetime() -> datetime:
    """Return the current date and time in the report time zone."""

    return datetime.now(ZoneInfo(REPORT_TIME_ZONE))


def create_input_signature(
    ode_text: str,
    x0: float,
    y0: float,
    x_end: float,
    steps: int,
    exact_solution_text: str,
    convergence_levels: int,
) -> str:
    """Create a stable signature used to detect stale Streamlit results."""

    payload = "|".join(
        [
            ode_text.strip(),
            f"{float(x0):.17g}",
            f"{float(y0):.17g}",
            f"{float(x_end):.17g}",
            str(int(steps)),
            exact_solution_text.strip(),
            str(int(convergence_levels)),
        ]
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def format_number(
    value: float | int | None,
    decimals: int = 3,
) -> str:
    """Format displayed values with three decimals and × 10ⁿ notation."""

    return format_display_number(value, decimals)



def validate_finite_real(value: Any, value_name: str) -> float:
    """Convert a scalar to a finite real float or raise ValueError."""

    try:
        complex_value = complex(value)
    except (TypeError, ValueError, OverflowError) as error:
        raise ValueError(f"{value_name} is not a valid numerical value.") from error

    if abs(complex_value.imag) > ZERO_TOLERANCE:
        raise ValueError(f"{value_name} is complex; a real value is required.")

    real_value = float(complex_value.real)
    if not math.isfinite(real_value):
        raise ValueError(f"{value_name} is NaN or infinity.")
    return real_value


def round_numeric_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Return a display copy with numeric columns rounded to six decimals."""

    rounded = dataframe.copy()
    numeric_columns = rounded.select_dtypes(include=[np.number]).columns
    rounded[numeric_columns] = rounded[numeric_columns].round(DISPLAY_DECIMALS)
    return rounded


def serialize_array(values: np.ndarray | list[float] | tuple[float, ...]) -> str:
    """Serialize a numerical sequence for Excel summary cells."""

    array = np.asarray(values, dtype=float).reshape(-1)
    return "[" + ", ".join(f"{value:.15g}" for value in array) + "]"


# =============================================================================
# Safe symbolic parsing
# =============================================================================
def _reject_unsupported_symbolic_constructs(
    expression: sp.Expr,
    allowed_symbols: set[sp.Symbol],
    expression_name: str,
) -> None:
    """Reject symbolic constructs inappropriate for numerical evaluation."""

    unsupported_atoms = (
        AppliedUndef,
        sp.Derivative,
        sp.Integral,
        sp.Sum,
        sp.Product,
        sp.Limit,
    )

    if expression.has(*unsupported_atoms):
        raise ValueError(
            f"{expression_name} contains unsupported symbolic operations."
        )

    if isinstance(expression, Relational) or expression.has(Relational):
        raise ValueError(
            f"{expression_name} must be an expression, not an equation or inequality."
        )

    unexpected_symbols = expression.free_symbols.difference(allowed_symbols)
    if unexpected_symbols:
        names = ", ".join(sorted(str(symbol) for symbol in unexpected_symbols))
        raise ValueError(
            f"{expression_name} contains unsupported variable(s): {names}."
        )

    if expression.has(sp.zoo, sp.oo, -sp.oo, sp.nan):
        raise ValueError(f"{expression_name} contains a non-finite symbolic value.")


def parse_ode_expression(ode_text: str) -> tuple[sp.Expr, Callable[[float, float], Any]]:
    """Parse y' = f(x, y) and return the expression and numerical function."""

    if not isinstance(ode_text, str) or not ode_text.strip():
        raise ValueError("The differential equation f(x, y) cannot be empty.")

    text = ode_text.strip().replace("^", "**")
    if "=" in text:
        raise ValueError(
            "Enter only the right-hand side f(x, y), not an equation with '='."
        )

    try:
        expression = sp.sympify(text, locals=ALLOWED_FUNCTION_NAMES, evaluate=True)
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The differential equation has an invalid format. Use SymPy syntax, "
            "for example: x + y, y - x**2 + 1, or sin(x) - y."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError("The differential equation could not be interpreted.")

    _reject_unsupported_symbolic_constructs(
        expression,
        allowed_symbols={X_SYMBOL, Y_SYMBOL},
        expression_name="The differential equation",
    )

    try:
        numerical_function = sp.lambdify(
            (X_SYMBOL, Y_SYMBOL), expression, modules=["numpy"]
        )
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The differential equation could not be converted to a numerical function."
        ) from error

    return expression, numerical_function


def parse_exact_solution(
    exact_solution_text: str,
) -> tuple[sp.Expr | None, Callable[[float], Any] | None]:
    """Parse an optional exact solution y(x) supplied by the user."""

    if not exact_solution_text.strip():
        return None, None

    text = exact_solution_text.strip().replace("^", "**")
    if "=" in text:
        left, right = text.split("=", maxsplit=1)
        if left.strip().lower() not in {"y", "y(x)"}:
            raise ValueError(
                "For the exact solution, enter only y(x) or use the form y = expression."
            )
        text = right.strip()

    try:
        expression = sp.sympify(text, locals=EXACT_SOLUTION_NAMES, evaluate=True)
    except (sp.SympifyError, TypeError, ValueError, SyntaxError) as error:
        raise ValueError(
            "The exact solution has an invalid format. Example: 2*exp(x) - x - 1."
        ) from error

    if not isinstance(expression, sp.Expr):
        raise ValueError("The exact solution could not be interpreted.")

    _reject_unsupported_symbolic_constructs(
        expression,
        allowed_symbols={X_SYMBOL},
        expression_name="The exact solution",
    )

    try:
        numerical_function = sp.lambdify(X_SYMBOL, expression, modules=["numpy"])
    except (TypeError, ValueError, NameError) as error:
        raise ValueError(
            "The exact solution could not be converted to a numerical function."
        ) from error

    return expression, numerical_function


def evaluate_ode(
    numerical_function: Callable[[float, float], Any],
    x_value: float,
    y_value: float,
) -> float:
    """Evaluate f(x, y) safely at one point."""

    try:
        raw_value = numerical_function(x_value, y_value)
    except (ValueError, TypeError, OverflowError, ZeroDivisionError, FloatingPointError) as error:
        raise ValueError(
            f"The differential equation is undefined at x = {x_value:.12g}, "
            f"y = {y_value:.12g}. Reason: {error}"
        ) from error

    return validate_finite_real(
        raw_value,
        f"f({x_value:.12g}, {y_value:.12g})",
    )


def evaluate_exact_solution(
    numerical_function: Callable[[float], Any],
    x_value: float,
) -> float:
    """Evaluate an optional exact solution safely."""

    try:
        raw_value = numerical_function(x_value)
    except (ValueError, TypeError, OverflowError, ZeroDivisionError, FloatingPointError) as error:
        raise ValueError(
            f"The exact solution is undefined at x = {x_value:.12g}. Reason: {error}"
        ) from error

    return validate_finite_real(raw_value, f"Exact solution at x = {x_value:.12g}")


# =============================================================================
# Classical RK4 algorithm
# =============================================================================
def run_rk4_grid(
    ode_function: Callable[[float, float], Any],
    x0: float,
    y0: float,
    x_end: float,
    steps: int,
    exact_function: Callable[[float], Any] | None = None,
    store_history: bool = True,
) -> tuple[float, tuple[RK4Iteration, ...]]:
    """Run the classical fourth-order Runge-Kutta method on a uniform grid.

    For each step:

        k1 = f(x_n, y_n)
        k2 = f(x_n + h/2, y_n + h*k1/2)
        k3 = f(x_n + h/2, y_n + h*k2/2)
        k4 = f(x_n + h, y_n + h*k3)
        y_(n+1) = y_n + h*(k1 + 2*k2 + 2*k3 + k4)/6

    No library ODE solver is used.
    """

    if steps < 1:
        raise ValueError("The number of steps must be at least 1.")

    step_size = (x_end - x0) / steps
    if not math.isfinite(step_size) or abs(step_size) <= ZERO_TOLERANCE:
        raise ValueError("The computed step size is zero or numerically unusable.")

    current_x = float(x0)
    current_y = float(y0)
    history: list[RK4Iteration] = []

    for iteration in range(1, steps + 1):
        k1 = evaluate_ode(ode_function, current_x, current_y)

        k2_x = current_x + 0.5 * step_size
        k2_y = current_y + 0.5 * step_size * k1
        if not all(math.isfinite(value) for value in (k2_x, k2_y)):
            raise ValueError(
                "RK4 produced a non-finite second-stage point "
                f"at iteration {iteration}."
            )
        k2 = evaluate_ode(ode_function, k2_x, k2_y)

        k3_x = current_x + 0.5 * step_size
        k3_y = current_y + 0.5 * step_size * k2
        if not all(math.isfinite(value) for value in (k3_x, k3_y)):
            raise ValueError(
                "RK4 produced a non-finite third-stage point "
                f"at iteration {iteration}."
            )
        k3 = evaluate_ode(ode_function, k3_x, k3_y)

        k4_x = current_x + step_size
        k4_y = current_y + step_size * k3
        if not all(math.isfinite(value) for value in (k4_x, k4_y)):
            raise ValueError(
                "RK4 produced a non-finite fourth-stage point "
                f"at iteration {iteration}."
            )
        k4 = evaluate_ode(ode_function, k4_x, k4_y)

        weighted_slope = (k1 + 2.0 * k2 + 2.0 * k3 + k4) / 6.0
        increment = step_size * weighted_slope
        next_x = x0 + iteration * step_size
        next_y = current_y + increment

        if not all(
            math.isfinite(value)
            for value in (weighted_slope, increment, next_x, next_y)
        ):
            raise ValueError(
                "RK4 produced a non-finite solution approximation "
                f"at iteration {iteration}."
            )

        exact_y_next: float | None = None
        signed_error: float | None = None
        absolute_error: float | None = None
        relative_error: float | None = None

        if exact_function is not None:
            exact_y_next = evaluate_exact_solution(exact_function, next_x)
            signed_error = next_y - exact_y_next
            absolute_error = abs(signed_error)
            if abs(exact_y_next) > RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
                relative_error = absolute_error / abs(exact_y_next)

        if store_history:
            history.append(
                RK4Iteration(
                    iteration=iteration,
                    x_n=float(current_x),
                    y_n=float(current_y),
                    step_size=float(step_size),
                    k1=float(k1),
                    k2_x=float(k2_x),
                    k2_y=float(k2_y),
                    k2=float(k2),
                    k3_x=float(k3_x),
                    k3_y=float(k3_y),
                    k3=float(k3),
                    k4_x=float(k4_x),
                    k4_y=float(k4_y),
                    k4=float(k4),
                    weighted_slope=float(weighted_slope),
                    increment=float(increment),
                    x_next=float(next_x),
                    y_next=float(next_y),
                    step_change=float(abs(next_y - current_y)),
                    exact_y_next=exact_y_next,
                    signed_error=signed_error,
                    absolute_error=absolute_error,
                    relative_error=relative_error,
                    operation=(
                        f"k1 = f({current_x:.12g}, {current_y:.12g}) = {k1:.12g}; "
                        f"k2 = f({k2_x:.12g}, {k2_y:.12g}) = {k2:.12g}; "
                        f"k3 = f({k3_x:.12g}, {k3_y:.12g}) = {k3:.12g}; "
                        f"k4 = f({k4_x:.12g}, {k4_y:.12g}) = {k4:.12g}; "
                        f"y_{iteration} = {current_y:.12g} + "
                        f"({step_size:.12g}/6)({k1:.12g} + 2({k2:.12g}) + "
                        f"2({k3:.12g}) + {k4:.12g}) = {next_y:.12g}"
                    ),
                    status="Completed",
                )
            )

        current_x = float(next_x)
        current_y = float(next_y)

    return current_y, tuple(history)

def build_convergence_analysis(
    ode_function: Callable[[float, float], Any],
    exact_function: Callable[[float], Any] | None,
    x0: float,
    y0: float,
    x_end: float,
    base_steps: int,
    levels: int,
) -> tuple[ConvergenceRecord, ...]:
    """Refine the grid and estimate the expected fourth-order convergence."""

    raw_records: list[dict[str, float | int | None]] = []

    for level in range(1, levels + 1):
        steps = base_steps * (2 ** (level - 1))
        final_y, _ = run_rk4_grid(
            ode_function=ode_function,
            x0=x0,
            y0=y0,
            x_end=x_end,
            steps=steps,
            exact_function=None,
            store_history=False,
        )

        exact_final: float | None = None
        absolute_error: float | None = None
        if exact_function is not None:
            exact_final = evaluate_exact_solution(exact_function, x_end)
            absolute_error = abs(final_y - exact_final)

        raw_records.append(
            {
                "level": level,
                "steps": steps,
                "step_size": (x_end - x0) / steps,
                "final_approximation": final_y,
                "exact_final_value": exact_final,
                "absolute_error": absolute_error,
                "successive_difference": None,
                "observed_order": None,
            }
        )

    for index in range(1, len(raw_records)):
        raw_records[index]["successive_difference"] = abs(
            float(raw_records[index]["final_approximation"])
            - float(raw_records[index - 1]["final_approximation"])
        )

    if exact_function is not None:
        for index in range(1, len(raw_records)):
            previous_error = raw_records[index - 1]["absolute_error"]
            current_error = raw_records[index]["absolute_error"]
            if (
                previous_error is not None
                and current_error is not None
                and float(previous_error) > ZERO_TOLERANCE
                and float(current_error) > ZERO_TOLERANCE
            ):
                raw_records[index]["observed_order"] = math.log(
                    float(previous_error) / float(current_error), 2.0
                )
    else:
        for index in range(2, len(raw_records)):
            previous_difference = raw_records[index - 1]["successive_difference"]
            current_difference = raw_records[index]["successive_difference"]
            if (
                previous_difference is not None
                and current_difference is not None
                and float(previous_difference) > ZERO_TOLERANCE
                and float(current_difference) > ZERO_TOLERANCE
            ):
                raw_records[index]["observed_order"] = math.log(
                    float(previous_difference) / float(current_difference), 2.0
                )

    return tuple(ConvergenceRecord(**record) for record in raw_records)


def error_result(message: str, input_signature: str = "") -> RK4Result:
    """Create a consistent failed result."""

    return RK4Result(
        status="error",
        success=False,
        method=METHOD_NAME,
        message=message,
        stopping_reason="Execution stopped because input validation or numerical evaluation failed.",
        ode_text="",
        ode_expression=None,
        exact_solution_text="",
        exact_solution_expression=None,
        x0=None,
        y0=None,
        x_end=None,
        steps=0,
        step_size=None,
        direction="Not available",
        iterations=(),
        final_x=None,
        final_y=None,
        exact_final_y=None,
        signed_final_error=None,
        absolute_final_error=None,
        relative_final_error=None,
        maximum_absolute_error=None,
        rmse=None,
        convergence_records=(),
        latest_observed_order=None,
        warnings=(),
        input_signature=input_signature,
        execution_datetime=current_report_datetime(),
    )


def solve_rk4_method(
    ode_text: str,
    x0_input: Any,
    y0_input: Any,
    x_end_input: Any,
    steps_input: Any,
    exact_solution_text: str = "",
    convergence_levels_input: Any = DEFAULT_CONVERGENCE_LEVELS,
) -> RK4Result:
    """Validate inputs, solve the IVP, analyze error, and return all results."""

    try:
        x0 = validate_finite_real(x0_input, "Initial x value x0")
        y0 = validate_finite_real(y0_input, "Initial value y0")
        x_end = validate_finite_real(x_end_input, "Final x value")

        try:
            steps = int(steps_input)
        except (TypeError, ValueError, OverflowError) as error:
            raise ValueError("The number of steps must be an integer.") from error

        try:
            convergence_levels = int(convergence_levels_input)
        except (TypeError, ValueError, OverflowError) as error:
            raise ValueError("Convergence levels must be an integer.") from error

        if steps < MIN_STEPS or steps > MAX_STEPS:
            raise ValueError(
                f"The number of steps must be between {MIN_STEPS} and {MAX_STEPS}."
            )
        if not MIN_CONVERGENCE_LEVELS <= convergence_levels <= MAX_CONVERGENCE_LEVELS:
            raise ValueError(
                "Convergence levels must be between "
                f"{MIN_CONVERGENCE_LEVELS} and {MAX_CONVERGENCE_LEVELS}."
            )
        if abs(x_end - x0) <= ZERO_TOLERANCE:
            raise ValueError("The final x value must be different from x0.")

        input_signature = create_input_signature(
            ode_text,
            x0,
            y0,
            x_end,
            steps,
            exact_solution_text,
            convergence_levels,
        )

        ode_expression, ode_function = parse_ode_expression(ode_text)
        exact_expression, exact_function = parse_exact_solution(exact_solution_text)

        warnings: list[str] = []

        if exact_function is not None:
            exact_initial = evaluate_exact_solution(exact_function, x0)
            initial_mismatch = abs(exact_initial - y0)
            initial_scale = max(1.0, abs(y0), abs(exact_initial))
            if initial_mismatch > 1.0e-8 * initial_scale:
                raise ValueError(
                    "The optional exact solution does not satisfy the supplied initial "
                    f"condition: y({x0:.12g}) = {exact_initial:.12g}, but y0 = {y0:.12g}."
                )

            exact_derivative = sp.diff(exact_expression, X_SYMBOL)
            residual_expression = sp.simplify(
                exact_derivative - ode_expression.subs(Y_SYMBOL, exact_expression)
            )
            if residual_expression != 0:
                try:
                    residual_function = sp.lambdify(
                        X_SYMBOL, residual_expression, modules=["numpy"]
                    )
                    check_points = np.linspace(x0, x_end, 5)
                    residual_values = [
                        abs(
                            validate_finite_real(
                                residual_function(point), "Exact solution residual"
                            )
                        )
                        for point in check_points
                    ]
                    if max(residual_values) > 1.0e-7:
                        raise ValueError(
                            "The optional exact solution does not satisfy y' = f(x, y) "
                            "over the requested interval."
                        )
                except ValueError:
                    raise
                except (TypeError, OverflowError, ZeroDivisionError, FloatingPointError):
                    warnings.append(
                        "The exact solution could not be fully verified symbolically; "
                        "it will be used only for numerical comparison."
                    )

        final_y, iterations = run_rk4_grid(
            ode_function=ode_function,
            x0=x0,
            y0=y0,
            x_end=x_end,
            steps=steps,
            exact_function=exact_function,
            store_history=True,
        )

        exact_final_y: float | None = None
        signed_final_error: float | None = None
        absolute_final_error: float | None = None
        relative_final_error: float | None = None
        maximum_absolute_error: float | None = None
        rmse: float | None = None

        if exact_function is not None:
            exact_final_y = evaluate_exact_solution(exact_function, x_end)
            signed_final_error = final_y - exact_final_y
            absolute_final_error = abs(signed_final_error)
            if abs(exact_final_y) > RELATIVE_ERROR_DENOMINATOR_TOLERANCE:
                relative_final_error = absolute_final_error / abs(exact_final_y)

            errors = np.array(
                [record.signed_error for record in iterations if record.signed_error is not None],
                dtype=float,
            )
            if errors.size:
                maximum_absolute_error = float(np.max(np.abs(errors)))
                rmse = float(np.sqrt(np.mean(errors**2)))

        convergence_records = build_convergence_analysis(
            ode_function=ode_function,
            exact_function=exact_function,
            x0=x0,
            y0=y0,
            x_end=x_end,
            base_steps=steps,
            levels=convergence_levels,
        )

        observed_orders = [
            record.observed_order
            for record in convergence_records
            if record.observed_order is not None and math.isfinite(record.observed_order)
        ]
        latest_observed_order = observed_orders[-1] if observed_orders else None

        if any(abs(record.y_next) >= VALUE_MAGNITUDE_WARNING for record in iterations):
            warnings.append(
                "The numerical solution reached a very large magnitude. RK4 may be "
                "unstable for this step size or differential equation."
            )

        if latest_observed_order is not None and latest_observed_order < 3.0:
            warnings.append(
                "The observed refinement behavior is weaker than the expected "
                "fourth-order convergence. Reduce the step size and inspect possible "
                "instability or nonsmooth behavior."
            )

        if exact_function is None:
            warnings.append(
                "No exact solution was supplied. True global errors are unavailable; "
                "the convergence table therefore uses successive final-value differences."
            )

        direction = "Forward integration" if x_end > x0 else "Backward integration"
        step_size = (x_end - x0) / steps

        return RK4Result(
            status="success",
            success=True,
            method=METHOD_NAME,
            message="Execution completed successfully.",
            stopping_reason=(
                "The requested number of uniform RK4 steps was completed and "
                "the final grid point was reached."
            ),
            ode_text=ode_text.strip(),
            ode_expression=ode_expression,
            exact_solution_text=exact_solution_text.strip(),
            exact_solution_expression=exact_expression,
            x0=x0,
            y0=y0,
            x_end=x_end,
            steps=steps,
            step_size=step_size,
            direction=direction,
            iterations=iterations,
            final_x=x_end,
            final_y=final_y,
            exact_final_y=exact_final_y,
            signed_final_error=signed_final_error,
            absolute_final_error=absolute_final_error,
            relative_final_error=relative_final_error,
            maximum_absolute_error=maximum_absolute_error,
            rmse=rmse,
            convergence_records=convergence_records,
            latest_observed_order=latest_observed_order,
            warnings=tuple(warnings),
            input_signature=input_signature,
            execution_datetime=current_report_datetime(),
        )

    except (ValueError, TypeError, ArithmeticError, OverflowError) as error:
        signature = ""
        try:
            signature = create_input_signature(
                str(ode_text),
                float(x0_input),
                float(y0_input),
                float(x_end_input),
                int(steps_input),
                str(exact_solution_text),
                int(convergence_levels_input),
            )
        except (TypeError, ValueError, OverflowError):
            pass
        return error_result(str(error), signature)


# =============================================================================
# DataFrame builders
# =============================================================================
def iterations_dataframe(result: RK4Result) -> pd.DataFrame:
    """Return the complete RK4 stage and update table."""

    rows = []
    for record in result.iterations:
        rows.append(
            {
                "Iteration": record.iteration,
                "x_n": record.x_n,
                "y_n": record.y_n,
                "h": record.step_size,
                "k1": record.k1,
                "k2 x": record.k2_x,
                "k2 y": record.k2_y,
                "k2": record.k2,
                "k3 x": record.k3_x,
                "k3 y": record.k3_y,
                "k3": record.k3,
                "k4 x": record.k4_x,
                "k4 y": record.k4_y,
                "k4": record.k4,
                "Weighted Slope": record.weighted_slope,
                "RK4 Increment": record.increment,
                "x_(n+1)": record.x_next,
                "y_(n+1)": record.y_next,
                "Step Change": record.step_change,
                "Exact y_(n+1)": record.exact_y_next,
                "Signed Error": record.signed_error,
                "Absolute Error": record.absolute_error,
                "Relative Error": record.relative_error,
                "Operation": record.operation,
                "Status": record.status,
            }
        )
    return pd.DataFrame(rows)

def solution_values_dataframe(result: RK4Result) -> pd.DataFrame:
    """Return all numerical solution points including the initial condition."""

    initial_exact = None
    if result.exact_solution_expression is not None:
        exact_function = sp.lambdify(
            X_SYMBOL, result.exact_solution_expression, modules=["numpy"]
        )
        initial_exact = evaluate_exact_solution(exact_function, float(result.x0))

    rows = [
        {
            "Point": 0,
            "x": result.x0,
            "RK4 Approximation": result.y0,
            "Exact Solution": initial_exact,
            "Absolute Error": 0.0 if initial_exact is not None else None,
        }
    ]
    for record in result.iterations:
        rows.append(
            {
                "Point": record.iteration,
                "x": record.x_next,
                "RK4 Approximation": record.y_next,
                "Exact Solution": record.exact_y_next,
                "Absolute Error": record.absolute_error,
            }
        )
    return pd.DataFrame(rows)

def error_analysis_dataframe(result: RK4Result) -> pd.DataFrame:
    """Return pointwise global errors when an exact solution is available."""

    if result.exact_solution_expression is None:
        return pd.DataFrame(
            {"Message": ["No exact solution was supplied; true global error is unavailable."]}
        )
    return pd.DataFrame(
        [
            {
                "Iteration": record.iteration,
                "x": record.x_next,
                "RK4 Approximation": record.y_next,
                "Exact Solution": record.exact_y_next,
                "Signed Error": record.signed_error,
                "Absolute Error": record.absolute_error,
                "Relative Error": record.relative_error,
            }
            for record in result.iterations
        ]
    )

def convergence_dataframe(result: RK4Result) -> pd.DataFrame:
    """Return grid-refinement convergence results."""

    return pd.DataFrame(
        [
            {
                "Level": record.level,
                "Steps": record.steps,
                "h": record.step_size,
                "Final Approximation": record.final_approximation,
                "Exact Final Value": record.exact_final_value,
                "Absolute Error": record.absolute_error,
                "Successive Difference": record.successive_difference,
                "Observed Order": record.observed_order,
            }
            for record in result.convergence_records
        ]
    )


def method_formula_dataframe(result: RK4Result) -> pd.DataFrame:
    """Return the classical RK4 formulas for the Excel report."""

    return pd.DataFrame(
        {
            "Item": [
                "Initial-Value Problem",
                "Uniform Step Size",
                "First Stage",
                "Second Stage",
                "Third Stage",
                "Fourth Stage",
                "Weighted Slope",
                "RK4 Update",
                "Expected Global Order",
                "Expected Local Truncation Error",
            ],
            "Formula": [
                "y' = f(x, y), y(x0) = y0",
                "h = (x_end - x0) / N",
                "k1 = f(x_n, y_n)",
                "k2 = f(x_n + h/2, y_n + h*k1/2)",
                "k3 = f(x_n + h/2, y_n + h*k2/2)",
                "k4 = f(x_n + h, y_n + h*k3)",
                "k_weighted = (k1 + 2*k2 + 2*k3 + k4)/6",
                "y_(n+1) = y_n + h*k_weighted",
                "O(h^4)",
                "O(h^5) per step",
            ],
        }
    )

# =============================================================================
# Scientific plots
# =============================================================================
def create_solution_figure(result: RK4Result) -> Figure:
    """Plot the RK4 numerical solution and optional exact solution."""

    solution_df = solution_values_dataframe(result)
    figure, axis = plt.subplots(figsize=(10, 6))
    axis.plot(
        solution_df["x"],
        solution_df["RK4 Approximation"],
        marker="o",
        linewidth=2,
        label="RK4 Approximation",
    )

    if result.exact_solution_expression is not None:
        exact_function = sp.lambdify(
            X_SYMBOL, result.exact_solution_expression, modules=["numpy"]
        )
        x_values = np.linspace(float(result.x0), float(result.x_end), 500)
        y_values = np.array(
            [evaluate_exact_solution(exact_function, value) for value in x_values],
            dtype=float,
        )
        axis.plot(x_values, y_values, linewidth=2, label="Exact Solution")

    stage_x = []
    stage_y = []
    for record in result.iterations:
        stage_x.extend([record.k2_x, record.k3_x, record.k4_x])
        stage_y.extend([record.k2_y, record.k3_y, record.k4_y])
    axis.scatter(stage_x, stage_y, s=24, marker="x", label="RK4 Stage Points", zorder=4)
    axis.scatter([result.x0], [result.y0], s=90, marker="s", label="Initial Condition", zorder=5)
    axis.scatter(
        [result.final_x], [result.final_y], s=110, marker="*",
        label=f"Final RK4 Value = {result.final_y:.6g}", zorder=6,
    )
    axis.axhline(0.0, linewidth=1)
    axis.axvline(0.0, linewidth=1)
    axis.set_title("Fourth-Order Runge–Kutta Numerical Solution")
    axis.set_xlabel("x")
    axis.set_ylabel("y")
    axis.grid(True)
    axis.legend()
    figure.tight_layout()
    return figure

def create_error_figure(result: RK4Result) -> Figure | None:
    """Plot pointwise absolute global error on a semilog axis."""

    if result.exact_solution_expression is None:
        return None

    dataframe = error_analysis_dataframe(result)
    errors = np.maximum(
        dataframe["Absolute Error"].to_numpy(dtype=float),
        np.finfo(float).tiny,
    )

    figure, axis = plt.subplots(figsize=(10, 6))
    axis.semilogy(dataframe["x"], errors, marker="o", linewidth=2)
    axis.set_title("Fourth-Order Runge–Kutta Method Global Error")
    axis.set_xlabel("x")
    axis.set_ylabel("Absolute Error (Log Scale)")
    axis.grid(True, which="both")
    figure.tight_layout()
    return figure


def create_convergence_figure(result: RK4Result) -> Figure:
    """Plot convergence against absolute step size on log-log axes."""

    dataframe = convergence_dataframe(result)
    h_values = np.abs(dataframe["h"].to_numpy(dtype=float))

    if result.exact_solution_expression is not None:
        metric = dataframe["Absolute Error"].to_numpy(dtype=float)
        label = "Absolute Final Error"
    else:
        metric = dataframe["Successive Difference"].to_numpy(dtype=float)
        label = "Successive Final-Value Difference"

    valid = (
        np.isfinite(metric)
        & (metric > 0.0)
        & np.isfinite(h_values)
        & (h_values > 0.0)
    )

    figure, axis = plt.subplots(figsize=(10, 6))
    if np.any(valid):
        axis.loglog(h_values[valid], metric[valid], marker="o", linewidth=2, label=label)
        reference_h = h_values[valid]
        reference_metric = metric[valid]
        reference_line = reference_metric[-1] * (reference_h / reference_h[-1]) ** 4
        axis.loglog(
            reference_h,
            reference_line,
            linestyle="--",
            label="Second-Order Reference O(h⁴)",
        )
    else:
        axis.text(
            0.5,
            0.5,
            "Insufficient nonzero data for a log-log convergence plot.",
            ha="center",
            va="center",
            transform=axis.transAxes,
        )

    axis.set_title("Fourth-Order Runge–Kutta Method Grid-Refinement Convergence")
    axis.set_xlabel("|h|")
    axis.set_ylabel("Error Indicator")
    axis.grid(True, which="both")
    axis.legend()
    figure.tight_layout()
    return figure


# =============================================================================
# Excel report
# =============================================================================
def format_workbook(workbook: Any) -> None:
    """Apply readable formatting to all workbook sheets."""

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000000E+00"

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 55)


def generate_excel_report(result: RK4Result) -> bytes:
    """Create a formatted in-memory XLSX report for a successful result."""

    if not result.success:
        raise ValueError("Only successful RK4 results can be exported.")

    summary_rows = [
        ("Method", result.method),
        ("Status", result.status),
        ("Message", result.message),
        ("Differential Equation", f"y' = {result.ode_text}"),
        ("Initial Condition", f"y({result.x0:.15g}) = {result.y0:.15g}"),
        ("Final x", result.x_end),
        ("Direction", result.direction),
        ("Number of Steps", result.steps),
        ("Step Size h", result.step_size),
        ("Final RK4 Approximation", result.final_y),
        ("Exact Solution", result.exact_solution_text or "Not supplied"),
        ("Exact Final Value", result.exact_final_y),
        ("Signed Final Error", result.signed_final_error),
        ("Absolute Final Error", result.absolute_final_error),
        ("Relative Final Error", result.relative_final_error),
        ("Maximum Absolute Error", result.maximum_absolute_error),
        ("RMSE", result.rmse),
        ("Expected Global Order", 4),
        ("Latest Observed Order", result.latest_observed_order),
        ("Warnings", " | ".join(result.warnings) if result.warnings else "None"),
        ("Stopping Reason", result.stopping_reason),
        (
            "Execution Date",
            result.execution_datetime.strftime("%Y-%m-%d %H:%M:%S %Z"),
        ),
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["Property", "Value"])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        method_formula_dataframe(result).to_excel(
            writer, sheet_name="Method Formula", index=False
        )
        iterations_dataframe(result).to_excel(
            writer, sheet_name="Iteration Results", index=False
        )
        solution_values_dataframe(result).to_excel(
            writer, sheet_name="Solution Values", index=False
        )
        error_analysis_dataframe(result).to_excel(
            writer, sheet_name="Error Analysis", index=False
        )
        convergence_dataframe(result).to_excel(
            writer, sheet_name="Convergence Analysis", index=False
        )

        workbook = writer.book
        format_workbook(workbook)

        iteration_sheet = workbook["Iteration Results"]
        if result.iterations:
            chart = LineChart()
            chart.title = "RK4 Numerical Solution"
            chart.y_axis.title = "y"
            chart.x_axis.title = "x"
            max_row = len(result.iterations) + 1
            x_reference = Reference(iteration_sheet, min_col=17, min_row=2, max_row=max_row)
            y_reference = Reference(iteration_sheet, min_col=18, min_row=1, max_row=max_row)
            chart.add_data(y_reference, titles_from_data=True)
            chart.set_categories(x_reference)
            chart.height = 8
            chart.width = 15
            iteration_sheet.add_chart(chart, "T2")

        convergence_sheet = workbook["Convergence Analysis"]
        if len(result.convergence_records) > 1:
            chart = LineChart()
            chart.title = "Grid-Refinement Convergence"
            chart.y_axis.title = (
                "Absolute Error"
                if result.exact_solution_expression is not None
                else "Successive Difference"
            )
            chart.x_axis.title = "Steps"
            max_row = len(result.convergence_records) + 1
            categories = Reference(
                convergence_sheet, min_col=2, min_row=2, max_row=max_row
            )
            metric_col = 6 if result.exact_solution_expression is not None else 7
            data = Reference(
                convergence_sheet, min_col=metric_col, min_row=1, max_row=max_row
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 15
            convergence_sheet.add_chart(chart, "J2")

    output.seek(0)
    return finalize_excel_report_with_visible_charts(output.getvalue())


# =============================================================================
# Streamlit renderers
# =============================================================================

def finalize_excel_report_with_visible_charts(report_bytes: bytes) -> bytes:
    """Place existing workbook charts on Summary so they are immediately visible.

    The report data and worksheets are preserved. If Excel chart post-processing is
    unavailable for any reason, the original report is returned unchanged rather
    than risking a damaged workbook.
    """
    try:
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_workbook

        workbook = _load_workbook(_BytesIO(report_bytes))
        if "Summary" not in workbook.sheetnames:
            return report_bytes

        summary_sheet = workbook["Summary"]
        collected_charts = list(summary_sheet._charts)
        summary_sheet._charts = []

        for worksheet in workbook.worksheets:
            if worksheet is summary_sheet:
                continue
            if worksheet._charts:
                collected_charts.extend(list(worksheet._charts))
                worksheet._charts = []

        # Put every chart in the existing Summary worksheet, below one another.
        # D2 keeps the first chart visible beside the main results when the report opens.
        for chart_index, chart in enumerate(collected_charts):
            anchor_row = 2 + chart_index * 19
            summary_sheet.add_chart(chart, f"D{anchor_row}")

        workbook.active = workbook.sheetnames.index("Summary")
        output = _BytesIO()
        workbook.save(output)
        return output.getvalue()
    except Exception:
        return report_bytes


def render_final_result(result: RK4Result) -> None:
    """Render the final result card."""

    if not result.success:
        st.error(result.message)
        return
    st.success(result.message)
    metric_columns = st.columns(2)
    metric_columns[0].metric("Final x", format_number(result.final_x))
    metric_columns[1].metric("RK4 Approximation", format_number(result.final_y))
    st.markdown(f"**Differential Equation:** `y' = {result.ode_text}`")
    st.markdown(
        f"**Initial Condition:** `y({format_number(result.x0)}) = {format_number(result.y0)}`"
    )
    st.markdown(f"**Number of Steps:** {result.steps}")
    st.markdown(f"**Step Size:** {format_number(result.step_size)}")
    st.markdown(f"**Direction:** {result.direction}")
    if result.exact_final_y is not None:
        st.markdown(f"**Exact Final Value:** {format_number(result.exact_final_y)}")
        st.markdown(f"**Absolute Final Error:** {format_number(result.absolute_final_error)}")
        st.markdown(f"**Relative Final Error:** {format_number(result.relative_final_error)}")
    st.markdown(f"**Stopping Reason:** {result.stopping_reason}")
    for warning in result.warnings:
        st.warning(warning)

def render_method_summary(result: RK4Result) -> None:
    """Render the RK4 formulas and problem summary."""

    st.subheader("Method Formula and Problem Setup")
    column_1, column_2 = st.columns(2)
    with column_1:
        st.latex(r"y' = f(x,y), \qquad y(x_0)=y_0")
        st.latex(r"k_1=f(x_n,y_n)")
        st.latex(r"k_2=f\left(x_n+\frac{h}{2},y_n+\frac{h}{2}k_1\right)")
    with column_2:
        st.latex(r"k_3=f\left(x_n+\frac{h}{2},y_n+\frac{h}{2}k_2\right)")
        st.latex(r"k_4=f(x_n+h,y_n+h k_3)")
        st.latex(r"y_{n+1}=y_n+\frac{h}{6}(k_1+2k_2+2k_3+k_4)")

    summary_df = pd.DataFrame(
        {
            "Property": [
                "Right-Hand Side f(x, y)", "Initial x", "Initial y", "Final x",
                "Steps", "Uniform h", "Expected Global Accuracy",
            ],
            "Value": [
                result.ode_text, result.x0, result.y0, result.x_end,
                result.steps, result.step_size, "Fourth order: O(h^4)",
            ],
        }
    )
    st.dataframe(round_numeric_dataframe(summary_df), use_container_width=True, hide_index=True)

def render_iteration_table(result: RK4Result) -> None:
    """Render the full RK4 stage table and detailed operations."""

    st.subheader("RK4 Iteration Table")
    dataframe = iterations_dataframe(result)
    st.dataframe(round_numeric_dataframe(dataframe), use_container_width=True, hide_index=True)
    st.caption(
        "Each accepted update uses four slope evaluations. Step Change is the "
        "change between successive approximations and is not the true global "
        "error unless an exact solution is supplied."
    )
    with st.expander("Detailed RK4 Stage Operations"):
        for record in result.iterations:
            st.markdown(f"**Iteration {record.iteration}:** `{record.operation}`")
            st.caption(
                f"k1 = {record.k1:.12g}; k2 = {record.k2:.12g}; "
                f"k3 = {record.k3:.12g}; k4 = {record.k4:.12g}; "
                f"weighted slope = {record.weighted_slope:.12g}; "
                f"increment = {record.increment:.12g}."
            )

def render_solution_graph(result: RK4Result) -> None:
    """Render the numerical solution graph."""

    st.subheader("Numerical Solution Graph")
    try:
        figure = create_solution_figure(result)
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)
    except (ValueError, TypeError, ArithmeticError, OverflowError) as error:
        st.warning(f"The solution graph could not be displayed: {error}")


def render_error_analysis(result: RK4Result) -> None:
    """Render pointwise and summary error analysis."""

    st.subheader("Error Analysis")
    if result.exact_solution_expression is None:
        st.info(
            "Enter an exact solution y(x) to calculate true global error, relative "
            "error, maximum absolute error, and RMSE."
        )
        return

    metric_columns = st.columns(4)
    metric_columns[0].metric(
        "Final Absolute Error", format_number(result.absolute_final_error)
    )
    metric_columns[1].metric(
        "Final Relative Error", format_number(result.relative_final_error)
    )
    metric_columns[2].metric(
        "Maximum Absolute Error", format_number(result.maximum_absolute_error)
    )
    metric_columns[3].metric("RMSE", format_number(result.rmse))

    dataframe = error_analysis_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(dataframe), use_container_width=True, hide_index=True
    )

    figure = create_error_figure(result)
    if figure is not None:
        st.pyplot(figure, use_container_width=True)
        plt.close(figure)


def render_convergence_analysis(result: RK4Result) -> None:
    """Render grid-refinement convergence analysis."""

    st.subheader("Convergence Analysis")
    st.markdown(
        "The classical RK4 method has **fourth-order global convergence** "
        "under the usual smoothness and stability conditions. Halving the step "
        "size should reduce the global error by approximately a factor of sixteen "
        "once the asymptotic range is reached."
    )

    dataframe = convergence_dataframe(result)
    st.dataframe(
        round_numeric_dataframe(dataframe), use_container_width=True, hide_index=True
    )

    if result.latest_observed_order is not None:
        st.metric("Latest Observed Order", format_number(result.latest_observed_order))

    figure = create_convergence_figure(result)
    st.pyplot(figure, use_container_width=True)
    plt.close(figure)


def render_excel_download(result: RK4Result) -> None:
    """Generate and render the Excel download button."""

    st.subheader("Excel Report")
    report_signature = result.input_signature

    if (
        st.session_state.get("rk4_excel_signature") != report_signature
        or "rk4_excel_report" not in st.session_state
    ):
        try:
            st.session_state.rk4_excel_report = generate_excel_report(result)
            st.session_state.rk4_excel_signature = report_signature
        except (ValueError, OSError, TypeError, ArithmeticError) as error:
            st.error(f"The Excel report could not be generated: {error}")
            return

    report_bytes = st.session_state.get("rk4_excel_report")
    if not report_bytes:
        st.error("The Excel report is unavailable.")
        return

    timestamp = result.execution_datetime.strftime("%Y%m%d_%H%M%S")
    filename = f"fourth_order_runge_kutta_report_{timestamp}.xlsx"
    st.download_button(
        label="Download Excel Report",
        data=report_bytes,
        file_name=filename,
        mime=EXCEL_MIME_TYPE,
        use_container_width=True,
        key="rk4_download_button",
    )


# =============================================================================
# Streamlit page
# =============================================================================
def render_page() -> None:
    """Render the complete Fourth-Order Runge–Kutta Method Streamlit page."""

    st.set_page_config(
        page_title="Fourth-Order Runge–Kutta Method Solver | Numerical Methods",
        page_icon="∫",
        layout="wide",
    )
    load_css()
    navbar(active_page="solver")

    st.html(
        """
        <section class="solver-hero">
            <div>
                <div class="page-label">ORDINARY DIFFERENTIAL EQUATIONS TOOL</div>
                <h1>Fourth-Order Runge–Kutta Method Solver</h1>
                <p>
                    Solve a first-order initial-value problem using the classical
                    fourth-order Runge–Kutta method. Review every stage slope,
                    intermediate stage point, weighted update, optional exact-solution
                    error, convergence result, graph, and report.
                </p>
                <div class="method-actions">
                    <a href="/Runge_Kutta_4" target="_self"
                       class="btn-outline-ui">Review Lesson →</a>
                    <a href="/Runge_Kutta_4_Quiz" target="_self"
                       class="btn-primary-ui">Take Quiz →</a>
                </div>
            </div>
        </section>
        """
    )

    # Match the centered Bisection solver content width.
    left_margin, main_area, right_margin = st.columns([0.035, 0.93, 0.035])
    with main_area:
        st.markdown(
            '<main class="solver-wrapper solver-streamlit-area">',
            unsafe_allow_html=True,
        )
    
        guide_column, conditions_column = st.columns(2)

        with guide_column:
            with st.container(border=True):
                st.subheader('How to Write the Function')
                st.markdown(
                    """
                Enter only the right-hand side of **y′ = f(x, y)**, without an equals sign.

                - Use only **x** and **y** as variables.
                - Powers: write `x**2`, not **x^2**.
                - Multiplication: write `2*x` or `x*y`, not `2x` or `xy`.
                - Use lowercase functions such as **sin(x)**, **exp(x)**, **sqrt(x)**, and **log(x)**.
                - The optional exact solution must be written as a function of **x** only.
                    """
                )

        with conditions_column:
            with st.container(border=True):
                st.subheader('Before Solving')
                st.markdown(
                    """
                - Supply the initial condition **y(x₀) = y₀** and a final point different from **x₀**.
                - The number of steps must be a positive integer.
                - Classical RK4 evaluates four stages per step and is fourth-order accurate for sufficiently smooth problems.
                - The optional exact solution must satisfy the supplied initial condition.
                - The differential equation must remain finite at all four stage points.
                    """
                )

        input_column, result_column = st.columns([1.35, 1.0])
    
        with input_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Input</h3>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    '<div class="input-label-ui">Right-hand side f(x, y)</div>',
                    unsafe_allow_html=True,
                )
                ode_text = st.text_input(
                    "Right-hand side f(x, y)",
                    value=DEFAULT_ODE,
                    placeholder="Example: x + y or y - x**2 + 1",
                    label_visibility="collapsed",
                    key="rk4_ode_function",
                )
    
                first_row = st.columns(3)
                with first_row[0]:
                    st.markdown('<div class="input-label-ui">Initial x₀</div>', unsafe_allow_html=True)
                    x0 = st.number_input(
                        "Initial x0", value=DEFAULT_X0, format="%.12g",
                        label_visibility="collapsed", key="rk4_x0"
                    )
                with first_row[1]:
                    st.markdown('<div class="input-label-ui">Initial y₀</div>', unsafe_allow_html=True)
                    y0 = st.number_input(
                        "Initial y0", value=DEFAULT_Y0, format="%.12g",
                        label_visibility="collapsed", key="rk4_y0"
                    )
                with first_row[2]:
                    st.markdown('<div class="input-label-ui">Final x</div>', unsafe_allow_html=True)
                    x_end = st.number_input(
                        "Final x", value=DEFAULT_X_END, format="%.12g",
                        label_visibility="collapsed", key="rk4_x_end"
                    )
    
                second_row = st.columns(2)
                with second_row[0]:
                    st.markdown('<div class="input-label-ui">Number of steps N</div>', unsafe_allow_html=True)
                    steps = st.number_input(
                        "Number of steps", min_value=MIN_STEPS, max_value=MAX_STEPS,
                        value=DEFAULT_STEPS, step=1, label_visibility="collapsed",
                        key="rk4_steps"
                    )
                with second_row[1]:
                    st.markdown('<div class="input-label-ui">Convergence levels</div>', unsafe_allow_html=True)
                    convergence_levels = st.number_input(
                        "Convergence levels", min_value=MIN_CONVERGENCE_LEVELS,
                        max_value=MAX_CONVERGENCE_LEVELS,
                        value=DEFAULT_CONVERGENCE_LEVELS, step=1,
                        label_visibility="collapsed", key="rk4_convergence_levels"
                    )
    
                st.markdown(
                    '<div class="input-label-ui">Exact solution y(x) — optional</div>',
                    unsafe_allow_html=True,
                )
                exact_solution_text = st.text_input(
                    "Exact solution y(x)", value="",
                    placeholder="Example for y' = x + y, y(0)=1: 2*exp(x) - x - 1",
                    label_visibility="collapsed", key="rk4_exact_solution"
                )
    
                if float(x_end) != float(x0):
                    preview_h = (float(x_end) - float(x0)) / int(steps)
                    st.caption(f"Computed uniform step size: h = {preview_h:.12g}")
    
                current_signature = create_input_signature(
                    ode_text, x0, y0, x_end, steps, exact_solution_text, convergence_levels
                )
    
                solve_clicked = st.button(
                    "Solve with RK4", type="primary", use_container_width=True,
                    key="rk4_solve_button"
                )
                if solve_clicked:
                    result = solve_rk4_method(
                        ode_text=ode_text,
                        x0_input=x0,
                        y0_input=y0,
                        x_end_input=x_end,
                        steps_input=steps,
                        exact_solution_text=exact_solution_text,
                        convergence_levels_input=convergence_levels,
                    )
                    st.session_state.rk4_result = result
                    st.session_state.rk4_result_signature = current_signature
                    st.session_state.pop("rk4_excel_report", None)
                    st.session_state.pop("rk4_excel_signature", None)
    
                with st.expander("Example Inputs"):
                    st.code(
                        "ODE: x + y\n"
                        "x0 = 0\n"
                        "y0 = 1\n"
                        "x_end = 1\n"
                        "Steps = 10\n"
                        "Exact solution: 2*exp(x) - x - 1",
                        language=None,
                    )
                    st.code(
                        "ODE: y - x**2 + 1\n"
                        "x0 = 0\n"
                        "y0 = 0.5\n"
                        "x_end = 2\n"
                        "Steps = 20\n"
                        "Exact solution: (x + 1)**2 - 0.5*exp(x)",
                        language=None,
                    )
    
        with result_column:
            with st.container(border=True):
                st.markdown(
                    '<h3 class="solver-box-title">Final Result</h3>',
                    unsafe_allow_html=True,
                )
                saved_result = st.session_state.get("rk4_result")
                saved_signature = st.session_state.get("rk4_result_signature")
    
                if saved_result is None:
                    st.info("Enter the IVP data and select Solve with RK4.")
                elif saved_signature != current_signature:
                    st.warning(
                        "The inputs have changed. Select Solve with RK4 "
                        "to update the result."
                    )
                else:
                    render_final_result(saved_result)
    
        saved_result = st.session_state.get("rk4_result")
        saved_signature = st.session_state.get("rk4_result_signature")
    
        if saved_result is not None and saved_signature == current_signature:
            if saved_result.success:
                st.divider()
                render_method_summary(saved_result)
                st.divider()
                render_iteration_table(saved_result)
                st.divider()
                render_solution_graph(saved_result)
                st.divider()
                render_error_analysis(saved_result)
                st.divider()
                render_convergence_analysis(saved_result)
                st.divider()
                render_excel_download(saved_result)
    
        st.markdown("</main>", unsafe_allow_html=True)

    st.html(
        """
        <footer class="footer-ui">
            <div>NM • © 2026 Numerical Methods</div>
            <div>Fourth-Order Runge-Kutta • Ordinary Differential Equations</div>
        </footer>
        """
    )


if __name__ == "__main__":
    render_page()
